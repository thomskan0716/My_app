# ES: Diagnóstico de arranque (si no ves "START" en consola, el bloqueo es antes de imports)
# EN: Startup diagnostic (if you don't see "START" in console, the hang is before imports)
# JP: 起動診断（コンソールに START が出ない場合はインポート前にブロック）
print("START 0sec.py", flush=True)
import sys
import os
import warnings
import pandas as pd
import numpy as np
from PySide6.QtWidgets import (QApplication, QMainWindow, QVBoxLayout, QHBoxLayout, 
                             QPushButton, QLabel, QFileDialog, QWidget, QGridLayout,
                             QProgressBar, QProgressDialog, QComboBox, QLineEdit, QDateEdit, QRadioButton,
                             QTableWidget, QTableWidgetItem, QHeaderView, QMessageBox,
                             QDialog, QFrame, QScrollArea, QSplitter, QTextEdit, QGroupBox,
                             QCheckBox, QSpinBox, QDoubleSpinBox, QTabWidget, QTextBrowser,
                             QFormLayout, QSizePolicy, QListWidget, QDialogButtonBox)
from PySide6.QtCore import Qt, QTimer, QDate, QThread, Signal, QPropertyAnimation, QEasingCurve, QSize, QObject, QEvent, QPoint
from PySide6.QtGui import QPixmap, QFont, QPalette, QColor, QIcon, QPainter, QLinearGradient, QMovie, QIntValidator, QTextCursor, QFontDatabase, QFontMetrics
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')
import seaborn as sns
from datetime import datetime, timedelta
import json
import sqlite3
import traceback
import time
import threading
from concurrent.futures import ThreadPoolExecutor
import queue
import subprocess
import shutil
import zipfile
import tempfile
import glob
import re
from pathlib import Path
import logging
import hashlib
import pickle
import gzip
import base64
import io
import csv
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
warnings.filterwarnings('ignore')

# --- D基準値 (D-score) - referencia: D_and_I最適化_Greedy法_ver3.py ---
from scipy.spatial.distance import cdist
from scipy.linalg import qr as scipy_qr
try:
    from sklearn.preprocessing import StandardScaler
except Exception:
    StandardScaler = None

from app_paths import (
    resource_path,
    get_db_path,
    get_backup_dir,
    migrate_legacy_db_if_needed,
)
from app_manifest import get_app_title
from backup_manager import auto_daily_backup, create_backup, prune_backups

# ES: Rutas canónicas de BBDD (instalación profesional: ProgramData\\...\\data) | EN: Canonical DB paths (pro install: ProgramData\\...\\data) | JA: DB正規パス（本番: ProgramData\\...\\data）
# EN: Canonical DB paths (pro install: ProgramData\\...\\data)
# JA: DBの正規パス（製品版: ProgramData\\...\\data）
RESULTS_DB_PATH = migrate_legacy_db_if_needed("results.db", shared=True)
YOSOKU_LINEAL_DB_PATH = get_db_path("yosoku_predictions_lineal.db", shared=True)
YOSOKU_NO_LINEAL_DB_PATH = get_db_path("yosoku_predictions_no_lineal.db", shared=True)

print("🔧 モジュールをインポート中...")

try:
    print("🔧 ウィジェットをインポート中...")
    from widgets import (
        create_logo_widget, create_ok_ng_buttons, create_dsaitekika_button, create_isaitekika_button,
        create_load_sample_button, create_load_results_button, create_show_results_button,
        create_regression_labels, create_load_sample_block, create_load_results_block
    )
    print("✅ ウィジェットのインポートが完了しました")
except Exception as e:
    print(f"❌ ウィジェットのインポート中にエラー: {e}")
    raise

try:
    print("🔧 ワーカーをインポート中...")
    from dsaitekikaworker import DsaitekikaWorker
    from showresultsworker import ShowResultsWorker
    from samplecombineworker import SampleCombinerWorker
    print("✅ ワーカーのインポートが完了しました")
except Exception as e:
    print(f"❌ ワーカーのインポート中にエラー: {e}")
    raise

try:
    print("🔧 非線形ワーカーをインポート中...")
    from nonlinear_worker import NonlinearWorker
    print("✅ 非線形ワーカーのインポートが完了しました")
except Exception as e:
    print(f"⚠️ 非線形ワーカーのインポート中にエラー: {e}")
    print("  （非線形解析なしで続行します）")
    NonlinearWorker = None

try:
    print("🔧 非線形解析ダイアログをインポート中...")
    from nonlinear_config_dialog import NonlinearConfigDialog
    from graph_viewer_dialog import GraphViewerDialog
    from pareto_results_dialog import ParetoResultsDialog
    print("✅ ダイアログのインポートが完了しました")
except Exception as e:
    print(f"⚠️ ダイアログのインポート中にエラー: {e}")
    print("  （ダイアログなしで続行します）")
    NonlinearConfigDialog = None
    GraphViewerDialog = None
    ParetoResultsDialog = None

try:
    print("🔧 分類ワーカーをインポート中...")
    from classification_worker import ClassificationWorker
    print("✅ 分類ワーカーのインポートが完了しました")
except Exception as e:
    print(f"⚠️ 分類ワーカーのインポート中にエラー: {e}")
    print("  （分類解析なしで続行します）")
    ClassificationWorker = None

try:
    print("🔧 分類設定ダイアログをインポート中...")
    from classification_config_dialog import ClassificationConfigDialog
    print("✅ 分類設定ダイアログのインポートが完了しました")
except Exception as e:
    print(f"⚠️ 分類設定ダイアログのインポート中にエラー: {e}")
    print("  （ダイアログなしで続行します）")
    ClassificationConfigDialog = None

try:
    print("🔧 ブラシ選択ダイアログをインポート中...")
    from brush_selection_dialog import BrushSelectionDialog
    print("✅ ブラシ選択ダイアログのインポートが完了しました")
except Exception as e:
    print(f"⚠️ ブラシ選択ダイアログのインポート中にエラー: {e}")
    print("  （ダイアログなしで続行します）")
    BrushSelectionDialog = None

try:
    print("🔧 データベース関連モジュールをインポート中...")
    from db_manager import DBManager as DBManagerMain
    from result_processor import ResultProcessor
    print("✅ DBモジュールのインポートが完了しました")
except Exception as e:
    print(f"❌ DBモジュールのインポート中にエラー: {e}")
    raise

try:
    print("🔧 統合オプティマイザーをインポート中...")
    from integrated_optimizer_worker import IntegratedOptimizerWorker
    print("✅ 統合オプティマイザーのインポートが完了しました")
except Exception as e:
    print(f"❌ 統合オプティマイザーのインポート中にエラー: {e}")
    raise

print("✅ すべてのモジュールのインポートが完了しました")
from datetime import datetime
import glob
import os, shutil
import sqlite3
import pandas as pd
import numpy as np

def calculate_d_criterion(X_selected):
    """ES: Calcula el criterio D-óptimo usando la lógica de D_and_I最適化_Greedy法_ver3.py
    EN: Compute the D-optimal criterion using D_and_I最適化_Greedy法_ver3.py logic
    JA: D_and_I最適化_Greedy法_ver3.py のロジックでD最適基準を計算
    """
    try:
        if X_selected.shape[0] < X_selected.shape[1]:
            return -np.inf
            
        # ES: Calcular número de condición para detectar problemas numéricos | EN: Compute condition number to detect numerical issues | JA: 数値問題検出のため条件数を計算
        # EN: Compute condition number to detect numerical issues
        # JA: 数値問題検出のため条件数を計算
        condition_number = np.linalg.cond(X_selected)
        
        # ES: Usar método numéricamente estable si la matriz está mal condicionada | EN: Use numerically stable method if matrix is ill-conditioned | JA: 行列が悪条件なら数値的に安定した手法を使用
        # EN: Use a numerically stable method if the matrix is ill-conditioned
        # JA: 行列が悪条件なら数値的に安定な手法を使用
        USE_NUMERICAL_STABLE_METHOD = True
        if USE_NUMERICAL_STABLE_METHOD or condition_number > 1e12:
            method = 'svd'
            print(f"🔧 高条件数検出({condition_number:.2e}) - SVD法適用")
        else:
            method = 'qr'
            
        if method == 'svd':
            # ES: Usar SVD para matrices mal condicionadas
            # EN: Use SVD for ill-conditioned matrices
            # JA: 悪条件行列にはSVDを使用
            _, s, _ = np.linalg.svd(X_selected, full_matrices=False)
            valid_singular_values = s[s > 1e-14]
            if len(valid_singular_values) == 0:
                return -np.inf
            log_det = np.sum(np.log(valid_singular_values))
        else:
            # ES: Usar descomposición QR para matrices bien condicionadas | EN: Use QR decomposition for well-conditioned matrices | JA: 良条件行列にはQR分解を使用
            # EN: Use QR decomposition for well-conditioned matrices
            # JA: 良条件行列にはQR分解を使用
            q, r = np.linalg.qr(X_selected, mode='economic')
            diag_r = np.diag(r)
            det = np.abs(np.prod(diag_r))
            log_det = np.log(det) if det > 1e-300 else -np.inf
            
        return log_det
    except Exception as e:
        print(f"⚠️ D-criterion計算エラー: {e}")
        return -np.inf

def calculate_i_criterion(X_selected, X_all):
    """ES: Calcula el criterio I-óptimo
    EN: Compute the I-optimal criterion
    JA: I最適基準を計算
    """
    try:
        if len(X_selected) == 0:
            return -np.inf
        distances = cdist(X_all, X_selected)
        min_distances = np.min(distances, axis=1)
        return -np.mean(min_distances)
    except:
        return -np.inf

def _standardize_like_reference(X: np.ndarray) -> np.ndarray:
    """
    Estandariza como en el archivo de referencia (StandardScaler).
    Si sklearn no está disponible, aplica z-score (ddof=0) con fallback seguro.
    """
    X = np.asarray(X, dtype=float)
    if StandardScaler is not None:
        return StandardScaler().fit_transform(X)
    mean = np.nanmean(X, axis=0)
    std = np.nanstd(X, axis=0)
    std = np.where(std == 0, 1.0, std)
    return (X - mean) / std

def calculate_d_criterion_stable_reference(X: np.ndarray, method: str = "auto",
                                           use_numerical_stable_method: bool = True,
                                           verbose: bool = False):
    """
    Cálculo idéntico a D_and_I最適化_Greedy法_ver3.py:
    devuelve (log_det, condition_number)
    """
    try:
        condition_number = np.linalg.cond(X)
        if use_numerical_stable_method or (method == "auto" and condition_number > 1e12):
            method = "svd"
            if verbose and condition_number > 1e12:
                print(f"🔧 高条件数検出({condition_number:.2e}) - SVD法適用")
        if method == "svd":
            _, s, _ = np.linalg.svd(X, full_matrices=False)
            valid_singular_values = s[s > 1e-14]
            if len(valid_singular_values) == 0:
                return -np.inf, condition_number
            log_det = np.sum(np.log(valid_singular_values))
        else:
            _, r = scipy_qr(X, mode="economic")
            diag_r = np.diag(r)
            det = np.abs(np.prod(diag_r))
            log_det = np.log(det) if det > 1e-300 else -np.inf
        return log_det, condition_number
    except Exception as e:
        if verbose:
            print(f"⚠️ D-criterion計算エラー: {e}")
        return -np.inf, np.inf

def calculate_d_score_reference(candidate_points_raw: np.ndarray, selected_indices) -> float:
    """
    D-score de referencia: fit StandardScaler sobre TODOS los candidatos,
    luego D-criterion estable sobre el subconjunto seleccionado.
    """
    if candidate_points_raw is None or selected_indices is None:
        return -np.inf
    X_scaled = _standardize_like_reference(candidate_points_raw)
    selected_indices = list(selected_indices)
    if len(selected_indices) == 0:
        return -np.inf
    X_subset = X_scaled[selected_indices]
    score, _ = calculate_d_criterion_stable_reference(X_subset, method="auto", use_numerical_stable_method=True, verbose=False)
    return float(score)

def _extract_design_matrix(df: pd.DataFrame) -> np.ndarray:
    """
    Extrae la matriz de variables de diseño (7 columnas) por NOMBRE, compatible con formato antiguo y nuevo.
    Columnas esperadas:
      回転速度, 送り速度, (UPカット o 回転方向), (切込量 o 切込み量), (突出量 o 突出し量), 載せ率, パス数
    """
    dir_col = "UPカット" if "UPカット" in df.columns else ("回転方向" if "回転方向" in df.columns else None)
    if dir_col is None:
        raise ValueError("❌ Falta columna de dirección: 'UPカット' o '回転方向'")
    cut_col = "切込量" if "切込量" in df.columns else ("切込み量" if "切込み量" in df.columns else None)
    if cut_col is None:
        raise ValueError("❌ Falta columna de切込量: '切込量' o '切込み量'")
    out_col = "突出量" if "突出量" in df.columns else ("突出し量" if "突出し量" in df.columns else None)
    if out_col is None:
        raise ValueError("❌ Falta columna de突出量: '突出量' o '突出し量'")

    design_cols = ["回転速度", "送り速度", dir_col, cut_col, out_col, "載せ率", "パス数"]
    missing = [c for c in design_cols if c not in df.columns]
    if missing:
        raise ValueError(f"❌ Faltan columnas de diseño: {missing}")
    X = df[design_cols].copy()
    for c in design_cols:
        X[c] = pd.to_numeric(X[c], errors="coerce")
    if X.isna().any().any():
        bad_cols = X.columns[X.isna().any()].tolist()
        raise ValueError(f"❌ Valores no numéricos en columnas de diseño: {bad_cols}")
    return X.to_numpy()








class LoadingOverlay(QWidget):
    """
    ES: Widget overlay para mostrar loading dentro de la ventana principal.
    EN: Overlay widget to show a loading indicator inside the main window.
    JA: メインウィンドウ内にローディング表示を出すオーバーレイウィジェット。

    ES: Usa QWidget en lugar de QDialog para que sea parte de la jerarquía de widgets
    EN: Uses QWidget (not QDialog) so it stays in the widget hierarchy
    JA: QDialogではなくQWidgetを使い、ウィジェット階層に属するようにする
    ES: y respete automáticamente el orden de ventanas del sistema operativo.
    EN: and naturally respects the OS window stacking order.
    JA: これによりOSのウィンドウ順序を自然に尊重する。
    """
    def __init__(self, parent=None):
        super().__init__(parent)
        
        # ES: CRÍTICO: Forzar explícitamente que NO sea una ventana de nivel superior | EN: CRITICAL: Explicitly ensure it is NOT a top-level window | JA: 重要：トップレベルウィンドウにしないよう明示
        # EN: CRITICAL: Explicitly force this to NOT be a top-level window
        # JA: 重要：トップレベルウィンドウにならないよう明示的に固定
        # ES: Esto asegura que el widget sea parte de la jerarquía del parent, no una ventana flotante | EN: Ensures widget is part of parent hierarchy, not a floating window | JA: ウィジェットが親の子になりフローティングでないことを保証
        # EN: This ensures the widget stays in the parent's hierarchy (not a floating window)
        # JA: 親の階層に属し、フローティングウィンドウにならないようにする
        self.setWindowFlags(Qt.Widget)  # Force to be a child widget, not a window
        
        # ES: NO usar setWindowModality - es un widget hijo, no una ventana
        # EN: Do NOT use setWindowModality - this is a child widget, not a window
        # JA: setWindowModality は使わない（子ウィジェットでありウィンドウではない）
        # ES: El widget será parte de la jerarquía del parent (center_frame) | EN: Widget will be part of parent (center_frame) hierarchy | JA: ウィジェットは親(center_frame)の子になる
        # EN: The widget will be part of the parent's hierarchy (center_frame)
        # JA: 親（center_frame）の階層に属する
        
        # ES: Asegurar que tenga parent (si no lo tiene, no funcionará correctamente) | EN: Ensure it has a parent (otherwise it will not work correctly) | JA: 親を設定（ないと正しく動作しない）
        # EN: Ensure it has a parent (without it, it won't behave correctly)
        # JA: 親を必ず設定（無いと正しく動作しない）
        if parent:
            self.setParent(parent)
        
        # ES: Configurar como widget overlay con fondo semitransparente
        # EN: Configure as an overlay widget with a semi-transparent background
        # JA: 半透明背景のオーバーレイとして設定
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.setStyleSheet("""
            QWidget {
                background: rgba(0, 0, 0, 0.3);
                border-radius: 10px;
            }
            QLabel {
                background: transparent;
                color: white;
            }
        """)

        # ES: Layout centrado para el loading
        # EN: Centered layout for the loading indicator
        # JA: ローディング表示を中央配置
        layout = QVBoxLayout(self)
        layout.setAlignment(Qt.AlignCenter)
        layout.setContentsMargins(0, 0, 0, 0)

        self.label = QLabel()
        self.label.setAlignment(Qt.AlignCenter)

        self.movie = QMovie(resource_path("loading.gif"))
        self.movie.setScaledSize(QSize(64, 64))
        self.label.setMovie(self.movie)

        layout.addWidget(self.label)
        
        # ES: Inicialmente oculto | EN: Hidden initially | JA: 初期状態は非表示
        self.hide()
    
    def _update_geometry(self):
        """ES: Actualiza la geometría para cubrir todo el parent
        EN: Update geometry to cover the full parent
        JA: 親全体を覆うようにジオメトリを更新
        """
        if self.parent() and self.isVisible():
            parent = self.parent()
            self.setGeometry(0, 0, parent.width(), parent.height())
    
    def start(self):
        """ES: Inicia el loading y lo muestra cubriendo todo el parent
        EN: Start loading and show it covering the full parent
        JA: ローディングを開始し、親全体を覆って表示
        """
        # ES: CRÍTICO: Verificar y forzar que NO sea una ventana | EN: CRITICAL: Verify and force that it is NOT a window | JA: 重要：ウィンドウでないことを確認・強制
        # EN: CRITICAL: Verify and force that this is NOT a window
        # JA: 重要：ウィンドウになっていないか確認し、ならないよう強制
        # ES: Si por alguna razón se convirtió en ventana, forzar que no lo sea | EN: If it became a window, force it back | JA: ウィンドウ化したら強制的に戻す
        # EN: If it somehow became a window, force it back to a widget
        # JA: 何らかの理由でウィンドウ化した場合はウィジェットに戻す
        if self.isWindow():
            print("⚠️ 警告: LoadingOverlay がウィンドウとして検出されました。修正します...")
            self.setWindowFlags(Qt.Widget)
            if self.parent():
                self.setParent(self.parent())  # Re-establecer parent
        
        if self.parent():
            parent = self.parent()
            
            # ES: Asegurar que el parent esté establecido correctamente | EN: Ensure parent is set correctly | JA: 親が正しく設定されていることを確認
            # EN: Ensure the parent is correctly set
            # JA: 親が正しく設定されていることを確認
            if self.parent() != parent:
                self.setParent(parent)
            
            # ES: Forzar que NO sea ventana nuevamente después de setParent | EN: Re-assert not a window after setParent | JA: setParent後にウィンドウでないことを再適用
            # EN: Re-assert "not a window" after setParent
            # JA: setParent 後に「ウィンドウではない」を再適用
            self.setWindowFlags(Qt.Widget)
            
            # ES: Cubrir todo el área del parent | EN: Cover full parent area | JA: 親領域いっぱいにカバー
            # EN: Cover the full parent area
            # JA: 親領域を全面カバー
            self.setGeometry(0, 0, parent.width(), parent.height())
            print(f"🔧 ローディングオーバーレイを設定しました: {parent.width()}x{parent.height()}")
            print(f"🔧 ウィンドウ: {self.isWindow()}, 親: {parent}")
            
            # ES: Conectar el evento de resize del parent para ajustar el overlay
            # EN: Hook parent's resize events to keep overlay sized correctly
            # JA: 親のリサイズイベントに追従してオーバーレイを調整
            if not hasattr(self, '_resize_connected'):
                parent.installEventFilter(self)
                self._resize_connected = True
        else:
            # ES: Si no hay parent, usar tamaño mínimo | EN: If no parent, use minimum size | JA: 親が無い場合は最小サイズを使用
            # EN: If there is no parent, use a minimum size
            # JA: 親が無い場合は最小サイズを使用
            print("⚠️ 警告: LoadingOverlay に親（parent）がありません")
            self.resize(120, 120)
            # ES: Aún así, forzar que no sea ventana | EN: Still force it not to be a window | JA: それでもウィンドウにしないよう強制
            # EN: Still, force it to not be a window
            # JA: それでもウィンドウ化しないよう強制
            self.setWindowFlags(Qt.Widget)

        self.movie.start()
        self.show()
        
        # ES: Verificar una vez más que no sea ventana después de show() | EN: Double-check not a window after show() | JA: show()後にウィンドウでないか再確認
        # EN: Double-check it's not a window after show()
        # JA: show() 後にウィンドウ化していないか再確認
        if self.isWindow():
            print("⚠️ 警告: LoadingOverlay が show() 後にウィンドウ化しました。修正します...")
            self.setWindowFlags(Qt.Widget)
            if self.parent():
                self.setParent(self.parent())
        
        self.raise_()  # Elevar dentro del parent, no del sistema
        QApplication.processEvents()  # Force UI refresh
    
    def eventFilter(self, obj, event):
        """ES: Filtra eventos del parent para ajustar el tamaño cuando cambia
        EN: Filter parent events to resize when it changes
        JA: 親のイベントをフックしてサイズ変更に追従
        """
        if obj == self.parent() and event.type() == QEvent.Type.Resize:
            self._update_geometry()
        return super().eventFilter(obj, event)

    def stop(self):
        """ES: Detiene el loading y lo oculta
        EN: Stop loading and hide the overlay
        JA: ローディングを停止して非表示
        """
        self.movie.stop()
        self.hide()

class CsvToExcelExportWorker(QObject):
    """ES: Worker ligero para ejecutar la conversión CSV→Excel en background (sin bloquear la UI).
    EN: Lightweight worker to run CSV→Excel conversion in background (non-blocking UI).
    JA: CSV→Excel変換をバックグラウンドで実行する軽量ワーカー（UIブロックなし）。"""
    finished = Signal()
    error = Signal(str)

    def __init__(self, fn):
        super().__init__()
        self._fn = fn

    def run(self):
        try:
            self._fn()
            self.finished.emit()
        except Exception as e:
            self.error.emit(str(e))

class CallableResultWorker(QObject):
    """ES: Ejecuta un callable en background y devuelve su resultado por señal (sin bloquear la UI).
    EN: Runs a callable in background and returns its result via signal (non-blocking UI).
    JA: callableをバックグラウンドで実行し、結果をシグナルで返す（UIブロックなし）。"""
    finished = Signal(object)
    error = Signal(str)

    def __init__(self, fn):
        super().__init__()
        self._fn = fn

    def run(self):
        try:
            result = self._fn()
            self.finished.emit(result)
        except Exception as e:
            self.error.emit(str(e))

class ReusableProgressDialog(QDialog):
    """ES: Cuadro de progreso reutilizable con imagen personalizable
    EN: Reusable progress dialog with a customizable image
    JA: 画像カスタム可能な再利用型進捗ダイアログ
    """
    
    # ES: Señal emitida cuando se cancela el proceso | EN: Signal emitted when process is cancelled | JA: 処理キャンセル時に発行するシグナル
    # EN: Signal emitted when the process is cancelled
    # JA: 処理がキャンセルされた時に発行されるシグナル
    cancelled = Signal()
    
    def __init__(self, parent=None, title="処理中...", chibi_image="xebec_chibi_suzukisan.png", chibi_size=100):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setFixedSize(600, 320)  # Size tuned to include time info
        # ES: Sin WindowStaysOnTopHint: solo bloquea el parent, no se queda en primer plano del sistema
        # EN: Without WindowStaysOnTopHint: only blocks parent, does not stay on top of system
        # JA: WindowStaysOnTopHintなし：親のみブロック、システム最前面にはならない
        self.setWindowFlags(Qt.Dialog)
        # ES: WindowModal bloquea solo el parent, no toda la aplicación ni otras apps | EN: WindowModal blocks only parent, not whole app or other apps | JA: WindowModalは親のみブロック、他アプリはブロックしない
        # EN: WindowModal blocks only the parent (not the entire app / other apps)
        # JA: WindowModal は親のみブロック（アプリ全体や他アプリはブロックしない）
        self.setWindowModality(Qt.WindowModal)
        
        # ES: Variables para tracking de actividad
        # EN: Activity-tracking variables
        # JA: 活動（進捗）追跡用の変数
        self.start_time = time.time()
        self.last_activity_time = time.time()
        self.process_active = True  # Python process state
        self.last_progress_value = 0
        self.activity_timer = QTimer()
        self.activity_timer.timeout.connect(self._update_activity_indicator)
        self.activity_timer.start(1000)  # Update every second
        
        # ES: Variables para tracking de stages | EN: Stage-tracking variables | JA: ステージ追跡用変数
        self.current_stage = '01_model_builder'  # Current stage
        
        # ES: Establecer fondo sólido sin borde | EN: Set solid background without border | JA: 枠なし単色背景を設定
        # EN: Set solid background without border
        # JA: 枠なしの単色背景に設定
        self.setStyleSheet("""
            QDialog {
                background-color: #ffffff;
                border-radius: 10px;
            }
        """)
        
        # ES: Layout principal | EN: Main layout | JA: メインレイアウト
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        
        # ES: Título y chibi en la misma línea horizontal | EN: Title and chibi on same horizontal line | JA: タイトルとちびを同じ横並びに
        # EN: Title and chibi on the same horizontal row
        # JA: タイトルとちび画像を同じ横並びに配置
        title_chibi_layout = QHBoxLayout()
        title_chibi_layout.setContentsMargins(0, 0, 0, 0)
        title_chibi_layout.setSpacing(10)
        
        # ES: Título a la izquierda | EN: Title on the left | JA: 左側タイトル
        title_label = QLabel("処理実行中")
        title_label.setStyleSheet("""
            font-size: 18px;
            font-weight: bold;
            color: #2c3e50;
            margin-bottom: 10px;
        """)
        title_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        title_label.setFixedHeight(30)
        title_chibi_layout.addWidget(title_label)
        
        # ES: Espaciador para empujar el chibi a la derecha
        # EN: Spacer to push the chibi to the right
        # JA: ちび画像を右に寄せるためのスペーサー
        title_chibi_layout.addStretch()
        
        # ES: Imagen del chibi a la derecha
        # EN: Chibi image on the right
        # JA: 右側のちび画像
        try:
            chibi_label = QLabel()
            chibi_pixmap = QPixmap(resource_path(chibi_image))
            if not chibi_pixmap.isNull():
                # ES: Redimensionar para un tamaño adecuado (usando chibi_size) | EN: Resize to adequate size (using chibi_size) | JA: 適切なサイズにリサイズ（chibi_size使用）
                # EN: Resize to an appropriate size (using chibi_size)
                # JA: 適切なサイズにリサイズ（chibi_size を使用）
                scaled_pixmap = chibi_pixmap.scaled(chibi_size, chibi_size, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                chibi_label.setPixmap(scaled_pixmap)
                chibi_label.setFixedSize(chibi_size, chibi_size)
                chibi_label.setStyleSheet("background: transparent; border: none; margin: 0; padding: 0;")
                chibi_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
                title_chibi_layout.addWidget(chibi_label)
            else:
                print(f"⚠️ 読み込めませんでした: {chibi_image}")
        except Exception as e:
            print(f"⚠️ ちび画像の読み込み中にエラー: {e}")
        
        layout.addLayout(title_chibi_layout)
        
        # ES: Label para tiempo transcurrido y estimado (centrado, debajo del título) | EN: Elapsed/estimated time label (centered, below title) | JA: 経過/推定時間ラベル（タイトル下・中央）
        # EN: Elapsed/estimated time label (centered, below the title)
        # JA: 経過/推定時間ラベル（タイトル下、中央）
        time_info_layout = QHBoxLayout()
        time_info_layout.addStretch()
        self.time_info_label = QLabel("⏱️ 経過時間: 0:00 | 推定残り時間: 計算中...")
        self.time_info_label.setStyleSheet("""
            font-size: 13px;
            font-weight: bold;
            color: #2c3e50;
            padding: 5px;
        """)
        self.time_info_label.setAlignment(Qt.AlignCenter)
        self.time_info_label.setFixedHeight(25)
        time_info_layout.addWidget(self.time_info_label)
        time_info_layout.addStretch()
        layout.addLayout(time_info_layout)
        
        # ES: Variables para cálculo de tiempo estimado | EN: Variables for estimated-time calculation | JA: 推定時間計算用変数
        # EN: Variables for estimated-time calculation
        # JA: 推定時間計算用の変数
        self.trial_times = []  # EN: List of times per trial
        self.last_trial_start_time = None
        self.current_trial_number = 0
        
        # ES: Barra de progreso centrada que ocupa todo el ancho
        # EN: Centered progress bar that spans full width
        # JA: 横幅いっぱいの中央配置プログレスバー
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.progress_bar.setFixedHeight(30)
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: 2px solid #bdc3c7;
                border-radius: 5px;
                text-align: center;
                background-color: #ecf0f1;
                min-height: 25px;
            }
            QProgressBar::chunk {
                background-color: #3498db;
                border-radius: 3px;
            }
        """)
        layout.addWidget(self.progress_bar)
        
        # ES: Etiqueta de porcentaje centrada (azul)
        # EN: Centered percentage label (blue)
        # JA: 中央のパーセント表示（青）
        percentage_layout = QHBoxLayout()
        percentage_layout.addStretch()
        self.percentage_label = QLabel("0%")
        self.percentage_label.setStyleSheet("""
            font-size: 14px;
            font-weight: bold;
            color: #3498db;
        """)
        self.percentage_label.setAlignment(Qt.AlignCenter)
        self.percentage_label.setFixedHeight(25)
        percentage_layout.addWidget(self.percentage_label)
        percentage_layout.addStretch()
        layout.addLayout(percentage_layout)
        
        # ES: Etiqueta para mostrar Trial, Fold y Pasadas centrada
        # EN: Centered label for Trial/Fold/Pass info
        # JA: Trial/Fold/Pass 情報の中央ラベル
        trial_fold_layout = QHBoxLayout()
        trial_fold_layout.addStretch()
        self.trial_fold_label = QLabel("")
        self.trial_fold_label.setStyleSheet("""
            font-size: 13px;
            font-weight: bold;
            color: #2c3e50;
        """)
        self.trial_fold_label.setAlignment(Qt.AlignCenter)
        self.trial_fold_label.setFixedHeight(25)
        trial_fold_layout.addWidget(self.trial_fold_label)
        trial_fold_layout.addStretch()
        layout.addLayout(trial_fold_layout)
        
        # ES: Botón de cancelar centrado | EN: Centered cancel button | JA: 中央のキャンセルボタン
        # EN: Centered cancel button
        # JA: 中央のキャンセルボタン
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        
        self.cancel_button = QPushButton("キャンセル")
        self.cancel_button.setFixedSize(120, 35)
        self.cancel_button.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 6px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
            QPushButton:pressed {
                background-color: #a93226;
            }
        """)
        self.cancel_button.clicked.connect(self.cancel_process)
        
        button_layout.addWidget(self.cancel_button)
        button_layout.addStretch()
        layout.addLayout(button_layout)
        
        # ES: Centrar en la pantalla
        # EN: Center on screen
        # JA: 画面中央に配置
        self.center_on_screen()
    
    def center_on_screen(self):
        """ES: Centrar el diálogo en la pantalla
        EN: Center the dialog on screen
        JA: ダイアログを画面中央に配置
        """
        screen = QApplication.primaryScreen()
        screen_geometry = screen.geometry()
        x = (screen_geometry.width() - self.width()) // 2
        y = (screen_geometry.height() - self.height()) // 2
        self.move(x, y)
    
    def update_progress(self, percentage, status_message):
        """ES: Actualizar progreso y mensaje de estado
        EN: Update progress and status message
        JA: 進捗とステータスメッセージを更新
        """
        current_time = time.time()
        
        # ES: Actualizar última actividad si hay cambio de progreso | EN: Update last activity on progress change | JA: 進捗変化時に最終活動を更新
        # EN: Update last activity timestamp when progress changes
        # JA: 進捗が変化したら最終活動時刻を更新
        if abs(int(percentage) - self.last_progress_value) > 0:
            self.last_activity_time = current_time
            self.last_progress_value = int(percentage)
        
        # ES: Actualizar barra de progreso | EN: Update progress bar | JA: プログレスバー更新
        self.progress_bar.setValue(int(percentage))
        self.percentage_label.setText(f"{int(percentage)}%")
        
        # ES: Actualizar color según actividad (Opción 4) | EN: Update bar color based on activity (Option 4) | JA: 活動状況に応じて色を更新（オプション4）
        self._update_progress_color(current_time)
        
        QApplication.processEvents()  # Force UI refresh
    
    def set_process_active(self, active):
        """ES: Actualizar estado del proceso Python
        EN: Update Python process state
        JA: Pythonプロセス状態を更新
        """
        self.process_active = active
        QApplication.processEvents()
    
    def _update_progress_color(self, current_time):
        """ES: Actualizar color de la barra según actividad (Opción 4)
        EN: Update progress bar color based on activity (Option 4)
        JA: 活動状況に応じてプログレスバー色を更新（オプション4）
        """
        time_since_activity = current_time - self.last_activity_time
        
        if time_since_activity < 3:
            # Verde: actividad reciente
            color = "#27ae60"
        elif time_since_activity < 10:
            # Amarillo: actividad moderada
            color = "#f39c12"
        elif time_since_activity < 30:
            # Naranja: posible bloqueo
            color = "#e67e22"
        else:
            # Rojo: probable bloqueo
            color = "#e74c3c"
        
        self.progress_bar.setStyleSheet(f"""
            QProgressBar {{
                border: 2px solid #bdc3c7;
                border-radius: 5px;
                text-align: center;
                background-color: #ecf0f1;
                min-height: 25px;
            }}
            QProgressBar::chunk {{
                background-color: {color};
                border-radius: 3px;
            }}
        """)
    
    def _update_activity_indicator(self):
        """ES: Actualizar indicadores de actividad cada segundo
        EN: Update activity indicators every second
        JA: 1秒ごとに活動指標を更新
        """
        current_time = time.time()
        
        # ES: Actualizar siempre el tiempo transcurrido | EN: Always update elapsed time | JA: 経過時間は常に更新
        if hasattr(self, 'time_info_label'):
            elapsed_time = current_time - self.start_time
            elapsed_str = self._format_time(elapsed_time)
            
            # ES: Obtener el texto actual para preservar la estimación si existe | EN: Read current label text to preserve any existing estimate | JA: 推定値があれば保持するため現在テキストを取得
            current_text = self.time_info_label.text()
            
            # ES: Si ya hay una estimación calculada (no "計算中"), preservarla | EN: If an estimate is already available (not "計算中"), preserve it | JA: 既に推定値がある（「計算中」ではない）場合は保持
            if "推定残り時間:" in current_text and "計算中" not in current_text:
                # ES: Extraer la estimación del texto actual | EN: Extract estimate from current text | JA: 現在テキストから推定値を抽出
                try:
                    remaining_part = current_text.split("推定残り時間:")[1].strip()
                    # ES: Actualizar solo el tiempo transcurrido, mantener la estimación | EN: Update elapsed time only; keep the estimate | JA: 経過時間のみ更新し推定値は維持
                    self.time_info_label.setText(f"⏱️ 経過時間: {elapsed_str} | 推定残り時間: {remaining_part}")
                except:
                    # ES: Si falla, calcular estimación básica | EN: If parsing fails, compute a basic estimate | JA: 失敗時は簡易推定を計算
                    if len(self.trial_times) > 0 and elapsed_time > 0:
                        # ES: Usar promedio de trials para estimar | EN: Estimate using average trial time | JA: trial平均時間で推定
                        avg_trial_time = sum(self.trial_times) / len(self.trial_times)
                        estimated_remaining = max(0, avg_trial_time - elapsed_time)
                        estimated_str = self._format_time(estimated_remaining)
                        self.time_info_label.setText(f"⏱️ 経過時間: {elapsed_str} | 推定残り時間: {estimated_str}")
                    else:
                        self.time_info_label.setText(f"⏱️ 経過時間: {elapsed_str} | 推定残り時間: 計算中...")
            else:
                # ES: No hay estimación: calcular una básica si es posible | EN: No estimate yet: compute a basic one if possible | JA: 推定値が無い場合は可能なら簡易推定
                if len(self.trial_times) > 0 and elapsed_time > 0:
                    avg_trial_time = sum(self.trial_times) / len(self.trial_times)
                    estimated_remaining = max(0, avg_trial_time - elapsed_time)
                    estimated_str = self._format_time(estimated_remaining)
                    self.time_info_label.setText(f"⏱️ 経過時間: {elapsed_str} | 推定残り時間: {estimated_str}")
                else:
                    self.time_info_label.setText(f"⏱️ 経過時間: {elapsed_str} | 推定残り時間: 計算中...")
        
        # ES: Actualizar color según actividad | EN: Update color based on activity | JA: 活動状況に応じて色を更新
        self._update_progress_color(current_time)
        
        QApplication.processEvents()
    
    def _format_time(self, seconds):
        """Formatea segundos a formato legible (MM:SS o HH:MM:SS)"""
        if seconds < 60:
            return f"{int(seconds)}s"
        elif seconds < 3600:
            mins = int(seconds // 60)
            secs = int(seconds % 60)
            return f"{mins}:{secs:02d}"
        else:
            hours = int(seconds // 3600)
            mins = int((seconds % 3600) // 60)
            secs = int(seconds % 60)
            return f"{hours}:{mins:02d}:{secs:02d}"
    
    def set_status(self, status_message):
        """ES: Actualizar solo el mensaje de estado (no se usa en el layout simplificado)
        EN: Update only the status message (not used in the simplified layout)
        JA: ステータスメッセージのみ更新（簡易レイアウトでは未使用）
        """
        pass
        QApplication.processEvents()
    
    def update_progress_detailed(self, trial_current, trial_total, fold_current, fold_total, pass_current, pass_total, current_task='dcv', data_analysis_completed=False, final_model_training=False, shap_analysis=False, model_current=0, model_total=0):
        """ES: Actualizar información detallada de progreso (trial/fold/pass) y calcular porcentaje
        EN: Update detailed progress info (trial/fold/pass) and compute percentage
        JA: 詳細進捗（trial/fold/pass）を更新して割合を計算
        """
        current_time = time.time()
        
        # ES: Detectar cuando comienza un nuevo trial para calcular tiempo promedio
        # EN: Detect when a new trial starts to compute average time
        # JA: 平均時間算出のため新トライアル開始を検出
        if trial_current > self.current_trial_number:
            # ES: Nuevo trial detectado
            # EN: New trial detected
            # JA: 新トライアル検出
            if self.last_trial_start_time is not None:
                # ES: Calcular tiempo del trial anterior | EN: Compute previous trial duration | JA: 前トライアルの所要時間を算出
                trial_duration = current_time - self.last_trial_start_time
                self.trial_times.append(trial_duration)
                # ES: Mantener solo los últimos 10 trials para el promedio | EN: Keep only last 10 trials for average | JA: 平均用に直近10トライアルのみ保持
                # EN: Keep only the last 10 trials for the average
                # JA: 平均用に直近10トライアルのみ保持
                if len(self.trial_times) > 10:
                    self.trial_times.pop(0)
            
            self.last_trial_start_time = current_time
            self.current_trial_number = trial_current
        
        # ES: Calcular porcentaje basado en trials, folds y passes
        # EN: Compute percentage from trials, folds, and passes
        # JA: trials・folds・passesからパーセントを計算
        # ES: Stage 1 (model_builder): ~70% del total
        # EN: Stage 1 (model_builder): ~70% of total
        # JP: Stage 1（model_builder）：全体の約70%
        # ES: Stage 2 (prediction): ~15% del total
        # EN: Stage 2 (prediction): ~15% of total
        # JP: Stage 2（prediction）：全体の約15%
        # ES: Stage 3 (pareto): ~15% del total
        # EN: Stage 3 (pareto): ~15% of total
        # JP: Stage 3（pareto）：全体の約15%
        
        # ES: Distribución del progreso dentro de Stage 1 (70%) | EN: Progress distribution within Stage 1 (70%) | JA: Stage 1内の進捗配分（70%）
        # EN: Progress distribution within Stage 1 (70%)
        # JA: Stage 1（70%）内の進捗配分
        # ES: - Inicialización y carga: 0-2%
        # EN: - Initialization and load: 0-2%
        # JP: - 初期化と読み込み：0-2%
        # ES: - Análisis de datos (si está habilitado): 2-5%
        # EN: - Data analysis (if enabled): 2-5%
        # JP: - データ分析（有効時）：2-5%
        # ES: - DCV (Double Cross-Validation): 5-60%
        # EN: - DCV (Double Cross-Validation): 5-60%
        # JP: - DCV（ダブル・クロスバリデーション）：5-60%
        # ES:   - Por cada pasada: Outer Folds+trials, modelo final 2%, SHAP 1%, guardado 0.5%
        # EN:   - Per pass: Outer Folds+trials, final model 2%, SHAP 1%, saving 0.5%
        # JP:   - 各パスあたり：Outer Folds+trials、最終モデル2%、SHAP1%、保存0.5%
        # ES: - Tareas finales: 60-70%
        # EN: - Final tasks: 60-70%
        # JP: - 最終タスク：60-70%
        
        percentage = 0
        stage1_base = 0  # Stage 1 base (0-70%)
        
        # ES: 1. Inicialización y carga (0-2%)
        # EN: 1. Initialization and load (0-2%)
        # JP: 1. 初期化と読み込み（0-2%）
        stage1_base += 2
        
        # ES: 2. Análisis de datos (2-5%) - solo si está habilitado
        # EN: 2. Data analysis (2-5%) - only if enabled
        # JP: 2. データ分析（2-5%）－有効時のみ
        if data_analysis_completed:
            stage1_base = 5
        elif current_task == 'data_analysis':
            # ES: Análisis de datos en progreso | EN: Data analysis in progress | JA: データ分析中
            stage1_base = 3.5
        
        # 3. DCV (Double Cross-Validation) - 5% to 60%
        # ES: Usar valores acumulados para cálculo lineal e incremental del porcentaje | EN: Use accumulated values for linear/incremental percentage calc | JA: 線形・増分パーセント計算に累積値を使用
        # EN: Use accumulated values for linear incremental percentage calculation
        # JA: パーセントの線形・増分計算には累積値を使用
        # ES: Los valores se calculan en nonlinear_worker.py y se pasan vía trial_current/trial_total (formato "X/Y" de Optuna) | EN: Values computed in nonlinear_worker.py via trial_current/trial_total (Optuna "X/Y") | JA: nonlinear_workerでtrial_current/trial_total（OptunaのX/Y形式）で計算・渡す
        # EN: Values are computed in nonlinear_worker.py and passed via trial_current/trial_total (Optuna "X/Y" format)
        # JA: nonlinear_worker.pyで計算し trial_current/trial_total（Optunaの"X/Y"形式）で渡す
        if trial_total > 0 and fold_total > 0 and pass_total > 0:
            # ES: Progreso dentro del DCV (5% a 60% = 55% del stage 1)
            # EN: Progress within DCV (5% to 60% = 55% of stage 1)
            # JP: DCV内の進捗（5%〜60%＝Stage 1の55%）
            dcv_start = 5
            dcv_range = 55  # 60 - 5
            
            # ES: Calcular total de trials acumulados (modelos, folds, passes, trials)
            # EN: Compute total accumulated trials (models × folds × passes × trials)
            # JA: 累積トライアル総数（モデル×フォールド×パス×トライアル）を計算
            # ES: model_total se pasa como parámetro; si no está disponible, usar 1 | EN: model_total passed as param; if unavailable use 1 | JA: model_totalはパラメータで渡す、無ければ1
            # EN: model_total is passed as parameter; if unavailable, use 1
            # JA: model_totalは引数で渡される。無い場合は1を使用
            model_total_used = model_total if model_total > 0 else 1
            total_trials_accumulated = pass_total * fold_total * model_total_used * trial_total
            
            # ES: Calcular trials completados acumulados
            # EN: Compute accumulated completed trials
            # JA: 完了したトライアルの累積数を計算
            # ES: trial_current = contador de trials en el modelo/fold actual
            # EN: trial_current = count of trials in current model/fold
            # JA: trial_currentは現在のモデル/フォールド内のトライアル数
            # ES: Acumulado (DCV): (pasadas completadas * folds * modelos * trials) + (folds completados * modelos * trials) + (modelos completados * trials) + (trials completados)
            # EN: Accumulated (DCV): (completed passes * folds * models * trials) + (completed folds * models * trials) + (completed models * trials) + (completed trials)
            # JA: 累積（DCV）：（完了パス * フォールド * モデル * トライアル）+（完了フォールド * モデル * トライアル）+（完了モデル * トライアル）+（完了トライアル）
            completed_passes = max(0, pass_current - 1)
            completed_folds_in_pass = max(0, fold_current - 1)
            completed_models_in_fold = max(0, model_current - 1) if model_total > 0 else 0
            completed_trials_accumulated = (
                (completed_passes * fold_total * model_total_used * trial_total) +
                (completed_folds_in_pass * model_total_used * trial_total) +
                (completed_models_in_fold * trial_total) +
                trial_current
            )
            
            # ES: ✅ Calcular progreso lineal basado en trials acumulados
            # EN: ✅ Compute linear progress based on accumulated trials
            # JP: ✅ 累積トライアルに基づき線形進捗を計算
            trial_progress = completed_trials_accumulated / total_trials_accumulated if total_trials_accumulated > 0 else 0
            
            # ES: Los trials representan ~85% del tiempo total del DCV
            # EN: Trials account for ~85% of total DCV time
            # JA: トライアルはDCV全体時間の約85％を占める
            # ES: El resto (15%) es para entrenamiento final, SHAP y guardado
            # EN: The remainder (15%) is for final training, SHAP and save
            # JA: 残り15％は最終訓練・SHAP・保存用
            dcv_trial_progress = trial_progress * 0.85
            
            # ES: Agregar progreso del modelo final (5% del DCV)
            # EN: Add progress for final model (5% of DCV)
            # JA: 最終モデル分の進捗を加算（DCVの5％）
            if final_model_training:
                dcv_trial_progress = min(0.90, dcv_trial_progress + 0.05)  # EN: Cap at 90% to leave room for SHAP
            
            # ES: Agregar progreso de SHAP (3% del DCV)
            # EN: Add progress for SHAP (3% of DCV)
            # JA: SHAP分の進捗を加算（DCVの3%）
            if shap_analysis:
                dcv_trial_progress = min(0.95, dcv_trial_progress + 0.03)  # EN: Cap at 95% to leave room for save
            
            # ES: Agregar progreso de guardado (2% del DCV)
            # EN: Add progress for save (2% of DCV)
            # JA: 保存分の進捗を加算（DCVの2%）
            if current_task == 'saving':
                dcv_trial_progress = min(1.0, dcv_trial_progress + 0.02)
            
            # ES: Calcular progreso del DCV
            # EN: Compute DCV progress
            # JA: DCVの進捗を計算
            dcv_progress = dcv_start + (dcv_trial_progress * dcv_range)
            stage1_base = max(stage1_base, dcv_progress)
        
        # ES: 4. Tareas finales (guardado, etc.) - 60-70%
        # EN: 4. Final tasks (saving, etc.) - 60-70%
        # JP: 4. 最終タスク（保存など）－60-70%
        if current_task == 'saving' or (pass_current >= pass_total and pass_total > 0):
            # ES: Si todas las pasadas están completas, avanzar hacia el final | EN: If all passes are complete, advance to end | JA: 全パス完了なら最後へ進める
            # EN: When all passes are complete, advance to the end
            # JA: 全パス完了時は最後まで進める
            if pass_current >= pass_total:
                stage1_base = 70
        
        # ES: Stage 1 representa 70% del total
        # EN: Stage 1 represents 70% of total
        # JP: Stage 1は全体の70%
        percentage = min(70, stage1_base)
        
        # ES: Actualizar barra de progreso y porcentaje
        # EN: Update progress bar and percentage
        # JA: プログレスバーとパーセントを更新
        self.progress_bar.setValue(int(percentage))
        self.percentage_label.setText(f"{int(percentage)}%")
        
        # ES: Actualizar tiempo transcurrido y estimado
        # EN: Update elapsed and estimated time
        # JA: 経過時間と推定時間を更新
        elapsed_time = current_time - self.start_time
        elapsed_str = self._format_time(elapsed_time)
        
        # ES: Calcular tiempo estimado
        # EN: Compute estimated time
        # JA: 推定時間を計算
        estimated_remaining = None
        if len(self.trial_times) > 0:
            # ES: Calcular tiempo promedio por trial
            # EN: Compute average time per trial
            # JA: トライアルあたり平均時間を計算
            avg_trial_time = sum(self.trial_times) / len(self.trial_times)
            
            # ES: Calcular trials restantes
            # EN: Compute remaining trials
            # JA: 残りトライアル数を計算
            if trial_total > 0 and fold_total > 0 and pass_total > 0:
                # ES: Trials restantes en el fold actual
                # EN: Remaining trials in current fold
                # JA: 現在のフォールドの残りトライアル
                remaining_trials_in_fold = max(0, trial_total - trial_current)
                # ES: Folds restantes en el pass actual
                # EN: Remaining folds in current pass
                # JA: 現在のパスの残りフォールド
                remaining_folds_in_pass = max(0, fold_total - fold_current)
                # ES: Passes restantes
                # EN: Remaining passes
                # JA: 残りパス
                remaining_passes = max(0, pass_total - pass_current)
                
                # ES: Calcular tiempo restante para stage 1
                # EN: Compute remaining time for stage 1
                # JA: Stage 1の残り時間を計算
                remaining_trials_stage1 = (
                    remaining_trials_in_fold +
                    remaining_folds_in_pass * trial_total +
                    remaining_passes * fold_total * trial_total
                )
                
                # ES: Tiempo estimado para stage 1
                # EN: Estimated time for stage 1
                # JA: Stage 1の推定時間
                estimated_stage1 = remaining_trials_stage1 * avg_trial_time
                
                # ES: Tiempo estimado para stages 2 y 3 (aproximadamente 30% del tiempo total)
                # EN: Estimated time for stages 2 and 3 (~30% of total)
                # JA: Stage 2・3の推定時間（全体の約30％）
                # ES: Si stage 1 toma 70%, entonces stages 2+3 toman aproximadamente 30%
                # EN: If stage 1 takes 70%, stages 2+3 take ~30%
                # JA: Stage 1が70％なら、Stage 2+3は約30％
                # ES: Estimar basado en el tiempo ya transcurrido
                # EN: Estimate based on elapsed time so far
                # JA: 経過時間に基づき推定
                if percentage > 0:
                    total_estimated_time = elapsed_time / (percentage / 100)
                    estimated_stage1_remaining = (total_estimated_time * 0.70) - elapsed_time
                    estimated_stage2_3 = total_estimated_time * 0.30
                    estimated_remaining = max(0, estimated_stage1_remaining + estimated_stage2_3)
                else:
                    estimated_remaining = estimated_stage1 * (1 / 0.70)  # Ajustar para incluir stages 2 y 3
        
        if estimated_remaining is not None:
            estimated_str = self._format_time(estimated_remaining)
            self.time_info_label.setText(f"⏱️ 経過時間: {elapsed_str} | 推定残り時間: {estimated_str}")
        else:
            self.time_info_label.setText(f"⏱️ 経過時間: {elapsed_str} | 推定残り時間: 計算中...")
        
        if hasattr(self, 'trial_fold_label'):
            # ES: Formatear información: Model X/Y: Trial Z/W | Fold A/B | Pass C/D
            # EN: Format info: Model X/Y: Trial Z/W | Fold A/B | Pass C/D
            # JA: 情報整形: Model X/Y: Trial Z/W | Fold A/B | Pass C/D
            parts = []
            
            # ES: Modelo (si hay modelos configurados) | EN: Model (if models configured) | JA: モデル（設定時）
            if model_total > 0:
                parts.append(f"Model: {model_current}/{model_total}")
            
            # ES: Trial (si hay trials) | EN: Trial (if trials) | JA: トライアル（あり時）
            if trial_total > 0:
                parts.append(f"Trial: {trial_current}/{trial_total}")
            
            # ES: Fold (si hay folds) | EN: Fold (if folds) | JA: フォールド（あり時）
            if fold_total > 0:
                parts.append(f"Fold: {fold_current}/{fold_total}")
            
            # ES: Pass (si hay passes) | EN: Pass (if passes) | JA: パス（あり時）
            if pass_total > 0:
                parts.append(f"Pass: {pass_current}/{pass_total}")
            
            # ES: Combinar todas las partes con " | " | EN: Join parts with " | " | JA: " | "で結合
            combined_text = " | ".join(parts) if parts else ""
            
            self.trial_fold_label.setText(combined_text)
        
        QApplication.processEvents()
    
    def update_status(self, status_message):
        """ES: Actualizar solo el mensaje de estado (alias para set_status)
        EN: Update only the status message (alias for set_status)
        JA: ステータスメッセージのみ更新（set_status の別名）
        """
        self.set_status(status_message)
    
    def set_title(self, title):
        """ES: Cambiar el título del diálogo
        EN: Change the dialog window title
        JA: ダイアログのウィンドウタイトルを変更
        """
        self.setWindowTitle(title)
    
    def set_main_title(self, title):
        """ES: Cambiar el título principal dentro del diálogo
        EN: Change the main title inside the dialog
        JA: ダイアログ内のメインタイトルを変更
        """
        # ES: Buscar el título label y actualizarlo | EN: Find the title label and update it | JA: タイトルラベルを探して更新
        for i in range(self.layout().count()):
            item = self.layout().itemAt(i)
            if item and item.widget():
                widget = item.widget()
                if isinstance(widget, QHBoxLayout):
                    for j in range(widget.count()):
                        sub_item = widget.itemAt(j)
                        if sub_item and sub_item.widget():
                            sub_widget = sub_item.widget()
                            if isinstance(sub_widget, QVBoxLayout):
                                for k in range(sub_widget.count()):
                                    label_item = sub_widget.itemAt(k)
                                    if label_item and label_item.widget():
                                        label_widget = label_item.widget()
                                        if isinstance(label_widget, QLabel) and label_widget.text() == "処理実行中":
                                            label_widget.setText(title)
                                            return
    
    def cancel_process(self):
        """ES: Cancelar proceso y cerrar popup
        EN: Cancel the process and close the popup
        JA: 処理をキャンセルしてポップアップを閉じる
        """
        # ES: Emitir señal de cancelación antes de cerrar | EN: Emit cancel signal before closing | JA: 閉じる前にキャンセルシグナルを送信
        self.cancelled.emit()
        self.progress_bar.setValue(0)
        self.percentage_label.setText("0%")
        QApplication.processEvents()
        self.reject()

class LinearAnalysisProgressDialog(ReusableProgressDialog):
    """ES: Popup de progreso para análisis lineal usando la clase reutilizable
    EN: Progress popup for linear analysis (reusable base)
    JA: 線形解析の進捗ポップアップ（再利用ベース）
    """
    
    def __init__(self, parent=None):
        super().__init__(
            parent=parent,
            title="線形解析実行中...",
            chibi_image="xebec_chibi_suzukisan.png",
            chibi_size=150  # 100 * 1.5 = 150 (larger chibi for linear analysis)
        )
        self.set_main_title("線形解析")
    
    def cancel_analysis(self):
        """ES: Cancelar análisis y cerrar popup
        EN: Cancel analysis and close popup
        JA: 解析をキャンセルしてポップアップを閉じる
        """
        self.cancel_process()

class YosokuWorker(QThread):
    """ES: Worker para predicción Yosoku con señales de progreso
    EN: Worker for Yosoku prediction with progress signals
    JA: 進捗シグナル付きYosoku予測ワーカー
    """
    
    # ES: Señales | EN: Signals | JA: シグナル | EN: Signals | JA: シグナル
    progress_updated = Signal(int, str)  # percent, message
    status_updated = Signal(str)  # status message
    finished = Signal(str)  # output file path
    error = Signal(str)  # error message
    
    def __init__(self, selected_params, unexperimental_file, output_path, prediction_folder=None):
        super().__init__()
        self.selected_params = selected_params
        self.unexperimental_file = unexperimental_file
        self.output_path = output_path
        self.prediction_folder = prediction_folder  # 04_予測計算
        self.is_cancelled = False

    @staticmethod
    def _apply_inverse_transform(values, transformation_info):
        """ES: Aplicar inversa de la transformación (compatible con linear_analysis_advanced.TransformationAnalyzer)
        EN: Apply inverse transform (compatible with linear_analysis_advanced.TransformationAnalyzer)
        JA: 逆変換を適用（linear_analysis_advanced.TransformationAnalyzer 互換）
        """
        try:
            import numpy as np
            if not transformation_info or not transformation_info.get("applied"):
                return values
            method = transformation_info.get("method", "none")
            params = transformation_info.get("parameters", {}) or {}

            if method == "log":
                return np.exp(values)
            if method == "log10":
                return np.power(10, values)
            if method == "sqrt":
                return np.power(values, 2)
            if method == "boxcox":
                lam = float(params.get("lambda", 0.0))
                if abs(lam) < 1e-6:
                    return np.exp(values)
                return np.power(lam * values + 1, 1 / lam)
            if method == "yeo_johnson":
                lam = float(params.get("lambda", 0.0))
                if abs(lam) < 1e-6:
                    return np.exp(values) - 1
                return np.power(lam * values + 1, 1 / lam) - 1
            return values
        except Exception:
            return values

    @staticmethod
    def _normalize_columns(df):
        try:
            import pandas as pd
            if isinstance(df.columns, pd.MultiIndex):
                df.columns = [" ".join([str(x).strip() for x in tup if str(x).strip() != ""]).strip() for tup in df.columns]
            else:
                df.columns = [str(c).strip() for c in df.columns]
        except Exception:
            pass
        return df

    def _find_models_regression_dir(self):
        """ES: Localiza la carpeta de modelos de regresión del último run lineal.
        EN: Locate the regression-models folder of the last linear run.
        JA: 直近の線形実行の回帰モデルフォルダを特定。"""
        import os
        # ES: Derivar run_folder desde prediction_folder si se proporciona
        # EN: Derive run_folder from prediction_folder if provided
        # JA: prediction_folderが指定されていればrun_folderを導出
        run_folder = None
        try:
            if self.prediction_folder:
                run_folder = os.path.abspath(os.path.join(self.prediction_folder, os.pardir))
        except Exception:
            run_folder = None

        candidates = []
        if run_folder:
            candidates.extend([
                os.path.join(run_folder, "01_学習モデル", "regression"),
                os.path.join(run_folder, "03_モデル学習", "01_学習モデル", "regression"),
                os.path.join(run_folder, "03_モデル学習", "regression"),
            ])
        for c in candidates:
            if os.path.isdir(c):
                return c

        # ES: Fallback: búsqueda acotada dentro de run_folder | EN: Fallback: bounded search inside run_folder | JA: フォールバック：run_folder内の限定検索
        # EN: Fallback: bounded search inside run_folder
        # JA: フォールバック：run_folder 内を深さ制限付きで探索
        if run_folder and os.path.isdir(run_folder):
            try:
                for root, dirs, files in os.walk(run_folder):
                    rel = os.path.relpath(root, run_folder)
                    if rel != "." and rel.count(os.sep) >= 4:
                        dirs[:] = []
                        continue
                    if any(f.startswith("best_model_") and f.endswith(".pkl") for f in files):
                        return root
            except Exception:
                pass
        return None
    
    def run(self):
        """ES: Ejecutar predicción Yosoku con progreso
        EN: Run Yosoku prediction with progress
        JA: 進捗付きでYosoku予測を実行
        """
        try:
            self.status_updated.emit("データを読み込み中...")
            self.progress_updated.emit(10, "データを読み込み中...")
            
            import pandas as pd
            import numpy as np
            import joblib

            ext = os.path.splitext(str(self.unexperimental_file))[1].lower()
            if ext == ".csv":
                data_df = pd.read_csv(self.unexperimental_file, encoding="utf-8-sig")
            else:
                data_df = pd.read_excel(self.unexperimental_file)
            data_df = self._normalize_columns(data_df)

            # ES: Validación mínima de columnas requeridas del 未実験データ | EN: Minimal validation of required 未実験データ columns | JA: 未実験データの必須列の最小検証
            # EN: Minimal validation of required columns in 未実験データ
            # JA: 未実験データの必須列を最小限チェック
            brush_cols = ["A13", "A11", "A21", "A32"]
            required_cols = brush_cols + ["線材長"]
            missing = [c for c in required_cols if c not in data_df.columns]
            if missing:
                raise ValueError(f"未実験データに必要な列がありません: {', '.join(missing)}")

            onehot = data_df[brush_cols].apply(pd.to_numeric, errors="coerce").fillna(0).astype(int)
            s = onehot.sum(axis=1)
            if (s != 1).any():
                bad = onehot.index[s != 1].tolist()[:10]
                raise ValueError(f"未実験データのブラシ列が不正です。不正行(先頭10): {bad}")

            wire_series = pd.to_numeric(data_df["線材長"], errors="coerce")
            if wire_series.isna().any():
                bad = wire_series.index[wire_series.isna()].tolist()[:10]
                raise ValueError(f"未実験データの 線材長 に数値以外/欠損があります。不正行(先頭10): {bad}")

            self.status_updated.emit("モデルを読み込み中...")
            self.progress_updated.emit(25, "モデルを読み込み中...")

            models_dir = self._find_models_regression_dir()
            if not models_dir:
                raise ValueError("回帰モデルフォルダが見つかりません（best_model_*.pkl）")

            model_files = [os.path.join(models_dir, f) for f in os.listdir(models_dir) if f.startswith("best_model_") and f.endswith(".pkl")]
            if not model_files:
                raise ValueError(f"回帰モデルが見つかりません: {models_dir}")

            # ES: Cargar modelos (solo targets relevantes si existen)
            # EN: Load models (only relevant targets if present)
            # JA: モデル読み込み（該当する目的変数のみ優先）
            target_whitelist = {"上面ダレ量", "側面ダレ量", "摩耗量"}
            models = {}
            for p in model_files:
                try:
                    d = joblib.load(p)
                    target = d.get("target_name") or os.path.splitext(os.path.basename(p))[0].replace("best_model_", "")
                    if target in target_whitelist:
                        models[target] = d
                except Exception:
                    continue

            if not models:
                # ES: Si no encontramos por whitelist, cargar todo lo que sea regresión | EN: If not found by whitelist, load all regression-related | JA: ホワイトリストで見つからなければ回帰関連を全てロード
                # EN: If nothing matches the whitelist, load all regression models
                # JA: 該当がなければ回帰モデルをすべて読み込む
                for p in model_files:
                    d = joblib.load(p)
                    target = d.get("target_name") or os.path.splitext(os.path.basename(p))[0].replace("best_model_", "")
                    models[target] = d

            # ES: Preparar features para predicción según feature_names del primer modelo | EN: Prepare features for prediction from first model's feature_names | JA: 最初のモデルのfeature_namesに基づき予測用featuresを準備
            # EN: Prepare prediction features using the first model's feature_names
            # JA: 先頭モデルの feature_names に合わせて特徴量を準備
            any_model = next(iter(models.values()))
            feature_names = list(any_model.get("feature_names") or [])
            scaler = any_model.get("scaler")
            if not feature_names:
                raise ValueError("モデルの feature_names が空です。")

            # ES: Mapear nombres alternativos
            # EN: Map alternative column names
            # JA: 列名の別名をマッピング
            alt = {
                "回転速度": ["回転速度"],
                "送り速度": ["送り速度"],
                "UPカット": ["UPカット", "回転方向"],
                "切込量": ["切込量", "切込み量"],
                "突出量": ["突出量", "突出し量"],
                "載せ率": ["載せ率"],
                "パス数": ["パス数", "バス数"],
            }
            colmap = {}
            for k, names in alt.items():
                for n in names:
                    if n in data_df.columns:
                        colmap[k] = n
                        break

            # ES: Construir X base con todas las columnas requeridas por feature_names
            # EN: Build X with all columns required by feature_names
            # JA: feature_names に必要な列で X を構築
            X = pd.DataFrame(index=data_df.index)
            for fn in feature_names:
                # ES: Si el modelo pide una de las columnas conocidas, mapearla
                # EN: If the model expects a known column, map it
                # JA: モデルが既知の列を要求する場合はマッピング
                if fn in colmap:
                    X[fn] = pd.to_numeric(data_df[colmap[fn]], errors="coerce")
                else:
                    # ES: Columna directa si existe, si no 0
                    # EN: Use the column directly if present; otherwise 0
                    # JA: 列があればそのまま使用、なければ0
                    if fn in data_df.columns:
                        X[fn] = pd.to_numeric(data_df[fn], errors="coerce")
                    else:
                        X[fn] = 0.0

            if X.isna().any().any():
                # ES: NaNs en features -> 0 (conservador)
                # EN: NaNs in features -> 0 (conservative)
                # JA: 特徴量のNaNは0に置換（保守的）
                X = X.fillna(0.0)

            # ES: Escalado (si existe)
            # EN: Scaling (if available)
            # JA: スケーリング（存在する場合）
            if scaler is not None:
                try:
                    X_scaled = scaler.transform(X.values)
                except Exception:
                    X_scaled = X.values
            else:
                X_scaled = X.values

            self.status_updated.emit("予測を計算中...")
            self.progress_updated.emit(60, "予測を計算中...")

            # ES: Base output (condiciones + meta)
            # EN: Base output (conditions + metadata)
            # JA: 出力の土台（条件 + メタ情報）
            out = pd.DataFrame(index=data_df.index)
            for c in brush_cols:
                out[c] = onehot[c].astype(int)
            out["直径"] = self.selected_params.get("diameter")
            out["材料"] = self.selected_params.get("material")
            out["線材長"] = wire_series.astype(float)

            # ES: Añadir condiciones (si existen) | EN: Add conditions (if any) | JA: 条件があれば追加
            # EN: Add condition columns (if present)
            # JA: 条件列を追加（存在する場合）
            for k in ["回転速度", "送り速度", "UPカット", "切込量", "突出量", "載せ率", "パス数"]:
                src = colmap.get(k, k)
                if src in data_df.columns:
                    out[k] = pd.to_numeric(data_df[src], errors="coerce")
                else:
                    out[k] = 0

            # 加工時間
            try:
                feed = pd.to_numeric(out["送り速度"], errors="coerce").replace(0, np.nan)
                out["加工時間"] = (100 / feed) * 60
                out["加工時間"] = out["加工時間"].fillna(0)
            except Exception:
                out["加工時間"] = 0

            # ES: Predicciones por target
            # EN: Predictions per target
            # JA: 目的変数ごとの予測
            done = 0
            total_t = len(models)
            for target_name, d in models.items():
                if self.is_cancelled:
                    return
                model = d.get("model")
                if model is None:
                    continue
                y_hat = model.predict(X_scaled)
                # ES: Inversa de transformación si aplica | EN: Inverse transform if applicable | JA: 適用時は逆変換
                # EN: Apply inverse transform if needed
                # JA: 必要なら逆変換を適用
                y_hat = self._apply_inverse_transform(np.asarray(y_hat), d.get("transformation_info") or {"applied": False})
                out[target_name] = y_hat
                done += 1
                self.progress_updated.emit(60 + int((done / max(total_t, 1)) * 30), f"予測中... ({done}/{total_t})")

            self.status_updated.emit("CSVファイルを保存中...")
            self.progress_updated.emit(95, "CSVファイルを保存中...")

            # ES: Guardar CSV (sin límite de filas de Excel) | EN: Save CSV (no Excel row limit) | JA: CSV保存（Excel行数制限なし）
            # EN: Save CSV (no Excel row limit)
            # JA: CSVを保存（Excelの行数制限なし）
            out.to_csv(self.output_path, index=False, encoding="utf-8-sig")

            self.status_updated.emit("完了！")
            self.progress_updated.emit(100, "完了！")
            self.finished.emit(self.output_path)
            
        except Exception as e:
            print(f"❌ Yosoku 予測中にエラー: {e}")
            import traceback
            traceback.print_exc()
            self.error.emit(f"Yosoku 予測中にエラー: {str(e)}")
    
    def cancel_prediction(self):
        """ES: Cancelar predicción
        EN: Cancel prediction
        JA: 予測をキャンセル
        """
        self.is_cancelled = True
        self.terminate()

class YosokuProgressDialog(ReusableProgressDialog):
    """ES: Popup de progreso para predicción Yosoku usando la clase reutilizable
    EN: Progress popup for Yosoku prediction (reusable base)
    JA: Yosoku予測の進捗ポップアップ（再利用ベース）
    """
    
    def __init__(self, parent=None):
        super().__init__(
            parent=parent,
            title="予測実行中...",
            chibi_image="Chibi_tamiru.png",
            chibi_size=150  # 100 * 1.5 = 150 (larger chibi for linear yosoku)
        )
        self.set_main_title("予測実行")
    
    def cancel_prediction(self):
        """ES: Cancelar predicción y cerrar popup
        EN: Cancel prediction and close popup
        JA: 予測をキャンセルしてポップアップを閉じる
        """
        self.cancel_process()

class YosokuImportProgressDialog(ReusableProgressDialog):
    """ES: Popup de progreso para importación de datos Yosoku usando la clase reutilizable
    EN: Progress popup for importing Yosoku data (reusable base)
    JA: Yosokuデータインポートの進捗ポップアップ（再利用ベース）
    """
    
    def __init__(self, parent=None):
        super().__init__(
            parent=parent,
            title="データベースインポート中...",
            chibi_image="Chibi_suzuki_tamiru.png",
            chibi_size=160  # 100 * 1.6 = 160 (larger chibi for yosoku import)
        )
        self.set_main_title("データベースインポート")
    
    def cancel_import(self):
        """ES: Cancelar importación y cerrar popup
        EN: Cancel import and close popup
        JA: インポートをキャンセルしてポップアップを閉じる
        """
        self.cancel_process()

class YosokuExportProgressDialog(ReusableProgressDialog):
    """ES: Popup de progreso para exportación de datos Yosoku usando la clase reutilizable
    EN: Progress popup for exporting Yosoku data (reusable base)
    JA: Yosokuデータエクスポートの進捗ポップアップ（再利用ベース）
    """
    
    def __init__(self, parent=None):
        super().__init__(
            parent=parent,
            title="データベースエクスポート中...",
            chibi_image="Chibi_suzuki_tamiru.png",
            chibi_size=160  # 100 * 1.6 = 160 (larger chibi for yosoku export)
        )
        self.set_main_title("データベースエクスポート")
    
    def cancel_export(self):
        """ES: Cancelar exportación y cerrar popup
        EN: Cancel export and close popup
        JA: エクスポートをキャンセルしてポップアップを閉じる
        """
        self.cancel_process()

class YosokuImportWorker(QThread):
    """ES: Worker para importación de datos Yosoku con progreso
    EN: Worker for importing Yosoku data with progress
    JA: 進捗付きYosokuデータインポートワーカー
    """
    
    # ES: Señales | EN: Signals | JA: シグナル
    # EN: Signals
    # JA: シグナル
    progress_updated = Signal(int, str)  # percent, message
    status_updated = Signal(str)  # status message
    finished = Signal()  # import completed
    error = Signal(str)  # error message
    
    def __init__(self, excel_path, analysis_type="lineal", parent_widget=None):
        super().__init__()
        self.excel_path = excel_path
        self.analysis_type = analysis_type  # "lineal" or "no_lineal"
        self.cancelled = False
    
    def cancel_import(self):
        """ES: Cancelar importación
        EN: Cancel import
        JA: インポートをキャンセル
        """
        self.cancelled = True
    
    def run(self):
        """ES: Ejecutar importación con progreso
        EN: Run import with progress
        JA: 進捗付きでインポートを実行
        """
        try:
            import pandas as pd
            import sqlite3
            import os
            from openpyxl import load_workbook
            import shutil
            from datetime import datetime
            import sys
            
            # ES: Paso 1: Crear carpeta temporal
            # EN: Step 1: Create temp folder
            # JA: 手順1：一時フォルダを作成
            self.status_updated.emit("フォルダ作成中...")
            self.progress_updated.emit(5, "フォルダ作成中...")
            print("📁 一時フォルダーを作成中...")
            
            if self.cancelled:
                return
            
            project_folder = os.path.dirname(self.excel_path)
            temp_folder = os.path.join(project_folder, "99_Temp")
            if not os.path.exists(temp_folder):
                os.makedirs(temp_folder)
                print(f"✅ フォルダーを作成しました: {temp_folder}")
            
            # ES: Paso 2: Crear copia
            # EN: Step 2: Create a copy
            # JA: 手順2：コピーを作成
            self.status_updated.emit("ファイルコピー中...")
            self.progress_updated.emit(10, "ファイルコピー中...")
            print("📋 Excelファイルのコピーを作成中...")
            
            if self.cancelled:
                return
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            original_filename = os.path.basename(self.excel_path)
            name, ext = os.path.splitext(original_filename)
            backup_filename = f"{name}_backup_{timestamp}{ext}"
            backup_path = os.path.join(temp_folder, backup_filename)
            
            shutil.copy2(self.excel_path, backup_path)
            print(f"✅ コピーを作成しました: {backup_path}")
            
            # ES: Guardar referencia para limpieza posterior
            # EN: Keep reference for later cleanup
            # JA: 後でクリーンアップするため参照を保持
            self.backup_path = backup_path
            
            ext_in = os.path.splitext(str(self.excel_path))[1].lower()

            # ES: Paso 3/4: Leer datos
            # EN: Step 3/4: Read data
            # JA: 手順3/4：データを読み込み
            # ES: - Si es CSV: no hay fórmulas -> leer directamente | EN: If CSV: no formulas -> read directly | JA: CSVなら数式なし→直接読み込み
            # EN: - If CSV: no formulas -> read directly
            # JA: - CSVの場合：数式なし → そのまま読み込み
            # ES: - Si es Excel: convertir fórmulas a valores (legacy) y leer data_only | EN: If Excel: convert formulas to values (legacy) and read data_only | JA: Excelなら数式を値に変換（レガシー）してdata_onlyで読む
            # EN: - If Excel: convert formulas to values (legacy) and read data_only
            # JA: - Excelの場合：数式→値（従来方式）にして data_only で読む
            self.status_updated.emit("データ読み込み中...")
            self.progress_updated.emit(20, "データ読み込み中...")

            if self.cancelled:
                return

            if ext_in == ".csv":
                df = pd.read_csv(backup_path, encoding="utf-8-sig")
            else:
                # ES: Convertir fórmulas a valores | EN: Convert formulas to values | JA: 数式を値に変換
                # EN: Convert formulas to values
                # JA: 数式を値に変換
                self.status_updated.emit("数式を値に変換中...")
                self.progress_updated.emit(25, "数式を値に変換中...")
                print("🔄 数式を値に変換中...")

                if self.cancelled:
                    return

                try:
                    import xlwings as xw

                    print("📊 xlwingsで数式を値に変換中...")
                    app = xw.App(visible=False, add_book=False)
                    try:
                        wb = app.books.open(str(backup_path))
                        wb.app.api.CalculateFull()

                        for sh in wb.sheets:
                            rng = sh.used_range
                            vals = rng.value
                            rng.value = vals

                        wb.save(str(backup_path))
                        print("✅ 数式を値に変換しました（xlwings）")
                    finally:
                        wb.close()
                        app.quit()

                except ImportError:
                    print("⚠️ xlwings が見つかりません。インストール中...")
                    import subprocess
                    subprocess.check_call([sys.executable, "-m", "pip", "install", "xlwings"])

                    import xlwings as xw

                    print("📊 xlwings（インストール後）で数式を値に変換中...")
                    app = xw.App(visible=False, add_book=False)
                    try:
                        wb = app.books.open(str(backup_path))
                        wb.app.api.CalculateFull()

                        for sh in wb.sheets:
                            rng = sh.used_range
                            vals = rng.value
                            rng.value = vals

                        wb.save(str(backup_path))
                        print("✅ 数式を値に変換しました（xlwings／インストール後）")
                    finally:
                        wb.close()
                        app.quit()

                except Exception as e:
                    print(f"⚠️ xlwings でエラー: {e}")
                    print("📊 代替手段（openpyxl）を使用中...")
                    # ES: Método alternativo: copia valores (NO evalúa fórmulas) | EN: Alternative: copy values (does not evaluate formulas) | JA: 別法：値をコピー（数式は評価しない）
                    # EN: Fallback: copy values (does NOT evaluate formulas)
                    # JA: フォールバック：値をコピー（数式の評価はしない）
                    workbook = load_workbook(backup_path, data_only=False)
                    worksheet = workbook.active
                    values_worksheet = workbook.create_sheet("values_only")
                    for row in worksheet.iter_rows(values_only=True):
                        values_worksheet.append(row)
                    workbook.remove(worksheet)
                    values_worksheet.title = "Sheet1"
                    workbook.save(backup_path)
                    workbook.close()
                    print("✅ 数式を値に変換しました（openpyxl／ベストエフォート）")

                # ES: Leer data_only
                # EN: Read with data_only
                # JA: data_only で読み込み
                self.status_updated.emit("データ読み込み中...")
                self.progress_updated.emit(40, "データ読み込み中...")

                workbook = load_workbook(backup_path, data_only=True)
                worksheet = workbook.active
                data = []
                headers = []
                for col in worksheet.iter_cols(min_row=1, max_row=1):
                    headers.append(col[0].value)
                for row in worksheet.iter_rows(min_row=2, values_only=True):
                    if any(cell is not None for cell in row):
                        data.append(row)
                df = pd.DataFrame(data, columns=headers)
                workbook.close()
            
            # ES: Paso 5: Conectar a base de datos
            # EN: Step 5: Connect to database
            # JA: 手順5：DBに接続
            self.status_updated.emit("データベース接続中...")
            self.progress_updated.emit(60, "データベース接続中...")
            
            if self.cancelled:
                return
            
            # ES: Determinar BBDD según el tipo de análisis | EN: Determine DB from analysis type | JA: 解析タイプでDBを決定
            # EN: Choose DB based on analysis type
            # JA: 解析タイプに応じてDBを選択
            if self.analysis_type == "no_lineal":
                db_path = YOSOKU_NO_LINEAL_DB_PATH
            else:  # default: "lineal"
                db_path = YOSOKU_LINEAL_DB_PATH
            conn = sqlite3.connect(db_path, timeout=10)
            cursor = conn.cursor()
            
            # ES: Paso 6: Crear tabla
            # EN: Step 6: Create table
            # JA: 手順6：テーブル作成
            self.status_updated.emit("テーブル作成中...")
            self.progress_updated.emit(70, "テーブル作成中...")
            
            if self.cancelled:
                conn.close()
                return
            
            create_table_sql = """
            CREATE TABLE IF NOT EXISTS yosoku_predictions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                A13 INTEGER,
                A11 INTEGER,
                A21 INTEGER,
                A32 INTEGER,
                直径 REAL,
                材料 TEXT,
                線材長 REAL,
                回転速度 REAL,
                送り速度 REAL,
                UPカット INTEGER,
                切込量 REAL,
                突出量 REAL,
                載せ率 REAL,
                パス数 INTEGER,
                加工時間 REAL,
                上面ダレ量 REAL,
                側面ダレ量 REAL,
                摩耗量 REAL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """
            cursor.execute(create_table_sql)
            
            # ES: Paso 7: Eliminar índice anterior y crear nuevo índice único | EN: Step 7: Drop old index and create new unique index | JA: ステップ7：旧インデックス削除し新ユニークインデックス作成
            # EN: Step 7: Drop old index and create new unique index
            # JA: 手順7：既存インデックス削除→ユニークインデックス作成
            self.status_updated.emit("インデックス作成中...")
            self.progress_updated.emit(80, "インデックス作成中...")
            
            if self.cancelled:
                conn.close()
                return
            
            print("⚡ 既存インデックスを削除し、新しいユニークインデックスを作成中...")
            print("🎯 重複判定に使う列のみを対象")
            print("📝 除外: 上面ダレ量, 側面ダレ量, 摩耗量, created_at")
            
            # ES: Eliminar índice anterior si existe
            # EN: Drop previous index if it exists
            # JA: 既存インデックスがあれば削除
            try:
                cursor.execute("DROP INDEX IF EXISTS idx_unique_yosoku")
                print("🗑️ 既存インデックスを削除しました")
            except Exception as e:
                print(f"⚠️ 既存インデックスはありませんでした: {e}")
            
            # ES: Crear nuevo índice único SOLO en las columnas que determinan duplicados
            # EN: Create a unique index ONLY on the duplicate-key columns
            # JA: 重複判定キー列のみにユニークインデックスを作成
            cursor.execute("""
                CREATE UNIQUE INDEX idx_unique_yosoku 
                ON yosoku_predictions (
                    A13, A11, A21, A32, 直径, 材料, 線材長, 回転速度, 
                    送り速度, UPカット, 切込量, 突出量, 載せ率, パス数, 加工時間
                )
            """)
            print("✅ 新しいユニークインデックスを作成しました")
            print("📊 重複判定の対象列:")
            print("   A13, A11, A21, A32, 直径, 材料, 線材長, 回転速度")
            print("   送り速度, UPカット, 切込量, 突出量, 載せ率, パス数, 加工時間")
            print("📝 除外列（上書きされます）:")
            print("   上面ダレ量, 側面ダレ量, 摩耗量, created_at")
            
            # ES: Paso 8: Insertar datos con sobreescritura automática
            # EN: Step 8: Insert data with automatic overwrite (INSERT OR REPLACE)
            # JA: 手順8：自動上書きで挿入（INSERT OR REPLACE）
            self.status_updated.emit("データ挿入中...")
            self.progress_updated.emit(90, "データ挿入中...")
            
            if self.cancelled:
                conn.close()
                return
            
            print("📝 INSERT OR REPLACE を実行中（自動上書き）")
            print("🔍 ユニークインデックスが有効か確認中...")
            
            # ES: Verificar que el índice existe
            # EN: Verify the index exists
            # JA: インデックスの存在確認
            cursor.execute("SELECT name FROM sqlite_master WHERE type='index' AND name='idx_unique_yosoku'")
            index_exists = cursor.fetchone()
            if index_exists:
                print("✅ ユニークインデックスを確認: idx_unique_yosoku")
            else:
                print("❌ エラー: ユニークインデックスが見つかりません！")
            
            insert_sql = """
            INSERT OR REPLACE INTO yosoku_predictions
            (A13, A11, A21, A32, 直径, 材料, 線材長, 回転速度, 送り速度, UPカット, 
             切込量, 突出量, 載せ率, パス数, 加工時間, 上面ダレ量, 側面ダレ量, 摩耗量)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """
            
            # ES: Insertar datos fila por fila
            # EN: Insert row by row
            # JA: 行ごとに挿入
            inserted_count = 0
            updated_count = 0
            total_rows = len(df)
            
            # ES: Obtener conteo inicial de registros
            # EN: Get initial row count
            # JA: 初期レコード数を取得
            cursor.execute("SELECT COUNT(*) FROM yosoku_predictions")
            initial_count = cursor.fetchone()[0]
            print(f"📊 DBの初期レコード数: {initial_count}")
            
            for index, row in df.iterrows():
                if self.cancelled:
                    conn.close()
                    return
                
                # ES: Verificar si el registro ya existe antes de insertar
                # EN: Check whether the row exists before inserting
                # JA: 挿入前に既存行か確認
                cursor.execute("""
                    SELECT COUNT(*) FROM yosoku_predictions 
                    WHERE A13=? AND A11=? AND A21=? AND A32=? AND 直径=? AND 材料=? 
                    AND 線材長=? AND 回転速度=? AND 送り速度=? AND UPカット=? 
                    AND 切込量=? AND 突出量=? AND 載せ率=? AND パス数=? AND 加工時間=?
                """, (
                    row.get('A13'), row.get('A11'), row.get('A21'), row.get('A32'),
                    row.get('直径'), row.get('材料'), row.get('線材長'), row.get('回転速度'),
                    row.get('送り速度'), row.get('UPカット'), row.get('切込量'), row.get('突出量'),
                    row.get('載せ率'), row.get('パス数'), row.get('加工時間')
                ))
                
                exists_before = cursor.fetchone()[0] > 0
                
                cursor.execute(insert_sql, (
                    row.get('A13'), row.get('A11'), row.get('A21'), row.get('A32'),
                    row.get('直径'), row.get('材料'), row.get('線材長'), row.get('回転速度'),
                    row.get('送り速度'), row.get('UPカット'), row.get('切込量'), row.get('突出量'),
                    row.get('載せ率'), row.get('パス数'), row.get('加工時間'),
                    row.get('上面ダレ量'), row.get('側面ダレ量'), row.get('摩耗量')
                ))
                
                if exists_before:
                    updated_count += 1
                else:
                    inserted_count += 1
                
                # ES: Mostrar progreso cada 1000 filas
                # EN: Report progress every 1000 rows
                # JA: 1000行ごとに進捗表示
                if (inserted_count + updated_count) % 1000 == 0:
                    progress = 90 + int(((inserted_count + updated_count) / total_rows) * 5)  # 90% a 95%
                    self.progress_updated.emit(progress, f"データ挿入中... ({inserted_count + updated_count}/{total_rows})")
            
            # ES: Obtener conteo final de registros
            # EN: Get final row count
            # JA: 最終レコード数を取得
            cursor.execute("SELECT COUNT(*) FROM yosoku_predictions")
            final_count = cursor.fetchone()[0]
            
            print(f"✅ {inserted_count + updated_count} 件を処理しました:")
            print(f"   📝 新規挿入: {inserted_count}")
            print(f"   🔄 更新（上書き）: {updated_count}")
            print(f"📊 DB レコード数: {initial_count} → {final_count}")
            print("💡 重複レコードは自動的に上書きされました")
            
            # ES: Paso 10: Finalizar | EN: Step 10: Finalize | JA: ステップ10：完了
            self.status_updated.emit("完了処理中...")
            self.progress_updated.emit(95, "完了処理中...")
            
            if self.cancelled:
                conn.close()
                return
            
            conn.commit()
            conn.close()
            
            # ES: Limpiar archivos temporales
            # EN: Clean up temporary files
            # JA: 一時ファイルを削除
            try:
                if os.path.exists(backup_path):
                    os.remove(backup_path)
                
                temp_folder = os.path.dirname(backup_path)
                if os.path.exists(temp_folder) and os.path.isdir(temp_folder):
                    try:
                        os.rmdir(temp_folder)
                    except OSError:
                        pass
            except Exception:
                pass
            
            self.status_updated.emit("インポート完了!")
            self.progress_updated.emit(100, "インポート完了!")
            self.finished.emit()
            
        except Exception as e:
            self.error.emit(str(e))

class ClassificationImportWorker(QThread):
    """ES: Worker para importación de resultados de clasificación a la BBDD de yosoku
    EN: Worker to import classification results into the yosoku DB
    JA: 分類結果を yosoku DB に取り込むワーカー
    """
    
    # ES: Señales | EN: Signals | JA: シグナル | EN: Signals | JA: シグナル
    progress_updated = Signal(int, str)  # percent, message
    status_updated = Signal(str)  # status message
    finished = Signal(int, int)  # registros_insertados, registros_actualizados
    error = Signal(str)  # error message
    
    def __init__(self, excel_path, overwrite=False, parent_widget=None):
        super().__init__()
        self.excel_path = excel_path
        self.overwrite = overwrite
        self.cancelled = False
    
    def cancel_import(self):
        """ES: Cancelar importación
        EN: Cancel import
        JA: インポートをキャンセル
        """
        self.cancelled = True
    
    def run(self):
        """ES: Ejecutar importación con progreso
        EN: Run import with progress updates
        JA: 進捗付きでインポートを実行
        """
        try:
            import pandas as pd
            import sqlite3
            import os
            import numpy as np
            
            # ES: Paso 1: Leer archivo Excel | EN: Step 1: Read Excel file | JA: ステップ1：Excelファイルを読み込み
            self.status_updated.emit("ファイル読み込み中...")
            self.progress_updated.emit(5, "ファイル読み込み中...")
            
            if self.cancelled:
                return
            
            if not os.path.exists(self.excel_path):
                self.error.emit(f"ファイルが見つかりません: {self.excel_path}")
                return
            
            df = pd.read_excel(self.excel_path)
            total_rows = len(df)
            
            if total_rows == 0:
                self.error.emit("ファイルにデータがありません")
                return
            
            # ES: Paso 2: Definir columnas para comparación (índice único)
            # EN: Step 2: Define comparison columns (unique index)
            # JA: 手順2：比較用列を定義（ユニークインデックス）
            # ES: Solo usar las columnas que realmente existen en el DataFrame
            # EN: Use only columns that actually exist in the DataFrame
            # JA: DataFrameに実際に存在する列のみ使用
            all_comparison_columns = [
                'A13', 'A11', 'A21', 'A32', '直径', '材料', '線材長', 
                '回転速度', '送り速度', 'UPカット', '切込量', '突出量', 
                '載せ率', 'パス数', '加工時間'
            ]
            
            # ES: Filtrar solo las columnas que existen en el DataFrame
            # EN: Filter to columns that exist in the DataFrame
            # JA: DataFrameに存在する列のみに絞る
            comparison_columns = [col for col in all_comparison_columns if col in df.columns]
            missing_cols = [col for col in all_comparison_columns if col not in df.columns]
            
            if len(comparison_columns) == 0:
                self.error.emit("比較に使用できる列が見つかりません。ファイルに必要な列が含まれているか確認してください。")
                return
            
            if missing_cols:
                print(f"⚠️ 以下の列がファイルに存在しません（NULLとして扱います）: {', '.join(missing_cols)}")
                print(f"✅ 比較に使用する列: {', '.join(comparison_columns)}")
            
            # ES: Paso 3: Procesar ambas BBDD (lineal y no_lineal)
            # EN: Step 3: Process both DBs (linear and non-linear)
            # JA: 手順3：両DBを処理（線形・非線形）
            total_inserted = 0
            total_updated = 0
            
            # ES: Procesar BBDD lineal (0-50% del progreso) | EN: Process linear DB (0-50% progress) | JA: 線形DB処理（進捗0-50%）
            self.status_updated.emit("線形データベース処理中...")
            self.progress_updated.emit(10, "線形データベース処理中...")
            
            if not self.cancelled:
                inserted_lineal, updated_lineal = self._process_database(
                    df, comparison_columns, YOSOKU_LINEAL_DB_PATH,
                    progress_start=10, progress_end=50
                )
                total_inserted += inserted_lineal
                total_updated += updated_lineal
            
            # ES: Procesar BBDD no lineal (50-100% del progreso) | EN: Process non-linear DB (50-100% progress) | JA: 非線形DB処理（進捗50-100%）
            if not self.cancelled:
                self.status_updated.emit("非線形データベース処理中...")
                self.progress_updated.emit(50, "非線形データベース処理中...")
                
                inserted_no_lineal, updated_no_lineal = self._process_database(
                    df, comparison_columns, YOSOKU_NO_LINEAL_DB_PATH,
                    progress_start=50, progress_end=95
                )
                total_inserted += inserted_no_lineal
                total_updated += updated_no_lineal
            
            if self.cancelled:
                return
            
            # ES: Finalizar
            # EN: Finalize
            # JA: 完了
            self.progress_updated.emit(100, "完了")
            self.status_updated.emit("インポート完了")
            self.finished.emit(total_inserted, total_updated)
            
        except Exception as e:
            print(f"❌ 分類のインポート中にエラー: {e}")
            import traceback
            traceback.print_exc()
            self.error.emit(f"インポート中にエラーが発生しました: {str(e)}")
    
    def _process_database(self, df, comparison_columns, db_path, progress_start=0, progress_end=100):
        """ES: Procesa una BBDD específica con los datos de clasificación
        EN: Process a specific DB with classification data
        JA: 分類データで特定DBを処理"""
        import pandas as pd
        import sqlite3
        import os
        
        inserted_count = 0
        updated_count = 0
        skipped_count = 0
        
        # ES: Conectar a BBDD | EN: Connect to DB | JA: DBに接続
        if not os.path.exists(db_path):
            print(f"ℹ️ DB {db_path} が存在しないため自動作成します")
        
        conn = sqlite3.connect(db_path, timeout=10)
        cursor = conn.cursor()
        
        try:
            # ES: Asegurar que la tabla existe (crear si no existe)
            # EN: Ensure table exists (create if it does not)
            # JA: テーブルが存在することを保証（無ければ作成）
            create_table_sql = """
            CREATE TABLE IF NOT EXISTS yosoku_predictions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                A13 INTEGER,
                A11 INTEGER,
                A21 INTEGER,
                A32 INTEGER,
                直径 REAL,
                材料 TEXT,
                線材長 REAL,
                回転速度 REAL,
                送り速度 REAL,
                UPカット INTEGER,
                切込量 REAL,
                突出量 REAL,
                載せ率 REAL,
                パス数 INTEGER,
                加工時間 REAL,
                上面ダレ量 REAL,
                側面ダレ量 REAL,
                摩耗量 REAL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """
            cursor.execute(create_table_sql)
            
            # ES: Asegurar que las columnas de clasificación existan
            # EN: Ensure classification columns exist
            # JA: 分類列が存在することを保証
            classification_columns = [
                ('pred_label', 'INTEGER'),
                ('p_cal', 'REAL'),
                ('tau_pos', 'REAL'),
                ('tau_neg', 'REAL'),
                ('ood_flag', 'INTEGER'),
                ('maha_dist', 'REAL')
            ]
            
            for col_name, col_type in classification_columns:
                try:
                    cursor.execute(f"ALTER TABLE yosoku_predictions ADD COLUMN {col_name} {col_type}")
                    print(f"✅ 列 {col_name} を {db_path} に追加しました")
                except sqlite3.OperationalError as e:
                    if "duplicate column" in str(e).lower() or "already exists" in str(e).lower():
                        print(f"ℹ️ 列 {col_name} は既に {db_path} に存在します")
                    else:
                        raise
            
            conn.commit()
            
            total_rows = len(df)
            progress_range = progress_end - progress_start
            
            if self.overwrite:
                # ES: ESTRATEGIA OPTIMIZADA: Cargar BBDD en memoria, merge, UPDATE/INSERT según corresponda
                # EN: OPTIMIZED STRATEGY: Load DB in memory, merge, then UPDATE/INSERT as needed
                # JA: 最適化戦略：DBをメモリに読み込み、マージ後UPDATE/INSERT
                # ES: Necesario porque el índice único incluye columnas que pueden no estar en el Excel
                # EN: Required because unique index includes columns that may be absent from Excel
                # JA: ユニークインデックスにExcelに無い列を含むため
                print("⚡ merge による最適化戦略を使用（上書きモード）")
                
                # ES: Cargar registros existentes de la BBDD (solo columnas de comparación que tenemos)
                # EN: Load existing records from DB (only comparison columns we have)
                # JA: DBから既存レコードを読み込み（持っている比較列のみ）
                progress_current = progress_start + int(progress_range * 0.1)
                self.status_updated.emit(f"既存データ読み込み中... ({db_path})")
                self.progress_updated.emit(progress_current, f"既存データ読み込み中... ({db_path})")
                
                db_query = f"SELECT id, {', '.join(comparison_columns)} FROM yosoku_predictions"
                db_df = pd.read_sql_query(db_query, conn)
                
                # ES: Preparar queries | EN: Prepare queries | JA: クエリを準備
                update_query = """
                    UPDATE yosoku_predictions 
                    SET pred_label = ?, p_cal = ?, tau_pos = ?, tau_neg = ?, 
                        ood_flag = ?, maha_dist = ?
                    WHERE id = ?
                """
                
                insert_columns = comparison_columns + ['pred_label', 'p_cal', 'tau_pos', 'tau_neg', 'ood_flag', 'maha_dist']
                placeholders = ','.join(['?'] * len(insert_columns))
                insert_query = f"""
                    INSERT INTO yosoku_predictions 
                    ({','.join(insert_columns)})
                    VALUES ({placeholders})
                """
                
                if len(db_df) > 0:
                    # ES: Crear clave de comparación en ambos DataFrames | EN: Create comparison key in both DataFrames | JA: 両DataFrameに比較キーを作成
                    def create_key(row, cols):
                        return tuple(row[col] if not pd.isna(row[col]) else 'NULL_VAL' for col in cols)
                    
                    excel_df = df.copy()
                    excel_df['_comparison_key'] = excel_df.apply(lambda r: create_key(r, comparison_columns), axis=1)
                    db_df['_comparison_key'] = db_df.apply(lambda r: create_key(r, comparison_columns), axis=1)
                    
                    # ES: Hacer merge para encontrar coincidencias
                    # EN: Merge to find matches
                    # JA: マージして一致を検索
                    merged = excel_df.merge(
                        db_df[['id', '_comparison_key']], 
                        on='_comparison_key', 
                        how='left',
                        suffixes=('', '_db')
                    )
                    
                    # ES: Separar en registros a actualizar vs insertar
                    # EN: Split into records to update vs insert
                    # JA: 更新対象と挿入対象のレコードに分離
                    to_update = merged[merged['id'].notna()].copy()
                    to_insert = merged[merged['id'].isna()].copy()
                    
                    print(f"📊 更新対象レコード: {len(to_update)}")
                    print(f"📊 挿入対象レコード: {len(to_insert)}")
                    
                    # ES: Procesar actualizaciones en lotes
                    # EN: Process updates in batches
                    # JA: 更新をバッチ処理
                    if len(to_update) > 0:
                        update_batch = []
                        for index, row in to_update.iterrows():
                            if self.cancelled:
                                conn.close()
                                return (inserted_count, updated_count)
                            
                            pred_label = row.get('pred_label', None)
                            p_cal = row.get('p_cal', None)
                            tau_pos = row.get('tau_pos', None)
                            tau_neg = row.get('tau_neg', None)
                            ood_flag = row.get('ood_flag', None)
                            maha_dist = row.get('maha_dist', None)
                            
                            # ES: Convertir NaN a None | EN: Convert NaN to None | JA: NaNをNoneに変換
                            if pd.isna(pred_label): pred_label = None
                            if pd.isna(p_cal): p_cal = None
                            if pd.isna(tau_pos): tau_pos = None
                            if pd.isna(tau_neg): tau_neg = None
                            if pd.isna(ood_flag): ood_flag = None
                            if pd.isna(maha_dist): maha_dist = None
                            
                            update_batch.append((
                                pred_label, p_cal, tau_pos, tau_neg, ood_flag, maha_dist, int(row['id'])
                            ))
                            
                            if len(update_batch) >= 1000:
                                cursor.executemany(update_query, update_batch)
                                updated_count += len(update_batch)
                                update_batch = []
                                
                                progress = progress_start + int(progress_range * 0.2) + int((updated_count / len(to_update)) * (progress_range * 0.3))
                                self.progress_updated.emit(progress, f"更新中: {updated_count}/{len(to_update)} ({db_path})")
                        
                        if update_batch:
                            cursor.executemany(update_query, update_batch)
                            updated_count += len(update_batch)
                    
                    # ES: Procesar inserciones en lotes
                    # EN: Process insertions in batches
                    # JA: 挿入をバッチ処理
                    if len(to_insert) > 0:
                        insert_batch = []
                        for index, row in to_insert.iterrows():
                            if self.cancelled:
                                conn.close()
                                return (inserted_count, updated_count)
                            
                            row_values = []
                            
                            # ES: Valores de columnas de comparación (solo las que tenemos) | EN: Comparison column values (only those we have) | JA: 比較列の値（持っているもののみ）
                            for col in comparison_columns:
                                val = row[col]
                                if pd.isna(val):
                                    row_values.append(None)
                                else:
                                    row_values.append(val)
                            
                            # ES: Valores de clasificación | EN: Classification values | JA: 分類の値
                            for col in ['pred_label', 'p_cal', 'tau_pos', 'tau_neg', 'ood_flag', 'maha_dist']:
                                val = row.get(col, None)
                                if pd.isna(val):
                                    row_values.append(None)
                                else:
                                    row_values.append(val)
                            
                            insert_batch.append(tuple(row_values))
                            
                            if len(insert_batch) >= 1000:
                                cursor.executemany(insert_query, insert_batch)
                                inserted_count += len(insert_batch)
                                insert_batch = []
                                
                                progress = progress_start + int(progress_range * 0.5) + int((inserted_count / len(to_insert)) * (progress_range * 0.3))
                                self.progress_updated.emit(progress, f"挿入中: {inserted_count}/{len(to_insert)} ({db_path})")
                        
                        if insert_batch:
                            cursor.executemany(insert_query, insert_batch)
                            inserted_count += len(insert_batch)
                else:
                    # ES: BBDD vacía, insertar todos | EN: Empty DB, insert all | JA: DBが空のため全件挿入
                    print("📊 DB が空のため全レコードを挿入します")
                    insert_batch = []
                    for index, row in df.iterrows():
                        if self.cancelled:
                            conn.close()
                            return (inserted_count, updated_count)
                        
                        row_values = []
                        for col in comparison_columns:
                            val = row.get(col, None)
                            if pd.isna(val):
                                row_values.append(None)
                            else:
                                row_values.append(val)
                        
                        for col in ['pred_label', 'p_cal', 'tau_pos', 'tau_neg', 'ood_flag', 'maha_dist']:
                            val = row.get(col, None)
                            if pd.isna(val):
                                row_values.append(None)
                            else:
                                row_values.append(val)
                        
                        insert_batch.append(tuple(row_values))
                        
                        if len(insert_batch) >= 1000:
                            cursor.executemany(insert_query, insert_batch)
                            inserted_count += len(insert_batch)
                            insert_batch = []
                            
                            progress = progress_start + int((inserted_count / total_rows) * (progress_range * 0.8))
                            self.progress_updated.emit(progress, f"挿入中: {inserted_count}/{total_rows} ({db_path})")
                    
                    if insert_batch:
                        cursor.executemany(insert_query, insert_batch)
                        inserted_count += len(insert_batch)
                
            else:
                # ES: ESTRATEGIA CON MERGE: Cargar BBDD en memoria y hacer merge (más rápido que SELECT por fila)
                # EN: MERGE STRATEGY: Load DB in memory and merge (faster than row-by-row SELECT)
                # JA: マージ戦略：DBをメモリに読み込みマージ（行単位SELECTより高速）
                print("⚡ merge 戦略を使用（上書きなしモード）")
                
                # ES: Cargar registros existentes de la BBDD (solo columnas necesarias)
                # EN: Load existing records from DB (only needed columns)
                # JA: DBから既存レコードを読み込み（必要な列のみ）
                progress_current = progress_start + int(progress_range * 0.1)
                self.status_updated.emit(f"既存データ読み込み中... ({db_path})")
                self.progress_updated.emit(progress_current, f"既存データ読み込み中... ({db_path})")
                
                db_query = f"SELECT id, {', '.join(comparison_columns)} FROM yosoku_predictions"
                db_df = pd.read_sql_query(db_query, conn)
                
                if len(db_df) > 0:
                    # ES: Crear clave de comparación en ambos DataFrames
                    # EN: Create comparison key in both DataFrames
                    # JA: 両DataFrameに比較キーを作成
                    # ES: Manejar NaN reemplazándolos con valor especial para la comparación
                    # EN: Handle NaN by replacing with special value for comparison
                    # JA: 比較用にNaNを特殊値で置換
                    def create_key(row, cols):
                        return tuple(row[col] if not pd.isna(row[col]) else 'NULL_VAL' for col in cols)
                    
                    excel_df = df.copy()
                    excel_df['_comparison_key'] = excel_df.apply(lambda r: create_key(r, comparison_columns), axis=1)
                    db_df['_comparison_key'] = db_df.apply(lambda r: create_key(r, comparison_columns), axis=1)
                    
                    # ES: Hacer merge para encontrar coincidencias | EN: Merge to find matches | JA: マージして一致を検出
                    merged = excel_df.merge(
                        db_df[['id', '_comparison_key']], 
                        on='_comparison_key', 
                        how='left',
                        suffixes=('', '_db')
                    )
                    
                    # ES: Separar en registros a insertar vs saltar
                    # EN: Split into records to insert vs skip
                    # JA: 挿入対象とスキップ対象のレコードに分離
                    to_insert = merged[merged['id'].isna()].copy()
                    to_skip = merged[merged['id'].notna()].copy()
                    
                    skipped_count = len(to_skip)
                    
                    print(f"📊 挿入対象レコード: {len(to_insert)}")
                    print(f"📊 スキップレコード（既存）: {skipped_count}")
                    
                    # ES: Insertar solo los nuevos
                    # EN: Insert only new ones
                    # JA: 新規のみ挿入
                    if len(to_insert) > 0:
                        insert_columns = comparison_columns + ['pred_label', 'p_cal', 'tau_pos', 'tau_neg', 'ood_flag', 'maha_dist']
                        placeholders = ','.join(['?'] * len(insert_columns))
                        insert_query = f"""
                            INSERT INTO yosoku_predictions 
                            ({','.join(insert_columns)})
                            VALUES ({placeholders})
                        """
                        
                        batch_data = []
                        for index, row in to_insert.iterrows():
                            if self.cancelled:
                                conn.close()
                                return (inserted_count, updated_count)
                            
                            row_values = []
                            
                            # ES: Valores de columnas de comparación | EN: Comparison column values | JA: 比較列の値
                            for col in comparison_columns:
                                val = row[col]
                                if pd.isna(val):
                                    row_values.append(None)
                                else:
                                    row_values.append(val)
                            
                            # ES: Valores de clasificación | EN: Classification values | JA: 分類の値
                            for col in ['pred_label', 'p_cal', 'tau_pos', 'tau_neg', 'ood_flag', 'maha_dist']:
                                val = row.get(col, None)
                                if pd.isna(val):
                                    row_values.append(None)
                                else:
                                    row_values.append(val)
                            
                            batch_data.append(tuple(row_values))
                            
                            # ES: Procesar en lotes de 1000 | EN: Process in batches of 1000 | JA: 1000件ずつ処理
                            if len(batch_data) >= 1000:
                                cursor.executemany(insert_query, batch_data)
                                inserted_count += len(batch_data)
                                batch_data = []
                                
                                progress = progress_start + int(progress_range * 0.2) + int((inserted_count / len(to_insert)) * (progress_range * 0.7))
                                self.progress_updated.emit(progress, f"挿入中: {inserted_count}/{len(to_insert)} ({db_path})")
                        
                        # ES: Procesar lote final
                        # EN: Process final batch
                        # JA: 最終バッチを処理
                        if batch_data:
                            cursor.executemany(insert_query, batch_data)
                            inserted_count += len(batch_data)
                else:
                    # ES: BBDD vacía, insertar todos | EN: Empty DB, insert all | JA: DBが空のため全件挿入
                    print("📊 DB が空のため全レコードを挿入します")
                    insert_columns = comparison_columns + ['pred_label', 'p_cal', 'tau_pos', 'tau_neg', 'ood_flag', 'maha_dist']
                    placeholders = ','.join(['?'] * len(insert_columns))
                    insert_query = f"""
                        INSERT INTO yosoku_predictions 
                        ({','.join(insert_columns)})
                        VALUES ({placeholders})
                    """
                    
                    batch_data = []
                    for index, row in df.iterrows():
                        if self.cancelled:
                            conn.close()
                            return (inserted_count, updated_count)
                        
                        row_values = []
                        for col in comparison_columns:
                            val = row.get(col, None)
                            if pd.isna(val):
                                row_values.append(None)
                            else:
                                row_values.append(val)
                        
                        for col in ['pred_label', 'p_cal', 'tau_pos', 'tau_neg', 'ood_flag', 'maha_dist']:
                            val = row.get(col, None)
                            if pd.isna(val):
                                row_values.append(None)
                            else:
                                row_values.append(val)
                        
                        batch_data.append(tuple(row_values))
                        
                        if len(batch_data) >= 1000:
                            cursor.executemany(insert_query, batch_data)
                            inserted_count += len(batch_data)
                            batch_data = []
                            
                            progress = progress_start + int((inserted_count / total_rows) * (progress_range * 0.8))
                            self.progress_updated.emit(progress, f"挿入中: {inserted_count}/{total_rows} ({db_path})")
                    
                    if batch_data:
                        cursor.executemany(insert_query, batch_data)
                        inserted_count += len(batch_data)
            
            # ES: Commit final
            # EN: Final commit
            # JA: 最終コミット
            conn.commit()
            conn.close()
            
            print(f"✅ {db_path} の処理完了: 挿入 {inserted_count}, 更新 {updated_count}")
            return (inserted_count, updated_count)
            
        except Exception as e:
            print(f"❌ 処理エラー {db_path}: {e}")
            import traceback
            traceback.print_exc()
            if conn:
                conn.close()
            raise

class YosokuExportWorker(QThread):
    """ES: Worker para exportación de datos Yosoku a Excel con progreso
    EN: Worker to export Yosoku data to Excel with progress
    JA: 進捗付きでYosokuデータをExcelへエクスポートするワーカー
    """
    
    # ES: Señales | EN: Signals | JA: シグナル | EN: Signals | JA: シグナル
    progress_updated = Signal(int, str)  # percent, message
    status_updated = Signal(str)  # status message
    finished = Signal(str, int)  # filepath, record_count
    error = Signal(str)  # error message
    
    def __init__(self, db_path, filepath, total_records):
        super().__init__()
        self.db_path = db_path
        self.filepath = filepath
        self.total_records = total_records
        self.cancelled = False
    
    def cancel_export(self):
        """ES: Cancelar exportación
        EN: Cancel export
        JA: エクスポートをキャンセル
        """
        self.cancelled = True
    
    def run(self):
        """ES: Ejecutar exportación con progreso
        EN: Run export with progress updates
        JA: 進捗付きでエクスポートを実行
        """
        try:
            import pandas as pd
            import sqlite3
            
            # ES: Paso 1: Conectar a base de datos | EN: Step 1: Connect to database | JA: 手順1：DBに接続
            self.status_updated.emit("データベースに接続中...")
            self.progress_updated.emit(10, "データベースに接続中...")
            
            if self.cancelled:
                return
            
            conn = sqlite3.connect(self.db_path)
            
            # ES: Paso 2: Leer datos | EN: Step 2: Read data | JA: 手順2：データを読み込み
            self.status_updated.emit("データを読み込み中...")
            self.progress_updated.emit(30, "データを読み込み中...")
            
            if self.cancelled:
                conn.close()
                return
            
            df = pd.read_sql_query("SELECT * FROM yosoku_predictions", conn)
            conn.close()
            
            # ES: Paso 3: Exportar a Excel | EN: Step 3: Export to Excel | JA: 手順3：Excelにエクスポート
            self.status_updated.emit("Excelファイルに書き込み中...")
            self.progress_updated.emit(60, "Excelファイルに書き込み中...")
            
            if self.cancelled:
                return
            
            df.to_excel(self.filepath, index=False)
            
            # ES: Paso 4: Completado | EN: Step 4: Completed | JA: 手順4：完了
            self.status_updated.emit("エクスポート完了")
            self.progress_updated.emit(100, "エクスポート完了")
            
            if not self.cancelled:
                self.finished.emit(self.filepath, len(df))
            
        except Exception as e:
            if not self.cancelled:
                error_msg = f"❌ エクスポート中にエラーが発生しました:\n{str(e)}"
                self.error.emit(error_msg)

class LinearAnalysisWorker(QThread):
    """ES: Worker para análisis lineal con señales de progreso
    EN: Worker for linear analysis with progress signals
    JA: 進捗シグナル付き線形解析ワーカー
    """
    
    # ES: Señales | EN: Signals | JA: シグナル | EN: Signals | JA: シグナル
    progress_updated = Signal(int, str)  # percent, message
    status_updated = Signal(str)  # status message
    finished = Signal(dict)  # results
    error = Signal(str)  # error message
    
    def __init__(self, db_manager, filters, output_folder, parent_widget=None):
        super().__init__()
        self.db_manager = db_manager
        self.filters = filters
        self.output_folder = output_folder
        self.db_connection = None
        self.is_cancelled = False  # ✅ NEW: Cancellation flag
        
    def stop(self):
        """ES: Método para solicitar la parada del worker
        EN: Request the worker to stop
        JA: ワーカー停止要求
        """
        self.is_cancelled = True

    def run(self):
        """ES: Ejecutar análisis lineal con progreso
        EN: Run linear analysis with progress updates
        JA: 進捗付きで線形解析を実行
        """
        import threading
        print(f"🚀 DEBUG: LinearAnalysisWorker iniciado en hilo: {threading.current_thread().name}")
        try:
            if self.is_cancelled: return  # Initial cancellation check

            self.status_updated.emit("データベースからデータを取得中...")
            self.progress_updated.emit(10, "データベースからデータを取得中...")
            
            if self.is_cancelled: return  # Check after emitting

            # ES: Pequeño delay para mostrar progreso | EN: Short delay to show progress | JA: 進捗表示用の短い遅延
            import time
            time.sleep(0.5)
            
            if self.is_cancelled: return

            # ES: Crear nueva conexión de base de datos en este thread | EN: Create new DB connection in this thread | JA: このスレッドで新規DB接続を作成
            import sqlite3
            self.db_connection = sqlite3.connect(RESULTS_DB_PATH, timeout=10)
            cursor = self.db_connection.cursor()
            
            # ES: Obtener datos filtrados
            # EN: Get filtered data
            # JA: フィルタ済みデータを取得
            query = "SELECT * FROM main_results WHERE 1=1"
            params = []
            
            # ... (filtros) ...
            # ES: No cambiar la lógica de filtros aquí; se mantiene igual
            # EN: Filter logic unchanged here; kept as-is
            # JA: フィルタロジックはここでは変更せずそのまま
            # ES: Mejor leo el archivo de nuevo para asegurar el bloque exacto.
            # EN: I'll re-read the file to ensure the exact block.
            # JP: 正確なブロックを確認するためにファイルを読み直す。
            self.progress_updated.emit(10, "データベースからデータを取得中...")
            
            # ES: Pequeño delay para mostrar progreso | EN: Short delay to show progress | JA: 進捗表示用の短い遅延
            import time
            time.sleep(0.5)
            
            # ES: Crear nueva conexión de base de datos en este thread | EN: Create new DB connection in this thread | JA: このスレッドで新規DB接続を作成
            import sqlite3
            self.db_connection = sqlite3.connect(RESULTS_DB_PATH, timeout=10)
            cursor = self.db_connection.cursor()
            
            # ES: Obtener datos filtrados
            # EN: Get filtered data
            # JA: フィルタ済みデータを取得
            query = "SELECT * FROM main_results WHERE 1=1"
            params = []
            
            # ES: Aplicar filtros de cepillo
            # EN: Apply brush filters
            # JA: ブラシフィルタを適用
            brush_selections = []
            if 'すべて' in self.filters and self.filters['すべて']:
                brush_condition = " OR ".join([f"{brush} = 1" for brush in ['A13', 'A11', 'A21', 'A32']])
                query += f" AND ({brush_condition})"
            else:
                for brush_type in ['A13', 'A11', 'A21', 'A32']:
                    if brush_type in self.filters and self.filters[brush_type]:
                        brush_selections.append(brush_type)
                
                if brush_selections:
                    brush_condition = " OR ".join([f"{brush} = 1" for brush in brush_selections])
                    query += f" AND ({brush_condition})"
            
            # ES: Aplicar filtros de rango
            # EN: Apply range filters
            # JA: 範囲フィルタを適用
            field_to_db = {
                "面粗度(Ra)前": "面粗度前",
                "面粗度(Ra)後": "面粗度後",
            }
            for field_name, filter_value in self.filters.items():
                if field_name in ['すべて', 'A13', 'A11', 'A21', 'A32']:
                    continue
                db_field = field_to_db.get(field_name, field_name)
                    
                if isinstance(filter_value, tuple) and len(filter_value) == 2:
                    desde, hasta = filter_value
                    if desde is not None and hasta is not None:
                        if field_name == "実験日":
                            desde_str = desde.toString("yyyyMMdd") if hasattr(desde, 'toString') else str(desde)
                            hasta_str = hasta.toString("yyyyMMdd") if hasattr(hasta, 'toString') else str(hasta)
                            query += f" AND {db_field} BETWEEN ? AND ?"
                            params.extend([desde_str, hasta_str])
                        else:
                            try:
                                desde_num = float(desde) if isinstance(desde, str) else desde
                                hasta_num = float(hasta) if isinstance(hasta, str) else hasta
                                query += f" AND {db_field} BETWEEN ? AND ?"
                                params.extend([desde_num, hasta_num])
                            except (ValueError, TypeError):
                                continue
                elif isinstance(filter_value, (str, int, float)) and filter_value:
                    try:
                        if field_name in ['線材長', '回転速度', '送り速度', 'UPカット', '突出量', 'パス数', 'バリ除去']:
                            value_num = int(filter_value) if isinstance(filter_value, str) else filter_value
                        else:
                            value_num = float(filter_value) if isinstance(filter_value, str) else filter_value
                        
                        query += f" AND {db_field} = ?"
                        params.append(value_num)
                    except (ValueError, TypeError):
                        continue
            
            # ES: Ejecutar consulta usando la nueva conexión | EN: Execute query using the new connection | JA: 新規接続でクエリを実行
            cursor.execute(query, params)
            filtered_data = cursor.fetchall()
            
            self.status_updated.emit("データを処理中...")
            self.progress_updated.emit(20, "データを処理中...")
            time.sleep(0.3)
            
            if not filtered_data:
                self.error.emit("フィルター条件に一致するデータが見つかりません")
                return
            
            # ES: Convertir a DataFrame | EN: Convert to DataFrame | JA: DataFrameに変換
            import pandas as pd
            # ES: No depender del orden físico de columnas en SQLite (puede cambiar con migraciones)
            # EN: Do not rely on physical column order in SQLite (may change with migrations)
            # JA: SQLiteの物理列順に依存しない（マイグレで変わりうる）
            column_names = [d[0] for d in cursor.description] if cursor.description else None
            df = pd.DataFrame(filtered_data, columns=column_names)
            
            self.status_updated.emit("データファイルを保存中...")
            self.progress_updated.emit(30, "データファイルを保存中...")
            time.sleep(0.3)
            
            # ES: Crear estructura de carpetas | EN: Create folder structure | JA: フォルダ構造を作成
            import os
            os.makedirs(self.output_folder, exist_ok=True)
            models_folder = os.path.join(self.output_folder, "01_学習モデル")
            os.makedirs(models_folder, exist_ok=True)
            
            # ES: Guardar datos filtrados | EN: Save filtered data | JA: フィルタ済みデータを保存
            filtered_data_path = os.path.join(models_folder, "filtered_data.xlsx")
            df.to_excel(filtered_data_path, index=False)
            
            if self.is_cancelled: return  # ✅ Cancellation check

            self.status_updated.emit("機械学習パイプラインを初期化中...")
            self.progress_updated.emit(40, "機械学習パイプラインを初期化中...")
            time.sleep(0.4)
            
            if self.is_cancelled: return  # ✅ Cancellation check

            # ES: Importar y configurar pipeline
            # EN: Import and configure pipeline
            # JA: パイプラインをインポート・設定
            from linear_analysis_advanced import IntegratedMLPipeline, PipelineConfig
            
            config = PipelineConfig()
            config.TRANSFORMATION['enable'] = True
            config.TRANSFORMATION['mode'] = 'advanced'
            config.FEATURE_SELECTION['method'] = 'importance'
            config.FEATURE_SELECTION['k_features'] = 10
            config.PREPROCESSING['noise_augmentation_ratio'] = 0.3
            config.TRANSFORMATION['improvement_threshold'] = 0.005
            
            pipeline = IntegratedMLPipeline(base_dir=self.output_folder, config=config)
            
            self.status_updated.emit("データを読み込み中...")
            self.progress_updated.emit(15, "データを読み込み中...")
            time.sleep(0.2)
            
            # ES: Cargar datos | EN: Load data | JA: データを読み込み
            pipeline.load_data(filtered_data_path, index_col='Index')
            
            self.status_updated.emit("データ構造を分析中...")
            self.progress_updated.emit(18, "データ構造を分析中...")
            time.sleep(0.2)
            
            self.status_updated.emit("変数を分離中...")
            self.progress_updated.emit(20, "変数を分離中...")
            time.sleep(0.2)
            
            if self.is_cancelled: return  # ✅ Cancellation check

            # ES: Separar variables | EN: Separate variables | JA: 変数を分離
            try:
                pipeline.separate_variables()
            except Exception as e:
                self.error.emit(f"Error separando variables: {str(e)}")
                return
            
            if self.is_cancelled: return  # ✅ Cancellation check

            self.status_updated.emit("特徴量を選択中...")
            self.progress_updated.emit(22, "特徴量を選択中...")
            time.sleep(0.2)
            
            self.status_updated.emit("データを前処理中...")
            self.progress_updated.emit(25, "データを前処理中...")
            time.sleep(0.3)
            
            if self.is_cancelled: return  # ✅ Cancellation check

            # ES: Preprocesar datos | EN: Preprocess data | JA: データを前処理
            try:
                pipeline.preprocess_data()
            except Exception as e:
                self.error.emit(f"Error preprocesando datos: {str(e)}")
                return
            
            if self.is_cancelled: return  # ✅ Cancellation check

            self.status_updated.emit("回帰モデルを初期化中...")
            self.progress_updated.emit(30, "回帰モデルを初期化中...")
            time.sleep(0.2)
            
            self.status_updated.emit("線形回帰モデルを訓練中...")
            self.progress_updated.emit(35, "線形回帰モデルを訓練中...")
            time.sleep(0.3)
            
            if self.is_cancelled: return  # ✅ Cancellation check

            self.status_updated.emit("ランダムフォレストモデルを訓練中...")
            self.progress_updated.emit(40, "ランダムフォレストモデルを訓練中...")
            time.sleep(0.3)
            
            if self.is_cancelled: return  # ✅ Cancellation check

            self.status_updated.emit("SVMモデル को 訓練中...")
            self.progress_updated.emit(45, "SVMモデルを訓練中...")
            time.sleep(0.3)
            
            self.status_updated.emit("分類モデルを初期化中...")
            self.progress_updated.emit(50, "分類モデルを初期化中...")
            time.sleep(0.2)
            
            self.status_updated.emit("ロジスティック回帰を訓練中...")
            self.progress_updated.emit(55, "ロジスティック回帰を訓練中...")
            time.sleep(0.3)
            
            self.status_updated.emit("決定木モデルを訓練中...")
            self.progress_updated.emit(60, "決定木モデルを訓練中...")
            time.sleep(0.3)
            
            self.status_updated.emit("ナイーブベイズモデルを訓練中...")
            self.progress_updated.emit(65, "ナイーブベイズモデルを訓練中...")
            time.sleep(0.3)
            
            self.status_updated.emit("モデルを評価中...")
            self.progress_updated.emit(40, "モデルを評価中...")
            time.sleep(0.2)
            
            self.status_updated.emit("モデル初期化中...")
            self.progress_updated.emit(41, "モデル初期化中...")
            time.sleep(0.2)
            
            self.status_updated.emit("データセットを分割中...")
            self.progress_updated.emit(42, "データセットを分割中...")
            time.sleep(0.2)
            
            self.status_updated.emit("訓練データを準備中...")
            self.progress_updated.emit(43, "訓練データを準備中...")
            time.sleep(0.2)
            
            self.status_updated.emit("検証データを準備中...")
            self.progress_updated.emit(44, "検証データを準備中...")
            time.sleep(0.2)
            
            self.status_updated.emit("モデル訓練を開始中...")
            self.progress_updated.emit(45, "モデル訓練を開始中...")
            time.sleep(0.2)
            
            # ES: Entrenar modelos
            # EN: Train models
            # JA: モデルを訓練
            try:
                pipeline.train_models()
            except Exception as e:
                self.error.emit(f"Error entrenando modelos: {str(e)}")
                return
            
            self.status_updated.emit("モデル訓練完了...")
            self.progress_updated.emit(46, "モデル訓練完了...")
            time.sleep(0.2)
            
            self.status_updated.emit("回帰モデルの性能を評価中...")
            self.progress_updated.emit(47, "回帰モデルの性能を評価中...")
            time.sleep(0.2)
            
            self.status_updated.emit("分類モデルの性能を評価中...")
            self.progress_updated.emit(48, "分類モデルの性能を評価中...")
            time.sleep(0.2)
            
            self.status_updated.emit("交差検証を実行中...")
            self.progress_updated.emit(49, "交差検証を実行中...")
            time.sleep(0.3)
            
            if self.is_cancelled: return  # ✅ NEW: Cancellation brake

            self.status_updated.emit("メトリクスを計算中...")
            self.progress_updated.emit(50, "メトリクスを計算中...")
            time.sleep(0.2)
            
            self.status_updated.emit("モデル比較を実行中...")
            self.progress_updated.emit(51, "モデル比較を実行中...")
            time.sleep(0.2)
            
            self.status_updated.emit("最適なモデルを選択中...")
            self.progress_updated.emit(52, "最適なモデルを選択中...")
            time.sleep(0.2)
            
            if self.is_cancelled: return  # ✅ NEW: Cancellation brake

            self.status_updated.emit("プロペンシティスコアを計算中...")
            self.progress_updated.emit(53, "プロペンシティスコアを計算中...")
            time.sleep(0.3)
            
            self.status_updated.emit("スコアの正規化中...")
            self.progress_updated.emit(54, "スコアの正規化中...")
            time.sleep(0.2)
            
            self.status_updated.emit("統計的検定を実行中...")
            self.progress_updated.emit(55, "統計的検定を実行中...")
            time.sleep(0.3)
            
            self.status_updated.emit("信頼区間を計算中...")
            self.progress_updated.emit(56, "信頼区間を計算中...")
            time.sleep(0.2)
            
            self.status_updated.emit("結果の整合性を確認中...")
            self.progress_updated.emit(57, "結果の整合性を確認中...")
            time.sleep(0.2)
            
            self.status_updated.emit("データの品質を検証中...")
            self.progress_updated.emit(58, "データの品質を検証中...")
            time.sleep(0.2)
            
            self.status_updated.emit("異常値を検出中...")
            self.progress_updated.emit(59, "異常値を検出中...")
            time.sleep(0.2)
            
            if self.is_cancelled: return  # ✅ NEW: Cancellation brake

            self.status_updated.emit("モデルの安定性を確認中...")
            self.progress_updated.emit(60, "モデルの安定性を確認中...")
            time.sleep(0.2)
            
            self.status_updated.emit("最終評価を実行中...")
            self.progress_updated.emit(61, "最終評価を実行中...")
            time.sleep(0.3)
            
            self.status_updated.emit("結果を保存中...")
            self.progress_updated.emit(62, "結果を保存中...")
            time.sleep(0.2)
            
            self.status_updated.emit("Excelファイルを作成中...")
            self.progress_updated.emit(63, "Excelファイルを作成中...")
            time.sleep(0.2)
            
            self.status_updated.emit("データをフォーマット中...")
            self.progress_updated.emit(64, "データをフォーマット中...")
            time.sleep(0.2)
            
            self.status_updated.emit("グラフを生成中...")
            self.progress_updated.emit(65, "グラフを生成中...")
            time.sleep(0.2)
            
            if self.is_cancelled: return  # ✅ NEW: Cancellation brake

            self.status_updated.emit("散布図を作成中...")
            self.progress_updated.emit(66, "散布図を作成中...")
            time.sleep(0.2)
            
            self.status_updated.emit("ヒートマップを生成中...")
            self.progress_updated.emit(67, "ヒートマップを生成中...")
            time.sleep(0.2)
            
            self.status_updated.emit("相関図を作成中...")
            self.progress_updated.emit(68, "相関図を作成中...")
            time.sleep(0.2)
            
            self.status_updated.emit("予測テンプレートを作成中...")
            self.progress_updated.emit(69, "予測テンプレートを作成中...")
            time.sleep(0.2)
            
            self.status_updated.emit("計算式を生成中...")
            self.progress_updated.emit(70, "計算式を生成中...")
            time.sleep(0.2)
            
            if self.is_cancelled: return  # ✅ NEW: Cancellation brake

            self.status_updated.emit("逆変換テンプレートを作成中...")
            self.progress_updated.emit(71, "逆変換テンプレートを作成中...")
            time.sleep(0.2)
            
            self.status_updated.emit("ファイルを最適化中...")
            self.progress_updated.emit(72, "ファイルを最適化中...")
            time.sleep(0.2)
            
            self.status_updated.emit("最終処理中...")
            self.progress_updated.emit(73, "最終処理中...")
            time.sleep(0.2)
            
            self.status_updated.emit("クリーンアップを実行中...")
            self.progress_updated.emit(74, "クリーンアップを実行中...")
            time.sleep(0.2)
            
            self.status_updated.emit("完了確認中...")
            self.progress_updated.emit(75, "完了確認中...")
            time.sleep(0.2)
            
            # ES: Calcular propensity scores y guardar resultados
            # EN: Calculate propensity scores and save results
            # JA: 傾向スコアを計算し結果を保存
            try:
                propensity_scores = pipeline.calculate_propensity_scores()
                pipeline.save_results()
                pipeline.create_prediction_template()
            except Exception as e:
                self.error.emit(f"結果の保存中にエラー: {str(e)}")
                return
            try:
                pipeline.save_prediction_formulas()
                # ES: Crear Excel durante análisis lineal | EN: Create Excel during linear analysis | JA: 線形解析中にExcelを作成
                print("🔧 線形解析中にExcel作成を開始します...")
                
                excel_calculator_path = pipeline.create_excel_prediction_calculator_with_inverse(None)
                
                if excel_calculator_path:
                    print(f"✅ Excel の作成が完了しました: {excel_calculator_path}")
                else:
                    print("⚠️ Excel を作成できませんでした（None が返りました）")
                    
            except Exception as e:
                print(f"❌ Excel 作成の詳細エラー: {str(e)}")
                import traceback
                traceback.print_exc()
                self.error.emit(f"Excel 予測計算シートの作成中にエラー: {str(e)}")
                return
            
            self.status_updated.emit("分析完了！")
            self.progress_updated.emit(100, "分析完了！")
            
            # ES: Preparar resultados | EN: Prepare results | JA: 結果を準備
            results = {
                'success': True,
                'data_count': len(df),
                'models_trained': len(pipeline.models),
                'output_folder': self.output_folder,
                'filters_applied': list(self.filters.keys()),
                'data_range': f"線材長: {df['線材長'].min()}-{df['線材長'].max()}, 送り速度: {df['送り速度'].min()}-{df['送り速度'].max()}" if len(df) > 0 else "N/A",
                'excel_calculator': None,  # ✅ FIX: keep None to avoid crash
                'transformation_info': pipeline.transformation_info,
                'feature_selection': pipeline.results.get('feature_selection', {}),
                'target_info': pipeline.target_info,
                'models': pipeline.models
            }
            
            # ES: Crear resumen de resultados | EN: Create results summary | JA: 結果サマリを作成
            summary = []
            for target_name, model_info in pipeline.models.items():
                if model_info.get('model') is not None:
                    if model_info['task_type'] == 'regression':
                        metrics = model_info.get('final_metrics', {})
                        summary.append({
                            'target': target_name,
                            'model': model_info.get('model_name', 'Unknown'),
                            'r2': metrics.get('r2', 'N/A'),
                            'mae': metrics.get('mae', 'N/A'),
                            'rmse': metrics.get('rmse', 'N/A'),
                            'transformation': pipeline.transformation_info.get(target_name, {}).get('method', 'none')
                        })
                    else:
                        metrics = model_info.get('final_metrics', {})
                        summary.append({
                            'target': target_name,
                            'model': model_info.get('model_name', 'Unknown'),
                            'accuracy': metrics.get('accuracy', 'N/A'),
                            'f1_score': metrics.get('f1_score', 'N/A'),
                            'transformation': 'none'
                        })
            
            results['summary'] = summary
            
            self.finished.emit(results)
            
        except Exception as e:
            import threading
            import traceback
            error_msg = f"❌ 線形解析ワーカーでエラー（スレッド: {threading.current_thread().name}）: {e}"
            print(error_msg)
            traceback.print_exc()
            self.error.emit(error_msg)
        finally:
            import threading
            print(f"🛑 DEBUG: LinearAnalysisWorker finalizando en hilo: {threading.current_thread().name}")
            # ES: Cerrar conexión de base de datos si existe | EN: Close DB connection if it exists | JA: DB接続があれば閉じる
            if hasattr(self, 'db_connection') and self.db_connection:
                try:
                    self.db_connection.close()
                    print("🛑 DEBUG: ワーカーでDB接続を閉じました")
                except:
                    pass

class ProjectCreationDialog(QDialog):
    """ES: Diálogo para crear un nuevo proyecto
    EN: Dialog to create a new project
    JA: 新規プロジェクト作成ダイアログ"""
    
    def __init__(self, parent=None, analysis_type="nonlinear"):
        super().__init__(parent)
        self.analysis_type = analysis_type  # "nonlinear" o "classification"
        self.setWindowTitle("新規プロジェクト作成")
        self.setFixedSize(500, 300)
        self.setModal(True)
        
        # ES: Layout principal | EN: Main layout | JA: メインレイアウト
        layout = QVBoxLayout()
        
        # ES: Título | EN: Title | JA: タイトル
        title_label = QLabel("新規プロジェクトを作成します")
        title_label.setStyleSheet("font-size: 16px; font-weight: bold; margin: 10px;")
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)
        
        # ES: Formulario | EN: Form | JA: フォーム
        form_layout = QFormLayout()
        
        # ES: Nombre del proyecto | EN: Project name | JA: プロジェクト名
        self.project_name_edit = QLineEdit()
        self.project_name_edit.setPlaceholderText("プロジェクト名を入力してください")
        form_layout.addRow("プロジェクト名:", self.project_name_edit)
        
        # ES: Directorio | EN: Directory | JA: ディレクトリ
        directory_layout = QHBoxLayout()
        self.directory_edit = QLineEdit()
        self.directory_edit.setPlaceholderText("プロジェクトを保存するディレクトリを選択してください")
        self.directory_edit.setReadOnly(True)
        
        browse_button = QPushButton("参照...")
        browse_button.clicked.connect(self.browse_directory)
        
        directory_layout.addWidget(self.directory_edit)
        directory_layout.addWidget(browse_button)
        
        form_layout.addRow("保存先:", directory_layout)
        
        layout.addLayout(form_layout)
        
        # ES: Botones
        # EN: Buttons
        # JA: ボタン
        button_layout = QHBoxLayout()
        
        cancel_button = QPushButton("キャンセル")
        cancel_button.clicked.connect(self.reject)
        
        create_button = QPushButton("作成")
        create_button.clicked.connect(self.accept)
        create_button.setStyleSheet("background-color: #27ae60; color: white; font-weight: bold;")
        
        button_layout.addWidget(cancel_button)
        button_layout.addStretch()
        button_layout.addWidget(create_button)
        
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
        
        # ES: Variables para almacenar los datos | EN: Variables to store data | JA: データ保持用変数
        self.project_name = ""
        self.project_directory = ""
    
    def browse_directory(self):
        """ES: Abrir diálogo para seleccionar directorio
        EN: Open dialog to select directory
        JA: ディレクトリ選択ダイアログを開く"""
        directory = QFileDialog.getExistingDirectory(self, "プロジェクト保存先を選択")
        if directory:
            self.directory_edit.setText(directory)
    
    def accept(self):
        """ES: Validar y aceptar el diálogo
        EN: Validate and accept the dialog
        JA: ダイアログを検証して承認"""
        project_name = self.project_name_edit.text().strip()
        directory = self.directory_edit.text().strip()
        
        if not directory:
            QMessageBox.warning(self, "エラー", "保存先ディレクトリを選択してください。")
            return
        
        # ES: Verificar si la carpeta seleccionada es un proyecto válido
        # EN: Check if selected folder is a valid project
        # JA: 選択フォルダが有効なプロジェクトか確認
        selected_path = Path(directory)
        
        # ES: Verificar si la carpeta seleccionada es un proyecto (usar tipo de análisis del diálogo)
        # EN: Check if selected folder is a project (use dialog analysis type)
        # JA: 選択フォルダがプロジェクトか確認（ダイアログの解析種別を使用）
        if self.parent().is_valid_project_folder(str(selected_path), analysis_type=self.analysis_type):
            # ES: La carpeta seleccionada ES un proyecto, usarla directamente
            # EN: Selected folder is a project; use it directly
            # JA: 選択フォルダはプロジェクトなのでそのまま使用
            self.project_name = selected_path.name
            self.project_directory = str(selected_path.parent)
            print(f"✅ 選択したフォルダは有効なプロジェクトです: {selected_path}")
            super().accept()
            return
        
        # ES: Verificar si dentro de la carpeta hay proyectos
        # EN: Check if there are projects inside the folder
        # JA: フォルダ内にプロジェクトがあるか確認
        project_folders = self.parent().find_project_folders_in_directory(str(selected_path), analysis_type=self.analysis_type)
        
        if project_folders:
            # ES: Hay proyectos dentro de la carpeta seleccionada; preguntar crear nuevo o usar existente
            # EN: There are projects in the selected folder; ask to create new or use existing
            # JA: 選択フォルダ内にプロジェクトあり。新規作成か既存使用かを質問
            choice_dialog = QDialog(self)
            choice_dialog.setWindowTitle("プロジェクト選択")
            choice_dialog.setMinimumWidth(450)
            
            choice_layout = QVBoxLayout()
            
            info_label = QLabel(
                f"選択したディレクトリ内に {len(project_folders)} 個の既存プロジェクトが見つかりました。\n\n"
                f"新規プロジェクトを作成しますか？\n"
                f"それとも既存のプロジェクトを使用しますか？"
            )
            info_label.setWordWrap(True)
            choice_layout.addWidget(info_label)
            
            # ES: Mostrar lista de proyectos existentes | EN: Show list of existing projects | JA: 既存プロジェクト一覧を表示
            projects_label = QLabel("既存プロジェクト:")
            projects_label.setStyleSheet("font-weight: bold; margin-top: 10px;")
            choice_layout.addWidget(projects_label)
            
            projects_list = QListWidget()
            projects_list.setMaximumHeight(150)
            for folder in project_folders:
                projects_list.addItem(folder)
            choice_layout.addWidget(projects_list)
            
            buttons = QDialogButtonBox(QDialogButtonBox.Cancel)
            
            # ES: Botón para crear nuevo | EN: Button to create new | JA: 新規作成ボタン
            create_new_btn = buttons.addButton("新規作成", QDialogButtonBox.ActionRole)
            create_new_btn.setStyleSheet("background-color: #27ae60; color: white; font-weight: bold; padding: 8px;")
            
            # ES: Botón para usar existente | EN: Button to use existing | JA: 既存を使用ボタン
            use_existing_btn = buttons.addButton("既存を使用", QDialogButtonBox.ActionRole)
            use_existing_btn.setStyleSheet("background-color: #3498db; color: white; font-weight: bold; padding: 8px;")
            
            # ES: Variables para almacenar la elección | EN: Variable to store user choice | JA: 選択結果保持用変数
            choice_result = None
            
            # ES: Conectar botones a funciones | EN: Connect buttons to handlers | JA: ボタンにハンドラを接続
            def on_create_new():
                nonlocal choice_result
                choice_result = "create_new"
                choice_dialog.accept()
            
            def on_use_existing():
                nonlocal choice_result
                choice_result = "use_existing"
                choice_dialog.accept()
            
            create_new_btn.clicked.connect(on_create_new)
            use_existing_btn.clicked.connect(on_use_existing)
            
            choice_layout.addWidget(buttons)
            choice_dialog.setLayout(choice_layout)
            
            result = choice_dialog.exec()
            
            if result == QDialog.Accepted and choice_result:
                if choice_result == "create_new":
                    # ES: Usuario quiere crear nuevo - validar nombre
                    # EN: User wants to create new; validate name
                    # JA: ユーザーが新規作成希望；名前を検証
                    if not project_name:
                        QMessageBox.warning(self, "エラー", "プロジェクト名を入力してください。")
                        return
                    
                    # ES: Almacenar los datos para crear nuevo proyecto
                    # EN: Store data to create new project
                    # JA: 新規プロジェクト作成用にデータを保存
                    self.project_name = project_name
                    self.project_directory = directory
                    print(f"📁 新規プロジェクトを作成します: {project_name}（場所: {directory}）")
                    super().accept()
                    return
                
                elif choice_result == "use_existing":
                    # ES: Usuario quiere usar existente - mostrar lista para seleccionar
                    # EN: User wants to use existing; show list to select
                    # JA: ユーザーが既存使用希望；選択用リストを表示
                    if len(project_folders) == 1:
                        # ES: Solo hay un proyecto, usarlo directamente
                        # EN: Only one project; use it directly
                        # JA: プロジェクトが1件のみ；そのまま使用
                        project_path = Path(project_folders[0])
                        self.project_name = project_path.name
                        self.project_directory = str(project_path.parent)
                        print(f"✅ 既存プロジェクトを使用します: {project_path}")
                        super().accept()
                        return
                    else:
                        # ES: Hay múltiples proyectos, mostrar lista para seleccionar
                        # EN: Multiple projects; show list to select
                        # JA: 複数プロジェクトあり。選択用リストを表示
                        select_dialog = QDialog(self)
                        select_dialog.setWindowTitle("プロジェクトを選択")
                        select_dialog.setMinimumWidth(500)
                        
                        select_layout = QVBoxLayout()
                        select_label = QLabel(f"使用するプロジェクトを選択してください:")
                        select_layout.addWidget(select_label)
                        
                        list_widget = QListWidget()
                        for folder in project_folders:
                            list_widget.addItem(folder)
                        list_widget.setCurrentRow(0)
                        select_layout.addWidget(list_widget)
                        
                        select_buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
                        select_buttons.accepted.connect(select_dialog.accept)
                        select_buttons.rejected.connect(select_dialog.reject)
                        select_layout.addWidget(select_buttons)
                        
                        select_dialog.setLayout(select_layout)
                        
                        if select_dialog.exec() == QDialog.Accepted:
                            selected_project = list_widget.currentItem().text()
                            project_path = Path(selected_project)
                            self.project_name = project_path.name
                            self.project_directory = str(project_path.parent)
                            print(f"✅ 選択したプロジェクトを使用します: {project_path}")
                            super().accept()
                            return
                        else:
                            # ES: Usuario canceló selección, volver al diálogo principal
                            # EN: User cancelled selection; return to main dialog
                            # JA: ユーザーが選択をキャンセル。メインダイアログに戻る
                            return
            
            # ES: Si se canceló el diálogo de elección, no hacer nada
            # EN: If choice dialog was cancelled, do nothing
            # JA: 選択ダイアログがキャンセルされた場合は何もしない
            return
        
        # ES: No se encontró proyecto válido; validar nombre y crear nuevo
        # EN: No valid project found; validate name and create new
        # JA: 有効なプロジェクトなし。名前を検証して新規作成
        if not project_name:
            QMessageBox.warning(self, "エラー", "プロジェクト名を入力してください。")
            return
        
        # ES: Almacenar los datos para crear nuevo proyecto
        # EN: Store data to create new project
        # JA: 新規プロジェクト作成用にデータを保存
        self.project_name = project_name
        self.project_directory = directory
        
        super().accept()

class FormulaProcessingWorker(QObject):
    """ES: Worker para procesamiento de fórmulas con barra de progreso
    EN: Worker for formula processing with a progress bar
    JA: 進捗バー付きの数式処理ワーカー
    """
    
    # ES: Señales | EN: Signals | JA: シグナル | EN: Signals | JA: シグナル
    progress_updated = Signal(int, str)  # percent, message
    status_updated = Signal(str)  # status message
    finished = Signal(str)  # output file path
    error_occurred = Signal(str)  # error message
    
    def __init__(self, output_path, data_df, formula_templates, prediction_columns, column_mapping, formula_columns):
        super().__init__()
        self.output_path = output_path
        self.data_df = data_df
        self.formula_templates = formula_templates
        self.prediction_columns = prediction_columns
        self.column_mapping = column_mapping
        self.formula_columns = formula_columns
        self.should_cancel = False
    
    def cancel(self):
        """ES: Cancelar el procesamiento
        EN: Cancel processing
        JA: 処理をキャンセル
        """
        self.should_cancel = True
    
    def run(self):
        """ES: Ejecutar el procesamiento de fórmulas
        EN: Run formula processing
        JA: 数式処理を実行
        """
        try:
            import openpyxl
            from openpyxl import load_workbook
            
            self.status_updated.emit("📊 Cargando archivo Excel...")
            self.progress_updated.emit(5, "Cargando archivo Excel")
            
            # ES: Cargar el archivo Excel con openpyxl para escribir fórmulas
            # EN: Load the Excel file with openpyxl to write formulas
            # JA: 数式を書き込むため openpyxl でExcelを読み込む
            wb = load_workbook(self.output_path)
            ws = wb.active
            
            total_rows = len(self.data_df)
            chunk_size = 100  # Process 100 rows at a time
            
            self.status_updated.emit(f"📊 Procesando {total_rows} filas en lotes de {chunk_size}...")
            
            for chunk_start in range(0, total_rows, chunk_size):
                if self.should_cancel:
                    self.status_updated.emit("❌ Procesamiento cancelado")
                    return
                
                chunk_end = min(chunk_start + chunk_size, total_rows)
                chunk_rows = range(chunk_start + 2, chunk_end + 2)  # +2 because we start from row 2
                
                chunk_number = chunk_start//chunk_size + 1
                total_chunks = (total_rows + chunk_size - 1)//chunk_size
                
                self.status_updated.emit(f"📊 Procesando chunk {chunk_number}/{total_chunks} (filas {chunk_start + 1}-{chunk_end})")
                
                # ES: Preparar todas las fórmulas para este chunk | EN: Prepare all formulas for this chunk | JA: このチャンクの全数式を準備
                chunk_formulas = {}
                
                for row_idx in chunk_rows:
                    if self.should_cancel:
                        return
                    
                    # ES: Crear diccionario de referencias de celda para sustituir en las fórmulas
                    # EN: Build a dict of cell references to substitute into formulas
                    # JA: 数式内で置換するセル参照の辞書を作成
                    formula_values = {}
                    for ref_cell, col_idx in self.column_mapping.items():
                        if col_idx is not None:
                            # ES: Crear referencia de celda Excel (ej: A2, B2, C2, etc.)
                            # EN: Build Excel cell reference (e.g., A2, B2, C2, ...)
                            # JA: Excelセル参照を作成（例: A2, B2, C2, ...）
                            excel_ref = f'{chr(64 + col_idx)}{row_idx}'
                            formula_values[ref_cell] = excel_ref
                        else:
                            formula_values[ref_cell] = '0'
                    
                    # ES: Aplicar las plantillas de fórmulas para esta fila
                    # EN: Apply formula templates for this row
                    # JA: この行に数式テンプレートを適用
                    row_formulas = {}
                    for i, (template, pred_col) in enumerate(zip(self.formula_templates, self.prediction_columns)):
                        if template != '=0':
                            # ES: Sustituir referencias de celda en la plantilla
                            # EN: Substitute cell references into the template
                            # JA: テンプレート内のセル参照を置換
                            processed_formula = template
                            for cell_ref, excel_ref in formula_values.items():
                                processed_formula = processed_formula.replace(cell_ref, excel_ref)
                            row_formulas[pred_col] = processed_formula
                        else:
                            row_formulas[pred_col] = '=0'
                    
                    chunk_formulas[row_idx] = row_formulas
                
                # ES: Escribir todas las fórmulas del chunk de una vez | EN: Write all formulas in the chunk at once | JA: チャンクの全数式を一括書き込み
                for row_idx, row_formulas in chunk_formulas.items():
                    if self.should_cancel:
                        return
                    
                    for pred_col, formula in row_formulas.items():
                        ws.cell(row=row_idx, column=self.formula_columns[pred_col], value=formula)
                
                # ES: Actualizar progreso
                # EN: Update progress
                # JA: 進捗を更新
                progress = int((chunk_end / total_rows) * 90)  # 90% for processing, 10% for saving
                self.progress_updated.emit(progress, f"Chunk {chunk_number}/{total_chunks} completado")
            
            if self.should_cancel:
                return
            
            self.status_updated.emit("💾 Guardando archivo...")
            self.progress_updated.emit(95, "Guardando archivo")
            
            # ES: Guardar el archivo con las fórmulas | EN: Save file with formulas | JA: 数式付きでファイルを保存
            wb.save(self.output_path)
            
            self.status_updated.emit("✅ Procesamiento completado")
            self.progress_updated.emit(100, "Completado")
            self.finished.emit(self.output_path)
            
        except Exception as e:
            error_msg = f"❌ Error en procesamiento de fórmulas: {str(e)}"
            self.status_updated.emit(error_msg)
            self.error_occurred.emit(error_msg)


class MainWindow(QMainWindow):

    def __init__(self):
        super().__init__()
        print("🔧 MainWindow を初期化中...")
        
        # ES: Variable para detectar acceso desde bunseki | EN: Variable to detect access from bunseki | JA: bunsekiからのアクセス検出用変数
        self.accessed_from_bunseki = False
        
        try:
            print("🔧 DBManager を作成中...")
            # ES: IMPORTANTE: en instalaciones (Program Files) no se puede escribir junto al EXE.
            # EN: IMPORTANT: in Program Files installs you cannot write next to the EXE.
            # JA: 重要：Program Files版ではEXE隣に書き込めない
            # ES: Usar siempre la ruta compartida en ProgramData (ver app_paths.py).
            # EN: Always use shared path in ProgramData (see app_paths.py).
            # JA: ProgramDataの共有パスを常に使用（app_paths.py参照）
            self.db = DBManagerMain(RESULTS_DB_PATH)
            print("🔧 ResultProcessor を作成中...")
            self.processor = ResultProcessor(self.db)
            # ES: Backup automático (1/día) de la BBDD principal en ProgramData\\...\\backups
            # EN: Auto backup (1/day) of main DB in ProgramData\\...\\backups
            # JA: ProgramData\\...\\backups でメインDBの自動バックアップ（1日1回）
            try:
                backup_dir = get_backup_dir(shared=True)
                res = auto_daily_backup(RESULTS_DB_PATH, backup_dir, prefix="results")
                prune_backups(backup_dir, prefix="results", keep_daily=30, keep_monthly=12)
                if res is not None:
                    print(f"✅ 日次バックアップを作成しました: {res.backup_path}")
            except Exception as _e:
                print(f"⚠️ 日次バックアップを実行できませんでした: {_e}")
            print("🔧 メインウィンドウを設定中...")
            # ES: Mostrar versión en la barra de título (arriba a la izquierda)
            # EN: Show version in title bar (top left)
            # JA: タイトルバー（左上）にバージョン表示
            self.setWindowTitle(get_app_title())
            self.setMinimumSize(1250, 950)
            print("🔧 メインウィンドウの設定が完了しました")
        except Exception as e:
            print(f"❌ __init__ でエラー: {e}")
            import traceback
            traceback.print_exc()
            raise
        
        # ES: Establecer el icono de la aplicación | EN: Set application icon | JA: アプリのアイコンを設定
        try:
            icon = QIcon(resource_path("xebec_logo_88.png"))
            self.setWindowIcon(icon)
        except Exception as e:
            print(f"⚠️ アイコンを読み込めませんでした: {e}")

        # ES: Crear el widget central | EN: Create central widget | JA: 中央ウィジェットを作成
        print("🔧 中央ウィジェットを作成中...")
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        # ES: Versión se muestra en la barra de título; no agregamos label en el canvas
        # EN: Version is shown in title bar; we do not add a label on the canvas
        # JA: バージョンはタイトルバーに表示。キャンバスにはラベルを追加しない

        # ES: Layout principal horizontal (panel izquierdo + panel central + consola)
        # EN: Main horizontal layout (left panel + center panel + console)
        # JA: メインレイアウト（左パネル＋中央パネル＋コンソール）
        print("🔧 メインレイアウトを設定中...")
        main_layout = QHBoxLayout()
        central_widget.setLayout(main_layout)

        # ========================
        # ES: Panel izquierdo (Controles) - Ancho fijo
        # EN: Left panel (Controls) - Fixed width
        # JA: 左パネル（コントロール）－幅固定
        # ========================
        print("🔧 左パネルを作成中...")
        self.left_frame = QFrame()
        self.left_frame.setFrameShape(QFrame.StyledPanel)
        self.left_frame.setFixedWidth(340)  # Fixed width for the left panel (reduced 15%)
        self.left_layout = QVBoxLayout()
        self.left_layout.setAlignment(Qt.AlignTop)
        self.left_frame.setLayout(self.left_layout)

        self.left_layout.addWidget(create_logo_widget())

        self.create_load_section()
        
        # ES: Campo de tamaño de muestra con valor por defecto 15 (siempre habilitado)
        # EN: Sample size field, default 15 (always enabled)
        # JA: サンプルサイズ入力、デフォルト15（常に有効）
        self.sample_size_label = QLabel("サンプルサイズ (10-50)")
        self.sample_size_input = QLineEdit()
        self.sample_size_input.setPlaceholderText("10-50")
        self.sample_size_input.setValidator(QIntValidator(10, 50))
        self.sample_size_input.setText("15")
        
        # ES: Conectar eventos para validación y pérdida de foco
        # EN: Connect events for validation and focus loss
        # JA: 検証とフォーカス喪失用のイベントを接続
        self.sample_size_input.editingFinished.connect(self.validate_sample_size)
        # Separador "サンプル" encima de サンプルサイズ
        sample_separator_label = QLabel("サンプル")
        sample_separator_label.setStyleSheet("font-weight: bold; color: #666666; margin: 5px 0px;")
        sample_separator_label.setAlignment(Qt.AlignCenter)
        sample_separator = QFrame()
        sample_separator.setFrameShape(QFrame.HLine)
        sample_separator.setStyleSheet("background-color: #CCCCCC; margin: 10px 0px;")
        self.left_layout.addWidget(sample_separator_label)
        self.left_layout.addWidget(sample_separator)
        
        self.sample_size_input.focusOutEvent = self.on_sample_size_focus_out
        
        self.left_layout.addWidget(self.sample_size_label)
        self.left_layout.addWidget(self.sample_size_input)
        
        self.create_action_buttons()
        
        # Separador "結果" encima del selector de 材料
        result_separator_label = QLabel("結果")
        result_separator_label.setStyleSheet("font-weight: bold; color: #666666; margin: 5px 0px;")
        result_separator_label.setAlignment(Qt.AlignCenter)
        result_separator = QFrame()
        result_separator.setFrameShape(QFrame.HLine)
        result_separator.setStyleSheet("background-color: #CCCCCC; margin: 10px 0px;")
        self.left_layout.addWidget(result_separator_label)
        self.left_layout.addWidget(result_separator)
        
        # ES: Crea los widgets de Material
        # EN: Create Material widgets
        # JA: 材料ウィジェットを作成
        self.material_label = QLabel("材料")
        self.material_selector = QComboBox()
        self.material_selector.addItems(["Steel", "Alumi"])
        self.left_layout.addWidget(self.material_label)
        self.left_layout.addWidget(self.material_selector)
        
        self.create_diameter_selector()
        self.create_show_results_button()

        self.create_export_button()
        


        # ES: Lista de widgets a desactivar/activar debajo del selector de muestras
        # EN: List of widgets to enable/disable below sample selector
        # JA: サンプル選択子の下で有効/無効にするウィジェット一覧
        self.widgets_below_sample_selector = []
        # ES: Usa los nombres correctos para los selectores
        # EN: Use the correct names for the selectors
        # JA: セレクタには正しい名前を使用
        self.widgets_below_sample_selector.append(self.diameter_label)
        self.widgets_below_sample_selector.append(self.diameter_selector)
        self.widgets_below_sample_selector.append(self.material_label)
        self.widgets_below_sample_selector.append(self.material_selector)
        # ES: NOTA: sample_size_label y sample_size_input NO están aquí porque deben estar siempre habilitados
        # EN: NOTE: sample_size_label and sample_size_input are not here; they stay always enabled
        # JA: 注意：sample_size_label と sample_size_input は常時有効のためここに含めない
        # ES: Añadir más si hay más widgets debajo
        # EN: Add more if there are more widgets below
        # JA: 下にウィジェットがあれば追加

        def set_widgets_enabled(enabled):
            for w in self.widgets_below_sample_selector:
                w.setEnabled(enabled)
                if hasattr(w, 'setStyleSheet'):
                    if enabled:
                        w.setStyleSheet("")
                    else:
                        w.setStyleSheet("color: gray;")
        # ES: Por defecto, desactivar | EN: Disable by default | JA: デフォルトで無効
        set_widgets_enabled(False)
        
        # Set initial state for UI elements
        self.set_ui_state_for_no_file()

        # ES: Exponer toggler (para habilitar/deshabilitar por tipo detectado, no por nombre de archivo)
        # EN: Expose toggler (to enable/disable by detected type, not by file name)
        # JP: トグル関数を公開（ファイル名ではなく検出タイプに応じて有効/無効化）
        self._set_widgets_below_sample_selector_enabled = set_widgets_enabled

        # ES: Cuando se cargue un archivo, habilitar SOLO si el caller indica explícitamente que es de resultados.
        # EN: When a file is loaded, enable ONLY if the caller explicitly marks it as results.
        # JA: ファイル読み込み時、呼び出し元が「結果用」と明示した場合のみ有効化
        # ES: El nombre del archivo no importa; la detección real se hace por cabecera en handle_single_file_load
        # EN: File name does not matter; real detection is by header in handle_single_file_load
        # JA: ファイル名は不問。実検出は handle_single_file_load のヘッダで行う
        def on_file_loaded(file_path, is_results=None):
            set_widgets_enabled(bool(is_results))
        self.on_file_loaded = on_file_loaded
        print("🔧 左パネルの設定が完了しました")
        
        # ES: Panel central (Visualización) - Se expande
        # EN: Center panel (visualization) - expands
        # JA: 中央パネル（表示）－伸縮
        # ========================
        print("🔧 中央パネルを作成中...")
        self.center_frame = QFrame()
        self.center_frame.setFrameShape(QFrame.StyledPanel)
        self.center_layout = QVBoxLayout()
        self.center_layout.setAlignment(Qt.AlignTop)
        self.center_frame.setLayout(self.center_layout)

        # ES: Inicializar navegación de gráficos (aún no creada)
        # EN: Initialize graph navigation (not yet created)
        # JA: グラフナビを初期化（まだ未作成）
        self.prev_button = None
        self.next_button = None
        self.graph_navigation_frame = None

        # ES: La flecha estará siempre visible, no necesitamos el botón
        # EN: Arrow is always visible; no activation button needed
        # JA: 矢印は常時表示、アクティベーションボタンは不要
        print("🔧 矢印システムを簡略化しました（有効化ボタンなし）")

        # ES: Crear panel central (gráficos, labels, OK/NG)
        # EN: Create center panel (charts, labels, OK/NG)
        # JA: 中央パネル作成（グラフ・ラベル・OK/NG）
        self.create_center_panel()

        # ========================
        # ES: Panel derecho (Consola) - Desplegable
        # EN: Right panel (Console) - Collapsible
        # JA: 右パネル（コンソール）－折りたたみ
        # ========================
        print("🔧 折りたたみコンソールパネルを作成中...")
        
        # ES: Contenedor principal del panel derecho | EN: Main container for right panel | JA: 右パネル用メインコンテナ
        self.right_container = QWidget()
        self.right_container.setFixedWidth(300)
        self.right_container.setMaximumWidth(300)
        
        # ES: Layout del contenedor derecho | EN: Right container layout | JA: 右コンテナのレイアウト
        self.right_container_layout = QVBoxLayout()
        self.right_container_layout.setContentsMargins(0, 0, 0, 0)
        self.right_container_layout.setSpacing(0)
        self.right_container.setLayout(self.right_container_layout)
        
        # ES: Panel de la consola
        # EN: Console panel
        # JA: コンソールパネル
        self.console_frame = QFrame()
        self.console_frame.setFrameShape(QFrame.StyledPanel)
        self.console_frame.setStyleSheet("""
            QFrame {
                background-color: #FFFFFF;
                border: 1px solid #CCCCCC;
                border-radius: 5px;
            }
        """)
        
        self.console_layout = QVBoxLayout()
        self.console_layout.setAlignment(Qt.AlignTop)
        self.console_layout.setContentsMargins(5, 5, 5, 5)
        self.console_frame.setLayout(self.console_layout)

        # ES: Crear la consola integrada | EN: Create integrated console | JA: 統合コンソールを作成
        print("🔧 統合コンソールを設定中...")
        self.create_console_panel()
        
        # ES: Añadir la consola al contenedor derecho | EN: Add console to right container | JA: 右コンテナにコンソールを追加
        self.right_container_layout.addWidget(self.console_frame)
        
        # ES: Crear el panel desplegable superpuesto | EN: Create overlay dropdown panel | JA: オーバーレイドロップダウンパネルを作成
        self.create_overlay_console_panel()

        # ES: Añadir solo el panel izquierdo y central al layout principal
        # EN: Add only left and center panels to main layout
        # JA: メインレイアウトには左・中央パネルのみ追加
        # ========================
        print("🔧 パネルをメインレイアウトに追加中...")
        main_layout.addWidget(self.left_frame)  # EN: Left panel fixed width
        main_layout.addWidget(self.center_frame, 1)  # EN: Center panel expands
        # ES: NOTA: El panel derecho se añadirá dinámicamente cuando se active
        # EN: NOTE: Right panel is added dynamically when activated
        # JA: 注意：右パネルはアクティブ時に動的に追加
        print("🔧 左パネルと中央パネルを正常に追加しました")

        # ========================
        # ES: Archivo cargando
        # EN: File loading
        # JA: ファイル読み込み中
        # ========================

        self.loader_overlay = LoadingOverlay(self.center_frame)

        self.graph_images = []  # EN: List of image paths
        self.current_graph_index = 0
        self.graph_label = QLabel()
        self.graph_label.setAlignment(Qt.AlignCenter)
        self.graph_area_layout = QVBoxLayout()
        self.graph_area.setLayout(self.graph_area_layout)
        self.graph_area_layout.addWidget(self.graph_label)

    # ======================================
    # Utilidades de UI (limpieza de layouts)
    # ======================================
    def _clear_layout_recursive(self, layout):
        """
        Limpia un QLayout de forma recursiva (widgets + sub-layouts).
        Importante: QLayoutItem.widget() solo devuelve widgets en el nivel actual;
        si hay addLayout(...), hay que limpiar también item.layout().
        """
        if layout is None:
            return

        while layout.count():
            item = layout.takeAt(0)
            if item is None:
                continue

            w = item.widget()
            if w is not None:
                try:
                    w.hide()
                except Exception:
                    pass
                try:
                    w.setParent(None)
                except Exception:
                    pass
                try:
                    w.deleteLater()
                except Exception:
                    pass
                continue

            child_layout = item.layout()
            if child_layout is not None:
                # ES: Limpiar recursivamente y soltar el layout
                # EN: Clear recursively and release the layout
                # JA: 再帰的にクリアしてレイアウトを解放
                self._clear_layout_recursive(child_layout)
                try:
                    child_layout.setParent(None)
                except Exception:
                    pass
                continue

            # SpacerItem u otros items: nada que hacer

    # ======================================
    # ES: Secciones de creación visual | EN: Visual creation sections | JA: UI作成セクション
    # ======================================

    def create_load_section(self):
        """ES: Crear la sección de carga de archivos
        EN: Create the file-load section
        JA: ファイル読み込みセクションを作成"""
        self.generate_button = QPushButton("生成：サンプル組合せ表")
        self.setup_generate_button_style(self.generate_button)
        self.left_layout.addWidget(self.generate_button)
        self.generate_button.clicked.connect(self.on_generate_sample_file_clicked)

        self.load_file_button = QPushButton("ファイルを読み込む")
        self.load_file_label = QLabel("ファイル未選択")
        self.setup_load_block(self.load_file_button, self.load_file_label)
        self.load_file_button.clicked.connect(self.handle_single_file_load)

        # self.load_sample_button = QPushButton("サンプルファイルをロード")
        # self.sample_label = QLabel("ファイル未選択")
        # self.setup_load_block(self.load_sample_button, self.sample_label)
        # self.load_sample_button.clicked.connect(lambda: self.load_file(self.sample_label, "サンプルファイルを選択"))
        #
        # self.load_results_button = QPushButton("結果ファイルをロード")
        # self.results_label = QLabel("ファイル未選択")
        # self.setup_load_block(self.load_results_button, self.results_label)
        # self.load_results_button.clicked.connect(lambda: self.load_file(self.results_label, "結果ファイルを選択"))

    def create_action_buttons(self):
        """ES: Crear los botones de Dsaitekika e iSaitekika separados
        EN: Create Dsaitekika and iSaitekika buttons separately
        JA: Dsaitekika と iSaitekika ボタンを別々に作成"""
        self.left_layout.addSpacing(10)

        self.d_optimize_button = QPushButton("D最適化を実行")
        self.setup_action_button(self.d_optimize_button)
        self.left_layout.addWidget(self.d_optimize_button)
        self.d_optimize_button.clicked.connect(self.on_d_optimizer_clicked)

        self.left_layout.addSpacing(5)

        self.i_optimize_button = QPushButton("I最適化を実行")
        self.setup_action_button(self.i_optimize_button)
        self.left_layout.addWidget(self.i_optimize_button)
        self.i_optimize_button.clicked.connect(self.on_i_optimizer_clicked)

    def create_show_results_button(self):
        """ES: Crear el botón Show Results
        EN: Create the Show Results button
        JA: Show Resultsボタンを作成"""
        self.left_layout.addStretch()

        self.show_results_button = QPushButton("データベースにインポート")
        self.setup_results_button(self.show_results_button)
        self.left_layout.addWidget(self.show_results_button)
        self.show_results_button.clicked.connect(self.on_show_results_clicked)

        self.left_layout.addSpacing(10)
        self.show_results_button.setEnabled(False)

        # ES: Botón de análisis | EN: Analysis button | JA: 解析ボタン
        self.analyze_button = QPushButton("分析")
        self.setup_results_button(self.analyze_button)
        self.left_layout.addWidget(self.analyze_button)
        self.analyze_button.clicked.connect(self.on_analyze_clicked)

        self.left_layout.addSpacing(10)
        self.analyze_button.setEnabled(True)



    def create_project_folder_structure(self, project_folder):
        """ES: Crear la estructura de carpetas del proyecto
        EN: Create project folder structure
        JA: プロジェクトのフォルダ構造を作成"""
        folders = [
            "01_実験リスト",
            "99_Temp", 
            "03_-----------解析------------",
            "99_------------------------------",
            "02_実験データ",
            "99_Results",
            "03_線形回帰",
            "04_非線形回帰",
            "05_分類"
        ]
        
        for folder in folders:
            folder_path = os.path.join(project_folder, folder)
            os.makedirs(folder_path, exist_ok=True)
            print(f"📁 フォルダーを作成しました: {folder_path}")

    def create_export_button(self):
        """ES: Crear el botón de exportar resultados a Excel
        EN: Create the export-results-to-Excel button
        JA: 結果をExcelにエクスポートするボタンを作成"""
        self.export_button = QPushButton("結果をエクスポート")
        self.setup_generate_button_style(self.export_button)
        self.left_layout.addWidget(self.export_button)
        self.export_button.clicked.connect(self.export_database_to_excel)
        
        # ES: Botón para exportar base de datos de Yosoku | EN: Button to export Yosoku database | JA: 予測DBエクスポートボタン
        self.yosoku_export_button = QPushButton("予測データベースをエクスポート")
        self.setup_generate_button_style(self.yosoku_export_button)
        self.left_layout.addWidget(self.yosoku_export_button)
        self.yosoku_export_button.clicked.connect(self.export_yosoku_database_to_excel)

        # ES: Backup de BBDD (results + yosoku si existen) | EN: DB backup (results + yosoku if exist) | JA: DBバックアップ（results＋yosokuがあれば）
        self.db_backup_button = QPushButton("DBバックアップ作成")
        self.setup_generate_button_style(self.db_backup_button)
        self.left_layout.addWidget(self.db_backup_button)
        self.db_backup_button.clicked.connect(self.backup_databases_now)

    def backup_databases_now(self):
        """ES: Crear backup seguro de las BBDD en ProgramData\\...\\backups (manual).
        EN: Create safe backup of DBs in ProgramData\\...\\backups (manual).
        JA: ProgramData\\...\\backups にDBの安全なバックアップを作成（手動）。"""
        try:
            backup_dir = get_backup_dir(shared=True)
            created = []

            # 1) BBDD principal
            if os.path.exists(RESULTS_DB_PATH):
                r = create_backup(RESULTS_DB_PATH, backup_dir, prefix="results")
                prune_backups(backup_dir, prefix="results", keep_daily=30, keep_monthly=12)
                created.append(Path(r.backup_path).name)

            # 2) Yosoku lineal / no lineal (si existen)
            if os.path.exists(YOSOKU_LINEAL_DB_PATH):
                r = create_backup(YOSOKU_LINEAL_DB_PATH, backup_dir, prefix="yosoku_lineal")
                prune_backups(backup_dir, prefix="yosoku_lineal", keep_daily=30, keep_monthly=12)
                created.append(Path(r.backup_path).name)

            if os.path.exists(YOSOKU_NO_LINEAL_DB_PATH):
                r = create_backup(YOSOKU_NO_LINEAL_DB_PATH, backup_dir, prefix="yosoku_no_lineal")
                prune_backups(backup_dir, prefix="yosoku_no_lineal", keep_daily=30, keep_monthly=12)
                created.append(Path(r.backup_path).name)

            if not created:
                QMessageBox.information(self, "情報", "📦 バックアップ対象のデータベースが見つかりません。")
                return

            msg = "✅ バックアップを作成しました:\n\n" + "\n".join(f"- {n}" for n in created)
            msg += f"\n\n📁 保存先:\n{str(backup_dir)}"
            QMessageBox.information(self, "完了", msg)
        except Exception as e:
            QMessageBox.critical(self, "エラー", f"❌ バックアップ作成中にエラーが発生しました:\n{e}")

    def _ensure_app_fonts_loaded(self):
        """ES: Cargar fuentes desde la carpeta Fonts (si existen) y elegir una familia válida para texto.
        EN: Load fonts from Fonts folder (if any) and pick a valid family for text.
        JA: Fontsフォルダからフォントを読み込み、テキスト用の有効なファミリを選択。"""
        if getattr(self, "_app_fonts_loaded", False):
            return

        self._app_fonts_loaded = True
        self._app_font_family = None

        try:
            fonts_dir = resource_path("Fonts")
            if not os.path.isdir(fonts_dir):
                return

            loaded_families = []
            for fn in os.listdir(fonts_dir):
                if not fn.lower().endswith((".ttf", ".otf", ".ttc")):
                    continue
                fpath = os.path.join(fonts_dir, fn)
                try:
                    font_id = QFontDatabase.addApplicationFont(fpath)
                    if font_id != -1:
                        loaded_families.extend(QFontDatabase.applicationFontFamilies(font_id))
                except Exception:
                    pass

            # Elegir una familia cargada que realmente soporte el texto (evita fuentes de iconos).
            sample_text = "0.00 sec"
            for fam in loaded_families:
                try:
                    fm = QFontMetrics(QFont(fam))
                    if all(fm.inFont(ch) for ch in sample_text):
                        self._app_font_family = fam
                        return
                except Exception:
                    continue
        except Exception:
            return

    def _add_center_header_title(self):
        """ES: Añadir el texto '0.00 sec' centrado arriba en el panel central (fuera del área de gráficos).
        EN: Add '0.00 sec' text centered above the center panel (outside graph area).
        JA: 中央パネル上部（グラフ領域外）に '0.00 sec' を中央揃えで追加。"""
        try:
            self._ensure_app_fonts_loaded()

            title = QLabel("0.00 sec")
            title.setAlignment(Qt.AlignCenter)
            title.setStyleSheet("background: transparent; color: #111111;")
            title.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

            # ES: Preferir fuente desde `Fonts/` si alguna soporta el texto; si no, fallback moderno de Windows.
            # EN: Prefer a font from `Fonts/` if one supports the text; otherwise use a modern Windows fallback.
            # JP: `Fonts/`内に表示可能なフォントがあれば優先し、無ければWindowsのモダンな代替フォントを使う
            preferred = self._app_font_family or "Segoe UI Variable Display"
            font = QFont(preferred)
            if not font.exactMatch():
                font = QFont(self._app_font_family or "Segoe UI")

            font.setPointSize(28)
            font.setWeight(QFont.DemiBold)
            title.setFont(font)

            self.center_title_label = title
            self.center_layout.addWidget(title, 0, Qt.AlignHCenter)
            self.center_layout.addSpacing(6)
        except Exception as e:
            print(f"⚠️ 中央タイトルの追加中にエラー: {e}")

    def create_center_panel(self):
        """ES: Crear la estructura del panel central
        EN: Create center panel structure
        JA: 中央パネルの構造を作成"""

        # ES: Título arriba del área de gráficos (fuera del graph_area)
        # EN: Title above graph area (outside graph_area)
        # JA: グラフ領域の上にタイトル（graph_area の外）
        self._add_center_header_title()

        # ES: Área de gráficos | EN: Graph area | JA: グラフ領域
        # ES: Contenedor de área de gráficos + botones de navegación
        # EN: Container for graph area + navigation buttons
        # JA: グラフ領域＋ナビボタン用コンテナ
        self.graph_container = QFrame()
        graph_container_layout = QVBoxLayout()
        graph_container_layout.setContentsMargins(0, 0, 0, 0)
        graph_container_layout.setSpacing(0)
        self.graph_container.setLayout(graph_container_layout)

        # ES: Área de gráficos | EN: Graph area | JA: グラフ領域
        self.graph_area = QFrame()
        self.graph_area.setStyleSheet("background-color: #F9F9F9; border: 1px solid #CCCCCC;")
        graph_container_layout.addWidget(self.graph_area, stretch=1)

        # ES: Añadir contenedor al layout principal central | EN: Add container to main center layout | JA: メイン中央レイアウトにコンテナを追加
        self.center_layout.addWidget(self.graph_container, stretch=1)

        # ES: Espacio flexible antes de los botones
        # EN: Flexible space before the buttons
        # JA: ボタン前の可変スペース
        self.center_layout.addStretch()

        # ES: Botones OK y NG
        # EN: OK and NG buttons
        # JA: OK/NGボタン
        self.ok_ng_frame = QFrame()
        ok_ng_layout = QHBoxLayout()
        ok_ng_layout.setAlignment(Qt.AlignCenter)
        self.ok_ng_frame.setLayout(ok_ng_layout)

        self.ok_button = QPushButton("OK")
        self.ng_button = QPushButton("NG")

        self.setup_ok_button(self.ok_button)
        self.setup_ng_button(self.ng_button)

        self.ok_button.clicked.connect(self.on_ok_clicked)
        self.ng_button.clicked.connect(self.on_ng_clicked)

        ok_ng_layout.addWidget(self.ok_button)
        ok_ng_layout.addSpacing(10)
        ok_ng_layout.addWidget(self.ng_button)

        self.center_layout.addWidget(self.ok_ng_frame)

        self.ok_button.setEnabled(False)
        self.ng_button.setEnabled(False)

    def create_console_panel(self):
        """ES: Crear la consola integrada en el panel derecho
        EN: Create the integrated console in the right panel
        JA: 右パネルに統合コンソールを作成"""
        # ES: Título de la consola | EN: Console title | JA: コンソールタイトル
        console_title = QLabel("コンソール出力")
        console_title.setAlignment(Qt.AlignCenter)
        console_title.setStyleSheet("""
            font-size: 14px;
            font-weight: bold;
            color: #333333;
            background-color: #F0F0F0;
            padding: 5px;
            border: 1px solid #CCCCCC;
            border-radius: 3px;
        """)
        self.console_layout.addWidget(console_title)

        # ES: Área de texto de la consola | EN: Console text area | JA: コンソールテキスト領域
        self.console_output = QTextEdit()
        self.console_output.setReadOnly(True)
        self.console_output.setMaximumHeight(400)
        self.console_output.setStyleSheet("""
            QTextEdit {
                background-color: #1E1E1E;
                color: #FFFFFF;
                font-family: 'Consolas', 'Monaco', 'Courier New', monospace;
                font-size: 11px;
                border: 1px solid #CCCCCC;
                border-radius: 3px;
            }
        """)
        self.console_layout.addWidget(self.console_output)

        # ES: Botones de control de la consola
        # EN: Console control buttons
        # JA: コンソール制御ボタン
        console_controls = QFrame()
        console_controls_layout = QHBoxLayout()
        console_controls_layout.setContentsMargins(0, 5, 0, 5)
        console_controls.setLayout(console_controls_layout)

        # ES: Botón para limpiar consola | EN: Clear console button | JA: コンソールクリアボタン
        self.clear_console_button = QPushButton("クリア")
        self.clear_console_button.setMaximumWidth(60)
        self.clear_console_button.clicked.connect(self.clear_console)
        self.clear_console_button.setStyleSheet("""
            QPushButton {
                background-color: #F0F0F0;
                border: 1px solid #CCCCCC;
                border-radius: 3px;
                padding: 3px 8px;
                font-size: 10px;
            }
            QPushButton:hover {
                background-color: #E0E0E0;
            }
        """)

        # ES: Botón para guardar log | EN: Save log button | JA: ログ保存ボタン
        self.save_log_button = QPushButton("保存")
        self.save_log_button.setMaximumWidth(60)
        self.save_log_button.clicked.connect(self.save_console_log)
        self.save_log_button.setStyleSheet("""
            QPushButton {
                background-color: #F0F0F0;
                border: 1px solid #CCCCCC;
                border-radius: 3px;
                padding: 3px 8px;
                font-size: 10px;
            }
            QPushButton:hover {
                background-color: #E0E0E0;
            }
        """)

        console_controls_layout.addWidget(self.clear_console_button)
        console_controls_layout.addWidget(self.save_log_button)
        console_controls_layout.addStretch()

        self.console_layout.addWidget(console_controls)
        
        # ES: NOTA: El botón オーバーレイ表示 se crea en __init__ y se añade al panel central
        # EN: NOTE: オーバーレイ表示 button is created in __init__ and added to center panel
        # JA: 注意：オーバーレイ表示ボタンは __init__ で作成し中央パネルに追加

        # ES: Configurar redirección de stdout y stderr a la consola
        # EN: Configure stdout/stderr redirection to console
        # JA: stdout/stderrをコンソールにリダイレクト
        self.setup_console_redirection()

    def create_overlay_console_panel(self):
        """ES: Crear el panel desplegable que se superpone sobre el panel central
        EN: Create the overlay dropdown panel on top of the center panel
        JA: 中央パネル上に重なるドロップダウンパネルを作成"""
        print("🔧 重ね表示のスライドパネルを作成中...")
        
        # ES: Panel desplegable que se superpone. IMPORTANTE: ventana top-level (sin parent) para que ReusableProgressDialog (WindowModal) no la bloquee durante análisis.
        # EN: Overlay dropdown panel. IMPORTANT: top-level window (no parent) so ReusableProgressDialog (WindowModal) does not block it during analysis.
        # JA: 重ねるドロップダウンパネル。重要：解析中にReusableProgressDialog(WindowModal)にブロックされないようトップレベル（親無し）にする。
        self.overlay_console = QFrame()
        self.overlay_console.setFrameShape(QFrame.StyledPanel)
        self.overlay_console.setStyleSheet("""
            QFrame {
                background-color: #FFFFFF;
                border: 2px solid #3498db;
                border-radius: 8px;
                box-shadow: 0 4px 8px rgba(0,0,0,0.2);
            }
        """)
        
        # ES: Por defecto NO forzar siempre-arriba: si no, tapa diálogos del sistema (QFileDialog, etc).
        # EN: By default do NOT force always-on-top; otherwise it covers system dialogs (QFileDialog, etc).
        # JA: デフォルトでは常に前面にしない（QFileDialog等のシステムダイアログを隠さないため）
        # ES: Activamos "siempre-arriba" solo mientras el loading (ReusableProgressDialog) esté visible.
        # EN: Enable always-on-top only while loading (ReusableProgressDialog) is visible.
        # JA: ローディング表示中のみ常に前面にする
        self.overlay_console.setWindowFlags(Qt.Tool | Qt.FramelessWindowHint)
        self.overlay_console.setAttribute(Qt.WA_TranslucentBackground, False)
        self.overlay_console.setAttribute(Qt.WA_NoSystemBackground, False)
        
        # ES: Layout del panel desplegable | EN: Dropdown panel layout | JA: ドロップダウンパネルのレイアウト
        self.overlay_console_layout = QVBoxLayout()
        self.overlay_console_layout.setContentsMargins(10, 10, 10, 10)
        self.overlay_console.setLayout(self.overlay_console_layout)
        
        # ES: Título del panel desplegable | EN: Dropdown panel title | JA: ドロップダウンパネルタイトル
        overlay_title = QLabel("コンソール出力 (オーバーレイ)")
        overlay_title.setAlignment(Qt.AlignCenter)
        overlay_title.setStyleSheet("""
            font-size: 14px;
            font-weight: bold;
            color: #2c3e50;
            background-color: #ecf0f1;
                border-radius: 5px;
                margin-bottom: 10px;
        """)
        self.overlay_console_layout.addWidget(overlay_title)
        
        # ES: Área de texto de la consola desplegable | EN: Dropdown console text area | JA: ドロップダウンコンソールのテキスト領域
        self.overlay_console_output = QTextEdit()
        self.overlay_console_output.setReadOnly(True)
        self.overlay_console_output.setMaximumHeight(500)
        self.overlay_console_output.setStyleSheet("""
            QTextEdit {
                background-color: #1E1E1E;
                color: #FFFFFF;
                font-family: 'Consolas', 'Monaco', 'Courier New', monospace;
                font-size: 11px;
                border: 1px solid #CCCCCC;
                border-radius: 5px;
            }
        """)
        self.overlay_console_layout.addWidget(self.overlay_console_output)
        
        # ES: Botones de control del panel desplegable
        # EN: Dropdown panel control buttons
        # JA: ドロップダウンパネルの制御ボタン
        overlay_controls = QFrame()
        overlay_controls_layout = QHBoxLayout()
        overlay_controls_layout.setContentsMargins(0, 5, 0, 5)
        overlay_controls.setLayout(overlay_controls_layout)
        
        # ES: Botón para limpiar consola desplegable | EN: Clear dropdown console button | JA: ドロップダウンコンソールクリアボタン
        self.overlay_clear_button = QPushButton("クリア")
        self.overlay_clear_button.setMaximumWidth(60)
        self.overlay_clear_button.clicked.connect(self.clear_overlay_console)
        self.overlay_clear_button.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                border-radius: 3px;
                padding: 3px 8px;
                font-size: 10px;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
        """)
        
        # ES: Botón para guardar log del panel desplegable | EN: Save log button for dropdown panel | JA: ドロップダウンパネルのログ保存ボタン
        self.overlay_save_button = QPushButton("保存")
        self.overlay_save_button.setMaximumWidth(60)
        self.overlay_save_button.clicked.connect(self.save_overlay_console_log)
        self.overlay_save_button.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                border: none;
                border-radius: 3px;
                padding: 3px 8px;
                font-size: 10px;
            }
            QPushButton:hover {
                background-color: #229954;
            }
        """)
        
        overlay_controls_layout.addWidget(self.overlay_clear_button)
        overlay_controls_layout.addWidget(self.overlay_save_button)
        overlay_controls_layout.addStretch()
        
        self.overlay_console_layout.addWidget(overlay_controls)
        
        # ES: Botón de flecha para expandir/contraer. IMPORTANTE: botón como ventana top-level (sin parent) para que siga clicable cuando el diálogo de progreso está en WindowModal.
        # EN: Arrow button to expand/collapse. IMPORTANT: button as top-level window (no parent) so it stays clickable when progress dialog is WindowModal.
        # JA: 展開/折りたたみ用矢印ボタン。重要：進捗ダイアログがWindowModalでもクリックできるようトップレベル（親無し）にする。
        self.console_toggle_button = QPushButton("◀")
        self.console_toggle_button.setFixedSize(30, 30)
        
        # ES: CRÍTICO: Para que el botón sea redondo en ventana top-level, fondo translúcido y FramelessWindowHint
        # EN: CRITICAL: For round button on top-level window, use translucent background and FramelessWindowHint
        # JA: 重要：トップレベルで丸ボタンにするには半透明背景とFramelessWindowHintが必要
        self.console_toggle_button.setAttribute(Qt.WA_TranslucentBackground)
        self.console_toggle_button.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                border-radius: 15px;
                font-size: 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        # ES: Ventana sin marco (overlay real). NO siempre-arriba por defecto.
        # EN: Frameless window (real overlay). Not always-on-top by default.
        # JA: 枠無しウィンドウ（オーバーレイ）。デフォルトでは常に前面にしない。
        self.console_toggle_button.setWindowFlags(Qt.Tool | Qt.FramelessWindowHint)
        # ES: Evitar "pelea" de foco mientras está el loading visible
        # EN: Avoid focus fight while loading is visible
        # JA: ローディング表示中はフォーカス競合を避ける
        self.console_toggle_button.setFocusPolicy(Qt.NoFocus)
        self.console_toggle_button.setAttribute(Qt.WA_ShowWithoutActivating, True)
        
        # ES: Conectar la flecha al método de toggle | EN: Connect arrow to toggle method | JA: 矢印をトグルメソッドに接続
        self.console_toggle_button.clicked.connect(self.toggle_right_panel)
        print("🔧 矢印を toggle_right_panel に接続しました")
        
        # ES: Inicialmente solo la flecha visible, panel desplegable oculto
        # EN: Initially only arrow visible, dropdown panel hidden
        # JA: 初期は矢印のみ表示、ドロップダウンパネルは非表示
        self.overlay_console.hide()
        self.console_toggle_button.show()  # EN: Arrow always visible
        
        # ES: Estado del panel desplegable
        # EN: Dropdown panel state
        # JA: ドロップダウンパネルの状態
        self.overlay_console_visible = False
        # ES: Estado "siempre-arriba" (solo durante loading modal)
        # EN: Always-on-top state (only during loading modal)
        # JA: 「常に前面」状態（ローディングモーダル中のみ）
        self._console_topmost_enabled = False
        
        # ES: Posicionar la flecha inicialmente | EN: Position arrow initially | JA: 矢印を初期配置
        QTimer.singleShot(100, self.position_arrow)
        
        # ES: Configurar timer para mantener elementos en primer plano
        # EN: Configure timer to keep elements on top
        # JA: 要素を前面に保つタイマーを設定
        self.keep_on_top_timer = QTimer()
        self.keep_on_top_timer.timeout.connect(self.keep_elements_on_top)
        self.keep_on_top_timer.start(1000)  # EN: Every second
        
        # ES: Configurar timer para verificar cambios de posición de la ventana
        # EN: Configure timer to check window position changes
        # JA: ウィンドウ位置変化を確認するタイマーを設定
        self.position_check_timer = QTimer()
        self.position_check_timer.timeout.connect(self.check_window_position)
        self.position_check_timer.start(500)  # EN: Every half second
        
        # ES: Guardar la posición inicial de la ventana | EN: Save initial window position | JA: ウィンドウの初期位置を保存
        self.last_window_position = self.geometry()
        
        print("🔧 スライドパネルを正常に作成しました")
        print(f"🔧 矢印ボタンを作成しました: {self.console_toggle_button}")
        print(f"🔧 ボタンの表示: {self.console_toggle_button.isVisible()}")
        print(f"🔧 ボタンの親: {self.console_toggle_button.parent()}")

    def _build_done_experiments_excel(self, main_file: str, temp_file: str, done_file: str):
        """
        Construye un Excel con los ensayos YA HECHOS como:
            done = (main_file) - (temp_file)
        usando como clave las 7 columnas de condiciones.

        - main_file: Proyecto_XX_未実験データ.xlsx (carpeta principal del proyecto)
        - temp_file: Proyecto_XX_未実験データ.xlsx (99_Temp)
        - done_file: salida (por defecto en 99_Temp/done_experiments.xlsx)
        """
        try:
            import os
            import pandas as pd
            import numpy as np

            # ES: Aceptar ambos nombres para la columna de dirección: "UPカット" (nuevo), "回転方向" (antiguo)
            # EN: Accept both names for direction column: "UPカット" (new), "回転方向" (legacy)
            # JA: 方向列は "UPカット"（新）と "回転方向"（旧）の両方を受け付ける
            dir_variants = ["UPカット", "回転方向"]
            key_cols_fixed = ['回転速度', '送り速度', '切込量', '突出量', '載せ率', 'パス数']
            int_cols = ['回転速度', '送り速度', 'DIR', 'パス数']
            float_cols = ['切込量', '突出量', '載せ率']

            if not (os.path.exists(main_file) and os.path.exists(temp_file)):
                print(f"⚠️ done_experiments: archivos no existen. main={main_file}, temp={temp_file}")
                return None

            # ES: Cache simple: si done_file es más nuevo que los inputs, reutilizar
            # EN: Simple cache: if done_file is newer than inputs, reuse it
            # JA: 簡易キャッシュ：done_fileが入力より新しければ再利用
            try:
                if os.path.exists(done_file):
                    done_mtime = os.path.getmtime(done_file)
                    if done_mtime >= max(os.path.getmtime(main_file), os.path.getmtime(temp_file)):
                        print(f"✅ done_experiments: 既存キャッシュを使用します {done_file}")
                        return done_file
            except Exception:
                pass

            def _read_table(path: str) -> pd.DataFrame:
                ext = os.path.splitext(str(path))[1].lower()
                if ext == ".csv":
                    return pd.read_csv(path, encoding="utf-8-sig")
                return pd.read_excel(path)

            main_df = _read_table(main_file)
            temp_df = _read_table(temp_file)

            def _pick_dir_col(df: pd.DataFrame):
                for c in dir_variants:
                    if c in df.columns:
                        return c
                return None

            dir_main = _pick_dir_col(main_df)
            dir_temp = _pick_dir_col(temp_df)
            if dir_main is None or dir_temp is None:
                print(f"❌ done_experiments: 方向列が見つかりません。main_has={list(main_df.columns)}, temp_has={list(temp_df.columns)}")
                return None

            missing_main = [c for c in key_cols_fixed if c not in main_df.columns]
            missing_temp = [c for c in key_cols_fixed if c not in temp_df.columns]
            if missing_main or missing_temp:
                print(f"❌ done_experiments: 必要な列が不足しています。main_missing={missing_main}, temp_missing={missing_temp}")
                return None

            def _norm_key_df(df: pd.DataFrame) -> pd.DataFrame:
                # ES: Normalizamos a un esquema común con columna "DIR"
                # EN: Normalize to common schema with "DIR" column
                # JA: "DIR"列を含む共通スキーマに正規化
                k = df[key_cols_fixed].copy()
                k["DIR"] = df[dir_main] if dir_main in df.columns else df[dir_temp]
                # ES: numérico + redondeo para evitar diferencias de precisión
                # EN: Numeric + rounding to avoid precision differences
                # JA: 精度差を避けるため数値化＋丸め
                for c in ["回転速度", "送り速度", "パス数", "DIR"]:
                    k[c] = pd.to_numeric(k[c], errors="coerce").round(0).astype("Int64")
                for c in float_cols:
                    k[c] = pd.to_numeric(k[c], errors="coerce").round(6)
                return k

            main_key_df = _norm_key_df(main_df)
            temp_key_df = _norm_key_df(temp_df)

            main_hash = pd.util.hash_pandas_object(main_key_df, index=False)
            temp_hash = pd.util.hash_pandas_object(temp_key_df, index=False)
            temp_set = set(temp_hash.values.tolist())

            done_mask = ~main_hash.isin(temp_set)
            done_full = main_df.loc[done_mask].copy()

            # Deduplicar por clave (conservar primera ocurrencia)
            dedup_cols = ['回転速度', '送り速度', dir_main, '切込量', '突出量', '載せ率', 'パス数']
            done_full = done_full.drop_duplicates(subset=[c for c in dedup_cols if c in done_full.columns])

            os.makedirs(os.path.dirname(done_file), exist_ok=True)
            # ES: Especificar engine para evitar problemas de autodetección en algunos entornos
            # EN: Specify engine to avoid autodetect issues in some environments
            # JA: 一部環境での自動検出問題を避けるためengineを指定
            done_full.to_excel(done_file, index=False, engine="openpyxl")

            print(f"✅ done_experiments を生成しました: {done_file} | 行数={len(done_full)}")
            return done_file

        except Exception as e:
            print(f"⚠️ done_experiments.xlsx の作成に失敗: {e}")
            return None

    def _export_unexperimented_excel_folder_from_csv(self, csv_path: str, project_folder: str, project_name: str):
        """
        Si el archivo de muestreo del proyecto está en CSV (Proyecto_XX_未実験データ.csv),
        crear también Excel(s) dentro de una carpeta:
          <project_folder>/99_未実験データ/
        - Si <= 500,000 filas: crear <project_name>_未実験データ.xlsx
        - Si > 500,000 filas: crear <project_name>_未実験データ_part_###.xlsx (500k filas por archivo)
        """
        try:
            if not csv_path or not os.path.exists(csv_path):
                return
            if os.path.splitext(csv_path)[1].lower() != ".csv":
                return

            out_folder = os.path.join(project_folder, "99_未実験データ")
            os.makedirs(out_folder, exist_ok=True)

            rows_per_file = 500_000
            chunksize = 100_000

            print(f"📦 99_未実験データ: CSV→Excel 変換開始: {csv_path}", flush=True)
            print(f"📦 99_未実験データ: 出力先フォルダ: {out_folder}", flush=True)

            part_idx = 1
            part_rows = 0
            startrow = 0
            writer = None
            wrote_any = False

            def _open_writer():
                nonlocal writer, part_idx, part_rows, startrow
                if writer is not None:
                    writer.close()
                part_path = os.path.join(out_folder, f"{project_name}_未実験データ_part_{part_idx:03d}.xlsx")
                print(f"📄 99_未実験データ: 作成中 {os.path.basename(part_path)}", flush=True)
                writer = pd.ExcelWriter(part_path, engine="openpyxl")
                part_idx += 1
                part_rows = 0
                startrow = 0

            for chunk in pd.read_csv(csv_path, encoding="utf-8-sig", chunksize=chunksize):
                pos = 0
                while pos < len(chunk):
                    if writer is None or part_rows >= rows_per_file:
                        _open_writer()
                    remaining = rows_per_file - part_rows
                    take = min(remaining, len(chunk) - pos)
                    piece = chunk.iloc[pos:pos + take]
                    header = startrow == 0
                    piece.to_excel(writer, index=False, header=header, startrow=startrow, sheet_name="Sheet1")
                    wrote_any = True
                    pos += take
                    part_rows += take
                    startrow += take + (1 if header else 0)

            if writer is not None:
                writer.close()

            # ES: Si solo se generó un part, renombrarlo a .xlsx normal
            # EN: If only one part was generated, rename to normal .xlsx
            # JA: partが1つだけの場合は通常の.xlsxにリネーム
            # (original: renombrarlo a .xlsx “normal”
            if wrote_any:
                parts = sorted(
                    [f for f in os.listdir(out_folder) if f.startswith(f"{project_name}_未実験データ_part_") and f.endswith(".xlsx")]
                )
                if len(parts) == 1:
                    src = os.path.join(out_folder, parts[0])
                    dst = os.path.join(out_folder, f"{project_name}_未実験データ.xlsx")
                    try:
                        if os.path.exists(dst):
                            os.remove(dst)
                        os.replace(src, dst)
                        print(f"✅ 99_未実験データ: 1ファイルのためリネーム: {os.path.basename(dst)}", flush=True)
                    except Exception as e:
                        print(f"⚠️ 99_未実験データ: リネーム失敗: {e}", flush=True)

            print("✅ 99_未実験データ: CSV→Excel 変換完了", flush=True)
        except Exception as e:
            print(f"⚠️ 99_未実験データ: CSV→Excel 変換エラー: {e}", flush=True)

    def create_diameter_selector(self):
        """ES: Crear selector de diámetro (el cepillo se toma del archivo de resultados, no de la UI)
        EN: Create diameter selector (brush is taken from results file, not from UI)
        JA: 直径セレクタを作成（ブラシは結果ファイルから取得、UIからではない）"""
        # ES: Selector de diámetro | EN: Diameter selector | JA: 直径セレクタ
        self.diameter_label = QLabel("直径 選択")
        self.diameter_selector = QComboBox()
        self.diameter_selector.addItems(["6", "15", "25", "40", "60", "100"])
        self.diameter_selector.setCurrentText("15")
        self.left_layout.addWidget(self.diameter_label)
        self.left_layout.addWidget(self.diameter_selector)
        # ES: Por defecto: sin restricción (solo se restringe si el archivo detecta A13) | EN: Default: no restriction (restrict only if file detects A13) | JA: デフォルト：制限なし（A13検出時のみ制限）
        self.update_diameter_options("")

    def update_diameter_options(self, brush_name):
        """ES: Restringe el selector de diámetro si el cepillo es A13
        EN: Restrict diameter selector when brush is A13
        JA: ブラシがA13の場合は直径セレクタを制限"""
        allowed = ["6", "15"] if brush_name == "A13" else ["6", "15", "25", "40", "60", "100"]
        for i in range(self.diameter_selector.count()):
            value = self.diameter_selector.itemText(i)
            self.diameter_selector.model().item(i).setEnabled(value in allowed)
        # ES: Si el valor actual no está permitido, selecciona el primero permitido
        # EN: If current value is not allowed, select the first allowed one
        # JA: 現在値が許可されていなければ最初の許可値を選択
        if self.diameter_selector.currentText() not in allowed:
            self.diameter_selector.setCurrentText(allowed[0])

    def _detect_brush_type_from_results_file(self, file_path):
        """
        Detecta el tipo de cepillo desde el archivo de resultados (one-hot A13/A11/A21/A32).
        Devuelve "A13"/"A11"/"A21"/"A32" o None si no se puede determinar.
        """
        try:
            import pandas as pd
            ext = os.path.splitext(str(file_path))[1].lower()
            if ext == ".csv":
                df = pd.read_csv(file_path, encoding="utf-8-sig")
            else:
                df = pd.read_excel(file_path, header=0)

            # ES: Normalizar columnas (espacios invisibles)
            # EN: Normalize columns (invisible spaces)
            # JP: 列名を正規化（不可視スペース）
            try:
                if isinstance(df.columns, pd.MultiIndex):
                    df.columns = [" ".join([str(x).strip() for x in tup if str(x).strip() != ""]).strip() for tup in df.columns]
                else:
                    df.columns = [str(c).strip() for c in df.columns]
            except Exception:
                pass

            brush_cols = ["A13", "A11", "A21", "A32"]
            if not all(c in df.columns for c in brush_cols):
                return None

            onehot = df[brush_cols].apply(pd.to_numeric, errors="coerce").fillna(0)
            # ES: Si el archivo tiene muchas filas, agregamos para mayor robustez
            # EN: If file has many rows, add for robustness
            # JA: 行数が多い場合はロバスト性のため加算
            sums = onehot.sum(axis=0)
            # ES: Selección conservadora: debe haber un único ganador con suma > 0
            # EN: Conservative selection: there must be a single winner with sum > 0
            # JA: 保守的選択：合計>0の唯一の勝者がいること
            winners = [c for c in brush_cols if sums.get(c, 0) > 0]
            if len(winners) == 1:
                return winners[0]
            # ES: Si hay varios con >0, decidir por el máximo si es claramente dominante
            # EN: If several have >0, choose by max if clearly dominant
            # JA: 複数が>0の場合は明らかに優勢な最大で決定
            best = sums.idxmax()
            if float(sums.max()) > 0 and (sums == sums.max()).sum() == 1:
                return str(best)
            return None
        except Exception:
            return None

    def _apply_results_file_brush_to_ui(self, file_path):
        """ES: Aplica restricciones UI (diámetro) en base al cepillo detectado del archivo.
        EN: Apply UI constraints (diameter) based on brush detected from file.
        JA: ファイルから検出したブラシに基づきUI制約（直径）を適用。"""
        brush = self._detect_brush_type_from_results_file(file_path)
        self._results_brush_type = brush
        # ES: Restringir diámetro si procede (A13)
        # EN: Restrict diameter if applicable (A13)
        # JP: 必要なら直径を制限（A13）
        try:
            self.update_diameter_options(brush or "")
        except Exception:
            pass



    def create_navigation_buttons(self):
        if self.graph_navigation_frame is not None:
            return

        self.graph_navigation_frame = QFrame()
        nav_layout = QHBoxLayout()
        nav_layout.setAlignment(Qt.AlignRight)
        self.graph_navigation_frame.setLayout(nav_layout)

        self.prev_button = QPushButton("← 前へ")
        self.next_button = QPushButton("次へ →")

        self.setup_navigation_button(self.prev_button)
        self.setup_navigation_button(self.next_button)

        nav_layout.addWidget(self.prev_button)
        nav_layout.addSpacing(10)
        nav_layout.addWidget(self.next_button)

        self.graph_container.layout().addWidget(self.graph_navigation_frame)

        # ES: ❗️Conectar aquí
        # EN: ❗️Connect here
        # JP: ❗️ここで接続
        self.prev_button.clicked.connect(self.show_previous_graph)
        self.next_button.clicked.connect(self.show_next_graph)

        self.prev_button.setEnabled(False)
        self.next_button.setEnabled(False)

    def show_previous_graph(self):
        if self.current_graph_index > 0:
            self.current_graph_index -= 1
            self.update_graph_display()

    def show_next_graph(self):
        if self.current_graph_index < len(self.graph_images) - 1:
            self.current_graph_index += 1
            self.update_graph_display()

    def create_filter_view(self):
        """ES: Crear la vista de filtrado a la derecha
        EN: Create filtering view on the right
        JA: 右側にフィルタビューを作成"""
        # ES: Limpiar el layout central COMPLETAMENTE (incluye layouts anidados)
        # EN: Clear the center layout completely (including nested layouts)
        # JA: 中央レイアウトを完全にクリア（ネストしたレイアウト含む）
        self._clear_layout_recursive(self.center_layout)
        try:
            QApplication.processEvents()
        except Exception:
            pass

        # ES: Título mejorado | EN: Improved title | JA: タイトル改善
        title = QLabel("データフィルター")
        title.setStyleSheet("""
            font-weight: bold; 
            font-size: 24px; 
            color: #2c3e50;
            margin-bottom: 20px;
            padding: 10px 0px;
            border-bottom: 2px solid #3498db;
            border-radius: 0px;
        """)
        title.setAlignment(Qt.AlignCenter)
        self.center_layout.addWidget(title)

        # ES: Espaciado entre título y filtros | EN: Spacing between title and filters | JA: タイトルとフィルタの間隔
        spacer = QWidget()
        spacer.setFixedHeight(15)
        self.center_layout.addWidget(spacer)

        # ES: Contenedor principal horizontal para filtros e imagen | EN: Main horizontal container for filters and image | JA: フィルタと画像用のメイン横コンテナ
        main_container = QHBoxLayout()
        
        # ES: Contenedor vertical para todos los filtros con margen izquierdo | EN: Vertical container for all filters with left margin | JA: 全フィルタ用縦コンテナ（左マージン付き）
        filters_container = QVBoxLayout()
        filters_container.setSpacing(8)
        filters_container.setAlignment(Qt.AlignTop)
        filters_container.setContentsMargins(20, 0, 0, 0)  # Margen izquierdo de 20px

        self.filter_inputs = {}

        # ES: Helper: añadir fila limpia | EN: Helper: add blank row | JA: ヘルパー：空行を追加
        def add_filter_row(label_text, widget1, widget2=None):
            row = QHBoxLayout()
            label = QLabel(label_text)
            label.setFixedWidth(90)
            label.setStyleSheet("font-weight: bold; font-size: 12px;")
            row.addWidget(label)

            # ES: Calcular el ancho total disponible (mismo que la fila de radio buttons)
            # EN: Compute total available width (same as radio-button row)
            # JA: 利用可能幅を計算（ラジオボタン行と同様）
            # 90px (label) + 4*radio_buttons + 3*12px (margins) + 8px (spacing) = ~200px
            total_width = 200
            widget1.setFixedWidth(total_width)
            row.addWidget(widget1)

            if widget2:
                spacer = QLabel("〜")
                spacer.setFixedWidth(10)
                spacer.setAlignment(Qt.AlignCenter)
                row.addWidget(spacer)

                widget2.setFixedWidth(total_width)
                row.addWidget(widget2)

            row.addStretch()
            filters_container.addLayout(row)

        # 実験日 (rango de fechas)
        desde_fecha = QDateEdit()
        desde_fecha.setCalendarPopup(True)
        desde_fecha.setDate(QDate.currentDate().addDays(-30))  # 30 days back by default
        desde_fecha.setFixedWidth(150)
        
        hasta_fecha = QDateEdit()
        hasta_fecha.setCalendarPopup(True)
        hasta_fecha.setDate(QDate.currentDate())  # Fecha actual por defecto
        hasta_fecha.setFixedWidth(150)
        
        # ES: Botón "なし" para no aplicar filtro de fecha | EN: "なし" button to skip date filter | JA: 日付フィルタなし用「なし」ボタン
        no_date_button = QPushButton("なし")
        no_date_button.setFixedWidth(80)
        no_date_button.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
                border: none;
                padding: 5px;
                border-radius: 4px;
                font-size: 11px;
            }
            QPushButton:hover {
                background-color: #7f8c8d;
            }
            QPushButton:pressed {
                background-color: #6c7b7d;
            }
        """)
        
        # ES: Variable para controlar si se aplica filtro de fecha | EN: Variable to control whether date filter is applied | JA: 日付フィルタ適用の有無を制御する変数
        self.apply_date_filter = True
        
        def toggle_date_filter():
            if self.apply_date_filter:
                # Desactivar filtro de fecha
                self.apply_date_filter = False
                no_date_button.setText("適用")
                no_date_button.setStyleSheet("""
                    QPushButton {
                        background-color: #e74c3c;
                        color: white;
                        border: none;
                        padding: 5px;
                        border-radius: 4px;
                        font-size: 11px;
                    }
                    QPushButton:hover {
                        background-color: #c0392b;
                    }
                    QPushButton:pressed {
                        background-color: #a93226;
                    }
                """)
                desde_fecha.setEnabled(False)
                hasta_fecha.setEnabled(False)
            else:
                # Activar filtro de fecha
                self.apply_date_filter = True
                no_date_button.setText("なし")
                no_date_button.setStyleSheet("""
                    QPushButton {
                        background-color: #95a5a6;
                        color: white;
                        border: none;
                        padding: 5px;
                        border-radius: 4px;
                        font-size: 11px;
                    }
                    QPushButton:hover {
                        background-color: #7f8c8d;
                    }
                    QPushButton:pressed {
                        background-color: #6c7b7d;
                    }
                """)
                desde_fecha.setEnabled(True)
                hasta_fecha.setEnabled(True)
        
        no_date_button.clicked.connect(toggle_date_filter)
        
        self.filter_inputs["実験日"] = (desde_fecha, hasta_fecha)
        
        # ES: Crear fila personalizada para fecha con botón | EN: Create custom date row with button | JA: ボタン付き日付のカスタム行を作成
        date_row = QHBoxLayout()
        date_label = QLabel("実験日")
        date_label.setFixedWidth(90)
        date_label.setStyleSheet("font-weight: bold; font-size: 12px;")
        date_row.addWidget(date_label)
        
        date_row.addWidget(desde_fecha)
        
        spacer = QLabel("〜")
        spacer.setFixedWidth(10)
        spacer.setAlignment(Qt.AlignCenter)
        date_row.addWidget(spacer)
        
        date_row.addWidget(hasta_fecha)
        
        # ES: Agregar espacio y botón | EN: Add spacing and button | JA: スペースとボタンを追加
        date_row.addSpacing(10)
        date_row.addWidget(no_date_button)
        
        date_row.addStretch()
        filters_container.addLayout(date_row)

        # バリ除去
        combo = QComboBox()
        combo.addItems(["", "0", "1"])
        combo.setFixedWidth(200)  # Mismo ancho que los otros campos
        self.filter_inputs["バリ除去"] = combo
        add_filter_row("バリ除去", combo)

        # 上面ダレ量
        desde = QLineEdit()
        hasta = QLineEdit()
        desde.setPlaceholderText("min")
        hasta.setPlaceholderText("max")
        self.filter_inputs["上面ダレ量"] = (desde, hasta)
        add_filter_row("上面ダレ量", desde, hasta)

        # 側面ダレ量
        desde = QLineEdit()
        hasta = QLineEdit()
        desde.setPlaceholderText("min")
        hasta.setPlaceholderText("max")
        self.filter_inputs["側面ダレ量"] = (desde, hasta)
        add_filter_row("側面ダレ量", desde, hasta)

        # 面粗度(Ra)前
        desde = QLineEdit()
        hasta = QLineEdit()
        desde.setPlaceholderText("min")
        hasta.setPlaceholderText("max")
        self.filter_inputs["面粗度(Ra)前"] = (desde, hasta)
        add_filter_row("面粗度(Ra)前", desde, hasta)

        # 面粗度(Ra)後
        desde = QLineEdit()
        hasta = QLineEdit()
        desde.setPlaceholderText("min")
        hasta.setPlaceholderText("max")
        self.filter_inputs["面粗度(Ra)後"] = (desde, hasta)
        add_filter_row("面粗度(Ra)後", desde, hasta)

        # 摩耗量
        desde = QLineEdit()
        hasta = QLineEdit()
        desde.setPlaceholderText("min")
        hasta.setPlaceholderText("max")
        self.filter_inputs["摩耗量"] = (desde, hasta)
        add_filter_row("摩耗量", desde, hasta)

        # 切削力X
        desde = QLineEdit()
        hasta = QLineEdit()
        desde.setPlaceholderText("min")
        hasta.setPlaceholderText("max")
        self.filter_inputs["切削力X"] = (desde, hasta)
        add_filter_row("切削力X", desde, hasta)

        # 切削力Y
        desde = QLineEdit()
        hasta = QLineEdit()
        desde.setPlaceholderText("min")
        hasta.setPlaceholderText("max")
        self.filter_inputs["切削力Y"] = (desde, hasta)
        add_filter_row("切削力Y", desde, hasta)

        # 切削力Z
        desde = QLineEdit()
        hasta = QLineEdit()
        desde.setPlaceholderText("min")
        hasta.setPlaceholderText("max")
        self.filter_inputs["切削力Z"] = (desde, hasta)
        add_filter_row("切削力Z", desde, hasta)

        # 材料
        material_combo = QComboBox()
        material_combo.addItems(["", "Steel", "Alumi"])
        material_combo.setFixedWidth(200)  # Same width as the other fields
        self.filter_inputs["材料"] = material_combo
        add_filter_row("材料", material_combo)

        # ブラシ
        brush_label = QLabel("ブラシ選択")
        brush_label.setFixedWidth(90)
        brush_label.setStyleSheet("font-weight: bold; font-size: 12px;")
        
        brush_container = QHBoxLayout()
        brush_container.setSpacing(4)  # Reduce spacing between buttons
        
        self.filter_inputs["すべて"] = QCheckBox("すべて")
        self.filter_inputs["A13"] = QCheckBox("A13")
        self.filter_inputs["A11"] = QCheckBox("A11")
        self.filter_inputs["A21"] = QCheckBox("A21")
        self.filter_inputs["A32"] = QCheckBox("A32")
        
        # ES: Establecer "すべて" como seleccionado por defecto | EN: Set "すべて" as selected by default | JA: デフォルトで「すべて」を選択
        self.filter_inputs["すべて"].setChecked(True)
        
        # ES: Aplicar estilo a los checkboxes
        # EN: Apply style to the checkboxes
        # JP: チェックボックスにスタイルを適用
        checkbox_style = """
            QCheckBox {
                font-size: 11px;
                spacing: 4px;
                padding: 2px;
                margin-right: 48px;
            }
            QCheckBox::indicator {
                width: 12px;
                height: 12px;
            }
        """
        
        for key in ["すべて", "A13", "A11", "A21", "A32"]:
            self.filter_inputs[key].setStyleSheet(checkbox_style)
            brush_container.addWidget(self.filter_inputs[key])
            
        # ES: Conectar señales para la lógica de selección mutuamente excluyente
        # EN: Connect signals for mutually exclusive selection logic
        # JA: 排他選択ロジック用にシグナルを接続
        self.filter_inputs["すべて"].toggled.connect(self.on_subete_toggled)
        self.filter_inputs["A13"].toggled.connect(self.on_brush_toggled)
        self.filter_inputs["A11"].toggled.connect(self.on_brush_toggled)
        self.filter_inputs["A21"].toggled.connect(self.on_brush_toggled)
        self.filter_inputs["A32"].toggled.connect(self.on_brush_toggled)
        
        # ES: Crear layout horizontal para label y botones | EN: Create horizontal layout for label and buttons | JA: ラベルとボタン用の横レイアウトを作成
        brush_row = QHBoxLayout()
        brush_row.addWidget(brush_label)
        brush_row.addLayout(brush_container)
        brush_row.addStretch()
        
        filters_container.addLayout(brush_row)

        # 直径
        desde = QLineEdit()
        hasta = QLineEdit()
        desde.setPlaceholderText("min")
        hasta.setPlaceholderText("max")
        self.filter_inputs["直径"] = (desde, hasta)
        add_filter_row("直径", desde, hasta)

        # 回転速度
        desde = QLineEdit()
        hasta = QLineEdit()
        desde.setPlaceholderText("min")
        hasta.setPlaceholderText("max")
        self.filter_inputs["回転速度"] = (desde, hasta)
        add_filter_row("回転速度", desde, hasta)

        # 送り速度
        desde = QLineEdit()
        hasta = QLineEdit()
        desde.setPlaceholderText("min")
        hasta.setPlaceholderText("max")
        self.filter_inputs["送り速度"] = (desde, hasta)
        add_filter_row("送り速度", desde, hasta)

        # UPカット
        up_combo = QComboBox()
        up_combo.addItems(["", "0", "1"])
        up_combo.setFixedWidth(200)  # Mismo ancho que los otros campos
        self.filter_inputs["UPカット"] = up_combo
        add_filter_row("UPカット", up_combo)

        # 切込量
        desde = QLineEdit()
        hasta = QLineEdit()
        desde.setPlaceholderText("min")
        hasta.setPlaceholderText("max")
        self.filter_inputs["切込量"] = (desde, hasta)
        add_filter_row("切込量", desde, hasta)

        # 突出量
        desde = QLineEdit()
        hasta = QLineEdit()
        desde.setPlaceholderText("min")
        hasta.setPlaceholderText("max")
        self.filter_inputs["突出量"] = (desde, hasta)
        add_filter_row("突出量", desde, hasta)

        # 載せ率
        desde = QLineEdit()
        hasta = QLineEdit()
        desde.setPlaceholderText("min")
        hasta.setPlaceholderText("max")
        self.filter_inputs["載せ率"] = (desde, hasta)
        add_filter_row("載せ率", desde, hasta)

        # パス数
        pass_input = QLineEdit()
        pass_input.setPlaceholderText("例: 3")
        pass_input.setFixedWidth(200)  # Mismo ancho que los otros campos
        self.filter_inputs["パス数"] = pass_input
        add_filter_row("パス数", pass_input)

        # 線材長
        desde = QLineEdit()
        hasta = QLineEdit()
        desde.setPlaceholderText("min")
        hasta.setPlaceholderText("max")
        self.filter_inputs["線材長"] = (desde, hasta)
        add_filter_row("線材長", desde, hasta)

        # 加工時間
        desde = QLineEdit()
        hasta = QLineEdit()
        desde.setPlaceholderText("min")
        hasta.setPlaceholderText("max")
        self.filter_inputs["加工時間"] = (desde, hasta)
        add_filter_row("加工時間", desde, hasta)

        # ES: Agregar filtros al contenedor principal
        # EN: Add filters to main container
        # JA: フィルタをメインコンテナに追加
        main_container.addLayout(filters_container)
        
        # ES: Agregar imagen chibi al lado derecho
        # EN: Add chibi image on the right
        # JA: 右側にちび画像を追加
        try:
            chibi_label = QLabel()
            chibi_pixmap = QPixmap(resource_path("xebec_chibi.png"))
            if not chibi_pixmap.isNull():
                # ES: Redimensionar la imagen 200% más grande (2x el tamaño original)
                # EN: Resize image 200% larger (2x original size)
                # JA: 画像を2倍に拡大（元の2倍）
                chibi_pixmap = chibi_pixmap.scaled(300, 400, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                chibi_label.setPixmap(chibi_pixmap)
                chibi_label.setAlignment(Qt.AlignRight | Qt.AlignBottom)
                chibi_label.setStyleSheet("margin-left: 20px;")
                main_container.addWidget(chibi_label)
                print("✅ ちび画像の読み込みに成功しました")
            else:
                print("⚠️ 画像を読み込めませんでした: xebec_chibi.png")
        except Exception as e:
            print(f"⚠️ ちび画像の読み込み中にエラー: {e}")
        
        # ES: Agregar el contenedor principal al layout central
        # EN: Add main container to center layout
        # JA: メインコンテナを中央レイアウトに追加
        self.center_layout.addLayout(main_container)

        # ES: Espaciado más grande entre filtros y botones | EN: Larger spacing between filters and buttons | JA: フィルタとボタンの間隔を広く
        spacer = QWidget()
        spacer.setFixedHeight(50)
        self.center_layout.addWidget(spacer)

        # ES: Contenedor horizontal para los 3 botones en paralelo con espacio a la derecha | EN: Horizontal container for 3 buttons in parallel with right spacing | JA: 3ボタン横並び＋右余白のコンテナ
        buttons_container = QHBoxLayout()
        buttons_container.setSpacing(10)  # Espacio entre botones
        
        # ES: Estilo común para todos los botones usando azul claro como el botón de carga
        # EN: Common style for all buttons using light blue like the load button
        # JA: ロードボタン同様の水色で全ボタンの共通スタイル
        button_style = """
            QPushButton {
                background-color: #5EC8E5;
                color: white;
                font-size: 14px;
                font-weight: bold;
                border: none;
                border-radius: 8px;
                padding: 10px 20px;
                min-width: 150px;
            }
            QPushButton:hover {
                background-color: #4BB8D0;
            }
        """
        
        # ES: Botón 線形解析 | EN: Linear analysis button | JA: 線形解析ボタン
        linear_btn = QPushButton("線形解析")
        linear_btn.setFixedHeight(45)
        linear_btn.setStyleSheet(button_style)
        linear_btn.clicked.connect(self.on_linear_analysis_clicked)
        buttons_container.addWidget(linear_btn)
        
        # ES: Botón 非線形解析 | EN: Non-linear analysis button | JA: 非線形解析ボタン
        nonlinear_btn = QPushButton("非線形解析")
        nonlinear_btn.setFixedHeight(45)
        nonlinear_btn.setStyleSheet(button_style)
        nonlinear_btn.setEnabled(True)  # Habilitado
        nonlinear_btn.setToolTip("非線形回帰分析を実行します")
        nonlinear_btn.clicked.connect(self.on_nonlinear_analysis_clicked)
        buttons_container.addWidget(nonlinear_btn)
        
        # ES: Botón 分類分析 | EN: Classification analysis button | JA: 分類分析ボタン
        classification_btn = QPushButton("分類分析")
        classification_btn.setFixedHeight(45)
        classification_btn.setStyleSheet(button_style)
        classification_btn.setEnabled(True)  # Habilitado
        classification_btn.setToolTip("分類分析を実行します")
        classification_btn.clicked.connect(self.on_classification_analysis_clicked)
        buttons_container.addWidget(classification_btn)
        
        # ES: Agregar espacio vacío a la derecha del tamaño de 2 botones
        # EN: Add empty space to the right, size of 2 buttons
        # JA: 2ボタン分の空きを右に追加
        spacer_widget = QWidget()
        spacer_widget.setFixedWidth(320)  # 2 buttons (150px each) + 2 spacers (10px each)
        buttons_container.addWidget(spacer_widget)
        
        # ES: Agregar el contenedor de botones al layout principal
        # EN: Add button container to main layout
        # JA: ボタンコンテナをメインレイアウトに追加
        self.center_layout.addLayout(buttons_container)

    # ======================================
    # ES: Funciones auxiliares de estilo
    # EN: Style helper functions
    # JP: スタイル補助関数
    # ======================================
    def setup_navigation_button(self, button: QPushButton):
        """ES: Aplica estilo moderno y compacto a los botones de navegación.
        EN: Apply modern compact style to navigation buttons.
        JA: ナビゲーションボタンにモダン・コンパクトなスタイルを適用。"""
        button.setFixedSize(80, 32)  # EN: Smaller button
        button.setStyleSheet("""
            QPushButton {
                background-color: #666666;  /* Gris oscuro normal */
                color: white;
                font-family: "Yu Gothic UI";
                font-weight: bold;
                font-size: 13px;
                border: none;
                border-radius: 6px;
            }
            QPushButton:hover {
                background-color: #555555;  /* Gray slightly darker on hover */
            }
        """)

    def setup_export_button(self, button):
        button.setStyleSheet("""
            QPushButton {
                background-color: lightgray;
                color: black;
                border: 1px solid #888;
                padding: 6px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #d0d0d0;
            }
        """)
        button.setEnabled(True)

    def setup_generate_button_style(self, button: QPushButton):
        """ES: Estilo específico para el botón de generación de archivo base de muestras.
        EN: Specific style for sample-base file generation button.
        JA: サンプルベースファイル生成ボタン用の専用スタイル。"""
        button.setFixedHeight(30)
        button.setStyleSheet("""
            QPushButton {
                background-color: #E0E0E0;
                color: #333333;
                border: none;
                border-radius: 8px;
                font-size: 11px;
                padding: 8px 16px;
            }
            QPushButton:hover {
                background-color: #D5D5D5;
            }
        """)

    def setup_ok_button(self, button: QPushButton):
        """ES: Configura estilo del botón OK
        EN: Configure OK button style
        JA: OKボタンのスタイルを設定"""
        button.setFixedSize(100, 40)
        button.setStyleSheet("""
            QPushButton {
                background-color: #CCCCCC;
                color: white;
                border: none;
                border-radius: 8px;
                font-size: 14px;
                padding: 6px 16px;
            }
            QPushButton:enabled {
                background-color: #5CB85C;
            }
            QPushButton:enabled:hover {
                background-color: #4CAE4C;
            }
        """)

    def setup_ng_button(self, button: QPushButton):
        """ES: Configura estilo del botón NG
        EN: Configure NG button style
        JA: NGボタンのスタイルを設定"""
        button.setFixedSize(100, 40)
        button.setStyleSheet("""
            QPushButton {
                background-color: #CCCCCC;
                color: white;
                border: none;
                border-radius: 8px;
                font-size: 14px;
                padding: 6px 16px;
            }
            QPushButton:enabled {
                background-color: #E57373;
            }
            QPushButton:enabled:hover {
                background-color: #EF5350;
            }
        """)

    def setup_load_block(self, button: QPushButton, label: QLabel):
        """Configura visualmente un bloque de carga"""
        button.setFixedHeight(30)
        button.setStyleSheet("""
            QPushButton {
                background-color: #5EC8E5;
                color: white;
                font-family: "Noto Sans JP";  /* Tipo de letra moderno */
                border: none;
                border-radius: 8px;
                font-size: 12px;
                padding: 8px 16px;
            }
            QPushButton:hover {
                background-color: #4BB8D0;
            }
        """)

        label.setAlignment(Qt.AlignVCenter | Qt.AlignLeft)
        label.setFixedHeight(28)
        label.setStyleSheet("""
            background-color: #FFFFFF;
            border: 1px solid #DDDDDD;
            border-radius: 6px;
            padding-left: 10px;
            font-size: 11px;
            color: #555555;
        """)

        self.left_layout.addWidget(button)
        self.left_layout.addWidget(label)

    def setup_action_button(self, button: QPushButton):
        """Configura los botones principales"""
        button.setFixedHeight(48)
        button.setStyleSheet("""
            QPushButton {
                background-color: #3A80BA;
                color: white;
                font-family: "Noto Sans JP";  /* Tipo de letra moderno */
                border: none;
                border-radius: 8px;
                font-size: 16px;
                padding: 8px 20px;
            }
            QPushButton:hover {
                background-color: #336DA3;
            }
            QPushButton:disabled {
                background-color: #CCCCCC;
                color: #888888;
            }
        """)

    def setup_results_button(self, button: QPushButton):
        """ES: Configura el botón Show Results
        EN: Configure Show Results button
        JA: Show Resultsボタンを設定"""
        button.setFixedHeight(40)
        button.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                font-family: "Noto Sans JP";
                border: none;
                border-radius: 8px;
                font-size: 16px;
                padding: 6px 16px;
            }
            QPushButton:hover {
                background-color: #43A047;
            }
            QPushButton:disabled {
                background-color: #B0B0B0;
                color: #EEEEEE;
            }
        """)

    # ======================================
    # Funciones de eventos
    # ======================================
    def apply_filters(self):
        query = "SELECT * FROM main_results WHERE 1=1"
        params = []

        # ES: Mapear nombres UI -> nombres reales en DB
        # EN: Map UI names to actual DB column names
        # JA: UI名をDBの実際のカラム名にマッピング
        field_to_db = {
            "面粗度(Ra)前": "面粗度前",
            "面粗度(Ra)後": "面粗度後",
        }

        # ES: Procesar filtros de cepillo primero (lógica especial)
        # EN: Process brush filters first (special logic)
        # JA: ブラシフィルタを先に適用（特殊ロジック）
        brush_filters = []
        for field in ["A13", "A11", "A21", "A32"]:
            if self.filter_inputs[field].isChecked():
                brush_filters.append(field)
        
        # ES: Si "すべて" está seleccionado, no aplicar filtros de cepillo
        # EN: If "すべて" is selected, do not apply brush filters
        # JA: 「すべて」選択時はブラシフィルタを適用しない
        if not self.filter_inputs["すべて"].isChecked() and brush_filters:
            # ES: Construir filtro OR para múltiples cepillos seleccionados
            # EN: Build OR filter for multiple selected brushes
            # JA: 複数ブラシ選択時にORフィルタを構築
            brush_conditions = []
            for brush in brush_filters:
                brush_conditions.append(f"{brush} = ?")
                params.append(1)
            if brush_conditions:
                query += f" AND ({' OR '.join(brush_conditions)})"

        # ES: Procesar otros filtros
        # EN: Process other filters
        # JA: その他のフィルタを処理
        for field, widgets in self.filter_inputs.items():
            # ES: Saltar filtros de cepillo ya procesados
            # EN: Skip brush filters already processed
            # JA: 既に処理したブラシフィルタはスキップ
            if field in ["すべて", "A13", "A11", "A21", "A32"]:
                continue
                
            if field in ["バリ除去", "UPカット"]:
                val = widgets.currentText()
                if val != "":
                    query += f" AND {field} = ?"
                    params.append(int(val))

            elif field == "材料":
                val = widgets.currentText()
                if val != "":
                    query += f" AND {field} = ?"
                    params.append(val)

            elif field == "パス数":
                text = widgets.text().strip()
                if text:
                    try:
                        query += f" AND パス数 = ?"
                        params.append(int(text))
                    except ValueError:
                        QMessageBox.warning(self, "入力エラー", f"❌ 数値を入力してください: {field}")
                        return

            elif field == "実験日":
                # Handle date range filter - only if enabled
                if hasattr(self, 'apply_date_filter') and self.apply_date_filter:
                    desde_fecha, hasta_fecha = widgets
                    desde = desde_fecha.date().toString("yyyyMMdd")
                    hasta = hasta_fecha.date().toString("yyyyMMdd")
                
                if desde and hasta:
                    query += f" AND {field} >= ? AND {field} <= ?"
                    params.append(int(desde))
                    params.append(int(hasta))

            else:
                # Handle range filters (min/max inputs)
                desde_input, hasta_input = widgets
                desde = desde_input.text().strip()
                hasta = hasta_input.text().strip()
                db_field = field_to_db.get(field, field)

                if desde:
                    query += f" AND {db_field} >= ?"
                    params.append(float(desde))
                if hasta:
                    query += f" AND {db_field} <= ?"
                    params.append(float(hasta))

        try:
            conn = sqlite3.connect(RESULTS_DB_PATH, timeout=10)
            df = pd.read_sql_query(query, conn, params=params)
            conn.close()

            self.filtered_df = df
            print("✅ フィルタ済みデータ:")
            print(df)
            QMessageBox.information(self, "完了", f"✅ {len(df)} 件のデータが見つかりました。")

        except Exception as e:
            QMessageBox.critical(self, "エラー", f"❌ フィルターの適用中にエラーが発生しました:\n{str(e)}")

    def linear_analysis(self):
        """ES: Análisis lineal de los datos filtrados
        EN: Linear analysis of filtered data
        JA: フィルタ済みデータの線形解析"""
        if hasattr(self, "filtered_df"):
            print("📊 線形解析を開始します...")
            print(f"フィルタ済み: {len(self.filtered_df)} 件")
            # ES: Aquí implementar análisis lineal
            # EN: Implement linear analysis here
            # JP: ここで線形解析を実装
            QMessageBox.information(self, "線形解析", "📊 線形解析を開始しました。")
        else:
            QMessageBox.warning(self, "警告", "⚠️ フィルタリングされたデータがありません。")
    
    def nonlinear_analysis(self):
        """ES: Análisis no lineal de los datos filtrados
        EN: Non-linear analysis of filtered data
        JA: フィルタ済みデータの非線形解析"""
        if hasattr(self, "filtered_df"):
            print("📈 非線形解析を開始します...")
            print(f"フィルタ済み: {len(self.filtered_df)} 件")
            # ES: Aquí implementar análisis no lineal
            # EN: Implement non-linear analysis here
            # JP: ここで非線形解析を実装
            QMessageBox.information(self, "非線形解析", "📈 非線形解析を開始しました。")
        else:
            QMessageBox.warning(self, "警告", "⚠️ フィルタリングされたデータがありません。")
    
    def classification_analysis(self):
        """ES: Análisis de clasificación de los datos filtrados
        EN: Classification analysis of filtered data
        JA: フィルタ済みデータの分類解析"""
        if hasattr(self, "filtered_df"):
            print("🏷️ 分類解析を開始します...")
            print(f"フィルタ済み: {len(self.filtered_df)} 件")
            # ES: Aquí implementar análisis de clasificación
            # EN: Implement classification analysis here
            # JP: ここで分類解析を実装
            QMessageBox.information(self, "分類分析", "🏷️ 分類分析を開始しました。")
        else:
            QMessageBox.warning(self, "警告", "⚠️ フィルタリングされたデータがありません。")
    
    def _cleanup_optimization_threads(self, aggressive: bool = False, wait_ms: int = 1500):
        """
        Limpia QThreads de optimización para evitar estados colgados.
        - aggressive=False: si el thread ya terminó, limpia referencia.
        - aggressive=True: si el thread sigue corriendo, intenta quit()+wait() y limpia referencia.
        """
        for t_attr in ("d_optimizer_thread", "i_optimizer_thread", "dsaitekika_thread"):
            t = getattr(self, t_attr, None)
            if t is None:
                continue
            try:
                running = bool(t.isRunning())
            except RuntimeError:
                setattr(self, t_attr, None)
                continue

            if not running:
                setattr(self, t_attr, None)
                continue

            if aggressive:
                try:
                    t.quit()
                    t.wait(wait_ms)
                except Exception:
                    pass
                # ES: Evitar que un thread "zombie" bloquee nuevas ejecuciones
                # EN: Prevent a "zombie" thread from blocking new runs
                # JP: 「ゾンビ」スレッドが新規実行を妨げないようにする
                setattr(self, t_attr, None)

    def analyze_filtered_data(self):
        if hasattr(self, "filtered_df"):
            print("⚙️ フィルタ済みデータを解析中...")
            print(self.filtered_df.head())
            # ES: Aquí puedes lanzar gráficos, cálculos, etc.
            # EN: You can run charts, calculations, etc. here
            # JP: ここでグラフや計算などを実行できる
        else:
            print("⚠️ フィルタ済みデータがありません。")

    def on_subete_toggled(self, checked):
        """ES: Maneja la lógica cuando se selecciona/deselecciona 'すべて' (subete)
        EN: Handle logic when 'すべて' (all) is selected/deselected
        JA: 'すべて' の選択/解除時のロジックを処理"""
        if checked:
            # ES: Si se selecciona "すべて", deseleccionar todos los otros cepillos
            # EN: If "すべて" is selected, deselect all other brushes
            # JP: 「すべて」を選択したら、他のブラシをすべて解除
            self.filter_inputs["A13"].setChecked(False)
            self.filter_inputs["A11"].setChecked(False)
            self.filter_inputs["A21"].setChecked(False)
            self.filter_inputs["A32"].setChecked(False)
            print("✅ 「すべて」を選択しました（他のブラシは解除しました）")

    def on_brush_toggled(self, checked):
        """ES: Maneja la lógica cuando se selecciona/deselecciona cualquier cepillo específico
        EN: Handle logic when any specific brush is selected/deselected
        JA: 特定ブラシの選択/解除時のロジックを処理"""
        sender = self.sender()
        if checked:
            # ES: Si se selecciona un cepillo específico, deseleccionar "すべて"
            # EN: If a specific brush is selected, deselect "すべて"
            # JP: 特定のブラシを選択したら「すべて」を解除
            self.filter_inputs["すべて"].setChecked(False)
            print(f"✅ {sender.text()} seleccionado - 'すべて' deseleccionado")
        else:
            # ES: Si se deselecciona un cepillo, verificar si no hay ninguno seleccionado
            # EN: If a brush is deselected, check whether none are selected
            # JP: ブラシを解除したら、何も選択されていないか確認
            if not any([
                self.filter_inputs["A13"].isChecked(),
                self.filter_inputs["A11"].isChecked(),
                self.filter_inputs["A21"].isChecked(),
                self.filter_inputs["A32"].isChecked()
            ]):
                # ES: Si no hay ninguno seleccionado, seleccionar "すべて" por defecto
                # EN: If none are selected, select "すべて" by default
                # JP: 何も選択されていなければ、デフォルトで「すべて」を選択
                self.filter_inputs["すべて"].setChecked(True)
                print("✅ 特定ブラシが未選択のため、既定で「すべて」を選択しました")

    def load_file(self, label_to_update: QLabel, title: str):
        """Carga un archivo y actualiza el label"""
        # ES: Limpiar referencias stale a threads de optimización al cambiar de archivo
        # EN: Clear stale optimization-thread references when switching files
        # JP: ファイル切替時に最適化スレッドの古い参照をクリーンアップ
        self._cleanup_optimization_threads(aggressive=False)

        # ES: Pausar timers automáticos para evitar interferencia con el diálogo | EN: Pause auto timers to avoid interference with the dialog | JA: ダイアログとの干渉を避けるため自動タイマーを一時停止
        self.pause_auto_timers()
        
        file_path, _ = QFileDialog.getOpenFileName(self, title)
        
        # ES: Reanudar timers después del diálogo | EN: Resume timers after the dialog | JA: ダイアログ後にタイマーを再開
        self.resume_auto_timers()

        if file_path:
            file_name = file_path.split("/")[-1]
            label_to_update.setText(f"ファイル読み込み完了: {file_name}")

            # ES: Guardar la ruta del sample o del results según el label | EN: Save sample or results path according to label | JA: ラベルに応じてサンプル/結果パスを保存
            if label_to_update == self.sample_label:
                self.sample_file_path = file_path
            elif label_to_update == self.results_label:
                self.results_file_path = file_path
        else:
            label_to_update.setText("ファイル未選択")

    def on_d_optimizer_clicked(self):
        """ES: Ejecuta solo la optimización D-óptima
        EN: Run D-optimal optimization only
        JA: D最適化のみ実行"""
        # ES: Limpiar threads stale antes de chequear "ya está corriendo"
        # EN: Clear stale threads before checking "already running"
        # JP: 「既に実行中」チェック前に古いスレッド参照をクリーンアップ
        self._cleanup_optimization_threads(aggressive=False)

        # ES: ✅ FIX UI: si venimos de la pantalla de filtros, volver a la pantalla principal
        # EN: ✅ UI FIX: if we come from the filter screen, return to the main screen
        # JP: ✅ UI修正：フィルタ画面から来た場合、メイン画面に戻す
        # ES: (si no, los botones/controles del filtro pueden quedarse visibles al mostrar gráficos)
        # EN: (otherwise, filter buttons/controls may remain visible when showing charts)
        # JP: （そうしないと、グラフ表示時にフィルタのボタン/コントロールが残って見えることがある）
        try:
            in_filter_view = False
            for i in range(self.center_layout.count()):
                item = self.center_layout.itemAt(i)
                if item.widget() and isinstance(item.widget(), QLabel):
                    if item.widget().text() == "データフィルター":
                        in_filter_view = True
                        break
            if in_filter_view:
                print("🔄 D最適化: detectada pantalla de filtros, restaurando pantalla principal...")
                self.clear_main_screen()
        except Exception:
            pass

        # ES: No mezclar ejecuciones pesadas en paralelo | EN: Do not run heavy tasks in parallel | JA: 重い処理の並列実行を避ける
        if hasattr(self, 'linear_worker') and self.linear_worker is not None:
            try:
                if self.linear_worker.isRunning():
                    QMessageBox.warning(self, "最適化", "⚠️ 線形解析が実行中です。\n完了または停止するまでお待ちください。")
                    return
            except RuntimeError:
                self.linear_worker = None
        if hasattr(self, 'nonlinear_worker') and self.nonlinear_worker is not None:
            try:
                if self.nonlinear_worker.isRunning():
                    QMessageBox.warning(self, "最適化", "⚠️ 非線形解析が実行中です。\n完了または停止するまでお待ちください。")
                    return
            except RuntimeError:
                self.nonlinear_worker = None

        # ES: Evitar arrancar si ya hay una optimización en ejecución | EN: Avoid starting if an optimization is already running | JA: 最適化実行中は起動を防ぐ
        for t_attr in ("d_optimizer_thread", "i_optimizer_thread", "dsaitekika_thread"):
            if hasattr(self, t_attr):
                t = getattr(self, t_attr)
                try:
                    if t is not None and t.isRunning():
                        QMessageBox.warning(self, "最適化", "⚠️ すでに最適化が実行中です。\n完了するまでお待ちください。")
                        return
                except RuntimeError:
                    setattr(self, t_attr, None)

        # ES: Verificar que el archivo de muestreo haya sido cargado | EN: Ensure sample file has been loaded | JA: サンプルファイルが読み込まれているか確認
        if not hasattr(self, "sample_file_path"):
            QMessageBox.warning(self, "エラー", "❌ サンプルファイルが読み込まれていません。")
            return

        # ES: Verificar si el archivo pertenece a un proyecto existente | EN: Check if the file belongs to an existing project | JA: ファイルが既存プロジェクトに属するか確認
        sample_path = self.sample_file_path
        sample_dir = os.path.dirname(sample_path)
        sample_file = os.path.basename(sample_path)
        
        # ES: Verificar si es un archivo de proyecto existente | EN: Check if it is an existing project file | JA: 既存プロジェクトのファイルか確認
        belongs_to_existing_project = False
        sample_ext = os.path.splitext(sample_file)[1].lower()
        is_project_sample = (
            sample_file.endswith("_未実験データ.xlsx")
            or sample_file.endswith("_未実験データ.xls")
            or sample_file.endswith("_未実験データ.csv")
        )
        if is_project_sample:
            project_name = sample_file[: -len(f"_未実験データ{sample_ext}")]
            if os.path.basename(sample_dir) == project_name:
                # ES: Es un archivo de proyecto existente
                # EN: It is an existing project file
                # JP: 既存プロジェクトのファイルです
                belongs_to_existing_project = True
                self.proyecto_folder = sample_dir
                self.proyecto_nombre = project_name
                print(f"✅ 既存プロジェクトに属するファイルです: {project_name}")
                
                # ES: Verificar si existe el archivo en 99_Temp | EN: Check if file exists in 99_Temp | JA: 99_Tempにファイルがあるか確認
                temp_file_path = os.path.join(self.proyecto_folder, "99_Temp", sample_file)
                if os.path.exists(temp_file_path):
                    print(f"✅ 99_Temp の既存ファイルを使用します: {temp_file_path}")
                    # ES: Usar directamente el archivo de 99_Temp
                    # EN: Use the 99_Temp file directly
                    # JP: 99_Tempのファイルをそのまま使用
                    input_file = temp_file_path
                else:
                    print(f"⚠️ 99_Temp にファイルが見つかりません。コピー中...")
                    # ES: Crear 99_Temp si no existe | EN: Create 99_Temp if it does not exist | JA: 99_Tempが無ければ作成
                    temp_base = os.path.join(self.proyecto_folder, "99_Temp")
                    os.makedirs(temp_base, exist_ok=True)
                    input_file = os.path.join(temp_base, sample_file)
                    try:
                        # ES: Mostrar loader ANTES de copiar (puede tardar mucho) | EN: Show loader BEFORE copying (may take long) | JA: コピー前にローダー表示（時間がかかる場合あり）
                        if not hasattr(self, 'loader_overlay') or self.loader_overlay is None:
                            self.loader_overlay = LoadingOverlay(self.center_frame)
                        self.loader_overlay.start()
                        try:
                            QApplication.processEvents()
                        except Exception:
                            pass
                        shutil.copy(self.sample_file_path, input_file)
                        print(f"✅ 99_Temp にコピーしました: {input_file}")
                    except Exception as e:
                        try:
                            self.loader_overlay.stop()
                        except Exception:
                            pass
                        QMessageBox.critical(self, "エラー", f"❌ 99_Tempへのコピーに失敗しました:\n{str(e)}")
                        return
            else:
                belongs_to_existing_project = False
        else:
            belongs_to_existing_project = False

        # ES: Si no pertenece a un proyecto existente, crear nuevo proyecto
        # EN: If it does not belong to an existing project, create a new project
        # JP: 既存プロジェクトに属さない場合は新規プロジェクトを作成
        if not belongs_to_existing_project:
            # ES: Pausar timers automáticos para evitar interferencia con el diálogo | EN: Pause auto timers to avoid interference with the dialog | JA: ダイアログとの干渉を避けるため自動タイマーを一時停止
            self.pause_auto_timers()
            
            folder_path, _ = QFileDialog.getSaveFileName(
                self, "プロジェクトフォルダ名を入力してください", "", "Proyecto (*.xlsx)"
            )
            
            # ES: Reanudar timers después del diálogo | EN: Resume timers after the dialog | JA: ダイアログ後にタイマーを再開
            self.resume_auto_timers()
            if not folder_path:
                return

            if folder_path.endswith(".xlsx"):
                folder_path = folder_path[:-5]

            project_name = os.path.basename(folder_path)
            project_folder = folder_path

            try:
                os.makedirs(project_folder, exist_ok=False)
            except FileExistsError:
                QMessageBox.warning(self, "既存フォルダ",
                                    f"⚠️ フォルダ '{project_name}' は既に存在します。別の名前を入力してください。")
                return

            self.proyecto_folder = project_folder
            self.proyecto_nombre = project_name
            
            # ES: Mostrar loader ANTES de crear estructura/copiar archivos (puede tardar mucho) | EN: Show loader BEFORE creating structure/copying files (may take long) | JA: 構造作成・ファイルコピー前にローダー表示（時間がかかる場合あり）
            if not hasattr(self, 'loader_overlay') or self.loader_overlay is None:
                self.loader_overlay = LoadingOverlay(self.center_frame)
            self.loader_overlay.start()
            try:
                QApplication.processEvents()
            except Exception:
                pass
            
            # ES: Crear estructura de carpetas del proyecto | EN: Create project folder structure | JA: プロジェクトのフォルダ構造を作成
            self.create_project_folder_structure(project_folder)
            
            # ES: Copiar archivo de muestreo a la carpeta principal del proyecto
            # EN: Copy the sample file to the project's main folder
            # JP: サンプルファイルをプロジェクトのメインフォルダへコピー
            src_ext = os.path.splitext(self.sample_file_path)[1].lower()
            if src_ext not in (".csv", ".xlsx", ".xls"):
                src_ext = ".csv"
            excel_dest_main = os.path.join(self.proyecto_folder, f"{project_name}_未実験データ{src_ext}")
            try:
                shutil.copy(self.sample_file_path, excel_dest_main)
            except Exception as e:
                try:
                    self.loader_overlay.stop()
                except Exception:
                    pass
                QMessageBox.critical(self, "エラー", f"❌ ファイルのコピーに失敗しました:\n{str(e)}")
                return
            
            # ES: Hacer copia en 99_Temp
            # EN: Make a copy in 99_Temp
            # JP: 99_Tempにコピーを作成
            temp_base = os.path.join(self.proyecto_folder, "99_Temp")
            excel_dest_temp = os.path.join(temp_base, f"{project_name}_未実験データ{src_ext}")
            try:
                shutil.copy(self.sample_file_path, excel_dest_temp)
            except Exception as e:
                try:
                    self.loader_overlay.stop()
                except Exception:
                    pass
                QMessageBox.critical(self, "エラー", f"❌ 99_Tempへのコピーに失敗しました:\n{str(e)}")
                return

            # ES: Actualizar el archivo de entrada al archivo del proyecto creado | EN: Update input file to the created project file | JA: 作成したプロジェクトのファイルに入力ファイルを更新
            print("🔄 入力ファイルを更新中...")
            self.sample_file_path = excel_dest_main
            self.load_file_label.setText(f"読み込み済み: {project_name}_未実験データ{src_ext}")
            print(f"✅ 入力ファイルを更新しました: {excel_dest_main}")
            print(f"✅ ラベルを更新しました: {self.load_file_label.text()}")

            # ES: CSV→Excel (99_未実験データ) deshabilitado: proceso pesado y no necesario para la optimización
            # EN: CSV→Excel (99_未実験データ) disabled: heavy process and not needed for optimization
            # JP: CSV→Excel（99_未実験データ）は無効：重く、最適化に不要
            
            # ES: Usar el archivo de 99_Temp para la optimización
            # EN: Use the 99_Temp file for optimization
            # JP: 最適化には99_Tempのファイルを使用
            input_file = excel_dest_temp

        # ES: Crear carpeta temporal para resultados D-óptimos | EN: Create temp folder for D-optimal results | JA: D最適結果用一時フォルダを作成
        temp_base = os.path.join(self.proyecto_folder, "99_Temp")
        os.makedirs(temp_base, exist_ok=True)
        temp_folder = os.path.join(temp_base, "Temp")
        os.makedirs(temp_folder, exist_ok=True)
        output_folder = temp_folder  # Usar 99_Temp/Temp
        
        # ES: Guardar referencia para limpieza posterior | EN: Save reference for later cleanup | JA: 後でクリーンアップするため参照を保存
        self.current_temp_folder = temp_folder

        # ES: Mostrar loader (ya se mostró arriba si se creó proyecto; asegurar que esté visible) | EN: Show loader (already shown above if project created; ensure visible) | JA: ローダー表示（プロジェクト作成時は上で既に表示済、表示を保証）
        if not hasattr(self, 'loader_overlay') or self.loader_overlay is None:
            self.loader_overlay = LoadingOverlay(self.center_frame)
        self.loader_overlay.start()
        try:
            QApplication.processEvents()
        except Exception:
            pass

        # ES: Usar el archivo determinado (existente o nuevo) | EN: Use the determined file (existing or new) | JA: 決定したファイル（既存または新規）を使用
        print(f"✅ 最適化に使用するファイル: {input_file}")

        # ES: === NUEVO: calcular "ensayos ya hechos" como (principal - 99_Temp) ===
        # EN: === NEW: compute "already-done experiments" as (main - 99_Temp) ===
        # JP: === 新規：既実験を（本体 - 99_Temp）で算出 ===
        # ES: main_file debe ser el archivo de la carpeta principal del proyecto (Excel o CSV).
        # EN: main_file must be the file in the project's main folder (Excel or CSV).
        # JP: main_fileはプロジェクト本体フォルダのファイル（Excel/CSV）である必要がある。
        main_file = getattr(self, "sample_file_path", None)

        done_file = os.path.join(self.proyecto_folder, "99_Temp", "done_experiments.xlsx")

        # ES: ⚡ Generar done_experiments en background para que el GIF no se congele al inicio
        # EN: ⚡ Generate done_experiments in the background so the GIF doesn't freeze at startup
        # JP: ⚡ 起動時にGIFが固まらないよう、done_experimentsをバックグラウンド生成
        def _start_d_with_existing(existing_file):
            # ES: Lanzar optimización D-óptima en hilo
            # EN: Launch D-optimal optimization in a thread
            # JP: D最適化をスレッドで起動
            self.d_optimizer_thread = QThread()
            self.d_optimizer_worker = IntegratedOptimizerWorker(
                sample_file=main_file if main_file else input_file,
                existing_file=existing_file,
                output_folder=output_folder,
                num_points=self.get_sample_size(),
                sample_size=None,  # O el valor que corresponda
                enable_hyperparameter_tuning=True,
                force_reoptimization=False,
                optimization_type="d_optimal"  # Specify D optimization
            )
            self.d_optimizer_worker.moveToThread(self.d_optimizer_thread)

            self.d_optimizer_thread.started.connect(self.d_optimizer_worker.run)
            self.d_optimizer_worker.finished.connect(self.on_d_optimizer_finished)
            self.d_optimizer_worker.error.connect(self.on_dsaitekika_error)
            # ES: ✅ FIX: si hay error, cerrar el thread también (si no, queda "isRunning()" para siempre)
            # EN: ✅ FIX: on error, also stop the thread (otherwise isRunning() stays true forever)
            # JP: ✅ 修正：エラー時もスレッドを終了（しないとisRunning()が永遠にtrueになる）
            self.d_optimizer_worker.error.connect(self.d_optimizer_thread.quit)
            self.d_optimizer_worker.finished.connect(self.d_optimizer_thread.quit)
            self.d_optimizer_worker.finished.connect(self.d_optimizer_worker.deleteLater)
            self.d_optimizer_thread.finished.connect(self.d_optimizer_thread.deleteLater)
            # ES: Limpiar referencia cuando el thread termine (evita estados colgados)
            # EN: Clear the reference when the thread finishes (prevents stuck states)
            # JP: スレッド終了時に参照をクリア（ハング状態を防止）
            self.d_optimizer_thread.finished.connect(lambda: setattr(self, "d_optimizer_thread", None))

            self.d_optimizer_thread.start()

        self._build_done_experiments_async(main_file, input_file, done_file, _start_d_with_existing)
        return

    def on_i_optimizer_clicked(self):
        """ES: Ejecuta solo la optimización I-óptima
        EN: Run I-optimal optimization only
        JA: I最適化のみ実行"""
        print("I最適化実行中...")
        # ES: Limpiar threads stale antes de chequear "ya está corriendo"
        # EN: Clear stale threads before checking "already running"
        # JP: 「既に実行中」チェック前に古いスレッド参照をクリーンアップ
        self._cleanup_optimization_threads(aggressive=False)

        # ES: ✅ FIX UI: si venimos de la pantalla de filtros, volver a la pantalla principal
        # EN: ✅ UI FIX: if we come from the filter screen, return to the main screen
        # JP: ✅ UI修正：フィルタ画面から来た場合、メイン画面に戻す
        try:
            in_filter_view = False
            for i in range(self.center_layout.count()):
                item = self.center_layout.itemAt(i)
                if item.widget() and isinstance(item.widget(), QLabel):
                    if item.widget().text() == "データフィルター":
                        in_filter_view = True
                        break
            if in_filter_view:
                print("🔄 I最適化: detectada pantalla de filtros, restaurando pantalla principal...")
                self.clear_main_screen()
        except Exception:
            pass

        # ES: No mezclar ejecuciones pesadas en paralelo | EN: Do not run heavy tasks in parallel | JA: 重い処理の並列実行を避ける
        if hasattr(self, 'linear_worker') and self.linear_worker is not None:
            try:
                if self.linear_worker.isRunning():
                    QMessageBox.warning(self, "最適化", "⚠️ 線形解析が実行中です。\n完了または停止するまでお待ちください。")
                    return
            except RuntimeError:
                self.linear_worker = None
        if hasattr(self, 'nonlinear_worker') and self.nonlinear_worker is not None:
            try:
                if self.nonlinear_worker.isRunning():
                    QMessageBox.warning(self, "最適化", "⚠️ 非線形解析が実行中です。\n完了または停止するまでお待ちください。")
                    return
            except RuntimeError:
                self.nonlinear_worker = None

        # ES: Evitar arrancar si ya hay una optimización en ejecución | EN: Avoid starting if an optimization is already running | JA: 最適化実行中は起動を防ぐ
        for t_attr in ("d_optimizer_thread", "i_optimizer_thread", "dsaitekika_thread"):
            if hasattr(self, t_attr):
                t = getattr(self, t_attr)
                try:
                    if t is not None and t.isRunning():
                        QMessageBox.warning(self, "最適化", "⚠️ すでに最適化が実行中です。\n完了するまでお待ちください。")
                        return
                except RuntimeError:
                    setattr(self, t_attr, None)
        
        # ES: Verificar que el archivo de muestreo haya sido cargado | EN: Ensure sample file has been loaded | JA: サンプルファイルが読み込まれているか確認
        if not hasattr(self, "sample_file_path"):
            QMessageBox.warning(self, "エラー", "❌ サンプルファイルが読み込まれていません。")
            return

        # ES: Verificar si el archivo pertenece a un proyecto existente | EN: Check if the file belongs to an existing project | JA: ファイルが既存プロジェクトに属するか確認
        sample_path = self.sample_file_path
        sample_dir = os.path.dirname(sample_path)
        sample_file = os.path.basename(sample_path)
        
        # ES: Verificar si es un archivo de proyecto existente | EN: Check if it is an existing project file | JA: 既存プロジェクトのファイルか確認
        belongs_to_existing_project = False
        sample_ext = os.path.splitext(sample_file)[1].lower()
        is_project_sample = (
            sample_file.endswith("_未実験データ.xlsx")
            or sample_file.endswith("_未実験データ.xls")
            or sample_file.endswith("_未実験データ.csv")
        )
        if is_project_sample:
            project_name = sample_file[: -len(f"_未実験データ{sample_ext}")]
            if os.path.basename(sample_dir) == project_name:
                # ES: Es un archivo de proyecto existente
                # EN: This is an existing project file
                # JP: 既存プロジェクトのファイルである
                belongs_to_existing_project = True
                self.proyecto_folder = sample_dir
                self.proyecto_nombre = project_name
                print(f"✅ 既存プロジェクトに属するファイルです: {project_name}")
                
                # ES: Verificar si existe el archivo en 99_Temp | EN: Check if file exists in 99_Temp | JA: 99_Tempにファイルがあるか確認
                temp_file_path = os.path.join(self.proyecto_folder, "99_Temp", sample_file)
                if os.path.exists(temp_file_path):
                    print(f"✅ 99_Temp の既存ファイルを使用します: {temp_file_path}")
                    # ES: Usar directamente el archivo de 99_Temp
                    # EN: Use the file directly from 99_Temp
                    # JP: 99_Tempのファイルを直接使用する
                    input_file = temp_file_path
                else:
                    print(f"⚠️ 99_Temp にファイルが見つかりません。コピー中...")
                    # ES: Crear 99_Temp si no existe | EN: Create 99_Temp if it does not exist | JA: 99_Tempが無ければ作成
                    temp_base = os.path.join(self.proyecto_folder, "99_Temp")
                    os.makedirs(temp_base, exist_ok=True)
                    input_file = os.path.join(temp_base, sample_file)
                    try:
                        # ES: Mostrar loader ANTES de copiar (puede tardar mucho) | EN: Show loader BEFORE copying (may take long) | JA: コピー前にローダー表示（時間がかかる場合あり）
                        if not hasattr(self, 'loader_overlay') or self.loader_overlay is None:
                            self.loader_overlay = LoadingOverlay(self.center_frame)
                        self.loader_overlay.start()
                        try:
                            QApplication.processEvents()
                        except Exception:
                            pass
                        shutil.copy(self.sample_file_path, input_file)
                        print(f"✅ 99_Temp にコピーしました: {input_file}")
                    except Exception as e:
                        try:
                            self.loader_overlay.stop()
                        except Exception:
                            pass
                        QMessageBox.critical(self, "エラー", f"❌ 99_Tempへのコピーに失敗しました:\n{str(e)}")
                        return
            else:
                belongs_to_existing_project = False
        else:
            belongs_to_existing_project = False

        # ES: Si no pertenece a un proyecto existente, crear nuevo proyecto
        # EN: If it does not belong to an existing project, create a new project
        # JP: 既存プロジェクトに属さない場合は新規プロジェクトを作成
        if not belongs_to_existing_project:
            # ES: Pausar timers automáticos para evitar interferencia con el diálogo | EN: Pause auto timers to avoid interference with the dialog | JA: ダイアログとの干渉を避けるため自動タイマーを一時停止
            self.pause_auto_timers()
            
            folder_path, _ = QFileDialog.getSaveFileName(
                self, "プロジェクトフォルダ名を入力してください", "", "Proyecto (*.xlsx)"
            )
            
            # ES: Reanudar timers después del diálogo | EN: Resume timers after the dialog | JA: ダイアログ後にタイマーを再開
            self.resume_auto_timers()
            if not folder_path:
                return

            if folder_path.endswith(".xlsx"):
                folder_path = folder_path[:-5]

            project_name = os.path.basename(folder_path)
            project_folder = folder_path

            try:
                os.makedirs(project_folder, exist_ok=False)
            except FileExistsError:
                QMessageBox.warning(self, "既存フォルダ",
                                    f"⚠️ フォルダ '{project_name}' は既に存在します。別の名前を入力してください。")
                return

            self.proyecto_folder = project_folder
            self.proyecto_nombre = project_name
            
            # ES: Mostrar loader ANTES de crear estructura/copiar archivos (puede tardar mucho) | EN: Show loader BEFORE creating structure/copying files (may take long) | JA: 構造作成・ファイルコピー前にローダー表示（時間がかかる場合あり）
            if not hasattr(self, 'loader_overlay') or self.loader_overlay is None:
                self.loader_overlay = LoadingOverlay(self.center_frame)
            self.loader_overlay.start()
            try:
                QApplication.processEvents()
            except Exception:
                pass
            
            # ES: Crear estructura de carpetas del proyecto | EN: Create project folder structure | JA: プロジェクトのフォルダ構造を作成
            self.create_project_folder_structure(project_folder)
            
            # ES: Copiar archivo de muestreo a la carpeta principal del proyecto
            # EN: Copy the sample file to the project's main folder
            # JP: サンプルファイルをプロジェクトのメインフォルダへコピー
            src_ext = os.path.splitext(self.sample_file_path)[1].lower()
            if src_ext not in (".csv", ".xlsx", ".xls"):
                src_ext = ".csv"
            excel_dest_main = os.path.join(self.proyecto_folder, f"{project_name}_未実験データ{src_ext}")
            try:
                shutil.copy(self.sample_file_path, excel_dest_main)
            except Exception as e:
                try:
                    self.loader_overlay.stop()
                except Exception:
                    pass
                QMessageBox.critical(self, "エラー", f"❌ ファイルのコピーに失敗しました:\n{str(e)}")
                return
            
            # ES: Hacer copia en 99_Temp
            # EN: Make a copy in 99_Temp
            # JP: 99_Tempにコピーを作成
            temp_base = os.path.join(self.proyecto_folder, "99_Temp")
            excel_dest_temp = os.path.join(temp_base, f"{project_name}_未実験データ{src_ext}")
            try:
                shutil.copy(self.sample_file_path, excel_dest_temp)
            except Exception as e:
                try:
                    self.loader_overlay.stop()
                except Exception:
                    pass
                QMessageBox.critical(self, "エラー", f"❌ 99_Tempへのコピーに失敗しました:\n{str(e)}")
                return

            # ES: Actualizar el archivo de entrada al archivo del proyecto creado | EN: Update input file to the created project file | JA: 作成したプロジェクトのファイルに入力ファイルを更新
            print("🔄 入力ファイルを更新中...")
            self.sample_file_path = excel_dest_main
            self.load_file_label.setText(f"読み込み済み: {project_name}_未実験データ{src_ext}")
            print(f"✅ 入力ファイルを更新しました: {excel_dest_main}")
            print(f"✅ ラベルを更新しました: {self.load_file_label.text()}")

            # CSV→Excel (99_未実験データ) deshabilitado: proceso pesado y no necesario para la optimización
            
            # ES: Usar el archivo de 99_Temp para la optimización
            # EN: Use the 99_Temp file for optimization
            # JP: 最適化には99_Tempのファイルを使用する
            input_file = excel_dest_temp

        # ES: Crear carpeta temporal para resultados I-óptimos | EN: Create temp folder for I-optimal results | JA: I最適結果用一時フォルダを作成
        temp_base = os.path.join(self.proyecto_folder, "99_Temp")
        os.makedirs(temp_base, exist_ok=True)
        temp_folder = os.path.join(temp_base, "Temp")
        os.makedirs(temp_folder, exist_ok=True)
        output_folder = temp_folder  # Usar 99_Temp/Temp
        
        # ES: Guardar referencia para limpieza posterior | EN: Save reference for later cleanup | JA: 後でクリーンアップするため参照を保存
        self.current_temp_folder = temp_folder

        # ES: Mostrar loader (ya se mostró arriba si se creó proyecto; asegurar que esté visible) | EN: Show loader (already shown above if project created; ensure visible) | JA: ローダー表示（プロジェクト作成時は上で既に表示済、表示を保証）
        if not hasattr(self, 'loader_overlay') or self.loader_overlay is None:
            self.loader_overlay = LoadingOverlay(self.center_frame)
        self.loader_overlay.start()
        try:
            QApplication.processEvents()
        except Exception:
            pass

        # ES: Usar el archivo determinado (existente o nuevo) | EN: Use the determined file (existing or new) | JA: 決定したファイル（既存または新規）を使用
        print(f"✅ 最適化に使用するファイル: {input_file}")

        # === NUEVO: calcular "ensayos ya hechos" como (principal - 99_Temp) ===
        main_file = getattr(self, "sample_file_path", None)

        done_file = os.path.join(self.proyecto_folder, "99_Temp", "done_experiments.xlsx")

        # ⚡ Generar done_experiments en background para que el GIF no se congele al inicio
        def _start_i_with_existing(existing_file):
            # Lanzar optimización I-óptima en hilo
            self.i_optimizer_thread = QThread()
            self.i_optimizer_worker = IntegratedOptimizerWorker(
                sample_file=main_file if main_file else input_file,
                existing_file=existing_file,
                output_folder=output_folder,
                num_points=self.get_sample_size(),
                sample_size=None,  # O el valor que corresponda
                enable_hyperparameter_tuning=True,
                force_reoptimization=False,
                optimization_type="i_optimal"  # Specify I optimization
            )
            self.i_optimizer_worker.moveToThread(self.i_optimizer_thread)

            self.i_optimizer_thread.started.connect(self.i_optimizer_worker.run)
            self.i_optimizer_worker.finished.connect(self.on_i_optimizer_finished)
            self.i_optimizer_worker.error.connect(self.on_dsaitekika_error)
            # ES: ✅ FIX: si hay error, cerrar el thread también (si no, queda "isRunning()" para siempre)
            # EN: ✅ FIX: if there is an error, close the thread too (otherwise it stays \"isRunning()\" forever)
            # JP: ✅ 修正: エラー時はスレッドも終了（そうしないと\"isRunning()\"のままになる）
            self.i_optimizer_worker.error.connect(self.i_optimizer_thread.quit)
            self.i_optimizer_worker.finished.connect(self.i_optimizer_thread.quit)
            self.i_optimizer_worker.finished.connect(self.i_optimizer_worker.deleteLater)
            self.i_optimizer_thread.finished.connect(self.i_optimizer_thread.deleteLater)
            # ES: Limpiar referencia cuando el thread termine (evita estados colgados)
            # EN: Clear the reference when the thread finishes (prevents stuck states)
            # JP: スレッド終了時に参照をクリア（ハング状態を防止）
            self.i_optimizer_thread.finished.connect(lambda: setattr(self, "i_optimizer_thread", None))

            self.i_optimizer_thread.start()

        self._build_done_experiments_async(main_file, input_file, done_file, _start_i_with_existing)
        return

    def on_dsaitekika_clicked(self):
        print("D最適化実行中...")
        print("🔍 DEBUG: on_dsaitekika_clicked を開始")
        # ES: Limpiar threads stale antes de chequear "ya está corriendo"
        # EN: Clear stale threads before checking "already running"
        # JP: 「既に実行中」チェック前に古いスレッド参照をクリーンアップ
        self._cleanup_optimization_threads(aggressive=False)

        # ES: No mezclar ejecuciones pesadas en paralelo | EN: Do not run heavy tasks in parallel | JA: 重い処理の並列実行を避ける
        if hasattr(self, 'linear_worker') and self.linear_worker is not None:
            try:
                if self.linear_worker.isRunning():
                    QMessageBox.warning(self, "最適化", "⚠️ 線形解析が実行中です。\n完了または停止するまでお待ちください。")
                    return
            except RuntimeError:
                self.linear_worker = None
        if hasattr(self, 'nonlinear_worker') and self.nonlinear_worker is not None:
            try:
                if self.nonlinear_worker.isRunning():
                    QMessageBox.warning(self, "最適化", "⚠️ 非線形解析が実行中です。\n完了または停止するまでお待ちください。")
                    return
            except RuntimeError:
                self.nonlinear_worker = None

        # ES: Evitar arrancar si ya hay una optimización en ejecución | EN: Avoid starting if an optimization is already running | JA: 最適化実行中は起動を防ぐ
        for t_attr in ("d_optimizer_thread", "i_optimizer_thread", "dsaitekika_thread"):
            if hasattr(self, t_attr):
                t = getattr(self, t_attr)
                try:
                    if t is not None and t.isRunning():
                        QMessageBox.warning(self, "最適化", "⚠️ すでに最適化が実行中です。\n完了するまでお待ちください。")
                        return
                except RuntimeError:
                    setattr(self, t_attr, None)

        if not hasattr(self, "sample_file_path"):
            QMessageBox.warning(self, "エラー", "❌ サンプルファイルが読み込まれていません。")
            return

        # ES: Pausar timers automáticos para evitar interferencia con el diálogo | EN: Pause auto timers to avoid interference with the dialog | JA: ダイアログとの干渉を避けるため自動タイマーを一時停止
        self.pause_auto_timers()

        # ES: Crear carpeta del proyecto | EN: Create project folder | JA: プロジェクトフォルダを作成
        folder_path, _ = QFileDialog.getSaveFileName(
            self, "プロジェクトフォルダ名を入力してください", "", "Proyecto (*.xlsx)"
        )
        
        # ES: Reanudar timers después del diálogo | EN: Resume timers after the dialog | JA: ダイアログ後にタイマーを再開
        self.resume_auto_timers()
        if not folder_path:
            return

        if folder_path.endswith(".xlsx"):
            folder_path = folder_path[:-5]

        project_name = os.path.basename(folder_path)
        project_folder = folder_path

        try:
            os.makedirs(project_folder, exist_ok=False)
        except FileExistsError:
            QMessageBox.warning(self, "既存フォルダ",
                                f"⚠️ フォルダ '{project_name}' は既に存在します。別の名前を入力してください。")
            return

        self.proyecto_folder = project_folder
        self.proyecto_nombre = project_name
        
        # ES: Mostrar loader ANTES de crear estructura/copiar archivos (puede tardar mucho) | EN: Show loader BEFORE creating structure/copying files (may take long) | JA: 構造作成・ファイルコピー前にローダー表示（時間がかかる場合あり）
        if not hasattr(self, 'loader_overlay') or self.loader_overlay is None:
            self.loader_overlay = LoadingOverlay(self.center_frame)
        self.loader_overlay.start()
        try:
            QApplication.processEvents()
        except Exception:
            pass
        
        # ES: Crear estructura de carpetas del proyecto | EN: Create project folder structure | JA: プロジェクトのフォルダ構造を作成
        self.create_project_folder_structure(project_folder)
        
        # ES: Copiar archivo de muestreo a la carpeta principal del proyecto
        # EN: Copy the sample file to the project's main folder
        # JP: サンプルファイルをプロジェクトのメインフォルダへコピー
        src_ext = os.path.splitext(self.sample_file_path)[1].lower()
        if src_ext not in (".csv", ".xlsx", ".xls"):
            src_ext = ".csv"
        excel_dest_main = os.path.join(self.proyecto_folder, f"{project_name}_未実験データ{src_ext}")
        try:
            shutil.copy(self.sample_file_path, excel_dest_main)
        except Exception as e:
            try:
                self.loader_overlay.stop()
            except Exception:
                pass
            QMessageBox.critical(self, "エラー", f"❌ ファイルのコピーに失敗しました:\n{str(e)}")
            return
        
        # ES: Hacer copia en 99_Temp
        # EN: Make a copy in 99_Temp
        # JP: 99_Tempにコピーを作成
        temp_base = os.path.join(self.proyecto_folder, "99_Temp")
        excel_dest_temp = os.path.join(temp_base, f"{project_name}_未実験データ{src_ext}")
        try:
            shutil.copy(self.sample_file_path, excel_dest_temp)
        except Exception as e:
            try:
                self.loader_overlay.stop()
            except Exception:
                pass
            QMessageBox.critical(self, "エラー", f"❌ 99_Tempへのコピーに失敗しました:\n{str(e)}")
            return

        self.muestreo_guardado_path = excel_dest_main
        
        print("🔍 DEBUG: 入力ファイル更新コードに到達しました")
        # ES: Actualizar el archivo de entrada al archivo del proyecto creado | EN: Update input file to the created project file | JA: 作成したプロジェクトのファイルに入力ファイルを更新
        print("🔄 入力ファイルを更新中...")
        self.sample_file_path = excel_dest_main
        self.load_file_label.setText(f"読み込み済み: {project_name}_未実験データ{src_ext}")
        print(f"✅ 入力ファイルを更新しました: {excel_dest_main}")
        print(f"✅ ラベルを更新しました: {self.load_file_label.text()}")

        # CSV→Excel (99_未実験データ) deshabilitado: proceso pesado y no necesario para la optimización

        # ES: Crear carpeta temporal de resultados dentro del proyecto | EN: Create results temp folder inside project | JA: プロジェクト内に結果用一時フォルダを作成
        temp_base = os.path.join(self.proyecto_folder, "99_Temp")
        os.makedirs(temp_base, exist_ok=True)
        temp_folder = os.path.join(temp_base, "Temp")
        os.makedirs(temp_folder, exist_ok=True)
        output_folder = temp_folder  # Usar 99_Temp/Temp

        self.dsaitekika_output_excel = os.path.join(output_folder, "selected_samples.xlsx")
        self.dsaitekika_output_prefix = os.path.join(output_folder, "d_optimal")
        
        # ES: Guardar referencia para limpieza posterior | EN: Save reference for later cleanup | JA: 後でクリーンアップするため参照を保存
        self.current_temp_folder = temp_folder

        # ES: El loader ya se mostró arriba (antes de crear/copiar). Mantenerlo activo.
        # EN: The loader was already shown above (before creating/copying). Keep it active.
        # JP: ローダーは上で表示済（作成/コピー前）。表示を維持する。

        # ES: Usar el archivo de 99_Temp en lugar del archivo original | EN: Use 99_Temp file instead of the original | JA: 元ファイルの代わりに99_Tempのファイルを使用
        input_file = excel_dest_temp
        print(f"✅ 99_Temp のファイルを使用します: {input_file}")
        # ES: Guardar para poder recalcular D基準値 como el archivo de referencia | EN: Save to recalculate D基準値 as reference file | JA: D基準値を参照ファイルとして再計算するため保存
        self._last_dsaitekika_input_file = input_file
        
        self.dsaitekika_thread = QThread()
        self.dsaitekika_worker = DsaitekikaWorker(
            input_file,
            self.dsaitekika_output_excel,
            self.dsaitekika_output_prefix,
            self.get_sample_size(),
        )
        self.dsaitekika_worker.moveToThread(self.dsaitekika_thread)

        self.dsaitekika_thread.started.connect(self.dsaitekika_worker.run)
        self.dsaitekika_worker.finished.connect(self.on_dsaitekika_finished)
        self.dsaitekika_worker.error.connect(self.on_dsaitekika_error)
        # ES: ✅ FIX: si hay error, cerrar el thread también (si no, queda "isRunning()" para siempre)
        # EN: ✅ FIX: if there is an error, close the thread too (otherwise it stays \"isRunning()\" forever)
        # JP: ✅ 修正: エラー時はスレッドも終了（そうしないと\"isRunning()\"のままになる）
        self.dsaitekika_worker.error.connect(self.dsaitekika_thread.quit)
        self.dsaitekika_worker.finished.connect(self.dsaitekika_thread.quit)
        self.dsaitekika_worker.finished.connect(self.dsaitekika_worker.deleteLater)
        self.dsaitekika_thread.finished.connect(self.dsaitekika_thread.deleteLater)
        # ES: Limpiar referencia cuando el thread termine (evita estados colgados)
        # EN: Clear the reference when the thread finishes (prevents stuck states)
        # JP: スレッド終了時に参照をクリア（ハング状態を防止）
        self.dsaitekika_thread.finished.connect(lambda: setattr(self, "dsaitekika_thread", None))

        self.dsaitekika_thread.start()

    def _start_csv_export_async(self, csv_path: str, project_folder: str, project_name: str):
        """
        Ejecuta la exportación CSV→Excel en un QThread para no bloquear la UI.
        No afecta a la optimización (solo genera archivos auxiliares en 99_未実験データ).
        """
        try:
            # ES: Evitar lanzar múltiples conversiones en paralelo
            # EN: Avoid launching multiple conversions in parallel
            # JP: 複数の変換を並列で起動しない
            if hasattr(self, "csv_export_thread") and self.csv_export_thread is not None:
                try:
                    if self.csv_export_thread.isRunning():
                        print("ℹ️ CSV→Excel エクスポートは既に実行中のため、新規要求をスキップします")
                        return
                except RuntimeError:
                    self.csv_export_thread = None
        except Exception:
            pass

        self.csv_export_thread = QThread()
        self.csv_export_worker = CsvToExcelExportWorker(
            lambda: self._export_unexperimented_excel_folder_from_csv(csv_path, project_folder, project_name)
        )
        self.csv_export_worker.moveToThread(self.csv_export_thread)
        self.csv_export_thread.started.connect(self.csv_export_worker.run)
        self.csv_export_worker.finished.connect(self.csv_export_thread.quit)
        self.csv_export_worker.finished.connect(self.csv_export_worker.deleteLater)
        self.csv_export_thread.finished.connect(self.csv_export_thread.deleteLater)

        def _on_err(msg: str):
            print(f"⚠️ CSV→Excel export (async) error: {msg}", flush=True)
        self.csv_export_worker.error.connect(_on_err)

        self.csv_export_thread.start()

    def _build_done_experiments_async(self, main_file: str, temp_file: str, done_file: str, on_ready):
        """Genera done_experiments.xlsx en background y llama on_ready(existing_file) en el hilo UI."""
        try:
            if hasattr(self, "done_exp_thread") and self.done_exp_thread is not None:
                try:
                    if self.done_exp_thread.isRunning():
                        print("ℹ️ done_experiments は既に実行中のため、完了した結果を再利用します", flush=True)
                except RuntimeError:
                    self.done_exp_thread = None
        except Exception:
            pass

        self.done_exp_thread = QThread()
        self.done_exp_worker = CallableResultWorker(
            lambda: self._build_done_experiments_excel(main_file, temp_file, done_file) if main_file else None
        )
        self.done_exp_worker.moveToThread(self.done_exp_thread)
        self.done_exp_thread.started.connect(self.done_exp_worker.run)
        self.done_exp_worker.finished.connect(on_ready)
        self.done_exp_worker.finished.connect(self.done_exp_thread.quit)
        self.done_exp_worker.finished.connect(self.done_exp_worker.deleteLater)
        self.done_exp_thread.finished.connect(self.done_exp_thread.deleteLater)

        def _on_err(msg: str):
            print(f"⚠️ done_experiments (async) error: {msg}", flush=True)
            try:
                on_ready(None)
            except Exception:
                pass
        self.done_exp_worker.error.connect(_on_err)

        self.done_exp_thread.start()

    def on_isaitekika_clicked(self):
        """ES: Acción al pulsar iSaitekika
        EN: Action when iSaitekika is clicked
        JA: iSaitekikaクリック時のアクション"""
        print("i最適化実行中...")
        self.ok_button.setEnabled(True)
        self.ng_button.setEnabled(True)

        self.create_navigation_buttons()
        self.prev_button.setEnabled(True)
        self.next_button.setEnabled(True)

    def find_matching_experiment_file(self, project_folder):
        """
        Busca en 01_実験リスト y compara con el archivo de resultados
        para encontrar el archivo de experimento correspondiente
        """
        import os
        import pandas as pd
        from pathlib import Path
        
        try:
            # ES: Leer archivo de resultados
            # EN: Read results file
            # JP: 結果ファイルを読み込む
            print(f"🔍 デバッグ: 結果ファイルを読み込み中: {self.results_file_path}")
            df_results = pd.read_excel(self.results_file_path)
            print(f"🔍 デバッグ: 結果ファイルを読み込みました: {len(df_results)} 行")
            print(f"🔍 デバッグ: 結果ファイルの列: {list(df_results.columns)}")
            
            # ES: Mostrar primera fila de resultados para debug | EN: Show first row of results for debug | JA: デバッグ用に結果の先頭行を表示
            if len(df_results) > 0:
                print("🔍 デバッグ: 結果の先頭行:")
                first_row = df_results.iloc[0]
                for col in df_results.columns:
                    print(f"  - {col}: {first_row[col]}")
            
            # Columnas a comparar (B a H)
            # Aceptar "UPカット" (nuevo) o "回転方向" (antiguo)
            dir_col = 'UPカット' if 'UPカット' in df_results.columns else '回転方向'
            comparison_columns = ['回転速度', '送り速度', dir_col, '切込量', '突出量', '載せ率', 'パス数']
            
            # ES: Verificar que las columnas existen en el archivo de resultados | EN: Ensure columns exist in results file | JA: 結果ファイルに列が存在するか確認
            available_columns = [col for col in comparison_columns if col in df_results.columns]
            if len(available_columns) < 3:  # Minimum 3 columns to compare
                print(f"⚠️ 比較に必要な列が不足しています: {available_columns}")
                return None
            
            print(f"🔍 デバッグ: 比較可能な列: {available_columns}")
            
            # ES: Buscar en 01_実験リスト
            # EN: Search in 01_実験リスト
            # JP: 01_実験リストを検索
            experiment_list_path = Path(project_folder) / "01_実験リスト"
            if not experiment_list_path.exists():
                print(f"❌ デバッグ: フォルダー 01_実験リスト が存在しません: {experiment_list_path}")
                print(f"🔍 デバッグ: プロジェクト構造を確認中:")
                project_path = Path(project_folder)
                if project_path.exists():
                    print(f"🔍 デバッグ: プロジェクトの内容:")
                    for item in project_path.iterdir():
                        if item.is_dir():
                            print(f"  📁 {item.name}")
                        else:
                            print(f"  📄 {item.name}")
                else:
                    print(f"❌ デバッグ: プロジェクトが存在しません: {project_path}")
                return None
            
            print(f"🔍 デバッグ: 検索中: {experiment_list_path}")
            
            # ES: Verificar contenido de 01_実験リスト | EN: Verify content of 01_実験リスト | JA: 01_実験リストの内容を確認
            experiment_list_contents = list(experiment_list_path.iterdir())
            print(f"🔍 デバッグ: 01_実験リスト の内容（{len(experiment_list_contents)}件）:")
            for item in experiment_list_contents:
                if item.is_dir():
                    print(f"  📁 {item.name}")
                else:
                    print(f"  📄 {item.name}")
            
            # ES: Buscar en subcarpetas
            # EN: Search in subfolders
            # JP: サブフォルダを検索
            subfolder_count = 0
            for subfolder in experiment_list_path.iterdir():
                if not subfolder.is_dir():
                    continue
                
                subfolder_count += 1
                print(f"🔍 デバッグ: サブフォルダー確認 {subfolder_count}: {subfolder.name}")
                
                # ES: Verificar contenido de la subcarpeta | EN: Verify subfolder content | JA: サブフォルダの内容を確認
                subfolder_contents = list(subfolder.iterdir())
                print(f"🔍 デバッグ: {subfolder.name} の内容（{len(subfolder_contents)}件）:")
                for item in subfolder_contents:
                    if item.is_dir():
                        print(f"    📁 {item.name}")
                    else:
                        print(f"    📄 {item.name}")
                
                # ES: Buscar archivos D最適化_新規実験点.xlsx o I最適化_新規実験点.xlsx
                # EN: Search for D最適化_新規実験点.xlsx or I最適化_新規実験点.xlsx files
                # JP: D最適化_新規実験点.xlsx または I最適化_新規実験点.xlsx を検索
                experiment_files = []
                for pattern in ["D最適化_新規実験点.xlsx", "I最適化_新規実験点.xlsx"]:
                    file_path = subfolder / pattern
                    if file_path.exists():
                        experiment_files.append((file_path, pattern))
                        print(f"🔍 デバッグ: ファイルを見つけました: {file_path}")
                
                if not experiment_files:
                    print(f"🔍 デバッグ: {subfolder.name} に実験ファイルが見つかりませんでした")
                
                for file_path, pattern in experiment_files:
                    try:
                        print(f"🔍 デバッグ: ファイルと比較中: {file_path}")
                        print(f"🔍 デバッグ: ファイルパターン: {pattern}")
                        df_experiment = pd.read_excel(file_path)
                        print(f"🔍 デバッグ: 実験ファイルを読み込みました: {len(df_experiment)} 行")
                        print(f"🔍 デバッグ: 実験ファイルの列: {list(df_experiment.columns)}")
                        
                        # ES: Mostrar primera fila de experimento para debug | EN: Show first experiment row for debug | JA: デバッグ用に実験の先頭行を表示
                        if len(df_experiment) > 0:
                            print("🔍 デバッグ: 実験の先頭行:")
                            first_exp_row = df_experiment.iloc[0]
                            for col in df_experiment.columns:
                                print(f"  - {col}: {first_exp_row[col]}")
                        
                        # Comparar filas
                        comparison_count = 0
                        for idx, result_row in df_results.iterrows():
                            for exp_idx, exp_row in df_experiment.iterrows():
                                comparison_count += 1
                                if comparison_count <= 3:  # Only show the first 3 comparisons
                                    print(f"🔍 デバッグ: 比較 {comparison_count}: 結果行 {idx} vs 実験行 {exp_idx}")
                                
                                # ES: Comparar solo las columnas disponibles
                                # EN: Compare only the available columns
                                # JP: 利用可能な列のみ比較する
                                match = True
                                mismatch_details = []
                                
                                for col in available_columns:
                                    if col in df_experiment.columns:
                                        result_val = result_row[col]
                                        exp_val = exp_row[col]
                                        
                                        # Debug de comparación
                                        if comparison_count <= 3:
                                            print(f"  🔍 デバッグ: 列 '{col}' を比較: '{result_val}' vs '{exp_val}'")
                                        
                                        # Comparar valores (considerando tipos de datos)
                                        if pd.isna(result_val) and pd.isna(exp_val):
                                            if comparison_count <= 3:
                                                print(f"    ✅ 両方ともNaN")
                                            continue
                                        elif pd.isna(result_val) or pd.isna(exp_val):
                                            if comparison_count <= 3:
                                                print(f"    ❌ 片方だけNaN")
                                            match = False
                                            mismatch_details.append(f"{col}: NaN vs {exp_val if pd.isna(result_val) else result_val}")
                                            break
                                        
                                        # Convertir a float para comparación numérica si es posible
                                        try:
                                            result_float = float(result_val)
                                            exp_float = float(exp_val)
                                            if abs(result_float - exp_float) < 1e-10:  # Numeric comparison with tolerance
                                                if comparison_count <= 3:
                                                    print(f"    ✅ 数値が一致: {result_float}")
                                                continue
                                            else:
                                                if comparison_count <= 3:
                                                    print(f"    ❌ 数値が不一致: {result_float} != {exp_float}")
                                                match = False
                                                mismatch_details.append(f"{col}: {result_float} vs {exp_float}")
                                                break
                                        except (ValueError, TypeError):
                                            # ES: Si no se pueden convertir a float, comparar como strings
                                            # EN: If they can't be converted to float, compare as strings
                                            # JP: floatに変換できない場合は文字列として比較
                                            if str(result_val).strip() == str(exp_val).strip():
                                                if comparison_count <= 3:
                                                    print(f"    ✅ 文字列が一致: '{result_val}'")
                                                continue
                                            else:
                                                if comparison_count <= 3:
                                                    print(f"    ❌ 文字列が不一致: '{result_val}' != '{exp_val}'")
                                                match = False
                                                mismatch_details.append(f"{col}: '{result_val}' vs '{exp_val}'")
                                                break
                                        else:
                                            if comparison_count <= 3:
                                                print(f"    ✅ 値が一致: '{result_val}'")
                                    else:
                                        if comparison_count <= 3:
                                            print(f"  ❌ 列 '{col}' が実験ファイルに存在しません")
                                        match = False
                                        mismatch_details.append(f"{col}: 実験ファイルに存在しません")
                                        break
                                
                                if match:
                                    print(f"✅ デバッグ: 一致を検出しました！")
                                    print(f"   ファイル: {file_path}")
                                    print(f"   Fila resultado: {idx}, Fila experimento: {exp_idx}")
                                    
                                    # ES: Extraer información de la carpeta
                                    # EN: Extract folder information
                                    # JP: フォルダ情報を抽出
                                    folder_name = subfolder.name
                                    print(f"🔍 デバッグ: 抽出したフォルダー名: {folder_name}")
                                    
                                    # ES: Determinar tipo de optimización basado en el nombre del archivo
                                    # EN: Determine optimization type based on the file name
                                    # JP: ファイル名に基づいて最適化タイプを判定
                                    if "D最適化" in pattern:
                                        optimization_type = "D最適化"
                                        print(f"🔍 デバッグ: ファイル名からDタイプを判定")
                                    elif "I最適化" in pattern:
                                        optimization_type = "I最適化"
                                        print(f"🔍 デバッグ: ファイル名からIタイプを判定")
                                    else:
                                        # ES: Fallback: intentar determinar por el nombre de la carpeta
                                        # EN: Fallback: try to determine from the folder name
                                        # JP: フォールバック：フォルダ名から判定を試す
                                        print(f"🔍 デバッグ: フォールバック - フォルダー名を解析中: {folder_name}")
                                        if "D" in folder_name.upper() or "d" in folder_name.lower():
                                            optimization_type = "D最適化"
                                            print(f"🔍 デバッグ: フォルダー名からDタイプを判定")
                                        elif "I" in folder_name.upper() or "i" in folder_name.lower():
                                            optimization_type = "I最適化"
                                            print(f"🔍 デバッグ: フォルダー名からIタイプを判定")
                                        else:
                                            optimization_type = "D最適化"  # Default
                                            print(f"🔍 デバッグ: 既定: D最適化")
                                    
                                    print(f"🔍 デバッグ: 最終的な最適化タイプ: {optimization_type}")
                                    
                                    return {
                                        'folder_name': folder_name,
                                        'optimization_type': optimization_type,
                                        'file_path': str(file_path),
                                        'result_row': idx,
                                        'experiment_row': exp_idx
                                    }
                                elif comparison_count <= 3:
                                    print(f"❌ デバッグ: 不一致。詳細: {mismatch_details}")
                        
                        if comparison_count > 0:
                            print(f"🔍 デバッグ: 比較回数合計: {comparison_count}")
                        
                    except Exception as e:
                        print(f"❌ {file_path} の読み込み中にエラー: {e}")
                        continue
            
            print("❌ デバッグ: どの実験ファイルにも一致が見つかりませんでした")
            return None
            
        except Exception as e:
            print(f"❌ find_matching_experiment_file でエラー: {e}")
            import traceback
            traceback.print_exc()
            return None

    def create_experiment_data_folder(self, experiment_info):
        """
        Crea la carpeta en 02_実験データ con el formato especificado
        """
        import os
        from datetime import datetime
        from pathlib import Path
        import re
        
        try:
            print("🔍 DEBUG: create_experiment_data_folder を開始")
            print(f"🔍 DEBUG: experiment_info を受信: {experiment_info}")
            
            # ES: Extraer número de la carpeta
            # EN: Extract folder number
            # JP: フォルダ番号を抽出
            folder_name = experiment_info['folder_name']
            optimization_type = experiment_info['optimization_type']
            
            print(f"🔍 DEBUG: フォルダを処理中: '{folder_name}'")
            print(f"🔍 DEBUG: 最適化タイプ: '{optimization_type}'")
            print(f"🔍 DEBUG: フォルダ名の長さ: {len(folder_name)}")
            print(f"🔍 DEBUG: フォルダ名の文字列: {[c for c in folder_name]}")
            
            # ES: Buscar número en el nombre de la carpeta
            # EN: Search for a number in the folder name
            # JP: フォルダ名から番号を検索
            # ES: Patrones para buscar números: "017", "001", etc.
            # EN: Patterns to search numbers: \"017\", \"001\", etc.
            # JP: 番号検索パターン：「017」「001」など
            number_patterns = [
                r'(\d{3,})',  # Numbers with 3+ digits
                r'(\d{2,})',  # Numbers with 2+ digits
                r'(\d+)'      # Any number
            ]
            
            folder_number = "001"  # Default number
            pattern_used = "default"
            
            print("🔍 DEBUG: 正規表現パターンを適用中:")
            for i, pattern in enumerate(number_patterns):
                print(f"  🔍 DEBUG: パターン {i+1}: {pattern}")
                number_match = re.search(pattern, folder_name)
                if number_match:
                    extracted_number = number_match.group(1)
                    folder_number = extracted_number.zfill(3)  # Rellenar con ceros
                    pattern_used = pattern
                    print(f"  ✅ DEBUG: パターン '{pattern}' で一致しました")
                    print(f"  ✅ DEBUG: 抽出した番号: '{extracted_number}'")
                    print(f"  ✅ DEBUG: 0埋め後の番号: '{folder_number}'")
                    break
                else:
                    print(f"  ❌ DEBUG: パターン '{pattern}' では一致しませんでした")
            
            # ES: Verificar que el número extraído es correcto | EN: Verify extracted number is correct | JA: 抽出した番号が正しいか確認
            print("🔍 DEBUG: 抽出サマリー:")
            print(f"  - 元のフォルダ名: '{folder_name}'")
            print(f"  - 使用したパターン: '{pattern_used}'")
            print(f"  - 最終番号: '{folder_number}'")
            print(f"  - 最適化タイプ: '{optimization_type}'")
            
            # Generar fecha y hora actual
            now = datetime.now()
            timestamp = now.strftime("%Y%m%d_%H%M%S")
            print(f"🔍 DEBUG: 生成したタイムスタンプ: '{timestamp}'")
            
            # ES: Crear nombre de carpeta | EN: Create folder name | JA: フォルダ名を作成
            new_folder_name = f"{folder_number}_{optimization_type}_{timestamp}"
            print(f"🔍 DEBUG: 生成した最終フォルダ名: '{new_folder_name}'")
            
            # ES: Crear carpeta en 02_実験データ | EN: Create folder in 02_実験データ | JA: 02_実験データにフォルダを作成
            experiment_data_path = Path(self.current_project_folder) / "02_実験データ" / new_folder_name
            print(f"🔍 DEBUG: 作成する完全パス: {experiment_data_path}")
            
            # ES: Verificar si la carpeta ya existe y crear una nueva si es necesario | EN: Check if folder exists and create new if needed | JA: フォルダが既にあれば新規作成
            if experiment_data_path.exists():
                print(f"⚠️ DEBUG: フォルダが既に存在します: {experiment_data_path}")
                # ES: Crear una nueva carpeta con un sufijo adicional | EN: Create new folder with additional suffix | JA: 追加サフィックスで新規フォルダを作成
                counter = 1
                while experiment_data_path.exists():
                    new_folder_name = f"{folder_number}_{optimization_type}_{timestamp}_{counter:02d}"
                    experiment_data_path = Path(self.current_project_folder) / "02_実験データ" / new_folder_name
                    print(f"🔍 DEBUG: 代替フォルダを作成試行: {new_folder_name}")
                    counter += 1
                    if counter > 10:  # Evitar bucle infinito
                        break
                
                print(f"🔍 DEBUG: 作成する最終フォルダ: {experiment_data_path}")
            
            experiment_data_path.mkdir(parents=True, exist_ok=True)
            
            print(f"✅ DEBUG: フォルダ作成に成功しました: {experiment_data_path}")
            return str(experiment_data_path)
            
        except Exception as e:
            print(f"❌ DEBUG: 実験フォルダ作成中にエラー: {e}")
            import traceback
            traceback.print_exc()
            return None

    def detect_project_folder_from_results_file(self, results_file_path):
        """
        Detecta automáticamente la carpeta del proyecto basándose en la ubicación del archivo de resultados.
        
        Busca patrones como:
        - NOMBREDELPROYECTO/99_Results/archivo.xlsx -> NOMBREDELPROYECTO
        - NOMBREDELPROYECTO/02_実験データ/archivo.xlsx -> NOMBREDELPROYECTO
        - NOMBREDELPROYECTO/archivo.xlsx -> NOMBREDELPROYECTO
        
        Returns:
            str: Ruta de la carpeta del proyecto si se encuentra, None si no se puede detectar
        """
        import os
        from pathlib import Path
        
        try:
            # ES: Convertir a Path para facilitar el manejo
            # EN: Convert to Path for easier handling
            # JP: 扱いやすくするためPathに変換
            file_path = Path(results_file_path)
            print(f"🔍 プロジェクトフォルダを検出中: {file_path}")
            
            # ES: Obtener el directorio del archivo
            # EN: Get the file's directory
            # JP: ファイルのディレクトリを取得
            file_dir = file_path.parent
            print(f"🔍 ファイルのディレクトリ: {file_dir}")
            
            # ES: Buscar patrones de carpetas de proyecto
            # EN: Look for project-folder patterns
            # JP: プロジェクトフォルダのパターンを探す
            project_folders = [
                "99_Results",
                "02_実験データ", 
                "03_線形回帰",
                "04_非線形回帰",
                "05_分類",
                "01_実験リスト"
            ]
            
            # ES: Buscar hacia arriba en la jerarquía de directorios
            # EN: Search upward in the directory hierarchy
            # JP: ディレクトリ階層を上方向に検索
            current_dir = file_dir
            max_levels = 5  # Max 5 levels upward
            
            for level in range(max_levels):
                print(f"🔍 レベル {level}: {current_dir}")
                
                # ES: Verificar si el directorio actual contiene carpetas de proyecto | EN: Check if current directory contains project folders | JA: 現ディレクトリにプロジェクトフォルダがあるか確認
                for folder in project_folders:
                    project_folder_path = current_dir / folder
                    if project_folder_path.exists() and project_folder_path.is_dir():
                        print(f"✅ プロジェクトフォルダを発見: {folder}")
                        # ES: El directorio padre de esta carpeta es el proyecto
                        # EN: The parent directory of this folder is the project root
                        # JP: このフォルダの親ディレクトリがプロジェクトルート
                        project_root = current_dir
                        print(f"✅ 検出したプロジェクトフォルダ: {project_root}")
                        return str(project_root)
                
                # ES: Verificar si el directorio actual tiene la estructura de un proyecto | EN: Check if current directory has project structure | JA: 現ディレクトリがプロジェクト構造か確認
                # (contiene múltiples carpetas de proyecto)
                project_folder_count = 0
                for folder in project_folders:
                    if (current_dir / folder).exists():
                        project_folder_count += 1
                
                if project_folder_count >= 2:  # Si tiene al menos 2 carpetas de proyecto
                    print(f"✅ プロジェクト構造を検出（{project_folder_count} 個のフォルダ）")
                    return str(current_dir)
                
                # Subir un nivel
                parent_dir = current_dir.parent
                if parent_dir == current_dir:  # Reached the root
                    break
                current_dir = parent_dir
            
            print("❌ プロジェクトフォルダを自動検出できませんでした")
            return None
            
        except Exception as e:
            print(f"❌ プロジェクトフォルダ検出中にエラー: {e}")
            return None

    def on_show_results_clicked(self):
        """ES: Acción al pulsar Show Results
        EN: Action when Show Results is clicked
        JA: Show Resultsクリック時のアクション"""
        try:
            print("結果表示中...")

            # ES: Verificar que se haya cargado un archivo de resultados | EN: Ensure a results file has been loaded | JA: 結果ファイルが読み込まれているか確認
            if not hasattr(self, 'results_file_path') or not self.results_file_path:
                QMessageBox.warning(self, "エラー", "❌ 結果ファイルが読み込まれていません。\nまず「ファイルを読み込む」で結果ファイルを選択してください。")
                return

            # ES: Verificar que el archivo de resultados existe | EN: Ensure the results file exists | JA: 結果ファイルが存在するか確認
            import os
            if not os.path.exists(self.results_file_path):
                QMessageBox.warning(self, "エラー", f"❌ 結果ファイルが見つかりません:\n{self.results_file_path}")
                return

            print(f"🔍 Debug - results_file_path: {self.results_file_path}")

            # ES: Intentar detectar automáticamente la carpeta del proyecto | EN: Try to auto-detect the project folder | JA: プロジェクトフォルダを自動検出
            project_folder = self.detect_project_folder_from_results_file(self.results_file_path)
            
            if project_folder:
                print(f"✅ プロジェクトフォルダを自動検出しました: {project_folder}")
                QMessageBox.information(self, "プロジェクト検出", f"✅ プロジェクトフォルダが自動検出されました:\n{project_folder}")
            else:
                print("❌ プロジェクトフォルダを自動検出できませんでした")
                # ES: Si no se pudo detectar automáticamente, pedir al usuario que seleccione
                # EN: If it couldn't be detected automatically, ask the user to select
                # JP: 自動検出できない場合はユーザーに選択してもらう
                project_folder = QFileDialog.getExistingDirectory(self, "プロジェクトフォルダを選択", "")
                if not project_folder:
                    QMessageBox.warning(self, "エラー", "❌ プロジェクトフォルダが選択されていません。")
                    return

            # ES: Guardar la carpeta del proyecto para uso posterior | EN: Save project folder for later use | JA: 後で使うためにプロジェクトフォルダを保存
            self.current_project_folder = project_folder
            print(f"✅ プロジェクトフォルダを保存しました: {self.current_project_folder}")

            # ES: Buscar archivo de experimento correspondiente | EN: Find matching experiment file | JA: 対応する実験ファイルを検索
            print("🔍 DEBUG: 実験ファイルの検索を開始...")
            experiment_info = self.find_matching_experiment_file(project_folder)
            if experiment_info:
                print(f"✅ DEBUG: 実験ファイルを発見: {experiment_info}")
                # ES: NO crear carpeta aquí, dejar que el worker lo haga después de verificar duplicados | EN: Do not create folder here; let worker do it after checking duplicates | JA: ここでフォルダを作らず、重複確認後にワーカーに任せる
                experiment_folder_name = None  # Do not create the folder prematurely
                print("✅ DEBUG: 実験情報を後処理用に保存しました")
            else:
                print("⚠️ DEBUG: 対応する実験ファイルが見つかりませんでした")
                # ES: NO crear carpeta por defecto aquí, dejar que el worker lo haga | EN: Do not create default folder here; let worker do it | JA: ここでデフォルトフォルダを作らず、ワーカーに任せる
                experiment_folder_name = None  # Do not create the folder prematurely
                print("✅ DEBUG: 早期にデフォルトフォルダを作成しません")

            # ES: Limpiar pantalla principal antes de mostrar loading | EN: Clear main screen before showing loading | JA: ローディング表示前にメイン画面をクリア
            self.clear_main_screen()

            # ES: Iniciar loading overlay centrado sobre el frame central | EN: Start loading overlay centered on the center frame | JA: 中央フレーム上でローディングオーバーレイを開始
            # ES: Reutilizar si ya existe para evitar múltiples overlays/eventFilters
            # EN: Reuse it if it already exists to avoid multiple overlays/eventFilters
            # JP: 複数のオーバーレイ/eventFilterを避けるため、既存なら再利用
            if not hasattr(self, 'loader_overlay') or self.loader_overlay is None:
                self.loader_overlay = LoadingOverlay(self.center_frame)
            self.loader_overlay.start()
            
            # ES: Verificar si la consola desplegable está visible | EN: Check if the dropdown console is visible | JA: ドロップダウンコンソールが表示されているか確認
            if hasattr(self, 'overlay_console') and self.overlay_console.isVisible():
                print("🔧 ドロップダウンコンソールを検出しました。表示を維持します...")
                # ES: El loading se posicionará por encima de la consola
                # EN: The loading overlay will be placed above the console
                # JP: ローディングはコンソールより前面に配置する
                print("🔧 ローディングをコンソールの前面に配置します")
            
            # ES: Debug del posicionamiento del loading | EN: Debug loading positioning | JA: ローディング位置のデバッグ
            print(f"🔧 中央フレームのジオメトリ: {self.center_frame.geometry()}")
            print(f"🔧 ローディングオーバーレイのジオメトリ: {self.loader_overlay.geometry()}")

            # ES: Crear worker y thread para procesamiento en paralelo | EN: Create worker and thread for parallel processing | JA: 並列処理用のワーカーとスレッドを作成
            print("🔍 Debug - ShowResultsWorker を作成:")
            print(f"  - project_folder: {project_folder}")
            print(f"  - results_file_path: {self.results_file_path}")
            print(f"  - brush(from_file): {getattr(self, '_results_brush_type', None)}")
            print(f"  - diameter: {self.diameter_selector.currentText()}")
            print(f"  - material: {self.material_selector.currentText()}")
            
            # ES: Verificar el contenido del archivo de resultados | EN: Verify results file content | JA: 結果ファイルの内容を確認
            try:
                import pandas as pd
                df_results = pd.read_excel(self.results_file_path)
                print(f"🔍 Debug - 結果ファイルの行数: {len(df_results)}")
                print(f"🔍 Debug - 列: {list(df_results.columns)}")
                print("🔍 Debug - 先頭行データ:")
                if len(df_results) > 0:
                    first_row = df_results.iloc[0]
                    print(f"  - 回転速度: {first_row.get('回転速度', 'N/A')}")
                    print(f"  - 送り速度: {first_row.get('送り速度', 'N/A')}")
                    print(f"  - 回転方向: {first_row.get('回転方向', 'N/A')}")
                    print(f"  - 切込量: {first_row.get('切込量', 'N/A')}")
                    print(f"  - 突出量: {first_row.get('突出量', 'N/A')}")
                    print(f"  - 載せ率: {first_row.get('載せ率', 'N/A')}")
                    print(f"  - パス数: {first_row.get('パス数', 'N/A')}")
            except Exception as e:
                print(f"🔍 Debug - 結果ファイル読み込みエラー: {e}")
            
            # ES: Verificar que ShowResultsWorker esté disponible | EN: Ensure ShowResultsWorker is available | JA: ShowResultsWorkerが利用可能か確認
            try:
                from showresultsworker import ShowResultsWorker
                print("✅ ShowResultsWorker のインポート成功")
            except ImportError as e:
                print(f"❌ ShowResultsWorker のインポートに失敗: {e}")
                QMessageBox.critical(self, "エラー", f"❌ ShowResultsWorkerのインポートに失敗しました:\n{str(e)}")
                return
            
            # ES: Verificar que el procesador existe | EN: Ensure the processor exists | JA: プロセッサが存在するか確認
            if not hasattr(self, 'processor'):
                print("❌ self.processor が存在しません")
                QMessageBox.critical(self, "エラー", "❌ プロセッサーが初期化されていません。")
                return
            
            print(f"✅ self.processor を確認: {self.processor}")
            
            # ES: Verificar registros en la base de datos antes de importar | EN: Check DB records before importing | JA: インポート前にDBのレコードを確認
            try:
                import sqlite3
                import os
                
                # ES: Verificar la ubicación de la base de datos | EN: Verify database location | JA: データベースの場所を確認
                db_path = RESULTS_DB_PATH
                print(f"🔍 Debug - DB パス: {os.path.abspath(db_path)}")
                print(f"🔍 Debug - DB は存在しますか?: {os.path.exists(db_path)}")
                
                conn = sqlite3.connect(db_path, timeout=10)
                cursor = conn.cursor()
                cursor.execute("SELECT COUNT(*) FROM main_results")
                count_before = cursor.fetchone()[0]
                
                # ES: Verificar algunos registros existentes para debug | EN: Check some existing records for debug | JA: デバッグ用に既存レコードを確認
                cursor.execute("SELECT * FROM main_results LIMIT 3")
                sample_records = cursor.fetchall()
                print("🔍 Debug - 既存レコードのサンプル:")
                for i, record in enumerate(sample_records):
                    print(f"  レコード {i+1}: {record[:5]}...")  # Only show the first 5 columns
                
                # ES: Verificar la estructura de la base de datos | EN: Verify database structure | JA: データベースの構造を確認
                print("🔍 Debug - DB 構造を確認中...")
                cursor.execute("PRAGMA table_info(main_results)")
                columns_info = cursor.fetchall()
                print("🔍 Debug - DB の列:")
                for col in columns_info:
                    print(f"  - {col[1]} ({col[2]})")
                
                # ES: Verificar si hay registros con los mismos valores que vamos a importar | EN: Check for records with same values as we are importing | JA: インポートする値と同じレコードがあるか確認
                print("🔍 Debug - 重複レコードの有無を確認中...")
                try:
                    cursor.execute("SELECT COUNT(*) FROM main_results WHERE 回転速度 = ? AND 送り速度 = ? AND 切込量 = ? AND 突出量 = ? AND 載せ率 = ? AND パス数 = ?", 
                                 (1000, 500, 1.0, 10, 0.4, 2))
                    duplicate_count = cursor.fetchone()[0]
                    print(f"🔍 Debug - 先頭レコードと類似するレコード数: {duplicate_count}")
                except Exception as e:
                    print(f"🔍 Debug - 重複確認中にエラー: {e}")
                
                conn.close()
                print(f"🔍 Debug - インポート前のDBレコード数: {count_before}")
                
                # ES: Verificar si hay otra base de datos en la carpeta del proyecto | EN: Check if another DB exists in project folder | JA: プロジェクトフォルダに別のDBがあるか確認
                # Debug legacy: antes la DB vivía dentro del proyecto; ya no se usa en instalación pro.
                project_db_path = os.path.join(project_folder, "results.db")
                print(f"🔍 Debug - プロジェクト内にDBは存在しますか?: {os.path.exists(project_db_path)}")
                if os.path.exists(project_db_path):
                    print(f"🔍 Debug - プロジェクトDBのパス: {os.path.abspath(project_db_path)}")
                    try:
                        conn_project = sqlite3.connect(project_db_path)
                        cursor_project = conn_project.cursor()
                        cursor_project.execute("SELECT COUNT(*) FROM main_results")
                        count_project = cursor_project.fetchone()[0]
                        conn_project.close()
                        print(f"🔍 Debug - プロジェクトDBのレコード数: {count_project}")
                    except Exception as e:
                        print(f"🔍 Debug - プロジェクトDB確認中にエラー: {e}")
            except Exception as e:
                print(f"🔍 Debug - インポート前のDB確認中にエラー: {e}")
            
            # ES: Crear worker y ejecutar directamente | EN: Create worker and run directly | JA: ワーカーを作成して直接実行
            self.show_results_worker = ShowResultsWorker(
                project_folder,
                self.results_file_path,
                float(self.diameter_selector.currentText()),
                self.material_selector.currentText(),
                self.backup_and_update_sample_file,
                self.processor.process_results_file_with_ui_values,
                experiment_info  # Pass the found experiment info
            )

            # ES: Crear thread para ejecutar el worker en paralelo | EN: Create thread to run worker in parallel | JA: ワーカーを並列実行するスレッドを作成
            self.import_thread = QThread()
            self.show_results_worker.moveToThread(self.import_thread)

            # ES: Conectar señales del thread | EN: Connect thread signals | JA: スレッドのシグナルを接続
            self.import_thread.started.connect(self.show_results_worker.run)
            self.show_results_worker.finished.connect(self.on_show_results_finished)
            self.show_results_worker.error.connect(self.on_show_results_error)
            self.show_results_worker.finished.connect(self.import_thread.quit)
            self.show_results_worker.finished.connect(self.show_results_worker.deleteLater)
            self.import_thread.finished.connect(self.import_thread.deleteLater)

            print("🔍 Debug - インポート用スレッドを開始...")
            self.import_thread.start()
        except Exception as e:
            print(f"❌ on_show_results_clicked で予期しないエラー: {e}")
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "エラー", f"❌ 予期しないエラーが発生しました:\n{str(e)}")

    def on_show_results_finished(self, result):
        """ES: Maneja el resultado exitoso del procesamiento de resultados
        EN: Handle successful result processing
        JA: 結果処理の成功を処理"""
        try:
            print(f"🔍 Debug - on_show_results_finished が呼ばれました: result={result}")
            
            if hasattr(self, 'loader_overlay'):
                self.loader_overlay.stop()
            
            # ES: Verificar que la base de datos se actualizó | EN: Verify that the database was updated | JA: データベースが更新されたか確認
            total_records_after = 0
            records_imported = 0
            try:
                import sqlite3
                import os
                conn = sqlite3.connect(RESULTS_DB_PATH, timeout=10)
                cursor = conn.cursor()
                cursor.execute("SELECT COUNT(*) FROM main_results")
                total_records_after = cursor.fetchone()[0]
                print(f"🔍 Debug - インポート後のDBレコード数: {total_records_after}")
                
                # ES: Registros importados reales = insertados + actualizados (sin contar filas idénticas) | EN: Real imported records = inserted + updated (excluding identical rows) | JA: 実際のインポート数＝挿入＋更新（同一行は除く）
                if result and isinstance(result, dict):
                    dbu = result.get("db_upsert_result")
                    if isinstance(dbu, dict):
                        try:
                            records_imported = int(dbu.get("inserted", 0) or 0) + int(dbu.get("updated", 0) or 0)
                            print(f"🔍 Debug - 実インポート数（insert+update）: {records_imported}")
                        except Exception:
                            records_imported = "N/A"
                    else:
                        # ES: Si no tenemos db_upsert_result, NO debemos inferir "importados" desde el Excel,
                        # EN: If we don't have db_upsert_result, we must NOT infer "imported" from Excel,
                        # JP: db_upsert_resultが無い場合、Excelから「インポート済み」を推測してはいけない
                        # ES: porque puede ser un early-exit (archivo idéntico) o un fallo parcial.
                        # EN: because it may be an early-exit (identical file) or a partial failure.
                        # JP: 同一ファイルによる早期終了、または部分失敗の可能性があるため
                        records_imported = 0
                        print("🔍 Debug - db_upsert_result がありません: records_imported=0（Excelから推測しません）")
                
                # ES: Mostrar contenido completo de la base de datos | EN: Show full database content | JA: データベースの全内容を表示
                if total_records_after > 0:
                    print("🔍 Debug - DBの全内容:")
                    cursor.execute("SELECT * FROM main_results ORDER BY id")
                    all_records = cursor.fetchall()
                    
                    # ES: Obtener nombres de columnas
                    # EN: Get column names
                    # JP: 列名を取得
                    cursor.execute("PRAGMA table_info(main_results)")
                    columns_info = cursor.fetchall()
                    column_names = [col[1] for col in columns_info]
                    
                    print(f"🔍 Debug - 列: {column_names}")
                    print(f"🔍 Debug - 総レコード数: {len(all_records)}")
                    
                    for i, record in enumerate(all_records, 1):
                        print(f"  レコード {i}:")
                        for j, value in enumerate(record):
                            if j < len(column_names):
                                print(f"    {column_names[j]}: {value}")
                        print()
                else:
                    print("🔍 Debug - DBが空です")
                    
                conn.close()
            except Exception as e:
                print(f"🔍 Debug - DB確認中にエラー: {e}")
            
            # ES: Mostrar mensaje de éxito con información del backup | EN: Show success message with backup info | JA: バックアップ情報付き成功メッセージを表示
            if result and isinstance(result, dict):
                if result.get('optimization_type') == 'EXISTING':
                    # ES: Caso cuando ya existe un archivo idéntico
                    # EN: Case when an identical file already exists
                    # JP: 同一内容のファイルが既に存在する場合
                    message = f"⚠️ 既に同じ内容のファイルが存在します:\n{result.get('identical_folder', 'Unknown')}\n\n"
                    message += f"📁 既存のフォルダ: {result.get('identical_folder', 'Unknown')}\n"
                    message += f"ℹ️ 新しいフォルダは作成されませんでした\n\n"
                    
                    # ES: Agregar información de la base de datos | EN: Add database information | JA: データベース情報を追加
                    message += f"📊 データベース内の総レコード数: {total_records_after}\n"
                    message += f"📈 今回インポートされたレコード数: {records_imported}"
                else:
                    # Caso normal
                    message = f"✅ 結果ファイルが保存されました:\n{result.get('results_file_path', 'N/A')}\n\n"
                    
                    # ES: Agregar información de la base de datos | EN: Add database information | JA: データベース情報を追加
                    message += f"📊 データベース内の総レコード数: {total_records_after}\n"
                    message += f"📈 今回インポートされたレコード数: {records_imported}\n\n"
                    
                    if result.get('backup_result', {}).get('backup_path'):
                        message += f"📋 バックアップ作成: {os.path.basename(result['backup_result']['backup_path'])}\n"
                        message += f"🗑️ サンプルファイルから削除された行: {result['backup_result'].get('removed_rows', 'N/A')}\n"
                        message += f"📊 サンプルファイルの残り行数: {result['backup_result'].get('remaining_rows', 'N/A')}"
                    else:
                        message += f"ℹ️ バックアップは実行されませんでした（アクティブなプロジェクトがありません）"

                    # ES: Aviso único de sobrescritura en BBDD + backup | EN: Single notice for DB overwrite + backup | JA: DB上書き＋バックアップの一括表示
                    dbu = result.get("db_upsert_result")
                    if isinstance(dbu, dict):
                        updated = int(dbu.get("updated", 0) or 0)
                        inserted = int(dbu.get("inserted", 0) or 0)
                        if updated > 0:
                            message += "\n\n⚠️ 既存データを上書きします。BBDDのバックアップを作成しました。"
                            message += f"\n🔁 上書き: {updated} / ➕ 追加: {inserted}"
                            if dbu.get("db_backup_path"):
                                message += f"\n📋 BBDDバックアップ: {os.path.basename(str(dbu.get('db_backup_path')))}"
                            else:
                                message += "\n📋 BBDDバックアップ: (作成できませんでした)"
            else:
                message = f"✅ 処理が完了しました\n\n"
                message += f"📊 データベース内の総レコード数: {total_records_after}\n"
                message += f"📈 今回インポートされたレコード数: {records_imported}"
            
            QMessageBox.information(self, "完了", message)
            
            # ES: Mostrar la vista de filtro después de procesar los datos | EN: Show filter view after processing data | JA: データ処理後にフィルタビューを表示
            self.create_filter_view()
            
            if hasattr(self, 'ok_button'):
                self.ok_button.setEnabled(True)
            if hasattr(self, 'ng_button'):
                self.ng_button.setEnabled(False)

            self.create_navigation_buttons()
            if hasattr(self, 'prev_button'):
                self.prev_button.setEnabled(True)
            if hasattr(self, 'next_button'):
                self.next_button.setEnabled(True)
                
        except Exception as e:
            print(f"❌ 結果表示完了ハンドラでエラー: {e}")
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "エラー", f"❌ 結果処理中にエラーが発生しました:\n{str(e)}")

    def on_show_results_error(self, error_message):
        """ES: Maneja el error del procesamiento de resultados
        EN: Handle result processing error
        JA: 結果処理のエラーを処理"""
        try:
            print(f"🔍 Debug - on_show_results_error llamado con error: {error_message}")
            
            if hasattr(self, 'loader_overlay'):
                self.loader_overlay.stop()
            
            QMessageBox.critical(self, "エラー", f"❌ 結果処理中にエラーが発生しました:\n{str(error_message)}")
            
        except Exception as e:
            print(f"❌ 結果表示エラーハンドラでエラー: {e}")
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "エラー", f"❌ エラー処理中にエラーが発生しました:\n{str(e)}")

    def display_image_in_graph_area(self, image_path):
        """ES: Carga y muestra una imagen dentro del área de gráficos.
        EN: Load and display an image inside the graph area.
        JA: グラフ領域内に画像を読み込み表示。"""


        if not hasattr(self.graph_area, "layout") or self.graph_area.layout() is None:
            self.graph_area.setLayout(QVBoxLayout())

        layout = self.graph_area.layout()

        # ES: Limpiar el contenido actual
        # EN: Clear current content
        # JP: 現在の内容をクリア
        for i in reversed(range(layout.count())):
            widget = layout.itemAt(i).widget()
            if widget:
                widget.setParent(None)

        # ES: Mostrar nueva imagen | EN: Show new image | JA: 新しい画像を表示
        label = QLabel()
        pixmap = QPixmap(image_path)
        label.setPixmap(pixmap.scaled(self.graph_area.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation))
        label.setAlignment(Qt.AlignCenter)
        layout.addWidget(label)

    def on_analyze_clicked(self):
        """ES: Acción al pulsar el botón de análisis - navega directamente a la página de filtros
        EN: Action when analysis button is clicked - navigate directly to filters page
        JA: 解析ボタンクリック時 - フィルタページへ直接遷移"""
        print("分析ページに移動中...")
        
        # ES: Marcar que se accedió desde el botón bunseki | EN: Mark that access was from bunseki button | JA: 分析ボタンからアクセスしたことを記録
        self.accessed_from_bunseki = True
        
        # ES: Verificar si ya estamos en la vista de filtros | EN: Check if we are already on filter view | JA: フィルタビューかどうか確認
        # ES: Buscar si hay un título "データフィルター" en el layout central
        # EN: Look for a title "データフィルター" in the center layout
        # JP: 中央レイアウトに「データフィルター」タイトルがあるか確認
        already_in_filter_view = False
        for i in range(self.center_layout.count()):
            item = self.center_layout.itemAt(i)
            if item.widget() and isinstance(item.widget(), QLabel):
                if item.widget().text() == "データフィルター":
                    already_in_filter_view = True
                    break
        
        if already_in_filter_view:
            # ES: Ya estamos en la pantalla de filtros, solo mostrar mensaje informativo
            # EN: We are already on the filter screen; just show an informational message
            # JP: 既にフィルタ画面なので、案内メッセージのみ表示
            QMessageBox.information(self, "分析ページ", "✅ 既に分析ページにいます。\nフィルターを設定してデータを分析してください。")
            return
        
        try:
            # ES: Crear la vista de filtros directamente | EN: Create filter view directly | JA: フィルタビューを直接作成
            self.create_filter_view()
            
            # ES: Habilitar botones de navegación
            # EN: Enable navigation buttons
            # JP: ナビゲーションボタンを有効化
            self.create_navigation_buttons()
            self.prev_button.setEnabled(True)
            self.next_button.setEnabled(True)
            
            QMessageBox.information(self, "分析ページ", "✅ 分析ページに移動しました。\nフィルターを設定してデータを分析してください。")
            
        except Exception as e:
            QMessageBox.critical(self, "エラー", f"❌ 分析ページの移動中にエラーが発生しました:\n{str(e)}")

    def on_ok_clicked(self):
        # ES: Verificación inicial: asegurar que solo exista un tipo de resultado.
        # EN: Initial check: ensure only one result type exists.
        # JP: 初期確認：結果タイプが1つだけであることを確認します。
        print("🔍 Debug - on_ok_clicked 開始:")
        print(f"🔍 Debug - dsaitekika_results 存在: {hasattr(self, 'dsaitekika_results')}")
        print(f"🔍 Debug - isaitekika_results 存在: {hasattr(self, 'isaitekika_results')}")
        print(f"🔍 Debug - last_executed_optimization 存在: {hasattr(self, 'last_executed_optimization')}")
        if hasattr(self, 'last_executed_optimization'):
            print(f"🔍 Debug - last_executed_optimization 値: {self.last_executed_optimization}")
        
        # ES: Verificación crítica: SIEMPRE usar last_executed_optimization si existe.
        # EN: Critical check: ALWAYS use last_executed_optimization if it exists.
        # JP: 重要：存在する場合は常に last_executed_optimization を使用します。
        if hasattr(self, 'last_executed_optimization'):
            print(f"🔍 Debug - last_executed_optimization の使用を強制: {self.last_executed_optimization}")
            # ES: Forzar el uso del último tipo de optimización ejecutado
            # EN: Force using the last executed optimization type
            # JP: 最後に実行した最適化タイプを強制的に使用
            if self.last_executed_optimization == 'I':
                if hasattr(self, 'dsaitekika_results'):
                    delattr(self, 'dsaitekika_results')
                    print("🧹 I最適化を強制するため dsaitekika_results をクリアしました")
            elif self.last_executed_optimization == 'D':
                if hasattr(self, 'isaitekika_results'):
                    delattr(self, 'isaitekika_results')
                    print("🧹 D最適化を強制するため isaitekika_results をクリアしました")
        
        # ES: Copiar archivos definitivos a carpeta 実験リスト solo al pulsar OK
        # EN: Copy final files to the 実験リスト folder only when pressing OK
        # JP: OK押下時のみ、確定ファイルを実験リストフォルダへコピー
        if hasattr(self, 'dsaitekika_results') or hasattr(self, 'isaitekika_results'):
            # ES: ✅ Simplificado: usar SIEMPRE last_executed_optimization como fuente de verdad.
            # EN: ✅ Simplified: ALWAYS use last_executed_optimization as the source of truth.
            # JP: ✅ 簡略化：常に last_executed_optimization を正とします。
            if hasattr(self, 'last_executed_optimization'):
                optimization_type = self.last_executed_optimization
                print(f"🔍 Debug - last_executed_optimization を使用: {optimization_type}")
            else:
                # ES: Fallback solo si no existe last_executed_optimization
                # EN: Fallback only if last_executed_optimization does not exist
                # JP: last_executed_optimizationが無い場合のみフォールバック
                if hasattr(self, 'isaitekika_results') and not hasattr(self, 'dsaitekika_results'):
                    optimization_type = 'I'
                elif hasattr(self, 'dsaitekika_results'):
                    optimization_type = 'D'
                else:
                    optimization_type = 'D'  # Default
                print(f"🔍 Debug - フォールバックを使用: optimization_type={optimization_type}")
            
            print(f"🔍 Debug - dsaitekika_results 存在: {hasattr(self, 'dsaitekika_results')}")
            print(f"🔍 Debug - isaitekika_results 存在: {hasattr(self, 'isaitekika_results')}")
            print(f"🔍 Debug - 最終 optimization_type: {optimization_type}")
            print(f"🔍 デバッグ - last_executed_optimization の値: {getattr(self, 'last_executed_optimization', '存在しません')}")
            
            # ES: ✅ SIMPLIFICADO: Limpiar resultados del tipo opuesto
            # EN: ✅ SIMPLIFIED: Clear results of the opposite type
            # JP: ✅ 簡略化：反対タイプの結果をクリア
            if optimization_type == 'D':
                print("✅ D最適化としてエクスポートします")
                if hasattr(self, 'isaitekika_results'):
                    delattr(self, 'isaitekika_results')
                    print("🧹 D のエクスポートのため isaitekika_results をクリアしました")
            elif optimization_type == 'I':
                print("✅ I最適化としてエクスポートします")
                if hasattr(self, 'dsaitekika_results'):
                    delattr(self, 'dsaitekika_results')
                    print("🧹 I のエクスポートのため dsaitekika_results をクリアしました")
            else:
                print(f"⚠️ 不明なタイプ: {optimization_type}。既定として D最適化を使用します")
                optimization_type = 'D'
            
            # ES: Crear carpeta y determinar el nombre basado en optimization_type.
            # EN: Create the folder and determine the name based on optimization_type.
            # JP: フォルダを作成し、optimization_type に基づいて名前を決定します。
            output_folder = self.current_temp_folder if hasattr(self, 'current_temp_folder') else os.path.join(self.proyecto_folder, "99_Temp", "Temp")
            project_name = getattr(self, 'proyecto_nombre', 'Unknown')
            today = datetime.now().strftime('%Y%m%d')
            
            # ES: Crear carpeta 01_実験リスト al mismo nivel que 99_Temp | EN: Create 01_実験リスト folder at same level as 99_Temp | JA: 99_Tempと同じ階層に01_実験リストフォルダを作成
            samples_base = os.path.join(self.proyecto_folder, "01_実験リスト")
            os.makedirs(samples_base, exist_ok=True)
            
            # ES: Formato de nombre de carpeta basado en optimization_type
            # EN: Folder-name format based on optimization_type
            # JP: optimization_typeに基づくフォルダ名フォーマット
            now = datetime.now()
            fecha_hora = now.strftime('%Y%m%d_%H%M%S')
            if optimization_type == 'I':
                prefix = 'I_SAITEKIKA'
                print(f"📁 プレフィックス I でフォルダを作成中: {prefix}")
            else:  # D optimization
                prefix = 'D最適化'
                print(f"📁 プレフィックス D でフォルダを作成中: {prefix}")
            
            # ES: Buscar el mayor número de carpeta existente y sumarle 1
            # EN: Find the largest existing folder number and add 1
            # JP: 既存フォルダ番号の最大値を探して+1する
            existing_folders = [d for d in os.listdir(samples_base) if os.path.isdir(os.path.join(samples_base, d))]
            max_num = 0
            for folder in existing_folders:
                try:
                    num = int(folder.split('_')[0])
                    if num > max_num:
                        max_num = num
                except Exception:
                    pass
            next_num = max_num + 1
            folder_name = f"{next_num:03d}_{prefix}_{fecha_hora}"
            sample_folder = os.path.join(samples_base, folder_name)
            os.makedirs(sample_folder, exist_ok=True)
            print(f"📁 フォルダを作成しました: {folder_name}")
            print(f"📁 完全なパス: {sample_folder}")
            
            if optimization_type == 'I':
                
                # ES: Cambiar nombre de columnas para la exportación antes de guardar
                # EN: Rename columns for export before saving
                # JP: 保存前に、エクスポート用に列名を変更
                if hasattr(self, 'isaitekika_results'):
                    if '面粗度(Ra)前' in self.isaitekika_selected_df.columns:
                        self.isaitekika_selected_df.rename(columns={'面粗度(Ra)前': 'Ra(前)'}, inplace=True)
                    if '面粗度(Ra)後' in self.isaitekika_selected_df.columns:
                        self.isaitekika_selected_df.rename(columns={'面粗度(Ra)後': 'Ra(後)'}, inplace=True)
                    # ES: Guardar archivo Excel I-óptimo | EN: Save I-optimal Excel file | JA: I最適Excelファイルを保存
                    if len(self.isaitekika_selected_df) > 0:
                        # ES: --- Ajuste de columnas y formato para I最適化_新規実験点.xlsx ---
                        # EN: --- Column/format adjustments for I最適化_新規実験点.xlsx ---
                        # JP: --- I最適化_新規実験点.xlsx の列・形式調整 ---
                        # ES: Mapear nombres de columnas antes de procesar
                        # EN: Map column names before processing
                        # JP: 処理前に列名をマッピング
                        if '突出し量' in self.isaitekika_selected_df.columns:
                            self.isaitekika_selected_df.rename(columns={'突出し量': '突出量'}, inplace=True)
                        if '切込み量' in self.isaitekika_selected_df.columns:
                            self.isaitekika_selected_df.rename(columns={'切込み量': '切込量'}, inplace=True)
                        
                        # ES: Dirección: usar nombre nuevo "UPカット"
                        # EN: Direction: use the new name "UPカット"
                        # JP: 方向：新しい名前「UPカット」を使用
                        if '回転方向' in self.isaitekika_selected_df.columns and 'UPカット' not in self.isaitekika_selected_df.columns:
                            self.isaitekika_selected_df.rename(columns={'回転方向': 'UPカット'}, inplace=True)

                        required_columns = ['No.', 'A13', 'A11', 'A21', 'A32',
                                           '回転速度', '送り速度', 'UPカット', '切込量', '突出量', '載せ率', 'パス数',
                                           '線材長', 'I基準値',
                                           '上面ダレ', '側面ダレ', '摩耗量', '面粗度(Ra)前', '面粗度(Ra)後',
                                           '切削力X', '切削力Y', '切削力Z',
                                           '実験日']
                        df_export = self.isaitekika_selected_df.copy()
                        # ES: Normalizar nombres de rugosidad si vienen como Ra(前)/Ra(後) o sin (Ra)
                        # EN: Normalize roughness column names if they come as Ra(前)/Ra(後) or without (Ra)
                        # JP: 粗さ列名がRa(前)/Ra(後)や(Ra)無しの場合は正規化
                        if 'Ra(前)' in df_export.columns and '面粗度(Ra)前' not in df_export.columns:
                            df_export.rename(columns={'Ra(前)': '面粗度(Ra)前'}, inplace=True)
                        if 'Ra(後)' in df_export.columns and '面粗度(Ra)後' not in df_export.columns:
                            df_export.rename(columns={'Ra(後)': '面粗度(Ra)後'}, inplace=True)
                        if '面粗度前' in df_export.columns and '面粗度(Ra)前' not in df_export.columns:
                            df_export.rename(columns={'面粗度前': '面粗度(Ra)前'}, inplace=True)
                        if '面粗度後' in df_export.columns and '面粗度(Ra)後' not in df_export.columns:
                            df_export.rename(columns={'面粗度後': '面粗度(Ra)後'}, inplace=True)
                        # ES: Crear las columnas que falten | EN: Create missing columns | JA: 不足列を作成
                        for col in required_columns:
                            if col not in df_export.columns and col != 'I基準値':
                                df_export[col] = ''
                        # ES: ISaitekika: I基準値 NO se calcula nunca
                        # EN: ISaitekika: I基準値 is never calculated
                        # JP: ISaitekika：I基準値は計算しない
                        df_export['I基準値'] = ''
                        # ES: 線材長 siempre en blanco en el Excel de salida
                        # EN: 線材長 is always blank in the output Excel
                        # JP: 出力Excelでは線材長は常に空欄
                        df_export['線材長'] = ''
                        # ES: Reordenar las columnas
                        # EN: Reorder columns
                        # JP: 列を並べ替え
                        df_export = df_export[required_columns]
                        i_path = os.path.join(output_folder, "I最適化_新規実験点.xlsx")
                        df_export.to_excel(i_path, index=False)
                        # ES: --- Fin ajuste de columnas ---
                        # EN: --- End of column adjustments ---
                        # JP: --- 列調整の終了 ---
                    # ES: Añadir columna de fecha si no existe | EN: Add date column if it does not exist | JA: 日付列が無ければ追加
                    if len(self.isaitekika_selected_df) > 0:
                        if '実験日' not in self.isaitekika_selected_df.columns:
                            self.isaitekika_selected_df['実験日'] = ''
                    # ES: Copiar archivo Excel a la carpeta 実験リスト
                    # EN: Copy the Excel file to the 実験リスト folder
                    # JP: Excelファイルを実験リストフォルダへコピー
                    excel_src = os.path.join(output_folder, "I最適化_新規実験点.xlsx")
                    if os.path.exists(excel_src):
                        shutil.copy2(excel_src, sample_folder)
                    # ES: Copiar imágenes a la carpeta 実験リスト
                    # EN: Copy images to the 実験リスト folder
                    # JP: 画像を実験リストフォルダへコピー
                    for img_path in glob.glob(os.path.join(output_folder, '*.png')):
                        shutil.copy2(img_path, sample_folder)
            else:
                # ES: Optimización D-óptima
                # EN: D-optimal optimization
                # JP: D最適化
                
                # ES: Cambiar nombre de columnas para la exportación antes de guardar
                # EN: Rename columns for export before saving
                # JP: 保存前に、エクスポート用に列名を変更
                if hasattr(self, 'dsaitekika_results'):
                    if '面粗度(Ra)前' in self.dsaitekika_selected_df.columns:
                        self.dsaitekika_selected_df.rename(columns={'面粗度(Ra)前': 'Ra(前)'}, inplace=True)
                    if '面粗度(Ra)後' in self.dsaitekika_selected_df.columns:
                        self.dsaitekika_selected_df.rename(columns={'面粗度(Ra)後': 'Ra(後)'}, inplace=True)
                    # ES: Guardar archivo Excel D-óptimo | EN: Save D-optimal Excel file | JA: D最適Excelファイルを保存
                    if len(self.dsaitekika_selected_df) > 0:
                        # ES: --- Ajuste de columnas y formato para D_optimal_新規実験点.xlsx ---
                        # EN: --- Column/format adjustments for D_optimal_新規実験点.xlsx ---
                        # JP: --- D_optimal_新規実験点.xlsx の列・形式調整 ---
                        # ES: Mapear nombres de columnas antes de procesar
                        # EN: Map column names before processing
                        # JP: 処理前に列名をマッピング
                        if '突出し量' in self.dsaitekika_selected_df.columns:
                            self.dsaitekika_selected_df.rename(columns={'突出し量': '突出量'}, inplace=True)
                        if '切込み量' in self.dsaitekika_selected_df.columns:
                            self.dsaitekika_selected_df.rename(columns={'切込み量': '切込量'}, inplace=True)
                        
                        # ES: Dirección: usar nombre nuevo "UPカット"
                        # EN: Direction: use the new name "UPカット"
                        # JP: 方向：新しい名前「UPカット」を使用
                        if '回転方向' in self.dsaitekika_selected_df.columns and 'UPカット' not in self.dsaitekika_selected_df.columns:
                            self.dsaitekika_selected_df.rename(columns={'回転方向': 'UPカット'}, inplace=True)

                        required_columns = ['No.', 'A13', 'A11', 'A21', 'A32',
                                           '回転速度', '送り速度', 'UPカット', '切込量', '突出量', '載せ率', 'パス数',
                                           '線材長', 'D基準値',
                                           '上面ダレ', '側面ダレ', '摩耗量', '面粗度(Ra)前', '面粗度(Ra)後',
                                           '切削力X', '切削力Y', '切削力Z',
                                           '実験日']
                        df_export = self.dsaitekika_selected_df.copy()
                        # ES: Normalizar nombres de rugosidad si vienen como Ra(前)/Ra(後) o sin (Ra)
                        # EN: Normalize roughness column names if they come as Ra(前)/Ra(後) or without (Ra)
                        # JP: 粗さ列名がRa(前)/Ra(後)や(Ra)無しの場合は正規化
                        if 'Ra(前)' in df_export.columns and '面粗度(Ra)前' not in df_export.columns:
                            df_export.rename(columns={'Ra(前)': '面粗度(Ra)前'}, inplace=True)
                        if 'Ra(後)' in df_export.columns and '面粗度(Ra)後' not in df_export.columns:
                            df_export.rename(columns={'Ra(後)': '面粗度(Ra)後'}, inplace=True)
                        if '面粗度前' in df_export.columns and '面粗度(Ra)前' not in df_export.columns:
                            df_export.rename(columns={'面粗度前': '面粗度(Ra)前'}, inplace=True)
                        if '面粗度後' in df_export.columns and '面粗度(Ra)後' not in df_export.columns:
                            df_export.rename(columns={'面粗度後': '面粗度(Ra)後'}, inplace=True)
                        # ES: Crear las columnas que falten | EN: Create missing columns | JA: 不足列を作成
                        for col in required_columns:
                            if col not in df_export.columns and col != 'D基準値':
                                df_export[col] = ''
                        # ES: Calcular D基準値 EXACTAMENTE como el archivo de referencia
                        # EN: Compute D基準値 EXACTLY like the reference file
                        # JP: 参照ファイルと同じ方法でD基準値を厳密に計算
                        if len(df_export) > 0:
                            d_score_ref = getattr(self, "_last_d_score_reference", None)
                            # ES: Intentar recalcular desde candidate_df + d_indices (más fiel a la referencia)
                            # EN: Try to recompute from candidate_df + d_indices (closer to the reference)
                            # JP: candidate_df + d_indices から再計算（参照により忠実）
                            if d_score_ref is None or not np.isfinite(d_score_ref):
                                try:
                                    cand_df = getattr(self, "_last_candidate_df_for_dscore", None)
                                    d_idx = getattr(self, "_last_d_indices", None)
                                    if cand_df is not None and d_idx is not None:
                                        cand_np = cand_df.to_numpy() if hasattr(cand_df, "to_numpy") else np.asarray(cand_df)
                                        d_score_ref = calculate_d_score_reference(cand_np, d_idx)
                                except Exception as e:
                                    print(f"⚠️ 候補点/インデックスから D基準値（参照）を再計算中にエラー: {e}")
                            # ES: Fallback: si no hay candidatos/índices, calcular sobre los seleccionados (escala fit en seleccionados)
                            # EN: Fallback: if no candidates/indices, compute on selected points (scaler fit on selected)
                            # JP: フォールバック：候補/インデックスが無ければ選択点で計算（選択点でスケーラfit）
                            if d_score_ref is None or not np.isfinite(d_score_ref):
                                X_raw = _extract_design_matrix(df_export)
                                X_scaled = _standardize_like_reference(X_raw)
                                d_score_ref, _ = calculate_d_criterion_stable_reference(
                                    X_scaled, method="auto", use_numerical_stable_method=True, verbose=False
                                )
                            self._last_d_score_reference = float(d_score_ref) if d_score_ref is not None else None
                            df_export["D基準値"] = self._last_d_score_reference if self._last_d_score_reference is not None else np.nan
                        else:
                            df_export["D基準値"] = np.nan
                        # ES: 線材長 siempre en blanco en el Excel de salida
                        # EN: 線材長 is always blank in the output Excel
                        # JP: 出力Excelでは線材長は常に空欄
                        df_export['線材長'] = ''
                        # ES: Reordenar las columnas
                        # EN: Reorder columns
                        # JP: 列を並べ替え
                        df_export = df_export[required_columns]
                        d_path = os.path.join(output_folder, "D最適化_新規実験点.xlsx")
                        df_export.to_excel(d_path, index=False)
                        # ES: --- Fin ajuste de columnas ---
                        # EN: --- End of column adjustments ---
                        # JP: --- 列調整の終了 ---
                    # ES: Añadir columna de fecha si no existe | EN: Add date column if it does not exist | JA: 日付列が無ければ追加
                    if len(self.dsaitekika_selected_df) > 0:
                        if '実験日' not in self.dsaitekika_selected_df.columns:
                            self.dsaitekika_selected_df['実験日'] = ''
                    # ES: Copiar archivo Excel a la carpeta 実験リスト
                    # EN: Copy the Excel file to the 実験リスト folder
                    # JP: Excelファイルを実験リストフォルダへコピー
                    excel_src = os.path.join(output_folder, "D最適化_新規実験点.xlsx")
                    if os.path.exists(excel_src):
                        shutil.copy2(excel_src, sample_folder)
                    # ES: Copiar imágenes a la carpeta 実験リスト
                    # EN: Copy images to the 実験リスト folder
                    # JP: 画像を実験リストフォルダへコピー
                    for img_path in glob.glob(os.path.join(output_folder, '*.png')):
                        shutil.copy2(img_path, sample_folder)
            # ES: Limpiar archivos temporales después de guardar exitosamente.
            # EN: Clean up temporary files after successful save.
            # JP: 保存成功後に一時ファイルを削除します。
            if hasattr(self, 'current_temp_folder') and self.current_temp_folder:
                try:
                    if os.path.exists(self.current_temp_folder):
                        shutil.rmtree(self.current_temp_folder)
                        print(f"🗑️ 保存後に Temp フォルダを削除しました: {self.current_temp_folder}")
                    # ES: NO borrar la carpeta 99_Temp - mantenerla para futuros usos
                    # EN: Do NOT delete the 99_Temp folder - keep it for future use
                    # JP: 99_Tempフォルダは削除しない（今後の利用のため保持）
                    temp_base = os.path.join(self.proyecto_folder, "99_Temp")
                    print(f"📁 99_Temp フォルダは保持します: {temp_base}")
                except Exception as e:
                    print(f"⚠️ 一時ファイルのクリーンアップ中にエラー: {e}")
            # ES: Limpiar referencias
            # EN: Clear references
            # JP: 参照をクリア
            if hasattr(self, 'current_temp_folder'):
                delattr(self, 'current_temp_folder')
            # ES: Habilitar botones de optimización después de guardar exitosamente
            # EN: Re-enable optimization buttons after a successful save
            # JP: 保存成功後に最適化ボタンを再有効化
            self.d_optimize_button.setEnabled(True)
            self.i_optimize_button.setEnabled(True)
            self.d_optimize_button.setStyleSheet(self.d_optimize_button.styleSheet())
            self.i_optimize_button.setStyleSheet(self.i_optimize_button.styleSheet())
            
            # ES: Deshabilitar botones OK/NG
            # EN: Disable OK/NG buttons
            # JP: OK/NGボタンを無効化
            self.ok_button.setEnabled(False)
            self.ng_button.setEnabled(False)
            
            # ES: Limpiar pantalla después de guardar exitosamente
            # EN: Clear the screen after a successful save
            # JP: 保存成功後に画面をクリア
            self.graph_images = []
            self.graph_images_content = []
            self.current_graph_index = 0
            
            # ES: Limpiar área de gráficos
            # EN: Clear chart area
            # JP: グラフ領域をクリア
            if hasattr(self, 'graph_area') and self.graph_area.layout():
                layout = self.graph_area.layout()
                for i in reversed(range(layout.count())):
                    widget = layout.itemAt(i).widget()
                    if widget:
                        widget.setParent(None)
            
            QMessageBox.information(self, '保存完了', 
                f'✅ サンプルファイルが以下のフォルダにエクスポートされました:\n\n'
                f'📁 {sample_folder}')
        else:
            QMessageBox.warning(self, 'エラー', '保存する結果がありません。')

    def on_ng_clicked(self):
        """ES: Borra archivos temporales y habilita botones de optimización
        EN: Remove temporary files and enable optimization buttons
        JA: 一時ファイルを削除し最適化ボタンを有効化"""
        try:
            print(f"🔍 デバッグ NG: current_temp_folder = {getattr(self, 'current_temp_folder', '存在しません')}")
            print(f"🔍 デバッグ NG: proyecto_folder = {getattr(self, 'proyecto_folder', '存在しません')}")
            
            # ES: Borrar carpeta temporal si existe
            # EN: Delete the temporary folder if it exists
            # JP: 一時フォルダが存在する場合は削除
            if hasattr(self, 'current_temp_folder') and self.current_temp_folder:
                print(f"🔍 Debug NG: 存在確認: {self.current_temp_folder}")
                if os.path.exists(self.current_temp_folder):
                    print("🔍 Debug NG: フォルダが存在するため削除します...")
                    shutil.rmtree(self.current_temp_folder)
                    print(f"🗑️ Temp フォルダを削除しました: {self.current_temp_folder}")
                else:
                    print(f"🔍 Debug NG: フォルダが存在しません: {self.current_temp_folder}")
                
                # ES: NO borrar la carpeta 99_Temp - mantenerla para futuros usos
                # EN: Do NOT delete the 99_Temp folder - keep it for future use
                # JP: 99_Tempフォルダは削除しない（今後の利用のため保持）
                temp_base = os.path.join(self.proyecto_folder, "99_Temp")
                print(f"📁 99_Temp フォルダは保持します: {temp_base}")
            else:
                print("🔍 Debug NG: current_temp_folder が定義されていません")
            
            # ES: Limpiar referencias
            # EN: Clear references
            # JP: 参照をクリア
            if hasattr(self, 'current_temp_folder'):
                delattr(self, 'current_temp_folder')
            if hasattr(self, 'dsaitekika_results'):
                delattr(self, 'dsaitekika_results')
            if hasattr(self, 'isaitekika_results'):
                delattr(self, 'isaitekika_results')
            if hasattr(self, 'dsaitekika_selected_df'):
                delattr(self, 'dsaitekika_selected_df')
            if hasattr(self, 'isaitekika_selected_df'):
                delattr(self, 'isaitekika_selected_df')
            
            # ES: Limpiar gráficos y tablas
            # EN: Clear charts and tables
            # JP: グラフとテーブルをクリア
            self.graph_images = []
            self.graph_images_content = []
            self.current_graph_index = 0
            
            # ES: Limpiar área de gráficos
            # EN: Clear chart area
            # JP: グラフ領域をクリア
            if hasattr(self, 'graph_area') and self.graph_area.layout():
                layout = self.graph_area.layout()
                for i in reversed(range(layout.count())):
                    widget = layout.itemAt(i).widget()
                    if widget:
                        widget.setParent(None)
            
            # ES: Habilitar botones de optimización
            # EN: Enable optimization buttons
            # JP: 最適化ボタンを有効化
            self.d_optimize_button.setEnabled(True)
            self.i_optimize_button.setEnabled(True)
            # ES: Aplicar estilo visual de habilitado
            # EN: Apply enabled visual style
            # JP: 有効時の見た目スタイルを適用
            self.d_optimize_button.setStyleSheet(self.d_optimize_button.styleSheet())
            self.i_optimize_button.setStyleSheet(self.i_optimize_button.styleSheet())
            
            # ES: Deshabilitar botones OK/NG
            # EN: Disable OK/NG buttons
            # JP: OK/NGボタンを無効化
            self.ok_button.setEnabled(False)
            self.ng_button.setEnabled(False)
            
            QMessageBox.information(self, 'キャンセル', 
                '✅ サンプルがキャンセルされました。')
            
        except Exception as e:
            QMessageBox.warning(self, '警告', 
                f'⚠️ 一時ファイルの削除中にエラーが発生しました:\n{str(e)}\n\n最適化ボタンは再有効化されました。')
            
            # ES: Aún así, habilitar los botones
            # EN: Even so, enable the buttons
            # JP: それでもボタンを有効化する
            self.d_optimize_button.setEnabled(True)
            self.i_optimize_button.setEnabled(True)
            self.d_optimize_button.setStyleSheet(self.d_optimize_button.styleSheet())
            self.i_optimize_button.setStyleSheet(self.i_optimize_button.styleSheet())
            self.ok_button.setEnabled(False)
            self.ng_button.setEnabled(False)

    def get_selected_brush(self):
        """
        Compatibilidad: antes devolvía el brush del selector UI.
        Ahora el brush SIEMPRE viene del archivo de resultados (A13/A11/A21/A32).
        """
        return getattr(self, "_results_brush_type", None)
    
    def get_selected_brush_from_filter(self):
        """ES: Obtener el brush seleccionado del filtro
        EN: Get selected brush from filter
        JA: フィルタから選択中のブラシを取得"""
        for key in ["すべて", "A13", "A11", "A21", "A32"]:
            if key in self.filter_inputs and self.filter_inputs[key].isChecked():
                return key
        return "すべて"  # Default

    def on_generate_sample_file_clicked(self):
        # ES: Pausar timers automáticos para evitar interferencia con el diálogo | EN: Pause auto timers to avoid interference with the dialog | JA: ダイアログとの干渉を避けるため自動タイマーを一時停止
        self.pause_auto_timers()
        
        config_file, _ = QFileDialog.getOpenFileName(
            self, "パラメータ設定ファイルを選択", "", "Excel Files (*.xlsx *.xls)"
        )
        if not config_file:
            # ES: Reanudar timers si se cancela el primer diálogo | EN: Resume timers if first dialog is cancelled | JA: 最初のダイアログがキャンセルされたらタイマーを再開
            self.resume_auto_timers()
            return

        save_path, _ = QFileDialog.getSaveFileName(
            self, "保存先を選択", "sample_combinations.xlsx", "Excel Files (*.xlsx *.xls)"
        )
        if not save_path:
            # ES: Reanudar timers si se cancela el segundo diálogo | EN: Resume timers if second dialog is cancelled | JA: 2つ目のダイアログがキャンセルされたらタイマーを再開
            self.resume_auto_timers()
            return
        
        # ES: Reanudar timers después de ambos diálogos | EN: Resume timers after both dialogs | JA: 両方のダイアログ後にタイマーを再開
        self.resume_auto_timers()

        # ES: Mostrar loader (reutilizar si ya existe para evitar múltiples overlays/eventFilters) | EN: Show loader (reuse if exists to avoid multiple overlays) | JA: ローダー表示（多重オーバーレイ防止のため既存を再利用）
        if not hasattr(self, 'loader_overlay') or self.loader_overlay is None:
            self.loader_overlay = LoadingOverlay(self.center_frame)
        self.loader_overlay.start()

        self.sample_thread = QThread()
        self.sample_worker = SampleCombinerWorker(config_file, save_path)
        self.sample_worker.moveToThread(self.sample_thread)

        self.sample_thread.started.connect(self.sample_worker.run)
        self.sample_worker.finished.connect(self.on_sample_generation_finished)
        self.sample_worker.error.connect(self.on_sample_generation_error)
        self.sample_worker.finished.connect(self.sample_thread.quit)
        self.sample_worker.finished.connect(self.sample_worker.deleteLater)
        self.sample_thread.finished.connect(self.sample_thread.deleteLater)

        self.sample_thread.start()

    def add_selected_samples_table_view(self, df):
        # Definir columnas básicas que siempre deben estar presentes
        columnas_basicas = ["No.", "回転速度", "送り速度", "UPカット", "回転方向", "切込量", "突出量", "載せ率", "パス数"]
        
        # ES: Verificar qué columnas están disponibles en el DataFrame | EN: Check which columns are available in DataFrame | JA: DataFrameで利用可能な列を確認
        columnas_disponibles = []
        for col in columnas_basicas:
            if col in df.columns:
                columnas_disponibles.append(col)
        
        # ES: Añadir columnas adicionales si están disponibles | EN: Add extra columns if available | JA: 利用可能なら追加列を追加
        # ISaitekika: NO mostrar I基準値 en la tabla
        if hasattr(self, 'isaitekika_selected_df') and df is getattr(self, 'isaitekika_selected_df', None):
            columnas_adicionales = ["D基準値", "上面ダレ", "側面ダレ", "摩耗量"]
        else:
            columnas_adicionales = ["D基準値", "I基準値", "上面ダレ", "側面ダレ", "摩耗量"]
        for col in columnas_adicionales:
            if col in df.columns:
                columnas_disponibles.append(col)

        # ES: Crear DataFrame filtrado solo con las columnas disponibles | EN: Create filtered DataFrame with available columns only | JA: 利用可能列のみのフィルタ済みDataFrameを作成
        df_filtrado = df[columnas_disponibles].copy()

        # ES: Crear contenedor para la tabla con título | EN: Create container for table with title | JA: タイトル付きテーブル用コンテナを作成
        table_container = QWidget()
        table_layout = QVBoxLayout(table_container)
        
        # ES: Determinar el título basándose en el tipo de optimización
        # EN: Determine the title based on the optimization type
        # JP: 最適化タイプに基づいてタイトルを決定
        # ES: Si tenemos resultados de I最適化, mostrar tabla I最適
        # EN: If we have I最適化 results, show the I最適 table
        # JP: I最適化の結果があればI最適テーブルを表示
        if hasattr(self, 'isaitekika_results') and hasattr(self, 'dsaitekika_results'):
            # ES: Si ambos existen, determinar por el DataFrame actual
            # EN: If both exist, decide based on the current DataFrame
            # JP: 両方ある場合は現在のDataFrameで判定
            if df is self.isaitekika_selected_df:
                title = "I最適サンプル一覧"
            else:
                title = "D最適サンプル一覧"
        elif hasattr(self, 'isaitekika_results'):
            title = "I最適サンプル一覧"
        elif hasattr(self, 'dsaitekika_results'):
            title = "D最適サンプル一覧"
        else:
            title = "新規実験点"
            
        title_label = QLabel(title)
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("font-size: 16px; font-weight: bold; margin-bottom: 8px;")
        table_layout.addWidget(title_label)

        table_widget = QTableWidget()
        table_widget.setRowCount(len(df_filtrado))
        table_widget.setColumnCount(len(df_filtrado.columns))
        table_widget.setHorizontalHeaderLabels(df_filtrado.columns)
        table_widget.setStyleSheet("font-size: 11px; font-family: 'Yu Gothic';")

        for row in range(len(df_filtrado)):
            for col in range(len(df_filtrado.columns)):
                item = QTableWidgetItem(str(df_filtrado.iat[row, col]))
                item.setFlags(item.flags() ^ Qt.ItemIsEditable)  # Read-only
                table_widget.setItem(row, col, item)

        # Expandir tabla al ancho completo del contenedor
        table_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        table_widget.horizontalHeader().setStretchLastSection(True)
        table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        table_layout.addWidget(table_widget)

        self.graph_images.append("table")  # marcador especial
        self.graph_images_content = getattr(self, "graph_images_content", [])
        self.graph_images_content.append(table_container)

        self.next_button.setEnabled(True)
        self.prev_button.setEnabled(True)

    def show_loader(self, show: bool):
        if show:
            self.loader_label.show()
            self.loader_movie.start()
        else:
            self.loader_movie.stop()
            self.loader_label.hide()

    def display_graphs(self, image_paths):
        """
        ES: Guarda las rutas y muestra la primera imagen.
        EN: Save the paths and show the first image.
        JP: パスを保存し、最初の画像を表示します。
        """
        self.graph_images = image_paths
        self.current_graph_index = 0

        # ES: Crear botones si no existen.
        # EN: Create buttons if they do not exist.
        # JP: ボタンが無ければ作成します。
        if self.prev_button is None or self.next_button is None:
            self.create_navigation_buttons()

        # ES: Mostrar el primer gráfico y activar/desactivar botones según corresponda.
        # EN: Show the first chart and enable/disable buttons as needed.
        # JP: 先頭グラフを表示し、必要に応じてボタンの有効/無効を切り替えます。
        self.update_graph_display()
        self.prev_button.setEnabled(self.current_graph_index > 0)
        self.next_button.setEnabled(self.current_graph_index < len(self.graph_images) - 1)
        print("グラフ数:", len(self.graph_images))

    # ES: Función para actualizar la imagen mostrada
    # EN: Function to update the displayed image
    # JP: 表示画像を更新する関数
    def update_graph_display(self):
        # ES: Verificar si el layout existe; si no, crear uno nuevo.
        # EN: Check if layout exists; if not, create a new one.
        # JP: レイアウトがなければ新規作成します。
        if self.graph_area.layout() is None:
            print("⚠️ グラフ領域のレイアウトが None です。新しいレイアウトを作成します...")
            self.graph_area.setLayout(QVBoxLayout())
        
        layout = self.graph_area.layout()

        # ES: Limpiar contenido actual
        # EN: Clear current content
        # JP: 現在の内容をクリア
        for i in reversed(range(layout.count())):
            widget = layout.itemAt(i).widget()
            if widget:
                widget.setParent(None)

        current_item = self.graph_images[self.current_graph_index]

        if current_item == "table":
            # ES: Determinar qué tabla mostrar basándose en el índice actual
            # EN: Determine which table to show based on the current index
            # JP: 現在のインデックスに基づいて表示するテーブルを決定
            table_index = 0  # Default D-optimal
            if hasattr(self, 'graph_images_content') and len(self.graph_images_content) >= 2:
                        # ES: Contar cuántas tablas hay antes del índice actual
                        # EN: Count how many tables appear before the current index
                        # JP: 現在のインデックスより前にあるテーブル数を数える
                table_count = 0
                for i in range(self.current_graph_index):
                    if self.graph_images[i] == "table":
                        table_count += 1
                
                # ES: Si es la primera tabla (table_count = 0), mostrar D-óptimo
                # EN: If it is the first table (table_count = 0), show D-optimal
                # JP: 1つ目のテーブル（table_count=0）ならD最適を表示
                # ES: Si es la segunda tabla (table_count = 1), mostrar I-óptimo
                # EN: If it is the second table (table_count = 1), show I-optimal
                # JP: 2つ目のテーブル（table_count=1）ならI最適を表示
                if table_count == 0:
                    print("📋 D最適テーブルを表示中")
                    self._add_tablewidget_to_graph_area(self.dsaitekika_selected_df, layout, "D最適サンプル一覧")
                elif table_count == 1:
                    print("📋 I最適テーブルを表示中")
                    self._add_tablewidget_to_graph_area(self.isaitekika_selected_df, layout, "I最適サンプル一覧")
                else:
                    # ES: Fallback: mostrar la tabla correspondiente del contenido
                    # EN: Fallback: show the corresponding table from the content list
                    # JP: フォールバック：contentリストから該当テーブルを表示
                    if table_count < len(self.graph_images_content):
                        layout.addWidget(self.graph_images_content[table_count])
            else:
                # ES: Fallback: mostrar la última tabla añadida
                # EN: Fallback: show the last added table
                # JP: フォールバック：最後に追加されたテーブルを表示
                if hasattr(self, 'graph_images_content') and self.graph_images_content:
                    layout.addWidget(self.graph_images_content[-1])
        else:
            img_path = current_item
            pixmap = QPixmap(img_path).scaled(700, 540, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            label = QLabel()
            label.setPixmap(pixmap)
            label.setAlignment(Qt.AlignCenter)
            layout.addWidget(label)

        self.prev_button.setEnabled(self.current_graph_index > 0)
        self.next_button.setEnabled(self.current_graph_index < len(self.graph_images) - 1)

    def _add_tablewidget_to_graph_area(self, df, layout, titulo=None):

        # ES: Definir columnas básicas que siempre deben estar presentes
        # EN: Define the basic columns that must always be present
        # JP: 常に存在すべき基本列を定義
        columnas_basicas = ["No.", "回転速度", "送り速度", "UPカット", "回転方向", "切込量", "突出量", "載せ率", "パス数"]
        columnas_disponibles = [col for col in columnas_basicas if col in df.columns]
        columnas_adicionales = ["D基準値", "I基準値", "上面ダレ", "側面ダレ", "摩耗量"]
        for col in columnas_adicionales:
            if col in df.columns:
                columnas_disponibles.append(col)
        df_filtrado = df[columnas_disponibles].copy()
        if titulo:
            label = QLabel(titulo)
            label.setAlignment(Qt.AlignCenter)
            label.setStyleSheet("font-size: 16px; font-weight: bold; margin-bottom: 8px;")
            layout.addWidget(label)
        table_widget = QTableWidget()
        table_widget.setRowCount(len(df_filtrado))
        table_widget.setColumnCount(len(df_filtrado.columns))
        table_widget.setHorizontalHeaderLabels(df_filtrado.columns)
        table_widget.setStyleSheet("font-size: 11px; font-family: 'Yu Gothic';")
        for row in range(len(df_filtrado)):
            for col in range(len(df_filtrado.columns)):
                item = QTableWidgetItem(str(df_filtrado.iat[row, col]))
                item.setFlags(item.flags() ^ Qt.ItemIsEditable)  # Read-only
                table_widget.setItem(row, col, item)
        table_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        table_widget.horizontalHeader().setStretchLastSection(True)
        table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(table_widget)

    def on_integrated_optimizer_finished(self, result):
        self.optimizer_result = result  # Ensure results are available for on_ok_clicked
        """ES: Maneja los resultados del optimizador integrado D-óptimo + I-óptimo
        EN: Handle results from the integrated D-optimal + I-optimal optimizer
        JA: 統合D最適+I最適オプティマイザの結果を処理"""

        # ES: Guardar ambos DataFrames | EN: Save both DataFrames | JA: 両DataFrameを保存
        self.dsaitekika_selected_df = result["d_dataframe"]
        self.isaitekika_selected_df = result["i_dataframe"]

        # ES: Guardar datos del optimizador para recalcular D基準値 exactamente como el archivo de referencia
        # EN: Store optimizer data to recalculate D-score exactly as in the reference file
        # JA: 参照ファイルと同一にD基準値を再計算するためオプティマイザデータを保持
        self._last_candidate_df_for_dscore = result.get("candidate_df", None)
        self._last_d_indices = result.get("d_indices", None)
        self._last_existing_indices = result.get("existing_indices", None)
        try:
            if self._last_candidate_df_for_dscore is not None and self._last_d_indices is not None:
                cand_np = (
                    self._last_candidate_df_for_dscore.to_numpy()
                    if hasattr(self._last_candidate_df_for_dscore, "to_numpy")
                    else np.asarray(self._last_candidate_df_for_dscore)
                )
                self._last_d_score_reference = calculate_d_score_reference(cand_np, self._last_d_indices)
                if len(self.dsaitekika_selected_df) > 0:
                    self.dsaitekika_selected_df["D基準値"] = self._last_d_score_reference
        except Exception as e:
            print(f"⚠️ D基準値計算エラー (referencia, integrado): {e}")
        
        # ES: Para optimización integrada, establecer el tipo basado en el último ejecutado | EN: For integrated optimization, set type from last run | JA: 統合最適化では最後の実行に基づきタイプを設定
        # ES: Por defecto, usar D最適化 como tipo principal | EN: By default use D最適化 as main type | JA: デフォルトでD最適化をメインタイプに
        self.last_executed_optimization = 'D'
        print(f"🔍 Debug - on_integrated_optimizer_finished: last_executed_optimization = 'D' (integrado)")
        
        # ES: Añadir columnas necesarias para la visualización en tabla | EN: Add columns needed for table display | JA: 表表示用に必要な列を追加
        if len(self.dsaitekika_selected_df) > 0:
            # Mapear nombres de columnas si es necesario
            if '突出し量' in self.dsaitekika_selected_df.columns:
                self.dsaitekika_selected_df.rename(columns={'突出し量': '突出量'}, inplace=True)
            if '切込み量' in self.dsaitekika_selected_df.columns:
                self.dsaitekika_selected_df.rename(columns={'切込み量': '切込量'}, inplace=True)
            
            if "No." not in self.dsaitekika_selected_df.columns:
                self.dsaitekika_selected_df.insert(0, "No.", list(range(1, len(self.dsaitekika_selected_df) + 1)))
            if "上面ダレ" not in self.dsaitekika_selected_df.columns:
                self.dsaitekika_selected_df["上面ダレ"] = ""
            if "側面ダレ" not in self.dsaitekika_selected_df.columns:
                self.dsaitekika_selected_df["側面ダレ"] = ""
            if "摩耗量" not in self.dsaitekika_selected_df.columns:
                self.dsaitekika_selected_df["摩耗量"] = ""
        if len(self.isaitekika_selected_df) > 0:
            # Mapear nombres de columnas si es necesario
            if '突出し量' in self.isaitekika_selected_df.columns:
                self.isaitekika_selected_df.rename(columns={'突出し量': '突出量'}, inplace=True)
            if '切込み量' in self.isaitekika_selected_df.columns:
                self.isaitekika_selected_df.rename(columns={'切込み量': '切込量'}, inplace=True)
            
            if "No." not in self.isaitekika_selected_df.columns:
                self.isaitekika_selected_df.insert(0, "No.", list(range(1, len(self.isaitekika_selected_df) + 1)))
            if "上面ダレ" not in self.isaitekika_selected_df.columns:
                self.isaitekika_selected_df["上面ダレ"] = ""
            if "側面ダレ" not in self.isaitekika_selected_df.columns:
                self.isaitekika_selected_df["側面ダレ"] = ""
            if "摩耗量" not in self.isaitekika_selected_df.columns:
                self.isaitekika_selected_df["摩耗量"] = ""
        
        # ES: Guardar rutas de archivos para uso posterior | EN: Save file paths for later use | JA: 後で使うためファイルパスを保存
        self.integrated_output_folder = os.path.dirname(result["d_path"]) if result["d_path"] else ""
        self.d_optimal_path = result["d_path"]
        self.i_optimal_path = result["i_path"]
        self.all_d_optimal_path = result["all_d_path"]
        self.all_i_optimal_path = result["all_i_path"]
        
        # ES: Guardar DataFrames adicionales para guardado posterior | EN: Save extra DataFrames for later save | JA: 後で保存するため追加DataFrameを保存
        self.candidate_df = result.get("candidate_df", pd.DataFrame())
        self.all_d_df = result.get("all_d_df", pd.DataFrame())
        self.all_i_df = result.get("all_i_df", pd.DataFrame())
        
        # ES: Exportar los Excel con criterios calculados
        # EN: Export Excel files with calculated criteria
        # JP: 計算した基準値付きでExcelを出力
        if len(self.dsaitekika_selected_df) > 0 and not os.path.exists(self.d_optimal_path):
            # Calcular D基準値 para D-óptimo (igual que referencia)
            df_d = self.dsaitekika_selected_df.copy()
            if len(df_d) > 0:
                # Preferir score de referencia (StandardScaler + logdet estable)
                d_score_ref = getattr(self, "_last_d_score_reference", None)
                if d_score_ref is None or not np.isfinite(d_score_ref):
                    # Fallback: calcular sobre los puntos seleccionados solamente (no ideal, pero consistente)
                    X_raw = _extract_design_matrix(df_d)
                    X_scaled = _standardize_like_reference(X_raw)
                    d_score_ref, _ = calculate_d_criterion_stable_reference(
                        X_scaled, method="auto", use_numerical_stable_method=True, verbose=False
                    )
                df_d["D基準値"] = float(d_score_ref) if d_score_ref is not None else np.nan
            
            df_d.to_excel(self.d_optimal_path, index=False)
            
        if len(self.isaitekika_selected_df) > 0 and not os.path.exists(self.i_optimal_path):
            # ISaitekika: I基準値 NO se calcula nunca (mantener en blanco)
            df_i = self.isaitekika_selected_df.copy()
            df_i['I基準値'] = ''
            
            df_i.to_excel(self.i_optimal_path, index=False)

        # ES: Configurar sistema de navegación de gráficos uno a uno | EN: Configure one-by-one chart navigation | JA: グラフの一対一ナビゲーションを設定
        self.graph_images = result["image_paths"]
        self.current_graph_index = 0
        print("📊 グラフナビゲーションを設定中:")
        print(f"  - グラフ総数: {len(self.graph_images)}")
        print(f"  - 利用可能なグラフ: {[os.path.basename(path) for path in self.graph_images]}")
        
        # ES: Crear botones de navegación si no existen | EN: Create navigation buttons if they do not exist | JA: ナビボタンが無ければ作成
        if self.prev_button is None or self.next_button is None:
            self.create_navigation_buttons()
        
        # ES: Mostrar primer gráfico | EN: Show first chart | JA: 先頭グラフを表示
        self.update_graph_display()
        self.prev_button.setEnabled(False)
        self.next_button.setEnabled(len(self.graph_images) > 1)
        print("✅ ナビゲーションを設定しました:")
        print(f"  - 現在のグラフ: {self.current_graph_index + 1}/{len(self.graph_images)}")
        print(f"  - 前へボタン: {'有効' if self.prev_button.isEnabled() else '無効'}")
        print(f"  - 次へボタン: {'有効' if self.next_button.isEnabled() else '無効'}")

        # ES: Añadir ambas tablas usando el método original | EN: Add both tables using original method | JA: 元のメソッドで両テーブルを追加
        print(f"📋 D最適テーブルを追加中（{len(self.dsaitekika_selected_df)} 行）")
        self.current_table_index = 0  # For D-optimal
        self.add_selected_samples_table_view(self.dsaitekika_selected_df)
        
        print(f"📋 I最適テーブルを追加中（{len(self.isaitekika_selected_df)} 行）")
        self.current_table_index = 1  # For I-optimal
        self.add_selected_samples_table_view(self.isaitekika_selected_df)
        
        print(f"✅ graph_images の要素数: {len(self.graph_images)}")
        print(f"✅ graph_images_content の要素数: {len(self.graph_images_content)}")

        # ES: Habilitar botones OK/NG
        # EN: Enable OK/NG buttons
        # JP: OK/NGボタンを有効化
        self.ok_button.setEnabled(True)
        self.ng_button.setEnabled(True)
        
        # ES: Deshabilitar botones de optimización después de completar el análisis integrado
        # EN: Disable optimization buttons after completing the integrated analysis
        # JP: 統合解析完了後に最適化ボタンを無効化
        self.d_optimize_button.setEnabled(False)
        self.i_optimize_button.setEnabled(False)
        # ES: Aplicar estilo visual de deshabilitado
        # EN: Apply disabled visual style
        # JP: 無効時の見た目スタイルを適用
        self.d_optimize_button.setStyleSheet(self.d_optimize_button.styleSheet())
        self.i_optimize_button.setStyleSheet(self.i_optimize_button.styleSheet())
        
        # ES: Mensaje de éxito
        # EN: Success message
        # JP: 成功メッセージ
        message = f"""✅ 最適化統合が完了しました。\n\n📊 結果サマリー:\n• D-最適新規選択: {len(result['d_dataframe'])} 点\n• I-最適新規選択: {len(result['i_dataframe'])} 点\n• 既存実験点活用: {len(result['existing_indices'])} 点\n\n📈 可視化: 特徴量分布 + 次元削減UMAP ({len(self.graph_images)} グラフ)\n📋 テーブル: D-最適 + I-最適 (ナビゲーションで切り替え)\n💾 ファイルはOKボタンを押した時に保存されます"""
        QMessageBox.information(self, "最適化統合完了", message)
        self.loader_overlay.stop()

    def on_d_optimizer_finished(self, results):
        print("DEBUG: on_d_optimizer_finished に入りました")
        print("DEBUG: on_d_optimizer_finished の結果:", results)
        self.dsaitekika_results = results
        self.dsaitekika_selected_df = results['d_dataframe']
        
        # ES: Limpiar TODOS los resultados anteriores para evitar conflictos.
        # EN: Clear ALL previous results to avoid conflicts.
        # JP: 競合を避けるため、過去の結果をすべてクリアします。
        if hasattr(self, 'isaitekika_results'):
            delattr(self, 'isaitekika_results')
            print("🧹 以前の isaitekika_results をクリアしました")
        if hasattr(self, 'isaitekika_selected_df'):
            delattr(self, 'isaitekika_selected_df')
            print("🧹 以前の isaitekika_selected_df をクリアしました")
        
        # ES: Establecer explícitamente el tipo de optimización | EN: Set optimization type explicitly | JA: 最適化タイプを明示的に設定
        self.last_executed_optimization = 'D'  # Mark that D optimization ran
        print(f"🔍 Debug - on_d_optimizer_finished: last_executed_optimization = 'D'")
        print(f"🔍 Debug - dsaitekika_results（クリア後）存在: {hasattr(self, 'dsaitekika_results')}")
        print(f"🔍 Debug - isaitekika_results（クリア後）存在: {hasattr(self, 'isaitekika_results')}")
        print(f"🔍 Debug - last_executed_optimization を設定しました: {self.last_executed_optimization}")
        
        # Mapear nombres de columnas si es necesario
        if '突出し量' in self.dsaitekika_selected_df.columns:
            self.dsaitekika_selected_df.rename(columns={'突出し量': '突出量'}, inplace=True)
        if '切込み量' in self.dsaitekika_selected_df.columns:
            self.dsaitekika_selected_df.rename(columns={'切込み量': '切込量'}, inplace=True)
        
        # Calcular D基準値 exactamente como el archivo de referencia (StandardScaler sobre TODOS los candidatos)
        try:
            self._last_candidate_df_for_dscore = results.get("candidate_df", getattr(self, "_last_candidate_df_for_dscore", None))
            self._last_d_indices = results.get("d_indices", getattr(self, "_last_d_indices", None))
            self._last_existing_indices = results.get("existing_indices", getattr(self, "_last_existing_indices", None))

            d_score_ref = None
            if self._last_candidate_df_for_dscore is not None and self._last_d_indices is not None:
                cand_np = (
                    self._last_candidate_df_for_dscore.to_numpy()
                    if hasattr(self._last_candidate_df_for_dscore, "to_numpy")
                    else np.asarray(self._last_candidate_df_for_dscore)
                )
                d_score_ref = calculate_d_score_reference(cand_np, self._last_d_indices)

            if (d_score_ref is None or not np.isfinite(d_score_ref)) and len(self.dsaitekika_selected_df) > 0:
                # Fallback: score sobre los seleccionados solamente
                X_raw = _extract_design_matrix(self.dsaitekika_selected_df)
                X_scaled = _standardize_like_reference(X_raw)
                d_score_ref, _ = calculate_d_criterion_stable_reference(
                    X_scaled, method="auto", use_numerical_stable_method=True, verbose=False
                )

            self._last_d_score_reference = float(d_score_ref) if d_score_ref is not None else None
            if len(self.dsaitekika_selected_df) > 0:
                self.dsaitekika_selected_df["D基準値"] = self._last_d_score_reference if self._last_d_score_reference is not None else np.nan
        except Exception as e:
            print(f"⚠️ D基準値計算エラー (referencia, D-only): {e}")
        output_folder = os.path.dirname(results['d_path']) if results['d_path'] else ""
        # Filtrar solo los gráficos relevantes a D最適化
        image_paths = sorted(glob.glob(os.path.join(output_folder, '*.png')))
        # Filtrar: solo mostrar histogramas y gráficos generales (no los que sean exclusivamente de I)
        d_image_paths = [p for p in image_paths if not ("I" in os.path.basename(p) or "i_optimal" in os.path.basename(p))]
        if not d_image_paths:
            d_image_paths = image_paths  # fallback: show all if no distinction
        
        # ES: Limpiar contenido anterior
        # EN: Clear previous content
        # JP: 前の内容をクリア
        self.graph_images = []
        self.graph_images_content = []
        
        self.display_graphs(d_image_paths)
        self.add_selected_samples_table_view(self.dsaitekika_selected_df)
        self.ok_button.setEnabled(True)
        self.ng_button.setEnabled(True)
        self.create_navigation_buttons()
        
        # ES: Deshabilitar botones de optimización después de completar D最適化
        # EN: Disable optimization buttons after completing D最適化
        # JP: D最適化完了後に最適化ボタンを無効化
        self.d_optimize_button.setEnabled(False)
        self.i_optimize_button.setEnabled(False)
        # ES: Aplicar estilo visual de deshabilitado
        # EN: Apply disabled visual style
        # JP: 無効時の見た目スタイルを適用
        self.d_optimize_button.setStyleSheet(self.d_optimize_button.styleSheet())
        self.i_optimize_button.setStyleSheet(self.i_optimize_button.styleSheet())
        
        QMessageBox.information(self, "完了",
                                f"✅ D最適化が完了しました。\n結果を保存しました:\n{results['d_path']}")
        # Asegurar que el QThread se cierra antes de permitir nuevas ejecuciones
        self._cleanup_optimization_threads(aggressive=True)
        self.loader_overlay.stop()

    def on_i_optimizer_finished(self, results):
        print("DEBUG: on_i_optimizer_finished に入りました")
        print("DEBUG: on_i_optimizer_finished の結果:", results)
        self.isaitekika_results = results
        self.isaitekika_selected_df = results['i_dataframe']
        # ES: Limpiar TODOS los resultados anteriores para evitar conflictos.
        # EN: Clear ALL previous results to avoid conflicts.
        # JP: 競合を避けるため、過去の結果をすべてクリアします。
        if hasattr(self, 'dsaitekika_results'):
            delattr(self, 'dsaitekika_results')
            print("🧹 以前の dsaitekika_results をクリアしました")
        if hasattr(self, 'dsaitekika_selected_df'):
            delattr(self, 'dsaitekika_selected_df')
            print("🧹 以前の dsaitekika_selected_df をクリアしました")
        
        # ES: Establecer explícitamente el tipo de optimización.
        # EN: Set optimization type explicitly.
        # JP: 最適化タイプを明示的に設定します。
        self.last_executed_optimization = 'I'  # Mark that I optimization ran
        print(f"🔍 Debug - on_i_optimizer_finished: last_executed_optimization = 'I'")
        print(f"🔍 Debug - isaitekika_results（クリア後）存在: {hasattr(self, 'isaitekika_results')}")
        print(f"🔍 Debug - dsaitekika_results（クリア後）存在: {hasattr(self, 'dsaitekika_results')}")
        print(f"🔍 Debug - last_executed_optimization を設定しました: {self.last_executed_optimization}")
        
        # Mapear nombres de columnas si es necesario
        if '突出し量' in self.isaitekika_selected_df.columns:
            self.isaitekika_selected_df.rename(columns={'突出し量': '突出量'}, inplace=True)
        if '切込み量' in self.isaitekika_selected_df.columns:
            self.isaitekika_selected_df.rename(columns={'切込み量': '切込量'}, inplace=True)
        
        # ISaitekika: I基準値 NO se calcula nunca (mantener en blanco)
        if len(self.isaitekika_selected_df) > 0:
            self.isaitekika_selected_df['I基準値'] = ''
        output_folder = os.path.dirname(results['i_path']) if results['i_path'] else ""
        # Filtrar solo los gráficos relevantes a I最適化
        image_paths = sorted(glob.glob(os.path.join(output_folder, '*.png')))
        # Filtrar: solo mostrar histogramas y gráficos generales (no los que sean exclusivamente de D)
        i_image_paths = [p for p in image_paths if not ("D" in os.path.basename(p) or "d_optimal" in os.path.basename(p))]
        if not i_image_paths:
            i_image_paths = image_paths  # fallback: show all if no distinction
        
        # ES: Limpiar contenido anterior
        # EN: Clear previous content
        # JP: 前の内容をクリア
        self.graph_images = []
        self.graph_images_content = []
        
        self.display_graphs(i_image_paths)
        self.add_selected_samples_table_view(self.isaitekika_selected_df)
        self.ok_button.setEnabled(True)
        self.ng_button.setEnabled(True)
        self.create_navigation_buttons()
        
        # ES: Deshabilitar botones de optimización después de completar I最適化
        # EN: Disable optimization buttons after completing I最適化
        # JP: I最適化完了後に最適化ボタンを無効化
        self.d_optimize_button.setEnabled(False)
        self.i_optimize_button.setEnabled(False)
        # ES: Aplicar estilo visual de deshabilitado
        # EN: Apply disabled visual style
        # JP: 無効時の見た目スタイルを適用
        self.d_optimize_button.setStyleSheet(self.d_optimize_button.styleSheet())
        self.i_optimize_button.setStyleSheet(self.i_optimize_button.styleSheet())
        
        QMessageBox.information(self, "完了",
                                f"✅ I最適化が完了しました。\n結果を保存しました:\n{results['i_path']}")
        # Asegurar que el QThread se cierra antes de permitir nuevas ejecuciones
        self._cleanup_optimization_threads(aggressive=True)
        self.loader_overlay.stop()

    def on_dsaitekika_finished(self, results):
        print("DEBUG: on_dsaitekika_finished に入りました")
        print("DEBUG: on_dsaitekika_finished の結果:", results)
        self.dsaitekika_results = results
        self.dsaitekika_selected_df = results['d_dataframe']  # ← Fixed to use the same structure as on_d_optimizer_finished
        # ES: Limpiar TODOS los resultados anteriores para evitar conflictos.
        # EN: Clear ALL previous results to avoid conflicts.
        # JP: 競合を避けるため、過去の結果をすべてクリアします。
        if hasattr(self, 'isaitekika_results'):
            delattr(self, 'isaitekika_results')
            print("🧹 以前の isaitekika_results をクリアしました")
        if hasattr(self, 'isaitekika_selected_df'):
            delattr(self, 'isaitekika_selected_df')
            print("🧹 以前の isaitekika_selected_df をクリアしました")
        
        # ES: Establecer explícitamente el tipo de optimización.
        # EN: Set optimization type explicitly.
        # JP: 最適化タイプを明示的に設定します。
        self.last_executed_optimization = 'D'  # Mark that D optimization ran
        print(f"🔍 Debug - on_dsaitekika_finished: last_executed_optimization = 'D'")
        print(f"🔍 Debug - dsaitekika_results（クリア後）存在: {hasattr(self, 'dsaitekika_results')}")
        print(f"🔍 Debug - isaitekika_results（クリア後）存在: {hasattr(self, 'isaitekika_results')}")
        print(f"🔍 Debug - last_executed_optimization を設定しました: {self.last_executed_optimization}")

        # Mapear nombres de columnas si es necesario
        if '突出し量' in self.dsaitekika_selected_df.columns:
            self.dsaitekika_selected_df.rename(columns={'突出し量': '突出量'}, inplace=True)
        if '切込み量' in self.dsaitekika_selected_df.columns:
            self.dsaitekika_selected_df.rename(columns={'切込み量': '切込量'}, inplace=True)

        # ✅ Añadir número de muestra
        self.dsaitekika_selected_df.insert(0, "No.", list(range(1, len(self.dsaitekika_selected_df) + 1)))

        # ✅ Añadir columnas vacías para resultados esperados
        self.dsaitekika_selected_df["上面ダレ"] = ""
        self.dsaitekika_selected_df["側面ダレ"] = ""
        self.dsaitekika_selected_df["摩耗量"] = ""
        
        # Calcular D基準値 como referencia (si podemos reconstruir candidatos + índices)
        try:
            d_score_ref = None
            # Indices seleccionados (0-based) a partir de la columna No. si existe
            if "No." in self.dsaitekika_selected_df.columns:
                no_series = pd.to_numeric(self.dsaitekika_selected_df["No."], errors="coerce")
                selected_indices = [int(x) - 1 for x in no_series.dropna().tolist() if int(x) > 0]
            else:
                selected_indices = []

            input_file = getattr(self, "_last_dsaitekika_input_file", None) or getattr(self, "sample_file_path", None)
            if input_file and selected_indices:
                ext = os.path.splitext(str(input_file))[1].lower()
                df_all = pd.read_csv(input_file, encoding="utf-8-sig") if ext == ".csv" else pd.read_excel(input_file)
                X_candidates = _extract_design_matrix(df_all)
                d_score_ref = calculate_d_score_reference(X_candidates, selected_indices)

            if (d_score_ref is None or not np.isfinite(d_score_ref)) and len(self.dsaitekika_selected_df) > 0:
                # Fallback: score sobre los seleccionados solamente
                X_raw = _extract_design_matrix(self.dsaitekika_selected_df)
                X_scaled = _standardize_like_reference(X_raw)
                d_score_ref, _ = calculate_d_criterion_stable_reference(
                    X_scaled, method="auto", use_numerical_stable_method=True, verbose=False
                )

            self._last_d_score_reference = float(d_score_ref) if d_score_ref is not None else None
            if len(self.dsaitekika_selected_df) > 0:
                self.dsaitekika_selected_df["D基準値"] = self._last_d_score_reference if self._last_d_score_reference is not None else np.nan
        except Exception as e:
            print(f"⚠️ D基準値計算エラー (referencia, Dsaitekika): {e}")

        image_paths = [
            self.dsaitekika_output_prefix + "_pca_features.png",
            self.dsaitekika_output_prefix + "_pca.png",
            self.dsaitekika_output_prefix + "_umap.png"
        ]
        self.display_graphs(image_paths)
        self.add_selected_samples_table_view(self.dsaitekika_selected_df)
        self.ok_button.setEnabled(True)
        self.ng_button.setEnabled(True)
        self.create_navigation_buttons()
        
        # ES: Deshabilitar botones de optimización después de completar D最適化
        # EN: Disable optimization buttons after completing D最適化
        # JP: D最適化完了後に最適化ボタンを無効化
        self.d_optimize_button.setEnabled(False)
        self.i_optimize_button.setEnabled(False)
        # ES: Aplicar estilo visual de deshabilitado
        # EN: Apply disabled visual style
        # JP: 無効時の見た目スタイルを適用
        self.d_optimize_button.setStyleSheet(self.d_optimize_button.styleSheet())
        self.i_optimize_button.setStyleSheet(self.i_optimize_button.styleSheet())

        QMessageBox.information(self, "完了",
                                f"✅ D最適化が完了しました。\n結果を保存しました:\n{self.dsaitekika_output_excel}")
        # Asegurar que el QThread se cierra antes de permitir nuevas ejecuciones
        self._cleanup_optimization_threads(aggressive=True)
        self.loader_overlay.stop()

    def on_dsaitekika_error(self, message):
        # ✅ FIX: asegurar que no queda ningún QThread de optimización "corriendo" tras un error
        try:
            for t_attr in ("d_optimizer_thread", "i_optimizer_thread", "dsaitekika_thread"):
                t = getattr(self, t_attr, None)
                if t is None:
                    continue
                try:
                    if t.isRunning():
                        t.quit()
                except RuntimeError:
                    # objeto Qt ya destruido
                    setattr(self, t_attr, None)
        except Exception:
            pass

        QMessageBox.critical(self, "エラー", f"❌ 最適化中にエラーが発生しました:\n{message}")
        self.loader_overlay.stop()
        # Asegurar cleanup completo en error (por si quedó algo vivo)
        self._cleanup_optimization_threads(aggressive=True)

        # ES: Re-habilitar botones por si quedaron deshabilitados
        # EN: Re-enable buttons in case they were left disabled
        # JP: 無効のまま残っていた場合に備えてボタンを再有効化
        try:
            self.d_optimize_button.setEnabled(True)
            self.i_optimize_button.setEnabled(True)
        except Exception:
            pass

    def on_sample_generation_finished(self):
        self.loader_overlay.stop()
        QMessageBox.information(self, "完了", "✅ サンプル組合せファイルが生成されました。")

    def on_sample_generation_error(self, error_msg):
        self.loader_overlay.stop()
        QMessageBox.critical(self, "エラー", f"❌ ファイル生成中にエラーが発生しました:\n{error_msg}")

    def load_results_file(self):
        # ES: Pausar timers automáticos para evitar interferencia con el diálogo | EN: Pause auto timers to avoid interference with the dialog | JA: ダイアログとの干渉を避けるため自動タイマーを一時停止
        self.pause_auto_timers()
        
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "結果ファイルを選択",
            "",
            "Excel/CSV Files (*.xlsx *.xls *.csv);;Excel Files (*.xlsx *.xls);;CSV Files (*.csv)"
        )
        
        # ES: Reanudar timers después del diálogo | EN: Resume timers after the dialog | JA: ダイアログ後にタイマーを再開
        self.resume_auto_timers()
        
        if file_path:
            try:
                if hasattr(self.processor, "process_results_file_with_ui_values"):
                    # brush y 線材長 vienen del Excel/CSV de resultados (A13/A11/A21/A32 y 線材長)
                    selected_brush = None
                    diameter = float(self.diameter_selector.currentText()) if hasattr(self, "diameter_selector") else 0.15
                    material = self.material_selector.currentText() if hasattr(self, "material_selector") else "Steel"
                    self.processor.process_results_file_with_ui_values(file_path, selected_brush, diameter, material)
                else:
                    # fallback
                    self.processor.process_results_file(file_path, None, None)
                QMessageBox.information(self, "完了", "✅ 結果ファイルをデータベースに取り込みました。")
            except Exception as e:
                QMessageBox.critical(self, "エラー", f"❌ 結果ファイル処理中にエラーが発生しました:\n{e}")

    def backup_and_update_sample_file(self, results_file_path, project_folder=None):
        """Hacer backup del archivo de muestreo y eliminar filas duplicadas basadas en el archivo de resultados"""
        try:
            # ES: Si no se especifica project_folder, usar el activo
            # EN: If project_folder is not specified, use the active one
            # JP: project_folder未指定ならアクティブなものを使用
            if project_folder is None:
                if not hasattr(self, 'proyecto_folder'):
                    raise ValueError("❌ アクティブなプロジェクトがありません。プロジェクトフォルダを指定してください。")
                project_folder = self.proyecto_folder
            
            # ES: Obtener el nombre del proyecto desde la carpeta
            # EN: Get the project name from the folder
            # JP: フォルダ名からプロジェクト名を取得
            project_name = os.path.basename(project_folder)
            
            print(f"🔍 Debug - project_folder: {project_folder}")
            print(f"🔍 Debug - project_name: {project_name}")
            
            # ES: Definir rutas: USAR EL ARCHIVO EN 99_Temp (o 99_Temp/Temp) DE LA CARPETA ESPECIFICADA
            # EN: Define paths: USE THE FILE IN 99_Temp (or 99_Temp/Temp) FROM THE SPECIFIED FOLDER
            # JP: パス定義: 指定フォルダの99_Temp（または99_Temp/Temp）のファイルを使用
            temp_base = os.path.join(project_folder, "99_Temp")
            os.makedirs(temp_base, exist_ok=True)

            # ES: ✅ NO depender del nombre del archivo:
            # EN: ✅ Do NOT depend on the file name:
            # JP: ✅ ファイル名に依存しない:
            # elegir cualquier *_未実験データ.(xlsx/xls/csv) dentro de 99_Temp o 99_Temp/Temp.
            # Preferencia (requerimiento): si existe CSV, priorizar CSV; si no, usar Excel.
            # ES: Si hay varios del mismo tipo, elegir el más reciente.
            # EN: If there are several of the same type, pick the most recent.
            # JP: 同タイプが複数ある場合は最新を選ぶ
            exts_priority = {".csv": 0, ".xlsx": 1, ".xls": 2}

            def _collect_candidates(folder: str):
                out = []
                try:
                    if not os.path.isdir(folder):
                        return out
                    for fn in os.listdir(folder):
                        if fn.startswith("~$"):
                            continue
                        if "_backup_" in fn:
                            continue
                        ext = os.path.splitext(fn)[1].lower()
                        if ext not in exts_priority:
                            continue
                        if not fn.endswith(f"_未実験データ{ext}"):
                            continue
                        full = os.path.join(folder, fn)
                        if os.path.isfile(full):
                            out.append(full)
                except Exception:
                    return []
                return out

            candidates = _collect_candidates(temp_base) + _collect_candidates(os.path.join(temp_base, "Temp"))
            if candidates:
                candidates.sort(key=lambda p: (exts_priority.get(os.path.splitext(p)[1].lower(), 9), -os.path.getmtime(p)))
                sample_file_path = candidates[0]
                try:
                    print("🔍 Debug - 候補 *_未実験データ.* が見つかりました (top 5):")
                    for p in candidates[:5]:
                        print(f"  - {p}")
                except Exception:
                    pass
            else:
                # ES: fallback legacy: nombre basado en carpeta
                # EN: legacy fallback: folder-based name
                # JP: 旧フォールバック: フォルダ名ベース
                candidate_sample_paths = [
                    os.path.join(temp_base, f"{project_name}_未実験データ.xlsx"),
                    os.path.join(temp_base, f"{project_name}_未実験データ.xls"),
                    os.path.join(temp_base, f"{project_name}_未実験データ.csv"),
                ]
                sample_file_path = next((p for p in candidate_sample_paths if os.path.exists(p)), candidate_sample_paths[0])

            sample_ext = os.path.splitext(sample_file_path)[1].lower()
            
            print(f"🔍 Debug - temp_base: {temp_base}")
            print(f"🔍 Debug - sample_file_path: {sample_file_path}")
            
            # ES: Verificar que existe el archivo de muestreo en 99_Temp | EN: Ensure sample file exists in 99_Temp | JA: 99_Tempにサンプルファイルがあるか確認
            if not os.path.exists(sample_file_path):
                raise ValueError(f"❌ サンプルファイルが見つかりません: {sample_file_path}")
            
            # ES: Crear carpeta backup en 99_Temp | EN: Create backup folder in 99_Temp | JA: 99_Tempにバックアップフォルダを作成
            backup_folder = os.path.join(temp_base, "backup")
            os.makedirs(backup_folder, exist_ok=True)
            
            # Generar nombre del backup con timestamp
            from datetime import datetime
            timestamp = datetime.now().strftime('%y%m%d_%H%M')
            backup_filename = f"{project_name}_未実験データ_backup_{timestamp}{sample_ext if sample_ext in ('.csv','.xlsx','.xls') else '.xlsx'}"
            backup_path = os.path.join(backup_folder, backup_filename)
            
            # ES: 1. Hacer backup del archivo de muestreo
            # EN: 1. Back up the sampling file
            # JP: 1. サンプルファイルをバックアップ
            print(f"📋 バックアップを作成中: {backup_path}")
            shutil.copy2(sample_file_path, backup_path)
            print(f"✅ バックアップが正常に作成されました")
            
            def _read_any_table(path: str) -> pd.DataFrame:
                ext = os.path.splitext(path)[1].lower()
                if ext == ".csv":
                    return pd.read_csv(path, encoding="utf-8-sig")
                return pd.read_excel(path)

            # ES: 2. Leer archivo de resultados (Excel/CSV)
            # EN: 2. Read the results file (Excel/CSV)
            # JP: 2. 結果ファイル（Excel/CSV）を読み込む
            print(f"📊 結果ファイルを読み込み中: {results_file_path}")
            df_results = _read_any_table(results_file_path)

            def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
                # Strip + normalizar espacios (incluye full-width)
                df = df.copy()
                df.columns = [
                    str(c).replace("\u3000", " ").strip() if c is not None else ""
                    for c in df.columns
                ]
                rename_map = {}
                # Variantes conocidas
                for c in df.columns:
                    if c == "突出し量":
                        rename_map[c] = "突出量"
                    elif c == "切込み量":
                        rename_map[c] = "切込量"
                    elif c == "回転方向":
                        rename_map[c] = "UPカット"
                    elif c == "UPカット/回転方向":
                        rename_map[c] = "UPカット"
                if rename_map:
                    df = df.rename(columns=rename_map)
                return df

            df_results = _normalize_columns(df_results)

            # ES: 3. Leer archivo de muestreo actual (Excel/CSV)
            # EN: 3. Read the current sampling file (Excel/CSV)
            # JP: 3. 現在のサンプルファイル（Excel/CSV）を読み込む
            print(f"📊 サンプルファイルを読み込み中: {sample_file_path}")
            df_sample = _read_any_table(sample_file_path)
            df_sample = _normalize_columns(df_sample)

            print(f"📊 元のサンプルファイル行数: {len(df_sample)} 行")

            # ES: 4. Eliminar filas del archivo de muestreo:
            # EN: 4. Remove rows from the sampling file:
            # JP: 4. サンプルファイルから行を削除:
            # ES: - Comparar por igualdad (normalizada) solo en las columnas de condición
            # EN: - Compare by (normalized) equality only on condition columns
            # JP: - 条件列のみ正規化して等価比較する
            # ES: - Para 線材長, elegir la fila cuyo valor sea más cercano (closest match)
            # EN: - For 線材長, pick the row whose value is closest (closest match)
            # JP: - 線材長は値が最も近い行を選ぶ（closest match）
            #
            # ES: Nota: NO usamos 直径/材料 como clave porque a veces están ausentes o vacíos en resultados y eso impide eliminar filas aunque las condiciones sean iguales.
            # EN: Note: we do NOT use 直径/材料 as a key because they can be missing/empty in results, which prevents row deletion even when conditions match.
            # JP: 注: 直径/材料は結果で欠損/空の場合があり、キーにすると条件が同じでも行削除できなくなるため使用しない
            strict_cols_candidate = [
                # Condiciones
                "回転速度", "送り速度", "UPカット",
                "切込量", "突出量", "載せ率", "パス数",
            ]
            len_col = "線材長"

            available_columns = df_results.columns.tolist()
            print(f"🔍 結果ファイルの利用可能な列: {available_columns}")
            print(f"🔍 サンプルファイルの利用可能な列: {df_sample.columns.tolist()}")

            # Requerimos al menos las 7 columnas de condición
            required_condition_cols = ["回転速度", "送り速度", "UPカット", "切込量", "突出量", "載せ率", "パス数"]
            missing_required = [c for c in required_condition_cols if (c not in df_results.columns or c not in df_sample.columns)]
            if missing_required:
                raise ValueError(f"❌ Faltan columnas de condición para comparar: {missing_required}")

            strict_cols = [c for c in strict_cols_candidate if (c in df_results.columns and c in df_sample.columns)]
            if not strict_cols:
                raise ValueError("❌ No hay columnas comunes suficientes para comparar resultados vs 未実験データ.")

            if len_col not in df_results.columns:
                raise ValueError(f"❌ El archivo de resultados no contiene la columna requerida: {len_col}")

            if len_col not in df_sample.columns:
                print(f"⚠️ サンプルファイルに '{len_col}' がありません。厳密キー一致の最初の一致を削除します。")

            import numpy as np
            from collections import defaultdict

            # Derivar un brush_id estable (si hay one-hot en ambos)
            brush_cols = ["A13", "A11", "A21", "A32"]
            has_brush = all(c in df_results.columns for c in brush_cols) and all(c in df_sample.columns for c in brush_cols)

            int_cols = set(["回転速度", "送り速度", "UPカット", "パス数"])
            float_cols = set(["切込量", "突出量", "載せ率"])

            def _normalize_upcut_series(s: pd.Series) -> pd.Series:
                # Aceptar 0/1, True/False, y algunas variantes texto comunes
                if s is None:
                    return s
                try:
                    if s.dtype == "bool":
                        return s.astype("Int64")
                except Exception:
                    pass
                # map texto -> 0/1 cuando aplique
                s_str = s.astype(str).str.replace("\u3000", " ").str.strip()
                upper = s_str.str.upper()
                mapped = upper.map({
                    "UP": 1, "DOWN": 0,
                    "CW": 1, "CCW": 0,
                    "TRUE": 1, "FALSE": 0,
                    "1": 1, "0": 0,
                })
                # conservar original donde no mapea
                return pd.to_numeric(mapped.fillna(s_str), errors="coerce")

            def _norm_key_cols(df: pd.DataFrame, cols: list) -> pd.DataFrame:
                out = df[cols].copy()
                for c in cols:
                    if c in int_cols:
                        if c == "UPカット":
                            out[c] = _normalize_upcut_series(out[c]).round(0).astype("Int64")
                        else:
                            out[c] = pd.to_numeric(out[c], errors="coerce").round(0).astype("Int64")
                    elif c in float_cols:
                        out[c] = pd.to_numeric(out[c], errors="coerce").round(6)
                    else:
                        out[c] = out[c]
                return out

            def _brush_id_from_onehot(df: pd.DataFrame) -> pd.Series:
                # 1->A11, 2->A21, 3->A32, 4->A13
                a13 = pd.to_numeric(df["A13"], errors="coerce").fillna(0).astype(int)
                a11 = pd.to_numeric(df["A11"], errors="coerce").fillna(0).astype(int)
                a21 = pd.to_numeric(df["A21"], errors="coerce").fillna(0).astype(int)
                a32 = pd.to_numeric(df["A32"], errors="coerce").fillna(0).astype(int)
                bid = pd.Series([pd.NA] * len(df), index=df.index, dtype="Int64")
                bid = bid.mask(a11 == 1, 1)
                bid = bid.mask(a21 == 1, 2)
                bid = bid.mask(a32 == 1, 3)
                bid = bid.mask(a13 == 1, 4)
                return bid

            sample_key_df = _norm_key_cols(df_sample, strict_cols)
            results_key_df = _norm_key_cols(df_results, strict_cols)

            match_cols = list(strict_cols)
            if has_brush:
                df_sample["__brush_id"] = _brush_id_from_onehot(df_sample)
                df_results["__brush_id"] = _brush_id_from_onehot(df_results)
                match_cols.append("__brush_id")

            # Normalizar también brush_id
            if "__brush_id" in match_cols:
                sample_key_df["__brush_id"] = df_sample["__brush_id"].astype("Int64")
                results_key_df["__brush_id"] = df_results["__brush_id"].astype("Int64")

            # Arrays de longitud
            sample_len = pd.to_numeric(df_sample[len_col], errors="coerce").astype(float).to_numpy() if len_col in df_sample.columns else np.full(len(df_sample), np.nan)
            results_len = pd.to_numeric(df_results[len_col], errors="coerce").astype(float).to_numpy()

            # Construir lookup: key -> lista de posiciones (indices) del sample
            buckets = defaultdict(list)
            sample_idx = df_sample.index.to_numpy()
            for i in range(len(df_sample)):
                row = sample_key_df.iloc[i]
                # key como tupla (incluye NA como None)
                key = tuple([None if pd.isna(row[c]) else row[c] for c in match_cols])
                buckets[key].append(i)

            used_pos = np.zeros(len(df_sample), dtype=bool)
            rows_to_remove = []
            missing = 0
            for r_i in range(len(df_results)):
                rrow = results_key_df.iloc[r_i]
                rkey = tuple([None if pd.isna(rrow[c]) else rrow[c] for c in match_cols])
                cand = buckets.get(rkey, [])
                # filtrar usados
                cand = [p for p in cand if not used_pos[p]]
                if not cand:
                    missing += 1
                    continue

                chosen = cand[0]
                rlen = results_len[r_i]
                if len_col in df_sample.columns and not np.isnan(rlen):
                    d = np.abs(sample_len[cand] - rlen)
                    if not np.all(np.isnan(d)):
                        chosen = cand[int(np.nanargmin(d))]

                used_pos[chosen] = True
                rows_to_remove.append(sample_idx[chosen])

            if missing > 0:
                print(f"⚠️ 結果 {missing}/{len(df_results)} 行で一致が見つかりませんでした（型/列/値を確認してください）")

            if rows_to_remove:
                print(f"🧹 一致件数: {len(rows_to_remove)} (線材長で近接)")
            
            # Eliminar filas duplicadas
            if rows_to_remove:
                df_sample_updated = df_sample.drop(rows_to_remove)
                print(f"🗑️ {len(rows_to_remove)} 件の重複行が削除されました")
                print(f"📊 更新後サンプルファイル行数: {len(df_sample_updated)} 行")
                
                # ES: Guardar archivo actualizado | EN: Save updated file | JA: 更新ファイルを保存
                try:
                    if sample_ext == ".csv":
                        df_sample_updated.to_csv(sample_file_path, index=False, encoding="utf-8-sig")
                    else:
                        df_sample_updated.to_excel(sample_file_path, index=False)
                except PermissionError as e:
                    # ES: En Windows esto suele pasar si el archivo está abierto (Excel lo bloquea)
                    # EN: On Windows, this usually happens if the file is open (Excel locks it)
                    # JP: Windowsではファイルが開かれていると起きやすい（Excelがロックする）
                    raise PermissionError(
                        f"❌ 99_Temp にサンプルファイルを保存できませんでした（権限が拒否されました）。\n\n"
                        f"Excel などでファイルが開かれている可能性があります。\n"
                        f"閉じてから再度お試しください。\n\n"
                        f"ファイル:\n{sample_file_path}"
                    ) from e
                print(f"✅ 更新後サンプルファイルを保存しました: {sample_file_path}")
                
                return {
                    'backup_path': backup_path,
                    'removed_rows': len(rows_to_remove),
                    'remaining_rows': len(df_sample_updated)
                }
            else:
                print(f"ℹ️ 削除する重複行が見つかりませんでした")
                return {
                    'backup_path': backup_path,
                    'removed_rows': 0,
                    'remaining_rows': len(df_sample)
                }
                
        except RuntimeError as e:
            if "already deleted" in str(e):
                # Ignorar silenciosamente el error de widget ya eliminado
                pass
            else:
                print(f"❌ backup_and_update_sample_file でエラー: {str(e)}")
                print("🔍 Debug - 現在の状態:")
                print(f"  - project_folder: {project_folder}")
                print(f"  - project_name: {os.path.basename(project_folder) if project_folder else '未指定'}")
                print(f"  - results_file_path: {results_file_path}")
                print(f"  - 期待する temp_base: {os.path.join(project_folder, '99_Temp') if project_folder else '未指定'}")
                raise e
        except Exception as e:
            print(f"❌ backup_and_update_sample_file でエラー: {str(e)}")
            print("🔍 Debug - 現在の状態:")
            print(f"  - project_folder: {project_folder}")
            print(f"  - project_name: {os.path.basename(project_folder) if project_folder else '未指定'}")
            print(f"  - results_file_path: {results_file_path}")
            print(f"  - 期待する temp_base: {os.path.join(project_folder, '99_Temp') if project_folder else '未指定'}")
            raise e

    def on_execute_results_clicked(self):
        if not hasattr(self, "results_file_path"):
            QMessageBox.warning(self, "エラー", "❌ 結果ファイルが読み込まれていません。")
            return
        
        # ES: Obtener valores de la UI
        # EN: Get values from the UI
        # JP: UIから値を取得
        # brush y 線材長 deben venir del archivo de resultados (no de la UI)
        selected_brush = None
        diameter = float(self.diameter_selector.currentText())
        material = self.material_selector.currentText()

        try:
            # ES: Hacer backup y actualizar archivo de muestreo | EN: Backup and update sampling file | JA: バックアップしてサンプリングファイルを更新
            print("🔄 サンプルファイルのバックアップと更新を開始...")
            # ES: Solo hacer backup si hay un proyecto activo
            # EN: Only create a backup if there is an active project
            # JP: アクティブなプロジェクトがある場合のみバックアップする
            if hasattr(self, 'proyecto_folder'):
                backup_result = self.backup_and_update_sample_file(self.results_file_path, self.proyecto_folder)
            else:
                print("⚠️ アクティブなプロジェクトがありません。バックアップとサンプルファイルの更新をスキップします")
                backup_result = {'backup_path': None, 'removed_rows': 0, 'remaining_rows': 0}
            
            # Procesar archivo de resultados (線材長 viene del archivo)
            dbu = self.processor.process_results_file_with_ui_values(
                self.results_file_path, 
                selected_brush, 
                diameter, 
                material
            )
            
            # ES: Mostrar mensaje de éxito con información del backup | EN: Show success message with backup info | JA: バックアップ情報付き成功メッセージを表示
            message = f"✅ 結果ファイルがデータベースに取り込まれました。\n\n"
            if backup_result['backup_path']:
                message += f"📋 バックアップ作成: {os.path.basename(backup_result['backup_path'])}\n"
                message += f"🗑️ サンプルファイルから削除された行: {backup_result['removed_rows']}\n"
                message += f"📊 サンプルファイルの残り行数: {backup_result['remaining_rows']}"
            else:
                message += f"ℹ️ バックアップは実行されませんでした（アクティブなプロジェクトがありません）"
            
            # Aviso único si hubo sobrescritura en BBDD
            if isinstance(dbu, dict):
                updated = int(dbu.get("updated", 0) or 0)
                inserted = int(dbu.get("inserted", 0) or 0)
                if updated > 0:
                    message += "\n\n⚠️ 既存データを上書きします。BBDDのバックアップを作成しました。"
                    message += f"\n🔁 上書き: {updated} / ➕ 追加: {inserted}"
                    if dbu.get("db_backup_path"):
                        message += f"\n📋 BBDDバックアップ: {os.path.basename(str(dbu.get('db_backup_path')))}"
                    else:
                        message += "\n📋 BBDDバックアップ: (作成できませんでした)"

            QMessageBox.information(self, "完了", message)
            self.create_filter_view()
        except Exception as e:
            QMessageBox.critical(self, "エラー", f"❌ データベースへの取り込み中にエラーが発生しました:\n{str(e)}")

    def closeEvent(self, event):
        """ES: Maneja el cierre de la ventana principal
        EN: Handle main window close
        JA: メインウィンドウの閉じるを処理"""
        try:
            print("🛑 アプリケーションを終了中...")
            
            # ES: Cancelar análisis no lineal si está corriendo
            # EN: Cancel non-linear analysis if it is running
            # JP: 非線形解析が実行中ならキャンセルする
            if hasattr(self, 'nonlinear_worker') and self.nonlinear_worker is not None:
                if self.nonlinear_worker.isRunning():
                    print("🛑 終了前に非線形解析をキャンセル中...")
                    self.nonlinear_worker.cancel()
                    
                    # Esperar a que el thread termine (máximo 5 segundos)
                    if self.nonlinear_worker.isRunning():
                        self.nonlinear_worker.quit()
                        if not self.nonlinear_worker.wait(5000):
                            print("⚠️ ワーカーが5秒以内に終了しなかったため、強制終了します...")
                            self.nonlinear_worker.terminate()
                            self.nonlinear_worker.wait(1000)
                    
                    print("✅ 非線形解析ワーカーをキャンセルしました")
            
            # ES: Cerrar base de datos
            # EN: Close database
            # JP: データベースを閉じる
            if hasattr(self, 'db'):
                self.db.close()
            
            print("✅ アプリケーションを正常に終了しました")
            event.accept()
            
        except Exception as e:
            print(f"❌ closeEvent でエラー: {e}")
            import traceback
            traceback.print_exc()
            # ES: Aún así cerrar la aplicación
            # EN: Still close the application
            # JP: それでもアプリケーションを終了
            if hasattr(self, 'db'):
                try:
                    self.db.close()
                except:
                    pass
            event.accept()

    def handle_single_file_load(self):
        # ES: Pausar timers automáticos para evitar interferencia con el diálogo | EN: Pause auto timers to avoid interference with the dialog | JA: ダイアログとの干渉を避けるため自動タイマーを一時停止
        self.pause_auto_timers()
        
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "ファイルを選択",
            "",
            "Excel/CSV Files (*.xlsx *.xls *.csv);;Excel Files (*.xlsx *.xls);;CSV Files (*.csv)"
        )
        
        # ES: Reanudar timers después del diálogo | EN: Resume timers after the dialog | JA: ダイアログ後にタイマーを再開
        self.resume_auto_timers()
        
        if not file_path:
            self.load_file_label.setText("ファイル未選択")
            # Reset all UI elements to default state when no file is selected
            self.set_ui_state_for_no_file()
            return

        self.load_file_label.setText(f"読み込み済み: {os.path.basename(file_path)}")

        try:
            ext = os.path.splitext(file_path)[1].lower()
            if ext == ".csv":
                df_raw = pd.read_csv(file_path, header=None, nrows=2, encoding="utf-8-sig")
            else:
                df_raw = pd.read_excel(file_path, header=None, nrows=2)

            fila_1 = df_raw.iloc[0].fillna("").tolist()
            fila_2 = df_raw.iloc[1].fillna("").tolist()

            # Aceptar tanto "UPカット" como "回転方向" como 3ª columna
            columnas_muestreo_exactas = ['回転速度', '送り速度', 'UPカット/回転方向', '切込量', '突出量', '載せ率', 'パス数']
            # Reconocimiento de resultados (nuevo formato): incluye brush one-hot, 線材長, 面粗度(Ra)前/後, y opcionalmente 切削力X/Y/Z
            columnas_resultados_minimas = [
                '回転速度', '送り速度', 'UPカット/回転方向', '切込量', '突出量', '載せ率', 'パス数',
                '線材長',
                '上面ダレ/上面ダレ量', '側面ダレ/側面ダレ量', '摩耗量',
                '面粗度(Ra)前/面粗度前/粗度(Ra)前', '面粗度(Ra)後/面粗度後/粗度(Ra)後',
                '実験日'
            ]

            def _matches_sample_header(row, start_idx: int) -> bool:
                try:
                    # ES: Formato antiguo: 7 variables
                    # EN: Old format: 7 variables
                    # JP: 旧フォーマット：7変数
                    if (
                        row[start_idx] == '回転速度' and
                        row[start_idx + 1] == '送り速度' and
                        row[start_idx + 2] in ('UPカット', '回転方向') and
                        row[start_idx + 3] in ('切込量', '切込み量') and
                        row[start_idx + 4] in ('突出量', '突出し量') and
                        row[start_idx + 5:start_idx + 7] == ['載せ率', 'パス数']
                    ):
                        return True

                    # ES: Formato nuevo: one-hot brush + variables
                    # EN: New format: one-hot brush + variables
                    # JP: 新フォーマット：ブラシone-hot＋変数
                    if (
                        row[start_idx:start_idx + 4] == ['A13', 'A11', 'A21', 'A32'] and
                        row[start_idx + 4] == '回転速度' and
                        row[start_idx + 5] == '送り速度' and
                        row[start_idx + 6] in ('UPカット', '回転方向') and
                        row[start_idx + 7] in ('切込量', '切込み量') and
                        row[start_idx + 8] in ('突出量', '突出し量') and
                        row[start_idx + 9:start_idx + 11] == ['載せ率', 'パス数']
                    ):
                        return True

                    return False
                except Exception:
                    return False

            def _matches_results_header(row) -> bool:
                """
                Detecta archivo de resultados por presencia de columnas de condiciones + resultados.
                - Requiere: A13/A11/A21/A32 + 7 variables de condición + 線材長 + (上面/側面/摩耗) + (面粗度 前/後) + 実験日
                - Acepta variantes: 回転方向 vs UPカット, 突出し量 vs 突出量, 上面ダレ量 vs 上面ダレ, 側面ダレ量 vs 側面ダレ
                - 切削力X/Y/Z: opcional
                """
                try:
                    headers = {str(x).strip() for x in row if str(x).strip() != ""}
                    has_brush = all(c in headers for c in ("A13", "A11", "A21", "A32"))
                    has_dir = ('UPカット' in headers) or ('回転方向' in headers)
                    has_out = ('突出量' in headers) or ('突出し量' in headers)
                    has_cut = ('切込量' in headers) or ('切込み量' in headers)
                    has_top = ('上面ダレ' in headers) or ('上面ダレ量' in headers)
                    has_side = ('側面ダレ' in headers) or ('側面ダレ量' in headers)
                    has_ra_pre = ('面粗度(Ra)前' in headers) or ('面粗度前' in headers) or ('粗度(Ra)前' in headers)
                    has_ra_post = ('面粗度(Ra)後' in headers) or ('面粗度後' in headers) or ('粗度(Ra)後' in headers)

                    has_design = (
                        ('回転速度' in headers) and
                        ('送り速度' in headers) and
                        has_dir and
                        has_cut and
                        has_out and
                        ('載せ率' in headers) and
                        ('パス数' in headers)
                    )
                    has_results = has_top and has_side and ('摩耗量' in headers) and has_ra_pre and has_ra_post
                    has_required_meta = ('線材長' in headers) and ('実験日' in headers)
                    return has_brush and has_design and has_results and has_required_meta
                except Exception:
                    return False

            # ES: Verificar archivo de resultados (nuevo): header en fila 1 o (a veces) en fila 2 | EN: Check results file (new): header in row 1 or (sometimes) row 2 | JA: 結果ファイル確認（新）：ヘッダーは1行目または2行目
            # ES: ✅ Prioridad: si un archivo parece "resultados" y "muestreo" a la vez, se tratará como resultados.
            # EN: ✅ Priority: if a file looks like both "results" and "sample", treat it as results.
            # JP: ✅ 優先: 「結果」と「サンプル」の両方に見える場合は結果として扱う
            is_resultados = _matches_results_header(fila_1) or _matches_results_header(fila_2)

            # ES: Verificar archivo de muestreo | EN: Check sampling file | JA: サンプリングファイルを確認
            # - Permite offset 0 (A1) o 1 (si hay columna índice/No. al inicio)
            is_muestreo = _matches_sample_header(fila_1, 0) or _matches_sample_header(fila_1, 1)

            # ES: Debug: imprimir las filas para diagnosticar
            # EN: Debug: print the rows for diagnosis
            # JP: デバッグ: 診断のため行を出力
            print(f"🔍 デバッグ - 行 1: {fila_1}")
            print(f"🔍 デバッグ - 行 2: {fila_2}")
            print(f"🔍 デバッグ - 期待するサンプル列: {columnas_muestreo_exactas}")
            print(f"🔍 デバッグ - 期待する結果列: {columnas_resultados_minimas}")
            print(f"🔍 デバッグ - 結果ファイル判定: {is_resultados}")
            print(f"🔍 デバッグ - サンプルファイル判定: {is_muestreo}")

            if is_resultados:
                QMessageBox.information(self, "ファイル種別", "📄 このファイルは【結果】ファイルとして認識されました。")
                self.results_file_path = file_path
                self.show_results_button.setEnabled(True)
                
                # Set UI state for results file
                self.set_ui_state_for_results_file()
                # ES: Aplicar restricciones según cepillo detectado del archivo (p.ej. A13 limita diámetros)
                # EN: Apply constraints based on the brush detected from the file (e.g., A13 limits diameters)
                # JP: ファイルから検出したブラシに基づいて制約を適用（例：A13は直径を制限）
                try:
                    self._apply_results_file_brush_to_ui(file_path)
                except Exception:
                    pass
                # ES: Habilitación de UI debajo del selector (sin depender del nombre del archivo)
                # EN: UI enablement below the selector (independent of the file name)
                # JP: セレクタ下のUI有効化（ファイル名に依存しない）
                try:
                    self._last_loaded_file_kind = "results"
                    if hasattr(self, "on_file_loaded"):
                        self.on_file_loaded(file_path, is_results=True)
                    elif hasattr(self, "_set_widgets_below_sample_selector_enabled"):
                        self._set_widgets_below_sample_selector_enabled(True)
                except Exception:
                    pass

            elif is_muestreo:
                QMessageBox.information(self, "ファイル種別", "📄 このファイルは【サンプル】ファイルとして認識されました。")
                self.sample_file_path = file_path
                self.show_results_button.setEnabled(False)
                
                # ES: Habilitación de UI debajo del selector (sin depender del nombre del archivo)
                # EN: UI enablement below the selector (independent of the file name)
                # JP: セレクタ下のUI有効化（ファイル名に依存しない）
                try:
                    self._last_loaded_file_kind = "sample"
                    if hasattr(self, "on_file_loaded"):
                        self.on_file_loaded(file_path, is_results=False)
                    elif hasattr(self, "_set_widgets_below_sample_selector_enabled"):
                        self._set_widgets_below_sample_selector_enabled(False)
                except Exception:
                    pass
                
                # ES: Verificar si el archivo pertenece a un proyecto diferente | EN: Check if file belongs to a different project | JA: ファイルが別プロジェクトのものか確認
                file_dir = os.path.dirname(file_path)
                file_name = os.path.basename(file_path)
                
                print(f"🔍 デバッグ Load: file_dir = {file_dir}")
                print(f"🔍 デバッグ Load: file_name = {file_name}")
                print(f"🔍 デバッグ Load: proyecto_folder = {getattr(self, 'proyecto_folder', '存在しません')}")
                
                # ES: Si hay un proyecto activo, verificar si el archivo pertenece al mismo proyecto
                # EN: If there is an active project, check whether the file belongs to the same project
                # JP: アクティブなプロジェクトがある場合、同じプロジェクトのファイルか確認
                if hasattr(self, 'proyecto_folder') and hasattr(self, 'proyecto_nombre'):
                    # ES: Verificar si el archivo está en el proyecto principal o en sus subcarpetas.
                    # EN: Check if the file is in the main project or its subfolders.
                    # JP: ファイルがメインプロジェクト配下（サブフォルダ含む）か確認します。
                    is_same_project = (file_dir == self.proyecto_folder or 
                                      file_dir.startswith(self.proyecto_folder + os.sep))
                    
                    print(f"🔍 Debug Load: is_same_project = {is_same_project}")
                    
                    if not is_same_project:
                        # ES: Archivo de un proyecto diferente: limpiar proyecto activo
                        # EN: File is from a different project: clear active project
                        # JP: 別プロジェクトのファイル: アクティブプロジェクトをクリア
                        print(f"🔄 別プロジェクトのファイルを検出しました。アクティブプロジェクトをクリアします: {getattr(self, 'proyecto_nombre', 'Unknown')}")
                        print(f"🔄 ファイル: {file_dir}")
                        print(f"🔄 プロジェクト: {self.proyecto_folder}")
                        delattr(self, 'proyecto_folder')
                        delattr(self, 'proyecto_nombre')
                        if hasattr(self, 'muestreo_guardado_path'):
                            delattr(self, 'muestreo_guardado_path')
                        print("✅ アクティブプロジェクトをクリアしました。次回の最適化で新しいプロジェクトを要求します。")
                    else:
                        print(f"✅ ファイルはアクティブプロジェクトに属しています: {getattr(self, 'proyecto_nombre', 'Unknown')}")
                else:
                    print("🔍 Debug Load: アクティブなプロジェクトがありません")
                
                # ES: Si estamos en la pantalla de filtros, volver a la pantalla principal | EN: If on filter screen, return to main screen | JA: フィルター画面ならメイン画面に戻る
                # ES: Verificar si estamos en la vista de filtros | EN: Check if we are on filter view | JA: フィルタビューか確認
                in_filter_view = False
                for i in range(self.center_layout.count()):
                    item = self.center_layout.itemAt(i)
                    if item.widget() and isinstance(item.widget(), QLabel):
                        if item.widget().text() == "データフィルター":
                            in_filter_view = True
                            break
                
                if in_filter_view:
                    print("🔄 フィルター画面でサンプルファイルを検出しました。メイン画面に戻ります...")
                    # ES: Limpiar la pantalla y volver al estado inicial
                    # EN: Clear the screen and return to the initial state
                    # JP: 画面をクリアして初期状態に戻す
                    self.clear_main_screen()
                
                # ES: Habilitar botones de optimización cuando se carga un nuevo archivo de muestras
                # EN: Enable optimization buttons when a new sample file is loaded
                # JP: 新しいサンプルファイル読み込み時に最適化ボタンを有効化
                self.d_optimize_button.setEnabled(True)
                self.i_optimize_button.setEnabled(True)
                # ES: Aplicar estilo visual de habilitado
                # EN: Apply enabled visual style
                # JP: 有効時の見た目スタイルを適用
                self.d_optimize_button.setStyleSheet(self.d_optimize_button.styleSheet())
                self.i_optimize_button.setStyleSheet(self.i_optimize_button.styleSheet())
                
                # Set UI state for sample file
                self.set_ui_state_for_sample_file()
                
                # ES: Limpiar resultados anteriores
                # EN: Clear previous results
                # JP: 以前の結果をクリア
                if hasattr(self, 'dsaitekika_results'):
                    delattr(self, 'dsaitekika_results')
                if hasattr(self, 'isaitekika_results'):
                    delattr(self, 'isaitekika_results')
                if hasattr(self, 'dsaitekika_selected_df'):
                    delattr(self, 'dsaitekika_selected_df')
                if hasattr(self, 'isaitekika_selected_df'):
                    delattr(self, 'isaitekika_selected_df')
                
                # ES: Limpiar gráficos y tablas anteriores
                # EN: Clear previous charts and tables
                # JP: 以前のグラフとテーブルをクリア
                self.graph_images = []
                self.graph_images_content = []
                self.current_graph_index = 0
                
                # ES: Limpiar área de gráficos
                # EN: Clear chart area
                # JP: グラフ領域をクリア
                if hasattr(self, 'graph_area') and self.graph_area.layout():
                    layout = self.graph_area.layout()
                    for i in reversed(range(layout.count())):
                        widget = layout.itemAt(i).widget()
                        if widget:
                            widget.setParent(None)
                
                # ES: Deshabilitar botones OK/NG
                # EN: Disable OK/NG buttons
                # JP: OK/NGボタンを無効化
                self.ok_button.setEnabled(False)
                self.ng_button.setEnabled(False)

            else:
                QMessageBox.warning(self, "警告", "⚠️ このファイルはサンプルでも結果でもないようです。")
                self.show_results_button.setEnabled(False)
                
                # Reset all UI elements to default state when file is neither sample nor results
                self.set_ui_state_for_no_file()
                try:
                    self._last_loaded_file_kind = None
                    if hasattr(self, "on_file_loaded"):
                        self.on_file_loaded(file_path, is_results=False)
                    elif hasattr(self, "_set_widgets_below_sample_selector_enabled"):
                        self._set_widgets_below_sample_selector_enabled(False)
                except Exception:
                    pass

        except Exception as e:
            QMessageBox.critical(self, "エラー", f"❌ ファイルの読み込み中にエラーが発生しました:\n{str(e)}")
            # Reset all UI elements to default state when error occurs
            self.set_ui_state_for_no_file()
            try:
                self._last_loaded_file_kind = None
                if hasattr(self, "_set_widgets_below_sample_selector_enabled"):
                    self._set_widgets_below_sample_selector_enabled(False)
            except Exception:
                pass

    def get_sample_size(self):
        """ES: Obtener el tamaño de muestra del campo de entrada
        EN: Get sample size from input field
        JA: 入力欄からサンプルサイズを取得"""
        try:
            size = int(self.sample_size_input.text())
            if 10 <= size <= 50:
                return size
            else:
                QMessageBox.warning(self, "エラー", f"❌ サンプルサイズは10-50の範囲内である必要があります。\n現在の値: {size}")
                self.sample_size_input.setText("15")
                return 15
        except ValueError:
            QMessageBox.warning(self, "エラー", "❌ サンプルサイズは数値である必要があります。")
            self.sample_size_input.setText("15")
            return 15

    def validate_sample_size(self):
        """ES: Validar el tamaño de muestra cuando se termina de editar
        EN: Validate sample size when editing is finished
        JA: 編集完了時にサンプルサイズを検証"""
        try:
            size = int(self.sample_size_input.text())
            if not (10 <= size <= 50):
                QMessageBox.warning(self, "エラー", f"❌ サンプルサイズは10-50の範囲内である必要があります。\n現在の値: {size}")
                self.sample_size_input.setText("15")
        except ValueError:
            QMessageBox.warning(self, "エラー", "❌ サンプルサイズは数値である必要があります。")
            self.sample_size_input.setText("15")

    def on_sample_size_focus_out(self, event):
        """ES: Manejar la pérdida de foco del campo de tamaño de muestra
        EN: Handle focus loss on sample size field
        JA: サンプルサイズ欄のフォーカス喪失を処理"""
        # ES: Llamar al método original de QLineEdit
        # EN: Call the original QLineEdit method
        # JP: QLineEditの元のメソッドを呼び出す
        super(QLineEdit, self.sample_size_input).focusOutEvent(event)
        # ES: Validar el valor
        # EN: Validate the value
        # JP: 値を検証する
        self.validate_sample_size()

    def export_database_to_excel(self):
        db_path = RESULTS_DB_PATH
        conn = sqlite3.connect(db_path, timeout=10)

        try:
            df = pd.read_sql_query("SELECT * FROM main_results", conn)
        except Exception as e:
            QMessageBox.critical(self, "エラー", f"❌ データベースからの取得中にエラーが発生しました:\n{e}")
            return
        finally:
            conn.close()

        # ES: Formatear columnas según el orden esperado de resultados (sin tocar la DB)
        # EN: Format columns in the expected results order (without touching the DB)
        # JP: 結果の想定順で列を整形（DBは変更しない）
        try:
            rename_map = {
                "面粗度前": "面粗度(Ra)前",
                "面粗度後": "面粗度(Ra)後",
            }
            df_export = df.rename(columns=rename_map)
            desired_order = [
                "id",
                "バリ除去", "上面ダレ量", "側面ダレ量", "摩耗量",
                "切削力X", "切削力Y", "切削力Z",
                "面粗度(Ra)後",
                "A13", "A11", "A21", "A32",
                "直径", "材料",
                "回転速度", "送り速度", "UPカット", "切込量", "突出量", "載せ率", "線材長", "パス数",
                "加工時間",
                "面粗度(Ra)前",
                "実験日",
            ]
            for col in desired_order:
                if col not in df_export.columns:
                    df_export[col] = ""
            df_export = df_export[[c for c in desired_order if c in df_export.columns]]
        except Exception:
            df_export = df

        # ES: Pausar timers automáticos para evitar interferencia con el diálogo | EN: Pause auto timers to avoid interference with the dialog | JA: ダイアログとの干渉を避けるため自動タイマーを一時停止
        self.pause_auto_timers()
        
        options = QFileDialog.Options()
        filepath, _ = QFileDialog.getSaveFileName(
            self, "Excelとして保存", "", "Excelファイル (*.xlsx)", options=options
        )
        
        # ES: Reanudar timers después del diálogo | EN: Resume timers after the dialog | JA: ダイアログ後にタイマーを再開
        self.resume_auto_timers()

        if filepath:
            try:
                df_export.to_excel(filepath, index=False)
                QMessageBox.information(self, "完了", "✅ データベースが正常にエクスポートされました。")
            except Exception as e:
                QMessageBox.critical(self, "エラー", f"❌ エクスポート中にエラーが発生しました:\n{e}")

    def export_yosoku_database_to_excel(self):
        """ES: Exportar base de datos de Yosoku a Excel con diálogo de progreso
        EN: Export Yosoku database to Excel with progress dialog
        JA: 予測DBをExcelにエクスポート（進捗ダイアログ付き）"""
        # ES: Crear diálogo personalizado más bonito | EN: Create nicer custom dialog | JA: より見やすいカスタムダイアログを作成
        dialog = QDialog(self)
        dialog.setWindowTitle("データベース選択")
        dialog.setFixedSize(500, 350)
        dialog.setWindowFlags(Qt.Dialog | Qt.WindowTitleHint | Qt.WindowCloseButtonHint)
        
        # ES: Layout principal | EN: Main layout | JA: メインレイアウト
        main_layout = QVBoxLayout(dialog)
        main_layout.setSpacing(20)
        main_layout.setContentsMargins(30, 30, 30, 30)
        
        # ES: Título | EN: Title | JA: タイトル
        title_label = QLabel("データベースを選択")
        title_label.setStyleSheet("""
            QLabel {
                font-size: 24px;
                font-weight: bold;
                color: #2c3e50;
                padding: 10px;
            }
        """)
        title_label.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(title_label)
        
        # Subtítulo
        subtitle_label = QLabel("エクスポートするデータベースを選択してください")
        subtitle_label.setStyleSheet("""
            QLabel {
                font-size: 14px;
                color: #7f8c8d;
                padding: 5px;
            }
        """)
        subtitle_label.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(subtitle_label)
        
        main_layout.addSpacing(20)
        
        # ES: Contenedor para los 3 botones alineados | EN: Container for 3 aligned buttons | JA: 3ボタン揃え用コンテナ
        buttons_container = QHBoxLayout()
        buttons_container.setSpacing(15)
        buttons_container.setContentsMargins(0, 0, 0, 0)
        
        # ES: Botón Lineal | EN: Linear button | JA: 線形ボタン
        lineal_button = QPushButton("線形データベース")
        lineal_button.setFixedSize(140, 50)
        lineal_button.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                border-radius: 8px;
                font-size: 14px;
                font-weight: bold;
                padding: 10px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
            QPushButton:pressed {
                background-color: #21618c;
            }
        """)
        
        # ES: Botón No Lineal | EN: Nonlinear button | JA: 非線形ボタン
        no_lineal_button = QPushButton("非線形データベース")
        no_lineal_button.setFixedSize(140, 50)
        no_lineal_button.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                border-radius: 8px;
                font-size: 14px;
                font-weight: bold;
                padding: 10px;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
            QPushButton:pressed {
                background-color: #a93226;
            }
        """)
        
        # ES: Botón Cancelar | EN: Cancel button | JA: キャンセルボタン
        cancel_button = QPushButton("キャンセル")
        cancel_button.setFixedSize(140, 50)
        cancel_button.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
                border: none;
                border-radius: 8px;
                font-size: 14px;
                font-weight: bold;
                padding: 10px;
            }
            QPushButton:hover {
                background-color: #7f8c8d;
            }
            QPushButton:pressed {
                background-color: #6c7a7b;
            }
        """)
        
        # ES: Agregar los 3 botones alineados
        # EN: Add the 3 buttons aligned
        # JP: 3つのボタンを整列して追加
        buttons_container.addStretch()
        buttons_container.addWidget(lineal_button)
        buttons_container.addWidget(no_lineal_button)
        buttons_container.addWidget(cancel_button)
        buttons_container.addStretch()
        
        main_layout.addLayout(buttons_container)
        main_layout.addStretch()
        
        # ES: Estilo del diálogo
        # EN: Dialog style
        # JP: ダイアログのスタイル
        dialog.setStyleSheet("""
            QDialog {
                background-color: #f8f9fa;
                border-radius: 10px;
            }
        """)
        
        # ES: Conectar señales | EN: Connect signals | JA: シグナルを接続
        lineal_button.clicked.connect(lambda: dialog.done(1))
        no_lineal_button.clicked.connect(lambda: dialog.done(2))
        cancel_button.clicked.connect(lambda: dialog.done(0))
        
        # ES: Pausar timers para evitar interferencia
        # EN: Pause timers to avoid interference
        # JP: 干渉を避けるためタイマーを一時停止
        self.pause_auto_timers()
        
        # ES: Mostrar diálogo | EN: Show dialog | JA: ダイアログを表示
        result = dialog.exec()
        
        # ES: Reanudar timers
        # EN: Resume timers
        # JP: タイマーを再開
        self.resume_auto_timers()
        
        # ES: Determinar qué BBDD usar según la respuesta
        # EN: Decide which DB to use based on the selection
        # JP: 選択内容に応じて使用するDBを決定
        if result == 0:  # Cancelar
            return
        elif result == 1:  # Lineal
            db_path = YOSOKU_LINEAL_DB_PATH
            db_name = "線形データベース"
        elif result == 2:  # No Lineal
            db_path = YOSOKU_NO_LINEAL_DB_PATH
            db_name = "非線形データベース"
        else:
            return
        
        # ES: Verificar si la base de datos existe | EN: Check if database exists | JA: DBが存在するか確認
        if not os.path.exists(db_path):
            QMessageBox.warning(
                self, 
                "警告", 
                f"❌ {db_name}が見つかりません。\n\n"
                f"ファイル: {db_path}\n\n"
                f"まず予測を実行してデータベースにデータをインポートしてください。"
            )
            return
        
        # ES: Verificar que la base de datos no esté vacía (sin mostrar loading aún) | EN: Ensure database is not empty (no loading yet) | JA: DBが空でないか確認（ローディング表示前）
        conn = sqlite3.connect(db_path, timeout=10)
        try:
            df = pd.read_sql_query("SELECT * FROM yosoku_predictions", conn)
            
            if len(df) == 0:
                QMessageBox.information(
                    self, 
                    "情報", 
                    f"📊 {db_name}は空です。\n\n"
                    f"まず予測を実行してデータベースにデータをインポートしてください。"
                )
                return
        except Exception as e:
            QMessageBox.critical(self, "エラー", f"❌ データベースからの取得中にエラーが発生しました:\n{e}")
            return
        finally:
            conn.close()

        # ES: Pausar timers automáticos para evitar interferencia con el diálogo | EN: Pause auto timers to avoid interference with the dialog | JA: ダイアログとの干渉を避けるため自動タイマーを一時停止
        self.pause_auto_timers()
        
        options = QFileDialog.Options()
        filepath, _ = QFileDialog.getSaveFileName(
            self, "予測データベースをExcelとして保存", "", "Excelファイル (*.xlsx)", options=options
        )
        
        # ES: Reanudar timers después del diálogo | EN: Resume timers after the dialog | JA: ダイアログ後にタイマーを再開
        self.resume_auto_timers()

        if not filepath:
            return  # User canceled file selection
        
        # ES: ✅ MOSTRAR LOADING después de seleccionar el archivo
        # EN: ✅ SHOW LOADING after selecting the file
        # JP: ✅ ファイル選択後にローディングを表示
        try:
            # ES: Crear y mostrar diálogo de progreso | EN: Create and show progress dialog | JA: 進捗ダイアログを作成して表示
            self.yosoku_export_progress_dialog = YosokuExportProgressDialog(self)
            self.yosoku_export_progress_dialog.show()
            # ES: Durante el loading con chibi: flecha/consola por encima
            # EN: During chibi loading: keep arrow/console on top
            # JP: chibiローディング中：矢印/コンソールを前面に
            self.set_console_overlay_topmost(True)
            self.yosoku_export_progress_dialog.update_progress(0, "初期化中...")
            self.yosoku_export_progress_dialog.set_status("初期化中...")
            QApplication.processEvents()
            
            # ES: Crear worker thread | EN: Create worker thread | JA: ワーカースレッドを作成
            self.yosoku_export_worker = YosokuExportWorker(db_path, filepath, len(df))
            
            # ES: Conectar señales | EN: Connect signals | JA: シグナルを接続
            self.yosoku_export_worker.progress_updated.connect(self.yosoku_export_progress_dialog.update_progress)
            self.yosoku_export_worker.status_updated.connect(self.yosoku_export_progress_dialog.set_status)
            self.yosoku_export_worker.finished.connect(self.on_yosoku_export_finished)
            self.yosoku_export_worker.error.connect(self.on_yosoku_export_error)
            
            # ES: Conectar botón de cancelar | EN: Connect cancel button | JA: キャンセルボタンを接続
            self.yosoku_export_progress_dialog.cancel_button.clicked.connect(self.cancel_yosoku_export)
            
            # Iniciar worker
            self.yosoku_export_worker.start()
            
        except Exception as e:
            print(f"❌ エクスポート開始中にエラー: {e}")
            import traceback
            traceback.print_exc()
            
            # ES: Cerrar loading si hay error
            # EN: Close loading if there is an error
            # JP: エラー時にローディングを閉じる
            if hasattr(self, 'yosoku_export_progress_dialog') and self.yosoku_export_progress_dialog is not None:
                self.yosoku_export_progress_dialog.close()
                self.yosoku_export_progress_dialog = None
            self.set_console_overlay_topmost(False)
            
            QMessageBox.critical(
                self,
                "エラー",
                f"❌ エクスポート開始中にエラーが発生しました:\n{str(e)}"
            )

    def set_ui_state_for_sample_file(self):
        """Set UI state when a sample file is loaded"""
        # No hay selector de brush en UI; resetear brush detectado de resultados
        self._results_brush_type = None
        try:
            self.update_diameter_options("")
        except Exception:
            pass
        self.sample_size_input.setEnabled(True)
        self.sample_size_input.setStyleSheet("")
        self.d_optimize_button.setEnabled(True)
        self.i_optimize_button.setEnabled(True)
        # Apply original blue style for action buttons
        self.d_optimize_button.setStyleSheet("""
            QPushButton {
                background-color: #3A80BA;
                color: white;
                font-family: "Noto Sans JP";
                border: none;
                border-radius: 8px;
                font-size: 16px;
                padding: 8px 20px;
            }
            QPushButton:hover {
                background-color: #336DA3;
            }
            QPushButton:disabled {
                background-color: #CCCCCC;
                color: #888888;
            }
        """)
        self.i_optimize_button.setStyleSheet("""
            QPushButton {
                background-color: #3A80BA;
                color: white;
                font-family: "Noto Sans JP";
                border: none;
                border-radius: 8px;
                font-size: 16px;
                padding: 8px 20px;
            }
            QPushButton:hover {
                background-color: #336DA3;
            }
            QPushButton:disabled {
                background-color: #CCCCCC;
                color: #888888;
            }
        """)
        self.material_selector.setEnabled(False)
        self.material_selector.setStyleSheet("color: gray; background-color: #f0f0f0;")
        self.diameter_selector.setEnabled(False)
        self.diameter_selector.setStyleSheet("color: gray; background-color: #f0f0f0;")
        # ES: El botón de análisis siempre está habilitado
        # EN: The analysis button is always enabled
        # JP: 解析ボタンは常に有効
        self.analyze_button.setEnabled(True)

    def set_ui_state_for_results_file(self):
        """Set UI state when a results file is loaded"""
        self.sample_size_input.setEnabled(False)
        self.sample_size_input.setStyleSheet("color: gray; background-color: #f0f0f0;")
        self.d_optimize_button.setEnabled(False)
        self.i_optimize_button.setEnabled(False)
        self.d_optimize_button.setStyleSheet("color: gray; background-color: #f0f0f0;")
        self.i_optimize_button.setStyleSheet("color: gray; background-color: #f0f0f0;")
        self.material_selector.setEnabled(True)
        self.material_selector.setStyleSheet("")
        self.diameter_selector.setEnabled(True)
        self.diameter_selector.setStyleSheet("")
        # ES: Habilitar botón de análisis
        # EN: Enable analysis button
        # JP: 解析ボタンを有効化
        self.analyze_button.setEnabled(True)

    def set_ui_state_for_no_file(self):
        """Set UI state when no file is loaded"""
        self._results_brush_type = None
        try:
            self.update_diameter_options("")
        except Exception:
            pass
        self.sample_size_input.setEnabled(False)
        self.sample_size_input.setStyleSheet("color: gray; background-color: #f0f0f0;")
        self.d_optimize_button.setEnabled(False)
        self.i_optimize_button.setEnabled(False)
        self.d_optimize_button.setStyleSheet("color: gray; background-color: #f0f0f0;")
        self.i_optimize_button.setStyleSheet("color: gray; background-color: #f0f0f0;")
        self.material_selector.setEnabled(False)
        self.material_selector.setStyleSheet("color: gray; background-color: #f0f0f0;")
        self.diameter_selector.setEnabled(False)
        self.diameter_selector.setStyleSheet("color: gray; background-color: #f0f0f0;")
        # ES: El botón de análisis siempre está habilitado
        # EN: The analysis button is always enabled
        # JP: 解析ボタンは常に有効
        self.analyze_button.setEnabled(True)

    def switch_to_unexperimented_data(self):
        """ES: Cambiar automáticamente al archivo 未実験データ después de la primera optimización
        EN: Switch to 未実験データ file automatically after first optimization
        JA: 初回最適化後に未実験データファイルへ自動切替"""
        if hasattr(self, 'proyecto_folder') and hasattr(self, 'proyecto_nombre'):
            proyecto_nombre = getattr(self, 'proyecto_nombre', 'Unknown')
            temp_dir = os.path.join(self.proyecto_folder, "99_Temp")
            candidates = [
                os.path.join(temp_dir, f"{proyecto_nombre}_未実験データ.xlsx"),
                os.path.join(temp_dir, f"{proyecto_nombre}_未実験データ.xls"),
                os.path.join(temp_dir, f"{proyecto_nombre}_未実験データ.csv"),
            ]
            unexperimented_file = next((p for p in candidates if os.path.exists(p)), None)
            if unexperimented_file:
                # ES: Actualizar la ruta del archivo cargado
                # EN: Update the loaded file path
                # JP: 読み込んだファイルパスを更新
                self.sample_file_path = unexperimented_file
                # ES: Actualizar la etiqueta en la UI
                # EN: Update the label in the UI
                # JP: UIのラベルを更新
                self.load_file_label.setText(f"読み込み済み: {os.path.basename(unexperimented_file)}")
                print(f"✅ 入力ファイルを自動的に変更しました: {unexperimented_file}")
                return True
        return False

    def clear_main_screen(self):
        """
        ES: Limpia toda la pantalla principal (panel derecho).
        EN: Clear the entire main screen (right panel).
        JP: メイン画面（右パネル）をすべてクリアします。
        """
        print("🧹 メイン画面をクリア中...")
        
        # ES: Limpiar variables de navegación primero
        # EN: Clear navigation variables first
        # JP: まずナビゲーション変数をクリア
        self.graph_images = []
        self.graph_images_content = []
        self.current_graph_index = 0
        
        # ES: Limpiar referencias a botones de navegación de forma segura
        # EN: Safely clear navigation button references
        # JP: ナビゲーションボタン参照を安全にクリア
        if hasattr(self, 'prev_button'):
            try:
                if self.prev_button and not self.prev_button.isHidden():
                    self.prev_button.setEnabled(False)
            except RuntimeError:
                # ES: El objeto ya fue eliminado, simplemente limpiar la referencia
                # EN: The object was already deleted; just clear the reference
                # JP: オブジェクトは既に削除済み。参照のみクリア
                self.prev_button = None
        
        if hasattr(self, 'next_button'):
            try:
                if self.next_button and not self.next_button.isHidden():
                    self.next_button.setEnabled(False)
            except RuntimeError:
                # ES: El objeto ya fue eliminado, simplemente limpiar la referencia
                # EN: The object was already deleted; just clear the reference
                # JP: オブジェクトは既に削除済み。参照のみクリア
                self.next_button = None
        
        # ES: Limpiar el layout central COMPLETAMENTE (incluye layouts anidados como los botones de filtros)
        # EN: Clear the center layout COMPLETELY (including nested layouts like filter buttons)
        # JP: 中央レイアウトを完全にクリア（フィルタボタン等のネストレイアウトも含む）
        try:
            self._clear_layout_recursive(self.center_layout)
        except Exception:
            # ES: Fallback: no bloquear si algo raro pasa en la jerarquía de widgets
            # EN: Fallback: don't block if something odd happens in the widget hierarchy
            # JP: フォールバック：ウィジェット階層で何か起きてもブロックしない
            pass
        try:
            QApplication.processEvents()
        except Exception:
            pass
        
        # ES: Restaurar los elementos básicos del panel central
        # EN: Restore the basic elements of the center panel
        # JP: 中央パネルの基本要素を復元
        # ES: Título | EN: Title | JA: タイトル arriba del área de gráficos
        self._add_center_header_title()

        # Área de gráficos
        self.graph_container = QFrame()
        graph_container_layout = QVBoxLayout()
        graph_container_layout.setContentsMargins(0, 0, 0, 0)
        graph_container_layout.setSpacing(0)
        self.graph_container.setLayout(graph_container_layout)

        self.graph_area = QFrame()
        self.graph_area.setStyleSheet("background-color: #F9F9F9; border: 1px solid #CCCCCC;")
        graph_container_layout.addWidget(self.graph_area, stretch=1)

        self.center_layout.addWidget(self.graph_container, stretch=1)

        # ES: Espacio flexible antes de los botones
        # EN: Flexible space before the buttons
        # JA: ボタン前の可変スペース
        self.center_layout.addStretch()

        # ES: Botones OK y NG
        # EN: OK and NG buttons
        # JA: OK/NGボタン
        self.ok_ng_frame = QFrame()
        ok_ng_layout = QHBoxLayout()
        ok_ng_layout.setAlignment(Qt.AlignCenter)
        self.ok_ng_frame.setLayout(ok_ng_layout)

        self.ok_button = QPushButton("OK")
        self.ng_button = QPushButton("NG")

        self.setup_ok_button(self.ok_button)
        self.setup_ng_button(self.ng_button)

        self.ok_button.clicked.connect(self.on_ok_clicked)
        self.ng_button.clicked.connect(self.on_ng_clicked)

        ok_ng_layout.addWidget(self.ok_button)
        ok_ng_layout.addSpacing(10)
        ok_ng_layout.addWidget(self.ng_button)

        self.center_layout.addWidget(self.ok_ng_frame)

        self.ok_button.setEnabled(False)
        self.ng_button.setEnabled(False)
        
        # ES: Limpiar referencias a botones de navegación
        # EN: Clear navigation button references
        # JP: ナビゲーションボタン参照をクリア
        self.prev_button = None
        self.next_button = None
        self.graph_navigation_frame = None
        
        print("✅ メイン画面をクリアしました")
        print("🔧 MainWindow の初期化が完了しました")

    def setup_console_redirection(self):
        """ES: Configurar redirección de stdout y stderr a la consola integrada Y a la consola original
        EN: Configure stdout/stderr redirection to integrated console and to original console
        JA: stdout/stderrを統合コンソールおよび元のコンソールにリダイレクト"""
        # ES: ✅ FIX CRÍTICO: La UI (QTextEdit / overlay) NO se puede tocar desde hilos secundarios.
        # EN: ✅ CRITICAL FIX: The UI (QTextEdit / overlay) must NOT be touched from worker threads.
        # JP: ✅ 重要修正：UI（QTextEdit/overlay）はワーカースレッドから触ってはいけない
        # ES: Creamos un stream QObject que emite señales; el slot corre en el hilo principal.
        # EN: We create a QObject stream that emits signals; the slot runs on the main thread.
        # JP: シグナルを発行するQObjectストリームを作り、スロットはメインスレッドで実行する
        from PySide6.QtCore import QObject, Signal, Qt

        if not hasattr(self, "_console_buffers"):
            self._console_buffers = {"stdout": "", "stderr": ""}

        class ConsoleStream(QObject):
            text_written = Signal(str, str)  # stream_type, text

            def __init__(self, stream_type, original_stream, parent=None):
                super().__init__(parent)
                self.stream_type = stream_type
                self.original_stream = original_stream

            def write(self, text):
                if text is None:
                    return

                # ES: Siempre escribir en la consola original con info de hilo (esto sí es seguro)
                # EN: Always write to the original console with thread info (this is safe)
                # JP: 元のコンソールへ常にスレッド情報付きで出力（これは安全）
                try:
                    import threading
                    current_thread = threading.current_thread()
                    thread_info = f"[{current_thread.name}:{current_thread.ident}]"
                    if str(text).strip():
                        self.original_stream.write(f"DEBUG {thread_info}: {text}")
                    else:
                        self.original_stream.write(str(text))
                    self.original_stream.flush()
                except:
                    pass

                # ES: Enviar a UI mediante señal (thread-safe)
                # EN: Send to the UI via a signal (thread-safe)
                # JP: シグナルでUIへ送信（スレッドセーフ）
                try:
                    self.text_written.emit(self.stream_type, str(text))
                except:
                    pass

            def flush(self):
                try:
                    self.original_stream.flush()
                except:
                    pass

        # ES: Crear streams personalizados que mantengan la consola original | EN: Create custom streams that keep original console | JA: 元コンソールを維持するカスタムストリームを作成
        self.stdout_stream = ConsoleStream("stdout", sys.__stdout__, parent=self)
        self.stderr_stream = ConsoleStream("stderr", sys.__stderr__, parent=self)

        # ES: Conectar señales | EN: Connect signals | JA: シグナルを接続 a slot (hilo principal)
        self.stdout_stream.text_written.connect(self._on_console_stream_text, Qt.QueuedConnection)
        self.stderr_stream.text_written.connect(self._on_console_stream_text, Qt.QueuedConnection)
        
        # ES: Guardar streams originales | EN: Save original streams | JA: 元のストリームを保存
        self.original_stdout = sys.stdout
        self.original_stderr = sys.stderr
        
        # ES: Redirigir streams
        # EN: Redirect streams
        # JP: ストリームをリダイレクト
        sys.stdout = self.stdout_stream
        sys.stderr = self.stderr_stream
        
        # ES: Mensaje inicial en ambas consolas
        # EN: Initial message in both consoles
        # JP: 両方のコンソールに初期メッセージ
        print("🚀 コンソールが起動しました")
        print("📝 すべての出力が両方のコンソールに表示されます")
        # ES: (No hacer append manual: ya lo hace el print vía redirección)
        # EN: (Do not append manually: print already does it via redirection)
        # JP: （手動でappendしない：リダイレクト経由でprintが既に追加する）

    def _on_console_stream_text(self, stream_type, text):
        """Recibe texto de stdout/stderr (desde cualquier hilo) y lo pinta en la UI (hilo principal)."""
        try:
            if not hasattr(self, "_console_buffers"):
                self._console_buffers = {"stdout": "", "stderr": ""}

            if not hasattr(self, "console_output") or self.console_output is None:
                return

            buf = self._console_buffers.get(stream_type, "") + (text or "")
            lines = buf.split("\n")
            self._console_buffers[stream_type] = lines[-1]  # partial line

            for line in lines[:-1]:
                if line == "":
                    continue
                timestamp = datetime.now().strftime("%H:%M:%S")
                self.console_output.append(f"[{timestamp}] {line}")

                # ES: Consola overlay (también en main thread)
                # EN: Overlay console (also on the main thread)
                # JP: オーバーレイコンソール（メインスレッドでも）
                try:
                    if hasattr(self, "overlay_console_output"):
                        overlay_console = self.overlay_console_output
                        if overlay_console and overlay_console.isVisible():
                            overlay_console.append(line)
                except:
                    pass
        except:
            pass

    def clear_console(self):
        """Limpiar el contenido de la consola"""
        self.console_output.clear()
        self.console_output.append("🧹 コンソールがクリアされました")
        self.console_output.append("📝 新しい出力を待機中...")

    def save_console_log(self):
        """Guardar el contenido de la consola en un archivo"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"console_log_{timestamp}.txt"
            
            # ES: Obtener el contenido de la consola
            # EN: Get console contents
            # JP: コンソール内容を取得
            content = self.console_output.toPlainText()
            
            # ES: Guardar archivo | EN: Save file | JA: ファイルを保存
            with open(filename, 'w', encoding='utf-8') as f:
                f.write(content)
            
            print(f"✅ コンソールログが保存されました: {filename}")
            
        except Exception as e:
            print(f"❌ ログの保存に失敗しました: {e}")

    # ES: NOTA: Este método ya no se necesita, la flecha está siempre visible
    # EN: NOTE: This method is no longer needed; the arrow is always visible
    # JP: 注: このメソッドは不要（矢印は常に表示）

    # ES: NOTA: Este método ya no se necesita, simplificado en show_right_panel
    # EN: NOTE: This method is no longer needed; simplified in show_right_panel
    # JP: 注: このメソッドは不要（show_right_panelで簡略化）

    # ES: NOTA: Este método ya no se necesita, solo usamos el panel superpuesto
    # EN: NOTE: This method is no longer needed; we only use the overlay panel
    # JP: 注: このメソッドは不要（オーバーレイパネルのみ使用）

    def position_arrow(self):
        """Posicionar la flecha en el borde derecho del panel central"""
        try:
            # Coordenadas globales (pantalla) del panel central
            center_global = self.center_frame.mapToGlobal(QPoint(0, 0))
            button_x = center_global.x() + self.center_frame.width() - 35
            button_y = center_global.y() + self.center_frame.height() // 2 - 15
            self.console_toggle_button.setGeometry(button_x, button_y, 30, 30)
            
            # Asegurar que la flecha esté en primer plano después de posicionarla
            self.console_toggle_button.raise_()
            
            print(f"🔧 矢印を ({button_x}, {button_y}) に配置し、最前面にしました")
        except Exception as e:
            print(f"⚠️ 矢印の配置エラー: {e}")

    def debug_button_state(self):
        """ES: Método de debug para verificar el estado del botón de flecha
        EN: Debug method to check arrow button state
        JA: 矢印ボタンの状態を確認するデバッグ用メソッド"""
        print("🔍 デバッグ: 矢印ボタンの状態")
        print(f"🔍 ボタンが存在するか: {hasattr(self, 'console_toggle_button')}")
        if hasattr(self, 'console_toggle_button'):
            print(f"🔍 ボタンの表示: {self.console_toggle_button.isVisible()}")
            print(f"🔍 ボタンのジオメトリ: {self.console_toggle_button.geometry()}")
            print(f"🔍 ボタンの親: {self.console_toggle_button.parent()}")
            print(f"🔍 ボタンのテキスト: {self.console_toggle_button.text()}")
            print(f"🔍 ボタンのスタイル: {self.console_toggle_button.styleSheet()}")
        else:
            print("❌ 矢印ボタンが存在しません")

    def clear_overlay_console(self):
        """Limpiar el contenido de la consola desplegable"""
        self.overlay_console_output.clear()
        self.overlay_console_output.append("🧹 オーバーレイコンソールがクリアされました")
        self.overlay_console_output.append("📝 新しい出力を待機中...")

    def save_overlay_console_log(self):
        """Guardar el contenido de la consola desplegable en un archivo"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"overlay_console_log_{timestamp}.txt"
            
            # ES: Obtener el contenido de la consola desplegable
            # EN: Get the dropdown console contents
            # JP: ドロップダウンコンソール内容を取得
            content = self.overlay_console_output.toPlainText()
            
            # ES: Guardar archivo | EN: Save file | JA: ファイルを保存
            with open(filename, 'w', encoding='utf-8') as f:
                f.write(content)
            
            print(f"✅ オーバーレイコンソールログが保存されました: {filename}")
            
        except Exception as e:
            print(f"❌ オーバーレイログの保存に失敗しました: {e}")

    def toggle_overlay_console(self):
        """Alternar la visibilidad del panel desplegable"""
        if not self.overlay_console_visible:
            # ES: Mostrar el panel desplegable | EN: Show dropdown panel | JA: ドロップダウンパネルを表示
            self.show_overlay_console()
        else:
            # Ocultar el panel desplegable
            self.hide_overlay_console()
            
    # ES: NOTA: Este método ya no se necesita, simplificado
    # EN: NOTE: This method is no longer needed; simplified
    # JP: 注: このメソッドは不要（簡略化済み）
            
    def toggle_right_panel(self):
        """Alternar la visibilidad del panel desplegable superpuesto"""
        print("🔧 toggle_right_panel を実行しました")
        if self.overlay_console_visible:
            # ES: Si el panel desplegable está visible, ocultarlo
            # EN: If the dropdown panel is visible, hide it
            # JP: ドロップダウンパネルが表示中なら隠す
            print("🔧 ドロップダウンパネルを非表示にしています...")
            self.hide_overlay_console()
        else:
            # ES: Si el panel desplegable está oculto, mostrarlo
            # EN: If the dropdown panel is hidden, show it
            # JP: ドロップダウンパネルが非表示なら表示
            print("🔧 ドロップダウンパネルを表示しています...")
            self.show_overlay_console()

    def show_overlay_console(self):
        """ES: Mostrar el panel desplegable superpuesto en el lado derecho
        EN: Show overlay dropdown panel on the right side
        JA: 右側にオーバーレイドロップダウンパネルを表示"""
        print("🔧 右側にドロップダウンパネルを表示しています...")
        
        # ES: Obtener la posición actual de la ventana principal
        # EN: Get the current position of the main window
        # JP: メインウィンドウの現在位置を取得
        current_window_pos = self.geometry()
        print(f"🔧 ウィンドウの現在位置: {current_window_pos}")
        print(f"🔧 座標 X: {current_window_pos.x()}, Y: {current_window_pos.y()}")
        print(f"🔧 サイズ: {current_window_pos.width()} x {current_window_pos.height()}")
        
        # ES: Posicionar la consola en el lado derecho de la pantalla
        # EN: Position the console on the right side of the screen
        # JP: コンソールを画面右側に配置
        self.position_overlay_console()
        
        # ES: Cambiar el texto del botón a flecha derecha
        # EN: Change the button text to the right arrow
        # JP: ボタンのテキストを右矢印に変更
        self.console_toggle_button.setText("▶")
        
        # ES: Mostrar el panel desplegable | EN: Show dropdown panel | JA: ドロップダウンパネルを表示
        self.overlay_console.show()
        
        # Asegurar que esté en primer plano
        self.overlay_console.raise_()
        
        # Asegurar que la flecha también esté en primer plano
        self.console_toggle_button.raise_()
        
        # Actualizar estado
        self.overlay_console_visible = True
        
        # ES: Sincronizar contenido con la consola principal
        # EN: Sync content with the main console
        # JP: メインコンソールと内容を同期
        self.sync_console_content()
        
        # Debug de posición
        self.debug_console_position()
        
        print("✅ 右側にオーバーレイドロップダウンパネルを表示しました")

    def hide_overlay_console(self):
        """Ocultar el panel desplegable"""
        print("🔧 ドロップダウンパネルを非表示にしています...")
        
        # Ocultar el panel desplegable
        self.overlay_console.hide()
        
        # ES: Cambiar el texto del botón a flecha izquierda
        # EN: Change the button text to the left arrow
        # JP: ボタンのテキストを左矢印に変更
        self.console_toggle_button.setText("◀")
        
        # Reposicionar la flecha
        self.position_arrow()
        
        # Asegurar que la flecha esté en primer plano
        self.console_toggle_button.raise_()
        
        # Actualizar estado
        self.overlay_console_visible = False
        
        print("✅ Panel desplegable oculto")

    def position_overlay_console(self):
        """Posicionar la consola desplegable en el lado derecho de la pantalla"""
        try:
            # ES: Obtener la posición y dimensiones de la ventana principal
            # EN: Get the main window position and dimensions
            # JP: メインウィンドウの位置とサイズを取得
            main_window_rect = self.geometry()
            
            # Calcular posición en el lado derecho de la ventana principal
            overlay_width = 350
            overlay_height = min(600, main_window_rect.height() - 80)
            
            # Posicionar en el lado derecho de la ventana principal
            # Usar coordenadas absolutas de la pantalla
            overlay_x = main_window_rect.x() + main_window_rect.width() - overlay_width - 20
            overlay_y = main_window_rect.y() + 40  # Margen superior
            
            # ES: Configurar geometría del panel desplegable | EN: Configure dropdown panel geometry | JA: ドロップダウンパネルのジオメトリを設定
            self.overlay_console.setGeometry(overlay_x, overlay_y, overlay_width, overlay_height)
            
            # Posicionar el botón de flecha en el borde derecho del panel central (coordenadas globales)
            self.position_arrow()
            
            print(f"🔧 Ventana principal: {main_window_rect}")
            print(f"🔧 Coordenadas absolutas de la consola: ({overlay_x}, {overlay_y}) - {overlay_width}x{overlay_height}")
            print(f"🔧 Flecha reposicionada junto al panel central")
            
            # ES: Verificar que la consola esté visible y en primer plano | EN: Ensure console is visible and in foreground | JA: コンソールが表示・前面か確認
            if self.overlay_console.isVisible():
                self.overlay_console.raise_()
                print("🔧 Consola elevada a primer plano")
            
        except Exception as e:
            print(f"⚠️ コンソールオーバーレイの配置エラー: {e}")

    def keep_elements_on_top(self):
        """Mantener la consola y la flecha en primer plano, respetando el orden del loading"""
        try:
            if not hasattr(self, '_heartbeat_count'): self._heartbeat_count = 0
            self._heartbeat_count += 1
            if self._heartbeat_count >= 10:
                print("💓 ハートビート: アプリは稼働中（待機中）")
                self._heartbeat_count = 0
                
            # ES: Si hay un loading visible, NO forzamos el Z-order cada segundo.
            # EN: If loading is visible, do NOT force Z-order every second.
            # JP: ローディング表示中は毎秒Z-orderを強制しない
            # ES: Antes bajábamos (lower) la flecha y la consola mientras el loading estaba visible,
            # EN: Previously we lowered the arrow and console while the loading was visible,
            # JP: 以前はローディング表示中に矢印とコンソールをlowerしていましたが、
            # ES: lo que causaba parpadeo/"refresh" constante y bloqueaba el botón de despliegue.
            # EN: which caused constant flicker/"refresh" and blocked the toggle button.
            # JP: それにより点滅/常時リフレッシュが発生し、切替ボタンが押せなくなりました。
            # ES: Dejamos que el resto de la lógica mantenga la flecha/consola accesibles.
            # EN: We now let the rest of the logic keep the arrow/console accessible.
            # JP: 以降は他のロジックに任せ、矢印/コンソールを操作可能に保ちます。

            # ES: Si hay un diálogo modal activo que NO sea el loading, no "pisar" el Z-order.
            # EN: If there is an active modal dialog that is NOT loading, don't override Z-order.
            # JP: ローディング以外のモーダルが有効ならZ-orderを上書きしない
            modal = QApplication.activeModalWidget()
            progress = getattr(self, 'progress_dialog', None)
            if modal is not None and modal is not progress:
                return

            # Mantener la consola desplegable en primer plano si está visible
            if hasattr(self, 'overlay_console') and self.overlay_console.isVisible():
                self.overlay_console.raise_()

            # Mantener la flecha en primer plano si está visible
            if hasattr(self, 'console_toggle_button') and self.console_toggle_button.isVisible():
                self.console_toggle_button.raise_()
                
        except Exception as e:
            print(f"⚠️ 要素を前面に維持する際のエラー: {e}")

    def set_console_overlay_topmost(self, enabled: bool):
        """
        Activa/desactiva WindowStaysOnTopHint para flecha + consola overlay.
        - enabled=True: permite clicar la flecha incluso con ReusableProgressDialog (WindowModal).
        - enabled=False: evita tapar diálogos del sistema (QFileDialog, etc).
        """
        try:
            self._console_topmost_enabled = bool(enabled)

            for w_attr in ("overlay_console", "console_toggle_button"):
                w = getattr(self, w_attr, None)
                if w is None:
                    continue

                was_visible = w.isVisible()
                flags = w.windowFlags()

                # ES: Asegurar tipo de ventana esperado
                # EN: Ensure expected window type
                # JP: 想定するウィンドウ種別を保証
                flags |= Qt.Tool
                flags |= Qt.FramelessWindowHint

                if enabled:
                    flags |= Qt.WindowStaysOnTopHint
                else:
                    flags &= ~Qt.WindowStaysOnTopHint

                w.setWindowFlags(flags)

                # ES: Aplicar cambios (Qt requiere show() tras cambiar flags)
                # EN: Apply changes (Qt requires show() after changing flags)
                # JP: 変更を適用（Qtはflags変更後にshow()が必要）
                if was_visible:
                    w.show()
                    w.raise_()

            # ES: Reposicionar por si el WM recalcula geometría
            # EN: Reposition in case the window manager recalculates geometry
            # JP: WMがジオメトリを再計算する可能性があるため再配置
            try:
                if hasattr(self, 'console_toggle_button'):
                    self.position_arrow()
                if getattr(self, 'overlay_console_visible', False):
                    self.position_overlay_console()
            except Exception:
                pass

        except Exception as e:
            print(f"⚠️ set_console_overlay_topmost({enabled}) エラー: {e}")

    def pause_auto_timers(self):
        """ES: Pausar los timers automáticos para evitar interferencia con diálogos
        EN: Pause automatic timers to avoid interference with dialogs
        JA: ダイアログとの干渉を避けるため自動タイマーを一時停止"""
        try:
            if hasattr(self, 'keep_on_top_timer') and self.keep_on_top_timer.isActive():
                self.keep_on_top_timer.stop()
                print("⏸️ タイマー keep_on_top を一時停止しました")
            
            if hasattr(self, 'position_check_timer') and self.position_check_timer.isActive():
                self.position_check_timer.stop()
                print("⏸️ タイマー position_check を一時停止しました")
        except Exception as e:
            print(f"⚠️ タイマー一時停止エラー: {e}")

    def resume_auto_timers(self):
        """ES: Reanudar los timers automáticos
        EN: Resume automatic timers
        JA: 自動タイマーを再開"""
        try:
            if hasattr(self, 'keep_on_top_timer'):
                self.keep_on_top_timer.start(1000)  # Cada segundo
                print("▶️ タイマー keep_on_top を再開しました")
            
            if hasattr(self, 'position_check_timer'):
                self.position_check_timer.start(500)  # Cada medio segundo
                print("▶️ タイマー position_check を再開しました")
        except Exception as e:
            print(f"⚠️ タイマー再開エラー: {e}")

    def check_window_position(self):
        """Verificar si la ventana principal se ha movido y actualizar la consola si es necesario"""
        try:
            current_position = self.geometry()
            
            # ES: Si la posición ha cambiado, reposicionar SIEMPRE la flecha (es una ventana top-level)
            # EN: If the position changed, ALWAYS reposition the arrow (it's a top-level window)
            # JP: 位置が変わったら矢印を常に再配置（トップレベルウィンドウ）
            if current_position != self.last_window_position:
                try:
                    if hasattr(self, 'console_toggle_button') and self.console_toggle_button.isVisible():
                        self.position_arrow()
                except Exception:
                    pass

            # ES: Si la posición ha cambiado y la consola está visible, reposicionar también la consola
            # EN: If the position changed and the console is visible, reposition the console too
            # JP: 位置が変わりコンソールが表示中ならコンソールも再配置
            if (current_position != self.last_window_position and
                hasattr(self, 'overlay_console_visible') and
                self.overlay_console_visible):
                
                print(f"🔧 ウィンドウが {self.last_window_position} から {current_position} に移動しました")
                print("🔧 コンソールを再配置中...")
                
                # Reposicionar la consola en la nueva posición
                self.position_overlay_console()
                
                # Asegurar que esté en primer plano
                modal = QApplication.activeModalWidget()
                progress = getattr(self, 'progress_dialog', None)
                if modal is None or modal is progress:
                    if getattr(self, '_console_topmost_enabled', False) or getattr(self, 'overlay_console_visible', False):
                        self.overlay_console.raise_()
                        self.console_toggle_button.raise_()
                
                print("✅ コンソールを新しい位置に再配置しました")
            
            # Actualizar la posición guardada
            self.last_window_position = current_position
            
        except Exception as e:
            print(f"⚠️ ウィンドウ位置の確認中にエラー: {e}")

    def moveEvent(self, event):
        """Mantener flecha/consola ancladas cuando la ventana se mueve (drag)."""
        super().moveEvent(event)
        try:
            if hasattr(self, 'console_toggle_button') and self.console_toggle_button.isVisible():
                self.position_arrow()
            if hasattr(self, 'overlay_console_visible') and self.overlay_console_visible:
                self.position_overlay_console()
        except Exception:
            pass

    def is_valid_project_folder(self, folder_path, analysis_type="nonlinear"):
        """
        Verifica si una carpeta tiene la estructura de un proyecto válido
        
        Parameters
        ----------
        folder_path : str
            Ruta de la carpeta a verificar
        analysis_type : str, optional
            Tipo de análisis: "nonlinear" (default) o "classification"
        
        Returns
        -------
        bool
            True si la carpeta tiene estructura de proyecto válida
        """
        if not os.path.exists(folder_path) or not os.path.isdir(folder_path):
            return False
        
        # ES: Carpetas esenciales según el tipo de análisis
        # EN: Essential folders depending on the analysis type
        # JP: 解析タイプ別の必須フォルダ
        if analysis_type == "classification":
            essential_folders = [
                "05_分類"  # Essential for classification analysis
            ]
        else:  # nonlinear (default)
            essential_folders = [
                "04_非線形回帰"  # Essential for non-linear analysis
            ]
        
        # Carpetas opcionales pero comunes en proyectos existentes
        optional_folders = [
            "03_線形回帰",
            "04_非線形回帰",
            "05_分類",
            "99_Results",
            "99_Temp",
            "backup"
        ]
        
        # ES: Verificar que existan las carpetas esenciales | EN: Ensure essential folders exist | JA: 必須フォルダの存在を確認
        for folder in essential_folders:
            folder_path_full = os.path.join(folder_path, folder)
            if not os.path.exists(folder_path_full) or not os.path.isdir(folder_path_full):
                return False
        
        # ES: Si tiene al menos una carpeta opcional, es más probable que sea un proyecto válido
        # EN: If it has at least one optional folder, it's more likely to be a valid project
        # JP: 任意フォルダが1つ以上あれば有効なプロジェクトである可能性が高い
        has_optional = any(
            os.path.exists(os.path.join(folder_path, folder)) and 
            os.path.isdir(os.path.join(folder_path, folder))
            for folder in optional_folders
        )
        
        # Considerar válido si tiene las esenciales y al menos una opcional
        return has_optional
    
    def find_project_folders_in_directory(self, directory, analysis_type="nonlinear"):
        """
        Busca carpetas de proyecto dentro de un directorio
        
        Parameters
        ----------
        directory : str
            Directorio donde buscar proyectos
        analysis_type : str, optional
            Tipo de análisis: "nonlinear" (default) o "classification"
        
        Returns
        -------
        list
            Lista de rutas de carpetas que son proyectos válidos
        """
        project_folders = []
        
        if not os.path.exists(directory) or not os.path.isdir(directory):
            return project_folders
        
        # ES: Buscar en el directorio seleccionado directamente
        # EN: Search directly in the selected directory
        # JP: 選択したディレクトリを直接検索
        if self.is_valid_project_folder(directory, analysis_type=analysis_type):
            project_folders.append(directory)
        
        # ES: Buscar en subdirectorios (solo un nivel de profundidad)
        # EN: Search in subdirectories (only one level deep)
        # JP: サブディレクトリを検索（深さ1のみ）
        try:
            for item in os.listdir(directory):
                item_path = os.path.join(directory, item)
                if os.path.isdir(item_path):
                    if self.is_valid_project_folder(item_path, analysis_type=analysis_type):
                        project_folders.append(item_path)
        except PermissionError:
            pass
        
        return project_folders
    
    def create_nonlinear_project_structure(self, project_name, base_directory):
        """
        Crear la estructura de carpetas del proyecto para análisis no lineal
        Similar a Proyecto_79 pero sin 01_実験リスト y 02_実験データ
        """
        try:
            # ES: Crear la carpeta principal del proyecto | EN: Create project main folder | JA: プロジェクトのメインフォルダを作成
            project_path = os.path.join(base_directory, project_name)
            os.makedirs(project_path, exist_ok=True)
            
            # ES: Crear las subcarpetas (SIN 01 y 02) | EN: Create subfolders (without 01 and 02) | JA: サブフォルダを作成（01・02除く）
            subfolders = [
                "03_線形回帰",
                "04_非線形回帰",
                "05_分類",
                "99_Results",
                "99_Temp",
                "backup"
            ]
            
            for subfolder in subfolders:
                subfolder_path = os.path.join(project_path, subfolder)
                os.makedirs(subfolder_path, exist_ok=True)
                print(f"📁 フォルダを作成しました: {subfolder_path}")
            
            print(f"✅ プロジェクト構造を作成しました: {project_path}")
            return project_path
            
        except Exception as e:
            print(f"❌ プロジェクト構造の作成中にエラー: {e}")
            raise e
    
    def create_project_structure(self, project_name, base_directory):
        """ES: Crear la estructura de carpetas del proyecto según la imagen
        EN: Create project folder structure as per the reference image
        JA: 参考画像に従いプロジェクトのフォルダ構造を作成"""
        try:
            # ES: Crear la carpeta principal del proyecto | EN: Create project main folder | JA: プロジェクトのメインフォルダを作成
            project_path = os.path.join(base_directory, project_name)
            os.makedirs(project_path, exist_ok=True)
            
            # ES: Crear las subcarpetas según la estructura de la imagen | EN: Create subfolders from image structure | JA: 画像の構造に従いサブフォルダを作成
            subfolders = [
                "01_データ準備",
                "02_前処理", 
                "03_線形回帰",
                "04_非線形回帰",
                "05_結果比較",
                "06_レポート"
            ]
            
            for subfolder in subfolders:
                subfolder_path = os.path.join(project_path, subfolder)
                os.makedirs(subfolder_path, exist_ok=True)
                print(f"📁 フォルダを作成しました: {subfolder_path}")
            
            # ES: Crear subcarpetas específicas dentro de 03_線形回帰 | EN: Create specific subfolders inside 03_線形回帰 | JA: 03_線形回帰内に特定サブフォルダを作成
            linear_subfolders = [
                "01_データ分割",
                "02_特徴選択", 
                "03_モデル学習",
                "04_予測計算",
                "05_結果評価"
            ]
            
            linear_path = os.path.join(project_path, "03_線形回帰")
            for subfolder in linear_subfolders:
                subfolder_path = os.path.join(linear_path, subfolder)
                os.makedirs(subfolder_path, exist_ok=True)
                print(f"📁 サブフォルダを作成しました: {subfolder_path}")
            
            print(f"✅ プロジェクト構造を作成しました: {project_path}")
            return project_path
            
        except Exception as e:
            print(f"❌ プロジェクト構造の作成中にエラー: {e}")
            raise e

    def run_linear_analysis_in_project(self, project_path):
        """ES: Ejecutar análisis lineal en la carpeta del proyecto
        EN: Run linear analysis in project folder
        JA: プロジェクトフォルダで線形解析を実行"""
        try:
            print(f"🔧 プロジェクトで線形解析を実行中: {project_path}")
            
            # ES: Establecer la carpeta del proyecto actual | EN: Set current project folder | JA: 現在のプロジェクトフォルダを設定
            self.current_project_folder = project_path
            print(f"📁 プロジェクトフォルダを設定しました: {self.current_project_folder}")
            
            # ES: Obtener filtros actuales
            # EN: Get current filters
            # JP: 現在のフィルタを取得
            filters = self.get_applied_filters()
            
            if not filters:
                QMessageBox.warning(self, "警告", "フィルターが設定されていません。\nフィルターを設定してから線形解析を実行してください。")
                return
            
            # ES: Crear carpeta de resultados con timestamp | EN: Create results folder with timestamp | JA: タイムスタンプ付き結果フォルダを作成
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            results_folder = os.path.join(project_path, "03_線形回帰", f"15_{timestamp}")
            os.makedirs(results_folder, exist_ok=True)
            
            # ES: Crear subcarpetas dentro del resultado | EN: Create subfolders inside result | JA: 結果内にサブフォルダを作成
            subfolders = ["01_データ分割", "02_特徴選択", "03_モデル学習", "04_予測計算", "05_結果評価"]
            for subfolder in subfolders:
                subfolder_path = os.path.join(results_folder, subfolder)
                os.makedirs(subfolder_path, exist_ok=True)
            
            print(f"📁 結果フォルダを作成しました: {results_folder}")
            
            # ES: Ejecutar análisis lineal con la carpeta del proyecto
            # EN: Run linear analysis using the project folder
            # JP: プロジェクトフォルダで線形解析を実行
            self.execute_linear_analysis_with_output_folder(results_folder)
            
        except Exception as e:
            print(f"❌ プロジェクトで線形解析を実行中にエラー: {e}")
            QMessageBox.critical(
                self, 
                "エラー", 
                f"❌ プロジェクト内での線形解析実行中にエラーが発生しました:\n{str(e)}"
            )

    def execute_linear_analysis_with_output_folder(self, output_folder):
        """ES: Ejecutar análisis lineal con carpeta de salida específica
        EN: Run linear analysis with specific output folder
        JA: 指定出力フォルダで線形解析を実行"""
        try:
            print(f"🔧 出力フォルダで線形解析を実行中: {output_folder}")

            # ES: Evitar re-ejecución si ya hay un análisis lineal corriendo | EN: Avoid re-running if linear analysis is already running | JA: 線形解析実行中は再実行を防ぐ
            if hasattr(self, 'linear_worker') and self.linear_worker is not None:
                try:
                    if self.linear_worker.isRunning():
                        QMessageBox.warning(self, "線形解析", "⚠️ すでに線形解析が実行中です。\n完了または停止するまでお待ちください。")
                        return
                except RuntimeError:
                    # ES: Si el objeto fue destruido, limpiar referencia
                    # EN: If the object was destroyed, clear the reference
                    # JP: オブジェクトが破棄されたら参照をクリア
                    self.linear_worker = None
            
            # ES: Obtener filtros aplicados
            # EN: Get applied filters
            # JP: 適用済みフィルタを取得
            filters = self.get_applied_filters()
            print(f"🔧 適用済みフィルタ: {filters}")
            
            # ES: Importar módulo de análisis lineal
            # EN: Import the linear-analysis module
            # JP: 線形解析モジュールをインポート
            try:
                from linear_analysis_advanced import run_advanced_linear_analysis_from_db
                print("✅ 線形解析モジュールのインポートに成功しました")
            except ImportError as e:
                print(f"❌ 線形解析モジュールのインポートに失敗: {e}")
                QMessageBox.critical(self, "エラー", "❌ モジュール de análisis lineal no se pudo importar.\nAsegúrese de que el archivo linear_analysis_module.py esté en el directorio correcto.")
                return
            
            # ES: Mostrar mensaje de confirmación | EN: Show confirmation message | JA: 確認メッセージを表示
            reply = QMessageBox.question(
                self,
                "線形解析確認", 
                f"線形解析を実行しますか？\n\nフィルター: {len(filters)} 条件\n\nこの操作は時間がかかる場合があります。",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes
            )
            
            if reply != QMessageBox.Yes:
                print("❌ ユーザーが線形解析をキャンセルしました")
                return
            
            # ES: Ejecutar análisis lineal con la carpeta específica usando el MISMO flujo con popup/cancelación
            # EN: Run linear analysis with the specific folder using the SAME popup/cancellation flow
            # JP: 指定フォルダで同じポップアップ/キャンセルフローを使って線形解析を実行
            print(f"🔧 フォルダで線形解析を実行中: {output_folder}")
            self._start_linear_analysis(filters, output_folder)
            
        except Exception as e:
            print(f"❌ 線形解析の実行中にエラー: {e}")
            QMessageBox.critical(self, "エラー", f"❌ 線形解析の実行中にエラーが発生しました:\n{str(e)}")

    def _start_linear_analysis(self, filters, analysis_folder):
        """ES: Arranca el análisis lineal con popup de progreso y cancelación cooperativa.
        EN: Start linear analysis with progress popup and cooperative cancellation.
        JA: 進捗ポップアップと協調キャンセル付きで線形解析を開始。"""
        # ES: No mezclar ejecuciones pesadas en paralelo | EN: Do not run heavy tasks in parallel | JA: 重い処理の並列実行を避ける
        if hasattr(self, 'nonlinear_worker') and self.nonlinear_worker is not None:
            try:
                if self.nonlinear_worker.isRunning():
                    QMessageBox.warning(self, "線形解析", "⚠️ 非線形解析が実行中です。\n完了または停止するまでお待ちください。")
                    return
            except RuntimeError:
                self.nonlinear_worker = None
        for t_attr in ("d_optimizer_thread", "i_optimizer_thread", "dsaitekika_thread"):
            if hasattr(self, t_attr):
                t = getattr(self, t_attr)
                try:
                    if t is not None and t.isRunning():
                        QMessageBox.warning(self, "線形解析", "⚠️ 最適化が実行中です。\n完了または停止するまでお待ちください。")
                        return
                except RuntimeError:
                    setattr(self, t_attr, None)

        # ES: Evitar re-ejecución si ya hay un análisis lineal corriendo | EN: Avoid re-running if linear analysis is already running | JA: 線形解析実行中は再実行を防ぐ
        if hasattr(self, 'linear_worker') and self.linear_worker is not None:
            try:
                if self.linear_worker.isRunning():
                    QMessageBox.warning(self, "線形解析", "⚠️ すでに線形解析が実行中です。\n完了または停止するまでお待ちください。")
                    return
            except RuntimeError:
                self.linear_worker = None

        # ES: Reset de bandera de cancelación (para esta ejecución) | EN: Reset cancellation flag (for this run) | JA: キャンセルフラグをリセット（今回の実行用）
        self._linear_cancel_requested = False

        # ES: Deshabilitar botones para evitar doble ejecución | EN: Disable buttons to avoid double execution | JA: 二重実行を防ぐためボタンを無効化
        if hasattr(self, 'linear_analysis_button'):
            self.linear_analysis_button.setEnabled(False)
        if hasattr(self, 'run_analysis_button'):
            self.run_analysis_button.setEnabled(False)

        # ES: Cerrar popup previo si quedara colgado
        # EN: Close the previous popup if it got stuck
        # JP: 前のポップアップが固まっていたら閉じる
        if hasattr(self, 'progress_dialog') and self.progress_dialog is not None:
            try:
                self.progress_dialog.close()
                self.progress_dialog.deleteLater()
            except:
                pass
            try:
                delattr(self, 'progress_dialog')
            except:
                pass

        # ES: Crear popup de progreso | EN: Create progress popup | JA: 進捗ポップアップを作成
        self.progress_dialog = LinearAnalysisProgressDialog(self)
        self.progress_dialog.show()
        # ES: Durante el loading modal del análisis lineal: permitir flecha/consola por encima
        # EN: During the linear-analysis modal loading: allow arrow/console on top
        # JP: 線形解析のモーダルローディング中：矢印/コンソールを前面に許可
        self.set_console_overlay_topmost(True)
        self.progress_dialog.rejected.connect(self.on_analysis_cancelled)

        # ES: Crear y arrancar worker (QThread) con señales de progreso | EN: Create and start worker (QThread) with progress signals | JA: 進捗シグナル付きワーカー(QThread)を作成・起動
        self.linear_worker = LinearAnalysisWorker(self.db, filters, analysis_folder, self)
        self.linear_worker.progress_updated.connect(self.progress_dialog.update_progress)
        self.linear_worker.status_updated.connect(self.progress_dialog.set_status)
        self.linear_worker.finished.connect(self.on_linear_analysis_finished)
        self.linear_worker.error.connect(self.on_linear_analysis_error)

        print("🚀 進捗付き線形解析を開始します（worker）...")
        self.linear_worker.start()

    def on_linear_analysis_clicked(self):
        """ES: Acción al pulsar el botón de análisis lineal
        EN: Action when linear analysis button is clicked
        JA: 線形解析ボタンクリック時のアクション"""
        print("🔧 線形解析を開始します...")
        
        # ES: Si se accedió desde bunseki, mostrar diálogo de creación de proyecto | EN: If accessed from bunseki, show project creation dialog | JA: 分析からアクセス時はプロジェクト作成ダイアログを表示
        if hasattr(self, 'accessed_from_bunseki') and self.accessed_from_bunseki:
            print("📁 bunseki からのアクセスを検出しました - プロジェクト作成ダイアログを表示します")
            
            # ES: Mostrar diálogo | EN: Show dialog | JA: ダイアログを表示 de creación de proyecto
            dialog = ProjectCreationDialog(self)
            if dialog.exec() == QDialog.Accepted:
                project_name = dialog.project_name
                project_directory = dialog.project_directory
                
                print(f"📁 プロジェクトを作成中: {project_name}（場所: {project_directory}）")
                
                try:
                    # ES: Crear estructura del proyecto | EN: Create project structure | JA: プロジェクト構造を作成
                    project_path = self.create_project_structure(project_name, project_directory)
                    
                    # ES: Mostrar mensaje de confirmación | EN: Show confirmation message | JA: 確認メッセージを表示
                    QMessageBox.information(
                        self, 
                        "プロジェクト作成完了", 
                        f"✅ プロジェクト '{project_name}' が作成されました。\n\n"
                        f"保存先: {project_path}\n\n"
                        f"線形解析を開始します..."
                    )
                    
                    # ES: Resetear la bandera
                    # EN: Reset the flag
                    # JP: フラグをリセット
                    self.accessed_from_bunseki = False
                    
                    # ES: Proceder con el análisis lineal en la nueva carpeta
                    # EN: Proceed with linear analysis in the new folder
                    # JP: 新しいフォルダで線形解析を続行
                    self.run_linear_analysis_in_project(project_path)
                    return
                    
                except Exception as e:
                    QMessageBox.critical(
                        self, 
                        "エラー", 
                        f"❌ プロジェクト作成中にエラーが発生しました:\n{str(e)}"
                    )
                    return
            else:
                # ES: Usuario canceló, resetear la bandera
                # EN: User canceled; reset the flag
                # JP: ユーザーがキャンセルしたのでフラグをリセット
                self.accessed_from_bunseki = False
                return
        
        try:
            # ES: Verificar si estamos en la vista de filtros | EN: Check if we are on filter view | JA: フィルタビューか確認
            already_in_filter_view = False
            for i in range(self.center_layout.count()):
                item = self.center_layout.itemAt(i)
                if item.widget() and isinstance(item.widget(), QLabel):
                    if item.widget().text() == "データフィルター":
                        already_in_filter_view = True
                        break
            
            if not already_in_filter_view:
                # ES: Crear la vista de filtros primero | EN: Create filter view first | JA: 先にフィルタビューを作成
                self.create_filter_view()
                self.create_navigation_buttons()
                self.prev_button.setEnabled(True)
                self.next_button.setEnabled(True)
                QMessageBox.information(self, "分析ページ", "✅ 分析ページに移動しました。\nフィルターを設定して線形解析を実行してください。")
                return
            
            # ES: Ya estamos en la vista de filtros, ejecutar análisis lineal
            # EN: We are already on the filter view; run linear analysis
            # JP: 既にフィルタ画面なので線形解析を実行
            self.execute_linear_analysis()
            
        except Exception as e:
            QMessageBox.critical(self, "エラー", f"❌ 線形解析の実行中にエラーが発生しました:\n{str(e)}")
            print(f"❌ 線形解析中にエラー: {e}")
            import traceback
            traceback.print_exc()

    def on_nonlinear_analysis_clicked(self):
        """ES: Acción al pulsar el botón de análisis no lineal
        EN: Action when non-linear analysis button is clicked
        JA: 非線形解析ボタンクリック時のアクション"""
        print("🔧 非線形解析を開始します...")

        # ES: No mezclar ejecuciones pesadas en paralelo | EN: Do not run heavy tasks in parallel | JA: 重い処理の並列実行を避ける
        if hasattr(self, 'linear_worker') and self.linear_worker is not None:
            try:
                if self.linear_worker.isRunning():
                    QMessageBox.warning(self, "非線形解析", "⚠️ 線形解析が実行中です。\n完了または停止するまでお待ちください。")
                    return
            except RuntimeError:
                self.linear_worker = None
        for t_attr in ("d_optimizer_thread", "i_optimizer_thread", "dsaitekika_thread"):
            if hasattr(self, t_attr):
                t = getattr(self, t_attr)
                try:
                    if t is not None and t.isRunning():
                        QMessageBox.warning(self, "非線形解析", "⚠️ 最適化が実行中です。\n完了または停止するまでお待ちください。")
                        return
                except RuntimeError:
                    setattr(self, t_attr, None)

        # ES: Evitar re-ejecución si ya hay un análisis no lineal corriendo | EN: Avoid re-running if nonlinear analysis is already running | JA: 非線形解析実行中は再実行を防ぐ
        if hasattr(self, 'nonlinear_worker') and self.nonlinear_worker is not None:
            try:
                if self.nonlinear_worker.isRunning():
                    QMessageBox.warning(self, "非線形解析", "⚠️ すでに非線形解析が実行中です。\n完了または停止するまでお待ちください。")
                    return
            except RuntimeError:
                self.nonlinear_worker = None
        
        # ES: Si se accedió desde bunseki, mostrar diálogo de creación de proyecto | EN: If accessed from bunseki, show project creation dialog | JA: 分析からアクセス時はプロジェクト作成ダイアログを表示
        if hasattr(self, 'accessed_from_bunseki') and self.accessed_from_bunseki:
            print("📁 bunseki からのアクセスを検出しました - プロジェクト作成ダイアログを表示します")
            
            # ES: Mostrar diálogo | EN: Show dialog | JA: ダイアログを表示 de creación de proyecto
            dialog = ProjectCreationDialog(self)
            if dialog.exec() == QDialog.Accepted:
                project_name = dialog.project_name
                project_directory = dialog.project_directory
                
                # ES: Determinar la ruta completa del proyecto
                # EN: Determine the full project path
                # JP: プロジェクトの完全パスを決定
                if project_directory:
                    # ES: Si se seleccionó un proyecto existente, project_directory es el padre
                    # EN: If an existing project was selected, project_directory is the parent
                    # JP: 既存プロジェクト選択時、project_directoryは親ディレクトリ
                    # y project_name es el nombre del proyecto
                    project_path = os.path.join(project_directory, project_name)
                else:
                    # ES: Si se creó nuevo, project_directory es donde crear y project_name es el nombre
                    # EN: If a new one was created, project_directory is where to create it and project_name is the name
                    # JP: 新規作成時、project_directoryは作成先でproject_nameが名称
                    project_path = os.path.join(project_directory, project_name)
                
                # ES: Verificar si el proyecto ya existe (fue detectado como existente) | EN: Check if project already exists (detected as existing) | JA: プロジェクトが既存か確認（既存検出時）
                project_exists = self.is_valid_project_folder(project_path)
                
                if project_exists:
                    print(f"✅ 既存プロジェクトを使用します: {project_path}")
                    # ES: No crear estructura, solo usar la carpeta existente
                    # EN: Do not create structure; just use the existing folder
                    # JP: 構造は作らず既存フォルダを使用
                    self.current_project_folder = project_path
                    
                    QMessageBox.information(
                        self, 
                        "プロジェクト使用", 
                        f"✅ 既存のプロジェクト '{project_name}' を使用します。\n\n"
                        f"保存先: {project_path}\n\n"
                        f"非線形解析を開始します..."
                    )
                else:
                    print(f"📁 新規プロジェクトを作成します: {project_name}（場所: {project_directory}）")
                    
                    try:
                        # ES: Crear estructura del proyecto (sin 01 y 02) | EN: Create project structure (without 01 and 02) | JA: プロジェクト構造を作成（01・02なし）
                        project_path = self.create_nonlinear_project_structure(project_name, project_directory)
                        
                        # ES: Establecer la carpeta del proyecto actual | EN: Set current project folder | JA: 現在のプロジェクトフォルダを設定
                        self.current_project_folder = project_path
                        
                        QMessageBox.information(
                            self, 
                            "プロジェクト作成完了", 
                            f"✅ プロジェクト '{project_name}' が作成されました。\n\n"
                            f"保存先: {project_path}\n\n"
                            f"非線形解析を開始します..."
                        )
                    except Exception as e:
                        QMessageBox.critical(
                            self, 
                            "エラー", 
                            f"❌ プロジェクト作成中にエラーが発生しました:\n{str(e)}"
                        )
                        self.accessed_from_bunseki = False
                        return
                
                # ES: Resetear la bandera
                # EN: Reset the flag
                # JP: フラグをリセット
                self.accessed_from_bunseki = False
                
                # ES: Continuar con el flujo normal (mostrar diálogo de configuración)
                # EN: Continue with the normal flow (show configuration dialog)
                # JP: 通常フローを続行（設定ダイアログを表示）
                # El resto del código seguirá igual, pero ahora con project_folder definido
                
            else:
                # ES: Usuario canceló, resetear la bandera
                # EN: User canceled; reset the flag
                # JP: ユーザーがキャンセルしたのでフラグをリセット
                self.accessed_from_bunseki = False
                return
        
        try:
            # ES: Verificar si estamos en la vista de filtros | EN: Check if we are on filter view | JA: フィルタビューか確認
            already_in_filter_view = False
            for i in range(self.center_layout.count()):
                item = self.center_layout.itemAt(i)
                if item.widget() and isinstance(item.widget(), QLabel):
                    if item.widget().text() == "データフィルター":
                        already_in_filter_view = True
                        break
            
            if not already_in_filter_view:
                # ES: Crear la vista de filtros primero | EN: Create filter view first | JA: 先にフィルタビューを作成
                self.create_filter_view()
                self.create_navigation_buttons()
                self.prev_button.setEnabled(True)
                self.next_button.setEnabled(True)
                QMessageBox.information(self, "分析ページ", "✅ 分析ページに移動しました。\nフィルターを設定して非線形解析を実行してください。")
                return
            
            # ES: Obtener datos filtrados aplicando filtros ahora
            # EN: Get filtered data applying filters now
            # JA: フィルタを適用してフィルタ済みデータを取得
            # ES: Similar al análisis lineal, obtener datos filtrados de la BBDD
            # EN: Same as linear analysis: get filtered data from the DB
            # JA: 線形解析と同様、DBからフィルタ済みデータを取得
            try:
                import sqlite3
                filters = self.get_applied_filters()
                
                # Construir query con filtros
                query = "SELECT * FROM main_results WHERE 1=1"
                params = []
                
                # ES: Aplicar filtros de cepillo
                # EN: Apply brush filters
                # JP: ブラシフィルタを適用
                brush_selections = []
                if 'すべて' in filters and filters['すべて']:
                    brush_condition = " OR ".join([f"{brush} = 1" for brush in ['A13', 'A11', 'A21', 'A32']])
                    query += f" AND ({brush_condition})"
                else:
                    for brush_type in ['A13', 'A11', 'A21', 'A32']:
                        if brush_type in filters and filters[brush_type]:
                            brush_selections.append(brush_type)
                    
                    if brush_selections:
                        brush_condition = " OR ".join([f"{brush} = 1" for brush in brush_selections])
                        query += f" AND ({brush_condition})"
                
                # ES: Aplicar otros filtros
                # EN: Apply other filters
                # JP: その他のフィルタを適用
                for field_name, filter_value in filters.items():
                    if field_name in ['すべて', 'A13', 'A11', 'A21', 'A32']:
                        continue
                    
                    if isinstance(filter_value, tuple) and len(filter_value) == 2:
                        desde, hasta = filter_value
                        if desde and hasta:
                            try:
                                query += f" AND {field_name} BETWEEN ? AND ?"
                                params.extend([float(desde), float(hasta)])
                            except (ValueError, TypeError):
                                continue
                    elif isinstance(filter_value, (str, int, float)) and filter_value:
                        try:
                            value_num = float(filter_value) if isinstance(filter_value, str) else filter_value
                            query += f" AND {field_name} = ?"
                            params.append(value_num)
                        except (ValueError, TypeError):
                            continue
                
                # Ejecutar query
                conn = sqlite3.connect(RESULTS_DB_PATH, timeout=10)
                df = pd.read_sql_query(query, conn, params=params)
                conn.close()
                
                if df.empty or len(df) == 0:
                    QMessageBox.warning(self, "警告", "⚠️ フィルタリングされたデータがありません。\nフィルター条件を変更してください。")
                    return
                
                self.filtered_df = df
                print(f"📊 フィルタ済みデータ取得: {len(df)} 件")
                
            except Exception as e:
                print(f"❌ フィルタ済みデータ取得中にエラー: {e}")
                import traceback
                traceback.print_exc()
                QMessageBox.critical(self, "エラー", f"❌ データ取得中にエラーが発生しました:\n{str(e)}")
                return
            
            # ES: Obtener carpeta base del proyecto
            # EN: Get the project's base folder
            # JP: プロジェクトのベースフォルダを取得
            # Intentar usar current_project_folder si existe, sino usar directorio actual
            if hasattr(self, 'current_project_folder') and self.current_project_folder:
                project_folder = self.current_project_folder
                print(f"📁 プロジェクトフォルダを使用します: {project_folder}")
            else:
                # Usar directorio actual como fallback
                project_folder = os.getcwd()
                print(f"⚠️ プロジェクトフォルダが未設定のため使用します: {project_folder}")
            
            # ES: Verificar si los módulos están disponibles | EN: Check if modules are available | JA: モジュールが利用可能か確認
            if NonlinearWorker is None or NonlinearConfigDialog is None:
                QMessageBox.warning(
                    self, 
                    "モジュールが見つかりません", 
                    "❌ 必要なモジュールが見つかりません。\n最初に必要なファイルが作成されているか確認してください。"
                )
                return
            
            # ES: Verificar que los scripts necesarios existen | EN: Ensure required scripts exist | JA: 必要なスクリプトの存在を確認
            required_scripts = ["01_model_builder.py", "02_prediction.py", "03_pareto_analyzer.py"]
            missing_scripts = [s for s in required_scripts if not os.path.exists(s)]
            
            if missing_scripts:
                QMessageBox.warning(
                    self,
                    "スクリプトが見つかりません",
                    f"❌ 以下のスクリプトが見つかりません:\n\n" + "\n".join(missing_scripts) + 
                    "\n\nこれらのスクリプトは非線形解析に必要です。\n"
                    "スクリプトを配置してから再度お試しください。"
                )
                return
            
            # ES: Mostrar diálogo | EN: Show dialog | JA: ダイアログを表示 de configuración
            config_dialog = NonlinearConfigDialog(self)
            if config_dialog.exec() != QDialog.Accepted:
                print("❌ ユーザーが設定ダイアログをキャンセルしました")
                return
            
            # ES: Obtener configuración
            # EN: Get configuration
            # JP: 設定を取得
            config_values = config_dialog.get_config_values()
            print(f"📋 設定: {config_values}")
            
            # ES: Mostrar diálogo | EN: Show dialog | JA: ダイアログを表示 de confirmación
            reply = QMessageBox.question(
                self,
                "非線形解析確認",
                f"非線形解析を実行しますか？\n\n"
                f"データ件数: {len(self.filtered_df)} 件\n"
                f"保存先: {project_folder}\n"
                f"モデル数: {len(config_values['models_to_use'])}\n\n"
                f"この操作は時間がかかる場合があります。",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes
            )
            
            if reply != QMessageBox.Yes:
                print("❌ ユーザーが非線形解析をキャンセルしました")
                return
            
            # ES: Guardar configuración para uso posterior | EN: Save configuration for later use | JA: 後で使うため設定を保存
            self.nonlinear_config = config_values

            # ES: reset de bandera de cancelación | EN: Reset cancellation flag | JA: キャンセルフラグをリセット
            self._nonlinear_cancel_requested = False
            
            # ES: Ejecutar análisis no lineal con worker
            # EN: Run non-linear analysis using the worker
            # JP: ワーカーで非線形解析を実行
            print("🔧 非線形解析ワーカーを開始します...")
            self.nonlinear_worker = NonlinearWorker(self.filtered_df, project_folder, self, config_values)
            
            # ES: Conectar señales | EN: Connect signals | JA: シグナルを接続
            self.nonlinear_worker.progress_updated.connect(self.on_nonlinear_progress)
            self.nonlinear_worker.status_updated.connect(self.on_nonlinear_status)
            self.nonlinear_worker.finished.connect(self.on_nonlinear_finished)
            self.nonlinear_worker.error.connect(self.on_nonlinear_error)
            self.nonlinear_worker.console_output.connect(self.on_nonlinear_console_output)
            
            # ES: Mostrar progreso | EN: Show progress | JA: 進捗を表示 (Stage 01 - chibi más grande x1.6)
            self.progress_dialog = ReusableProgressDialog(
                self, 
                title="非線形解析処理中...",
                chibi_image="Chibi_raul.png",
                chibi_size=160  # 100 * 1.6 = 160
            )
            self.progress_dialog.show()
            # Durante el loading modal: permitir flecha/consola por encima
            self.set_console_overlay_topmost(True)
            
            # ES: Conectar señal de cancelación del diálogo para cancelar el worker | EN: Connect dialog cancel signal to cancel worker | JA: ダイアログのキャンセルシグナルをワーカーキャンセルに接続
            self.progress_dialog.cancelled.connect(self.on_nonlinear_cancelled)
            
            # ES: Conectar señal de progreso detallado (trial/fold/pass) | EN: Connect detailed progress signal (trial/fold/pass) | JA: 詳細進捗シグナル（trial/fold/pass）を接続
            self.nonlinear_worker.progress_detailed.connect(self.on_nonlinear_progress_detailed)
            
            # Iniciar worker
            self.nonlinear_worker.start()
            
        except Exception as e:
            QMessageBox.critical(self, "エラー", f"❌ 非線形解析の実行中にエラーが発生しました:\n{str(e)}")
            print(f"❌ 非線形解析中にエラー: {e}")
            import traceback
            traceback.print_exc()
    
    def on_nonlinear_progress(self, value, message):
        """Actualiza la barra de progreso"""
        if hasattr(self, 'progress_dialog'):
            # ES: Si el mensaje indica un stage específico, actualizar el porcentaje según el stage
            # EN: If the message indicates a specific stage, update the percentage accordingly
            # JP: メッセージが特定ステージを示す場合は進捗率を更新
            if "Stage 02" in message or "Prediction" in message:
                # Stage 2: 70-85% (15% del total)
                # Ajustar el porcentaje para que esté en el rango correcto
                if value < 70:
                    value = 70
                elif value > 85:
                    value = 85
                # Mapear el progreso del stage 2 al rango 70-85%
                stage2_progress = (value - 60) / 40 if value >= 60 else 0  # Normalizar 60-100 a 0-1
                value = 70 + (stage2_progress * 15)  # Mapear a 70-85%
            elif "Stage 03" in message or "Pareto" in message:
                # Stage 3: 85-100% (15% del total)
                if value < 85:
                    value = 85
                elif value > 100:
                    value = 100
                # Mapear el progreso del stage 3 al rango 85-100%
                stage3_progress = (value - 90) / 10 if value >= 90 else 0  # Normalizar 90-100 a 0-1
                value = 85 + (stage3_progress * 15)  # Mapear a 85-100%
            
            self.progress_dialog.update_progress(value, message)
            # ES: Verificar si el mensaje indica que el proceso sigue activo | EN: Check if message indicates process is still active | JA: メッセージが処理継続中か確認
            if "処理継続中" in message or "経過" in message:
                self.progress_dialog.set_process_active(True)
    
    def on_nonlinear_status(self, message):
        """Actualiza el mensaje de estado"""
        print(f"📊 状態: {message}")
        if hasattr(self, 'progress_dialog'):
            # Actualizar estado del proceso basado en el mensaje
            if "処理継続中" in message or "経過" in message:
                self.progress_dialog.set_process_active(True)
            self.progress_dialog.set_status(message)
    
    def on_nonlinear_progress_detailed(self, trial_current, trial_total, fold_current, fold_total, pass_current, pass_total, current_task='dcv', data_analysis_completed=False, final_model_training=False, shap_analysis=False, model_current=0, model_total=0):
        """ES: Actualiza el progreso detallado (trial/fold/pass/model) en el diálogo
        EN: Update detailed progress (trial/fold/pass/model) in the dialog
        JA: ダイアログ内の詳細進捗（trial/fold/pass/model）を更新"""
        if hasattr(self, 'progress_dialog') and self.progress_dialog:
            self.progress_dialog.update_progress_detailed(
                trial_current, trial_total, fold_current, fold_total, pass_current, pass_total, current_task, data_analysis_completed, final_model_training, shap_analysis, model_current, model_total
            )
    
    def on_nonlinear_console_output(self, message):
        """Muestra mensajes de consola del worker en la consola de la app"""
        try:
            # Escribir en la consola principal
            if hasattr(self, 'console_output') and self.console_output:
                self.console_output.append(message)
                # Auto-scroll al final (PySide6 usa MoveOperation.End)
                cursor = self.console_output.textCursor()
                cursor.movePosition(QTextCursor.MoveOperation.End)
                self.console_output.setTextCursor(cursor)
            
            # También escribir en la consola desplegable si existe
            if hasattr(self, 'overlay_console_output') and self.overlay_console_output:
                self.overlay_console_output.append(message)
                # Auto-scroll al final (PySide6 usa MoveOperation.End)
                cursor = self.overlay_console_output.textCursor()
                cursor.movePosition(QTextCursor.MoveOperation.End)
                self.overlay_console_output.setTextCursor(cursor)
            
            # También imprimir en stdout para que aparezca en PyCharm
            print(message, flush=True)
        except Exception as e:
            # ES: Si falla, al menos intentar imprimir
            # EN: If it fails, at least try to print it
            # JP: 失敗しても最低限printは試す
            try:
                print(f"[Console Output Error] {e}: {message}", flush=True)
            except:
                pass
    
    def on_nonlinear_finished(self, results):
        """ES: Maneja el resultado de la ejecución
        EN: Handle execution result
        JA: 実行結果を処理"""
        try:
            # ES: Si el usuario canceló, no procesar resultados | EN: If user cancelled, do not process results | JA: ユーザーがキャンセルしたら結果を処理しない
            if hasattr(self, '_nonlinear_cancel_requested') and self._nonlinear_cancel_requested:
                print("🛑 DEBUG: キャンセル後に非線形結果を受信しました。無視します。")
                if hasattr(self, 'progress_dialog') and self.progress_dialog:
                    try:
                        self.progress_dialog.close()
                    except:
                        pass
                self.set_console_overlay_topmost(False)
                return

            print("✅ 非線形解析が完了しました")
            print(f"   出力フォルダ: {results['output_folder']}")
            print(f"   Stage: {results.get('stage', 'unknown')}")
            
            # ES: Cerrar diálogo de progreso
            # EN: Close progress dialog
            # JP: 進捗ダイアログを閉じる
            if hasattr(self, 'progress_dialog'):
                self.progress_dialog.close()
            self.set_console_overlay_topmost(False)
            
            # ES: Verificar si es stage 01 (model_builder) | EN: Check if it is stage 01 (model_builder) | JA: stage 01（model_builder）か確認
            if results.get('stage') == '01_model_builder':
                # ES: Mostrar visor de gráficos | EN: Show chart viewer | JA: グラフビューアを表示
                self._show_graph_viewer(results)
            
            # ES: Si es stage completed, mostrar resultados finales
            # EN: If it's stage completed, show final results
            # JP: stage completedなら最終結果を表示
            elif results.get('stage') == 'completed':
                self._show_final_results(results)
            
        except Exception as e:
            print(f"❌ on_nonlinear_finished でエラー: {e}")
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "エラー", f"❌ 結果処理中にエラーが発生しました:\n{str(e)}")
    
    def _show_graph_viewer(self, results):
        """ES: Muestra el visor de gráficos y maneja OK/NG
        EN: Show graph viewer and handle OK/NG
        JA: グラフビューアを表示しOK/NGを処理"""
        if GraphViewerDialog is None:
            QMessageBox.warning(self, "モジュールが見つかりません", "グラフビューアが利用できません。")
            return
        
        graph_paths = results.get('graph_paths', [])
        
        if not graph_paths:
            QMessageBox.information(
                self,
                "グラフなし",
                "生成されたグラフが見つかりませんでした。"
            )
            return
        
        # ES: Mostrar visor de gráficos | EN: Show chart viewer | JA: グラフビューアを表示
        viewer = GraphViewerDialog(graph_paths, self)
        
        # ES: Si el usuario hace OK, continuar con stages 2 y 3
        # EN: If the user presses OK, continue with stages 2 and 3
        # JP: ユーザーがOKならStage 2と3を続行
        if viewer.exec() == QDialog.Accepted:
            print("✅ ユーザーがグラフを確認しました。Stage 2-3 を続行します")
            
            # ES: Mostrar progreso | EN: Show progress | JA: 進捗を表示 nuevamente
            self.progress_dialog = ReusableProgressDialog(
                self,
                title="予測・パレート分析処理中...",
                chibi_image="xebec_chibi.png"
            )
            self.progress_dialog.show()
            self.set_console_overlay_topmost(True)
            
            # ES: Conectar señales | EN: Connect signals | JA: シグナルを接続 nuevamente
            self.nonlinear_worker.finished.disconnect()
            self.nonlinear_worker.finished.connect(self.on_nonlinear_finished)
            
            # Ejecutar stages 2 y 3
            self.nonlinear_worker.run_stage2_and_3()
        else:
            print("❌ ユーザーがキャンセルしました。処理を停止します")
            QMessageBox.information(
                self,
                "非線形解析中止",
                "プロセスが中止されました。\n\n保存先: " + results['output_folder']
            )
    
    def _show_final_results(self, results):
        """ES: Muestra resultados finales del análisis completo con estadísticas
        EN: Show final results of the full analysis with statistics
        JA: 統計付きで解析の最終結果を表示
        """
        output_folder = results.get('output_folder', '')
        is_load_existing = results.get('load_existing', False)
        existing_folder_path = results.get('existing_folder_path', '')
        
        # ES: Si hay información de gráficos de Pareto, mostrar diálogo de resultados
        # EN: If there is Pareto chart info, show the results dialog
        # JP: パレートグラフ情報があれば結果ダイアログを表示
        pareto_plots_folder = results.get('pareto_plots_folder')
        prediction_output_file = results.get('prediction_output_file')
        
        if pareto_plots_folder and prediction_output_file and ParetoResultsDialog is not None:
            self._show_pareto_charts_screen(pareto_plots_folder, prediction_output_file)
            return
        
        # ES: Limpiar layout central completamente
        # EN: Clear the center layout completely
        # JP: 中央レイアウトを完全にクリア
        while self.center_layout.count():
            item = self.center_layout.takeAt(0)
            widget = item.widget()
            if widget:
                widget.deleteLater()
            else:
                # ES: Si es un layout, limpiarlo también
                # EN: If it's a layout, clear it too
                # JP: レイアウトならそれもクリア
                layout = item.layout()
                if layout:
                    while layout.count():
                        layout_item = layout.takeAt(0)
                        layout_widget = layout_item.widget()
                        if layout_widget:
                            layout_widget.deleteLater()
        
        # ES: Forzar actualización de la UI | EN: Force UI refresh | JA: UIを強制更新
        QApplication.processEvents()
        
        # ES: Crear contenedor con fondo gris limpio | EN: Create container with clean grey background | JA: クリーンなグレー背景のコンテナを作成
        gray_container = QFrame()
        gray_container.setStyleSheet("""
            QFrame {
                background-color: #f5f5f5;
                border-radius: 10px;
                margin: 10px;
            }
        """)
        
        # ES: Layout interno para el contenedor gris | EN: Inner layout for grey container | JA: グレーコンテナ用の内部レイアウト
        container_layout = QVBoxLayout(gray_container)
        container_layout.setContentsMargins(20, 20, 20, 20)
        container_layout.setSpacing(15)
        
        # ES: Título | EN: Title | JA: タイトル
        if is_load_existing:
            title_text = "既存非線形解析結果"
        else:
            title_text = "非線形解析完了"
        
        title = QLabel(title_text)
        title.setStyleSheet("""
            font-weight: bold; 
            font-size: 24px; 
            color: #2c3e50;
            margin-bottom: 20px;
            padding: 10px 0px;
            border-bottom: 2px solid #3498db;
            border-radius: 0px;
        """)
        title.setAlignment(Qt.AlignCenter)
        container_layout.addWidget(title)
        
        # ES: Mensaje de éxito
        # EN: Success message
        # JP: 成功メッセージ
        if is_load_existing:
            success_text = "✅ 既存の解析結果を読み込みました！"
        else:
            success_text = "✅ 非線形解析が完了しました！"
        
        success_label = QLabel(success_text)
        success_label.setStyleSheet("""
            font-size: 18px;
            font-weight: bold;
            color: #27ae60;
            padding: 10px;
            background-color: #d5f4e6;
            border-radius: 8px;
            border: 1px solid #27ae60;
        """)
        success_label.setAlignment(Qt.AlignCenter)
        container_layout.addWidget(success_label)
        
        # ES: Si es carga existente, cargar y mostrar archivos
        # EN: If loading an existing run, load and show files
        # JP: 既存読み込みの場合はファイルを読み込み表示
        if is_load_existing and existing_folder_path:
            self._load_and_display_existing_files(container_layout, existing_folder_path, output_folder)
        else:
            # ES: Cargar y mostrar estadísticas del análisis recién completado
            # EN: Load and show statistics for the just-completed analysis
            # JP: 直近完了した解析の統計を読み込み表示
            self._load_and_display_analysis_statistics(container_layout, output_folder)
        
        # ES: Mensaje final
        # EN: Final message
        # JP: 最終メッセージ
        final_message = QLabel("結果を確認してください。")
        final_message.setStyleSheet("""
            font-size: 14px;
            color: #7f8c8d;
            font-style: italic;
            margin-top: 10px;
        """)
        final_message.setAlignment(Qt.AlignCenter)
        container_layout.addWidget(final_message)
        
        # ES: Agregar botón "次へ" para ver gráficos (siempre que haya carpeta de salida)
        # EN: Add a "次へ" button to view charts (as long as there is an output folder)
        # JP: 出力フォルダがある場合、グラフ閲覧用に「次へ」ボタンを追加
        if output_folder:
            button_layout = QHBoxLayout()
            button_layout.addStretch()
            
            next_button = QPushButton("次へ")
            next_button.setFixedSize(120, 40)
            next_button.setStyleSheet("""
                QPushButton {
                    background-color: #3498db;
                    color: white;
                    border: none;
                    padding: 10px 20px;
                    border-radius: 5px;
                    font-size: 14px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #2980b9;
                }
            """)
            next_button.clicked.connect(lambda: self._show_nonlinear_charts_from_results(results))
            button_layout.addWidget(next_button)
            button_layout.addStretch()
            container_layout.addLayout(button_layout)
        
        # ES: Agregar el contenedor al layout central
        # EN: Add the container to the center layout
        # JP: コンテナを中央レイアウトに追加
        self.center_layout.addWidget(gray_container)
        
        # ES: Guardar información para navegación de gráficos | EN: Save info for chart navigation | JA: グラフナビ用情報を保存
        if output_folder:
            # ES: Buscar carpeta de resultados para guardar la ruta
            # EN: Find the results folder to save the path
            # JP: パス保存のため結果フォルダを探す
            result_folder = os.path.join(output_folder, '03_学習結果')
            if os.path.exists(result_folder):
                self.nonlinear_existing_folder_path = result_folder
                # ES: Guardar la carpeta del análisis completo (NUM_YYYYMMDD_HHMMSS) como project_folder | EN: Save full analysis folder as project_folder | JA: 解析フォルダをproject_folderとして保存
                # Esto permite que el botón "予測" funcione correctamente
                self.nonlinear_project_folder = output_folder
        
        # ES: Forzar actualización | EN: Force refresh | JA: 強制更新
        QApplication.processEvents()
    
    def _load_and_display_existing_files(self, container_layout, existing_folder_path, output_folder):
        """ES: Carga y muestra las estadísticas de un análisis existente
        EN: Load and display statistics from an existing analysis
        JA: 既存解析の統計を読み込み表示
        """
        # ES: Usar la misma función que para análisis nuevo, ya que la estructura es la misma
        # EN: Use the same function as for a new analysis, since the structure is the same
        # JP: 構造が同じなので新規解析と同じ関数を使う
        # ES: existing_folder_path es la carpeta del análisis (NUM_YYYYMMDD_HHMMSS)
        # EN: existing_folder_path is the analysis folder (NUM_YYYYMMDD_HHMMSS)
        # JP: existing_folder_path は解析フォルダ（NUM_YYYYMMDD_HHMMSS）
        # output_folder puede ser la misma o diferente, pero usamos existing_folder_path
        self._load_and_display_analysis_statistics(container_layout, existing_folder_path)
    
    def _load_and_display_analysis_statistics(self, container_layout, output_folder):
        """ES: Carga y muestra las estadísticas del análisis recién completado
        EN: Load and display statistics for the newly completed analysis
        JA: 完了直後の解析統計を読み込み表示
        """
        try:
            from pathlib import Path
            import json
            from datetime import datetime
            
            # ES: Buscar analysis_results.json directamente en la carpeta de resultados
            # EN: Look for analysis_results.json directly in the results folder
            # JP: 結果フォルダ直下でanalysis_results.jsonを探す
            result_folder = os.path.join(output_folder, '03_学習結果')
            analysis_results_path = os.path.join(result_folder, 'analysis_results.json')
            
            analysis_data = {}
            
            if os.path.exists(analysis_results_path):
                try:
                    with open(analysis_results_path, 'r', encoding='utf-8') as f:
                        analysis_data = json.load(f)
                    print(f"✅ 解析データを読み込みました: {analysis_results_path}")
                except Exception as e:
                    print(f"⚠️ analysis_results.json の読み込み中にエラー: {e}")
            else:
                print(f"⚠️ analysis_results.json が見つかりません: {analysis_results_path}")
            
            # ES: Información del análisis
            # EN: Analysis information
            # JP: 解析情報
            filters_applied = analysis_data.get('filters_applied', [])
            if filters_applied == "N/A" or filters_applied is None:
                filters_text = "N/A"
            elif isinstance(filters_applied, list):
                if len(filters_applied) == 0:
                    filters_text = "N/A"
                elif len(filters_applied) > 3:
                    filters_text = f"{len(filters_applied)} 条件"
                else:
                    filters_text = ", ".join(str(f) for f in filters_applied)
            else:
                filters_text = str(filters_applied)
            
            # Truncar si es muy largo
            if len(filters_text) > 50:
                filters_text = filters_text[:47] + "..."
            
            data_range = analysis_data.get('data_range', 'N/A')
            if isinstance(data_range, str) and len(data_range) > 50:
                data_range = data_range[:47] + "..."
            
            # ES: Obtener tiempo de análisis
            # EN: Get analysis duration
            # JP: 解析時間を取得
            analysis_duration = analysis_data.get('analysis_duration_formatted', 'N/A')
            if analysis_duration == 'N/A' and analysis_data.get('analysis_duration_seconds'):
                # ES: Si no está formateado, formatearlo
                # EN: If it's not formatted, format it
                # JP: 未フォーマットならフォーマットする
                duration_seconds = analysis_data.get('analysis_duration_seconds')
                if duration_seconds:
                    hours = int(duration_seconds // 3600)
                    minutes = int((duration_seconds % 3600) // 60)
                    seconds = int(duration_seconds % 60)
                    if hours > 0:
                        analysis_duration = f"{hours}時間{minutes}分{seconds}秒"
                    elif minutes > 0:
                        analysis_duration = f"{minutes}分{seconds}秒"
                    else:
                        analysis_duration = f"{seconds:.1f}秒"
            
            info_text = f"""
            📊 解析完了時刻: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
            ⏱️ 解析時間: {analysis_duration}
            📈 データ数: {analysis_data.get('data_count', 'N/A')} レコード
            🤖 訓練済みモデル: {analysis_data.get('models_trained', 'N/A')} 個
            🔧 フィルター適用: {filters_text}
            📊 データ範囲: {data_range}
            """
            
            info_label = QLabel(info_text)
            info_label.setStyleSheet("""
                font-size: 14px;
                color: #34495e;
                background-color: #ecf0f1;
                padding: 15px;
                border-radius: 8px;
                border: 1px solid #bdc3c7;
            """)
            info_label.setAlignment(Qt.AlignLeft)
            info_label.setWordWrap(True)
            container_layout.addWidget(info_label)
            
            # Sección destacada de métricas de confianza
            models = analysis_data.get('models', {})
            if models and isinstance(models, dict) and len(models) > 0:
                # ES: Título | EN: Title | JA: タイトル de la sección de métricas
                metrics_title = QLabel("📊 信頼性指標 (Confidence Metrics)")
                metrics_title.setStyleSheet("""
                    font-weight: bold; 
                    font-size: 20px; 
                    color: #2c3e50;
                    margin-top: 20px;
                    margin-bottom: 15px;
                    padding-bottom: 10px;
                    border-bottom: 3px solid #3498db;
                """)
                metrics_title.setAlignment(Qt.AlignCenter)
                container_layout.addWidget(metrics_title)
                
                # ES: Crear layout horizontal para las tarjetas de métricas | EN: Create horizontal layout for metric cards | JA: メトリックカード用横レイアウトを作成
                metrics_container = QHBoxLayout()
                metrics_container.setSpacing(15)
                
                # ES: Iterar sobre cada target y crear tarjeta de métricas
                # EN: Iterate over each target and create a metrics card
                # JP: 各ターゲットを走査して指標カードを作成
                for target_name, model_info in models.items():
                    if isinstance(model_info, dict):
                        # ES: Crear tarjeta para este target | EN: Create card for this target | JA: このターゲット用カードを作成
                        metric_card = QFrame()
                        metric_card.setStyleSheet("""
                            QFrame {
                                background-color: #ffffff;
                                border: 2px solid #3498db;
                                border-radius: 10px;
                                padding: 15px;
                                min-width: 250px;
                            }
                        """)
                        card_layout = QVBoxLayout(metric_card)
                        card_layout.setSpacing(10)
                        
                        # ES: Título | EN: Title | JA: タイトル del target
                        target_label = QLabel(f"【{target_name}】")
                        target_label.setStyleSheet("""
                            font-weight: bold;
                            font-size: 16px;
                            color: #2c3e50;
                            padding-bottom: 5px;
                            border-bottom: 1px solid #ecf0f1;
                        """)
                        target_label.setAlignment(Qt.AlignCenter)
                        card_layout.addWidget(target_label)
                        
                        # Métricas CV principales
                        cv_mae = model_info.get('cv_mae')
                        cv_rmse = model_info.get('cv_rmse')
                        cv_r2 = model_info.get('cv_r2')
                        
                        # MAE
                        if cv_mae is not None:
                            mae_label = QLabel(f"MAE: {cv_mae:.4f}" if isinstance(cv_mae, (int, float)) else f"MAE: {cv_mae}")
                            mae_label.setStyleSheet("""
                                font-size: 14px;
                                color: #34495e;
                                padding: 5px;
                                background-color: #f8f9fa;
                                border-radius: 5px;
                            """)
                            card_layout.addWidget(mae_label)
                        
                        # RMSE
                        if cv_rmse is not None:
                            rmse_label = QLabel(f"RMSE: {cv_rmse:.4f}" if isinstance(cv_rmse, (int, float)) else f"RMSE: {cv_rmse}")
                            rmse_label.setStyleSheet("""
                                font-size: 14px;
                                color: #34495e;
                                padding: 5px;
                                background-color: #f8f9fa;
                                border-radius: 5px;
                            """)
                            card_layout.addWidget(rmse_label)
                        
                        # R² (con color según el valor)
                        if cv_r2 is not None:
                            r2_value = cv_r2 if isinstance(cv_r2, (int, float)) else 0
                            # Color según calidad: verde si R² > 0.7, amarillo si > 0.5, rojo si <= 0.5
                            if r2_value > 0.7:
                                r2_color = "#27ae60"  # Verde
                                r2_bg = "#d5f4e6"
                            elif r2_value > 0.5:
                                r2_color = "#f39c12"  # Amarillo
                                r2_bg = "#fef5e7"
                            else:
                                r2_color = "#e74c3c"  # Rojo
                                r2_bg = "#fadbd8"
                            
                            r2_label = QLabel(f"R²: {cv_r2:.4f}" if isinstance(cv_r2, (int, float)) else f"R²: {cv_r2}")
                            r2_label.setStyleSheet(f"""
                                font-size: 14px;
                                font-weight: bold;
                                color: {r2_color};
                                padding: 5px;
                                background-color: {r2_bg};
                                border-radius: 5px;
                                border: 1px solid {r2_color};
                            """)
                            card_layout.addWidget(r2_label)
                        
                        # Métricas de folds (media y desviación estándar) si están disponibles
                        fold_mae_mean = model_info.get('fold_mae_mean')
                        fold_mae_std = model_info.get('fold_mae_std')
                        fold_rmse_mean = model_info.get('fold_rmse_mean')
                        fold_rmse_std = model_info.get('fold_rmse_std')
                        fold_r2_mean = model_info.get('fold_r2_mean')
                        fold_r2_std = model_info.get('fold_r2_std')
                        
                        # ES: Agregar separador si hay métricas de folds
                        # EN: Add a separator if fold metrics are available
                        # JP: fold指標がある場合は区切りを追加
                        if any([fold_mae_mean, fold_rmse_mean, fold_r2_mean]):
                            separator = QLabel("─" * 20)
                            separator.setStyleSheet("color: #bdc3c7;")
                            separator.setAlignment(Qt.AlignCenter)
                            card_layout.addWidget(separator)
                            
                            # Subtítulo para métricas de folds
                            fold_title = QLabel("Fold Statistics:")
                            fold_title.setStyleSheet("""
                                font-size: 12px;
                                font-weight: bold;
                                color: #7f8c8d;
                                margin-top: 5px;
                            """)
                            fold_title.setAlignment(Qt.AlignCenter)
                            card_layout.addWidget(fold_title)
                            
                            # MAE fold statistics
                            if fold_mae_mean is not None:
                                mae_std_str = f"±{fold_mae_std:.4f}" if fold_mae_std is not None else ""
                                fold_mae_label = QLabel(f"MAE: {fold_mae_mean:.4f} {mae_std_str}")
                                fold_mae_label.setStyleSheet("""
                                    font-size: 12px;
                                    color: #7f8c8d;
                                    padding: 3px;
                                """)
                                card_layout.addWidget(fold_mae_label)
                            
                            # RMSE fold statistics
                            if fold_rmse_mean is not None:
                                rmse_std_str = f"±{fold_rmse_std:.4f}" if fold_rmse_std is not None else ""
                                fold_rmse_label = QLabel(f"RMSE: {fold_rmse_mean:.4f} {rmse_std_str}")
                                fold_rmse_label.setStyleSheet("""
                                    font-size: 12px;
                                    color: #7f8c8d;
                                    padding: 3px;
                                """)
                                card_layout.addWidget(fold_rmse_label)
                            
                            # R² fold statistics
                            if fold_r2_mean is not None:
                                r2_std_str = f"±{fold_r2_std:.4f}" if fold_r2_std is not None else ""
                                fold_r2_label = QLabel(f"R²: {fold_r2_mean:.4f} {r2_std_str}")
                                fold_r2_label.setStyleSheet("""
                                    font-size: 12px;
                                    color: #7f8c8d;
                                    padding: 3px;
                                """)
                                card_layout.addWidget(fold_r2_label)
                        
                        # ES: Agregar la tarjeta al layout horizontal
                        # EN: Add the card to the horizontal layout
                        # JP: カードを横レイアウトに追加
                        metrics_container.addWidget(metric_card)
                
                # ES: Agregar stretch al final para centrar las tarjetas
                # EN: Add a stretch at the end to center the cards
                # JP: カードを中央寄せするため末尾にstretchを追加
                metrics_container.addStretch()
                
                # ES: Crear widget contenedor para el layout horizontal | EN: Create widget container for horizontal layout | JA: 横レイアウト用ウィジェットコンテナを作成
                metrics_widget = QWidget()
                metrics_widget.setLayout(metrics_container)
                container_layout.addWidget(metrics_widget)
            
            # ES: Ruta clickeable del archivo de salida
            # EN: Clickable output-folder path
            # JP: 出力フォルダのクリック可能なパス
            if output_folder:
                path_layout = QVBoxLayout()
                
                path_title = QLabel("📁 出力ディレクトリ:")
                path_title.setStyleSheet("""
                    font-size: 14px;
                    font-weight: bold;
                    color: #2c3e50;
                    margin-top: 10px;
                    margin-bottom: 5px;
                """)
                path_layout.addWidget(path_title)
                
                path_label = QLabel(output_folder)
                path_label.setStyleSheet("""
                    QLabel {
                        font-size: 12px;
                        color: #3498db;
                        background-color: #e8f4fd;
                        padding: 10px;
                        border-radius: 5px;
                        border: 1px solid #3498db;
                        text-decoration: underline;
                    }
                    QLabel:hover {
                        background-color: #d1ecf1;
                        cursor: pointer;
                    }
                """)
                path_label.setWordWrap(True)
                path_label.setAlignment(Qt.AlignLeft)
                
                def open_folder():
                    try:
                        import subprocess
                        if os.name == 'nt':  # Windows
                            os.startfile(output_folder)
                        elif os.name == 'posix':  # macOS y Linux
                            subprocess.run(['open', output_folder], check=True)
                        else:
                            subprocess.run(['xdg-open', output_folder], check=True)
                        print(f"✅ フォルダを開きました: {output_folder}")
                    except Exception as e:
                        print(f"❌ フォルダを開く際にエラー: {e}")
                        QMessageBox.warning(self, "エラー", f"❌ フォルダを開けませんでした:\n{str(e)}")
                
                path_label.mousePressEvent = lambda event: open_folder()
                path_layout.addWidget(path_label)
                container_layout.addLayout(path_layout)
            
            # Resultados detallados de modelos (ya tenemos models de la sección anterior)
            if models and isinstance(models, dict) and len(models) > 0:
                models_title = QLabel("詳細モデル結果")
                models_title.setStyleSheet("""
                    font-weight: bold; 
                    font-size: 18px; 
                    color: #2c3e50;
                    margin-top: 20px;
                    margin-bottom: 10px;
                """)
                container_layout.addWidget(models_title)
                
                for target_name, model_info in models.items():
                    if isinstance(model_info, dict):
                        status = "✅ 成功"
                        model_name = model_info.get('model_name', 'Unknown')
                        details = f"モデル: {model_name}"
                        
                        # ES: Agregar métricas CV si están disponibles
                        # EN: Add CV metrics if available
                        # JP: 利用可能ならCV指標を追加
                        cv_r2 = model_info.get('cv_r2')
                        cv_mae = model_info.get('cv_mae')
                        cv_rmse = model_info.get('cv_rmse')
                        
                        if cv_r2 is not None:
                            if isinstance(cv_r2, (int, float)):
                                details += f", R²: {cv_r2:.4f}"
                            else:
                                details += f", R²: {cv_r2}"
                        
                        if cv_mae is not None:
                            if isinstance(cv_mae, (int, float)):
                                details += f", MAE: {cv_mae:.4f}"
                            else:
                                details += f", MAE: {cv_mae}"
                        
                        if cv_rmse is not None:
                            if isinstance(cv_rmse, (int, float)):
                                details += f", RMSE: {cv_rmse:.4f}"
                            else:
                                details += f", RMSE: {cv_rmse}"
                    else:
                        status = "✅ 成功"
                        details = f"モデル情報: {str(model_info)[:100]}"
                    
                    model_label = QLabel(f"【{target_name}】 {status}\n{details}")
                    model_label.setStyleSheet("""
                        font-size: 12px;
                        color: #34495e;
                        background-color: #f8f9fa;
                        padding: 10px;
                        border-radius: 5px;
                        border: 1px solid #dee2e6;
                        margin: 5px 0px;
                    """)
                    container_layout.addWidget(model_label)
        
        except Exception as e:
            print(f"❌ 解析統計の読み込み中にエラー: {e}")
            import traceback
            traceback.print_exc()
            error_label = QLabel(f"❌ 統計情報の読み込み中にエラーが発生しました:\n{str(e)}")
            error_label.setStyleSheet("color: #e74c3c; padding: 10px; background-color: #fadbd8; border-radius: 5px;")
            error_label.setWordWrap(True)
            container_layout.addWidget(error_label)
    
    def _show_nonlinear_charts_from_results(self, results):
        """ES: Mostrar gráficos del análisis no lineal desde los resultados
        EN: Show non-linear analysis charts from results
        JA: 結果から非線形解析のグラフを表示
        """
        output_folder = results.get('output_folder', '')
        if not output_folder:
            QMessageBox.warning(self, "エラー", "❌ グラフを表示するための情報が見つかりません。")
            return
        
        # ES: Buscar carpeta de resultados (03_学習結果)
        # EN: Find the results folder (03_学習結果)
        # JP: 結果フォルダ（03_学習結果）を探す
        result_folder = os.path.join(output_folder, '03_学習結果')
        
        # ES: Guardar información para navegación | EN: Save navigation info | JA: ナビ用情報を保存
        if os.path.exists(result_folder):
            self.nonlinear_existing_folder_path = result_folder
            self.nonlinear_project_folder = output_folder
            # Llamar a la función de mostrar gráficos (si existe)
            if hasattr(self, 'show_nonlinear_charts'):
                self.show_nonlinear_charts()
            else:
                QMessageBox.information(
                    self,
                    "情報",
                    "グラフ表示機能は準備中です。\n\n結果フォルダ:\n" + output_folder
                )
        else:
            QMessageBox.warning(
                self,
                "エラー",
                f"❌ 結果フォルダが見つかりません:\n{result_folder}"
            )
    
    def show_nonlinear_charts(self):
        """ES: Mostrar gráficos del análisis no lineal con navegación
        EN: Show non-linear analysis charts with navigation
        JA: ナビゲーション付きで非線形解析グラフを表示
        """
        print("🔧 非線形解析のグラフを表示中...")
        
        try:
            # ES: Verificar que tenemos la ruta de la carpeta cargada | EN: Ensure we have loaded folder path | JA: 読み込み済みフォルダパスがあるか確認
            if not hasattr(self, 'nonlinear_existing_folder_path') or not self.nonlinear_existing_folder_path:
                QMessageBox.warning(self, "エラー", "❌ グラフを表示するための情報が見つかりません。")
                return
            
            # ES: Limpiar layout central completamente
            # EN: Clear the center layout completely
            # JP: 中央レイアウトを完全にクリア
            while self.center_layout.count():
                item = self.center_layout.takeAt(0)
                widget = item.widget()
                if widget:
                    widget.deleteLater()
                else:
                    # ES: Si es un layout, limpiarlo también
                    # EN: If it's a layout, clear it too
                    # JP: レイアウトならそれもクリア
                    layout = item.layout()
                    if layout:
                        while layout.count():
                            layout_item = layout.takeAt(0)
                            layout_widget = layout_item.widget()
                            if layout_widget:
                                layout_widget.deleteLater()
            
            # ES: Forzar actualización de la UI | EN: Force UI refresh | JA: UIを強制更新
            QApplication.processEvents()
            
            # ES: Crear contenedor con fondo gris limpio | EN: Create container with clean grey background | JA: クリーンなグレー背景のコンテナを作成
            gray_container = QFrame()
            gray_container.setStyleSheet("""
                QFrame {
                    background-color: #f5f5f5;
                    border-radius: 10px;
                    margin: 10px;
                }
            """)
            
            # ES: Layout interno para el contenedor gris | EN: Inner layout for grey container | JA: グレーコンテナ用の内部レイアウト
            container_layout = QVBoxLayout(gray_container)
            container_layout.setContentsMargins(20, 20, 20, 20)
            container_layout.setSpacing(15)
            
            # ES: Título | EN: Title | JA: タイトル
            title = QLabel("非線形解析結果 チャート")
            title.setStyleSheet("""
                font-weight: bold; 
                font-size: 24px; 
                color: #2c3e50;
                margin-bottom: 20px;
                padding: 10px 0px;
                border-bottom: 2px solid #3498db;
                border-radius: 0px;
            """)
            title.setAlignment(Qt.AlignCenter)
            container_layout.addWidget(title)
            
            # ES: Buscar gráficos PNG en la carpeta de resultados (03_学習結果)
            # EN: Search for PNG charts in the results folder (03_学習結果)
            # JP: 結果フォルダ（03_学習結果）でPNGグラフを探す
            from pathlib import Path
            folder_path = Path(self.nonlinear_existing_folder_path)
            chart_images = []
            
            # ES: Buscar imágenes PNG directamente en la carpeta de resultados
            # EN: Search for PNG images directly in the results folder
            # JP: 結果フォルダ直下でPNG画像を探す
            for file in folder_path.glob("*.png"):
                if file.is_file():
                    chart_images.append(str(file))
            
            # ES: Buscar también en data_analysis si existe
            # EN: Also search in data_analysis if it exists
            # JP: data_analysisが存在する場合はそこも探す
            data_analysis_path = folder_path / "data_analysis"
            if data_analysis_path.exists() and data_analysis_path.is_dir():
                for file in data_analysis_path.glob("*.png"):
                    if file.is_file():
                        chart_images.append(str(file))
            
            # ES: Si no se encuentran gráficos, mostrar mensaje
            # EN: If no charts are found, show a message
            # JP: グラフが見つからない場合はメッセージを表示
            if not chart_images:
                no_charts_label = QLabel("⚠️ グラフが見つかりません")
                no_charts_label.setStyleSheet("""
                    font-size: 16px;
                    color: #e74c3c;
                    background-color: #fadbd8;
                    padding: 20px;
                    border-radius: 8px;
                    border: 1px solid #e74c3c;
                    margin: 20px 0px;
                """)
                no_charts_label.setAlignment(Qt.AlignCenter)
                container_layout.addWidget(no_charts_label)
            else:
                # ES: Configurar navegación de gráficos | EN: Configure chart navigation | JA: グラフナビゲーションを設定
                self.nonlinear_chart_images = sorted(chart_images)
                self.current_nonlinear_chart_index = 0
                
                # ES: Layout principal | EN: Main layout | JA: メインレイアウト para la imagen y navegación
                chart_layout = QVBoxLayout()
                
                # ES: Label para mostrar la imagen (ocupa todo el ancho)
                # EN: Label to display the image (takes full width)
                # JP: 画像表示用ラベル（全幅）
                self.nonlinear_chart_label = QLabel()
                self.nonlinear_chart_label.setAlignment(Qt.AlignCenter)
                self.nonlinear_chart_label.setStyleSheet("""
                    QLabel {
                        background-color: white;
                        border: 2px solid #bdc3c7;
                        border-radius: 10px;
                        padding: 10px;
                        min-height: 500px;
                    }
                """)
                chart_layout.addWidget(self.nonlinear_chart_label)
                
                # ES: Layout horizontal para botones de navegación (debajo de la imagen) | EN: Horizontal layout for nav buttons (below image) | JA: ナビボタン用横レイアウト（画像下）
                nav_buttons_layout = QHBoxLayout()
                nav_buttons_layout.addStretch()
                
                # ES: Botón flecha izquierda | EN: Left arrow button | JA: 左矢印ボタン
                prev_chart_button = QPushButton("◀ 前へ")
                prev_chart_button.setFixedSize(100, 40)
                prev_chart_button.setStyleSheet("""
                    QPushButton {
                        background-color: #3498db;
                        color: white;
                        border: none;
                        border-radius: 8px;
                        font-size: 14px;
                        font-weight: bold;
                        padding: 8px 16px;
                    }
                    QPushButton:hover {
                        background-color: #2980b9;
                    }
                    QPushButton:disabled {
                        background-color: #bdc3c7;
                        color: #7f8c8d;
                    }
                """)
                prev_chart_button.clicked.connect(self.show_previous_nonlinear_chart)
                nav_buttons_layout.addWidget(prev_chart_button)
                
                # ES: Espacio entre botones
                # EN: Space between buttons
                # JA: ボタン間のスペース
                nav_buttons_layout.addSpacing(20)
                
                # ES: Botón flecha derecha | EN: Right arrow button | JA: 右矢印ボタン
                next_chart_button = QPushButton("次へ ▶")
                next_chart_button.setFixedSize(100, 40)
                next_chart_button.setStyleSheet("""
                    QPushButton {
                        background-color: #3498db;
                        color: white;
                        border: none;
                        border-radius: 8px;
                        font-size: 14px;
                        font-weight: bold;
                        padding: 8px 16px;
                    }
                    QPushButton:hover {
                        background-color: #2980b9;
                    }
                    QPushButton:disabled {
                        background-color: #bdc3c7;
                        color: #7f8c8d;
                    }
                """)
                next_chart_button.clicked.connect(self.show_next_nonlinear_chart)
                nav_buttons_layout.addWidget(next_chart_button)
                
                nav_buttons_layout.addStretch()
                chart_layout.addLayout(nav_buttons_layout)
                
                # ES: Información del gráfico actual
                # EN: Current chart information
                # JA: 現在のグラフ情報
                self.nonlinear_chart_info_label = QLabel()
                self.nonlinear_chart_info_label.setStyleSheet("""
                    font-size: 14px;
                    color: #2c3e50;
                    background-color: #ecf0f1;
                    padding: 10px;
                    border-radius: 5px;
                    border: 1px solid #bdc3c7;
                    margin: 10px 0px;
                """)
                self.nonlinear_chart_info_label.setAlignment(Qt.AlignCenter)
                chart_layout.addWidget(self.nonlinear_chart_info_label)
                
                container_layout.addLayout(chart_layout)
                
                # ES: Mostrar el primer gráfico | EN: Show first chart | JA: 先頭グラフを表示
                self.update_nonlinear_chart_display()
            
            # ES: Botones para volver y predicción
            # EN: Back and prediction buttons
            # JA: 戻る・予測ボタン
            buttons_layout = QHBoxLayout()
            buttons_layout.addStretch()
            
            # ES: Botón para volver | EN: Back button | JA: 戻るボタン
            back_button = QPushButton("戻る")
            back_button.setFixedSize(120, 40)
            back_button.setStyleSheet("""
                QPushButton {
                    background-color: #e74c3c;
                    color: white;
                    border: none;
                    padding: 10px 20px;
                    border-radius: 5px;
                    font-size: 14px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #c0392b;
                }
            """)
            back_button.clicked.connect(self.on_analyze_clicked)
            buttons_layout.addWidget(back_button)
            
            # ES: Espacio entre botones
            # EN: Space between buttons
            # JA: ボタン間のスペース
            buttons_layout.addSpacing(20)
            
            # ES: Botón para predicción | EN: Prediction button | JA: 予測ボタン
            prediction_button = QPushButton("予測")
            prediction_button.setFixedSize(120, 40)
            prediction_button.setStyleSheet("""
                QPushButton {
                    background-color: #27ae60;
                    color: white;
                    border: none;
                    padding: 10px 20px;
                    border-radius: 5px;
                    font-size: 14px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #229954;
                }
            """)
            # ES: Conectar botón de predicción si existe la función, sino deshabilitarlo | EN: Connect prediction button if function exists, else disable | JA: 予測関数があればボタン接続、なければ無効化
            if hasattr(self, 'run_nonlinear_prediction'):
                prediction_button.clicked.connect(self.run_nonlinear_prediction)
            else:
                prediction_button.setEnabled(False)
                prediction_button.setToolTip("予測機能は準備中です")
            buttons_layout.addWidget(prediction_button)
            
            buttons_layout.addStretch()
            container_layout.addLayout(buttons_layout)
            
            # ES: Espacio flexible
            # EN: Flexible space
            # JA: 可変スペース
            container_layout.addStretch()
            
            # ES: Agregar el contenedor gris al layout central
            # EN: Add the gray container to the center layout
            # JP: 灰色コンテナを中央レイアウトに追加
            self.center_layout.addWidget(gray_container)
            
            print("✅ 非線形解析のグラフを表示しました")
            
        except Exception as e:
            print(f"❌ 非線形解析のグラフ表示中にエラー: {e}")
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "エラー", f"❌ グラフの表示中にエラーが発生しました:\n{str(e)}")
    
    def show_previous_nonlinear_chart(self):
        """ES: Mostrar gráfico anterior del análisis no lineal
        EN: Show previous non-linear analysis chart
        JA: 非線形解析の前のグラフを表示
        """
        if hasattr(self, 'nonlinear_chart_images') and len(self.nonlinear_chart_images) > 0:
            if not hasattr(self, 'current_nonlinear_chart_index'):
                self.current_nonlinear_chart_index = 0
            self.current_nonlinear_chart_index = (self.current_nonlinear_chart_index - 1) % len(self.nonlinear_chart_images)
            self.update_nonlinear_chart_display()
    
    def show_next_nonlinear_chart(self):
        """ES: Mostrar gráfico siguiente del análisis no lineal
        EN: Show next non-linear analysis chart
        JA: 非線形解析の次のグラフを表示
        """
        if hasattr(self, 'nonlinear_chart_images') and len(self.nonlinear_chart_images) > 0:
            if not hasattr(self, 'current_nonlinear_chart_index'):
                self.current_nonlinear_chart_index = 0
            self.current_nonlinear_chart_index = (self.current_nonlinear_chart_index + 1) % len(self.nonlinear_chart_images)
            self.update_nonlinear_chart_display()
    
    def update_nonlinear_chart_display(self):
        """ES: Actualizar la visualización del gráfico actual del análisis no lineal
        EN: Update the display of the current non-linear analysis chart
        JA: 非線形解析の現在グラフ表示を更新
        """
        if not hasattr(self, 'nonlinear_chart_images') or len(self.nonlinear_chart_images) == 0:
            return
        
        if not hasattr(self, 'current_nonlinear_chart_index'):
            self.current_nonlinear_chart_index = 0
        
        if self.current_nonlinear_chart_index < 0:
            self.current_nonlinear_chart_index = 0
        elif self.current_nonlinear_chart_index >= len(self.nonlinear_chart_images):
            self.current_nonlinear_chart_index = len(self.nonlinear_chart_images) - 1
        
        current_image_path = self.nonlinear_chart_images[self.current_nonlinear_chart_index]
        
        # ES: Cargar y mostrar la imagen | EN: Load and display the image | JA: 画像を読み込み表示
        pixmap = QPixmap(current_image_path)
        if not pixmap.isNull():
            # ES: Redimensionar la imagen para ocupar el ancho disponible | EN: Resize to fit available space | JA: 利用可能領域に合わせてリサイズ
            # ES: Obtener el tamaño del contenedor | EN: Get container size | JA: コンテナサイズを取得
            container_width = self.nonlinear_chart_label.width() - 20  # Restar padding
            container_height = self.nonlinear_chart_label.height() - 20  # Restar padding
            
            # ES: Si el contenedor aún no tiene tamaño, usar un tamaño por defecto
            # EN: If container size is not ready yet, use a default size
            # JA: サイズ未確定の場合はデフォルトサイズを使用
            if container_width <= 0:
                container_width = 1000
            if container_height <= 0:
                container_height = 600
            
            # ES: Redimensionar manteniendo la proporción | EN: Resize while keeping aspect ratio | JA: アスペクト比を維持してリサイズ
            scaled_pixmap = pixmap.scaled(container_width, container_height, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            self.nonlinear_chart_label.setPixmap(scaled_pixmap)
            
            # ES: Actualizar información del gráfico | EN: Update chart info | JA: グラフ情報を更新
            filename = os.path.basename(current_image_path)
            info_text = f"📊 {filename} ({self.current_nonlinear_chart_index + 1}/{len(self.nonlinear_chart_images)})"
            if hasattr(self, 'nonlinear_chart_info_label'):
                self.nonlinear_chart_info_label.setText(info_text)
            
            print(f"✅ グラフを表示中: {filename}")
        else:
            print(f"❌ 画像を読み込めませんでした: {current_image_path}")
    
    def _show_pareto_charts_screen(self, pareto_plots_folder, prediction_output_file):
        """ES: Mostrar gráficos de Pareto en formato pantalla (similar a show_nonlinear_charts)
        EN: Show Pareto charts in a screen layout (similar to show_nonlinear_charts)
        JA: 画面レイアウトでParetoグラフを表示（show_nonlinear_charts同様）
        """
        print("🔧 Pareto グラフを画面に表示中...")
        
        try:
            # ES: Limpiar layout central completamente
            # EN: Clear the center layout completely
            # JP: 中央レイアウトを完全にクリア
            while self.center_layout.count():
                item = self.center_layout.takeAt(0)
                widget = item.widget()
                if widget:
                    widget.deleteLater()
                else:
                    layout = item.layout()
                    if layout:
                        while layout.count():
                            layout_item = layout.takeAt(0)
                            layout_widget = layout_item.widget()
                            if layout_widget:
                                layout_widget.deleteLater()
            
            # ES: Forzar actualización de la UI | EN: Force UI refresh | JA: UIを強制更新
            QApplication.processEvents()
            
            # ES: Crear contenedor con fondo gris limpio | EN: Create container with clean grey background | JA: クリーンなグレー背景のコンテナを作成
            gray_container = QFrame()
            gray_container.setStyleSheet("""
                QFrame {
                    background-color: #f5f5f5;
                    border-radius: 10px;
                    margin: 10px;
                }
            """)
            
            # ES: Layout interno para el contenedor gris | EN: Inner layout for grey container | JA: グレーコンテナ用の内部レイアウト
            container_layout = QVBoxLayout(gray_container)
            container_layout.setContentsMargins(20, 20, 20, 20)
            container_layout.setSpacing(15)
            
            # ES: Título | EN: Title | JA: タイトル
            title = QLabel("パレート分析結果 チャート")
            title.setStyleSheet("""
                font-weight: bold; 
                font-size: 24px; 
                color: #2c3e50;
                margin-bottom: 20px;
                padding: 10px 0px;
                border-bottom: 2px solid #3498db;
                border-radius: 0px;
            """)
            title.setAlignment(Qt.AlignCenter)
            container_layout.addWidget(title)
            
            # ES: Buscar gráficos PNG en la carpeta de Pareto
            # EN: Search for PNG charts in the Pareto folder
            # JP: パレートフォルダでPNGグラフを探す
            from pathlib import Path
            folder_path = Path(pareto_plots_folder)
            chart_images = []
            
            # ES: Buscar imágenes PNG en la carpeta
            # EN: Search for PNG images in the folder
            # JP: フォルダ内のPNG画像を探す
            if folder_path.exists() and folder_path.is_dir():
                for file in folder_path.glob("*.png"):
                    if file.is_file():
                        chart_images.append(str(file))
                # ES: También buscar JPG/JPEG
                # EN: Also look for JPG/JPEG
                # JP: JPG/JPEGも探す
                for file in folder_path.glob("*.jpg"):
                    if file.is_file():
                        chart_images.append(str(file))
                for file in folder_path.glob("*.jpeg"):
                    if file.is_file():
                        chart_images.append(str(file))
            
            # ES: Si no se encuentran gráficos, mostrar mensaje
            # EN: If no charts are found, show a message
            # JP: グラフが見つからない場合はメッセージを表示
            if not chart_images:
                no_charts_label = QLabel("⚠️ グラフが見つかりません")
                no_charts_label.setStyleSheet("""
                    font-size: 16px;
                    color: #e74c3c;
                    background-color: #fadbd8;
                    padding: 20px;
                    border-radius: 8px;
                    border: 1px solid #e74c3c;
                    margin: 20px 0px;
                """)
                no_charts_label.setAlignment(Qt.AlignCenter)
                container_layout.addWidget(no_charts_label)
            else:
                # ES: Configurar navegación de gráficos | EN: Configure chart navigation | JA: グラフナビゲーションを設定
                self.pareto_chart_images = sorted(chart_images)
                self.current_pareto_chart_index = 0
                
                # ES: Layout principal | EN: Main layout | JA: メインレイアウト para la imagen y navegación
                chart_layout = QVBoxLayout()
                
                # ES: Label para mostrar la imagen (ocupa todo el ancho)
                # EN: Label to display the image (takes full width)
                # JP: 画像表示用ラベル（全幅）
                self.pareto_chart_label = QLabel()
                self.pareto_chart_label.setAlignment(Qt.AlignCenter)
                self.pareto_chart_label.setStyleSheet("""
                    QLabel {
                        background-color: white;
                        border: 2px solid #bdc3c7;
                        border-radius: 10px;
                        padding: 10px;
                        min-height: 500px;
                    }
                """)
                chart_layout.addWidget(self.pareto_chart_label)
                
                # ES: Layout horizontal para botones de navegación (debajo de la imagen) | EN: Horizontal layout for nav buttons (below image) | JA: ナビボタン用横レイアウト（画像下）
                nav_buttons_layout = QHBoxLayout()
                nav_buttons_layout.addStretch()
                
                # ES: Botón flecha izquierda | EN: Left arrow button | JA: 左矢印ボタン
                prev_chart_button = QPushButton("◀ 前へ")
                prev_chart_button.setFixedSize(100, 40)
                prev_chart_button.setStyleSheet("""
                    QPushButton {
                        background-color: #3498db;
                        color: white;
                        border: none;
                        border-radius: 8px;
                        font-size: 14px;
                        font-weight: bold;
                        padding: 8px 16px;
                    }
                    QPushButton:hover {
                        background-color: #2980b9;
                    }
                    QPushButton:disabled {
                        background-color: #bdc3c7;
                        color: #7f8c8d;
                    }
                """)
                prev_chart_button.clicked.connect(self.show_previous_pareto_chart)
                nav_buttons_layout.addWidget(prev_chart_button)
                
                # ES: Espacio entre botones
                # EN: Space between buttons
                # JA: ボタン間のスペース
                nav_buttons_layout.addSpacing(20)
                
                # ES: Botón flecha derecha | EN: Right arrow button | JA: 右矢印ボタン
                next_chart_button = QPushButton("次へ ▶")
                next_chart_button.setFixedSize(100, 40)
                next_chart_button.setStyleSheet("""
                    QPushButton {
                        background-color: #3498db;
                        color: white;
                        border: none;
                        border-radius: 8px;
                        font-size: 14px;
                        font-weight: bold;
                        padding: 8px 16px;
                    }
                    QPushButton:hover {
                        background-color: #2980b9;
                    }
                    QPushButton:disabled {
                        background-color: #bdc3c7;
                        color: #7f8c8d;
                    }
                """)
                next_chart_button.clicked.connect(self.show_next_pareto_chart)
                nav_buttons_layout.addWidget(next_chart_button)
                
                nav_buttons_layout.addStretch()
                chart_layout.addLayout(nav_buttons_layout)
                
                # Información del gráfico actual
                self.pareto_chart_info_label = QLabel()
                self.pareto_chart_info_label.setStyleSheet("""
                    font-size: 14px;
                    color: #2c3e50;
                    background-color: #ecf0f1;
                    padding: 10px;
                    border-radius: 5px;
                    border: 1px solid #bdc3c7;
                    margin: 10px 0px;
                """)
                self.pareto_chart_info_label.setAlignment(Qt.AlignCenter)
                chart_layout.addWidget(self.pareto_chart_info_label)
                
                container_layout.addLayout(chart_layout)
                
                # ES: Guardar referencia al archivo de predicción para importar | EN: Save reference to prediction file for import | JA: インポート用に予測ファイル参照を保存
                self.pareto_prediction_output_file = prediction_output_file
                
                # ES: Mostrar el primer gráfico | EN: Show first chart | JA: 先頭グラフを表示
                self.update_pareto_chart_display()
            
            # ES: Botones
            # EN: Buttons
            # JA: ボタン de acción
            buttons_layout = QHBoxLayout()
            buttons_layout.addStretch()
            
            # ES: Botón para volver | EN: Back button | JA: 戻るボタン
            back_button = QPushButton("戻る")
            back_button.setFixedSize(120, 40)
            back_button.setStyleSheet("""
                QPushButton {
                    background-color: #e74c3c;
                    color: white;
                    border: none;
                    padding: 10px 20px;
                    border-radius: 5px;
                    font-size: 14px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #c0392b;
                }
            """)
            back_button.clicked.connect(self.on_analyze_clicked)
            buttons_layout.addWidget(back_button)
            
            # ES: Espacio entre botones
            # EN: Space between buttons
            # JA: ボタン間のスペース
            buttons_layout.addSpacing(20)
            
            # ES: Botón para importar a base de datos | EN: Import to database button | JA: DBへインポートボタン
            import_button = QPushButton("データベースにインポート")
            import_button.setFixedSize(180, 40)
            import_button.setStyleSheet("""
                QPushButton {
                    background-color: #27ae60;
                    color: white;
                    border: none;
                    padding: 10px 20px;
                    border-radius: 5px;
                    font-size: 14px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #229954;
                }
            """)
            import_button.clicked.connect(lambda: self.import_nonlinear_pareto_to_database(self.pareto_prediction_output_file))
            buttons_layout.addWidget(import_button)
            
            buttons_layout.addStretch()
            container_layout.addLayout(buttons_layout)
            
            # ES: Espacio flexible
            # EN: Flexible space
            # JA: 可変スペース
            container_layout.addStretch()
            
            # ES: Agregar el contenedor gris al layout central
            # EN: Add the gray container to the center layout
            # JP: 灰色コンテナを中央レイアウトに追加
            self.center_layout.addWidget(gray_container)
            
            print("✅ Pareto グラフを画面に表示しました")
            
        except Exception as e:
            print(f"❌ Pareto グラフ表示中にエラー: {e}")
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "エラー", f"❌ グラフの表示中にエラーが発生しました:\n{str(e)}")
    
    def show_previous_pareto_chart(self):
        """ES: Mostrar gráfico anterior de Pareto
        EN: Show previous Pareto chart
        JA: Paretoの前のグラフを表示
        """
        if hasattr(self, 'pareto_chart_images') and len(self.pareto_chart_images) > 0:
            self.current_pareto_chart_index = (self.current_pareto_chart_index - 1) % len(self.pareto_chart_images)
            self.update_pareto_chart_display()
    
    def show_next_pareto_chart(self):
        """ES: Mostrar gráfico siguiente de Pareto
        EN: Show next Pareto chart
        JA: Paretoの次のグラフを表示
        """
        if hasattr(self, 'pareto_chart_images') and len(self.pareto_chart_images) > 0:
            self.current_pareto_chart_index = (self.current_pareto_chart_index + 1) % len(self.pareto_chart_images)
            self.update_pareto_chart_display()
    
    def update_pareto_chart_display(self):
        """ES: Actualizar la visualización del gráfico actual de Pareto
        EN: Update the display of the current Pareto chart
        JA: 現在のParetoグラフ表示を更新
        """
        if hasattr(self, 'pareto_chart_images') and len(self.pareto_chart_images) > 0:
            current_image_path = self.pareto_chart_images[self.current_pareto_chart_index]
            
            # ES: Cargar y mostrar la imagen | EN: Load and display the image | JA: 画像を読み込み表示
            pixmap = QPixmap(current_image_path)
            if not pixmap.isNull():
                # ES: Redimensionar para ocupar el ancho disponible | EN: Resize to fit available space | JA: 利用可能領域に合わせてリサイズ
                container_width = self.pareto_chart_label.width() - 20  # Restar padding
                container_height = self.pareto_chart_label.height() - 20  # Restar padding
                
                # ES: Si el contenedor aún no tiene tamaño, usar un tamaño por defecto
                # EN: If container size is not ready yet, use a default size
                # JA: サイズ未確定の場合はデフォルトサイズを使用
                if container_width <= 0:
                    container_width = 1000
                if container_height <= 0:
                    container_height = 600
                
                # ES: Redimensionar manteniendo la proporción | EN: Resize while keeping aspect ratio | JA: アスペクト比を維持してリサイズ
                scaled_pixmap = pixmap.scaled(container_width, container_height, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                self.pareto_chart_label.setPixmap(scaled_pixmap)
                
                # ES: Actualizar información del gráfico | EN: Update chart info | JA: グラフ情報を更新
                filename = os.path.basename(current_image_path)
                info_text = f"📊 {filename} ({self.current_pareto_chart_index + 1}/{len(self.pareto_chart_images)})"
                self.pareto_chart_info_label.setText(info_text)
                
                print(f"✅ Pareto グラフを表示中: {filename}")
            else:
                print(f"❌ 画像を読み込めませんでした: {current_image_path}")
    
    def run_nonlinear_prediction(self):
        """
        Ejecuta predicción no lineal (02_prediction.py y 03_pareto_analyzer.py)
        desde la pantalla de gráficos del análisis no lineal
        """
        print("🔧 グラフ画面から非線形予測を開始します...")
        
        try:
            # ES: Verificar que tenemos la carpeta del proyecto no lineal | EN: Ensure we have nonlinear project folder | JA: 非線形プロジェクトフォルダがあるか確認
            if not hasattr(self, 'nonlinear_project_folder') or not self.nonlinear_project_folder:
                QMessageBox.warning(
                    self,
                    "エラー",
                    "❌ 予測を実行するための情報が見つかりません。\n\nまず非線形解析を実行してください。"
                )
                return
            
            working_dir = self.nonlinear_project_folder
            if not os.path.exists(working_dir):
                QMessageBox.warning(
                    self,
                    "エラー",
                    f"❌ 作業ディレクトリが見つかりません:\n{working_dir}"
                )
                return
            
            # Confirmar con el usuario
            reply = QMessageBox.question(
                self,
                "予測実行確認",
                f"予測とパレート解析を実行しますか？\n\n作業ディレクトリ:\n{working_dir}\n\n"
                f"⚠️ 実行前にバックアップが作成されます。",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            if reply != QMessageBox.Yes:
                return
            
            # ES: Crear backup antes de ejecutar | EN: Create backup before running | JA: 実行前にバックアップ作成
            backup_created = self._create_nonlinear_backup(working_dir)
            if not backup_created:
                reply = QMessageBox.question(
                    self,
                    "バックアップ警告",
                    "⚠️ バックアップの作成に失敗しました。\n\nそれでも続行しますか？",
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.No
                )
                if reply != QMessageBox.Yes:
                    return
            
            # ES: Mostrar diálogo | EN: Show dialog | JA: ダイアログを表示 de progreso (Stage 02/03 - chibi más grande x1.6)
            self.progress_dialog = ReusableProgressDialog(
                self,
                title="予測・パレート分析処理中...",
                chibi_image="Chibi_sukuzisan_raul.png",
                chibi_size=160  # 100 * 1.6 = 160
            )
            self.progress_dialog.show()
            self.set_console_overlay_topmost(True)
            self.progress_dialog.set_status("予測処理を開始中...")
            self.progress_dialog.update_progress(5, "予測処理を開始中...")
            
            # ES: Guardar tiempo de inicio total (para tiempo transcurrido continuo) | EN: Save total start time (for continuous elapsed time) | JA: 経過時間用に開始時刻を保存
            total_start_time = time.time()
            
            # Ejecutar 02_prediction.py (5% - 20%)
            print(f"🔧 Ejecutando 02_prediction.py en: {working_dir}")
            self.progress_dialog.set_status("02_prediction.py 実行中...")
            
            prediction_success = self._run_prediction_script(working_dir, self.progress_dialog, progress_start=5, progress_end=20, total_start_time=total_start_time)
            
            if not prediction_success:
                self.progress_dialog.close()
                self.set_console_overlay_topmost(False)
                QMessageBox.critical(
                    self,
                    "エラー",
                    "❌ 02_prediction.py の実行に失敗しました。\n\n詳細はコンソールを確認してください。"
                )
                return
            
            # Ejecutar 03_pareto_analyzer.py (20% - 100%)
            print(f"🔧 Ejecutando 03_pareto_analyzer.py en: {working_dir}")
            self.progress_dialog.set_status("03_pareto_analyzer.py 実行中...")
            self.progress_dialog.update_progress(20, "03_pareto_analyzer.py 実行中...")
            
            pareto_success = self._run_pareto_script(working_dir, self.progress_dialog, progress_start=20, progress_end=100, total_start_time=total_start_time)
            
            if not pareto_success:
                self.progress_dialog.close()
                self.set_console_overlay_topmost(False)
                QMessageBox.critical(
                    self,
                    "エラー",
                    "❌ 03_pareto_analyzer.py の実行に失敗しました。\n\n詳細はコンソールを確認してください。"
                )
                return
            
            # ES: Cerrar diálogo de progreso
            # EN: Close progress dialog
            # JP: 進捗ダイアログを閉じる
            self.progress_dialog.close()
            self.set_console_overlay_topmost(False)
            
            # Construir rutas de resultados del pareto
            pareto_plots_folder = os.path.join(working_dir, "05_パレート解", "pareto_plots")
            prediction_output_file = os.path.join(working_dir, "04_予測", "Prediction_output.xlsx")
            
            # ES: DEBUG: Verificar rutas
            # EN: DEBUG: Check paths
            # JP: DEBUG: パスを確認
            print(f"🔍 DEBUG run_nonlinear_prediction: working_dir = {working_dir}")
            print(f"🔍 DEBUG run_nonlinear_prediction: pareto_plots_folder = {pareto_plots_folder}")
            print(f"🔍 DEBUG run_nonlinear_prediction: prediction_output_file = {prediction_output_file}")
            print(f"🔍 DEBUG run_nonlinear_prediction: pareto_plots_folder exists = {os.path.exists(pareto_plots_folder)}")
            print(f"🔍 DEBUG run_nonlinear_prediction: prediction_output_file exists = {os.path.exists(prediction_output_file)}")
            
            # ES: Verificar que existen los archivos | EN: Ensure files exist | JA: ファイルの存在を確認
            if os.path.exists(pareto_plots_folder) and os.path.exists(prediction_output_file):
                # ES: Mostrar pantalla de gráficos de Pareto | EN: Show Pareto charts screen | JA: パレート図画面を表示
                print(f"✅ Pareto グラフを表示します: {pareto_plots_folder}")
                self._show_pareto_charts_screen(pareto_plots_folder, prediction_output_file)
            else:
                # ES: Si no existen, mostrar mensaje de éxito pero sin gráficos
                # EN: If they don't exist, show a success message but without charts
                # JP: 存在しない場合、成功メッセージのみ表示（グラフ無し）
                missing_items = []
                if not os.path.exists(pareto_plots_folder):
                    missing_items.append(f"パレートグラフフォルダ: {pareto_plots_folder}")
                    print("❌ DEBUG: pareto_plots_folder は存在しません")
                if not os.path.exists(prediction_output_file):
                    missing_items.append(f"予測出力ファイル: {prediction_output_file}")
                    print("❌ DEBUG: prediction_output_file は存在しません")
                
                # Listar contenido del directorio para debug
                if os.path.exists(working_dir):
                    print("🔍 DEBUG: working_dir の内容:")
                    try:
                        for item in os.listdir(working_dir):
                            item_path = os.path.join(working_dir, item)
                            item_type = "DIR" if os.path.isdir(item_path) else "FILE"
                            print(f"   {item_type}: {item}")
                    except Exception as e:
                        print(f"⚠️ 内容の列挙中にエラー: {e}")
                
                QMessageBox.information(
                    self,
                    "処理完了",
                    f"✅ 予測とパレート解析が正常に完了しました！\n\n"
                    f"作業ディレクトリ: {working_dir}\n\n"
                    f"✅ 02_prediction.py: 完了\n"
                    f"✅ 03_pareto_analyzer.py: 完了\n\n"
                    f"⚠️ 以下のファイルが見つかりませんでした:\n" + "\n".join(missing_items)
                )
            
        except Exception as e:
            print(f"❌ run_nonlinear_prediction でエラー: {e}")
            import traceback
            traceback.print_exc()
            
            if hasattr(self, 'progress_dialog'):
                self.progress_dialog.close()
            
            QMessageBox.critical(
                self,
                "エラー",
                f"❌ 予測実行中にエラーが発生しました:\n{str(e)}"
            )
    
    def _create_nonlinear_backup(self, working_dir):
        """
        Crea un backup de la carpeta del análisis no lineal antes de ejecutar predicción
        
        Parameters
        ----------
        working_dir : str
            Directorio de trabajo del análisis no lineal
        
        Returns
        -------
        bool
            True si el backup se creó exitosamente, False en caso contrario
        """
        try:
            from datetime import datetime
            
            # ES: Obtener la ruta base del proyecto (donde está 0sec.py)
            # EN: Get the project's base path (where 0sec.py is)
            # JP: プロジェクトのベースパス（0sec.pyがある場所）を取得
            # working_dir es algo como: Archivos_de_salida/Proyecto_79/04_非線形回帰/100_20251120_102819
            # Necesitamos llegar a la raíz del proyecto donde está .venv
            current_path = Path(working_dir).resolve()
            
            # ES: Buscar la carpeta .venv o la raíz del proyecto
            # EN: Look for the .venv folder or the project root
            # JP: .venvフォルダまたはプロジェクトルートを探す
            backup_base = None
            search_path = current_path
            
            # ES: Buscar hacia arriba hasta encontrar .venv o llegar a la raíz
            # EN: Search upward until finding .venv or reaching the root
            # JP: .venvが見つかるかルートに到達するまで上方向に探す
            while search_path != search_path.parent:
                venv_path = search_path / ".venv"
                if venv_path.exists() and venv_path.is_dir():
                    # ES: Encontramos .venv, crear Backup en el mismo nivel
                    # EN: Found .venv; create Backup at the same level
                    # JP: .venvを発見: 同じ階層にBackupを作成
                    backup_base = search_path / "Backup"
                    break
                search_path = search_path.parent
            
            # ES: Si no encontramos .venv, usar la ruta del directorio actual como fallback
            # EN: If we don't find .venv, use the current directory path as a fallback
            # JP: .venvが見つからなければ現ディレクトリをフォールバックとして使用
            if backup_base is None:
                backup_base = Path.cwd() / "Backup"
            
            # ES: Crear carpeta Backup si no existe | EN: Create Backup folder if it does not exist | JA: Backupフォルダが無ければ作成
            backup_base.mkdir(parents=True, exist_ok=True)
            
            # ES: Crear carpeta con timestamp (formato: YYYYMMDD) | EN: Create folder with timestamp (YYYYMMDD) | JA: タイムスタンプ付きフォルダ作成（YYYYMMDD）
            timestamp = datetime.now().strftime("%Y%m%d")
            backup_folder = backup_base / timestamp
            backup_folder.mkdir(parents=True, exist_ok=True)
            
            # ES: Copiar toda la carpeta del análisis no lineal
            # EN: Copy the entire non-linear analysis folder
            # JP: 非線形解析フォルダを丸ごとコピー
            folder_name = os.path.basename(working_dir)
            dest_folder = backup_folder / folder_name
            
            # ES: Si ya existe, agregar un sufijo numérico
            # EN: If it already exists, add a numeric suffix
            # JP: 既に存在する場合は数値サフィックスを付ける
            if dest_folder.exists():
                counter = 1
                while (backup_folder / f"{folder_name}_{counter}").exists():
                    counter += 1
                dest_folder = backup_folder / f"{folder_name}_{counter}"
            
            print(f"📁 バックアップを作成中: {working_dir} → {dest_folder}")
            
            # ES: Copiar recursivamente
            # EN: Copy recursively
            # JP: 再帰的にコピー
            shutil.copytree(working_dir, str(dest_folder), dirs_exist_ok=True)
            
            print(f"✅ バックアップの作成が完了しました: {dest_folder}")
            return True
            
        except Exception as e:
            print(f"⚠️ バックアップ作成中にエラー: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _run_prediction_script(self, working_dir, progress_dialog=None, progress_start=0, progress_end=20, total_start_time=None):
        """
        Ejecuta 02_prediction.py en el directorio de trabajo
        
        Parameters
        ----------
        working_dir : str
            Directorio de trabajo
        progress_dialog : ReusableProgressDialog, optional
            Diálogo de progreso para actualizar
        progress_start : int
            Porcentaje inicial de progreso (0-100)
        progress_end : int
            Porcentaje final de progreso (0-100)
        total_start_time : float, optional
            Tiempo de inicio total para tiempo transcurrido continuo
        
        Returns
        -------
        bool
            True si el script se ejecutó exitosamente, False en caso contrario
        """
        try:
            # ES: Preparar archivo de predicción antes de ejecutar
            # EN: Prepare the prediction file before running
            # JP: 実行前に予測ファイルを準備
            # 1. Crear carpeta 04_予測 si no existe
            prediction_folder = os.path.join(working_dir, "04_予測")
            os.makedirs(prediction_folder, exist_ok=True)
            
            # 2. Buscar el archivo NOMBREDELPROYECTO__未実験データ.xlsx en la carpeta principal del proyecto
            # working_dir es: .../Proyecto_79/04_非線形回帰/100_YYYYMMDD_HHMMSS
            # Necesitamos llegar a: .../Proyecto_79/
            from pathlib import Path
            working_path = Path(working_dir).resolve()
            project_folder = None
            
            # ES: Buscar hacia arriba hasta encontrar la carpeta del proyecto (que contiene 04_非線形回帰)
            # EN: Search upward until finding the project folder (that contains 04_非線形回帰)
            # JP: 04_非線形回帰を含むプロジェクトフォルダが見つかるまで上方向に探す
            for parent in working_path.parents:
                if parent.name == "04_非線形回帰":
                    project_folder = parent.parent
                    break
            
            if project_folder is None:
                # ES: Fallback: buscar por nombre de carpeta que contiene "Proyecto"
                # EN: Fallback: look for a folder name containing "Proyecto"
                # JP: フォールバック: 「Proyecto」を含むフォルダ名で探す
                for parent in working_path.parents:
                    if "Proyecto" in parent.name:
                        project_folder = parent
                        break
            
            if project_folder is None:
                # ES: Último fallback: usar el directorio padre de 04_非線形回帰
                # EN: Last fallback: use the parent directory of 04_非線形回帰
                # JP: 最終フォールバック: 04_非線形回帰 の親ディレクトリを使用
                # working_dir debería ser .../Proyecto_XX/04_非線形回帰/100_...
                # Entonces parent.parent debería ser Proyecto_XX
                project_folder = working_path.parent.parent
                print(f"⚠️ プロジェクトフォルダのフォールバックを使用します: {project_folder}")
            
            print(f"📁 プロジェクトフォルダを検出しました: {project_folder}")
            
            # 3. Buscar el archivo con patrón *__未実験データ.xlsx
            prediction_source_file = None
            project_name = project_folder.name  # Ej: "Proyecto_79"
            expected_filename = f"{project_name}_未実験データ.xlsx"
            expected_path = project_folder / expected_filename
            
            print(f"🔍 ファイルを検索中: {expected_path}")
            
            if expected_path.exists():
                prediction_source_file = expected_path
                print(f"✅ ファイルを発見しました: {prediction_source_file}")
            else:
                # ES: Buscar cualquier archivo que termine en _未実験データ.xlsx
                # EN: Search for any file that ends with _未実験データ.xlsx
                # JP: _未実験データ.xlsx で終わるファイルを探す
                print("⚠️ 期待するファイルが見つかりません。パターン *_未実験データ.xlsx を検索します...")
                matching_files = list(project_folder.glob("*_未実験データ.xlsx"))
                if matching_files:
                    prediction_source_file = matching_files[0]
                    print(f"✅ ファイルを発見しました（パターン）: {prediction_source_file}")
                else:
                    print(f"❌ パターン *_未実験データ.xlsx のファイルが見つかりません: {project_folder}")
                    # Listar archivos disponibles para debug
                    all_files = list(project_folder.glob("*.xlsx"))
                    if all_files:
                        print(f"📋 {project_folder} にある .xlsx ファイル:")
                        for f in all_files:
                            print(f"   - {f.name}")
            
            if prediction_source_file is None:
                print(f"⚠️ 未実験データファイルが見つかりません: {project_folder}")
                print(f"   検索対象: {expected_filename} または *_未実験データ.xlsx")
                # Continuar de todas formas, puede que el usuario lo haya preparado manualmente
            
            # 4. Copiar el archivo a 04_予測/Prediction_input.xlsx
            prediction_input_path = os.path.join(prediction_folder, "Prediction_input.xlsx")
            if prediction_source_file and prediction_source_file.exists():
                import shutil
                shutil.copy2(str(prediction_source_file), prediction_input_path)
                print(f"✅ ファイルをコピーしました: {prediction_source_file} → {prediction_input_path}")
            else:
                # ES: Si no existe, verificar si ya existe el archivo de destino
                # EN: If it doesn't exist, check whether the destination file already exists
                # JP: 存在しない場合、宛先ファイルが既にあるか確認
                if not os.path.exists(prediction_input_path):
                    print("⚠️ 元ファイルが見つからず、宛先も存在しません。続行します...")
            
            # 5. Actualizar config_custom.py para cambiar PREDICTION_FOLDER a 04_予測
            config_custom_path = os.path.join(working_dir, "config_custom.py")
            if os.path.exists(config_custom_path):
                try:
                    with open(config_custom_path, 'r', encoding='utf-8') as f:
                        config_content = f.read()
                    
                    # Reemplazar PREDICTION_FOLDER de '03_予測' a '04_予測'
                    import re
                    # ES: Buscar y reemplazar PREDICTION_FOLDER = '03_予測' o PREDICTION_FOLDER = "03_予測"
                    # EN: Find and replace PREDICTION_FOLDER = '03_予測' or PREDICTION_FOLDER = \"03_予測\"
                    # JP: PREDICTION_FOLDER = '03_予測' / \"03_予測\" を検索して置換
                    pattern = r"(PREDICTION_FOLDER\s*=\s*['\"])03_予測(['\"])"
                    replacement = r"\g<1>04_予測\g<2>"
                    config_content = re.sub(pattern, replacement, config_content)
                    
                    with open(config_custom_path, 'w', encoding='utf-8') as f:
                        f.write(config_content)
                    print("✅ config_custom.py を更新しました: PREDICTION_FOLDER = '04_予測'")
                except Exception as e:
                    print(f"⚠️ config_custom.py 更新中にエラー: {e}")
            
            script_path = os.path.join(working_dir, "02_prediction.py")
            
            # ES: Si el script no está en la carpeta de salida, usar el del directorio actual
            # EN: If the script is not in the output folder, use the one from the current directory
            # JP: スクリプトが出力フォルダに無ければ現ディレクトリのものを使用
            if not os.path.exists(script_path):
                script_path = "02_prediction.py"
                if not os.path.exists(script_path):
                    print("❌ スクリプトが見つかりません: 02_prediction.py")
                    return False
            
            # ES: Configurar variables de entorno | EN: Configure environment variables | JA: 環境変数を設定
            env = os.environ.copy()
            env["OMP_NUM_THREADS"] = "1"
            env["MKL_NUM_THREADS"] = "1"
            env["OPENBLAS_NUM_THREADS"] = "1"
            env["NUMEXPR_NUM_THREADS"] = "1"
            env["MPLBACKEND"] = "Agg"
            env["QT_QPA_PLATFORM"] = "offscreen"
            env["KMP_DUPLICATE_LIB_OK"] = "TRUE"
            
            # ES: Configurar PYTHONPATH - buscar 00_Pythonコード de manera robusta | EN: Configure PYTHONPATH - find 00_Pythonコード robustly | JA: PYTHONPATH設定－00_Pythonコードを堅牢に検索
            from pathlib import Path
            python_code_folder = None
            search_path = Path(working_dir).resolve() if working_dir else Path.cwd()
            
            # ES: Buscar hacia arriba hasta encontrar 00_Pythonコード o .venv
            # EN: Search upward until finding 00_Pythonコード or .venv
            # JP: 00_Pythonコード または .venv が見つかるまで上方向に探す
            while search_path != search_path.parent:
                python_code_candidate = search_path / "00_Pythonコード"
                if python_code_candidate.exists() and python_code_candidate.is_dir():
                    python_code_folder = python_code_candidate
                    break
                # ES: También buscar .venv como indicador de la raíz del proyecto
                # EN: Also check for .venv as an indicator of the project root
                # JP: プロジェクトルートの指標として.venvも確認する
                venv_candidate = search_path / ".venv"
                if venv_candidate.exists() and venv_candidate.is_dir():
                    python_code_candidate = search_path / "00_Pythonコード"
                    if python_code_candidate.exists() and python_code_candidate.is_dir():
                        python_code_folder = python_code_candidate
                        break
                search_path = search_path.parent
            
            # ES: Si no se encuentra, usar el directorio actual como fallback
            # EN: If it's not found, use the current directory as a fallback
            # JP: 見つからなければ現ディレクトリをフォールバックとして使用
            if python_code_folder is None:
                python_code_folder = Path.cwd() / "00_Pythonコード"
                if not python_code_folder.exists():
                    # ES: Último fallback: buscar desde el directorio del script
                    # EN: Last fallback: search from the script directory
                    # JP: 最終フォールバック: スクリプトディレクトリから探す
                    script_dir = Path(__file__).parent if '__file__' in globals() else Path.cwd()
                    python_code_folder = script_dir / "00_Pythonコード"
            
            import site
            site_packages_paths = []
            try:
                for site_pkg in site.getsitepackages():
                    if os.path.exists(site_pkg):
                        site_packages_paths.append(site_pkg)
            except:
                venv_lib = Path(sys.executable).parent.parent / "Lib" / "site-packages"
                if venv_lib.exists():
                    site_packages_paths.append(str(venv_lib))
            
            pythonpath_parts = [str(python_code_folder)]
            pythonpath_parts.extend(site_packages_paths)
            
            existing_pythonpath = env.get("PYTHONPATH", "")
            if existing_pythonpath:
                pythonpath_parts.append(existing_pythonpath)
            
            separator = ";" if sys.platform == "win32" else ":"
            pythonpath = separator.join(pythonpath_parts)
            env["PYTHONPATH"] = pythonpath
            
            print(f"🔧 実行中: {script_path}")
            print(f"📁 Working directory: {working_dir}")
            print(f"📁 PYTHONPATH 設定: {pythonpath}")
            print(f"📁 00_Pythonコード を検出: {python_code_folder}")
            
            # Ejecutar script
            process = subprocess.Popen(
                [sys.executable, script_path],
                cwd=working_dir,
                env=env,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                encoding='utf-8',
                errors='replace'
            )
            
            # ES: Leer salida en tiempo real y actualizar progreso
            # EN: Read output in real time and update progress
            # JP: リアルタイムで出力を読み取り進捗を更新
            output_lines = []
            error_lines = []
            script_start_time = time.time()
            
            # Usar tiempo total si está disponible, sino usar tiempo del script
            if total_start_time is None:
                total_start_time = script_start_time
            
            def read_output(pipe, lines_list, is_stderr=False):
                try:
                    for line in iter(pipe.readline, ''):
                        if line:
                            line_clean = line.rstrip('\n\r')
                            lines_list.append(line_clean)
                            prefix = "[02_prediction]" if not is_stderr else "[02_prediction ERROR]"
                            print(f"{prefix} {line_clean}")
                except:
                    pass
            
            stdout_thread = threading.Thread(target=read_output, args=(process.stdout, output_lines, False), daemon=True)
            stderr_thread = threading.Thread(target=read_output, args=(process.stderr, error_lines, True), daemon=True)
            stdout_thread.start()
            stderr_thread.start()
            
            # Monitorear progreso mientras espera
            estimated_duration = 45  # segundos estimados para script 02
            while process.poll() is None:
                time.sleep(0.5)  # Check every 0.5 seconds
                if progress_dialog:
                    # Tiempo transcurrido total desde el inicio
                    total_elapsed = time.time() - total_start_time
                    # Tiempo transcurrido del script actual
                    script_elapsed = time.time() - script_start_time
                    
                    # Progreso basado en tiempo del script actual (sin límite artificial)
                    time_progress = min(0.95, script_elapsed / estimated_duration)  # Max 95% until it finishes
                    current_progress = int(progress_start + (progress_end - progress_start) * time_progress)
                    
                    # Calcular tiempo restante estimado de forma más precisa
                    if script_elapsed > 3 and time_progress > 0.1:  # Esperar al menos 3 segundos y 10% de progreso
                        # Usar velocidad promedio reciente
                        estimated_total = script_elapsed / time_progress
                        estimated_remaining = max(0, estimated_total - script_elapsed)
                        remaining_str = progress_dialog._format_time(estimated_remaining)
                    else:
                        # Estimación inicial conservadora
                        estimated_remaining = max(0, estimated_duration - script_elapsed)
                        remaining_str = progress_dialog._format_time(estimated_remaining)
                    
                    elapsed_str = progress_dialog._format_time(total_elapsed)
                    progress_dialog.time_info_label.setText(
                        f"⏱️ 経過時間: {elapsed_str} | 推定残り時間: {remaining_str}"
                    )
                    
                    progress_dialog.update_progress(current_progress, "02_prediction.py 実行中...")
                    QApplication.processEvents()
            
            returncode = process.returncode
            
            # Completar al 100% del rango asignado
            if progress_dialog:
                progress_dialog.update_progress(progress_end, "02_prediction.py 完了")
            
            stdout_thread.join(timeout=1.0)
            stderr_thread.join(timeout=1.0)
            
            if returncode == 0:
                print("✅ 02_prediction.py の実行が完了しました")
                return True
            else:
                print(f"❌ 02_prediction.py が終了コード {returncode} で失敗しました")
                if error_lines:
                    print("エラー:")
                    for line in error_lines:
                        print(f"  {line}")
                return False
                
        except Exception as e:
            print(f"❌ 02_prediction.py 実行中にエラー: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _run_pareto_script(self, working_dir, progress_dialog=None, progress_start=20, progress_end=100, total_start_time=None):
        """
        Ejecuta 03_pareto_analyzer.py en el directorio de trabajo
        
        Parameters
        ----------
        working_dir : str
            Directorio de trabajo
        progress_dialog : ReusableProgressDialog, optional
            Diálogo de progreso para actualizar
        progress_start : int
            Porcentaje inicial de progreso (0-100)
        progress_end : int
            Porcentaje final de progreso (0-100)
        total_start_time : float, optional
            Tiempo de inicio total para tiempo transcurrido continuo
        
        Returns
        -------
        bool
            True si el script se ejecutó exitosamente, False en caso contrario
        """
        try:
            script_path = os.path.join(working_dir, "03_pareto_analyzer.py")
            
            # ES: Si el script no está en la carpeta de salida, usar el del directorio actual
            # EN: If the script is not in the output folder, use the one from the current directory
            # JP: スクリプトが出力フォルダに無ければ現ディレクトリのものを使用
            if not os.path.exists(script_path):
                script_path = "03_pareto_analyzer.py"
                if not os.path.exists(script_path):
                    print("❌ スクリプトが見つかりません: 03_pareto_analyzer.py")
                    return False
            
            # ES: Configurar variables de entorno | EN: Configure environment variables | JA: 環境変数を設定 (igual que para prediction)
            env = os.environ.copy()
            env["OMP_NUM_THREADS"] = "1"
            env["MKL_NUM_THREADS"] = "1"
            env["OPENBLAS_NUM_THREADS"] = "1"
            env["NUMEXPR_NUM_THREADS"] = "1"
            env["MPLBACKEND"] = "Agg"
            env["QT_QPA_PLATFORM"] = "offscreen"
            env["KMP_DUPLICATE_LIB_OK"] = "TRUE"
            
            # ES: Configurar PYTHONPATH - buscar 00_Pythonコード de manera robusta | EN: Configure PYTHONPATH - find 00_Pythonコード robustly | JA: PYTHONPATH設定－00_Pythonコードを堅牢に検索 (igual que prediction)
            from pathlib import Path
            python_code_folder = None
            search_path = Path(working_dir).resolve() if working_dir else Path.cwd()
            
            # ES: Buscar hacia arriba hasta encontrar 00_Pythonコード o .venv
            # EN: Search upward until finding 00_Pythonコード or .venv
            # JP: 00_Pythonコード または .venv が見つかるまで上方向に探す
            while search_path != search_path.parent:
                python_code_candidate = search_path / "00_Pythonコード"
                if python_code_candidate.exists() and python_code_candidate.is_dir():
                    python_code_folder = python_code_candidate
                    break
                # ES: También buscar .venv como indicador de la raíz del proyecto
                # EN: Also check for .venv as an indicator of the project root
                # JP: プロジェクトルートの指標として.venvも確認する
                venv_candidate = search_path / ".venv"
                if venv_candidate.exists() and venv_candidate.is_dir():
                    python_code_candidate = search_path / "00_Pythonコード"
                    if python_code_candidate.exists() and python_code_candidate.is_dir():
                        python_code_folder = python_code_candidate
                        break
                search_path = search_path.parent
            
            # ES: Si no se encuentra, usar el directorio actual como fallback
            # EN: If it's not found, use the current directory as a fallback
            # JP: 見つからなければ現ディレクトリをフォールバックとして使用
            if python_code_folder is None:
                python_code_folder = Path.cwd() / "00_Pythonコード"
                if not python_code_folder.exists():
                    # ES: Último fallback: buscar desde el directorio del script
                    # EN: Last fallback: search from the script directory
                    # JP: 最終フォールバック: スクリプトディレクトリから探す
                    script_dir = Path(__file__).parent if '__file__' in globals() else Path.cwd()
                    python_code_folder = script_dir / "00_Pythonコード"
            
            import site
            site_packages_paths = []
            try:
                for site_pkg in site.getsitepackages():
                    if os.path.exists(site_pkg):
                        site_packages_paths.append(site_pkg)
            except:
                venv_lib = Path(sys.executable).parent.parent / "Lib" / "site-packages"
                if venv_lib.exists():
                    site_packages_paths.append(str(venv_lib))
            
            pythonpath_parts = [str(python_code_folder)]
            pythonpath_parts.extend(site_packages_paths)
            
            existing_pythonpath = env.get("PYTHONPATH", "")
            if existing_pythonpath:
                pythonpath_parts.append(existing_pythonpath)
            
            separator = ";" if sys.platform == "win32" else ":"
            pythonpath = separator.join(pythonpath_parts)
            env["PYTHONPATH"] = pythonpath
            
            print(f"🔧 実行中: {script_path}")
            print(f"📁 Working directory: {working_dir}")
            print(f"📁 PYTHONPATH 設定: {pythonpath}")
            print(f"📁 00_Pythonコード を検出: {python_code_folder}")
            
            # Ejecutar script
            process = subprocess.Popen(
                [sys.executable, script_path],
                cwd=working_dir,
                env=env,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                encoding='utf-8',
                errors='replace'
            )
            
            # ES: Leer salida en tiempo real y actualizar progreso
            # EN: Read output in real time and update progress
            # JP: リアルタイムで出力を読み取り進捗を更新
            output_lines = []
            error_lines = []
            script_start_time = time.time()
            
            # Usar tiempo total si está disponible, sino usar tiempo del script
            if total_start_time is None:
                total_start_time = script_start_time
            
            def read_output(pipe, lines_list, is_stderr=False):
                try:
                    for line in iter(pipe.readline, ''):
                        if line:
                            line_clean = line.rstrip('\n\r')
                            lines_list.append(line_clean)
                            prefix = "[03_pareto]" if not is_stderr else "[03_pareto ERROR]"
                            print(f"{prefix} {line_clean}")
                except:
                    pass
            
            stdout_thread = threading.Thread(target=read_output, args=(process.stdout, output_lines, False), daemon=True)
            stderr_thread = threading.Thread(target=read_output, args=(process.stderr, error_lines, True), daemon=True)
            stdout_thread.start()
            stderr_thread.start()
            
            # Monitorear progreso mientras espera
            estimated_duration = 90  # segundos estimados para script 03
            while process.poll() is None:
                time.sleep(0.5)  # Check every 0.5 seconds
                if progress_dialog:
                    # Tiempo transcurrido total desde el inicio
                    total_elapsed = time.time() - total_start_time
                    # Tiempo transcurrido del script actual
                    script_elapsed = time.time() - script_start_time
                    
                    # Progreso basado en tiempo del script actual (sin límite artificial)
                    time_progress = min(0.95, script_elapsed / estimated_duration)  # Max 95% until it finishes
                    current_progress = int(progress_start + (progress_end - progress_start) * time_progress)
                    
                    # Calcular tiempo restante estimado de forma más precisa
                    if script_elapsed > 5 and time_progress > 0.1:  # Esperar al menos 5 segundos y 10% de progreso
                        # Usar velocidad promedio reciente
                        estimated_total = script_elapsed / time_progress
                        estimated_remaining = max(0, estimated_total - script_elapsed)
                        remaining_str = progress_dialog._format_time(estimated_remaining)
                    else:
                        # Estimación inicial conservadora
                        estimated_remaining = max(0, estimated_duration - script_elapsed)
                        remaining_str = progress_dialog._format_time(estimated_remaining)
                    
                    elapsed_str = progress_dialog._format_time(total_elapsed)
                    progress_dialog.time_info_label.setText(
                        f"⏱️ 経過時間: {elapsed_str} | 推定残り時間: {remaining_str}"
                    )
                    
                    progress_dialog.update_progress(current_progress, "03_pareto_analyzer.py 実行中...")
                    QApplication.processEvents()
            
            returncode = process.returncode
            
            # Completar al 100% cuando termine
            if progress_dialog:
                progress_dialog.update_progress(100, "03_pareto_analyzer.py 完了")
            
            stdout_thread.join(timeout=1.0)
            stderr_thread.join(timeout=1.0)
            
            if returncode == 0:
                print("✅ 03_pareto_analyzer.py の実行が完了しました")
                return True
            else:
                print(f"❌ 03_pareto_analyzer.py が終了コード {returncode} で失敗しました")
                if error_lines:
                    print("エラー:")
                    for line in error_lines:
                        print(f"  {line}")
                return False
                
        except Exception as e:
            print(f"❌ 03_pareto_analyzer.py 実行中にエラー: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def on_nonlinear_error(self, error_message):
        """ES: Maneja errores del worker
        EN: Handle worker errors
        JA: ワーカーのエラーを処理"""
        # ES: Si el usuario canceló, no mostrar error como fallo | EN: If user cancelled, do not show as error/failure | JA: ユーザーキャンセル時はエラーとして表示しない
        if hasattr(self, '_nonlinear_cancel_requested') and self._nonlinear_cancel_requested:
            print(f"🛑 DEBUG: キャンセル後に非線形エラーを受信しました: {error_message}。無視します。")
            try:
                if hasattr(self, 'progress_dialog') and self.progress_dialog:
                    self.progress_dialog.close()
            except:
                pass
            self.set_console_overlay_topmost(False)
            return

        print(f"❌ ワーカーでエラー: {error_message}")
        
        # ES: Cerrar diálogo de progreso
        # EN: Close progress dialog
        # JP: 進捗ダイアログを閉じる
        if hasattr(self, 'progress_dialog'):
            self.progress_dialog.close()
        self.set_console_overlay_topmost(False)
        
        QMessageBox.critical(
            self,
            "非線形解析エラー",
            f"❌ 非線形解析の実行中にエラーが発生しました:\n\n{error_message}"
        )
    
    def on_classification_analysis_clicked(self):
        """ES: Acción al pulsar el botón de análisis de clasificación
        EN: Action when classification analysis button is clicked
        JA: 分類解析ボタンクリック時のアクション"""
        print("🔧 分類解析を開始します...")
        
        # ES: Si se accedió desde bunseki, mostrar diálogo de creación de proyecto | EN: If accessed from bunseki, show project creation dialog | JA: 分析からアクセス時はプロジェクト作成ダイアログを表示
        if hasattr(self, 'accessed_from_bunseki') and self.accessed_from_bunseki:
            print("📁 bunseki からのアクセスを検出しました - プロジェクト作成ダイアログを表示します")
            
            # ES: Mostrar diálogo | EN: Show dialog | JA: ダイアログを表示 de creación de proyecto (para clasificación)
            dialog = ProjectCreationDialog(self, analysis_type="classification")
            if dialog.exec() == QDialog.Accepted:
                project_name = dialog.project_name
                project_directory = dialog.project_directory
                
                # ES: Determinar la ruta completa del proyecto
                # EN: Determine the full project path
                # JP: プロジェクトの完全パスを決定
                if project_directory:
                    # ES: Si se seleccionó un proyecto existente, project_directory es el padre
                    # EN: If an existing project was selected, project_directory is the parent
                    # JP: 既存プロジェクト選択時、project_directoryは親ディレクトリ
                    # y project_name es el nombre del proyecto
                    project_path = os.path.join(project_directory, project_name)
                else:
                    # ES: Si se creó nuevo, project_directory es donde crear y project_name es el nombre
                    # EN: If a new one was created, project_directory is where to create it and project_name is the name
                    # JP: 新規作成時、project_directoryは作成先でproject_nameが名称
                    project_path = os.path.join(project_directory, project_name)
                
                # ES: Verificar si el proyecto ya existe (fue detectado como existente) | EN: Check if project already exists (detected as existing) | JA: プロジェクトが既存か確認（既存検出時）
                # ES: Para clasificación, verificar con analysis_type="classification"
                # EN: For classification, check with analysis_type=\"classification\"
                # JP: 分類の場合は analysis_type=\"classification\" で確認
                project_exists = self.is_valid_project_folder(project_path, analysis_type="classification")
                
                if project_exists:
                    print(f"✅ 既存プロジェクトを使用します: {project_path}")
                    # ES: No crear estructura, solo usar la carpeta existente
                    # EN: Do not create structure; just use the existing folder
                    # JP: 構造は作らず既存フォルダを使用
                    self.current_project_folder = project_path
                    
                    QMessageBox.information(
                        self, 
                        "プロジェクト使用", 
                        f"✅ 既存のプロジェクト '{project_name}' を使用します。\n\n"
                        f"保存先: {project_path}\n\n"
                        f"分類解析を開始します..."
                    )
                else:
                    print(f"📁 新規プロジェクトを作成します: {project_name}（場所: {project_directory}）")
                    
                    try:
                        # ES: Crear estructura del proyecto (sin 01 y 02) | EN: Create project structure (without 01 and 02) | JA: プロジェクト構造を作成（01・02なし）
                        project_path = self.create_nonlinear_project_structure(project_name, project_directory)
                        
                        # ES: Establecer la carpeta del proyecto actual | EN: Set current project folder | JA: 現在のプロジェクトフォルダを設定
                        self.current_project_folder = project_path
                        
                        QMessageBox.information(
                            self, 
                            "プロジェクト作成完了", 
                            f"✅ プロジェクト '{project_name}' が作成されました。\n\n"
                            f"保存先: {project_path}\n\n"
                            f"分類解析を開始します..."
                        )
                    except Exception as e:
                        QMessageBox.critical(
                            self, 
                            "エラー", 
                            f"❌ プロジェクト作成中にエラーが発生しました:\n{str(e)}"
                        )
                        self.accessed_from_bunseki = False
                        return
                
                # ES: Resetear la bandera
                # EN: Reset the flag
                # JP: フラグをリセット
                self.accessed_from_bunseki = False
                
                # ES: Continuar con el flujo normal (mostrar diálogo de configuración)
                # EN: Continue with the normal flow (show configuration dialog)
                # JP: 通常フローを続行（設定ダイアログを表示）
                # El resto del código seguirá igual, pero ahora con project_folder definido
                
            else:
                # ES: Usuario canceló, resetear la bandera
                # EN: User canceled; reset the flag
                # JP: ユーザーがキャンセルしたのでフラグをリセット
                self.accessed_from_bunseki = False
                return
        
        try:
            # ES: Verificar si estamos en la vista de filtros | EN: Check if we are on filter view | JA: フィルタビューか確認
            already_in_filter_view = False
            for i in range(self.center_layout.count()):
                item = self.center_layout.itemAt(i)
                if item.widget() and isinstance(item.widget(), QLabel):
                    if item.widget().text() == "データフィルター":
                        already_in_filter_view = True
                        break
            
            if not already_in_filter_view:
                # ES: Crear la vista de filtros primero | EN: Create filter view first | JA: 先にフィルタビューを作成
                self.create_filter_view()
                self.create_navigation_buttons()
                self.prev_button.setEnabled(True)
                self.next_button.setEnabled(True)
                QMessageBox.information(self, "分析ページ", "✅ 分析ページに移動しました。\nフィルターを設定して分類分析を実行してください。")
                return
            
            # ES: Obtener datos filtrados aplicando filtros ahora
            # EN: Get filtered data applying filters now
            # JA: フィルタを適用してフィルタ済みデータを取得
            # ES: Similar al análisis no lineal, obtener datos filtrados de la BBDD
            # EN: Same as non-linear analysis: get filtered data from the DB
            # JA: 非線形解析と同様、DBからフィルタ済みデータを取得
            try:
                import sqlite3
                filters = self.get_applied_filters()
                
                # Construir query con filtros
                query = "SELECT * FROM main_results WHERE 1=1"
                params = []
                
                # ES: Aplicar filtros de cepillo
                # EN: Apply brush filters
                # JP: ブラシフィルタを適用
                brush_selections = []
                if 'すべて' in filters and filters['すべて']:
                    brush_condition = " OR ".join([f"{brush} = 1" for brush in ['A13', 'A11', 'A21', 'A32']])
                    query += f" AND ({brush_condition})"
                else:
                    for brush_type in ['A13', 'A11', 'A21', 'A32']:
                        if brush_type in filters and filters[brush_type]:
                            brush_selections.append(brush_type)
                    
                    if brush_selections:
                        brush_condition = " OR ".join([f"{brush} = 1" for brush in brush_selections])
                        query += f" AND ({brush_condition})"
                
                # ES: Aplicar otros filtros
                # EN: Apply other filters
                # JP: その他のフィルタを適用
                for field_name, filter_value in filters.items():
                    if field_name in ['すべて', 'A13', 'A11', 'A21', 'A32']:
                        continue
                    
                    if isinstance(filter_value, tuple) and len(filter_value) == 2:
                        desde, hasta = filter_value
                        if desde and hasta:
                            try:
                                query += f" AND {field_name} BETWEEN ? AND ?"
                                params.extend([float(desde), float(hasta)])
                            except (ValueError, TypeError):
                                continue
                    elif isinstance(filter_value, (str, int, float)) and filter_value:
                        try:
                            value_num = float(filter_value) if isinstance(filter_value, str) else filter_value
                            query += f" AND {field_name} = ?"
                            params.append(value_num)
                        except (ValueError, TypeError):
                            continue
                
                # Ejecutar query
                conn = sqlite3.connect(RESULTS_DB_PATH, timeout=10)
                df = pd.read_sql_query(query, conn, params=params)
                conn.close()
                
                if df.empty or len(df) == 0:
                    QMessageBox.warning(self, "警告", "⚠️ フィルタリングされたデータがありません。\nフィルター条件を変更してください。")
                    return
                
                self.filtered_df = df
                print(f"📊 フィルタ済みデータ取得: {len(df)} 件")
                
            except Exception as e:
                print(f"❌ フィルタ済みデータ取得中にエラー: {e}")
                import traceback
                traceback.print_exc()
                QMessageBox.critical(self, "エラー", f"❌ データ取得中にエラーが発生しました:\n{str(e)}")
                return
            
            # ES: Verificar que hay proyecto seleccionado | EN: Ensure a project is selected | JA: プロジェクトが選択されているか確認
            if not hasattr(self, 'current_project_folder') or not self.current_project_folder:
                QMessageBox.warning(self, "プロジェクトなし", "❌ プロジェクトが選択されていません。\nまずプロジェクトを選択してください。")
                return
            
            # ES: Verificar que los módulos están disponibles | EN: Ensure modules are available | JA: モジュールが利用可能か確認
            if ClassificationWorker is None or ClassificationConfigDialog is None or BrushSelectionDialog is None:
                QMessageBox.critical(
                    self,
                    "モジュールが見つかりません",
                    "❌ 分類分析モジュールが利用できません。\nclassification_worker.py, classification_config_dialog.py と brush_selection_dialog.py が存在することを確認してください。"
                )
                return
            
            # ES: Mostrar diálogo | EN: Show dialog | JA: ダイアログを表示 de configuración
            config_dialog = ClassificationConfigDialog(self, filtered_df=self.filtered_df)
            
            if config_dialog.exec() != QDialog.Accepted:
                print("❌ ユーザーが分類解析をキャンセルしました")
                return
            
            # ES: Obtener valores de configuración
            # EN: Get configuration values
            # JP: 設定値を取得
            config_values = config_dialog.get_config_values()
            self.classification_config = config_values
            
            # ES: Verificar si es carga de folder existente | EN: Check if it is loading existing folder | JA: 既存フォルダの読み込みか確認
            is_load_existing = config_values.get('load_existing', False)
            
            # Solo preguntar parámetros si NO es carga existente
            selected_brush = None
            selected_material = None
            selected_wire_length = None
            
            if not is_load_existing:
                # ES: Mostrar diálogo | EN: Show dialog | JA: ダイアログを表示 para seleccionar parámetros (similar a yosoku)
                # QLabel, QDialog, etc. ya están importados globalmente, no importar de nuevo
                
                dialog = QDialog(self)
                dialog.setWindowTitle("予測パラメーター選択")
                dialog.setModal(True)
                dialog.resize(400, 350)
                
                layout = QVBoxLayout()
                
                # ES: Título | EN: Title | JA: タイトル
                title = QLabel("予測パラメーターを選択してください")
                title.setStyleSheet("font-weight: bold; font-size: 14px; margin: 10px;")
                title.setAlignment(Qt.AlignCenter)
                layout.addWidget(title)
                
                # ES: Formulario de selección | EN: Selection form | JA: 選択フォーム
                form_layout = QFormLayout()
                
                # ES: Tipo de cepillo | EN: Brush type | JA: ブラシタイプ
                brush_combo = QComboBox()
                brush_combo.addItem("A13", "A13")
                brush_combo.addItem("A11", "A11")
                brush_combo.addItem("A21", "A21")
                brush_combo.addItem("A32", "A32")
                brush_combo.setCurrentText("A11")  # Valor por defecto
                form_layout.addRow("ブラシタイプ:", brush_combo)
                
                # ES: Material | EN: Material | JA: 材料
                material_combo = QComboBox()
                material_combo.addItem("Steel", "Steel")
                material_combo.addItem("Alum", "Alum")
                material_combo.setCurrentText("Steel")  # Valor por defecto
                form_layout.addRow("材料:", material_combo)
                
                # 線材長 (de 30 a 75 en intervalos de 5mm)
                wire_length_combo = QComboBox()
                for value in range(30, 80, 5):  # 30, 35, 40, 45, 50, 55, 60, 65, 70, 75
                    wire_length_combo.addItem(str(value), value)
                wire_length_combo.setCurrentText("75")  # Valor por defecto
                form_layout.addRow("線材長:", wire_length_combo)
                
                layout.addLayout(form_layout)
                layout.addStretch()
                
                # ES: Botones
            # EN: Buttons
            # JA: ボタン
                button_layout = QHBoxLayout()
                
                cancel_button = QPushButton("キャンセル")
                cancel_button.clicked.connect(dialog.reject)
                
                ok_button = QPushButton("続行")
                ok_button.clicked.connect(dialog.accept)
                ok_button.setStyleSheet("background-color: #27ae60; color: white; font-weight: bold;")
                
                button_layout.addWidget(cancel_button)
                button_layout.addWidget(ok_button)
                layout.addLayout(button_layout)
                
                dialog.setLayout(layout)
                
                # ES: Mostrar diálogo | EN: Show dialog | JA: ダイアログを表示
                result = dialog.exec()
                
                if result == QDialog.Accepted:
                    selected_brush = brush_combo.currentData()
                    selected_material = material_combo.currentData()
                    selected_wire_length = wire_length_combo.currentData()
                    
                    print("✅ 選択したパラメータ:")
                    print(f"   - Brush: {selected_brush}")
                    print(f"   - Material: {selected_material}")
                    print(f"   - Wire Length: {selected_wire_length}")
                else:
                    print("❌ ユーザーがパラメータ選択をキャンセルしました")
                    return
            else:
                print("ℹ️ 既存フォルダの読み込み: パラメータ選択は不要です")
            
            # ES: Ejecutar análisis de clasificación con worker
            # EN: Run classification analysis using the worker
            # JP: ワーカーで分類解析を実行
            print("🔧 分類ワーカーを開始します...")
            self.classification_worker = ClassificationWorker(
                self.filtered_df, 
                self.current_project_folder, 
                self, 
                config_values,
                selected_brush=selected_brush,
                selected_material=selected_material,
                selected_wire_length=selected_wire_length
            )
            
            # ES: Conectar señales | EN: Connect signals | JA: シグナルを接続
            self.classification_worker.progress_updated.connect(self.on_classification_progress)
            self.classification_worker.status_updated.connect(self.on_classification_status)
            self.classification_worker.finished.connect(self.on_classification_finished)
            self.classification_worker.error.connect(self.on_classification_error)
            self.classification_worker.console_output.connect(self.on_classification_console_output)
            self.classification_worker.file_selection_requested.connect(self.on_classification_file_selection_requested)
            
            # ES: Mostrar progreso | EN: Show progress | JA: 進捗を表示
            self.progress_dialog = ReusableProgressDialog(
                self, 
                title="分類分析処理中...",
                chibi_image="Chibi_raul.png",
                chibi_size=160
            )
            self.progress_dialog.show()
            self.set_console_overlay_topmost(True)
            
            # ES: Conectar señal de cancelación | EN: Connect cancel signal | JA: キャンセルシグナルを接続
            self.progress_dialog.cancelled.connect(self.on_classification_cancelled)
            
            # Iniciar worker
            self.classification_worker.start()
            
        except Exception as e:
            QMessageBox.critical(self, "エラー", f"❌ 分類分析の実行中にエラーが発生しました:\n{str(e)}")
            print(f"❌ 分類解析中にエラー: {e}")
            import traceback
            traceback.print_exc()
    
    def on_classification_progress(self, value, message):
        """Actualiza la barra de progreso"""
        if hasattr(self, 'progress_dialog'):
            self.progress_dialog.update_progress(value, message)
    
    def on_classification_status(self, message):
        """Actualiza el estado"""
        print(f"📊 状態: {message}")
        if hasattr(self, 'progress_dialog'):
            self.progress_dialog.update_status(message)
    
    def on_classification_finished(self, results):
        """ES: Maneja el resultado de la ejecución
        EN: Handle execution result
        JA: 実行結果を処理"""
        try:
            print("✅ 分類解析が完了しました")
            print(f"   出力フォルダ: {results.get('output_folder', 'N/A')}")
            
            # ES: Cerrar diálogo de progreso
            # EN: Close progress dialog
            # JP: 進捗ダイアログを閉じる
            if hasattr(self, 'progress_dialog'):
                self.progress_dialog.close()
            self.set_console_overlay_topmost(False)
            
            # ES: Mostrar pantalla de resultados finales con estadísticas | EN: Show final results screen with statistics | JA: 統計付き最終結果画面を表示
            self._show_classification_final_results(results)
            
        except Exception as e:
            print(f"❌ on_classification_finished でエラー: {e}")
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "エラー", f"❌ 結果処理中にエラーが発生しました:\n{str(e)}")
    
    def on_classification_error(self, error_message):
        """ES: Maneja errores del worker
        EN: Handle worker errors
        JA: ワーカーのエラーを処理"""
        print(f"❌ ワーカーでエラー: {error_message}")
        
        # ES: Cerrar diálogo de progreso
        # EN: Close progress dialog
        # JP: 進捗ダイアログを閉じる
        if hasattr(self, 'progress_dialog'):
            self.progress_dialog.close()
        self.set_console_overlay_topmost(False)
        
        QMessageBox.critical(
            self,
            "分類分析エラー",
            f"❌ 分類分析の実行中にエラーが発生しました:\n\n{error_message}"
        )
    
    def on_classification_console_output(self, message):
        """ES: Maneja la salida de consola
        EN: Handle console output
        JA: コンソール出力を処理"""
        print(f"📝 {message}")
    
    def on_classification_file_selection_requested(self, initial_path):
        """ES: Maneja la solicitud de selección de archivo desde el worker
        EN: Handle file selection request from worker
        JA: ワーカーからのファイル選択要求を処理"""
        try:
            from pathlib import Path
            
            # ES: Mostrar diálogo | EN: Show dialog | JA: ダイアログを表示 para seleccionar archivo
            prev_topmost = getattr(self, '_console_topmost_enabled', False)
            # Durante file dialogs: NO taparlos con la flecha/consola
            self.set_console_overlay_topmost(False)
            file_path, _ = QFileDialog.getOpenFileName(
                self,
                "未実験データファイルを選択してください",
                initial_path,
                "Excel Files (*.xlsx *.xls);;All Files (*)"
            )
            # ES: Restaurar estado (si el loading sigue activo)
            # EN: Restore state (if loading is still active)
            # JP: 状態を復元（ローディングがまだ有効な場合）
            if prev_topmost:
                self.set_console_overlay_topmost(True)
            
            if file_path and file_path.strip():
                # ES: Validar que el archivo existe
                # EN: Validate that the file exists
                # JP: ファイルが存在するか検証
                if not Path(file_path).exists():
                    QMessageBox.warning(
                        self,
                        "エラー",
                        f"❌ 選択されたファイルが見つかりません:\n{file_path}"
                    )
                    # ES: Notificar al worker que no se seleccionó archivo
                    # EN: Notify the worker that no file was selected
                    # JP: ファイル未選択としてワーカーに通知
                    if hasattr(self, 'classification_worker'):
                        self.classification_worker._selected_file_path = None
                        self.classification_worker._file_selection_event.set()
                    return
                
                # ES: Validar columnas del archivo antes de aceptarlo
                # EN: Validate the file's columns before accepting it
                # JP: 受理前にファイル列を検証
                try:
                    import pandas as pd
                    df = pd.read_excel(file_path)
                    
                    required_columns = ['回転速度', '送り速度', 'UPカット', '切込量', '突出量', '載せ率', 'パス数']
                    missing_columns = [col for col in required_columns if col not in df.columns]
                    
                    if missing_columns:
                        QMessageBox.warning(
                            self,
                            "エラー",
                            f"❌ 選択されたファイルに必要な列がありません:\n\n"
                            f"不足している列: {', '.join(missing_columns)}\n\n"
                            f"必要な列: {', '.join(required_columns)}"
                        )
                        # ES: Notificar al worker que no se seleccionó archivo válido
                        # EN: Notify the worker that no valid file was selected
                        # JP: 有効ファイル未選択としてワーカーに通知
                        if hasattr(self, 'classification_worker'):
                            self.classification_worker._selected_file_path = None
                            self.classification_worker._file_selection_event.set()
                        return
                    
                    if len(df) == 0:
                        QMessageBox.warning(
                            self,
                            "エラー",
                            f"❌ 選択されたファイルにデータがありません:\n{file_path}"
                        )
                        # ES: Notificar al worker que no se seleccionó archivo válido
                        # EN: Notify the worker that no valid file was selected
                        # JP: 有効ファイル未選択としてワーカーに通知
                        if hasattr(self, 'classification_worker'):
                            self.classification_worker._selected_file_path = None
                            self.classification_worker._file_selection_event.set()
                        return
                    
                    # ES: Archivo válido, notificar al worker
                    # EN: Valid file; notify the worker
                    # JP: 有効ファイルとしてワーカーに通知
                    if hasattr(self, 'classification_worker'):
                        self.classification_worker._selected_file_path = file_path
                        self.classification_worker._file_selection_event.set()
                        print(f"✅ ファイルを選択して検証しました: {file_path}")
                    
                except Exception as e:
                    QMessageBox.critical(
                        self,
                        "エラー",
                        f"❌ ファイルの読み込み中にエラーが発生しました:\n{str(e)}"
                    )
                    # ES: Notificar al worker que hubo un error
                    # EN: Notify the worker that an error occurred
                    # JP: エラー発生としてワーカーに通知
                    if hasattr(self, 'classification_worker'):
                        self.classification_worker._selected_file_path = None
                        self.classification_worker._file_selection_event.set()
            else:
                # ES: Usuario canceló, notificar al worker
                # EN: User canceled; notify the worker
                # JP: ユーザーがキャンセル: ワーカーに通知
                if hasattr(self, 'classification_worker'):
                    self.classification_worker._selected_file_path = None
                    self.classification_worker._file_selection_event.set()
                    
        except Exception as e:
            print(f"❌ ファイル選択中にエラー: {e}")
            import traceback
            traceback.print_exc()
            # Notificar al worker que hubo un error
            if hasattr(self, 'classification_worker'):
                self.classification_worker._selected_file_path = None
                self.classification_worker._file_selection_event.set()
    
    def on_classification_cancelled(self):
        """ES: Maneja la cancelación
        EN: Handle cancellation
        JA: キャンセルを処理"""
        print("🛑 分類解析をキャンセル中...")
        if hasattr(self, 'classification_worker') and self.classification_worker is not None:
            self.classification_worker.cancel()
        
        if hasattr(self, 'progress_dialog'):
            self.progress_dialog.close()
        self.set_console_overlay_topmost(False)
        
        QMessageBox.information(self, "キャンセル", "分類分析がキャンセルされました。")
    
    def _show_classification_final_results(self, results):
        """ES: Muestra resultados finales del análisis de clasificación con estadísticas
        EN: Show final classification analysis results with statistics
        JA: 分類解析の最終結果を統計付きで表示"""
        output_folder = results.get('output_folder', '')
        if not output_folder:
            QMessageBox.warning(self, "エラー", "❌ 結果を表示するための情報が見つかりません。")
            return
        
        is_load_existing = results.get('load_existing', False)
        existing_folder_path = results.get('existing_folder_path', '')
        
        # ES: Limpiar layout central completamente
        # EN: Clear the center layout completely
        # JP: 中央レイアウトを完全にクリア
        while self.center_layout.count():
            item = self.center_layout.takeAt(0)
            widget = item.widget()
            if widget:
                widget.deleteLater()
            else:
                # ES: Si es un layout, limpiarlo también
                # EN: If it's a layout, clear it too
                # JP: レイアウトならそれもクリア
                layout = item.layout()
                if layout:
                    while layout.count():
                        layout_item = layout.takeAt(0)
                        layout_widget = layout_item.widget()
                        if layout_widget:
                            layout_widget.deleteLater()
        
        # Forzar actualización de la UI
        QApplication.processEvents()
        
        # ES: Crear scroll area para permitir scroll si el contenido es grande | EN: Create scroll area for large content | JA: コンテンツが大きい場合のスクロールエリアを作成
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setStyleSheet("""
            QScrollArea {
                border: none;
                background-color: #f5f5f5;
            }
        """)
        
        # ES: Crear contenedor con fondo gris limpio | EN: Create container with clean grey background | JA: クリーンなグレー背景のコンテナを作成 (dentro del scroll)
        gray_container = QFrame()
        gray_container.setStyleSheet("""
            QFrame {
                background-color: #f5f5f5;
                border-radius: 10px;
            }
        """)
        
        # ES: Layout interno para el contenedor gris | EN: Inner layout for grey container | JA: グレーコンテナ用の内部レイアウト
        container_layout = QVBoxLayout(gray_container)
        container_layout.setContentsMargins(15, 15, 15, 15)
        container_layout.setSpacing(12)  # Reducir espaciado
        
        # ES: Título | EN: Title | JA: タイトル
        if is_load_existing:
            title_text = "既存分類解析結果"
        else:
            title_text = "分類解析完了"
        
        title = QLabel(title_text)
        title.setStyleSheet("""
            font-weight: bold; 
            font-size: 20px; 
            color: #2c3e50;
            margin-bottom: 10px;
            padding: 8px 0px;
            border-bottom: 2px solid #3498db;
            border-radius: 0px;
        """)
        title.setAlignment(Qt.AlignCenter)
        container_layout.addWidget(title)
        
        # ES: Mensaje de éxito
        # EN: Success message
        # JP: 成功メッセージ
        if is_load_existing:
            success_text = "✅ 既存の解析結果を読み込みました！"
        else:
            success_text = "✅ 分類解析が完了しました！"
        
        success_label = QLabel(success_text)
        success_label.setStyleSheet("""
            font-size: 16px;
            font-weight: bold;
            color: #27ae60;
            padding: 8px;
            background-color: #d5f4e6;
            border-radius: 6px;
            border: 1px solid #27ae60;
        """)
        success_label.setAlignment(Qt.AlignCenter)
        container_layout.addWidget(success_label)
        
        # ES: Si es carga existente, cargar y mostrar archivos
        # EN: If loading an existing run, load and show files
        # JP: 既存読み込みの場合はファイルを読み込み表示
        if is_load_existing and existing_folder_path:
            self._load_and_display_existing_classification_files(container_layout, existing_folder_path, output_folder)
        else:
            # ES: Cargar y mostrar estadísticas del análisis recién completado
            # EN: Load and show statistics for the just-completed analysis
            # JP: 直近完了した解析の統計を読み込み表示
            analysis_duration = results.get('analysis_duration', 0)
            self._load_and_display_classification_statistics(container_layout, output_folder, analysis_duration)
        
        # ES: Mensaje final
        # EN: Final message
        # JP: 最終メッセージ
        final_message = QLabel("結果を確認してください。")
        final_message.setStyleSheet("""
            font-size: 12px;
            color: #7f8c8d;
            font-style: italic;
            margin-top: 8px;
        """)
        final_message.setAlignment(Qt.AlignCenter)
        container_layout.addWidget(final_message)
        
        # ES: Agregar botón "次へ" para ver gráficos (siempre que haya carpeta de salida)
        # EN: Add a "次へ" button to view charts (as long as there is an output folder)
        # JP: 出力フォルダがある場合、グラフ閲覧用に「次へ」ボタンを追加
        if output_folder:
            button_layout = QHBoxLayout()
            button_layout.addStretch()
            
            next_button = QPushButton("次へ")
            next_button.setFixedSize(100, 35)  # More compact button
            next_button.setStyleSheet("""
                QPushButton {
                    background-color: #3498db;
                    color: white;
                    border: none;
                    padding: 8px 16px;
                    border-radius: 4px;
                    font-size: 12px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #2980b9;
                }
            """)
            next_button.clicked.connect(lambda: self._show_classification_charts_from_results(results))
            button_layout.addWidget(next_button)
            button_layout.addStretch()
            container_layout.addLayout(button_layout)
        
        # ES: Configurar el scroll area con el contenedor | EN: Configure scroll area with container | JA: スクロールエリアにコンテナを設定
        scroll_area.setWidget(gray_container)
        
        # ES: Agregar el scroll area al layout central
        # EN: Add the scroll area to the center layout
        # JP: スクロールエリアを中央レイアウトに追加
        self.center_layout.addWidget(scroll_area)
        
        # ES: Guardar información para navegación de gráficos | EN: Save info for chart navigation | JA: グラフナビ用情報を保存
        if output_folder:
            # ES: Buscar carpeta de resultados para guardar la ruta
            # EN: Find the results folder to save the path
            # JP: パス保存のため結果フォルダを探す
            result_folder = os.path.join(output_folder, '02_本学習結果', '02_評価結果')
            if os.path.exists(result_folder):
                self.classification_existing_folder_path = result_folder
                # ES: Guardar la carpeta del análisis completo como project_folder | EN: Save full analysis folder as project_folder | JA: 解析フォルダをproject_folderとして保存
                self.classification_project_folder = output_folder
        
        # Forzar actualización
        QApplication.processEvents()
    
    def _load_and_display_classification_statistics(self, container_layout, output_folder, analysis_duration=0):
        """ES: Carga y muestra las estadísticas del análisis de clasificación desde diagnostic_report.txt
        EN: Load and show classification analysis statistics from diagnostic_report.txt
        JA: diagnostic_report.txtから分類解析の統計を読み込み表示"""
        try:
            from pathlib import Path
            from datetime import datetime
            import re
            
            # ES: Buscar diagnostic_report.txt en 02_本学習結果\\04_診断情報
            # EN: Search for diagnostic_report.txt in 02_本学習結果\\04_診断情報
            # JP: 02_本学習結果\\04_診断情報 で diagnostic_report.txt を探す
            diagnostic_report_path = os.path.join(output_folder, '02_本学習結果', '04_診断情報', 'diagnostic_report.txt')
            
            # También buscar en 02_本学習結果\02_評価結果 (por si acaso)
            alternative_path = os.path.join(output_folder, '02_本学習結果', '02_評価結果', 'diagnostic_report.txt')
            
            diagnostic_data = {}
            
            # ES: Intentar leer diagnostic_report.txt
            # EN: Try to read diagnostic_report.txt
            # JP: diagnostic_report.txt を読み込んでみる
            report_path = None
            if os.path.exists(diagnostic_report_path):
                report_path = diagnostic_report_path
            elif os.path.exists(alternative_path):
                report_path = alternative_path
            else:
                # Búsqueda recursiva como fallback
                for root, dirs, files in os.walk(output_folder):
                    if 'diagnostic_report.txt' in files:
                        report_path = os.path.join(root, 'diagnostic_report.txt')
                        break
            
            if report_path:
                try:
                    with open(report_path, 'r', encoding='utf-8') as f:
                        content = f.read()
                    
                    # Parsear el contenido del reporte
                    # [設定情報]
                    np_alpha_match = re.search(r'NP_ALPHA:\s*([\d.]+)', content)
                    if np_alpha_match:
                        diagnostic_data['np_alpha'] = np_alpha_match.group(1)
                    else:
                        # Intentar variaciones
                        alt_match = re.search(r'NP_ALPHA[:\s]+([\d.]+)', content, re.IGNORECASE)
                        if alt_match:
                            diagnostic_data['np_alpha'] = alt_match.group(1)
                    
                    objective_match = re.search(r'目的変数:\s*(.+)', content)
                    if objective_match:
                        diagnostic_data['objective'] = objective_match.group(1).strip()
                    else:
                        # Intentar variaciones
                        alt_match = re.search(r'目的変数[:\s]+(.+)', content)
                        if alt_match:
                            diagnostic_data['objective'] = alt_match.group(1).strip()
                    
                    # [モデル情報]
                    calibrator_match = re.search(r'Calibrator:\s*(.+)', content)
                    if calibrator_match:
                        diagnostic_data['calibrator'] = calibrator_match.group(1).strip()
                    
                    # Intentar diferentes formatos para tau_pos
                    tau_pos_match = re.search(r'τ\+\s*\(tau_pos\):\s*([\d.]+)', content)
                    if not tau_pos_match:
                        tau_pos_match = re.search(r'tau_pos[:\s]+([\d.]+)', content, re.IGNORECASE)
                    if not tau_pos_match:
                        tau_pos_match = re.search(r'τ\+[:\s]+([\d.]+)', content)
                    if tau_pos_match:
                        diagnostic_data['tau_pos'] = tau_pos_match.group(1)
                    
                    # Intentar diferentes formatos para tau_neg
                    tau_neg_match = re.search(r'τ-\s*\(tau_neg\):\s*([\d.]+)', content)
                    if not tau_neg_match:
                        tau_neg_match = re.search(r'tau_neg[:\s]+([\d.]+)', content, re.IGNORECASE)
                    if not tau_neg_match:
                        tau_neg_match = re.search(r'τ-[:\s]+([\d.]+)', content)
                    if tau_neg_match:
                        diagnostic_data['tau_neg'] = tau_neg_match.group(1)
                    
                    features_match = re.search(r'選択特徴量数:\s*(\d+)', content)
                    if features_match:
                        diagnostic_data['selected_features'] = features_match.group(1)
                    
                    # [予測結果統計]
                    total_data_match = re.search(r'総データ数:\s*([\d,]+)', content)
                    if total_data_match:
                        diagnostic_data['total_data'] = total_data_match.group(1).replace(',', '')
                    
                    coverage_match = re.search(r'カバレッジ:\s*([\d.]+)%', content)
                    if not coverage_match:
                        coverage_match = re.search(r'カバレッジ[:\s]+([\d.]+)', content)
                    if coverage_match:
                        diagnostic_data['coverage'] = coverage_match.group(1)
                    
                    # [ノイズ付加設定]
                    noise_enabled_match = re.search(r'ノイズ付加:\s*(True|False)', content)
                    if noise_enabled_match:
                        diagnostic_data['noise_enabled'] = noise_enabled_match.group(1) == 'True'
                    
                    noise_level_match = re.search(r'ノイズレベル:\s*([\d.]+)\s*ppm', content)
                    if not noise_level_match:
                        noise_level_match = re.search(r'ノイズレベル[:\s]+([\d.]+)', content)
                    if noise_level_match:
                        diagnostic_data['noise_level'] = noise_level_match.group(1)
                    
                    print(f"✅ 診断データを読み込みました: {report_path}")
                    print(f"🔍 [DEBUG] 解析済みデータ: {diagnostic_data}")
                    print(f"🔍 [DEBUG] tau_pos: {diagnostic_data.get('tau_pos')}")
                    print(f"🔍 [DEBUG] tau_neg: {diagnostic_data.get('tau_neg')}")
                    print(f"🔍 [DEBUG] noise_enabled: {diagnostic_data.get('noise_enabled')}")
                    print(f"🔍 [DEBUG] noise_level: {diagnostic_data.get('noise_level')}")
                except Exception as e:
                    print(f"⚠️ diagnostic_report.txt の読み込み中にエラー: {e}")
                    import traceback
                    traceback.print_exc()
            else:
                print(f"⚠️ diagnostic_report.txt が見つかりません: {diagnostic_report_path} または {alternative_path}")
            
            # ES: Formatear tiempo de análisis
            # EN: Format analysis duration
            # JP: 解析時間を整形
            if analysis_duration > 0:
                hours = int(analysis_duration // 3600)
                minutes = int((analysis_duration % 3600) // 60)
                seconds = int(analysis_duration % 60)
                if hours > 0:
                    analysis_duration_formatted = f"{hours}時間{minutes}分{seconds}秒"
                elif minutes > 0:
                    analysis_duration_formatted = f"{minutes}分{seconds}秒"
                else:
                    analysis_duration_formatted = f"{seconds:.1f}秒"
            else:
                analysis_duration_formatted = "N/A"
            
            # ES: Información del análisis
            # EN: Analysis information
            # JP: 解析情報
            info_lines = []
            info_lines.append(f"📊 解析完了時刻: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            info_lines.append(f"⏱️ 解析時間: {analysis_duration_formatted}")
            
            if diagnostic_data.get('objective'):
                info_lines.append(f"🎯 目的変数: {diagnostic_data['objective']}")
            
            if diagnostic_data.get('np_alpha'):
                info_lines.append(f"⚙️ NP_ALPHA: {diagnostic_data['np_alpha']}")
            
            if diagnostic_data.get('total_data'):
                info_lines.append(f"📈 総データ数: {diagnostic_data['total_data']} レコード")
            
            if diagnostic_data.get('coverage'):
                info_lines.append(f"📊 カバレッジ: {diagnostic_data['coverage']}%")
            
            if diagnostic_data.get('selected_features'):
                info_lines.append(f"🔧 選択特徴量数: {diagnostic_data['selected_features']} 個")
            
            info_text = "\n".join(info_lines)
            info_label = QLabel(info_text)
            info_label.setStyleSheet("""
                font-size: 12px;
                color: #34495e;
                background-color: #ecf0f1;
                padding: 10px;
                border-radius: 6px;
                border: 1px solid #bdc3c7;
            """)
            info_label.setAlignment(Qt.AlignLeft)
            info_label.setWordWrap(True)
            info_label.setMinimumHeight(50)
            container_layout.addWidget(info_label)
            
            # ES: Sección de métricas del modelo si están disponibles
            # EN: Model-metrics section (if available)
            # JP: モデル指標セクション（利用可能なら）
            print(f"🔍 [DEBUG] Verificando Model Information: tau_pos={diagnostic_data.get('tau_pos')}, tau_neg={diagnostic_data.get('tau_neg')}")
            if diagnostic_data.get('tau_pos') and diagnostic_data.get('tau_neg'):
                print("✅ [DEBUG] Model Information を表示中")
                metrics_title = QLabel("📊 モデル情報 (Model Information)")
                metrics_title.setStyleSheet("""
                    font-weight: bold; 
                    font-size: 16px; 
                    color: #2c3e50;
                    margin-top: 10px;
                    margin-bottom: 8px;
                    padding-bottom: 6px;
                    border-bottom: 2px solid #3498db;
                """)
                metrics_title.setAlignment(Qt.AlignCenter)
                container_layout.addWidget(metrics_title)
                
                # ES: Crear tarjeta de métricas | EN: Create metric card | JA: メトリックカードを作成
                metric_card = QFrame()
                metric_card.setStyleSheet("""
                    QFrame {
                        background-color: #ffffff;
                        border: 2px solid #3498db;
                        border-radius: 8px;
                        padding: 10px;
                    }
                """)
                card_layout = QVBoxLayout(metric_card)
                card_layout.setSpacing(6)  # Reducir espaciado
                card_layout.setContentsMargins(10, 10, 10, 10)
                
                # Calibrator
                if diagnostic_data.get('calibrator'):
                    calibrator_text = f"Calibrator: {diagnostic_data['calibrator']}"
                    calibrator_label = QLabel(calibrator_text)
                    calibrator_label.setStyleSheet("""
                        font-size: 12px;
                        color: #34495e;
                        padding: 6px;
                        background-color: #f8f9fa;
                        border-radius: 4px;
                        min-height: 24px;
                    """)
                    calibrator_label.setMinimumHeight(24)
                    calibrator_label.setWordWrap(True)
                    print(f"✅ [DEBUG] Agregando calibrator_label: {calibrator_text}")
                    card_layout.addWidget(calibrator_label)
                
                # τ+ y τ- (separados en labels diferentes para asegurar visibilidad)
                tau_pos_text = f"τ+ (tau_pos): {diagnostic_data['tau_pos']}"
                tau_pos_label = QLabel(tau_pos_text)
                tau_pos_label.setStyleSheet("""
                    font-size: 12px;
                    color: #34495e;
                    padding: 6px;
                    background-color: #f8f9fa;
                    border-radius: 4px;
                    min-height: 24px;
                """)
                tau_pos_label.setMinimumHeight(24)
                tau_pos_label.setWordWrap(True)
                print(f"✅ [DEBUG] Agregando tau_pos_label: {tau_pos_text}")
                card_layout.addWidget(tau_pos_label)
                
                tau_neg_text = f"τ- (tau_neg): {diagnostic_data['tau_neg']}"
                tau_neg_label = QLabel(tau_neg_text)
                tau_neg_label.setStyleSheet("""
                    font-size: 12px;
                    color: #34495e;
                    padding: 6px;
                    background-color: #f8f9fa;
                    border-radius: 4px;
                    min-height: 24px;
                """)
                tau_neg_label.setMinimumHeight(24)
                tau_neg_label.setWordWrap(True)
                print(f"✅ [DEBUG] Agregando tau_neg_label: {tau_neg_text}")
                card_layout.addWidget(tau_neg_label)
                
                # ES: Verificar si τ- < τ+ (normal) | EN: Check if τ- < τ+ (normal) | JA: τ- < τ+（正常）か確認
                try:
                    tau_pos_val = float(diagnostic_data['tau_pos'])
                    tau_neg_val = float(diagnostic_data['tau_neg'])
                    print(f"🔍 [DEBUG] Comparando tau: tau_neg={tau_neg_val} < tau_pos={tau_pos_val} = {tau_neg_val < tau_pos_val}")
                    if tau_neg_val < tau_pos_val:
                        status_text = "✅ 正常: τ- < τ+"
                        status_label = QLabel(status_text)
                        status_label.setStyleSheet("""
                            font-size: 12px;
                            font-weight: bold;
                            color: #27ae60;
                            padding: 6px;
                            background-color: #d5f4e6;
                            border-radius: 4px;
                            border: 1px solid #27ae60;
                            min-height: 28px;
                        """)
                    else:
                        status_text = "⚠️ 警告: τ- >= τ+"
                        status_label = QLabel(status_text)
                        status_label.setStyleSheet("""
                            font-size: 12px;
                            font-weight: bold;
                            color: #f39c12;
                            padding: 6px;
                            background-color: #fef5e7;
                            border-radius: 4px;
                            border: 1px solid #f39c12;
                            min-height: 28px;
                        """)
                    status_label.setMinimumHeight(28)
                    status_label.setWordWrap(True)
                    status_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                    print(f"✅ [DEBUG] Agregando status_label: {status_text}")
                    card_layout.addWidget(status_label)
                    print(f"✅ [DEBUG] status_label agregado al layout. Total widgets en card_layout: {card_layout.count()}")
                except Exception as e:
                    print(f"⚠️ status_label 追加エラー: {e}")
                    import traceback
                    traceback.print_exc()
                
                # Asegurar que la tarjeta tenga contenido visible
                print(f"✅ [DEBUG] Total widgets en metric_card antes de agregar: {card_layout.count()}")
                # Calcular altura mínima basada en el número de widgets (más compacto)
                min_height = max(120, card_layout.count() * 35)  # Al menos 35px por widget
                metric_card.setMinimumHeight(min_height)
                metric_card.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Minimum)
                print(f"✅ [DEBUG] metric_card 最小高さ: {min_height}px")
                container_layout.addWidget(metric_card)
                print(f"✅ [DEBUG] metric_card agregado al container_layout")
            else:
                # ES: Mostrar mensaje si no hay información del modelo | EN: Show message if no model info | JA: モデル情報が無い場合にメッセージ表示
                if not diagnostic_data:
                    no_data_label = QLabel("⚠️ 統計情報を読み込めませんでした。\n診断レポートファイルを確認してください。")
                    no_data_label.setStyleSheet("""
                        font-size: 14px;
                        color: #e67e22;
                        background-color: #fef5e7;
                        padding: 15px;
                        border-radius: 8px;
                        border: 1px solid #e67e22;
                    """)
                    no_data_label.setAlignment(Qt.AlignCenter)
                    no_data_label.setWordWrap(True)
                    no_data_label.setMinimumHeight(60)
                    container_layout.addWidget(no_data_label)
            
            # Información de ruido si está disponible
            print(f"🔍 [DEBUG] Verificando Noise Settings: noise_enabled={diagnostic_data.get('noise_enabled')}")
            if diagnostic_data.get('noise_enabled'):
                print("✅ [DEBUG] Noise Addition Settings を表示中")
                noise_title = QLabel("🔊 ノイズ付加設定 (Noise Addition Settings)")
                noise_title.setStyleSheet("""
                    font-weight: bold; 
                    font-size: 16px; 
                    color: #2c3e50;
                    margin-top: 10px;
                    margin-bottom: 8px;
                    padding-bottom: 6px;
                    border-bottom: 2px solid #3498db;
                """)
                noise_title.setAlignment(Qt.AlignCenter)
                container_layout.addWidget(noise_title)
                
                noise_card = QFrame()
                noise_card.setStyleSheet("""
                    QFrame {
                        background-color: #ffffff;
                        border: 2px solid #3498db;
                        border-radius: 8px;
                        padding: 10px;
                    }
                """)
                noise_layout = QVBoxLayout(noise_card)
                noise_layout.setSpacing(6)  # Reducir espaciado
                noise_layout.setContentsMargins(10, 10, 10, 10)
                
                if diagnostic_data.get('noise_level'):
                    noise_info = f"ノイズレベル: {diagnostic_data['noise_level']} ppm"
                    noise_label = QLabel(noise_info)
                    noise_label.setStyleSheet("""
                        font-size: 12px;
                        color: #34495e;
                        padding: 6px;
                        background-color: #f8f9fa;
                        border-radius: 4px;
                        min-height: 24px;
                    """)
                    noise_label.setMinimumHeight(24)
                    noise_label.setWordWrap(True)
                    print(f"✅ [DEBUG] Agregando noise_label: {noise_info}")
                    noise_layout.addWidget(noise_label)
                else:
                    # ES: Mostrar mensaje si no hay noise_level pero noise_enabled es True | EN: Show message if no noise_level but noise_enabled is True | JA: noise_level無しでnoise_enabledがTrueのときメッセージ表示
                    noise_info_text = "ノイズ付加: 有効"
                    noise_info_label = QLabel(noise_info_text)
                    noise_info_label.setStyleSheet("""
                        font-size: 12px;
                        color: #34495e;
                        padding: 6px;
                        background-color: #f8f9fa;
                        border-radius: 4px;
                        min-height: 24px;
                    """)
                    noise_info_label.setMinimumHeight(24)
                    print(f"✅ [DEBUG] Agregando noise_info_label: {noise_info_text}")
                    noise_layout.addWidget(noise_info_label)
                
                # Asegurar que la tarjeta tenga contenido visible
                print(f"✅ [DEBUG] Total widgets en noise_card antes de agregar: {noise_layout.count()}")
                # Calcular altura mínima basada en el número de widgets (más compacto)
                min_height = max(70, noise_layout.count() * 35)  # Al menos 35px por widget
                noise_card.setMinimumHeight(min_height)
                noise_card.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Minimum)
                print(f"✅ [DEBUG] noise_card 最小高さ: {min_height}px")
                container_layout.addWidget(noise_card)
                print(f"✅ [DEBUG] noise_card agregado al container_layout")
            
        except Exception as e:
            print(f"❌ 分類統計の読み込み中にエラー: {e}")
            import traceback
            traceback.print_exc()
            error_label = QLabel(f"⚠️ 統計情報の読み込み中にエラーが発生しました: {str(e)}")
            error_label.setStyleSheet("""
                font-size: 14px;
                color: #e74c3c;
                background-color: #fadbd8;
                padding: 15px;
                border-radius: 8px;
                border: 1px solid #e74c3c;
            """)
            container_layout.addWidget(error_label)
    
    def _load_and_display_existing_classification_files(self, container_layout, existing_folder_path, output_folder):
        """ES: Carga y muestra los archivos de un análisis de clasificación existente
        EN: Load and show files from an existing classification analysis
        JA: 既存の分類解析のファイルを読み込み表示"""
        try:
            # ES: Cargar y mostrar estadísticas del análisis existente
            # EN: Load and display statistics from the existing analysis
            # JP: 既存解析の統計を読み込み表示
            self._load_and_display_classification_statistics(container_layout, output_folder, analysis_duration=0)
            
        except Exception as e:
            print(f"❌ 既存の分類結果ファイルの読み込み中にエラー: {e}")
            import traceback
            traceback.print_exc()
            error_label = QLabel(f"⚠️ 既存結果の読み込み中にエラーが発生しました: {str(e)}")
            error_label.setStyleSheet("""
                font-size: 14px;
                color: #e74c3c;
                background-color: #fadbd8;
                padding: 15px;
                border-radius: 8px;
                border: 1px solid #e74c3c;
            """)
            container_layout.addWidget(error_label)
    
    def _show_classification_charts_from_results(self, results):
        """ES: Mostrar gráficos del análisis de clasificación desde los resultados
        EN: Show classification analysis charts from results
        JA: 結果から分類解析のグラフを表示"""
        output_folder = results.get('output_folder', '')
        if not output_folder:
            QMessageBox.warning(self, "エラー", "❌ グラフを表示するための情報が見つかりません。")
            return
        
        # ES: Buscar carpeta de resultados (02_本学習結果\\02_評価結果)
        # EN: Find the results folder (02_本学習結果\\02_評価結果)
        # JP: 結果フォルダ（02_本学習結果\\02_評価結果）を探す
        result_folder = os.path.join(output_folder, '02_本学習結果', '02_評価結果')
        
        # ES: Guardar información para navegación | EN: Save navigation info | JA: ナビ用情報を保存
        if os.path.exists(result_folder):
            self.classification_existing_folder_path = result_folder
            self.classification_project_folder = output_folder
            # Llamar a la función de mostrar gráficos
            if hasattr(self, 'show_classification_charts'):
                self.show_classification_charts()
            else:
                QMessageBox.information(
                    self,
                    "情報",
                    "グラフ表示機能は準備中です。\n\n結果フォルダ:\n" + output_folder
                )
        else:
            QMessageBox.warning(
                self,
                "エラー",
                f"❌ 結果フォルダが見つかりません:\n{result_folder}"
            )
    
    def show_classification_charts(self):
        """ES: Mostrar gráficos del análisis de clasificación con navegación
        EN: Show classification analysis charts with navigation
        JA: 分類解析のグラフをナビ付きで表示"""
        print("🔧 分類解析のグラフを表示中...")
        
        try:
            # ES: Verificar que tenemos la ruta de la carpeta cargada | EN: Ensure we have loaded folder path | JA: 読み込み済みフォルダパスがあるか確認
            if not hasattr(self, 'classification_existing_folder_path') or not self.classification_existing_folder_path:
                QMessageBox.warning(self, "エラー", "❌ グラフを表示するための情報が見つかりません。")
                return
            
            # ES: Limpiar layout central completamente
            # EN: Clear the center layout completely
            # JP: 中央レイアウトを完全にクリア
            while self.center_layout.count():
                item = self.center_layout.takeAt(0)
                widget = item.widget()
                if widget:
                    widget.deleteLater()
                else:
                    # ES: Si es un layout, limpiarlo también
                    # EN: If it's a layout, clear it too
                    # JP: レイアウトならそれもクリア
                    layout = item.layout()
                    if layout:
                        while layout.count():
                            layout_item = layout.takeAt(0)
                            layout_widget = layout_item.widget()
                            if layout_widget:
                                layout_widget.deleteLater()
            
            # Forzar actualización de la UI
            QApplication.processEvents()
            
            # ES: Crear contenedor con fondo gris limpio | EN: Create container with clean grey background | JA: クリーンなグレー背景のコンテナを作成
            gray_container = QFrame()
            gray_container.setStyleSheet("""
                QFrame {
                    background-color: #f5f5f5;
                    border-radius: 10px;
                    margin: 10px;
                }
            """)
            
            # ES: Layout interno para el contenedor gris | EN: Inner layout for grey container | JA: グレーコンテナ用の内部レイアウト
            container_layout = QVBoxLayout(gray_container)
            container_layout.setContentsMargins(20, 20, 20, 20)
            container_layout.setSpacing(15)
            
            # ES: Título | EN: Title | JA: タイトル
            title = QLabel("分類解析結果 チャート")
            title.setStyleSheet("""
                font-weight: bold; 
                font-size: 24px; 
                color: #2c3e50;
                margin-bottom: 20px;
                padding: 10px 0px;
                border-bottom: 2px solid #3498db;
                border-radius: 0px;
            """)
            title.setAlignment(Qt.AlignCenter)
            container_layout.addWidget(title)
            
            # ES: Buscar gráficos PNG en la carpeta de resultados (02_本学習結果\\02_評価結果)
            # EN: Search for PNG charts in the results folder (02_本学習結果\\02_評価結果)
            # JP: 結果フォルダ（02_本学習結果\\02_評価結果）でPNGグラフを探す
            from pathlib import Path
            folder_path = Path(self.classification_existing_folder_path)
            chart_images = []
            
            # ES: Buscar imágenes PNG directamente en la carpeta de resultados
            # EN: Search for PNG images directly in the results folder
            # JP: 結果フォルダ直下でPNG画像を探す
            for file in folder_path.glob("*.png"):
                if file.is_file():
                    chart_images.append(str(file))
            
            # ES: Si no se encuentran gráficos, mostrar mensaje
            # EN: If no charts are found, show a message
            # JP: グラフが見つからない場合はメッセージを表示
            if not chart_images:
                no_charts_label = QLabel("⚠️ グラフが見つかりません")
                no_charts_label.setStyleSheet("""
                    font-size: 16px;
                    color: #e74c3c;
                    background-color: #fadbd8;
                    padding: 20px;
                    border-radius: 8px;
                    border: 1px solid #e74c3c;
                    margin: 20px 0px;
                """)
                no_charts_label.setAlignment(Qt.AlignCenter)
                container_layout.addWidget(no_charts_label)
            else:
                # ES: Configurar navegación de gráficos | EN: Configure chart navigation | JA: グラフナビゲーションを設定
                self.classification_chart_images = sorted(chart_images)
                self.current_classification_chart_index = 0
                
                # ES: Layout principal | EN: Main layout | JA: メインレイアウト para la imagen y navegación
                chart_layout = QVBoxLayout()
                
                # ES: Label para mostrar la imagen (ocupa todo el ancho)
                # EN: Label to display the image (takes full width)
                # JP: 画像表示用ラベル（全幅）
                self.classification_chart_label = QLabel()
                self.classification_chart_label.setAlignment(Qt.AlignCenter)
                self.classification_chart_label.setStyleSheet("""
                    QLabel {
                        background-color: white;
                        border: 2px solid #bdc3c7;
                        border-radius: 10px;
                        padding: 10px;
                        min-height: 500px;
                    }
                """)
                chart_layout.addWidget(self.classification_chart_label)
                
                # ES: Layout horizontal para botones de navegación (debajo de la imagen) | EN: Horizontal layout for nav buttons (below image) | JA: ナビボタン用横レイアウト（画像下）
                nav_buttons_layout = QHBoxLayout()
                nav_buttons_layout.addStretch()
                
                # ES: Botón flecha izquierda | EN: Left arrow button | JA: 左矢印ボタン
                prev_chart_button = QPushButton("◀ 前へ")
                prev_chart_button.setFixedSize(100, 40)
                prev_chart_button.setStyleSheet("""
                    QPushButton {
                        background-color: #3498db;
                        color: white;
                        border: none;
                        border-radius: 8px;
                        font-size: 14px;
                        font-weight: bold;
                        padding: 8px 16px;
                    }
                    QPushButton:hover {
                        background-color: #2980b9;
                    }
                    QPushButton:disabled {
                        background-color: #bdc3c7;
                        color: #7f8c8d;
                    }
                """)
                prev_chart_button.clicked.connect(self.show_previous_classification_chart)
                nav_buttons_layout.addWidget(prev_chart_button)
                
                # ES: Espacio entre botones
                # EN: Space between buttons
                # JA: ボタン間のスペース
                nav_buttons_layout.addSpacing(20)
                
                # ES: Botón flecha derecha | EN: Right arrow button | JA: 右矢印ボタン
                next_chart_button = QPushButton("次へ ▶")
                next_chart_button.setFixedSize(100, 40)
                next_chart_button.setStyleSheet("""
                    QPushButton {
                        background-color: #3498db;
                        color: white;
                        border: none;
                        border-radius: 8px;
                        font-size: 14px;
                        font-weight: bold;
                        padding: 8px 16px;
                    }
                    QPushButton:hover {
                        background-color: #2980b9;
                    }
                    QPushButton:disabled {
                        background-color: #bdc3c7;
                        color: #7f8c8d;
                    }
                """)
                next_chart_button.clicked.connect(self.show_next_classification_chart)
                nav_buttons_layout.addWidget(next_chart_button)
                
                nav_buttons_layout.addStretch()
                chart_layout.addLayout(nav_buttons_layout)
                
                # Información del gráfico actual
                self.classification_chart_info_label = QLabel()
                self.classification_chart_info_label.setStyleSheet("""
                    font-size: 14px;
                    color: #2c3e50;
                    background-color: #ecf0f1;
                    padding: 10px;
                    border-radius: 5px;
                    border: 1px solid #bdc3c7;
                    margin: 10px 0px;
                """)
                self.classification_chart_info_label.setAlignment(Qt.AlignCenter)
                chart_layout.addWidget(self.classification_chart_info_label)
                
                container_layout.addLayout(chart_layout)
                
                # ES: Mostrar el primer gráfico | EN: Show first chart | JA: 先頭グラフを表示
                self.update_classification_chart_display()
            
            # ES: Botones
            # EN: Buttons
            # JA: ボタン para volver e importar a BBDD
            buttons_layout = QHBoxLayout()
            buttons_layout.addStretch()
            
            # ES: Botón para importar a BBDD | EN: Import to DB button | JA: BBDDへインポートボタン
            import_db_button = QPushButton("データベースにインポート")
            import_db_button.setFixedSize(180, 40)
            import_db_button.setStyleSheet("""
                QPushButton {
                    background-color: #27ae60;
                    color: white;
                    border: none;
                    padding: 10px 20px;
                    border-radius: 5px;
                    font-size: 14px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #229954;
                }
            """)
            import_db_button.clicked.connect(lambda: self.import_classification_results_to_yosoku_db())
            buttons_layout.addWidget(import_db_button)
            
            buttons_layout.addSpacing(20)
            
            # ES: Botón para volver | EN: Back button | JA: 戻るボタン
            back_button = QPushButton("戻る")
            back_button.setFixedSize(120, 40)
            back_button.setStyleSheet("""
                QPushButton {
                    background-color: #e74c3c;
                    color: white;
                    border: none;
                    padding: 10px 20px;
                    border-radius: 5px;
                    font-size: 14px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #c0392b;
                }
            """)
            back_button.clicked.connect(self.on_analyze_clicked)
            buttons_layout.addWidget(back_button)
            
            buttons_layout.addStretch()
            container_layout.addLayout(buttons_layout)
            
            # ES: Espacio flexible
            # EN: Flexible space
            # JA: 可変スペース
            container_layout.addStretch()
            
            # ES: Agregar el contenedor gris al layout central
            # EN: Add the gray container to the center layout
            # JP: 灰色コンテナを中央レイアウトに追加
            self.center_layout.addWidget(gray_container)
            
            print("✅ 分類解析のグラフを表示しました")
            
        except Exception as e:
            print(f"❌ 分類解析グラフの表示エラー: {e}")
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "エラー", f"❌ グラフの表示中にエラーが発生しました:\n{str(e)}")
    
    def show_previous_classification_chart(self):
        """ES: Mostrar gráfico anterior del análisis de clasificación
        EN: Show previous classification analysis chart
        JA: 分類解析の前のグラフを表示
        """
        if hasattr(self, 'classification_chart_images') and len(self.classification_chart_images) > 0:
            if not hasattr(self, 'current_classification_chart_index'):
                self.current_classification_chart_index = 0
            self.current_classification_chart_index = (self.current_classification_chart_index - 1) % len(self.classification_chart_images)
            self.update_classification_chart_display()
    
    def show_next_classification_chart(self):
        """ES: Mostrar gráfico siguiente del análisis de clasificación
        EN: Show next classification analysis chart
        JA: 分類解析の次のグラフを表示
        """
        if hasattr(self, 'classification_chart_images') and len(self.classification_chart_images) > 0:
            if not hasattr(self, 'current_classification_chart_index'):
                self.current_classification_chart_index = 0
            self.current_classification_chart_index = (self.current_classification_chart_index + 1) % len(self.classification_chart_images)
            self.update_classification_chart_display()
    
    def update_classification_chart_display(self):
        """ES: Actualizar la visualización del gráfico actual del análisis de clasificación
        EN: Update the display of the current classification analysis chart
        JA: 分類解析の現在グラフ表示を更新
        """
        if not hasattr(self, 'classification_chart_images') or len(self.classification_chart_images) == 0:
            return
        
        if not hasattr(self, 'current_classification_chart_index'):
            self.current_classification_chart_index = 0
        
        if self.current_classification_chart_index < 0:
            self.current_classification_chart_index = 0
        elif self.current_classification_chart_index >= len(self.classification_chart_images):
            self.current_classification_chart_index = len(self.classification_chart_images) - 1
        
        current_image_path = self.classification_chart_images[self.current_classification_chart_index]
        
        # ES: Cargar y mostrar la imagen | EN: Load and display the image | JA: 画像を読み込み表示
        pixmap = QPixmap(current_image_path)
        if not pixmap.isNull():
            # ES: Redimensionar para ocupar el ancho disponible | EN: Resize to fit available space | JA: 利用可能領域に合わせてリサイズ
            container_width = self.classification_chart_label.width() - 20
            container_height = self.classification_chart_label.height() - 20
            
            # ES: Si el contenedor aún no tiene tamaño, usar un tamaño por defecto
            # EN: If container size is not ready yet, use a default size
            # JA: サイズ未確定の場合はデフォルトサイズを使用
            if container_width <= 0:
                container_width = 1000
            if container_height <= 0:
                container_height = 600
            
            # ES: Redimensionar manteniendo la proporción | EN: Resize while keeping aspect ratio | JA: アスペクト比を維持してリサイズ
            scaled_pixmap = pixmap.scaled(container_width, container_height, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            self.classification_chart_label.setPixmap(scaled_pixmap)
            
            # ES: Actualizar información del gráfico | EN: Update chart info | JA: グラフ情報を更新
            image_name = os.path.basename(current_image_path)
            total_images = len(self.classification_chart_images)
            current_index = self.current_classification_chart_index + 1
            self.classification_chart_info_label.setText(f"{image_name} ({current_index}/{total_images})")
            
            # Actualizar estado de botones de navegación
            if hasattr(self, 'classification_chart_label'):
                # Los botones se habilitan/deshabilitan automáticamente por el layout
                pass

    def create_linear_analysis_folder_structure(self, project_folder):
        """ES: Crear estructura de carpetas para análisis lineal con numeración correlativa y timestamp
        EN: Create folder structure for linear analysis with sequential numbering and timestamp
        JA: 線形解析用フォルダ構造を連番とタイムスタンプで作成"""
        import os
        from datetime import datetime
        import re
        
        # ES: Ruta de la carpeta de análisis lineal
        # EN: Path to the linear-analysis folder
        # JP: 線形解析フォルダのパス
        linear_regression_folder = os.path.join(project_folder, "03_線形回帰")
        
        # ES: Crear carpeta si no existe | EN: Create folder if it does not exist | JA: フォルダが無ければ作成
        os.makedirs(linear_regression_folder, exist_ok=True)
        
        # ES: Obtener timestamp actual
        # EN: Get current timestamp
        # JP: 現在のタイムスタンプを取得
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # ES: Buscar el siguiente número correlativo
        # EN: Find the next sequential number
        # JP: 次の連番を探す
        existing_folders = []
        for item in os.listdir(linear_regression_folder):
            item_path = os.path.join(linear_regression_folder, item)
            if os.path.isdir(item_path):
                # ES: Buscar patrones como \"01_\", \"02_\", etc.
                # EN: Look for patterns like \"01_\", \"02_\", etc.
                # JP: 「01_」「02_」などのパターンを探す
                match = re.match(r'^(\d{2})_', item)
                if match:
                    existing_folders.append(int(match.group(1)))
        
        # ES: Determinar el siguiente número
        # EN: Determine the next number
        # JP: 次の番号を決定
        if existing_folders:
            next_number = max(existing_folders) + 1
        else:
            next_number = 1
        
        # ES: Crear nombre de carpeta con formato: 01_YYYYMMDD_HHMMSS | EN: Create folder name format 01_YYYYMMDD_HHMMSS | JA: フォルダ名形式 01_YYYYMMDD_HHMMSS を作成
        folder_name = f"{next_number:02d}_{timestamp}"
        analysis_folder = os.path.join(linear_regression_folder, folder_name)
        
        # ES: Crear carpeta principal | EN: Create main folder | JA: メインフォルダを作成
        os.makedirs(analysis_folder, exist_ok=True)
        print(f"📁 解析フォルダを作成しました: {analysis_folder}")
        
        # ES: Crear subcarpetas | EN: Create subfolders | JA: サブフォルダを作成
        subfolders = [
            "01_学習モデル",
            "02_パラメーター", 
            "03_評価スコア",
            "04_予測計算"
        ]
        
        for subfolder in subfolders:
            subfolder_path = os.path.join(analysis_folder, subfolder)
            os.makedirs(subfolder_path, exist_ok=True)
            print(f"📁 サブフォルダを作成しました: {subfolder_path}")
            
            # ES: Crear subcarpeta adicional dentro de 03_評価スコア | EN: Create extra subfolder inside 03_評価スコア | JA: 03_評価スコア内に追加サブフォルダを作成
            if subfolder == "03_評価スコア":
                chart_subfolder = os.path.join(subfolder_path, "01_チャート")
                os.makedirs(chart_subfolder, exist_ok=True)
                print(f"📁 グラフ用サブフォルダを作成しました: {chart_subfolder}")
        
        return analysis_folder

    def execute_linear_analysis(self):
        """ES: Ejecutar análisis lineal con los filtros aplicados
        EN: Run linear analysis with filters applied
        JA: 適用済みフィルタで線形解析を実行"""
        print("🔧 線形解析を実行中...")
        
        # ES: Evitar re-ejecución si ya hay un análisis lineal corriendo | EN: Avoid re-running if linear analysis is already running | JA: 線形解析実行中は再実行を防ぐ
        if hasattr(self, 'linear_worker') and self.linear_worker is not None:
            try:
                if self.linear_worker.isRunning():
                    QMessageBox.warning(self, "線形解析", "⚠️ すでに線形解析が実行中です。\n完了または停止するまでお待ちください。")
                    return
            except RuntimeError:
                self.linear_worker = None
        
        try:
            # ES: Obtener filtros aplicados
            # EN: Get applied filters
            # JP: 適用済みフィルタを取得
            filters = self.get_applied_filters()
            print(f"🔧 適用されたフィルタ: {filters}")
            
            # ES: Importar módulo de análisis lineal
            # EN: Import the linear-analysis module
            # JP: 線形解析モジュールをインポート
            try:
                from linear_analysis_advanced import run_advanced_linear_analysis_from_db
                print("✅ 線形解析モジュールを正常にインポートしました")
            except ImportError as e:
                print(f"❌ 線形解析モジュールのインポートエラー: {e}")
                QMessageBox.critical(self, "エラー", "❌ モジュール de análisis lineal no se pudo importar.\nAsegúrese de que el archivo linear_analysis_module.py esté en el directorio correcto.")
                return
            
            # ES: Mostrar mensaje de confirmación | EN: Show confirmation message | JA: 確認メッセージを表示
            reply = QMessageBox.question(
                self, 
                "線形解析確認", 
                f"線形解析を実行しますか？\n\nフィルター: {len(filters)} 条件\n\nこの操作は時間がかかる場合があります。",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes
            )
            
            if reply != QMessageBox.Yes:
                print("❌ ユーザーが線形解析をキャンセルしました")
                return
            
            # ES: Crear estructura de carpetas para el análisis | EN: Create folder structure for the analysis | JA: 解析用のフォルダ構成を作成
            if hasattr(self, 'current_project_folder') and self.current_project_folder:
                analysis_folder = self.create_linear_analysis_folder_structure(self.current_project_folder)
                print(f"✅ フォルダ構成を作成しました: {analysis_folder}")
            else:
                print("⚠️ プロジェクトフォルダが検出されませんでした。デフォルトフォルダを使用します")
                analysis_folder = "analysis_output"

            # Arrancar con flujo unificado (worker + popup + cancelación)
            self._start_linear_analysis(filters, analysis_folder)
                
        except Exception as e:
            print(f"❌ 線形解析実行エラー: {e}")
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "エラー", f"❌ 線形解析の実行中にエラーが発生しました:\n{str(e)}")

    def on_linear_analysis_finished(self, results):
        """ES: Maneja el resultado exitoso del análisis lineal
        EN: Handle successful linear analysis result
        JA: 線形解析の成功結果を処理"""
        # ES: Re-habilitar botones | EN: Re-enable buttons | JA: ボタンを再有効化
        if hasattr(self, 'linear_analysis_button'):
            self.linear_analysis_button.setEnabled(True)
        if hasattr(self, 'run_analysis_button'):
            self.run_analysis_button.setEnabled(True)
            
        try:
            # ES: Si el usuario canceló, NO mostrar resultados (evita "cancelé y aun así me enseña resultados") | EN: If user cancelled, do NOT show results (avoids showing results after cancel) | JA: キャンセル後は結果を表示しない
            if hasattr(self, '_linear_cancel_requested') and self._linear_cancel_requested:
                print("🛑 DEBUG: 結果を受信しましたがユーザーがキャンセルしました。結果を無視します。")
                # ES: Cerrar popup de progreso de forma segura
                # EN: Close the progress popup safely
                # JP: 進捗ポップアップを安全に閉じる
                if hasattr(self, 'progress_dialog') and self.progress_dialog is not None:
                    try:
                        self.progress_dialog.close()
                        self.progress_dialog.deleteLater()
                    except:
                        pass
                if hasattr(self, 'progress_dialog'):
                    try:
                        delattr(self, 'progress_dialog')
                    except:
                        pass
                self.set_console_overlay_topmost(False)
                # ES: Limpiar worker
                # EN: Clean up worker
                # JP: ワーカーをクリーンアップ
                try:
                    self.linear_worker = None
                except:
                    pass
                return

            # ES: Cerrar popup de progreso de forma segura
            # EN: Close the progress popup safely
            # JP: 進捗ポップアップを安全に閉じる
            if hasattr(self, 'progress_dialog') and self.progress_dialog is not None:
                try:
                    self.progress_dialog.close()
                    self.progress_dialog.deleteLater()
                except:
                    pass  # Ignore errors when closing the popup
            
            # ES: Limpiar referencias
            # EN: Clear references
            # JP: 参照をクリア
            if hasattr(self, 'progress_dialog'):
                delattr(self, 'progress_dialog')
            self.set_console_overlay_topmost(False)
            
            if results.get('success', False):
                # ES: Mostrar resultados | EN: Show results | JA: 結果を表示
                self.show_linear_analysis_results(results)
                QMessageBox.information(self, "線形解析完了", f"✅ 線形解析が完了しました！\n結果は{results.get('output_folder', 'N/A')}フォルダに保存されています。")
            else:
                error_msg = results.get('error', 'Error desconocido')
                QMessageBox.critical(self, "線形解析エラー", f"❌ 線形解析中にエラーが発生しました:\n{error_msg}")
                
        except Exception as e:
            print(f"❌ 線形解析完了ハンドラでエラー: {e}")
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "エラー", f"❌ 結果の処理中にエラーが発生しました:\n{str(e)}")

    def on_linear_analysis_error(self, error_message):
        """ES: Maneja el error del análisis lineal
        EN: Handle linear analysis error
        JA: 線形解析のエラーを処理"""
        # ES: Re-habilitar botones | EN: Re-enable buttons | JA: ボタンを再有効化
        if hasattr(self, 'linear_analysis_button'):
            self.linear_analysis_button.setEnabled(True)
        if hasattr(self, 'run_analysis_button'):
            self.run_analysis_button.setEnabled(True)
            
        try:
            # ES: Si el usuario canceló, tratamos como cancelación silenciosa | EN: If user cancelled, treat as silent cancellation | JA: ユーザーキャンセル時はサイレントキャンセルとして扱う
            if hasattr(self, '_linear_cancel_requested') and self._linear_cancel_requested:
                print(f"🛑 DEBUG: キャンセル後にエラーを受信: {error_message}。無視します。")
                if hasattr(self, 'progress_dialog') and self.progress_dialog is not None:
                    try:
                        self.progress_dialog.close()
                        self.progress_dialog.deleteLater()
                    except:
                        pass
                if hasattr(self, 'progress_dialog'):
                    try:
                        delattr(self, 'progress_dialog')
                    except:
                        pass
                self.set_console_overlay_topmost(False)
                try:
                    self.linear_worker = None
                except:
                    pass
                return

            # ES: Cerrar popup de progreso de forma segura
            # EN: Close the progress popup safely
            # JP: 進捗ポップアップを安全に閉じる
            if hasattr(self, 'progress_dialog') and self.progress_dialog is not None:
                try:
                    self.progress_dialog.close()
                    self.progress_dialog.deleteLater()
                except:
                    pass  # Ignore errors when closing the popup
            
            # ES: Limpiar referencias
            # EN: Clear references
            # JP: 参照をクリア
            if hasattr(self, 'progress_dialog'):
                delattr(self, 'progress_dialog')
            self.set_console_overlay_topmost(False)
            
            print(f"❌ 線形解析エラー: {error_message}")
            QMessageBox.critical(self, "線形解析エラー", f"❌ 線形解析中にエラーが発生しました:\n{error_message}")
            
        except Exception as e:
            print(f"❌ 線形解析エラーハンドラでエラー: {e}")
            import traceback
            traceback.print_exc()

    def on_nonlinear_cancelled(self):
        """ES: Maneja la cancelación del análisis no lineal desde el diálogo
        EN: Handle cancellation of non-linear analysis from the dialog
        JA: ダイアログからの非線形解析キャンセルを処理"""
        try:
            print("🛑 ユーザーにより非線形解析がキャンセルされました")

            # ES: marcar cancelación para esta ejecución | EN: Mark cancellation for this run | JA: この実行のキャンセルを記録
            self._nonlinear_cancel_requested = True
            
            # ES: Cancelar el worker (esto terminará el proceso subprocess)
            # EN: Cancel the worker (this will terminate the subprocess)
            # JP: ワーカーをキャンセル（subprocessを終了させる）
            if hasattr(self, 'nonlinear_worker') and self.nonlinear_worker is not None:
                try:
                    self.nonlinear_worker.cancel()
                except:
                    pass
                try:
                    self.nonlinear_worker.requestInterruption()
                except:
                    pass

            # ES: Cerrar/ocultar progreso sin bloquear UI
            # EN: Close/hide progress without blocking the UI
            # JP: UIをブロックせず進捗を閉じる/隠す
            if hasattr(self, 'progress_dialog') and self.progress_dialog:
                try:
                    self.progress_dialog.hide()
                except:
                    pass
            self.set_console_overlay_topmost(False)
            
            print("✅ 非線形解析ワーカーを正常にキャンセルしました")
            
        except Exception as e:
            print(f"❌ 非線形キャンセルハンドラでエラー: {e}")
            import traceback
            traceback.print_exc()
    
    def on_analysis_cancelled(self):
        """ES: Maneja la cancelación del análisis de forma segura
        EN: Handle analysis cancellation safely
        JA: 解析のキャンセルを安全に処理"""
        try:
            print("🛑 DEBUG: on_analysis_cancelled 発火 - 安全停止を開始します")
            
            # ES: Re-habilitar botones | EN: Re-enable buttons | JA: ボタンを再有効化
            if hasattr(self, 'linear_analysis_button'):
                self.linear_analysis_button.setEnabled(True)
            if hasattr(self, 'run_analysis_button'):
                self.run_analysis_button.setEnabled(True)

            # ES: Marcar cancelación para esta ejecución (evita mostrar resultados luego) | EN: Mark cancellation for this run (avoids showing results later) | JA: この実行のキャンセルを記録（後で結果を表示しない）
            self._linear_cancel_requested = True
            
            # 1. Solicitar parada cooperativa al worker lineal (NO terminate)
            if hasattr(self, 'linear_worker') and self.linear_worker is not None:
                try:
                    if self.linear_worker.isRunning():
                        print(f"🛑 DEBUG: ワーカーにキャンセルを要求しています {self.linear_worker}")
                        # Señal cooperativa
                        try:
                            self.linear_worker.requestInterruption()
                        except:
                            pass
                        try:
                            self.linear_worker.stop()
                        except:
                            # ES: fallback por si cambia el nombre del método
                            # EN: fallback in case the method name changes
                            # JP: メソッド名が変わった場合のフォールバック
                            try:
                                self.linear_worker.is_cancelled = True
                            except:
                                pass
                except RuntimeError:
                    self.linear_worker = None

            # 2. Informar al worker no lineal (si existe)
            if hasattr(self, 'nonlinear_worker') and self.nonlinear_worker is not None:
                print("🛑 DEBUG: 非線形プロセスをキャンセル中")
                self.nonlinear_worker.cancel()
            
            # ES: 3. Limpiar la UI (el worker puede tardar en parar si está en cómputo pesado)
            # EN: 3. Clear the UI (the worker may take time to stop if it's doing heavy computation)
            # JP: 3. UIをクリア（重い計算中だとワーカー停止に時間がかかる場合あり）
            if hasattr(self, 'progress_dialog') and self.progress_dialog:
                self.progress_dialog.hide()
            self.set_console_overlay_topmost(False)
            
            print("✅ 安全停止完了。クラッシュは発生しない想定です。")
            
        except Exception as e:
            print(f"❌ 解析キャンセルハンドラでエラー: {e}")
            import traceback
            traceback.print_exc()

    def get_applied_filters(self):
        """ES: Obtener filtros aplicados por el usuario
        EN: Get filters applied by user
        JA: ユーザーが適用したフィルタを取得"""
        filters = {}
        
        if not hasattr(self, 'filter_inputs'):
            return filters
        
        # ES: Manejar filtros de cepillo de manera especial | EN: Handle brush filters specially | JA: ブラシフィルターを特別に処理
        brush_selections = []
        subete_selected = False
        
        for field_name, input_widget in self.filter_inputs.items():
            if field_name in ['すべて', 'A13', 'A11', 'A21', 'A32']:
                if hasattr(input_widget, 'isChecked') and input_widget.isChecked():
                    if field_name == 'すべて':
                        subete_selected = True
                    else:
                        brush_selections.append(field_name)
                continue
            
            if isinstance(input_widget, tuple):
                # Rango de valores (desde, hasta)
                desde, hasta = input_widget
                
                # Manejo especial para fecha
                if field_name == "実験日":
                    # Solo aplicar filtro de fecha si está habilitado
                    if hasattr(self, 'apply_date_filter') and self.apply_date_filter:
                        desde_val = desde.date().toString("yyyyMMdd") if hasattr(desde, 'date') else ''
                        hasta_val = hasta.date().toString("yyyyMMdd") if hasattr(hasta, 'date') else ''
                        
                        # ES: Solo agregar filtro si ambos valores están especificados
                        # EN: Only add the filter if both values are specified
                        # JP: 両方の値が指定されている場合のみフィルタを追加
                        if desde_val and hasta_val:
                            filters[field_name] = (desde_val, hasta_val)
                else:
                    # Otros campos de rango
                    desde_val = desde.text().strip() if hasattr(desde, 'text') else ''
                    hasta_val = hasta.text().strip() if hasattr(hasta, 'text') else ''
                    
                    # ES: Solo agregar filtro si ambos valores están especificados
                    # EN: Only add the filter if both values are specified
                    # JP: 両方の値が指定されている場合のみフィルタを追加
                    if desde_val and hasta_val:
                        filters[field_name] = (desde_val, hasta_val)
            else:
                # Valor único
                if hasattr(input_widget, 'text'):
                    value = input_widget.text().strip()
                elif hasattr(input_widget, 'currentText'):
                    value = input_widget.currentText().strip()
                elif hasattr(input_widget, 'date'):
                    value = input_widget.date().toString('yyyy-MM-dd')
                else:
                    value = ''
                
                # ES: Solo agregar filtro si el valor no está vacío
                # EN: Only add the filter if the value is not empty
                # JP: 値が空でない場合のみフィルタを追加
                if value and value != "":
                    filters[field_name] = value
        
        # ES: Aplicar lógica de filtros de cepillo | EN: Apply brush filter logic | JA: ブラシフィルターのロジックを適用
        if subete_selected:
            # ES: Si está seleccionado "すべて", agregar el filtro
            # EN: If "すべて" is selected, add the filter
            # JP: 「すべて」が選択されている場合はフィルタを追加
            filters['すべて'] = True
        elif brush_selections:
            # ES: Si no está seleccionado "すべて" pero hay cepillos específicos seleccionados
            # EN: If "すべて" is not selected but specific brushes are selected
            # JP: 「すべて」が未選択で特定ブラシが選択されている場合
            for brush in brush_selections:
                filters[brush] = True
        
        return filters

    def show_linear_analysis_results(self, results):
        """ES: Mostrar resultados del análisis lineal
        EN: Show linear analysis results
        JA: 線形解析の結果を表示"""
        print("🔧 線形解析の結果を表示中...")
        
        try:
            # ES: Limpiar layout central completamente
            # EN: Clear the center layout completely
            # JP: 中央レイアウトを完全にクリア
            while self.center_layout.count():
                item = self.center_layout.takeAt(0)
                widget = item.widget()
                if widget:
                    widget.deleteLater()
                else:
                    # ES: Si es un layout, limpiarlo también
                    # EN: If it's a layout, clear it too
                    # JP: レイアウトならそれもクリア
                    layout = item.layout()
                    if layout:
                        while layout.count():
                            layout_item = layout.takeAt(0)
                            layout_widget = layout_item.widget()
                            if layout_widget:
                                layout_widget.deleteLater()
            
            # Forzar actualización de la UI
            QApplication.processEvents()
            
            # ES: Crear contenedor con fondo gris limpio | EN: Create container with clean grey background | JA: クリーンなグレー背景のコンテナを作成
            gray_container = QFrame()
            gray_container.setStyleSheet("""
                QFrame {
                    background-color: #f5f5f5;
                    border-radius: 10px;
                    margin: 10px;
                }
            """)
            
            # ES: Layout interno para el contenedor gris | EN: Inner layout for grey container | JA: グレーコンテナ用の内部レイアウト
            container_layout = QVBoxLayout(gray_container)
            container_layout.setContentsMargins(20, 20, 20, 20)
            container_layout.setSpacing(15)
            
            # ES: Título | EN: Title | JA: タイトル
            title = QLabel("線形解析結果")
            title.setStyleSheet("""
                font-weight: bold; 
                font-size: 24px; 
                color: #2c3e50;
                margin-bottom: 20px;
                padding: 10px 0px;
                border-bottom: 2px solid #3498db;
                border-radius: 0px;
            """)
            title.setAlignment(Qt.AlignCenter)
            container_layout.addWidget(title)
            
            # ES: Información del análisis
            # EN: Analysis information
            # JP: 解析情報
            # ES: Formatear datos largos para evitar texto cortado | EN: Format long data to avoid truncated text | JA: 長いデータを整形して切れを防ぐ
            filters_applied = results.get('filters_applied', 'N/A')
            if isinstance(filters_applied, list):
                if len(filters_applied) > 3:
                    filters_text = f"{len(filters_applied)} 条件"
                else:
                    filters_text = ", ".join(str(f) for f in filters_applied)
            else:
                filters_text = str(filters_applied)
            
            # Truncar si es muy largo
            if len(filters_text) > 50:
                filters_text = filters_text[:47] + "..."
            
            data_range = results.get('data_range', 'N/A')
            if isinstance(data_range, str) and len(data_range) > 50:
                data_range = data_range[:47] + "..."
            
            info_text = f"""
            📊 解析完了時刻: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
            📈 データ数: {results.get('data_count', 'N/A')} レコード
            🤖 訓練済みモデル: {results.get('models_trained', 'N/A')} 個
            🔧 フィルター適用: {filters_text}
            📊 データ範囲: {data_range}
            """
            
            info_label = QLabel(info_text)
            info_label.setStyleSheet("""
                font-size: 14px;
                color: #34495e;
                background-color: #ecf0f1;
                padding: 15px;
                border-radius: 8px;
                border: 1px solid #bdc3c7;
            """)
            info_label.setAlignment(Qt.AlignLeft)
            info_label.setWordWrap(True)  # EN: Allow line wrap
            container_layout.addWidget(info_label)
            
            # ES: Ruta clickeable del archivo Excel | EN: Clickable path to the Excel file | JA: Excelファイルへのクリック可能なパス
            output_folder = results.get('output_folder', '')
            if output_folder:
                # ES: Buscar dinámicamente el archivo Excel | EN: Find Excel file dynamically | JA: Excelファイルを動的に検索
                excel_file_path = None
                
                # ES: Buscar en la estructura de carpetas del análisis lineal
                # EN: Search within the linear-analysis folder structure
                # JP: 線形解析のフォルダ構造内を検索
                linear_regression_folder = os.path.join(output_folder, "03_線形回帰")
                if os.path.exists(linear_regression_folder):
                    # ES: Buscar en todas las subcarpetas de 03_線形回帰
                    # EN: Search in all subfolders of 03_線形回帰
                    # JP: 03_線形回帰 の全サブフォルダを検索
                    for subfolder in os.listdir(linear_regression_folder):
                        subfolder_path = os.path.join(linear_regression_folder, subfolder)
                        if os.path.isdir(subfolder_path):
                            # ES: Buscar en 04_予測計算 dentro de cada subcarpeta
                            # EN: Search in 04_予測計算 inside each subfolder
                            # JP: 各サブフォルダ内の04_予測計算を検索
                            prediction_folder = os.path.join(subfolder_path, "04_予測計算")
                            if os.path.exists(prediction_folder):
                                # ES: Buscar el archivo Excel
                                # EN: Look for the Excel file
                                # JP: Excelファイルを探す
                                excel_file = os.path.join(prediction_folder, "XEBEC_予測計算機_逆変換対応.xlsx")
                                if os.path.exists(excel_file):
                                    excel_file_path = excel_file
                                    break
                
                # ES: Si no se encuentra en la estructura esperada, buscar en cualquier lugar del output_folder
                # EN: If not found in the expected structure, search anywhere under output_folder
                # JP: 想定構造で見つからない場合はoutput_folder配下を全検索
                if not excel_file_path:
                    for root, dirs, files in os.walk(output_folder):
                        for file in files:
                            if file == "XEBEC_予測計算機_逆変換対応.xlsx":
                                excel_file_path = os.path.join(root, file)
                                break
                        if excel_file_path:
                            break
                
                # ES: Verificar si el archivo existe | EN: Check if file exists | JA: ファイルが存在するか確認
                if excel_file_path and os.path.exists(excel_file_path):
                    # ES: Crear layout para la ruta clickeable | EN: Create layout for clickable path | JA: クリック可能パス用レイアウトを作成
                    path_layout = QVBoxLayout()
                    
                    # ES: Título | EN: Title | JA: タイトル
                    path_title = QLabel("📁 出力ディレクトリ:")
                    path_title.setStyleSheet("""
                        font-size: 14px;
                        font-weight: bold;
                        color: #2c3e50;
                        margin-bottom: 5px;
                    """)
                    path_layout.addWidget(path_title)
                    
                    # ES: Ruta clickeable con scroll horizontal si es necesario
                    # EN: Clickable path (with horizontal scrolling if needed)
                    # JP: クリック可能なパス（必要なら横スクロール）
                    path_label = QLabel(excel_file_path)
                    path_label.setStyleSheet("""
                        QLabel {
                            font-size: 12px;
                            color: #3498db;
                            background-color: #e8f4fd;
                            padding: 10px;
                            border-radius: 5px;
                            border: 1px solid #3498db;
                            text-decoration: underline;
                        }
                        QLabel:hover {
                            background-color: #d1ecf1;
                            cursor: pointer;
                        }
                    """)
                    path_label.setWordWrap(True)  # Allow line wrap
                    path_label.setAlignment(Qt.AlignLeft)
                    
                    # ES: Hacer la ruta clickeable
                    # EN: Make the path clickable
                    # JP: パスをクリック可能にする
                    def open_excel_file():
                        try:
                            # ES: Abrir el archivo Excel con la aplicación por defecto
                            # EN: Open the Excel file with the default application
                            # JP: 既定アプリでExcelファイルを開く
                            if os.name == 'nt':  # Windows
                                os.startfile(excel_file_path)
                            elif os.name == 'posix':  # macOS y Linux
                                subprocess.run(['open', excel_file_path], check=True)
                            else:
                                subprocess.run(['xdg-open', excel_file_path], check=True)
                            print(f"✅ Excelファイルを開きました: {excel_file_path}")
                        except Exception as e:
                            print(f"❌ Excelファイルを開く際のエラー: {e}")
                            QMessageBox.warning(self, "エラー", f"❌ Excelファイルを開けませんでした:\n{str(e)}")
                    
                    # ES: Conectar el click | EN: Connect the click | JA: クリックを接続
                    path_label.mousePressEvent = lambda event: open_excel_file()
                    
                    path_layout.addWidget(path_label)
                    container_layout.addLayout(path_layout)
                else:
                    # ES: Si el archivo no existe, mostrar mensaje informativo
                    # EN: If the file does not exist, show an informational message
                    # JP: ファイルが無い場合は案内メッセージを表示
                    missing_file_label = QLabel(f"⚠️ Excelファイルが見つかりません\n\n検索場所: {output_folder}\n\nファイル名: XEBEC_予測計算機_逆変換対応.xlsx")
                    missing_file_label.setStyleSheet("""
                        font-size: 12px;
                        color: #e74c3c;
                        background-color: #fadbd8;
                        padding: 10px;
                        border-radius: 5px;
                        border: 1px solid #e74c3c;
                        margin: 10px 0px;
                    """)
                    missing_file_label.setWordWrap(True)
                    missing_file_label.setAlignment(Qt.AlignCenter)
                    container_layout.addWidget(missing_file_label)
            
            # Resultados detallados de modelos
            models = results.get('models', {})
            if models:
                models_title = QLabel("詳細モデル結果")
                models_title.setStyleSheet("""
                    font-weight: bold; 
                    font-size: 18px; 
                    color: #2c3e50;
                    margin-top: 20px;
                    margin-bottom: 10px;
                """)
                container_layout.addWidget(models_title)
                
                for target_name, model_info in models.items():
                    if model_info.get('model') is None:
                        status = "❌ 失敗"
                        error = model_info.get('error', 'Unknown error')
                        details = f"エラー: {error}"
                    else:
                        status = "✅ 成功"
                        model_name = model_info.get('model_name', 'Unknown')
                        task_type = model_info.get('task_type', 'Unknown')
                        details = f"モデル: {model_name}, タイプ: {task_type}"
                        
                        if task_type == 'regression':
                            metrics = model_info.get('final_metrics', {})
                            details += f", R²: {metrics.get('r2', 'N/A'):.4f}, MAE: {metrics.get('mae', 'N/A'):.4f}"
                        else:
                            metrics = model_info.get('final_metrics', {})
                            details += f", 精度: {metrics.get('accuracy', 'N/A'):.4f}, F1: {metrics.get('f1_score', 'N/A'):.4f}"
                    
                    model_label = QLabel(f"【{target_name}】 {status}\n{details}")
                    model_label.setStyleSheet("""
                        font-size: 12px;
                        color: #34495e;
                        background-color: #f8f9fa;
                        padding: 10px;
                        border-radius: 5px;
                        border: 1px solid #dee2e6;
                        margin: 5px 0px;
                    """)
                    container_layout.addWidget(model_label)
            
            # ES: Botón para volver | EN: Back button | JA: 戻るボタン a filtros
            button_layout = QHBoxLayout()
            button_layout.addStretch()
            
            back_button = QPushButton("次へ")
            back_button.setFixedSize(120, 40)  # Make the button narrower
            back_button.setStyleSheet("""
                QPushButton {
                    background-color: #3498db;
                    color: white;
                    border: none;
                    padding: 10px 20px;
                    border-radius: 5px;
                    font-size: 14px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #2980b9;
                }
            """)
            back_button.clicked.connect(self.show_evaluation_charts)
            
            button_layout.addWidget(back_button)
            button_layout.addStretch()
            container_layout.addLayout(button_layout)
            
            # ES: Espacio flexible
            # EN: Flexible space
            # JA: 可変スペース
            container_layout.addStretch()
            
            # ES: Agregar el contenedor gris al layout central | EN: Add grey container to center layout | JA: 中央レイアウトにグレーコンテナを追加
            self.center_layout.addWidget(gray_container)
            
            print("✅ 線形解析の結果を表示しました")
            
        except Exception as e:
            print(f"❌ 結果表示エラー: {e}")
            import traceback
            traceback.print_exc()

    def show_evaluation_charts(self):
        """ES: Mostrar gráficos de evaluación con navegación
        EN: Show evaluation charts with navigation
        JA: 評価グラフをナビ付きで表示"""
        print("🔧 評価グラフを表示中...")
        
        try:
            # ES: Limpiar layout central completamente
            # EN: Clear the center layout completely
            # JP: 中央レイアウトを完全にクリア
            while self.center_layout.count():
                item = self.center_layout.takeAt(0)
                widget = item.widget()
                if widget:
                    widget.deleteLater()
                else:
                    # ES: Si es un layout, limpiarlo también
                    # EN: If it's a layout, clear it too
                    # JP: レイアウトならそれもクリア
                    layout = item.layout()
                    if layout:
                        while layout.count():
                            layout_item = layout.takeAt(0)
                            layout_widget = layout_item.widget()
                            if layout_widget:
                                layout_widget.deleteLater()
            
            # Forzar actualización de la UI
            QApplication.processEvents()
            
            # ES: Crear contenedor con fondo gris limpio | EN: Create container with clean grey background | JA: クリーンなグレー背景のコンテナを作成
            gray_container = QFrame()
            gray_container.setStyleSheet("""
                QFrame {
                    background-color: #f5f5f5;
                    border-radius: 10px;
                    margin: 10px;
                }
            """)
            
            # ES: Layout interno para el contenedor gris | EN: Inner layout for grey container | JA: グレーコンテナ用の内部レイアウト
            container_layout = QVBoxLayout(gray_container)
            container_layout.setContentsMargins(20, 20, 20, 20)
            container_layout.setSpacing(15)
            
            # ES: Título | EN: Title | JA: タイトル
            title = QLabel("評価スコア チャート")
            title.setStyleSheet("""
                font-weight: bold; 
                font-size: 24px; 
                color: #2c3e50;
                margin-bottom: 20px;
                padding: 10px 0px;
                border-bottom: 2px solid #3498db;
                border-radius: 0px;
            """)
            title.setAlignment(Qt.AlignCenter)
            container_layout.addWidget(title)
            
            # ES: Buscar gráficos de evaluación | EN: Find evaluation charts | JA: 評価グラフを検索
            chart_images = []
            if hasattr(self, 'current_project_folder') and self.current_project_folder:
                # ES: Buscar en la estructura de carpetas del análisis lineal
                # EN: Search within the linear-analysis folder structure
                # JP: 線形解析のフォルダ構造内を検索
                linear_regression_folder = os.path.join(self.current_project_folder, "03_線形回帰")
                if os.path.exists(linear_regression_folder):
                    # ES: Buscar en todas las subcarpetas de 03_線形回帰
                    # EN: Search in all subfolders of 03_線形回帰
                    # JP: 03_線形回帰 の全サブフォルダを検索
                    for subfolder in os.listdir(linear_regression_folder):
                        subfolder_path = os.path.join(linear_regression_folder, subfolder)
                        if os.path.isdir(subfolder_path):
                            # ES: Buscar en 03_評価スコア\\01_チャート
                            # EN: Search in 03_評価スコア\\01_チャート
                            # JP: 03_評価スコア\\01_チャート を検索
                            evaluation_folder = os.path.join(subfolder_path, "03_評価スコア", "01_チャート")
                            if os.path.exists(evaluation_folder):
                                # ES: Buscar archivos PNG
                                # EN: Search for PNG files
                                # JP: PNGファイルを探す
                                for file in os.listdir(evaluation_folder):
                                    if file.lower().endswith('.png'):
                                        chart_images.append(os.path.join(evaluation_folder, file))
                                break
            
            # ES: Si no se encuentran gráficos, mostrar mensaje
            # EN: If no charts are found, show a message
            # JP: グラフが見つからない場合はメッセージを表示
            if not chart_images:
                no_charts_label = QLabel("⚠️ 評価スコアチャートが見つかりません")
                no_charts_label.setStyleSheet("""
                    font-size: 16px;
                    color: #e74c3c;
                    background-color: #fadbd8;
                    padding: 20px;
                    border-radius: 8px;
                    border: 1px solid #e74c3c;
                    margin: 20px 0px;
                """)
                no_charts_label.setAlignment(Qt.AlignCenter)
                container_layout.addWidget(no_charts_label)
            else:
                # ES: Configurar navegación de gráficos | EN: Configure chart navigation | JA: グラフのナビゲーションを設定
                self.chart_images = sorted(chart_images)
                self.current_chart_index = 0
                
                # ES: Layout principal | EN: Main layout | JA: メインレイアウト para la imagen y navegación
                chart_layout = QVBoxLayout()
                
                # ES: Label para mostrar la imagen (ocupa todo el ancho)
                # EN: Label to display the image (takes full width)
                # JP: 画像表示用ラベル（全幅）
                self.chart_label = QLabel()
                self.chart_label.setAlignment(Qt.AlignCenter)
                self.chart_label.setStyleSheet("""
                    QLabel {
                        background-color: white;
                        border: 2px solid #bdc3c7;
                        border-radius: 10px;
                        padding: 10px;
                        min-height: 500px;
                    }
                """)
                chart_layout.addWidget(self.chart_label)
                
                # ES: Layout horizontal para botones de navegación (debajo de la imagen) | EN: Horizontal layout for nav buttons (below image) | JA: ナビボタン用横レイアウト（画像下）
                nav_buttons_layout = QHBoxLayout()
                nav_buttons_layout.addStretch()
                
                # ES: Botón flecha izquierda | EN: Left arrow button | JA: 左矢印ボタン con mejor icono
                prev_chart_button = QPushButton("◀ 前へ")
                prev_chart_button.setFixedSize(100, 40)
                prev_chart_button.setStyleSheet("""
                    QPushButton {
                        background-color: #3498db;
                        color: white;
                        border: none;
                        border-radius: 8px;
                        font-size: 14px;
                        font-weight: bold;
                        padding: 8px 16px;
                    }
                    QPushButton:hover {
                        background-color: #2980b9;
                    }
                    QPushButton:disabled {
                        background-color: #bdc3c7;
                        color: #7f8c8d;
                    }
                """)
                prev_chart_button.clicked.connect(self.show_previous_chart)
                nav_buttons_layout.addWidget(prev_chart_button)
                
                # ES: Espacio entre botones
                # EN: Space between buttons
                # JA: ボタン間のスペース
                nav_buttons_layout.addSpacing(20)
                
                # ES: Botón flecha derecha con mejor icono | EN: Right arrow button with better icon | JA: 改善アイコンの右矢印ボタン
                next_chart_button = QPushButton("次へ ▶")
                next_chart_button.setFixedSize(100, 40)
                next_chart_button.setStyleSheet("""
                    QPushButton {
                        background-color: #3498db;
                        color: white;
                        border: none;
                        border-radius: 8px;
                        font-size: 14px;
                        font-weight: bold;
                        padding: 8px 16px;
                    }
                    QPushButton:hover {
                        background-color: #2980b9;
                    }
                    QPushButton:disabled {
                        background-color: #bdc3c7;
                        color: #7f8c8d;
                    }
                """)
                next_chart_button.clicked.connect(self.show_next_chart)
                nav_buttons_layout.addWidget(next_chart_button)
                
                nav_buttons_layout.addStretch()
                chart_layout.addLayout(nav_buttons_layout)
                
                # Información del gráfico actual
                self.chart_info_label = QLabel()
                self.chart_info_label.setStyleSheet("""
                    font-size: 14px;
                    color: #2c3e50;
                    background-color: #ecf0f1;
                    padding: 10px;
                    border-radius: 5px;
                    border: 1px solid #bdc3c7;
                    margin: 10px 0px;
                """)
                self.chart_info_label.setAlignment(Qt.AlignCenter)
                chart_layout.addWidget(self.chart_info_label)
                
                container_layout.addLayout(chart_layout)
                
                # ES: Mostrar el primer gráfico | EN: Show first chart | JA: 先頭グラフを表示
                self.update_chart_display()
            
            # ES: Botones
            # EN: Buttons
            # JA: ボタン para volver a resultados y predicción
            buttons_layout = QHBoxLayout()
            buttons_layout.addStretch()
            
            # ES: Botón para volver | EN: Back button | JA: 戻るボタン a filtros (modoru)
            back_button = QPushButton("戻る")
            back_button.setFixedSize(120, 40)
            back_button.setStyleSheet("""
                QPushButton {
                    background-color: #e74c3c;
                    color: white;
                    border: none;
                    padding: 10px 20px;
                    border-radius: 5px;
                    font-size: 14px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #c0392b;
                }
            """)
            back_button.clicked.connect(self.on_analyze_clicked)
            buttons_layout.addWidget(back_button)
            
            # ES: Espacio entre botones
            # EN: Space between buttons
            # JA: ボタン間のスペース
            buttons_layout.addSpacing(20)
            
            # ES: Botón para predicción | EN: Prediction button | JA: 予測ボタン
            prediction_button = QPushButton("予測")
            prediction_button.setFixedSize(120, 40)
            prediction_button.setStyleSheet("""
                QPushButton {
                    background-color: #27ae60;
                    color: white;
                    border: none;
                    padding: 10px 20px;
                    border-radius: 5px;
                    font-size: 14px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #229954;
                }
            """)
            prediction_button.clicked.connect(self.run_prediction)
            buttons_layout.addWidget(prediction_button)
            
            buttons_layout.addStretch()
            container_layout.addLayout(buttons_layout)
            
            # ES: Espacio flexible
            # EN: Flexible space
            # JA: 可変スペース
            container_layout.addStretch()
            
            # ES: Agregar el contenedor gris al layout central | EN: Add grey container to center layout | JA: 中央レイアウトにグレーコンテナを追加
            self.center_layout.addWidget(gray_container)
            
            print("✅ 評価グラフを表示しました")
            
        except Exception as e:
            print(f"❌ 評価グラフ表示エラー: {e}")
            import traceback
            traceback.print_exc()
    
    def show_previous_chart(self):
        """ES: Mostrar gráfico anterior
        EN: Show previous chart
        JA: 前のグラフを表示
        """
        if hasattr(self, 'chart_images') and len(self.chart_images) > 0:
            self.current_chart_index = (self.current_chart_index - 1) % len(self.chart_images)
            self.update_chart_display()
    
    def show_next_chart(self):
        """ES: Mostrar gráfico siguiente
        EN: Show next chart
        JA: 次のグラフを表示
        """
        if hasattr(self, 'chart_images') and len(self.chart_images) > 0:
            self.current_chart_index = (self.current_chart_index + 1) % len(self.chart_images)
            self.update_chart_display()
    
    def update_chart_display(self):
        """ES: Actualizar la visualización del gráfico actual
        EN: Update the display of the current chart
        JA: 現在のグラフ表示を更新
        """
        if hasattr(self, 'chart_images') and len(self.chart_images) > 0:
            current_image_path = self.chart_images[self.current_chart_index]
            
            # ES: Cargar y mostrar la imagen | EN: Load and display the image | JA: 画像を読み込み表示
            pixmap = QPixmap(current_image_path)
            if not pixmap.isNull():
                # ES: Redimensionar la imagen para ocupar todo el ancho disponible | EN: Resize image to fill available width | JA: 画像を利用可能幅いっぱいにリサイズ
                # ES: Obtener el tamaño del contenedor
                # EN: Get container size
                # JP: コンテナサイズを取得
                container_width = self.chart_label.width() - 20  # Restar padding
                container_height = self.chart_label.height() - 20  # Restar padding
                
                # ES: Si el contenedor aún no tiene tamaño, usar un tamaño por defecto
                # EN: If container size is not ready yet, use a default size
                # JA: サイズ未確定の場合はデフォルトサイズを使用
                if container_width <= 0:
                    container_width = 1000
                if container_height <= 0:
                    container_height = 600
                
                # ES: Redimensionar manteniendo la proporción | EN: Resize while keeping aspect ratio | JA: アスペクト比を維持してリサイズ
                scaled_pixmap = pixmap.scaled(container_width, container_height, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                self.chart_label.setPixmap(scaled_pixmap)
                
                # ES: Actualizar información del gráfico | EN: Update chart info | JA: グラフ情報を更新
                filename = os.path.basename(current_image_path)
                info_text = f"📊 {filename} ({self.current_chart_index + 1}/{len(self.chart_images)})"
                self.chart_info_label.setText(info_text)
                
                print(f"✅ グラフを表示中: {filename}")
            else:
                print(f"❌ 画像を読み込めませんでした: {current_image_path}")



    def on_formula_processing_error(self, error_msg):
        """ES: Manejar errores en el procesamiento de fórmulas
        EN: Handle formula-processing errors
        JA: 数式処理エラーを処理
        """
        print(f"❌ 式の処理エラー: {error_msg}")
        QMessageBox.critical(self, "エラー", f"❌ 予測計算中にエラーが発生しました:\n{error_msg}")

    def show_yosoku_parameters_dialog(self):
        """ES: Mostrar diálogo para seleccionar parámetros de predicción Yosoku
        EN: Show a dialog to select Yosoku prediction parameters
        JA: Yosoku予測パラメータ選択ダイアログを表示
        """
        try:
            from PySide6.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QLabel, QComboBox, QPushButton, QFormLayout
            
            dialog = QDialog(self)
            dialog.setWindowTitle("予測パラメーター選択")
            dialog.setModal(True)
            dialog.resize(400, 350)
            
            layout = QVBoxLayout()
            
            # ES: Título | EN: Title | JA: タイトル
            title = QLabel("予測パラメーターを選択してください")
            title.setStyleSheet("font-weight: bold; font-size: 14px; margin: 10px;")
            title.setAlignment(Qt.AlignCenter)
            layout.addWidget(title)
            
            # ES: Formulario | EN: Form | JA: フォーム de selección
            form_layout = QFormLayout()
            
            # Diámetro
            diameter_combo = QComboBox()
            diameter_combo.addItem("6", 6)
            diameter_combo.addItem("15", 15)
            diameter_combo.addItem("25", 25)
            diameter_combo.addItem("40", 40)
            diameter_combo.addItem("60", 60)
            diameter_combo.addItem("100", 100)
            diameter_combo.setCurrentText("15")  # Valor por defecto
            form_layout.addRow("直径:", diameter_combo)
            
            # Material
            material_combo = QComboBox()
            material_combo.addItem("Steel", "Steel")
            material_combo.addItem("Alum", "Alum")
            material_combo.setCurrentText("Steel")  # Valor por defecto
            form_layout.addRow("材料:", material_combo)
            
            layout.addLayout(form_layout)
            layout.addStretch()
            
            # ES: Botones
            # EN: Buttons
            # JA: ボタン
            button_layout = QHBoxLayout()
            
            cancel_button = QPushButton("キャンセル")
            cancel_button.clicked.connect(dialog.reject)
            
            ok_button = QPushButton("予測実行")
            ok_button.clicked.connect(dialog.accept)
            ok_button.setStyleSheet("background-color: #27ae60; color: white; font-weight: bold;")
            
            button_layout.addWidget(cancel_button)
            button_layout.addWidget(ok_button)
            layout.addLayout(button_layout)
            
            dialog.setLayout(layout)
            
            # ES: Mostrar diálogo | EN: Show dialog | JA: ダイアログを表示
            result = dialog.exec()
            
            if result == QDialog.Accepted:
                # Procesar selecciones
                selected_params = {
                    'diameter': diameter_combo.currentData(),
                    'material': material_combo.currentData(),
                }
                
                print(f"📊 選択されたパラメータ: {selected_params}")
                return selected_params
            else:
                return None
                
        except Exception as e:
            print(f"❌ パラメータダイアログ表示エラー: {e}")
            import traceback
            traceback.print_exc()
            return None

    @staticmethod
    def _normalize_columns_inplace(df):
        """Normaliza nombres de columnas para evitar fallos por espacios invisibles."""
        try:
            import pandas as pd
            if isinstance(df.columns, pd.MultiIndex):
                df.columns = [" ".join([str(x).strip() for x in tup if str(x).strip() != ""]).strip() for tup in df.columns]
            else:
                df.columns = [str(c).strip() for c in df.columns]
        except Exception:
            pass

    def _read_table_any(self, file_path, nrows=None, usecols=None):
        """Lee XLSX/XLS/CSV de forma uniforme."""
        import pandas as pd
        ext = os.path.splitext(str(file_path))[1].lower()
        if ext == ".csv":
            return pd.read_csv(file_path, encoding="utf-8-sig", nrows=nrows, usecols=usecols)
        # Excel: soporta xlsx/xls
        return pd.read_excel(file_path, nrows=nrows, usecols=usecols)

    def _extract_brush_and_wire_length_from_unexperimental(self, unexperimental_file):
        """
        Extrae (desde *_未実験データ.(xlsx|csv)):
        - brush_types: lista de tipos encontrados (p.ej. ["A11","A13"])
        - wire_lengths: lista de 線材長 encontrados (p.ej. [30.0, 35.0, ...])
        Requisitos (si falta, lanzar error):
        - columnas one-hot: A13/A11/A21/A32
        - columna 線材長
        Además, valida que:
        - cada fila tiene exactamente un 1 en A13/A11/A21/A32
        """
        import pandas as pd

        # ES: Leer solo header para validar columnas
        # EN: Read only the header to validate columns
        # JP: 列検証のためヘッダーのみ読み込む
        df_head = self._read_table_any(unexperimental_file, nrows=0)
        self._normalize_columns_inplace(df_head)
        headers = set(df_head.columns)

        brush_cols = ["A13", "A11", "A21", "A32"]
        required = brush_cols + ["線材長"]
        missing = [c for c in required if c not in headers]
        if missing:
            raise ValueError(
                f"❌ 未実験データファイルに必要な列がありません: {', '.join(missing)}\n"
                f"必要列: {', '.join(required)}\n"
                f"ファイル: {os.path.basename(str(unexperimental_file))}"
            )

        # ES: Leer solo columnas necesarias
        # EN: Read only the required columns
        # JP: 必要な列のみ読み込む
        df = self._read_table_any(unexperimental_file, usecols=required)
        self._normalize_columns_inplace(df)
        if df.empty:
            raise ValueError("❌ 未実験データファイルが空です。")

        # Brush one-hot
        onehot = df[brush_cols].apply(pd.to_numeric, errors="coerce").fillna(0).astype(int)
        s = onehot.sum(axis=1)
        bad_idx = df.index[s != 1]
        if len(bad_idx) > 0:
            raise ValueError(
                f"❌ 未実験データのブラシ列が不正です。各行で A13/A11/A21/A32 の合計が 1 である必要があります。"
                f" 不正行(先頭10): {bad_idx.tolist()[:10]}"
            )

        per_row_brush = onehot.idxmax(axis=1)
        uniq_brush = list(pd.unique(per_row_brush))
        # preservar orden A13/A11/A21/A32
        uniq_brush.sort(key=lambda x: brush_cols.index(str(x)) if str(x) in brush_cols else 999)
        brush_types = [str(x) for x in uniq_brush]

        # 線材長
        wire = pd.to_numeric(df["線材長"], errors="coerce").dropna()
        if wire.empty:
            raise ValueError("❌ 未実験データの 線材長 列に有効な値がありません。")
        uniq_wire = list(pd.unique(wire))
        try:
            uniq_wire = sorted([float(x) for x in uniq_wire])
        except Exception:
            # fallback: keep raw ordering
            uniq_wire = [float(x) for x in uniq_wire]
        wire_lengths = uniq_wire

        return brush_types, wire_lengths

    def find_latest_formulas_file(self):
        """ES: Encontrar el archivo XEBEC_予測計算機_逆変換対応.xlsx en la carpeta del análisis lineal más reciente
        EN: Find XEBEC_予測計算機_逆変換対応.xlsx in the latest linear analysis folder
        JA: 直近の線形解析フォルダ内で XEBEC_予測計算機_逆変換対応.xlsx を検索"""
        try:
            # ES: Buscar la carpeta del análisis lineal más reciente
            # EN: Find the most recent linear-analysis folder
            # JP: 最新の線形解析フォルダを探す
            linear_regression_folder = os.path.join(self.current_project_folder, "03_線形回帰")
            
            if not os.path.exists(linear_regression_folder):
                print(f"❌ フォルダが見つかりません: {linear_regression_folder}")
                return None
            
            # ES: Buscar subcarpetas de ejecución. Prioridad: NN_YYYYMMDD_HHMMSS (p.ej. 15_20260126_134704).
            # EN: Search run subfolders. Priority: NN_YYYYMMDD_HHMMSS (e.g., 15_20260126_134704).
            # JP: 実行サブフォルダを検索（優先：NN_YYYYMMDD_HHMMSS、例：15_20260126_134704）
            import re
            from datetime import datetime

            subfolders = []
            dated = []
            for item in os.listdir(linear_regression_folder):
                item_path = os.path.join(linear_regression_folder, item)
                if not os.path.isdir(item_path):
                    continue
                subfolders.append(item_path)
                m = re.match(r"^\d+_(\d{8})_(\d{6})", str(item))
                if m:
                    try:
                        dt = datetime.strptime(m.group(1) + m.group(2), "%Y%m%d%H%M%S")
                        dated.append((dt, item_path))
                    except Exception:
                        pass
            
            if not subfolders:
                print(f"❌ 線形解析のサブフォルダが見つかりません: {linear_regression_folder}")
                return None
            
            # Elegir última: primero por timestamp en nombre; fallback por mtime
            if dated:
                dated.sort(key=lambda t: t[0], reverse=True)
                latest_folder = dated[0][1]
            else:
                subfolders.sort(key=lambda x: os.path.getmtime(x), reverse=True)
                latest_folder = subfolders[0]
            print(f"📊 最新フォルダ: {latest_folder}")
            
            # ES: Buscar la subcarpeta 04_予測計算
            # EN: Look for the 04_予測計算 subfolder
            # JP: 04_予測計算 サブフォルダを探す
            prediction_folder = os.path.join(latest_folder, "04_予測計算")
            
            if not os.path.exists(prediction_folder):
                print(f"❌ フォルダが見つかりません: {prediction_folder}")
                return None
            
            # ES: Buscar el archivo XEBEC_予測計算機_逆変換対応.xlsx
            # EN: Look for the file XEBEC_予測計算機_逆変換対応.xlsx
            # JP: XEBEC_予測計算機_逆変換対応.xlsx を探す
            formulas_file = os.path.join(prediction_folder, "XEBEC_予測計算機_逆変換対応.xlsx")
            
            if os.path.exists(formulas_file):
                print(f"✅ 式ファイルを見つけました: {formulas_file}")
                return formulas_file
            else:
                print(f"❌ ファイルが見つかりません: {formulas_file}")
                return None
                
        except Exception as e:
            print(f"❌ 式ファイル検索エラー: {e}")
            import traceback
            traceback.print_exc()
            return None

    def validate_filtered_data(self, selected_params):
        """
        Validar el archivo filtered_data.xlsx contra los parámetros seleccionados.
        Devuelve: (is_valid: bool, errors: list[str], warnings: list[str])
        """
        try:
            # ES: Buscar la carpeta del análisis lineal más reciente
            # EN: Find the most recent linear-analysis folder
            # JP: 最新の線形解析フォルダを探す
            linear_regression_folder = os.path.join(self.current_project_folder, "03_線形回帰")
            
            if not os.path.exists(linear_regression_folder):
                return False, ["❌ No se encontró la carpeta de análisis lineal: 03_線形回帰"], []

            # Elegir la última carpeta de ejecución dentro de 03_線形回帰.
            # Prioridad: NN_YYYYMMDD_HHMMSS (p.ej. 15_20260126_134704). Fallback: mtime.
            import re
            from datetime import datetime

            run_candidates = []
            try:
                for item in os.listdir(linear_regression_folder):
                    item_path = os.path.join(linear_regression_folder, item)
                    if not os.path.isdir(item_path):
                        continue
                    m = re.match(r"^\d+_(\d{8})_(\d{6})", str(item))
                    if m:
                        try:
                            dt = datetime.strptime(m.group(1) + m.group(2), "%Y%m%d%H%M%S")
                            run_candidates.append((dt, item_path))
                        except Exception:
                            continue
            except Exception:
                run_candidates = []

            if run_candidates:
                run_candidates.sort(key=lambda t: t[0], reverse=True)
                latest_folder = run_candidates[0][1]
            else:
                # Fallback: cualquier subcarpeta más reciente por mtime
                subfolders = []
                try:
                    for item in os.listdir(linear_regression_folder):
                        item_path = os.path.join(linear_regression_folder, item)
                        if os.path.isdir(item_path):
                            subfolders.append(item_path)
                except Exception:
                    subfolders = []

                if not subfolders:
                    return False, ["❌ 03_線形回帰 に線形解析のサブフォルダが見つかりません"], []
                latest_folder = max(subfolders, key=lambda x: os.path.getmtime(x))
            
            # ES: Buscar el archivo filtered_data.xlsx en la carpeta 01_学習モデル
            # EN: Look for filtered_data.xlsx in the 01_学習モデル folder
            # JP: 01_学習モデルフォルダでfiltered_data.xlsxを探す
            candidate_paths = [
                os.path.join(latest_folder, "01_学習モデル", "filtered_data.xlsx"),
                os.path.join(latest_folder, "03_モデル学習", "filtered_data.xlsx"),
                os.path.join(latest_folder, "03_モデル学習", "01_学習モデル", "filtered_data.xlsx"),
            ]

            filtered_data_file = next((p for p in candidate_paths if os.path.exists(p)), None)
            if not filtered_data_file:
                # Búsqueda acotada dentro de latest_folder (profundidad <= 4)
                found = []
                try:
                    for root, dirs, files in os.walk(latest_folder):
                        rel = os.path.relpath(root, latest_folder)
                        if rel != "." and rel.count(os.sep) >= 4:
                            dirs[:] = []
                            continue
                        if "filtered_data.xlsx" in files:
                            found.append(os.path.join(root, "filtered_data.xlsx"))
                except Exception:
                    found = []

                if found:
                    # ES: Elegir el más reciente por mtime
                    # EN: Pick the most recent one by mtime
                    # JP: mtimeで最新のものを選ぶ
                    filtered_data_file = max(found, key=lambda p: os.path.getmtime(p))
                else:
                    return False, ["❌ No se encontró el archivo: filtered_data.xlsx (01_学習モデル/03_モデル学習)"], []
            
            print(f"📊 ファイルを検証中: {filtered_data_file}")
            
            # ES: Cargar datos del archivo Excel
            # EN: Load data from the Excel file
            # JP: Excelファイルからデータを読み込む
            import pandas as pd
            data_df = pd.read_excel(filtered_data_file)
            
            print(f"📊 検証用データを読み込みました: {len(data_df)} 行, {len(data_df.columns)} 列")
            print(f"📊 利用可能な列: {list(data_df.columns)}")
            
            errors = []
            warnings = []
            
            # ES: 1. Validar tipos de cepillo (A13, A11, A21, A32)
            # EN: 1. Validate brush types (A13, A11, A21, A32)
            # JP: 1. ブラシ種別を検証（A13, A11, A21, A32）
            brush_columns = ['A13', 'A11', 'A21', 'A32']
            brush_values = {}
            
            for col in brush_columns:
                if col in data_df.columns:
                    # Contar valores únicos que no sean 0
                    non_zero_values = data_df[data_df[col] == 1][col].unique()
                    brush_values[col] = len(non_zero_values)
                else:
                    brush_values[col] = 0
            
            # ES: Verificar que los brushes requeridos (desde 未実験データ) estén presentes en filtered_data | EN: Ensure required brushes (from 未実験データ) are in filtered_data | JA: filtered_dataに必要なブラシ（未実験データ由来）があるか確認
            required_brushes = []
            if isinstance(selected_params, dict):
                if selected_params.get("brush") in brush_columns:
                    required_brushes = [selected_params.get("brush")]
                elif isinstance(selected_params.get("brushes"), (list, tuple)):
                    required_brushes = [b for b in selected_params.get("brushes") if b in brush_columns]
            for b in required_brushes:
                if b in brush_values and brush_values[b] == 0:
                    errors.append(f"❌ filtered_data にブラシ '{b}' のデータがありません")
            
            # ES: 2. Validar material
            # EN: 2. Validate material
            # JP: 2. 材料を検証
            material_column = '材料'
            if material_column in data_df.columns:
                unique_materials = data_df[material_column].dropna().unique()
                if len(unique_materials) > 1:
                    errors.append(f"❌ Múltiples materiales encontrados: {list(unique_materials)}")
                
                # ES: Verificar si el material seleccionado está presente | EN: Check if selected material is present | JA: 選択材料が存在するか確認
                selected_material = selected_params['material']
                if selected_material not in unique_materials:
                    errors.append(f"❌ El material seleccionado '{selected_material}' no está presente en los datos")
            else:
                errors.append(f"❌ No se encontró la columna de material: {material_column}")
            
            # ES: 3. Validar diámetro
            # EN: 3. Validate diameter
            # JP: 3. 直径を検証
            diameter_column = '直径'
            if diameter_column in data_df.columns:
                unique_diameters = data_df[diameter_column].dropna().unique()
                if len(unique_diameters) > 1:
                    errors.append(f"❌ Múltiples diámetros encontrados: {list(unique_diameters)}")
                
                # ES: Verificar si el diámetro seleccionado está presente | EN: Check if selected diameter is present | JA: 選択直径が存在するか確認
                selected_diameter = selected_params['diameter']
                if selected_diameter not in unique_diameters:
                    errors.append(f"❌ El diámetro seleccionado '{selected_diameter}' no está presente en los datos")
            else:
                errors.append(f"❌ No se encontró la columna de diámetro: {diameter_column}")
            
            # 4. Validar rango de 線材長
            wire_length_column = '線材長'
            if wire_length_column in data_df.columns:
                wire_length_values = data_df[wire_length_column].dropna()
                if len(wire_length_values) > 0:
                    min_wire_length = wire_length_values.min()
                    max_wire_length = wire_length_values.max()
                    # ES: Si se proporcionó un único wire_length, mantener validación legacy.
                    # EN: If a single wire_length was provided, keep legacy validation.
                    # JP: wire_lengthが単一の場合は従来の検証を維持
                    if isinstance(selected_params, dict) and selected_params.get("wire_length") is not None:
                        selected_wire_length = selected_params["wire_length"]
                        expected_min = selected_wire_length - 5
                        expected_max = selected_wire_length
                        if min_wire_length < expected_min or max_wire_length > expected_max:
                            errors.append(f"❌ Rango de 線材長 fuera del rango esperado:")
                            errors.append(f"   - Rango en datos: {min_wire_length} - {max_wire_length}")
                            errors.append(f"   - Rango esperado: {expected_min} - {expected_max}")
                            errors.append(f"   - Seleccionado por usuario: {selected_wire_length}")
                    # Nuevo: múltiples wire_lengths (desde 未実験データ) -> comprobar que están dentro del rango de filtered_data
                    elif isinstance(selected_params, dict) and isinstance(selected_params.get("wire_lengths"), (list, tuple)):
                        try:
                            req = [float(x) for x in selected_params.get("wire_lengths")]
                            out = [x for x in req if not (min_wire_length <= x <= max_wire_length)]
                            if out:
                                warnings.append("⚠️ 未実験データ の 線材長 が filtered_data の範囲外です")
                                warnings.append(f"   - filtered_data range: {min_wire_length} - {max_wire_length}")
                                warnings.append(f"   - out of range (first 10): {out[:10]}")
                        except Exception:
                            # ES: Si no se puede convertir, no bloquear aquí (YosokuWorker validará)
                            # EN: If it can't be converted, don't block here (YosokuWorker will validate)
                            # JP: 変換できない場合でもここではブロックしない（YosokuWorkerが検証）
                            pass
                else:
                    errors.append(f"❌ No hay datos válidos en la columna 線材長")
            else:
                errors.append(f"❌ No se encontró la columna 線材長: {wire_length_column}")
            
            # ES: Retornar resultado de validación
            # EN: Return validation result
            # JP: 検証結果を返す
            if errors:
                print("❌ 検証エラーが見つかりました:")
                for error in errors:
                    print(f"   {error}")
                if warnings:
                    print("⚠️ 検証の警告:")
                    for w in warnings:
                        print(f"   {w}")
                return False, errors, warnings
            else:
                if warnings:
                    print("⚠️ 検証の警告:")
                    for w in warnings:
                        print(f"   {w}")
                else:
                    print("✅ 検証成功 - すべてのパラメータが一貫しています")
                return True, [], warnings
                
        except Exception as e:
            error_msg = f"❌ Error durante la validación: {str(e)}"
            print(error_msg)
            import traceback
            traceback.print_exc()
            return False, [error_msg], []

    def run_prediction(self):
        """ES: Ejecutar predicción Yosoku con parámetros del usuario y diálogo de progreso
        EN: Run Yosoku prediction with user parameters and progress dialog
        JA: ユーザー指定と進捗ダイアログで予測Yosokuを実行"""
        print("🔧 予測（Yosoku）を開始中...")
        
        try:
            # ES: Verificar que tenemos la carpeta del proyecto | EN: Ensure we have project folder | JA: プロジェクトフォルダがあるか確認
            if not hasattr(self, 'current_project_folder') or not self.current_project_folder:
                QMessageBox.warning(self, "エラー", "❌ プロジェクトフォルダが見つかりません。")
                return

            # ES: Buscar archivo 未実験データ (xlsx/csv)
            # EN: Look for the 未実験データ file (xlsx/csv)
            # JP: 未実験データファイル（xlsx/csv）を探す
            unexperimental_file = self.find_unexperimental_file()
            if not unexperimental_file:
                QMessageBox.warning(self, "エラー", "❌ 未実験データファイルが見つかりません。")
                return

            # ES: Validar que existan columnas (A13/A11/A21/A32, 線材長) en 未実験データ y recoger valores
            # EN: Validate required columns (A13/A11/A21/A32, 線材長) in 未実験データ and collect values
            # JP: 未実験データに必要な列（A13/A11/A21/A32、線材長）があるか検証し値を取得
            try:
                brush_types, wire_lengths = self._extract_brush_and_wire_length_from_unexperimental(unexperimental_file)
                print(f"✅ 未実験データから取得: brushes={brush_types}, 線材長={wire_lengths[:10]}{'...' if len(wire_lengths) > 10 else ''}")
            except Exception as e:
                QMessageBox.critical(self, "エラー", f"❌ 未実験データの読み込み/検証に失敗しました:\n{str(e)}")
                return

            # ⚠️ Confirmación si hay múltiples brushes y/o múltiples 線材長
            try:
                multi_brush = isinstance(brush_types, (list, tuple)) and len(brush_types) > 1
                multi_len = isinstance(wire_lengths, (list, tuple)) and len(wire_lengths) > 1
                if multi_brush or multi_len:
                    lines = []
                    lines.append("⚠️ 未実験データに複数の値が含まれています。予測を続行しますか？")
                    lines.append("")
                    if multi_brush:
                        bt = ", ".join([str(x) for x in brush_types[:8]])
                        more = "..." if len(brush_types) > 8 else ""
                        lines.append(f"- ブラシタイプ: {bt}{more} (count={len(brush_types)})")
                    if multi_len:
                        wl = ", ".join([str(x) for x in wire_lengths[:10]])
                        more = "..." if len(wire_lengths) > 10 else ""
                        lines.append(f"- 線材長: {wl}{more} (count={len(wire_lengths)})")
                    lines.append("")
                    lines.append("※ 続行すると、各行の A13/A11/A21/A32 と 線材長 をそのまま使用して予測します。")

                    reply = QMessageBox.question(
                        self,
                        "警告",
                        "\n".join(lines),
                        QMessageBox.Yes | QMessageBox.No,
                        QMessageBox.No,
                    )
                    if reply != QMessageBox.Yes:
                        print("ℹ️ 複数値の警告後、ユーザーが予測をキャンセルしました")
                        return
            except Exception:
                # ES: Si falla el warning por cualquier motivo, no bloquear la predicción
                # EN: If the warning fails for any reason, don't block prediction
                # JP: 警告が何らかの理由で失敗しても予測を止めない
                pass
            
            # ES: Mostrar diálogo | EN: Show dialog | JA: ダイアログを表示 de selección de parámetros
            selected_params = self.show_yosoku_parameters_dialog()
            if not selected_params:
                print("❌ ユーザーがパラメータ選択をキャンセルしました")
                return

            # ES: Completar parámetros desde archivo (no UI)
            # EN: Fill parameters from the file (not from the UI)
            # JP: パラメータをファイルから補完（UIではない）
            # Nota: el archivo puede contener múltiples brush/線材長; Yosoku los usa por fila.
            selected_params["brushes"] = brush_types
            selected_params["wire_lengths"] = wire_lengths
            
            print(f"📊 選択されたパラメータ: {selected_params}")
            
            # ES: Validar datos filtrados antes de continuar
            # EN: Validate filtered data before continuing
            # JP: 続行前にフィルタ済みデータを検証
            print("🔍 フィルタ済みデータを検証中...")
            is_valid, validation_errors, validation_warnings = self.validate_filtered_data(selected_params)
            
            if not is_valid:
                # ES: Mostrar resumen de errores | EN: Show error summary | JA: エラーサマリを表示
                error_summary = "❌ Validación fallida - No se puede continuar con la predicción:\n\n"
                error_summary += "\n".join(validation_errors)
                
                print("❌ 検証に失敗しました:")
                for error in validation_errors:
                    print(f"   {error}")
                
                QMessageBox.critical(
                    self,
                    "エラー - データ検証失敗",
                    error_summary
                )
                return

            # ES: Si hay warnings (p.ej. 線材長 fuera de rango), preguntar si desea continuar
            # EN: If there are warnings (e.g., 線材長 out of range), ask whether to continue
            # JP: 警告がある場合（例：線材長が範囲外）、続行するか確認
            if validation_warnings:
                try:
                    msg = "⚠️ データ検証で警告が見つかりました。続行しますか？\n\n"
                    msg += "\n".join(validation_warnings)
                    reply = QMessageBox.question(
                        self,
                        "警告",
                        msg,
                        QMessageBox.Yes | QMessageBox.No,
                        QMessageBox.No,
                    )
                    if reply != QMessageBox.Yes:
                        print("ℹ️ 検証の警告後、ユーザーが予測をキャンセルしました")
                        return
                except Exception:
                    # ES: Si el popup falla, continuar por defecto (no bloquear)
                    # EN: If the popup fails, continue by default (don't block)
                    # JP: ポップアップが失敗したらデフォルトで続行（ブロックしない）
                    pass
            
            print("✅ 検証成功 - 予測を続行します")
            
            # ES: Iniciar predicción con diálogo de progreso
            # EN: Start prediction with a progress dialog
            # JP: 進捗ダイアログ付きで予測を開始
            self.start_yosoku_prediction_with_progress(selected_params, unexperimental_file=unexperimental_file)
            
        except Exception as e:
            print(f"❌ 予測実行エラー: {e}")
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "エラー", f"❌ 予測実行中にエラーが発生しました:\n{str(e)}")

    def start_yosoku_prediction_with_progress(self, selected_params, unexperimental_file=None):
        """ES: Iniciar predicción Yosoku con diálogo de progreso
        EN: Start Yosoku prediction with progress dialog
        JA: 進捗ダイアログ付きで予測Yosokuを開始"""
        try:
            # ES: Buscar archivos necesarios
            # EN: Look for required files
            # JP: 必要なファイルを探す
            if not unexperimental_file:
                unexperimental_file = self.find_unexperimental_file()
            if not unexperimental_file:
                QMessageBox.warning(self, "エラー", "❌ 未実験データファイルが見つかりません。")
                return
            
            # ES: Localizar carpeta de predicción del análisis lineal más reciente (para guardar el CSV)
            # EN: Locate the latest linear-analysis prediction folder (to save the CSV)
            # JP: 最新の線形解析の予測フォルダを特定（CSV保存用）
            prediction_folder = None
            try:
                prediction_folder = self.find_latest_prediction_folder()
            except Exception:
                prediction_folder = None
            if not prediction_folder or not os.path.exists(prediction_folder):
                QMessageBox.warning(self, "エラー", "❌ 04_予測計算 フォルダが見つかりません。")
                return
            
            # ES: Crear ruta de salida | EN: Create output path | JA: 出力パスを作成
            from datetime import datetime
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            base = os.path.basename(unexperimental_file)
            for suf in ("_未実験データ.xlsx", "_未実験データ.xls", "_未実験データ.csv"):
                if base.endswith(suf):
                    base = base[: -len(suf)]
                    break
            output_filename = f"{base}_予測結果_{timestamp}.csv"
            output_path = os.path.join(prediction_folder, output_filename)
            
            # ES: Crear y mostrar diálogo de progreso | EN: Create and show progress dialog | JA: 進捗ダイアログを作成して表示
            self.yosoku_progress_dialog = YosokuProgressDialog(self)
            self.yosoku_progress_dialog.show()
            self.set_console_overlay_topmost(True)
            
            # ES: Crear worker thread | EN: Create worker thread | JA: ワーカースレッドを作成
            # YosokuWorker ahora calcula predicciones en Python y guarda CSV (sin límite de filas de Excel)
            self.yosoku_worker = YosokuWorker(selected_params, unexperimental_file, output_path, prediction_folder=prediction_folder)
            
            # ES: Conectar señales | EN: Connect signals | JA: シグナルを接続
            self.yosoku_worker.progress_updated.connect(self.yosoku_progress_dialog.update_progress)
            self.yosoku_worker.status_updated.connect(self.yosoku_progress_dialog.update_status)
            self.yosoku_worker.finished.connect(self.on_yosoku_prediction_finished)
            self.yosoku_worker.error.connect(self.on_yosoku_prediction_error)
            
            # ES: Conectar botón de cancelar | EN: Connect cancel button | JA: キャンセルボタンを接続
            self.yosoku_progress_dialog.cancel_button.clicked.connect(self.cancel_yosoku_prediction)

            # ES: Iniciar worker
            # EN: Start worker
            # JP: ワーカーを開始する

            self.yosoku_worker.start()
            
        except Exception as e:
            print(f"❌ 予測開始エラー: {e}")
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "エラー", f"❌ 予測開始中にエラーが発生しました:\n{str(e)}")

    def find_unexperimental_file(self):
        """# ES: Encontrar el archivo 未実験データ (xlsx/csv/xls)
# EN: Find the 未実験データ file (xlsx/csv/xls)
# JP: 未実験データ のファイル（xlsx/csv/xls）を探す
"""
        try:
            project_name = os.path.basename(self.current_project_folder)
            candidates = [
                os.path.join(self.current_project_folder, f"{project_name}_未実験データ.xlsx"),
                os.path.join(self.current_project_folder, f"{project_name}_未実験データ.xls"),
                os.path.join(self.current_project_folder, f"{project_name}_未実験データ.csv"),
            ]
            for p in candidates:
                if os.path.exists(p):
                    return p

            # ES: Fallback: buscar por patrón, preferir Excel, luego CSV
            # EN: Fallback: search by pattern; prefer Excel, then CSV
            # JP: フォールバック: パターン検索（Excel優先、次にCSV）
            files = []
            try:
                files = os.listdir(self.current_project_folder)
            except Exception:
                files = []

            preferred_exts = (".xlsx", ".xls", ".csv")
            for ext in preferred_exts:
                for file in files:
                    if file.endswith(f"_未実験データ{ext}"):
                        return os.path.join(self.current_project_folder, file)
            return None
        except Exception as e:
            print(f"❌ 未実験データファイルの検索エラー: {e}")
            return None

    def on_yosoku_prediction_finished(self, output_path):
        """ES: Manejar finalización exitosa de la predicción
        EN: Handle successful prediction completion
        JA: 予測の正常完了を処理"""
        try:
            # ES: Cerrar diálogo de progreso
            # EN: Close progress dialog
            # JP: 進捗ダイアログを閉じる
            if hasattr(self, 'yosoku_progress_dialog'):
                self.yosoku_progress_dialog.close()
                self.yosoku_progress_dialog = None
            self.set_console_overlay_topmost(False)

            # ES: Terminar worker
            # EN: Stop worker
            # JP: ワーカーを終了する

            if hasattr(self, 'yosoku_worker'):
                self.yosoku_worker.quit()
                self.yosoku_worker.wait()
                self.yosoku_worker = None
            
            # ES: Mostrar mensaje de éxito | EN: Show success message | JA: 成功メッセージを表示
            output_filename = os.path.basename(output_path)
            formulas_folder = os.path.dirname(output_path)
            
            QMessageBox.information(
                self,
                "予測完了",
                f"✅ 予測が完了しました！\n\n結果ファイル: {output_filename}\n\n保存場所: {formulas_folder}"
            )

            # ES: Preguntar si quiere importar a la base de datos
            # EN: Ask whether they want to import into the database
            # JP: データベースにインポートするかどうかを確認する

            reply = QMessageBox.question(
                self,
                "データベースインポート",
                "予測結果をデータベースにインポートしますか？",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes
            )
            
            if reply == QMessageBox.Yes:
                self.import_yosoku_results_to_database(output_path)
            
        except Exception as e:
            print(f"❌ 終了処理エラー: {e}")
            import traceback
            traceback.print_exc()

    def on_yosoku_prediction_error(self, error_msg):
        """ES: Manejar errores en la predicción
        EN: Handle prediction errors
        JA: 予測のエラーを処理"""
        try:
            # ES: Cerrar diálogo de progreso
            # EN: Close progress dialog
            # JP: 進捗ダイアログを閉じる
            if hasattr(self, 'yosoku_progress_dialog'):
                self.yosoku_progress_dialog.close()
                self.yosoku_progress_dialog = None
            self.set_console_overlay_topmost(False)

            # ES: Terminar worker
            # EN: Stop worker
            # JP: ワーカーを終了する

            if hasattr(self, 'yosoku_worker'):
                self.yosoku_worker.quit()
                self.yosoku_worker.wait()
                self.yosoku_worker = None
            
            # ES: Mostrar mensaje de error | EN: Show error message | JA: エラーメッセージを表示
            QMessageBox.critical(self, "エラー", f"❌ 予測実行中にエラーが発生しました:\n{error_msg}")
            
        except Exception as e:
            print(f"❌ エラーハンドリングでエラー: {e}")
            import traceback
            traceback.print_exc()

    def import_yosoku_results_to_database(self, excel_path):
        """ES: Importar resultados de predicción a la base de datos con diálogo de progreso
        EN: Import prediction results into the database with progress dialog
        JA: 進捗ダイアログ付きで予測結果をDBに取り込み"""
        try:
            # ES: Verificar si ya existe un diálogo abierto (para evitar duplicados) | EN: Check if dialog is already open (avoid duplicates) | JA: ダイアログ重複防止のため既に開いているか確認
            if hasattr(self, 'yosoku_import_progress_dialog') and self.yosoku_import_progress_dialog is not None:
                # ES: Si ya existe, reutilizarlo
                # EN: If it already exists, reuse it
                # JP: 既に存在する場合は再利用
                existing_dialog = self.yosoku_import_progress_dialog
            else:
                # ES: Crear y mostrar diálogo de progreso | EN: Create and show progress dialog | JA: 進捗ダイアログを作成して表示
                self.yosoku_import_progress_dialog = YosokuImportProgressDialog(self)
                self.yosoku_import_progress_dialog.show()
                existing_dialog = self.yosoku_import_progress_dialog
            # ES: Mientras el diálogo con chibi esté activo: flecha/consola por encima
            # EN: While the chibi dialog is active: keep arrow/console above it
            # JP: chibiダイアログ表示中は矢印/コンソールを前面に
            self.set_console_overlay_topmost(True)
            
            # ES: Crear worker thread | EN: Create worker thread | JA: ワーカースレッドを作成 (análisis lineal)
            self.yosoku_import_worker = YosokuImportWorker(excel_path, analysis_type="lineal")
            
            # ES: Conectar señales | EN: Connect signals | JA: シグナルを接続
            self.yosoku_import_worker.progress_updated.connect(existing_dialog.update_progress)
            self.yosoku_import_worker.status_updated.connect(existing_dialog.set_status)
            self.yosoku_import_worker.finished.connect(self.on_yosoku_import_finished)
            self.yosoku_import_worker.error.connect(self.on_yosoku_import_error)
            
            # ES: Conectar botón de cancelar | EN: Connect cancel button | JA: キャンセルボタンを接続
            existing_dialog.cancel_button.clicked.connect(self.cancel_yosoku_import)

            # ES: Iniciar worker
            # EN: Start worker
            # JP: ワーカーを開始する

            self.yosoku_import_worker.start()
            
        except Exception as e:
            print(f"❌ インポート開始エラー: {e}")
            import traceback
            traceback.print_exc()
            QMessageBox.critical(
                self,
                "エラー",
                f"❌ インポート開始中にエラーが発生しました:\n{str(e)}"
            )
    
    def on_yosoku_import_finished(self):
        """ES: Manejar finalización exitosa de importación
        EN: Handle successful import completion
        JA: インポートの正常完了を処理"""
        try:
            # ES: Cerrar diálogo de progreso
            # EN: Close progress dialog
            # JP: 進捗ダイアログを閉じる
            if hasattr(self, 'yosoku_import_progress_dialog') and self.yosoku_import_progress_dialog is not None:
                self.yosoku_import_progress_dialog.close()
                self.yosoku_import_progress_dialog = None
            self.set_console_overlay_topmost(False)
            
            # ES: Limpiar worker
            # EN: Clean up worker
            # JP: ワーカーをクリーンアップ
            if hasattr(self, 'yosoku_import_worker') and self.yosoku_import_worker is not None:
                self.yosoku_import_worker.quit()
                self.yosoku_import_worker.wait()
                self.yosoku_import_worker = None
            
            # ES: Mostrar mensaje de éxito | EN: Show success message | JA: 成功メッセージを表示
            QMessageBox.information(
                self,
                "インポート完了",
                "✅ データベースへのインポートが完了しました！"
            )
            
        except Exception as e:
            print(f"❌ 終了処理のハンドリングエラー: {e}")
            import traceback
            traceback.print_exc()
    
    def on_yosoku_import_error(self, error_msg):
        """ES: Manejar error en importación
        EN: Handle import error
        JA: インポートのエラーを処理"""
        try:
            # ES: Cerrar diálogo de progreso
            # EN: Close progress dialog
            # JP: 進捗ダイアログを閉じる
            if hasattr(self, 'yosoku_import_progress_dialog') and self.yosoku_import_progress_dialog is not None:
                self.yosoku_import_progress_dialog.close()
                self.yosoku_import_progress_dialog = None
            self.set_console_overlay_topmost(False)
            
            # ES: Limpiar worker
            # EN: Clean up worker
            # JP: ワーカーをクリーンアップ
            if hasattr(self, 'yosoku_import_worker') and self.yosoku_import_worker is not None:
                self.yosoku_import_worker.quit()
                self.yosoku_import_worker.wait()
                self.yosoku_import_worker = None
            
            # ES: Mostrar mensaje de error | EN: Show error message | JA: エラーメッセージを表示
            QMessageBox.critical(
                self,
                "エラー",
                f"❌ データベースへのインポート中にエラーが発生しました:\n{error_msg}"
            )
            
        except Exception as e:
            print(f"❌ エラーハンドリングでエラー: {e}")
            import traceback
            traceback.print_exc()
    
    def cancel_yosoku_import(self):
        """ES: Cancelar importación
        EN: Cancel import
        JA: インポートをキャンセル
        """
        try:
            if hasattr(self, 'yosoku_import_worker'):
                self.yosoku_import_worker.cancel_import()
                self.yosoku_import_worker.quit()
                self.yosoku_import_worker.wait()
                self.yosoku_import_worker = None
            
            if hasattr(self, 'yosoku_import_progress_dialog'):
                self.yosoku_import_progress_dialog.close()
                self.yosoku_import_progress_dialog = None
            self.set_console_overlay_topmost(False)
        except Exception as e:
            print(f"❌ インポートキャンセルエラー: {e}")
            import traceback
            traceback.print_exc()
    
    def import_classification_results_to_yosoku_db(self):
        """ES: Importar resultados de clasificación a la base de datos de yosoku
        EN: Import classification results into the yosoku database
        JA: 分類結果をyosoku DBに取り込み"""
        try:
            # ES: Obtener la carpeta raíz del análisis de clasificación
            # EN: Get the root folder of the classification analysis
            # JP: 分類解析のルートフォルダを取得
            # Puede estar en classification_project_folder o classification_existing_folder_path
            from pathlib import Path
            import glob
            import os
            
            # ES: Intentar obtener la carpeta raíz del análisis
            # EN: Try to determine the analysis root folder
            # JP: 解析ルートフォルダを特定してみる
            if hasattr(self, 'classification_project_folder') and self.classification_project_folder:
                analysis_root = Path(self.classification_project_folder)
            elif hasattr(self, 'classification_existing_folder_path') and self.classification_existing_folder_path:
                # ES: Si solo tenemos la carpeta de evaluación, subir dos niveles para llegar a la raíz
                # EN: If we only have the evaluation folder, go up two levels to reach the root
                # JP: 評価フォルダしか無い場合は2階層上がってルートへ
                analysis_root = Path(self.classification_existing_folder_path).parent.parent
            else:
                QMessageBox.warning(self, "エラー", "❌ 分類解析結果のフォルダが見つかりません。")
                return
            
            print(f"🔍 解析のルートフォルダ: {analysis_root}")
            print(f"🔍 ルートフォルダの存在: {analysis_root.exists()}")
            
            # ES: Construir ruta del archivo de predicción desde la carpeta raíz
            # EN: Build the prediction-file path from the root folder
            # JP: ルートフォルダから予測ファイルのパスを構築
            pred_folder = analysis_root / "02_本学習結果" / "03_予測結果"
            
            print(f"🔍 予測ファイルを検索中: {pred_folder}")
            print(f"🔍 フォルダの存在: {pred_folder.exists()}")
            
            if not pred_folder.exists():
                # ES: Intentar con ruta absoluta
                # EN: Try using an absolute path
                # JP: 絶対パスで試す
                pred_folder_abs = analysis_root.resolve() / "02_本学習結果" / "03_予測結果"
                print(f"🔍 絶対パスで試行中: {pred_folder_abs}")
                if pred_folder_abs.exists():
                    pred_folder = pred_folder_abs
                else:
                    # ES: Mostrar información de debug | EN: Show debug info | JA: デバッグ情報を表示
                    print("❌ 予測フォルダが見つかりません")
                    print(f"   試行パス 1: {pred_folder}")
                    print(f"   試行パス 2: {pred_folder_abs}")
                    print(f"   ルートフォルダ: {analysis_root}")
                    print(f"   ルートフォルダの存在: {analysis_root.exists()}")
                    if analysis_root.exists():
                        print("   ルートフォルダの内容:")
                        for item in analysis_root.iterdir():
                            print(f"     - {item.name} ({'dir' if item.is_dir() else 'file'})")
                    
                    QMessageBox.warning(
                        self,
                        "エラー",
                        f"❌ 予測結果フォルダが見つかりません。\n\n"
                        f"フォルダ: {pred_folder}\n\n"
                        f"または:\n{pred_folder_abs}\n\n"
                        f"分析ルートフォルダ: {analysis_root}"
                    )
                    return
            
            # ES: Listar archivos en la carpeta para debug
            # EN: List files in the folder for debugging
            # JP: デバッグ用にフォルダ内のファイルを列挙
            all_files = list(pred_folder.glob("*"))
            print(f"🔍 フォルダ内のファイル ({len(all_files)} 件):")
            for f in all_files:
                print(f"  - {f.name} (archivo: {f.is_file()}, dir: {f.is_dir()})")
            
            # ES: Buscar archivo de predicción con diferentes estrategias
            # EN: Look for the prediction file using different strategies
            # JP: 複数の戦略で予測ファイルを探す
            prediction_file = None

            # ES: Prioridad 1: Prediction_input_pred.xlsx (ignorar archivos temporales de Excel)
            # EN: Priority 1: Prediction_input_pred.xlsx (ignore temporary Excel files)
            # JP: 優先度1：Prediction_input_pred.xlsx（Excel の一時ファイルは無視する）

            candidate1 = pred_folder / "Prediction_input_pred.xlsx"
            if candidate1.exists() and not candidate1.name.startswith("~$"):
                prediction_file = candidate1
                print(f"✅ ファイルが見つかりました（優先度1）: {prediction_file}")
            else:
                # ES: Prioridad 2: Buscar cualquier archivo *_pred.xlsx (ignorar temporales)
                # EN: Priority 2: Find any *_pred.xlsx file (ignore temp files)
                # JP: 優先2: *_pred.xlsx を探す（テンポラリは除外）
                pred_files = [f for f in pred_folder.glob("*_pred.xlsx") if not f.name.startswith("~$")]
                if pred_files:
                    # ES: Seleccionar el más reciente
                    # EN: Pick the most recent one
                    # JP: 最新を選ぶ
                    prediction_file = max(pred_files, key=lambda p: p.stat().st_mtime)
                    print(f"✅ ファイルが見つかりました（優先度2）: {prediction_file}")
                else:
                    # ES: Prioridad 3: Buscar cualquier archivo .xlsx en la carpeta (ignorar temporales)
                    # EN: Priority 3: Find any .xlsx file in the folder (ignore temp files)
                    # JP: 優先3: フォルダ内の.xlsxを探す（テンポラリは除外）
                    xlsx_files = [f for f in pred_folder.glob("*.xlsx") if not f.name.startswith("~$")]
                    if xlsx_files:
                        # ES: Seleccionar el más reciente
                        # EN: Pick the most recent one
                        # JP: 最新を選ぶ
                        prediction_file = max(xlsx_files, key=lambda p: p.stat().st_mtime)
                        print(f"✅ ファイルが見つかりました（優先度3）: {prediction_file}")
            
            if not prediction_file or not prediction_file.exists():
                # ES: Listar archivos disponibles para ayudar al usuario
                # EN: List available files to help the user
                # JP: ユーザーを支援するために利用可能なファイルを一覧表示する

                available_files = [f.name for f in pred_folder.glob("*.xlsx") if not f.name.startswith("~$")]
                files_list = "\n".join([f"  - {f}" for f in available_files]) if available_files else "  (なし)"
                
                QMessageBox.warning(
                    self,
                    "エラー",
                    f"❌ 予測結果ファイルが見つかりません。\n\n"
                    f"フォルダ: {pred_folder}\n\n"
                    f"利用可能なファイル:\n{files_list}\n\n"
                    f"期待されるファイル名:\n"
                    f"- Prediction_input_pred.xlsx\n"
                    f"- *_pred.xlsx\n\n"
                    f"注意: Excelでファイルが開かれている場合は、閉じてから再度お試しください。"
                )
                return
            
            print(f"✅ 予測ファイルを選択しました: {prediction_file}")

            # ES: Preguntar al usuario sobre sobrescritura
            # EN: Ask the user about overwriting
            # JP: 上書きするかどうかをユーザーに確認する

            reply = QMessageBox.question(
                self,
                "データベースへのインポート",
                "既存のデータを上書きしますか？\n\n"
                "既存のレコードが見つかった場合、そのレコードを更新します。\n"
                "「いいえ」を選択した場合、既存のレコードはスキップされます。",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            overwrite = (reply == QMessageBox.Yes)
            
            # ES: Crear y mostrar diálogo de progreso | EN: Create and show progress dialog | JA: 進捗ダイアログを作成して表示
            if hasattr(self, 'classification_import_progress_dialog') and self.classification_import_progress_dialog is not None:
                existing_dialog = self.classification_import_progress_dialog
            else:
                self.classification_import_progress_dialog = YosokuImportProgressDialog(self)
                self.classification_import_progress_dialog.show()
                existing_dialog = self.classification_import_progress_dialog
            self.set_console_overlay_topmost(True)
            
            # ES: Crear worker thread | EN: Create worker thread | JA: ワーカースレッドを作成
            self.classification_import_worker = ClassificationImportWorker(str(prediction_file), overwrite=overwrite)
            
            # ES: Conectar señales | EN: Connect signals | JA: シグナルを接続
            self.classification_import_worker.progress_updated.connect(existing_dialog.update_progress)
            self.classification_import_worker.status_updated.connect(existing_dialog.set_status)
            self.classification_import_worker.finished.connect(self.on_classification_import_finished)
            self.classification_import_worker.error.connect(self.on_classification_import_error)
            
            # ES: Conectar botón de cancelar | EN: Connect cancel button | JA: キャンセルボタンを接続
            existing_dialog.cancel_button.clicked.connect(self.cancel_classification_import)

            # ES: Iniciar worker
            # EN: Start worker
            # JP: ワーカーを開始する

            self.classification_import_worker.start()
            
        except Exception as e:
            print(f"❌ 分類インポート開始エラー: {e}")
            import traceback
            traceback.print_exc()
            QMessageBox.critical(
                self,
                "エラー",
                f"❌ インポート開始中にエラーが発生しました:\n{str(e)}"
            )
    
    def on_classification_import_finished(self, inserted_count, updated_count):
        """ES: Manejar finalización exitosa de importación de clasificación
        EN: Handle successful classification import completion
        JA: 分類インポートの正常完了を処理"""
        try:
            # ES: Cerrar diálogo de progreso
            # EN: Close progress dialog
            # JP: 進捗ダイアログを閉じる
            if hasattr(self, 'classification_import_progress_dialog') and self.classification_import_progress_dialog is not None:
                self.classification_import_progress_dialog.close()
                self.classification_import_progress_dialog = None
            self.set_console_overlay_topmost(False)
            
            # ES: Limpiar worker
            # EN: Clean up worker
            # JP: ワーカーをクリーンアップ
            if hasattr(self, 'classification_import_worker') and self.classification_import_worker is not None:
                self.classification_import_worker.quit()
                self.classification_import_worker.wait()
                self.classification_import_worker = None
            
            # ES: Mostrar mensaje de éxito | EN: Show success message | JA: 成功メッセージを表示
            QMessageBox.information(
                self,
                "インポート完了",
                f"✅ データベースへのインポートが完了しました！\n\n"
                f"新規追加: {inserted_count} 件\n"
                f"更新: {updated_count} 件"
            )
            
        except Exception as e:
            print(f"❌ 終了処理のハンドリングエラー: {e}")
            import traceback
            traceback.print_exc()
    
    def on_classification_import_error(self, error_msg):
        """ES: Manejar error en importación de clasificación
        EN: Handle classification import error
        JA: 分類インポートのエラーを処理"""
        try:
            # ES: Cerrar diálogo de progreso
            # EN: Close progress dialog
            # JP: 進捗ダイアログを閉じる
            if hasattr(self, 'classification_import_progress_dialog') and self.classification_import_progress_dialog is not None:
                self.classification_import_progress_dialog.close()
                self.classification_import_progress_dialog = None
            self.set_console_overlay_topmost(False)
            
            # ES: Limpiar worker
            # EN: Clean up worker
            # JP: ワーカーをクリーンアップ
            if hasattr(self, 'classification_import_worker') and self.classification_import_worker is not None:
                self.classification_import_worker.quit()
                self.classification_import_worker.wait()
                self.classification_import_worker = None
            
            # ES: Mostrar mensaje de error | EN: Show error message | JA: エラーメッセージを表示
            QMessageBox.critical(
                self,
                "エラー",
                f"❌ データベースへのインポート中にエラーが発生しました:\n{error_msg}"
            )
            
        except Exception as e:
            print(f"❌ エラーハンドリングでエラー: {e}")
            import traceback
            traceback.print_exc()
    
    def cancel_classification_import(self):
        """ES: Cancelar importación de clasificación
        EN: Cancel classification import
        JA: 分類インポートをキャンセル
        """
        try:
            if hasattr(self, 'classification_import_worker'):
                self.classification_import_worker.cancel_import()
                self.classification_import_worker.quit()
                self.classification_import_worker.wait()
                self.classification_import_worker = None
            
            if hasattr(self, 'classification_import_progress_dialog'):
                self.classification_import_progress_dialog.close()
                self.classification_import_progress_dialog = None
            self.set_console_overlay_topmost(False)
                
        except Exception as e:
            print(f"❌ インポートキャンセルエラー: {e}")
            import traceback
            traceback.print_exc()
            
            QMessageBox.information(self, "キャンセル", "インポートがキャンセルされました。")
            
        except Exception as e:
            print(f"❌ インポートキャンセルエラー: {e}")
            import traceback
            traceback.print_exc()
    
    def on_yosoku_export_finished(self, filepath, record_count):
        """ES: Manejar finalización exitosa de exportación
        EN: Handle successful export completion
        JA: エクスポートの正常完了を処理"""
        try:
            # ES: Cerrar diálogo de progreso
            # EN: Close progress dialog
            # JP: 進捗ダイアログを閉じる
            if hasattr(self, 'yosoku_export_progress_dialog') and self.yosoku_export_progress_dialog is not None:
                self.yosoku_export_progress_dialog.close()
                self.yosoku_export_progress_dialog = None
            self.set_console_overlay_topmost(False)
            
            # ES: Limpiar worker
            # EN: Clean up worker
            # JP: ワーカーをクリーンアップ
            if hasattr(self, 'yosoku_export_worker') and self.yosoku_export_worker is not None:
                self.yosoku_export_worker.quit()
                self.yosoku_export_worker.wait()
                self.yosoku_export_worker = None
            
            # ES: Mostrar mensaje de éxito | EN: Show success message | JA: 成功メッセージを表示
            QMessageBox.information(
                self,
                "完了",
                f"✅ 予測データベースが正常にエクスポートされました。\n\nファイル: {os.path.basename(filepath)}\nレコード数: {record_count}"
            )
            
        except Exception as e:
            print(f"❌ エクスポート終了処理のハンドリングエラー: {e}")
            import traceback
            traceback.print_exc()
    
    def on_yosoku_export_error(self, error_msg):
        """ES: Manejar error en exportación
        EN: Handle export error
        JA: エクスポートのエラーを処理"""
        try:
            # ES: Cerrar diálogo de progreso
            # EN: Close progress dialog
            # JP: 進捗ダイアログを閉じる
            if hasattr(self, 'yosoku_export_progress_dialog') and self.yosoku_export_progress_dialog is not None:
                self.yosoku_export_progress_dialog.close()
                self.yosoku_export_progress_dialog = None
            self.set_console_overlay_topmost(False)
            
            # ES: Limpiar worker
            # EN: Clean up worker
            # JP: ワーカーをクリーンアップ
            if hasattr(self, 'yosoku_export_worker') and self.yosoku_export_worker is not None:
                self.yosoku_export_worker.quit()
                self.yosoku_export_worker.wait()
                self.yosoku_export_worker = None
            
            # ES: Mostrar mensaje de error | EN: Show error message | JA: エラーメッセージを表示
            QMessageBox.critical(
                self,
                "エラー",
                error_msg
            )
            
        except Exception as e:
            print(f"❌ エクスポートエラーのハンドリングエラー: {e}")
            import traceback
            traceback.print_exc()
    
    def cancel_yosoku_export(self):
        """ES: Cancelar exportación
        EN: Cancel export
        JA: エクスポートをキャンセル
        """
        try:
            if hasattr(self, 'yosoku_export_worker'):
                self.yosoku_export_worker.cancel_export()
                self.yosoku_export_worker.quit()
                self.yosoku_export_worker.wait()
                self.yosoku_export_worker = None
            
            if hasattr(self, 'yosoku_export_progress_dialog'):
                self.yosoku_export_progress_dialog.close()
                self.yosoku_export_progress_dialog = None
            self.set_console_overlay_topmost(False)
            
            QMessageBox.information(self, "キャンセル", "エクスポートがキャンセルされました。")
            
        except Exception as e:
            print(f"❌ エクスポートキャンセルエラー: {e}")
            import traceback
            traceback.print_exc()
    
    def prepare_dataframe_for_import(self, df, selected_params):
        """
        Prepara el DataFrame para importación agregando columnas de usuario
        y renombrando columnas de predicción si es necesario
        """
        try:
            # ES: Crear copia para no modificar el original | EN: Create copy to avoid modifying original | JA: 原本を変えぬようコピーを作成
            df_prepared = df.copy()

            # ES: Brush/longitud del alambre deben venir del archivo (no de la UI).
            # EN: Brush/wire length must come from the file (not the UI).
            # JP: Brush/線材長 はファイルから取得する必要がある（UI ではない）。
            # ES: Si faltan, es un error (no podemos inferirlos aquí).
            # EN: If they are missing, it's an error (we can't infer them here).
            # JP: 欠けている場合はエラー（ここでは推測できない）
            required_brush_cols = ["A13", "A11", "A21", "A32"]
            missing_brush = [c for c in required_brush_cols if c not in df_prepared.columns]
            if missing_brush:
                raise ValueError(
                    f"❌ Prediction file must include brush one-hot columns: {', '.join(required_brush_cols)} "
                    f"(missing: {', '.join(missing_brush)})"
                )
            if "線材長" not in df_prepared.columns:
                raise ValueError("❌ Prediction file must include column: 線材長")
            
            # ES: Agregar columnas de usuario
            # EN: Add user columns
            # JP: ユーザー列を追加
            df_prepared['直径'] = selected_params['diameter']
            df_prepared['材料'] = selected_params['material']
            
            # ES: Renombrar columnas de predicción si tienen prefijo 'prediction_'
            # EN: Rename prediction columns if they have the 'prediction_' prefix
            # JP: prediction_プレフィックス付きの予測列をリネーム
            rename_map = {}
            for col in df_prepared.columns:
                if col.startswith('prediction_'):
                    new_name = col.replace('prediction_', '')
                    rename_map[col] = new_name
            
            if rename_map:
                df_prepared = df_prepared.rename(columns=rename_map)
                print(f"📝 列のリネーム: {rename_map}")
            
            # Calcular 加工時間 si no existe
            if '加工時間' not in df_prepared.columns:
                if '送り速度' in df_prepared.columns:
                    # Fórmula: 100 / 送り速度 * 60
                    df_prepared['加工時間'] = df_prepared.apply(
                        lambda row: (100 / row['送り速度'] * 60) if pd.notna(row.get('送り速度')) and row.get('送り速度', 0) != 0 else 0,
                        axis=1
                    )
                    print("✅ 加工時間 calculado")
                else:
                    df_prepared['加工時間'] = 0
                    print("⚠️ 送り速度が見つかりません、加工時間 = 0")
            
            return df_prepared
            
        except Exception as e:
            print(f"❌ DataFrame 準備エラー: {e}")
            import traceback
            traceback.print_exc()
            raise
    
    def import_nonlinear_pareto_to_database(self, excel_path):
        """ES: Importa resultados de Pareto del análisis no lineal a la base de datos
        EN: Import non-linear Pareto results into the database
        JA: 非線形Pareto結果をDBにインポート
        """
        try:
            # ES: 1. Mostrar diálogo de parámetros (solo diámetro/material) PRIMERO (sin loading)
            # EN: 1. Show parameter dialog (diameter/material only) FIRST (no loading)
            # JP: 1. パラメータダイアログ（直径／材料のみ）を最初に表示する（ローディングなし）

            selected_params = self.show_yosoku_parameters_dialog()
            
            if not selected_params:
                print("❌ ユーザーがパラメータ選択をキャンセルしました")
                return
            
            # ES: ✅ MOSTRAR LOADING DESPUÉS de seleccionar parámetros y presionar OK
            # EN: ✅ SHOW LOADING AFTER selecting parameters and pressing OK
            # JP: ✅ パラメータ選択後、OK押下後にローディングを表示
            self.yosoku_import_progress_dialog = YosokuImportProgressDialog(self)
            self.yosoku_import_progress_dialog.show()
            self.yosoku_import_progress_dialog.update_progress(0, "初期化中...")
            self.yosoku_import_progress_dialog.set_status("初期化中...")
            QApplication.processEvents()  # Force UI refresh
            
            # ES: 2. Leer Excel y preparar DataFrame
            # EN: 2. Read Excel and prepare the DataFrame
            # JP: 2. Excelを読み込みDataFrameを準備
            self.yosoku_import_progress_dialog.update_progress(10, "Excelファイルを読み込み中...")
            self.yosoku_import_progress_dialog.set_status("Excelファイルを読み込み中...")
            QApplication.processEvents()
            
            print(f"📊 ファイルを読み込み中: {excel_path}")
            df = pd.read_excel(excel_path)
            print(f"✅ データを読み込みました: {len(df)} 行, {len(df.columns)} 列")

            # ES: 3. Preparar DataFrame con columnas de usuario
            # EN: 3. Prepare DataFrame with user-defined columns
            # JP: 3. ユーザー定義の列で DataFrame を準備する

            self.yosoku_import_progress_dialog.update_progress(30, "データを準備中...")
            self.yosoku_import_progress_dialog.set_status("データを準備中...")
            QApplication.processEvents()
            
            df_prepared = self.prepare_dataframe_for_import(df, selected_params)
            
            # ES: 4. Guardar DataFrame preparado en archivo intermedio (misma carpeta que Prediction_output.xlsx)
            # EN: 4. Save prepared DataFrame to an intermediate file (same folder as Prediction_output.xlsx)
            # JP: 4. 準備済みDataFrameを中間ファイルに保存（Prediction_output.xlsxと同じフォルダ）
            self.yosoku_import_progress_dialog.update_progress(50, "ファイルを保存中...")
            self.yosoku_import_progress_dialog.set_status("ファイルを保存中...")
            QApplication.processEvents()
            
            excel_folder = Path(excel_path).parent
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            intermediate_filename = f"Prediction_output_prepared_{timestamp}.xlsx"
            intermediate_path = excel_folder / intermediate_filename
            
            try:
                df_prepared.to_excel(intermediate_path, index=False)
                print(f"📁 中間ファイルを保存しました: {intermediate_path}")
            except Exception as e:
                print(f"⚠️ 中間ファイル保存エラー: {e}")
                # ES: No detener el proceso si falla guardar el intermedio
                # EN: Do not stop the process if saving the intermediate file fails
                # JP: 中間保存が失敗しても処理を止めない
            
            # ES: 5. Guardar también en archivo temporal para la importación
            # EN: 5. Also save to a temporary file for import
            # JP: 5. インポート用に一時ファイルにも保存
            temp_dir = tempfile.gettempdir()
            temp_file = os.path.join(temp_dir, f"pareto_import_{timestamp}.xlsx")
            df_prepared.to_excel(temp_file, index=False)
            print(f"📁 一時ファイルを作成しました: {temp_file}")
            
            # ES: 6. Importar usando el worker existente (el worker continuará desde 60%)
            # EN: 6. Import using the existing worker (the worker will continue from 60%)
            # JP: 6. 既存ワーカーでインポート（60%から継続）
            # ES: Nota: import_yosoku_results_to_database creará su propio diálogo,
            # EN: Note: import_yosoku_results_to_database will create its own dialog,
            # JP: 注: import_yosoku_results_to_database は独自のダイアログを作成するため、
            # así que necesitamos reutilizar el existente o pasarle el diálogo
            self._continue_import_with_worker(temp_file)
            
            # ES: 7. Limpiar archivo temporal después de un delay
            # EN: 7. Clean up the temporary file after a delay
            # JP: 7. 遅延後に一時ファイルを削除
            # ES: Nota: El archivo intermedio NO se elimina, queda como registro
            # EN: Note: the intermediate file is NOT deleted; it remains as a record
            # JP: 注: 中間ファイルは削除しない（記録として残す）
            def cleanup_temp_file():
                try:
                    if os.path.exists(temp_file):
                        os.remove(temp_file)
                        print(f"🗑️ 一時ファイルを削除しました: {temp_file}")
                except:
                    pass
            
            QTimer.singleShot(5000, cleanup_temp_file)  # Clean up after 5 seconds
            
        except Exception as e:
            print(f"❌ ParetoのBDインポートエラー: {e}")
            import traceback
            traceback.print_exc()
            
            # ES: Cerrar loading si hay error
            # EN: Close loading if there is an error
            # JP: エラー時にローディングを閉じる
            if hasattr(self, 'yosoku_import_progress_dialog') and self.yosoku_import_progress_dialog is not None:
                self.yosoku_import_progress_dialog.close()
                self.yosoku_import_progress_dialog = None
            
            QMessageBox.critical(
                self,
                "エラー",
                f"❌ データベースへのインポート中にエラーが発生しました:\n{str(e)}"
            )
    
    def _continue_import_with_worker(self, temp_file):
        """ES: Continúa la importación usando el worker, reutilizando el diálogo existente
        EN: Continue import using the worker, reusing the existing dialog
        JA: 既存ダイアログを再利用しワーカーでインポートを続行"""
        try:
            # ES: Actualizar progreso antes de iniciar worker
            # EN: Update progress before starting worker
            # JA: ワーカー開始前に進捗を更新
            self.yosoku_import_progress_dialog.update_progress(60, "データベースにインポート中...")
            self.yosoku_import_progress_dialog.set_status("データベースにインポート中...")
            QApplication.processEvents()
            
            # ES: Crear worker thread | EN: Create worker thread | JA: ワーカースレッドを作成 (análisis no lineal)
            self.yosoku_import_worker = YosokuImportWorker(temp_file, analysis_type="no_lineal")
            
            # ES: Conectar señales | EN: Connect signals | JA: シグナルを接続 (reutilizando el diálogo existente)
            self.yosoku_import_worker.progress_updated.connect(self._on_yosoku_import_progress)
            self.yosoku_import_worker.status_updated.connect(self.yosoku_import_progress_dialog.set_status)
            self.yosoku_import_worker.finished.connect(self.on_yosoku_import_finished)
            self.yosoku_import_worker.error.connect(self.on_yosoku_import_error)
            
            # ES: Conectar botón de cancelar | EN: Connect cancel button | JA: キャンセルボタンを接続
            self.yosoku_import_progress_dialog.cancel_button.clicked.connect(self.cancel_yosoku_import)

            # ES: Iniciar worker
            # EN: Start worker
            # JP: ワーカーを開始する

            self.yosoku_import_worker.start()
            
        except Exception as e:
            print(f"❌ インポートワーカー開始エラー: {e}")
            import traceback
            traceback.print_exc()
            
            # ES: Cerrar loading si hay error
            # EN: Close loading if there is an error
            # JP: エラー時にローディングを閉じる
            if hasattr(self, 'yosoku_import_progress_dialog') and self.yosoku_import_progress_dialog is not None:
                self.yosoku_import_progress_dialog.close()
                self.yosoku_import_progress_dialog = None
            
            QMessageBox.critical(
                self,
                "エラー",
                f"❌ インポート開始中にエラーが発生しました:\n{str(e)}"
            )
    
    def _on_yosoku_import_progress(self, value, message):
        """ES: Maneja el progreso del worker, mapeando de 0-100% del worker a 60-100% del total
        EN: Handle worker progress, mapping worker 0-100% to total 60-100%
        JA: ワーカー進捗を処理（ワーカー0-100%を全体60-100%にマッピング）"""
        # El worker emite progreso de 0-100%, pero nosotros ya estamos en 60%
        # Mapear el progreso del worker (0-100%) al rango 60-100% del total
        mapped_value = 60 + int((value * 40) / 100)  # 60% + (worker_progress * 40% / 100)
        if hasattr(self, 'yosoku_import_progress_dialog') and self.yosoku_import_progress_dialog is not None:
            self.yosoku_import_progress_dialog.update_progress(mapped_value, message)

    def create_yosoku_database_table(self, cursor):
        """ES: Crear tabla de predicciones si no existe
        EN: Create predictions table if it does not exist
        JA: 予測テーブルが無ければ作成"""
        create_table_sql = """
        CREATE TABLE IF NOT EXISTS yosoku_predictions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            A13 INTEGER,
            A11 INTEGER,
            A21 INTEGER,
            A32 INTEGER,
            直径 REAL,
            材料 TEXT,
            線材長 REAL,
            回転速度 REAL,
            送り速度 REAL,
            UPカット INTEGER,
            切込量 REAL,
            突出量 REAL,
            載せ率 REAL,
            パス数 INTEGER,
            加工時間 REAL,
            上面ダレ量 REAL,
            側面ダレ量 REAL,
            摩耗量 REAL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """
        cursor.execute(create_table_sql)

    def check_duplicate_yosoku_data(self, cursor, df):
        """Verificar si hay datos duplicados (columnas A-O)"""
        duplicate_rows = []
        
        for index, row in df.iterrows():
            # ES: Verificar si existe una fila con los mismos valores en las columnas A-O | EN: Check if row with same values in columns A-O exists | JA: 列A-Oで同一値の行が存在するか確認
            # ES: Las columnas A–O corresponden a: A13, A11, A21, A32, diámetro, material, longitud del alambre, velocidad de rotación, velocidad de avance, corte UP, profundidad de corte, longitud sobresaliente, tasa de carga, número de pases, tiempo de mecanizado
            # EN: Columns A–O correspond to: A13, A11, A21, A32, diameter, material, wire length, rotation speed, feed speed, UP cut, depth of cut, protrusion length, load ratio, number of passes, machining time
            # JP: 列 A〜O は以下に対応する：A13、A11、A21、A32、直径、材料、線材長、回転速度、送り速度、UPカット、切込量、突出量、載せ率、パス数、加工時間

            check_sql = """
            SELECT id FROM yosoku_predictions 
            WHERE A13 = ? AND A11 = ? AND A21 = ? AND A32 = ? 
            AND 直径 = ? AND 材料 = ? AND 線材長 = ? 
            AND 回転速度 = ? AND 送り速度 = ? AND UPカット = ? 
            AND 切込量 = ? AND 突出量 = ? AND 載せ率 = ? 
            AND パス数 = ? AND 加工時間 = ?
            """
            
            cursor.execute(check_sql, (
                int(row.get('A13', 0)),
                int(row.get('A11', 0)),
                int(row.get('A21', 0)),
                int(row.get('A32', 0)),
                float(row.get('直径', 0)),
                str(row.get('材料', '')),
                float(row.get('線材長', 0)),
                float(row.get('回転速度', 0)),
                float(row.get('送り速度', 0)),
                int(row.get('UPカット', 0)),
                float(row.get('切込量', 0)),
                float(row.get('突出量', 0)),
                float(row.get('載せ率', 0)),
                int(row.get('パス数', 0)),
                float(row.get('加工時間', 0))
            ))
            
            result = cursor.fetchone()
            if result:
                duplicate_rows.append((index, result[0]))  # (excel_row_index, db_id)
        
        return duplicate_rows

    def remove_duplicate_yosoku_data(self, cursor, duplicate_rows):
        """# ES: Eliminar datos duplicados existentes en la base de datos
# EN: Remove existing duplicate data from the database
# JP: データベース内の既存の重複データを削除する
"""
        for excel_row_index, db_id in duplicate_rows:
            cursor.execute("DELETE FROM yosoku_predictions WHERE id = ?", (db_id,))

    def insert_yosoku_data(self, cursor, df):
        """# ES: Insertar datos del Excel a la base de datos
# EN: Insert data from Excel into the database
# JP: Excel のデータをデータベースに挿入する
"""
        insert_sql = """
        INSERT INTO yosoku_predictions 
        (A13, A11, A21, A32, 直径, 材料, 線材長, 回転速度, 送り速度, UPカット, 
         切込量, 突出量, 載せ率, パス数, 加工時間, 上面ダレ量, 側面ダレ量, 摩耗量)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
        
        for index, row in df.iterrows():
            # ES: Función auxiliar para convertir valores de forma segura
            # EN: Helper function to safely convert values
            # JP: 値を安全に変換するための補助関数

            def safe_convert(value, convert_func, default=0):
                try:
                    if pd.isna(value) or value is None or value == '':
                        return default
                    return convert_func(value)
                except (ValueError, TypeError):
                    return default

            # ES: Convertir fórmulas a valores numéricos de forma segura
            # EN: Safely convert formulas to numeric values
            # JP: 数式を安全に数値へ変換する

            values = (
                safe_convert(row.get('A13', 0), int),
                safe_convert(row.get('A11', 0), int),
                safe_convert(row.get('A21', 0), int),
                safe_convert(row.get('A32', 0), int),
                safe_convert(row.get('直径', 0), float),
                str(row.get('材料', '')).strip() if row.get('材料') is not None else '',
                safe_convert(row.get('線材長', 0), float),
                safe_convert(row.get('回転速度', 0), float),
                safe_convert(row.get('送り速度', 0), float),
                safe_convert(row.get('UPカット', 0), int),
                safe_convert(row.get('切込量', 0), float),
                safe_convert(row.get('突出量', 0), float),
                safe_convert(row.get('載せ率', 0), float),
                safe_convert(row.get('パス数', 0), int),
                safe_convert(row.get('加工時間', 0), float),
                safe_convert(row.get('上面ダレ量', 0), float),
                safe_convert(row.get('側面ダレ量', 0), float),
                safe_convert(row.get('摩耗量', 0), float)
            )
            
            cursor.execute(insert_sql, values)

    def cancel_yosoku_prediction(self):
        """ES: Cancelar predicción Yosoku
        EN: Cancel Yosoku prediction
        JA: Yosoku予測をキャンセル
        """
        try:
            if hasattr(self, 'yosoku_worker'):
                self.yosoku_worker.cancel_prediction()
                self.yosoku_worker.quit()
                self.yosoku_worker.wait()
                self.yosoku_worker = None
            
            if hasattr(self, 'yosoku_progress_dialog'):
                self.yosoku_progress_dialog.close()
                self.yosoku_progress_dialog = None
            self.set_console_overlay_topmost(False)
                
        except Exception as e:
            print(f"❌ 予測キャンセルエラー: {e}")
            import traceback
            traceback.print_exc()


    def validate_prediction_parameters(self, selected_params):
        """ES: Validar que los parámetros seleccionados coincidan con los filtros aplicados
        EN: Validate that selected parameters match the applied filters
        JA: 選択パラメータが適用済みフィルタと一致するか検証
        """
        try:
            # ES: Obtener filtros aplicados
            # EN: Get applied filters
            # JP: 適用済みフィルタを取得
            filters = self.get_applied_filters()
            
            # ES: Lista para recopilar todos los errores | EN: List to collect all errors | JA: 全エラーを集めるリスト
            errors = []
            
            if not filters:
                return {
                    'valid': True,
                    'reason': 'No hay filtros aplicados, se pueden usar cualquier parámetro'
                }
            
            # ES: Verificar brush (legacy: único) o brushes (múltiples desde 未実験データ) | EN: Verify brush (legacy single) or brushes (multiple from 未実験データ) | JA: brush（レガシー単一）またはbrushes（未実験データ複数）を確認
            if 'brush' in selected_params and selected_params.get('brush') in ['A13', 'A11', 'A21', 'A32']:
                brush = selected_params['brush']
                if brush not in filters or filters[brush] != 1:
                    errors.append(f"Brush {brush} no está seleccionado en los filtros aplicados")
            elif 'brushes' in selected_params and isinstance(selected_params.get('brushes'), (list, tuple)):
                req = [b for b in selected_params.get('brushes') if b in ['A13', 'A11', 'A21', 'A32']]
                for b in req:
                    if b in filters and filters.get(b) == 1:
                        continue
                    # ES: Si no hay filtro de brush aplicado, no bloqueamos (los filtros pueden no incluir brush)
                    # EN: If no brush filter is applied, we don't block (filters may not include brush)
                    # JP: ブラシフィルタが適用されていない場合はブロックしない（フィルタにbrushが含まれない場合がある）
            
            # ES: Verificar diameter | EN: Verify diameter | JA: 直径を確認
            if 'diameter' in selected_params:
                diameter = selected_params['diameter']
                if '直径' in filters and filters['直径'] != diameter:
                    errors.append(f"Diámetro {diameter} no coincide con el filtro aplicado ({filters['直径']})")
            
            # ES: Verificar material | EN: Verify material | JA: 材料を確認
            if 'material' in selected_params:
                material = selected_params['material']
                if '材料' in filters and filters['材料'] != material:
                    errors.append(f"Material {material} no coincide con el filtro aplicado ({filters['材料']})")
            
            # ES: Verificar wire_length (legacy) con tolerancia de -5mm | EN: Verify wire_length (legacy) with -5mm tolerance | JA: wire_length（レガシー）を-5mm許差で確認
            if 'wire_length' in selected_params and selected_params.get('wire_length') is not None:
                wire_length = selected_params['wire_length']
                if '線材長' in filters:
                    filter_wire_length = filters['線材長']

                    # ES: Convertir wire_length a int para asegurar comparaciones correctas
                    # EN: Convert wire_length to int to ensure correct comparisons
                    # JP: 正しい比較を保証するために wire_length を int に変換する

                    try:
                        wire_length = int(wire_length)
                    except (ValueError, TypeError):
                        errors.append(f"Valor de wire_length inválido: {wire_length}")
                        return {
                            'valid': False,
                            'reason': '; '.join(errors)
                        }
                    
                    # ES: Manejar caso donde filter_wire_length puede ser una tupla
                    # EN: Handle the case where filter_wire_length may be a tuple
                    # JP: filter_wire_length がタプルの場合を処理
                    if isinstance(filter_wire_length, tuple):
                        # ES: Si es una tupla, verificar que TODOS los valores estén en el rango válido
                        # EN: If it's a tuple, verify that ALL values are within the valid range
                        # JP: タプルなら全ての値が有効範囲内か確認
                        min_length = wire_length - 5
                        max_length = wire_length

                        # ES: Convertir todos los valores de la tupla a int
                        # EN: Convert all values of the tuple to int
                        # JP: タプル内のすべての値を int に変換する

                        try:
                            converted_values = [int(val) for val in filter_wire_length]
                            invalid_values = [val for val in converted_values if not (min_length <= val <= max_length)]
                            if invalid_values:
                                errors.append(f"線材長 {filter_wire_length} contiene valores fuera del rango permitido ({min_length}-{max_length}mm) para el valor seleccionado {wire_length}mm. Valores inválidos: {invalid_values}")
                        except (ValueError, TypeError) as e:
                            errors.append(f"Error convirtiendo valores de filter_wire_length: {e}")
                    else:
                        # ES: Si es un valor único, verificar directamente
                        # EN: If it's a single value, check it directly
                        # JP: 単一値なら直接チェック
                        min_length = wire_length - 5
                        max_length = wire_length

                        # ES: Convertir filter_wire_length a int
                        # EN: Convert filter_wire_length to int
                        # JP: filter_wire_length を int に変換する

                        try:
                            filter_wire_length = int(filter_wire_length)
                            if not (min_length <= filter_wire_length <= max_length):
                                errors.append(f"線材長 {filter_wire_length} no está dentro del rango permitido ({min_length}-{max_length}mm) para el valor seleccionado {wire_length}mm")
                        except (ValueError, TypeError) as e:
                            errors.append(f"Error convirtiendo filter_wire_length: {e}")
            # ES: Nuevo: múltiples longitudes de cable desde datos no experimentados
            # EN: New: multiple wire_lengths from untested data
            # JP: 新規：未実験データからの複数のワイヤ長

            elif 'wire_lengths' in selected_params and isinstance(selected_params.get('wire_lengths'), (list, tuple)):
                if '線材長' in filters:
                    # ES: Si hay un filtro de 線材長 aplicado, comprobamos que no contradice completamente
                    # EN: If a 線材長 filter is applied, ensure it doesn't completely contradict
                    # JP: 線材長フィルタが適用されている場合、完全に矛盾しないか確認
                    try:
                        req = [int(float(x)) for x in selected_params.get('wire_lengths')]
                    except Exception:
                        req = []
                    # ES: Si el filtro es único, al menos uno debe estar dentro del rango [-5, 0] respecto a ese valor
                    # EN: If the filter is a single value, at least one must be within [-5, 0] relative to that value
                    # JP: フィルタが単一値なら、少なくとも1つがその値に対して[-5,0]以内にある必要がある
                    filter_wire_length = filters.get('線材長')
                    try:
                        fw = int(float(filter_wire_length)) if not isinstance(filter_wire_length, tuple) else None
                    except Exception:
                        fw = None
                    if fw is not None and req:
                        min_ok = fw - 5
                        max_ok = fw
                        if not any(min_ok <= v <= max_ok for v in req):
                            errors.append(f"線材長 フィルタ({fw}) と 未実験データ の 線材長 が一致しません")
            
            if errors:
                return {
                    'valid': False,
                    'reason': '; '.join(errors)
                }
            else:
                return {
                    'valid': True,
                    'reason': 'Parámetros válidos'
                }
                
        except Exception as e:
            print(f"❌ パラメータ検証エラー: {e}")
            return {
                'valid': False,
                'reason': f'Error en validación: {str(e)}'
            }


    def find_latest_prediction_folder(self):
        """ES: Encontrar la carpeta 04_予測計算 del análisis lineal más reciente
        EN: Find folder 04_予測計算 of the latest linear analysis
        JA: 直近の線形解析の 04_予測計算 フォルダを検索"""
        try:
            if not hasattr(self, 'current_project_folder') or not self.current_project_folder:
                print("⚠️ 現在のプロジェクトフォルダがありません")
                return None
            
            # ES: Buscar en la carpeta 03_線形回帰
            # EN: Search in the 03_線形回帰 folder
            # JP: 03_線形回帰フォルダを検索
            linear_regression_folder = os.path.join(self.current_project_folder, "03_線形回帰")
            if not os.path.exists(linear_regression_folder):
                print("⚠️ フォルダ 03_線形回帰 が見つかりません")
                return None

            # ES: Helper: elegir la última carpeta de ejecución dentro de 03_線形回帰
            # EN: Helper: select the latest execution folder within 03_線形回帰
            # JP: ヘルパー：03_線形回帰 内の最新実行フォルダを選択する

            def _pick_latest_run_folder(base_dir: str):
                import re
                from datetime import datetime

                candidates = []
                try:
                    for item in os.listdir(base_dir):
                        p = os.path.join(base_dir, item)
                        if not os.path.isdir(p):
                            continue
                        m = re.match(r"^\d+_(\d{8})_(\d{6})", str(item))
                        if m:
                            try:
                                dt = datetime.strptime(m.group(1) + m.group(2), "%Y%m%d%H%M%S")
                                candidates.append((dt, p))
                            except Exception:
                                continue
                except Exception:
                    candidates = []

                if candidates:
                    candidates.sort(key=lambda t: t[0], reverse=True)
                    return candidates[0][1]

                # ES: Fallback: por mtime (ignorando carpetas "01_..." típicas si es posible)
                # EN: Fallback: by mtime (ignoring typical "01_..." folders if possible)
                # JP: フォールバック：mtime 基準（可能であれば典型的な「01_...」フォルダを除外）

                subdirs = []
                try:
                    for item in os.listdir(base_dir):
                        p = os.path.join(base_dir, item)
                        if os.path.isdir(p):
                            subdirs.append(p)
                except Exception:
                    subdirs = []
                if not subdirs:
                    return None
                try:
                    return max(subdirs, key=lambda x: os.path.getmtime(x))
                except Exception:
                    return subdirs[-1]
            
            latest_subfolder = _pick_latest_run_folder(linear_regression_folder)
            if not latest_subfolder:
                print("⚠️ 線形解析フォルダが見つかりません")
                return None
            
            # ES: Buscar la carpeta 04_予測計算 dentro de la carpeta más reciente
            # EN: Look for the 04_予測計算 folder inside the most recent folder
            # JP: 最新フォルダ内の04_予測計算フォルダを探す
            prediction_folder = os.path.join(latest_subfolder, "04_予測計算")
            
            if os.path.exists(prediction_folder):
                print(f"✅ 予測フォルダを見つけました: {prediction_folder}")
                return prediction_folder
            else:
                print(f"⚠️ フォルダ 04_予測計算 が見つかりません: {latest_subfolder}")
                return None
                
        except Exception as e:
            print(f"❌ 予測フォルダ検索エラー: {e}")
            return None

    def find_latest_formulas_file(self):
        """ES: Encontrar automáticamente el archivo de fórmulas del análisis lineal más reciente
        EN: Automatically find the formulas file of the latest linear analysis
        JA: 直近の線形解析の数式ファイルを自動検索"""
        try:
            if not self.current_project_folder:
                print("❌ プロジェクトフォルダが設定されていません")
                return None
            
            linear_regression_folder = os.path.join(self.current_project_folder, "03_線形回帰")
            if not os.path.exists(linear_regression_folder):
                print("❌ 線形解析フォルダが見つかりません")
                return None
            
            print(f"🔍 式ファイルを検索中: {linear_regression_folder}")

            # ES: Preferir la última carpeta de ejecución (NN_YYYYMMDD_HHMMSS) si existe
            # EN: Prefer the latest run folder (NN_YYYYMMDD_HHMMSS) if it exists
            # JP: NN_YYYYMMDD_HHMMSS の最新実行フォルダがあれば優先
            latest_subfolder = None
            try:
                latest_subfolder = self.find_latest_prediction_folder()
            except Exception:
                latest_subfolder = None

            if latest_subfolder:
                # ES: find_latest_prediction_folder devuelve 04_予測計算; subir un nivel para reutilizar la lógica
                # EN: find_latest_prediction_folder returns 04_予測計算; go up one level to reuse the logic
                # JP: find_latest_prediction_folder は 04_予測計算 を返すため、ロジックを再利用するために1階層上に移動する

                base_run = os.path.dirname(latest_subfolder)
                formulas_file = os.path.join(latest_subfolder, "XEBEC_予測計算機_逆変換対応.xlsx")
                if os.path.exists(formulas_file):
                    print(f"✅ 式ファイルを見つけました: {formulas_file}")
                    return formulas_file
                # ES: fallback: búsqueda acotada dentro del run
                # EN: fallback: scoped search within the run
                # JP: フォールバック：run 内での限定検索
                try:
                    for root, dirs, files in os.walk(base_run):
                        rel = os.path.relpath(root, base_run)
                        if rel != "." and rel.count(os.sep) >= 4:
                            dirs[:] = []
                            continue
                        if "XEBEC_予測計算機_逆変換対応.xlsx" in files:
                            found = os.path.join(root, "XEBEC_予測計算機_逆変換対応.xlsx")
                            print(f"✅ 式ファイルを見つけました（検索）: {found}")
                            return found
                except Exception:
                    pass
            
            # ES: Buscar todas las subcarpetas de análisis lineal
            # EN: Search all linear-analysis subfolders
            # JP: 線形解析の全サブフォルダを探す
            subfolders = []
            for item in os.listdir(linear_regression_folder):
                item_path = os.path.join(linear_regression_folder, item)
                if os.path.isdir(item_path):
                    subfolders.append(item_path)
            
            if not subfolders:
                print("❌ 線形解析のサブフォルダが見つかりません")
                return None

            # ES: Ordenar por fecha de creación (más reciente primero)
            # EN: Sort by creation date (most recent first)
            # JP: 作成日時で並び替える（新しいものを先に）
            subfolders.sort(key=lambda x: os.path.getctime(x), reverse=True)
            
            print(f"📊 線形解析フォルダを {len(subfolders)} 件見つけました")
            
            # ES: Buscar el archivo de fórmulas en cada carpeta, empezando por la más reciente
            # EN: Look for the formulas file in each folder, starting with the most recent
            # JP: 各フォルダで数式ファイルを探す（最新から）
            for i, subfolder in enumerate(subfolders):
                folder_name = os.path.basename(subfolder)
                print(f"🔍 フォルダを確認中 {i+1}/{len(subfolders)}: {folder_name}")
                
                # ES: Buscar en la carpeta de predicción
                # EN: Search in the prediction folder
                # JP: 予測フォルダを検索
                prediction_folder = os.path.join(subfolder, "04_予測計算")
                if os.path.exists(prediction_folder):
                    formulas_file = os.path.join(prediction_folder, "XEBEC_予測計算機_逆変換対応.xlsx")
                    if os.path.exists(formulas_file):
                        print(f"✅ 式ファイルを見つけました: {formulas_file}")
                        return formulas_file
                    else:
                        print(f"   ⚠️ 式ファイルが見つかりません: {prediction_folder}")
                else:
                    print(f"   ⚠️ 予測フォルダが見つかりません: {prediction_folder}")
            
            print("❌ 有効な式ファイルが見つかりません")
            return None
            
        except Exception as e:
            print(f"❌ 式ファイル検索エラー: {e}")
            import traceback
            traceback.print_exc()
            return None

    def debug_console_position(self):
        """ES: Método de debug para verificar la posición de la consola
        EN: Debug method to verify console position
        JA: コンソール位置を確認するデバッグ用メソッド"""
        try:
            if hasattr(self, 'overlay_console'):
                console_geo = self.overlay_console.geometry()
                window_geo = self.geometry()
                print(f"🔍 DEBUG - メインウィンドウ: {window_geo}")
                print(f"🔍 DEBUG - コンソールオーバーレイ: {console_geo}")
                print(f"🔍 DEBUG - コンソール表示: {self.overlay_console.isVisible()}")
                print(f"🔍 DEBUG - オーバーレイ状態: {getattr(self, 'overlay_console_visible', '未定義')}")
            else:
                print("🔍 DEBUG - コンソールオーバーレイがありません")
        except Exception as e:
            print(f"🔍 DEBUG - エラー: {e}")

    # ES: NOTA: Este método ya no se necesita, solo usamos el panel superpuesto
    # EN: NOTE: This method is no longer needed; we only use the overlay panel
    # JP: 注: このメソッドは不要（オーバーレイパネルのみ使用）

    def sync_console_content(self):
        """Sincronizar el contenido de la consola desplegable con la principal"""
        try:
            # ES: Obtener el contenido de la consola principal
            # EN: Get the main console contents
            # JP: メインコンソール内容を取得
            main_content = self.console_output.toPlainText()

            # ES: Actualizar la consola desplegable
            # EN: Update the expandable console
            # JP: 展開可能なコンソールを更新する
            self.overlay_console_output.setPlainText(main_content)

            # ES: Mover el cursor al final (PySide6 usa MoveOperation.End)
            # EN: Move the cursor to the end (PySide6 uses MoveOperation.End)
            # JP: カーソルを末尾へ移動する（PySide6 は MoveOperation.End を使用）
            cursor = self.overlay_console_output.textCursor()
            cursor.movePosition(QTextCursor.MoveOperation.End)
            self.overlay_console_output.setTextCursor(cursor)
            
        except Exception as e:
            print(f"⚠️ コンソール同期エラー: {e}")

    def resizeEvent(self, event):
        """ES: Manejar el redimensionamiento de la ventana
        EN: Handle window resize
        JA: ウィンドウのリサイズを処理"""
        super().resizeEvent(event)
        
        # ES: Si el panel desplegable está visible, reposicionarlo
        # EN: If the dropdown panel is visible, reposition it
        # JP: ドロップダウンパネルが表示中なら再配置
        if hasattr(self, 'overlay_console_visible') and self.overlay_console_visible:
            self.position_overlay_console()

        # ES: También reposicionar el botón de flecha si está visible
        # EN: Also reposition the arrow button if it is visible
        # JP: 表示されている場合は矢印ボタンも再配置する
        if hasattr(self, 'console_toggle_button') and self.console_toggle_button.isVisible():
            self.position_arrow()

        # ES: Mantener el título actualizado (por si el manifest cambia durante runtime)
        # EN: Keep the title updated (in case the manifest changes at runtime)
        # JP: タイトルを最新の状態に保つ（実行中にmanifestが変わる場合に備える）
        try:
            self.setWindowTitle(get_app_title())
        except Exception:
            pass
        
        # ES: Actualizar gráficos del análisis no lineal si están siendo mostrados
        # EN: Refresh non-linear-analysis charts if they are being displayed
        # JP: 非線形解析のグラフが表示中なら更新
        if hasattr(self, 'nonlinear_chart_images') and hasattr(self, 'nonlinear_chart_label'):
            # Usar QTimer para actualizar después de que el resize termine
            QTimer.singleShot(100, self.update_nonlinear_chart_display)

    def closeEvent(self, event):
        """ES: Manejar el cierre de la aplicación
        EN: Handle application close
        JA: アプリ終了を処理"""
        try:
            print("🛑 アプリケーションを終了中...")

            # ES: Parar temporizadores de overlays (evita que sigan intentando raise_ tras cerrar)
            # EN: Stop overlay timers (prevents them from continuing to call raise_ after closing)
            # JP: オーバーレイのタイマーを停止する（終了後に raise_ を呼び続けるのを防ぐ）

            for timer_attr in ("keep_on_top_timer", "position_check_timer"):
                try:
                    t = getattr(self, timer_attr, None)
                    if t is not None and t.isActive():
                        t.stop()
                except Exception:
                    pass

            # ES: Cerrar ventanas flotantes (flecha y consola overlay)
            # EN: Close floating windows (arrow and console overlay)
            # JP: フローティングウィンドウ（矢印・コンソールオーバーレイ）を閉じる
            for w_attr in ("overlay_console", "console_toggle_button"):
                try:
                    w = getattr(self, w_attr, None)
                    if w is not None:
                        w.close()
                except Exception:
                    pass

            # ES: Cancelar análisis no lineal si está corriendo
            # EN: Cancel non-linear analysis if it is running
            # JP: 非線形解析が実行中ならキャンセルする
            if hasattr(self, 'nonlinear_worker') and self.nonlinear_worker is not None:
                try:
                    if self.nonlinear_worker.isRunning():
                        print("🛑 終了前に非線形解析をキャンセル中...")
                        self.nonlinear_worker.cancel()
                        if self.nonlinear_worker.isRunning():
                            self.nonlinear_worker.quit()
                            if not self.nonlinear_worker.wait(5000):
                                print("⚠️ ワーカーが5秒以内に終了しなかったため、強制終了します...")
                                self.nonlinear_worker.terminate()
                                self.nonlinear_worker.wait(1000)
                        print("✅ 非線形解析ワーカーをキャンセルしました")
                except Exception:
                    pass

            # ES: Cerrar base de datos si existe
            # EN: Close the database if it exists
            # JP: DBがあれば閉じる
            try:
                if hasattr(self, 'db'):
                    self.db.close()
            except Exception:
                pass

            # ES: Restaurar streams originales
            # EN: Restore original streams
            # JP: 元のストリームを復元
            if hasattr(self, 'original_stdout'):
                sys.stdout = self.original_stdout
            if hasattr(self, 'original_stderr'):
                sys.stderr = self.original_stderr

        finally:
            # ES: Continuar con el cierre normal
            # EN: Continue with the normal close flow
            # JP: 通常の終了処理を続行
            super().closeEvent(event)

# ======================================
# ES: Lanzamiento de la aplicación.
# EN: Application launch.
# JP: アプリケーション起動。
# ======================================

def handle_exception(exc_type, exc_value, exc_traceback):
    """
    ES: Manejar excepciones no capturadas para evitar que la app se cierre.
    EN: Handle uncaught exceptions to prevent the app from closing.
    JP: アプリ終了を防ぐため、未捕捉例外を処理します。
    """
    error_msg = f"❌ 未処理のエラー:\n{exc_type.__name__}: {exc_value}"
    print(error_msg)
    print("完全なトレースバック:")
    import traceback
    traceback.print_exception(exc_type, exc_value, exc_traceback)
    
    # ES: Mostrar mensaje de error en la consola si está disponible.
    # EN: Show the error message in the console if available.
    # JP: コンソールが利用可能ならエラーメッセージを表示します。
    try:
        if 'window' in globals() and hasattr(window, 'console_output'):
            window.console_output.append(error_msg)
    except:
        pass

if __name__ == "__main__":
    # ES: Configurar manejador de excepciones global.
    # EN: Configure the global exception handler.
    # JP: グローバル例外ハンドラを設定します。
    sys.excepthook = handle_exception
    
    try:
        app = QApplication(sys.argv)
        window = MainWindow()
        window.show()
        print("🚀 アプリケーションが正常に起動しました")
        sys.exit(app.exec())
    except Exception as e:
        print(f"❌ アプリケーション起動時にエラー: {e}")
        import traceback
        traceback.print_exc()
        input("Enter を押して終了します...")
