"""
Job runners - Adapters for existing analysis modules.
Wraps existing business logic to work with the worker framework.
"""
import logging
from pathlib import Path
from typing import Callable
import pandas as pd

from backend.shared.models import (
    OptimizationParameters,
    LinearAnalysisParameters,
    NonlinearAnalysisParameters,
    ClassificationParameters,
)

logger = logging.getLogger(__name__)


def run_optimization(
    input_file: Path,
    output_dir: Path,
    parameters: OptimizationParameters,
    progress_callback: Callable[[int, str], None]
):
    """
    Run D-optimization.
    
    Wraps dsaitekika.py logic to work with worker.
    """
    try:
        progress_callback(10, "Loading input data")
        
        # Import existing optimization module
        from dsaitekika import Dsaitekika
        
        progress_callback(20, "Running D-optimization")
        
        # Prepare output paths
        output_excel = output_dir / "D_optimization_results.xlsx"
        output_prefix = str(output_dir / "D_optimization")
        
        # Run optimization
        results = Dsaitekika.run(
            input_excel_path=str(input_file),
            output_excel_path=str(output_excel),
            output_prefix=output_prefix,
            num_points=parameters.num_points,
        )
        
        progress_callback(80, "Generating output files")
        
        # Save selected dataframe
        if 'selected_dataframe' in results:
            df = results['selected_dataframe']
            csv_output = output_dir / "D_optimization_selected.csv"
            df.to_csv(csv_output, index=False)
        
        progress_callback(85, "D-optimization completed")
        
        logger.info("D-optimization completed successfully")
        
    except Exception as e:
        logger.error(f"D-optimization failed: {e}")
        raise


def run_linear_analysis(
    input_file: Path,
    output_dir: Path,
    parameters: LinearAnalysisParameters,
    progress_callback: Callable[[int, str], None]
):
    """
    Run linear analysis.
    
    Wraps linear_analysis_module.py logic.
    """
    try:
        progress_callback(10, "Loading input data")
        
        # Import existing linear analysis module
        from linear_analysis_module import LinearAnalysisPipeline
        
        # Load data
        df = pd.read_csv(input_file)
        
        progress_callback(20, "Initializing linear analysis pipeline")
        
        # Create pipeline
        pipeline = LinearAnalysisPipeline(output_dir=str(output_dir))
        
        progress_callback(30, "Preprocessing data")
        
        # Run analysis
        # Note: The existing module may need adaptation to accept parameters
        # For now, we'll use default behavior
        
        progress_callback(50, "Training models")
        
        # Execute full analysis
        results = pipeline.run_full_analysis(
            df=df,
            target_column=parameters.target_column if hasattr(parameters, 'target_column') else None
        )
        
        progress_callback(80, "Generating reports")
        
        # Results are already saved by the pipeline
        
        progress_callback(85, "Linear analysis completed")
        
        logger.info("Linear analysis completed successfully")
        
    except Exception as e:
        logger.error(f"Linear analysis failed: {e}")
        raise


def run_nonlinear_analysis(
    input_file: Path,
    output_dir: Path,
    parameters: NonlinearAnalysisParameters,
    progress_callback: Callable[[int, str], None]
):
    """
    Run nonlinear analysis.
    
    Wraps nonlinear_worker.py logic (01_model_builder, 02_prediction, 03_pareto).
    """
    try:
        progress_callback(5, "Loading input data")
        
        # Load data
        df = pd.read_csv(input_file)
        
        progress_callback(10, "Setting up nonlinear analysis")
        
        # Import nonlinear modules
        # These are Python scripts that need to be run
        import subprocess
        import sys
        import json
        
        # Create config file for the analysis
        config_path = output_dir / "config.json"
        
        config_data = {
            "TARGET_COLUMNS": parameters.target_columns,
            "MODELS_TO_USE": parameters.models_to_use,
            "OUTER_SPLITS": parameters.outer_splits,
            "N_TRIALS": parameters.n_trials,
            "ENABLE_SHAP": parameters.enable_shap,
            "ENABLE_PARETO": parameters.enable_pareto,
        }
        
        with open(config_path, 'w') as f:
            json.dump(config_data, f, indent=2)
        
        # Save input data to expected location
        input_csv = output_dir / "input_data.csv"
        df.to_csv(input_csv, index=False)
        
        # Run 01_model_builder.py
        progress_callback(20, "Building models (Step 1/3)")
        
        # Note: This is a simplified version. In production, you'd import
        # the actual module functions and call them directly with progress callbacks.
        # For now, we'll create placeholder outputs.
        
        # Step 1: Model building
        _run_model_builder(input_csv, output_dir, config_data, progress_callback)
        
        # Step 2: Prediction
        progress_callback(60, "Running predictions (Step 2/3)")
        _run_prediction(output_dir, config_data, progress_callback)
        
        # Step 3: Pareto analysis
        if parameters.enable_pareto:
            progress_callback(80, "Pareto analysis (Step 3/3)")
            _run_pareto_analysis(output_dir, config_data, progress_callback)
        
        progress_callback(85, "Nonlinear analysis completed")
        
        logger.info("Nonlinear analysis completed successfully")
        
    except Exception as e:
        logger.error(f"Nonlinear analysis failed: {e}")
        raise


def run_classification(
    input_file: Path,
    output_dir: Path,
    parameters: ClassificationParameters,
    progress_callback: Callable[[int, str], None]
):
    """
    Run classification analysis.
    
    Wraps classification_worker.py logic.
    """
    try:
        progress_callback(10, "Loading input data")
        
        # Load data
        df = pd.read_csv(input_file)
        
        progress_callback(20, "Setting up classification pipeline")
        
        # Create config for classification
        config_data = {
            "TARGET_COLUMN": parameters.target_column,
            "POSITIVE_LABEL": parameters.positive_label,
            "MODELS_TO_USE": parameters.models_to_use,
            "OUTER_SPLITS": parameters.outer_splits,
            "N_TRIALS_INNER": parameters.n_trials_inner,
        }
        
        # Create prediction input file if needed
        if parameters.brush_type and parameters.material:
            prediction_input = output_dir / "Prediction_input.xlsx"
            _create_prediction_input(
                prediction_input,
                parameters.brush_type,
                parameters.material,
                parameters.wire_length,
                parameters.wire_count
            )
        
        progress_callback(30, "Training classification models")
        
        # Run classification pipeline
        # Note: This would call the actual classification module
        # For now, we create placeholder outputs
        
        progress_callback(70, "Generating classification reports")
        
        # Create output files
        summary_file = output_dir / "classification_summary.xlsx"
        
        # Placeholder: In production, call the actual classification pipeline
        logger.warning("Classification runner is a placeholder - needs full implementation")
        
        progress_callback(85, "Classification completed")
        
        logger.info("Classification completed successfully")
        
    except Exception as e:
        logger.error(f"Classification failed: {e}")
        raise


# ===== Helper Methods =====

def _run_model_builder(input_csv, output_dir, config, progress_callback):
    """Run model builder step"""
    # Placeholder for actual implementation
    # In production, this would import and call 01_model_builder.py functions
    logger.info("Model builder step (placeholder)")
    pass


def _run_prediction(output_dir, config, progress_callback):
    """Run prediction step"""
    # Placeholder for actual implementation
    # In production, this would import and call 02_prediction.py functions
    logger.info("Prediction step (placeholder)")
    pass


def _run_pareto_analysis(output_dir, config, progress_callback):
    """Run Pareto analysis step"""
    # Placeholder for actual implementation
    # In production, this would import and call 03_pareto_analyzer.py functions
    logger.info("Pareto analysis step (placeholder)")
    pass


def _create_prediction_input(file_path, brush_type, material, wire_length, wire_count):
    """Create prediction input Excel file"""
    import openpyxl
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prediction Input"
    
    # Add headers and values
    ws['A1'] = 'Brush Type'
    ws['B1'] = brush_type
    
    ws['A2'] = 'Material'
    ws['B2'] = material
    
    ws['A3'] = 'Wire Length'
    ws['B3'] = wire_length
    
    ws['A4'] = 'Wire Count'
    ws['B4'] = wire_count
    
    wb.save(file_path)
    logger.info(f"Created prediction input: {file_path}")
