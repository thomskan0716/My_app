#!/usr/bin/env python
# coding: utf-8

# In[2]:


#V60 (A11, A21, A32除外版)
"""
変換対応統合機械学習パイプライン
逆変換対応Excel予測計算機付き
A11, A21, A32を特徴量から除外
"""

import pandas as pd
import numpy as np
import os
import json
import joblib
import warnings
from pathlib import Path
from typing import Dict, List, Tuple, Any, Optional, Union

# 統計・機械学習
from sklearn.model_selection import KFold, GridSearchCV, cross_val_score, StratifiedKFold
from sklearn.preprocessing import StandardScaler, LabelEncoder, RobustScaler, OneHotEncoder, PowerTransformer
from sklearn.linear_model import (
    LinearRegression, Ridge, Lasso, ElasticNet,
    LogisticRegression, RidgeClassifier
)
from sklearn.discriminant_analysis import LinearDiscriminantAnalysis
from sklearn.ensemble import RandomForestRegressor, RandomForestClassifier
from sklearn.feature_selection import SelectKBest, f_classif, f_regression
from sklearn.metrics import (
    mean_absolute_error, mean_squared_error, r2_score,
    classification_report, confusion_matrix, f1_score,
    precision_recall_curve, roc_curve, auc,
    accuracy_score, precision_score, recall_score
)

# 統計的検定
from scipy import stats
from scipy.stats import shapiro, kstest, anderson, boxcox, yeojohnson

# 可視化
import matplotlib.pyplot as plt
import matplotlib as mpl
import seaborn as sns
plt.style.use('default')

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

def set_japanese_font():
    """日本語フォント設定"""
    try:
        if os.name == 'nt':
            fonts = ['MS Gothic', 'Yu Gothic', 'Meiryo']
        else:
            fonts = ['IPAexGothic', 'Hiragino Sans', 'Noto Sans CJK JP', 'IPAGothic', 'VL Gothic']
        for font in fonts:
            try:
                mpl.rcParams['font.family'] = font
                mpl.rcParams['font.size'] = 12
                mpl.rcParams['figure.dpi'] = 300
                mpl.rcParams['savefig.dpi'] = 300
                break
            except:
                continue
    except Exception as e:
        pass

set_japanese_font()
warnings.filterwarnings('ignore')

class TransformationAnalyzer:
    """変換分析システム"""

    def __init__(self, mode='simple', alpha=0.05):
        self.mode = mode
        self.alpha = alpha
        self.transformation_history = {}

        self.log_favorable_patterns = [
            '濃度', '速度', '量', '率', '密度', '粘度',
            '導電率', '反応速度', '拡散係数', '表面積',
            '摩耗', 'ダレ', '除去', '収率', '効率'
        ]

    def analyze_and_transform(self, y: pd.Series, X: pd.DataFrame, target_name: str) -> Tuple[pd.Series, Dict]:
        """変換分析と実行"""
        if self.mode == 'simple':
            return self._simple_transformation_analysis(y, target_name)
        else:
            return self._advanced_transformation_analysis(y, X, target_name)

    def _simple_transformation_analysis(self, y: pd.Series, target_name: str) -> Tuple[pd.Series, Dict]:
        """簡潔版変換分析"""
        transformation_info = {
            'applied': False,
            'method': 'none',
            'parameters': {},
            'original_stats': self._calculate_basic_stats(y),
            'reason': 'no_transformation'
        }

        try:
            if len(y) <= 3:
                transformation_info['reason'] = 'insufficient_data'
                return y, transformation_info

            if y.min() <= 0:
                transformation_info['reason'] = 'negative_values'
                return y, transformation_info

            # 正規性検定
            sample_size = min(len(y), 5000)
            y_sample = y.sample(sample_size, random_state=42) if len(y) > 5000 else y
            stat, p_value = shapiro(y_sample)

            if np.isnan(p_value) or np.isinf(p_value):
                transformation_info['reason'] = 'test_failed'
                return y, transformation_info

            if p_value >= self.alpha:
                transformation_info['reason'] = 'already_normal'
                return y, transformation_info

            # Log変換試行
            y_log = np.log(y)
            y_log_sample = y_log.sample(sample_size, random_state=42) if len(y_log) > 5000 else y_log
            stat_log, p_log = shapiro(y_log_sample)

            if np.isnan(p_log) or np.isinf(p_log):
                transformation_info['reason'] = 'log_test_failed'
                return y, transformation_info

            # 改善判定
            improvement_threshold = 0.01
            p_value_safe = max(p_value, 1e-10)
            p_log_safe = max(p_log, 1e-10)

            if p_log_safe > p_value_safe + improvement_threshold:
                transformation_info.update({
                    'applied': True,
                    'method': 'log',
                    'parameters': {'base': 'natural'},
                    'transformed_stats': self._calculate_basic_stats(y_log),
                    'original_p_value': p_value,
                    'transformed_p_value': p_log,
                    'reason': 'significant_improvement'
                })
                return y_log, transformation_info
            else:
                transformation_info['reason'] = 'insufficient_improvement'
                return y, transformation_info

        except Exception as e:
            transformation_info['reason'] = f'error: {str(e)}'
            return y, transformation_info

    def _advanced_transformation_analysis(self, y: pd.Series, X: pd.DataFrame, target_name: str) -> Tuple[pd.Series, Dict]:
        """高度版変換分析"""
        chemical_preference = self._check_chemical_rationality(target_name, y)
        transformation_candidates = self._generate_transformation_candidates(y)
        statistical_scores = self._evaluate_statistical_properties(transformation_candidates)
        model_performance_scores = self._evaluate_model_performance(transformation_candidates, X)

        best_transformation = self._select_best_transformation(
            statistical_scores, model_performance_scores, chemical_preference, target_name
        )

        transformation_info = self._create_detailed_transformation_info(
            best_transformation, transformation_candidates, y, target_name
        )

        transformed_data = transformation_candidates[best_transformation['name']]
        return transformed_data, transformation_info

    def _calculate_basic_stats(self, data: pd.Series) -> dict:
        """基本統計量計算"""
        return {
            'mean': float(np.mean(data)),
            'std': float(np.std(data)),
            'min': float(np.min(data)),
            'max': float(np.max(data)),
            'skewness': float(stats.skew(data)),
            'kurtosis': float(stats.kurtosis(data))
        }

    def _check_chemical_rationality(self, target_name: str, y: pd.Series) -> dict:
        """化学的合理性チェック"""
        preference_score = 0
        reasons = []

        for pattern in self.log_favorable_patterns:
            if pattern in target_name:
                preference_score += 2
                reasons.append(f"化学測定量({pattern})は対数正規分布が自然")

        if y.max() / y.min() > 100:
            preference_score += 3
            reasons.append(f"広い値域({y.min():.2e}-{y.max():.2e})は対数スケールが適切")

        if y.min() > 0:
            preference_score += 1
            reasons.append("正値のみ→対数変換可能")
        else:
            preference_score -= 2
            reasons.append("負値含有→対数変換困難")

        skewness = stats.skew(y)
        if skewness > 1:
            preference_score += 2
            reasons.append(f"強い右偏り(歪度:{skewness:.2f})→対数変換で改善期待")
        elif skewness > 0.5:
            preference_score += 1
            reasons.append(f"右偏り(歪度:{skewness:.2f})→対数変換で軽度改善期待")

        return {
            'preference_score': preference_score,
            'reasons': reasons,
            'recommendation': 'log_preferred' if preference_score >= 3 else 'neutral'
        }

    def _generate_transformation_candidates(self, y: pd.Series) -> dict:
        """変換候補の生成"""
        candidates = {'original': y.copy()}
        candidate_info = {'original': {'method': 'none', 'parameters': {}}}

        try:
            if y.min() > 0:
                candidates['log'] = np.log(y)
                candidate_info['log'] = {'method': 'log', 'parameters': {'base': 'natural'}}

                candidates['log10'] = np.log10(y)
                candidate_info['log10'] = {'method': 'log10', 'parameters': {'base': 10}}

                candidates['sqrt'] = np.sqrt(y)
                candidate_info['sqrt'] = {'method': 'sqrt', 'parameters': {}}

            if y.min() > 0:
                try:
                    transformed_boxcox, lambda_boxcox = boxcox(y)
                    candidates['boxcox'] = pd.Series(transformed_boxcox, index=y.index)
                    candidate_info['boxcox'] = {'method': 'boxcox', 'parameters': {'lambda': lambda_boxcox}}
                except:
                    pass

            try:
                pt = PowerTransformer(method='yeo-johnson', standardize=False)
                transformed_yj = pt.fit_transform(y.values.reshape(-1, 1)).flatten()
                candidates['yeo_johnson'] = pd.Series(transformed_yj, index=y.index)
                candidate_info['yeo_johnson'] = {
                    'method': 'yeo_johnson',
                    'parameters': {'lambda': pt.lambdas_[0]},
                    'transformer_object': pt
                }
            except:
                pass

        except Exception as e:
            pass

        self.candidate_info = candidate_info
        return candidates

    def _evaluate_statistical_properties(self, candidates: dict) -> dict:
        """統計的性質の評価"""
        results = {}

        for name, data in candidates.items():
            if len(data) < 3:
                continue

            scores = {}

            try:
                sample_size = min(len(data), 5000)
                sample_data = data.sample(sample_size, random_state=42) if len(data) > 5000 else data

                stat_sw, p_sw = shapiro(sample_data)
                scores['shapiro_p'] = p_sw

                standardized = (sample_data - np.mean(sample_data)) / np.std(sample_data)
                stat_ks, p_ks = kstest(standardized, 'norm')
                scores['ks_p'] = p_ks

                result_ad = anderson(standardized, dist='norm')
                critical_5pct = result_ad.critical_values[2]
                scores['anderson_pass'] = result_ad.statistic < critical_5pct

                scores['skewness'] = abs(stats.skew(sample_data))
                scores['kurtosis'] = abs(stats.kurtosis(sample_data))

                normality_score = 0
                if scores['shapiro_p'] > self.alpha:
                    normality_score += 3
                if scores['ks_p'] > self.alpha:
                    normality_score += 2
                if scores['anderson_pass']:
                    normality_score += 2
                if scores['skewness'] < 0.5:
                    normality_score += 1
                if scores['kurtosis'] < 1:
                    normality_score += 1

                scores['normality_score'] = normality_score

            except Exception as e:
                scores = {'normality_score': 0}

            results[name] = scores

        return results

    def _evaluate_model_performance(self, candidates: dict, X: pd.DataFrame) -> dict:
        """モデル性能による評価"""
        results = {}

        if len(X) < 10:
            return {name: {'r2': 0, 'rmse': float('inf')} for name in candidates.keys()}

        for name, y_transformed in candidates.items():
            try:
                model = LinearRegression()
                model.fit(X, y_transformed)
                y_pred = model.predict(X)

                r2 = r2_score(y_transformed, y_pred)
                rmse = np.sqrt(mean_squared_error(y_transformed, y_pred))

                residuals = y_transformed - y_pred
                if len(residuals) > 3:
                    try:
                        sample_size = min(len(residuals), 5000)
                        residual_sample = residuals.sample(sample_size, random_state=42) if len(residuals) > 5000 else residuals
                        _, residual_p = shapiro(residual_sample)
                    except:
                        residual_p = 0
                else:
                    residual_p = 0

                results[name] = {
                    'r2': r2,
                    'rmse': rmse,
                    'residual_normality_p': residual_p
                }

            except Exception as e:
                results[name] = {'r2': 0, 'rmse': float('inf'), 'residual_normality_p': 0}

        return results

    def _select_best_transformation(self, statistical_scores: dict,
                                   model_scores: dict, chemical_preference: dict,
                                   target_name: str) -> dict:
        """最適変換の選択"""
        transformation_rankings = {}

        for name in statistical_scores.keys():
            if name not in model_scores:
                continue

            score = 0
            details = []

            stat_score = statistical_scores[name].get('normality_score', 0)
            score += stat_score * 0.4
            details.append(f"正規性:{stat_score}")

            r2 = model_scores[name].get('r2', 0)
            model_score = min(max(r2, 0) * 10, 5)
            score += model_score * 0.35
            details.append(f"R²:{r2:.3f}")

            residual_p = model_scores[name].get('residual_normality_p', 0)
            residual_score = 3 if residual_p > 0.05 else 1
            score += residual_score * 0.15
            details.append(f"残差正規p:{residual_p:.3f}")

            if name in ['log', 'log10']:
                chem_score = chemical_preference['preference_score']
            elif name == 'original':
                chem_score = 2
            else:
                chem_score = 1
            score += chem_score * 0.1
            details.append(f"化学合理性:{chem_score}")

            transformation_rankings[name] = {
                'total_score': score,
                'details': details,
                'statistical': statistical_scores[name],
                'model_performance': model_scores[name]
            }

        best_name = max(transformation_rankings.keys(),
                       key=lambda x: transformation_rankings[x]['total_score'])

        return {
            'name': best_name,
            'score': transformation_rankings[best_name]['total_score'],
            'all_rankings': transformation_rankings
        }

    def _create_detailed_transformation_info(self, best_result: dict, candidates: dict,
                                           original_y: pd.Series, target_name: str) -> dict:
        """詳細な変換情報の作成"""
        best_name = best_result['name']

        transformation_info = {
            'applied': best_name != 'original',
            'method': best_name,
            'parameters': self.candidate_info[best_name]['parameters'].copy(),
            'original_stats': self._calculate_basic_stats(original_y),
            'transformed_stats': self._calculate_basic_stats(candidates[best_name]),
            'selection_score': best_result['score'],
            'all_candidates': {name: self.candidate_info[name] for name in candidates.keys()},
            'ranking': best_result['all_rankings']
        }

        if 'transformer_object' in transformation_info['parameters']:
            transformer = transformation_info['parameters'].pop('transformer_object')
            transformation_info['parameters']['lambda'] = transformer.lambdas_[0]
            transformation_info['yj_transformer'] = transformer

        return transformation_info

class InverseTransformer:
    """逆変換クラス"""

    @staticmethod
    def create_inverse_formula(transformation_info: dict, prediction_cell: str) -> str:
        """Excel用逆変換式の生成"""
        if not transformation_info['applied']:
            return f'={prediction_cell}'

        method = transformation_info['method']
        params = transformation_info['parameters']

        if method == 'log':
            return f'=EXP({prediction_cell})'

        elif method == 'log10':
            return f'=POWER(10,{prediction_cell})'

        elif method == 'sqrt':
            return f'=POWER({prediction_cell},2)'

        elif method == 'boxcox':
            lambda_val = params['lambda']
            if abs(lambda_val) < 1e-6:
                return f'=EXP({prediction_cell})'
            else:
                return f'=POWER({lambda_val}*{prediction_cell}+1,1/{lambda_val})'

        elif method == 'yeo_johnson':
            lambda_val = params['lambda']
            if abs(lambda_val) < 1e-6:
                return f'=EXP({prediction_cell})-1'
            else:
                return f'=POWER({lambda_val}*{prediction_cell}+1,1/{lambda_val})-1'

        else:
            return f'={prediction_cell}'

    @staticmethod
    def apply_inverse_transform(values: np.ndarray, transformation_info: dict) -> np.ndarray:
        """Python環境での逆変換実行"""
        if not transformation_info['applied']:
            return values

        method = transformation_info['method']
        params = transformation_info['parameters']

        if method == 'log':
            return np.exp(values)

        elif method == 'log10':
            return np.power(10, values)

        elif method == 'sqrt':
            return np.power(values, 2)

        elif method == 'boxcox':
            lambda_val = params['lambda']
            if abs(lambda_val) < 1e-6:
                return np.exp(values)
            else:
                return np.power(lambda_val * values + 1, 1/lambda_val)

        elif method == 'yeo_johnson':
            lambda_val = params['lambda']
            if abs(lambda_val) < 1e-6:
                return np.exp(values) - 1
            else:
                return np.power(lambda_val * values + 1, 1/lambda_val) - 1

        else:
            return values

class PipelineConfig:
    """パイプライン設定クラス"""

    TARGET_COLUMNS = ['バリ除去', '摩耗量', '上面ダレ量', '側面ダレ量']

    TARGET_TYPES = {
        'バリ除去': 'classification',
        '摩耗量': 'regression',
        '上面ダレ量': 'regression',
        '側面ダレ量': 'regression'
    }

    FEATURE_COLUMNS = [
        '送り速度', 'UPカット', '切込量',
        '突出量', '載せ率', '回転速度', 'パス数'
    ]

    CATEGORICAL_FEATURES = []

    INNER_CV_SPLITS = 10
    OUTER_CV_SPLITS = 10
    RANDOM_STATE = 42

    FEATURE_SELECTION = {
        'enable': True,
        'method': 'importance',
        'k_features': 10,
        'correlation_threshold': 0.95,
        'mandatory_features': []
    }

    TRANSFORMATION = {
        'enable': True,
        'mode': 'advanced',
        'alpha': 0.05,
        'improvement_threshold': 0.01
    }

    PREPROCESSING = {
        'use_robust_scaling': True,
        'missing_threshold': 0.5,
        'enable_noise_augmentation': False,
        'noise_ppm_range': (1, 10),
        'noise_augmentation_ratio': 0.2,
        'noise_feature_selection_method': 'manual',
        'noise_target_features': ['切込量', '突出量', '載せ率', '送り速度', '回転速度']
    }

    VISUALIZATION = {
        'dpi': 300,
        'save_format': 'png',
        'create_summary_plots': True,
        'show_metrics_on_plots': True
    }

class SmartFeatureSelector:
    """特徴選択クラス"""

    def __init__(self, method='importance', k=10, random_state=42, mandatory_features=None):
        self.method = method
        self.k = k
        self.random_state = random_state
        self.selected_features_ = None
        self.feature_scores_ = None
        self.mandatory_features = mandatory_features or []

    def fit_transform(self, X, y, feature_names, problem_type='classification'):
        """特徴選択の実行"""
        if self.mandatory_features:
            available_mandatory = [f for f in self.mandatory_features if f in feature_names]
        else:
            available_mandatory = []

        if self.method == 'importance':
            return self._importance_selection(X, y, feature_names, problem_type, available_mandatory)
        elif self.method == 'statistical':
            return self._statistical_selection(X, y, feature_names, problem_type, available_mandatory)
        else:
            return X, feature_names, None

    def _importance_selection(self, X, y, feature_names, problem_type, mandatory_features):
        """Random Forest重要度による特徴選択"""
        if problem_type == 'regression':
            model = RandomForestRegressor(n_estimators=100, random_state=self.random_state)
        else:
            model = RandomForestClassifier(n_estimators=100, random_state=self.random_state)

        model.fit(X, y)
        importances = model.feature_importances_

        feature_scores = dict(zip(feature_names, importances))
        sorted_features = sorted(feature_scores.items(), key=lambda x: x[1], reverse=True)

        selected_features = []

        for mandatory_feature in mandatory_features:
            if mandatory_feature in feature_names:
                selected_features.append(mandatory_feature)

        remaining_slots = self.k - len(selected_features)
        if remaining_slots > 0:
            for feature, importance in sorted_features:
                if feature not in selected_features and len(selected_features) < self.k:
                    selected_features.append(feature)

        if len(selected_features) > self.k:
            non_mandatory = [f for f in selected_features if f not in mandatory_features]
            non_mandatory_with_scores = [(f, feature_scores[f]) for f in non_mandatory]
            non_mandatory_with_scores.sort(key=lambda x: x[1], reverse=True)

            keep_count = self.k - len(mandatory_features)
            selected_features = mandatory_features + [f for f, _ in non_mandatory_with_scores[:keep_count]]

        selected_indices = [feature_names.index(f) for f in selected_features]
        X_selected = X[:, selected_indices]

        self.selected_features_ = selected_features
        self.feature_scores_ = feature_scores
        return X_selected, selected_features, feature_scores

    def _statistical_selection(self, X, y, feature_names, problem_type, mandatory_features):
        """統計的検定による特徴選択"""
        if problem_type == 'regression':
            score_func = f_regression
        else:
            score_func = f_classif

        available_mandatory = [f for f in mandatory_features if f in feature_names]
        effective_k = min(self.k, X.shape[1])

        if available_mandatory:
            mandatory_indices = [feature_names.index(f) for f in available_mandatory]
            non_mandatory_indices = [i for i in range(len(feature_names)) if i not in mandatory_indices]

            if non_mandatory_indices:
                X_non_mandatory = X[:, non_mandatory_indices]
                non_mandatory_names = [feature_names[i] for i in non_mandatory_indices]

                remaining_k = max(1, effective_k - len(available_mandatory))
                remaining_k = min(remaining_k, len(non_mandatory_names))

                selector = SelectKBest(score_func=score_func, k=remaining_k)
                X_selected_non_mandatory = selector.fit_transform(X_non_mandatory, y)

                selected_mask_non_mandatory = selector.get_support()
                selected_non_mandatory = [non_mandatory_names[i] for i in range(len(non_mandatory_names)) if selected_mask_non_mandatory[i]]

                selected_features = available_mandatory + selected_non_mandatory

                all_selected_indices = mandatory_indices + [non_mandatory_indices[i] for i in range(len(non_mandatory_indices)) if selected_mask_non_mandatory[i]]
                X_selected = X[:, all_selected_indices]

                all_scores = np.zeros(len(feature_names))
                all_scores[non_mandatory_indices] = selector.scores_
                feature_scores = dict(zip(feature_names, all_scores))
            else:
                selected_features = available_mandatory
                mandatory_indices = [feature_names.index(f) for f in available_mandatory]
                X_selected = X[:, mandatory_indices]
                feature_scores = dict(zip(feature_names, np.zeros(len(feature_names))))
        else:
            selector = SelectKBest(score_func=score_func, k=effective_k)
            X_selected = selector.fit_transform(X, y)

            selected_mask = selector.get_support()
            selected_features = [feature_names[i] for i in range(len(feature_names)) if selected_mask[i]]

            scores = selector.scores_
            feature_scores = dict(zip(feature_names, scores))

        self.selected_features_ = selected_features
        self.feature_scores_ = feature_scores
        return X_selected, selected_features, feature_scores

class IntegratedMLPipeline:
    """統合機械学習パイプライン"""

    def __init__(self, base_dir: str = ".", config: PipelineConfig = None):
        """初期化"""
        self.base_dir = Path(base_dir)
        self.config = config or PipelineConfig()
        np.random.seed(self.config.RANDOM_STATE)

        self.transformation_analyzer = TransformationAnalyzer(
            mode=self.config.TRANSFORMATION['mode'],
            alpha=self.config.TRANSFORMATION['alpha']
        )

        self.dirs = {
            'raw_data': self.base_dir / '01_raw_data',  # No se exporta
            'preprocessed': self.base_dir / '02_preprocessed',  # No se exporta
            'models': self.base_dir / '01_学習モデル',  # 03_models -> 01_学習モデル
            'parameters': self.base_dir / '02_パラメーター',  # 04_parameters -> 02_パラメーター
            'results': self.base_dir / '03_評価スコア',  # 05_results -> 03_評価スコア
            'predictions': self.base_dir / '04_予測計算'  # 06_predictions -> 04_予測計算
        }

        self.dirs['models_regression'] = self.dirs['models'] / 'regression'
        self.dirs['models_classification'] = self.dirs['models'] / 'classification'
        self.dirs['evaluation_graphs'] = self.dirs['results'] / '01_チャート'  # evaluation_graphs -> 01_チャート

        # Solo crear las carpetas que se van a exportar (no crear raw_data ni preprocessed)
        folders_to_create = ['models', 'parameters', 'results', 'predictions', 'models_regression', 'models_classification', 'evaluation_graphs']
        for folder_name in folders_to_create:
            self.dirs[folder_name].mkdir(parents=True, exist_ok=True)

        self.data = {}
        self.target_info = {}
        self.models = {}
        self.encoders = {}
        self.scalers = {}
        self.results = {}
        self.transformation_info = {}

    def load_data(self, file_path: str, index_col: str = 'Index') -> pd.DataFrame:
        """データ読み込み"""
        try:
            df = pd.read_excel(file_path)
            df.columns = df.columns.str.strip()

            index_found = False
            actual_index_col = None

            if index_col in df.columns:
                actual_index_col = index_col
                index_found = True
            else:
                for col in df.columns:
                    if str(col).lower().strip() == index_col.lower():
                        actual_index_col = col
                        index_found = True
                        break

                if not index_found:
                    for col in df.columns:
                        if index_col.lower() in str(col).lower() or str(col).lower() in index_col.lower():
                            actual_index_col = col
                            index_found = True
                            break

            if index_found and actual_index_col is not None:
                df.set_index(actual_index_col, inplace=True)
            else:
                df.index = range(len(df))

            self.data['raw'] = df
            return df

        except Exception as e:
            raise

    def separate_variables(self, target_columns: List[str] = None,
                          feature_columns: List[str] = None) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """変数分離"""
        df = self.data['raw']

        if target_columns is None:
            target_columns = [col for col in self.config.TARGET_COLUMNS if col in df.columns]

        if feature_columns is None:
            if hasattr(self.config, 'FEATURE_COLUMNS') and self.config.FEATURE_COLUMNS:
                feature_columns = [col for col in self.config.FEATURE_COLUMNS if col in df.columns]
            else:
                feature_columns = [col for col in df.columns if col not in target_columns]

        targets = df[target_columns].copy()
        features = df[feature_columns].copy()

        self.data['targets'] = targets
        self.data['features'] = features

        self._determine_task_types(targets)

        return targets, features

    def _determine_task_types(self, targets: pd.DataFrame):
        """タスク種別判定"""
        for col in targets.columns:
            if hasattr(self.config, 'TARGET_TYPES') and col in self.config.TARGET_TYPES:
                task_type = self.config.TARGET_TYPES[col]
            else:
                y = targets[col].dropna()
                unique_count = y.nunique()

                if unique_count <= 2:
                    task_type = 'binary_classification' if unique_count == 2 else 'constant'
                elif unique_count <= 10 and y.dtype.kind in 'iub':
                    task_type = 'multiclass_classification'
                else:
                    task_type = 'regression'

            y = targets[col].dropna()
            cv = y.std() / y.mean() if y.mean() != 0 else 0

            self.target_info[col] = {
                'task_type': task_type,
                'unique_count': y.nunique(),
                'dtype': str(y.dtype),
                'cv': cv,
                'range': [float(y.min()), float(y.max())],
                'stats': {
                    'mean': float(y.mean()),
                    'std': float(y.std())
                }
            }

    def preprocess_data(self) -> Dict[str, pd.DataFrame]:
        """統合前処理実行"""
        targets = self.data['targets'].copy()
        features = self.data['features'].copy()

        # 欠損値処理
        missing_info = {}
        threshold = self.config.PREPROCESSING['missing_threshold']

        for col in features.columns:
            missing_ratio = features[col].isnull().sum() / len(features)

            if missing_ratio > threshold:
                features.drop(columns=[col], inplace=True)
                missing_info[col] = 'dropped'
            elif missing_ratio > 0:
                if features[col].dtype in ['int64', 'float64']:
                    features[col].fillna(features[col].median(), inplace=True)
                    missing_info[col] = 'median_imputed'
                else:
                    features[col].fillna(features[col].mode()[0], inplace=True)
                    missing_info[col] = 'mode_imputed'

        # カテゴリカル変数のOneHotEncoding
        features_encoded = self._encode_categorical_features(features)

        # 高相関特徴量除去
        features_final = self._remove_high_correlation_features(features_encoded)

        # ノイズデータ生成
        if self.config.PREPROCESSING['enable_noise_augmentation']:
            features_final = self._generate_noise_augmented_data(features_final)
            targets = self.data['targets'].copy()

        # 特徴選択
        if self.config.FEATURE_SELECTION['enable']:
            features_final = self._apply_feature_selection_representative(features_final, targets)

        # スケーリング
        features_scaled = self._apply_scaling(features_final)

        target_nan_info = {}
        for col in targets.columns:
            nan_count = targets[col].isnull().sum()
            total_count = len(targets)
            target_nan_info[col] = {
                'nan_count': nan_count,
                'valid_count': total_count - nan_count,
                'nan_ratio': nan_count / total_count
            }

        self.data['processed'] = {
            'targets': targets,
            'features': features_scaled,
            'features_original': features_final,
            'target_nan_info': target_nan_info
        }

        self._save_preprocessing_params(missing_info)

        return self.data['processed']

    def _apply_feature_selection_representative(self, features, targets):
        """代表的な目的変数での特徴選択"""
        target_nan_counts = {col: targets[col].isnull().sum() for col in targets.columns}
        representative_target = min(target_nan_counts.keys(), key=lambda x: target_nan_counts[x])

        if len(features) != len(targets):
            if len(features) > len(targets):
                features_for_selection = features.iloc[:len(targets)]
                valid_mask = ~targets[representative_target].isnull()
                X_rep = features_for_selection.loc[valid_mask]
            else:
                targets_for_selection = targets.iloc[:len(features)]
                valid_mask = ~targets_for_selection[representative_target].isnull()
                X_rep = features.loc[valid_mask]
        else:
            valid_mask = ~targets[representative_target].isnull()
            X_rep = features.loc[valid_mask]

        y_rep = targets[representative_target].loc[valid_mask]

        task_type = self.target_info[representative_target]['task_type']

        if 'classification' in task_type:
            if task_type == 'binary_classification':
                y_for_selection = y_rep
            else:
                le = LabelEncoder()
                y_for_selection = le.fit_transform(y_rep)
            problem_type = 'classification'
        else:
            y_for_selection = y_rep
            problem_type = 'regression'

        mandatory_features = self.config.FEATURE_SELECTION.get('mandatory_features', [])

        selector = SmartFeatureSelector(
            method=self.config.FEATURE_SELECTION['method'],
            k=min(self.config.FEATURE_SELECTION['k_features'], features.shape[1]),
            random_state=self.config.RANDOM_STATE,
            mandatory_features=mandatory_features
        )

        X_selected, selected_features, feature_scores = selector.fit_transform(
            X_rep.values, y_for_selection,
            X_rep.columns.tolist(), problem_type
        )

        self.results['feature_selection'] = {
            'method': self.config.FEATURE_SELECTION['method'],
            'selected_features': selected_features,
            'feature_scores': feature_scores,
            'representative_target': representative_target,
            'mandatory_features': mandatory_features
        }

        return features[selected_features]

    def _generate_noise_augmented_data(self, features):
        """ppmレベルノイズ付きデータ生成"""
        if not self.config.PREPROCESSING['enable_noise_augmentation']:
            return features

        noise_ratio = self.config.PREPROCESSING['noise_augmentation_ratio']
        ppm_range = self.config.PREPROCESSING['noise_ppm_range']
        selection_method = self.config.PREPROCESSING['noise_feature_selection_method']
        target_features = self.config.PREPROCESSING['noise_target_features']

        original_size = len(features)
        augment_size = int(original_size * noise_ratio)

        if selection_method == 'manual':
            noise_features = [f for f in target_features if f in features.columns]
        else:
            noise_features = features.select_dtypes(include=[np.number]).columns.tolist()

        if not noise_features:
            return features

        augment_indices = np.random.choice(original_size, augment_size, replace=True)
        augmented_data = features.iloc[augment_indices].copy()

        for col in noise_features:
            if col in augmented_data.columns:
                ppm_levels = np.random.uniform(ppm_range[0], ppm_range[1], augment_size)
                noise_ratios = ppm_levels * 1e-6
                original_values = augmented_data[col].values
                noise = original_values * noise_ratios * np.random.normal(0, 1, augment_size)
                augmented_data[col] = original_values + noise

        features_reset = features.reset_index(drop=True)
        augmented_data_reset = augmented_data.reset_index(drop=True)
        features_with_noise = pd.concat([features_reset, augmented_data_reset], ignore_index=True)

        if hasattr(self, 'data') and 'targets' in self.data:
            targets_original = self.data['targets'].reset_index(drop=True)
            targets_augmented = targets_original.iloc[augment_indices].reset_index(drop=True)
            targets_with_noise = pd.concat([targets_original, targets_augmented], ignore_index=True)

            self.data['targets'] = targets_with_noise

        return features_with_noise

    def _encode_categorical_features(self, features):
        """カテゴリカル変数のOneHotEncoding"""
        categorical_columns = []
        if hasattr(self.config, 'CATEGORICAL_FEATURES') and self.config.CATEGORICAL_FEATURES:
            categorical_columns = [col for col in self.config.CATEGORICAL_FEATURES if col in features.columns]
        else:
            categorical_columns = features.select_dtypes(include=['object', 'category']).columns.tolist()

        if len(categorical_columns) > 0:
            encoder = OneHotEncoder(drop='first', sparse_output=False, handle_unknown='ignore')
            categorical_data = features[categorical_columns]
            encoded_data = encoder.fit_transform(categorical_data)
            feature_names = encoder.get_feature_names_out(categorical_columns)
            encoded_df = pd.DataFrame(encoded_data, columns=feature_names, index=features.index)
            features_encoded = features.drop(columns=categorical_columns)
            features_encoded = pd.concat([features_encoded, encoded_df], axis=1)
            self.encoders['onehot'] = encoder
        else:
            features_encoded = features

        return features_encoded

    def _remove_high_correlation_features(self, features):
        """高相関特徴量除去"""
        numeric_features = features.select_dtypes(include=[np.number])
        if len(numeric_features.columns) > 1:
            corr_matrix = numeric_features.corr().abs()
            upper_tri = np.triu(np.ones_like(corr_matrix, dtype=bool), k=1)
            high_corr_pairs = np.where((corr_matrix > self.config.FEATURE_SELECTION['correlation_threshold']) & upper_tri)

            mandatory_features = self.config.FEATURE_SELECTION.get('mandatory_features', [])

            drop_features = []
            for i, j in zip(high_corr_pairs[0], high_corr_pairs[1]):
                feature1 = corr_matrix.index[i]
                feature2 = corr_matrix.columns[j]

                if feature1 in mandatory_features and feature2 in mandatory_features:
                    continue
                elif feature1 in mandatory_features:
                    if feature2 not in drop_features:
                        drop_features.append(feature2)
                elif feature2 in mandatory_features:
                    if feature1 not in drop_features:
                        drop_features.append(feature1)
                else:
                    if feature2 not in drop_features:
                        drop_features.append(feature2)

            features = features.drop(columns=drop_features)

        return features

    def _apply_scaling(self, features):
        """スケーリングの適用"""
        numeric_columns = features.select_dtypes(include=[np.number]).columns

        if self.config.PREPROCESSING['use_robust_scaling']:
            scaler = RobustScaler()
        else:
            scaler = StandardScaler()

        features_scaled = features.copy()
        if len(numeric_columns) > 0:
            features_scaled[numeric_columns] = scaler.fit_transform(features[numeric_columns])

        self.scalers['features'] = scaler
        return features_scaled

    def train_models(self) -> Dict[str, Dict]:
        """モデル訓練"""
        targets = self.data['processed']['targets']
        features = self.data['processed']['features']

        best_models = {}

        for target_col in targets.columns:
            task_type = self.target_info[target_col]['task_type']

            valid_mask = ~targets[target_col].isnull()
            X = features[valid_mask]
            y_raw = targets[target_col][valid_mask]

            try:
                if task_type == 'regression':
                    if self.config.TRANSFORMATION['enable']:
                        y_processed, transformation_info = self.transformation_analyzer.analyze_and_transform(
                            y_raw, X, target_col
                        )
                        self.transformation_info[target_col] = transformation_info
                    else:
                        y_processed = y_raw
                        self.transformation_info[target_col] = {
                            'applied': False, 'method': 'none', 'parameters': {},
                            'reason': 'transformation_disabled'
                        }

                    best_model = self._train_regression_models(X, y_processed, target_col)

                elif 'classification' in task_type:
                    best_model = self._train_classification_models(X, y_raw, target_col)
                    self.transformation_info[target_col] = {
                        'applied': False, 'method': 'none', 'parameters': {},
                        'reason': 'classification_task'
                    }
                else:
                    best_model = {
                        'model': None, 'model_name': 'ConstantTarget',
                        'task_type': task_type, 'error': 'constant_target'
                    }
                    self.transformation_info[target_col] = {
                        'applied': False, 'method': 'none', 'parameters': {},
                        'reason': 'constant_target'
                    }

                best_models[target_col] = best_model

            except Exception as e:
                best_models[target_col] = {
                    'model': None, 'model_name': 'TrainingFailed',
                    'task_type': task_type, 'error': str(e)
                }
                self.transformation_info[target_col] = {
                    'applied': False, 'method': 'none', 'parameters': {},
                    'reason': f'training_error: {str(e)}'
                }

        self.models = best_models
        return best_models

    def _train_regression_models(self, X: pd.DataFrame, y: pd.Series, target_name: str) -> Dict:
        """回帰モデル訓練"""
        models = {
            'LinearRegression': LinearRegression(),
            'Ridge': Ridge(random_state=self.config.RANDOM_STATE),
            'Lasso': Lasso(random_state=self.config.RANDOM_STATE, max_iter=2000),
            'ElasticNet': ElasticNet(random_state=self.config.RANDOM_STATE, max_iter=2000),
            'RandomForest': RandomForestRegressor(random_state=self.config.RANDOM_STATE, n_estimators=100)
        }

        param_grids = {
            'Ridge': {'alpha': [0.1, 1.0, 10.0, 100.0]},
            'Lasso': {'alpha': [0.01, 0.1, 1.0, 10.0]},
            'ElasticNet': {'alpha': [0.01, 0.1, 1.0], 'l1_ratio': [0.1, 0.5, 0.9]},
            'RandomForest': {'n_estimators': [50, 100, 200], 'max_depth': [5, 10, None]}
        }

        best_model_name, model_scores, best_model_config = self._perform_double_cv_regression(
            X, y, models, param_grids
        )

        # RandomForest回避ロジック
        linear_models = ['LinearRegression', 'Ridge', 'Lasso', 'ElasticNet']

        if best_model_name == 'RandomForest':
            linear_scores = {k: v for k, v in model_scores.items() if k in linear_models and 'error' not in v}
            if linear_scores:
                actual_best_model_name = min(linear_scores.keys(), key=lambda x: linear_scores[x]['cv_mae_mean'])
                best_model_name = actual_best_model_name
                best_model_config = {'best_params': {}}

                if best_model_name in param_grids:
                    best_params_linear = self._get_best_params_for_linear_model(X, y, models[best_model_name], param_grids[best_model_name])
                    best_model_config['best_params'] = best_params_linear

        if best_model_name in param_grids:
            final_model = models[best_model_name].set_params(**best_model_config['best_params'])
        else:
            final_model = models[best_model_name]

        final_model.fit(X, y)

        y_pred = final_model.predict(X)
        mae = mean_absolute_error(y, y_pred)
        rmse = np.sqrt(mean_squared_error(y, y_pred))
        r2 = r2_score(y, y_pred)
        residuals = y - y_pred

        prediction_formula = self._extract_prediction_formula(final_model, X.columns, target_name)

        model_path = self.dirs['models_regression'] / f'best_model_{target_name}.pkl'
        model_save_data = {
            'model': final_model,
            'feature_names': X.columns.tolist(),
            'prediction_formula': prediction_formula,
            'scaler': self.scalers.get('features'),
            'target_name': target_name,
            'transformation_info': self.transformation_info.get(target_name, {})
        }
        joblib.dump(model_save_data, model_path)

        if self.config.VISUALIZATION['create_summary_plots']:
            metrics_dict = {'MAE': mae, 'RMSE': rmse, 'R²': r2}
            self._plot_regression_results_enhanced(y, y_pred, residuals, target_name, metrics_dict)

        return {
            'model': final_model,
            'model_name': best_model_name,
            'model_path': str(model_path),
            'best_params': best_model_config.get('best_params', {}),
            'cv_scores': model_scores,
            'double_cv_results': best_model_config,
            'final_metrics': {'mae': mae, 'rmse': rmse, 'r2': r2},
            'prediction_formula': prediction_formula,
            'task_type': 'regression',
            'transformation_info': self.transformation_info.get(target_name, {})
        }

    def _perform_double_cv_regression(self, X, y, models, param_grids):
        """真のダブルクロスバリデーション（回帰）"""
        outer_cv = KFold(n_splits=self.config.OUTER_CV_SPLITS, shuffle=True, random_state=self.config.RANDOM_STATE)
        inner_cv = KFold(n_splits=self.config.INNER_CV_SPLITS, shuffle=True, random_state=self.config.RANDOM_STATE)

        model_results = {}

        for model_name, model in models.items():
            outer_scores = []
            best_params_list = []

            for fold_idx, (train_idx, test_idx) in enumerate(outer_cv.split(X, y)):
                X_train_outer, X_test_outer = X.iloc[train_idx], X.iloc[test_idx]
                y_train_outer, y_test_outer = y.iloc[train_idx], y.iloc[test_idx]

                if model_name in param_grids:
                    grid_search = GridSearchCV(
                        model, param_grids[model_name],
                        cv=inner_cv, scoring='neg_mean_absolute_error', n_jobs=-1
                    )
                    grid_search.fit(X_train_outer, y_train_outer)
                    best_model = grid_search.best_estimator_
                    best_params_list.append(grid_search.best_params_)
                else:
                    best_model = model
                    best_model.fit(X_train_outer, y_train_outer)
                    best_params_list.append({})

                y_pred_outer = best_model.predict(X_test_outer)
                mae_outer = mean_absolute_error(y_test_outer, y_pred_outer)
                outer_scores.append(mae_outer)

            mean_mae = np.mean(outer_scores)
            std_mae = np.std(outer_scores)

            model_results[model_name] = {
                'cv_mae_mean': mean_mae,
                'cv_mae_std': std_mae,
                'outer_scores': outer_scores,
                'best_params_list': best_params_list
            }

        best_model_name = min(model_results.keys(), key=lambda x: model_results[x]['cv_mae_mean'])

        best_params = {}
        if model_results[best_model_name]['best_params_list'][0]:
            param_counts = {}
            for params in model_results[best_model_name]['best_params_list']:
                param_str = str(sorted(params.items()))
                param_counts[param_str] = param_counts.get(param_str, 0) + 1
            most_common_params_str = max(param_counts.keys(), key=lambda x: param_counts[x])
            best_params = dict(eval(most_common_params_str))

        return best_model_name, model_results, {'best_params': best_params}

    def _extract_prediction_formula(self, model, feature_names, target_name):
        """予測式の抽出"""
        formula_info = {
            'model_type': type(model).__name__,
            'target_name': target_name,
            'feature_names': feature_names.tolist(),
            'formula_string': None,
            'coefficients': None,
            'intercept': None
        }

        if hasattr(model, 'coef_') and hasattr(model, 'intercept_'):
            coefficients = model.coef_
            intercept = model.intercept_

            formula_parts = [f"{intercept:.6f}"]
            for coef, feature in zip(coefficients, feature_names):
                if coef >= 0:
                    formula_parts.append(f" + {coef:.6f} * {feature}")
                else:
                    formula_parts.append(f" - {abs(coef):.6f} * {feature}")

            formula_string = f"{target_name} = " + "".join(formula_parts)

            formula_info.update({
                'formula_string': formula_string,
                'coefficients': coefficients.tolist(),
                'intercept': float(intercept)
            })

        return formula_info

    def _train_classification_models(self, X: pd.DataFrame, y: pd.Series, target_name: str) -> Dict:
        """分類モデル訓練"""
        class_counts = y.value_counts()

        if len(class_counts) < 2:
            return {'model': None, 'model_name': 'SingleClass', 'task_type': 'classification', 'error': 'single_class'}

        if class_counts.min() < 5:
            return {'model': None, 'model_name': 'InsufficientSamples', 'task_type': 'classification', 'error': 'insufficient_samples'}

        le = LabelEncoder()
        y_encoded = le.fit_transform(y)

        model = LogisticRegression(random_state=self.config.RANDOM_STATE, max_iter=2000)
        model.fit(X, y_encoded)

        y_pred = model.predict(X)
        y_proba = model.predict_proba(X) if hasattr(model, 'predict_proba') else None

        accuracy = accuracy_score(y_encoded, y_pred)
        f1 = f1_score(y_encoded, y_pred, average='weighted')

        model_path = self.dirs['models_classification'] / f'best_model_{target_name}.pkl'
        model_save_data = {
            'model': model,
            'label_encoder': le,
            'feature_names': X.columns.tolist(),
            'target_name': target_name
        }
        joblib.dump(model_save_data, model_path)

        return {
            'model': model,
            'label_encoder': le,
            'model_name': 'LogisticRegression',
            'model_path': str(model_path),
            'final_metrics': {'accuracy': accuracy, 'f1_score': f1},
            'probabilities': y_proba,
            'task_type': 'classification'
        }

    def _get_best_params_for_linear_model(self, X, y, model, param_grid):
        """線形回帰モデルのベストパラメータ取得"""
        if not param_grid:
            return {}

        inner_cv = KFold(n_splits=self.config.INNER_CV_SPLITS, shuffle=True, random_state=self.config.RANDOM_STATE)
        grid_search = GridSearchCV(model, param_grid, cv=inner_cv, scoring='neg_mean_absolute_error', n_jobs=-1)
        grid_search.fit(X, y)
        return grid_search.best_params_

    def _plot_regression_results_enhanced(self, y_true: pd.Series, y_pred: np.ndarray,
                                         residuals: np.ndarray, target_name: str,
                                         metrics_dict: Dict[str, float]):
        """回帰結果の可視化"""
        fig, axes = plt.subplots(1, 2, figsize=(16, 7))

        axes[0].scatter(y_true, y_pred, alpha=0.6, s=50, color='blue')
        axes[0].plot([y_true.min(), y_true.max()], [y_true.min(), y_true.max()], 'r--', lw=2)
        axes[0].set_xlabel('実測値')
        axes[0].set_ylabel('予測値')
        axes[0].set_title(f'{target_name}: 実測値 vs 予測値')
        axes[0].grid(True, alpha=0.3)

        metrics_text = [f'{metric} = {value:.4f}' for metric, value in metrics_dict.items()]
        metrics_str = '\n'.join(metrics_text)
        axes[0].text(0.05, 0.95, metrics_str, transform=axes[0].transAxes,
                    bbox=dict(boxstyle='round,pad=0.5', facecolor='white', alpha=0.8),
                    verticalalignment='top', fontsize=11, fontweight='bold')

        axes[1].scatter(y_pred, residuals, alpha=0.6, s=50, color='green')
        axes[1].axhline(y=0, color='r', linestyle='--', linewidth=2)
        axes[1].set_xlabel('予測値')
        axes[1].set_ylabel('残差 (実測値 - 予測値)')
        axes[1].set_title(f'{target_name}: 残差プロット')
        axes[1].grid(True, alpha=0.3)

        plt.tight_layout()

        save_path = self.dirs['evaluation_graphs'] / f'regression_enhanced_{target_name}.png'
        plt.savefig(save_path, dpi=self.config.VISUALIZATION['dpi'], bbox_inches='tight')
        plt.close()

    def save_prediction_formulas(self):
        """予測式の一括保存"""
        formulas = {}
        for target_col, model_info in self.models.items():
            if 'prediction_formula' in model_info and model_info['prediction_formula']:
                formulas[target_col] = model_info['prediction_formula']

        if formulas:
            formula_path = self.dirs['parameters'] / 'prediction_formulas.json'
            with open(formula_path, 'w', encoding='utf-8') as f:
                json.dump(formulas, f, indent=2, ensure_ascii=False)

            readable_path = self.dirs['parameters'] / 'prediction_formulas_readable.txt'
            with open(readable_path, 'w', encoding='utf-8') as f:
                f.write("=== 予測式一覧 ===\n\n")
                for target, formula_info in formulas.items():
                    f.write(f"【{target}】\n")
                    f.write(f"モデル: {formula_info['model_type']}\n")
                    if formula_info['formula_string']:
                        f.write(f"予測式: {formula_info['formula_string']}\n")
                    else:
                        f.write("予測式: 非線形モデル（数式表現不可）\n")
                    f.write("\n" + "-"*50 + "\n\n")

    def save_results(self):
        """結果保存"""
        results_summary = []

        for target_col, model_info in self.models.items():
            row = {
                'Target': target_col,
                'Task_Type': model_info['task_type'],
                'Best_Model': model_info.get('model_name', 'Unknown'),
                'Best_Params': str(model_info.get('best_params', {}))
            }

            transformation_info = self.transformation_info.get(target_col, {})
            row['Transformation_Applied'] = transformation_info.get('applied', False)
            row['Transformation_Method'] = transformation_info.get('method', 'none')

            if model_info.get('model') is not None:
                if model_info['task_type'] == 'regression':
                    metrics = model_info.get('final_metrics', {})
                    row.update({
                        'MAE': metrics.get('mae', 'N/A'),
                        'RMSE': metrics.get('rmse', 'N/A'),
                        'R2_Score': metrics.get('r2', 'N/A'),
                        'Status': 'Success'
                    })
                else:
                    metrics = model_info.get('final_metrics', {})
                    row.update({
                        'Accuracy': metrics.get('accuracy', 'N/A'),
                        'F1_Score': metrics.get('f1_score', 'N/A'),
                        'Status': 'Success'
                    })
            else:
                error_info = model_info.get('error', 'unknown_error')
                row.update({'Status': 'Failed', 'Error': error_info})

            results_summary.append(row)

        results_df = pd.DataFrame(results_summary)
        results_df.to_excel(self.dirs['results'] / 'evaluation_scores.xlsx', index=False)

    def _save_preprocessing_params(self, missing_info: dict):
        """前処理パラメータの保存"""
        encoding_info = {}
        if 'onehot' in self.encoders:
            encoder = self.encoders['onehot']
            encoding_info['onehot'] = {
                'feature_names': encoder.get_feature_names_out().tolist(),
                'categories': [cat.tolist() for cat in encoder.categories_]
            }

        preprocessing_params = {
            'missing_value_treatment': missing_info,
            'onehot_encoding': encoding_info,
            'feature_selection': self.results.get('feature_selection', {}),
            'scaling_method': 'robust' if self.config.PREPROCESSING['use_robust_scaling'] else 'standard',
            'transformation_info': self.transformation_info
        }

        with open(self.dirs['parameters'] / 'preprocessing_params.json', 'w', encoding='utf-8') as f:
            json.dump(preprocessing_params, f, indent=2, ensure_ascii=False)

        if self.encoders:
            joblib.dump(self.encoders, self.dirs['parameters'] / 'encoders.pkl')

    def create_prediction_template(self):
        """予測テンプレート作成"""
        features = self.data['processed']['features_original']
        template_data = {}

        for col in features.columns:
            if features[col].dtype in ['int64', 'float64']:
                template_data[col] = [features[col].min(), features[col].mean(), features[col].max()]
            else:
                unique_values = features[col].unique()[:3]
                template_data[col] = unique_values.tolist()

        max_len = max(len(v) for v in template_data.values())
        for key, values in template_data.items():
            while len(values) < max_len:
                values.append(values[-1])

        template_df = pd.DataFrame(template_data)
        template_df.index.name = 'Sample'
        template_df.to_excel(self.dirs['predictions'] / 'prediction_template.xlsx')

        return template_df

    def calculate_propensity_scores(self) -> Dict[str, np.ndarray]:
        """傾向スコア計算"""
        propensity_scores = {}

        for target_col, model_info in self.models.items():
            if (model_info.get('model') is not None and
                model_info['task_type'] == 'classification' and
                'probabilities' in model_info):

                proba = model_info['probabilities']

                if proba is not None:
                    if proba.shape[1] == 2:
                        propensity_scores[target_col] = proba[:, 1]
                    else:
                        propensity_scores[target_col] = np.max(proba, axis=1)

        return propensity_scores

    def run_full_pipeline(self, file_path: str, target_columns: List[str] = None,
                         feature_columns: List[str] = None,
                         index_col: str = 'Index') -> Dict[str, Any]:
        """完全パイプライン実行"""
        try:
            self.load_data(file_path, index_col)
            self.separate_variables(target_columns, feature_columns)
            self.preprocess_data()
            self.train_models()
            propensity_scores = self.calculate_propensity_scores()
            self.save_results()
            self.create_prediction_template()
            self.save_prediction_formulas()
            excel_calculator_path = self.create_excel_prediction_calculator_with_inverse()

            return {
                'models': self.models,
                'target_info': self.target_info,
                'transformation_info': self.transformation_info,
                'propensity_scores': propensity_scores,
                'feature_selection': self.results.get('feature_selection', {}),
                'directories': {k: str(v) for k, v in self.dirs.items()},
                'excel_calculator': str(excel_calculator_path) if excel_calculator_path else None
            }

        except Exception as e:
            raise

    def create_excel_prediction_calculator_with_inverse(self, parent_widget=None):
        """逆変換対応Excel予測計算機作成"""
        if not OPENPYXL_AVAILABLE:
            print("⚠️ openpyxl no disponible, no se puede crear la calculadora Excel")
            return None

        try:
            print(f"🔧 Creando Excel calculator...")
            print(f"🔧 Modelos disponibles: {len(self.models) if hasattr(self, 'models') else 'No disponible'}")
            print(f"🔧 Transformación disponible: {len(self.transformation_info) if hasattr(self, 'transformation_info') else 'No disponible'}")
            
            prediction_info = self.load_models_for_excel_prediction()

            if not prediction_info:
                print("⚠️ No hay modelos compatibles para Excel")
                return None

            wb = Workbook()
            wb.remove(wb.active)

            ws_main = wb.create_sheet("予測計算機")
            self._create_main_prediction_sheet_with_inverse(ws_main, prediction_info)

            ws_params = wb.create_sheet("モデルパラメーター")
            self._create_parameters_sheet_with_transformation(ws_params, prediction_info)

            ws_manual = wb.create_sheet("使用方法")
            self._create_manual_sheet_with_transformation(ws_manual, prediction_info)

            wb.active = ws_main

            excel_file_path = self.dirs['predictions'] / 'XEBEC_予測計算機_逆変換対応.xlsx'
            wb.save(excel_file_path)

            print(f"✅ Excel creado exitosamente: {excel_file_path}")
            return excel_file_path
            
        except Exception as e:
            print(f"❌ Error creating Excel calculator: {e}")
            import traceback
            traceback.print_exc()
            return None


    def load_models_for_excel_prediction(self) -> Dict[str, Dict]:
        """Excel予測用モデル情報読み込み"""
        excel_prediction_info = {}

        regression_model_dir = self.dirs['models_regression']
        if regression_model_dir.exists():
            for model_file in regression_model_dir.glob('best_model_*.pkl'):
                target_name = model_file.stem.replace('best_model_', '')

                try:
                    model_data = joblib.load(model_file)
                    model = model_data['model']

                    if hasattr(model, 'coef_') and hasattr(model, 'intercept_'):
                        excel_info = self._extract_excel_prediction_info_with_transformation(
                            model, model_data, target_name, 'regression'
                        )
                        excel_prediction_info[target_name] = excel_info
                except Exception as e:
                    pass

        classification_model_dir = self.dirs['models_classification']
        if classification_model_dir.exists():
            for model_file in classification_model_dir.glob('best_model_*.pkl'):
                target_name = model_file.stem.replace('best_model_', '')

                try:
                    model_data = joblib.load(model_file)
                    model = model_data['model']

                    if hasattr(model, 'coef_') and hasattr(model, 'intercept_'):
                        excel_info = self._extract_excel_prediction_info_with_transformation(
                            model, model_data, target_name, 'classification'
                        )
                        excel_prediction_info[target_name] = excel_info
                except Exception as e:
                    pass

        return excel_prediction_info

    def _extract_excel_prediction_info_with_transformation(self, model, model_data, target_name, model_type):
        """Excel予測に必要な情報抽出"""
        feature_names = model_data.get('feature_names', [])
        coefficients = model.coef_
        intercept = model.intercept_

        scaler = self.scalers.get('features')
        scaling_params = {}

        if scaler is not None:
            feature_columns = self.data['processed']['features'].columns
            if hasattr(scaler, 'center_') and hasattr(scaler, 'scale_'):
                scaling_params = {
                    'method': 'robust',
                    'centers': dict(zip(feature_columns, scaler.center_)),
                    'scales': dict(zip(feature_columns, scaler.scale_))
                }
            elif hasattr(scaler, 'mean_') and hasattr(scaler, 'scale_'):
                scaling_params = {
                    'method': 'standard',
                    'means': dict(zip(feature_columns, scaler.mean_)),
                    'stds': dict(zip(feature_columns, scaler.scale_))
                }

        transformation_info = model_data.get('transformation_info', {'applied': False, 'method': 'none', 'parameters': {}})

        classification_info = {}
        if model_type == 'classification':
            label_encoder = model_data.get('label_encoder')

            if label_encoder:
                classification_info = {
                    'classes': label_encoder.classes_.tolist()
                }

            if coefficients.ndim > 1:
                coefficients = coefficients[0]

        return {
            'target_name': target_name,
            'model_type': model_type,
            'model_name': type(model).__name__,
            'feature_names': feature_names,
            'coefficients': coefficients.tolist() if hasattr(coefficients, 'tolist') else coefficients,
            'intercept': float(intercept) if hasattr(intercept, '__float__') else intercept,
            'scaling_params': scaling_params,
            'classification_info': classification_info,
            'transformation_info': transformation_info
        }

    def _create_main_prediction_sheet_with_inverse(self, ws, prediction_info):
        """メイン予測シート作成（逆変換対応）"""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        input_fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
        output_fill = PatternFill(start_color="F0F8E7", end_color="F0F8E7", fill_type="solid")
        transform_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))

        ws['A1'] = 'XEBEC加工条件予測システム（逆変換対応版）'
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:J1')

        all_features = set()
        for info in prediction_info.values():
            all_features.update(info['feature_names'])
        all_features = sorted(list(all_features))

        original_features = self.data['processed']['features_original']

        row = 3
        ws[f'A{row}'] = '【入力データ】'
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1

        ws[f'A{row}'] = '特徴量'
        ws[f'B{row}'] = '入力値'
        ws[f'C{row}'] = '単位/範囲'
        ws[f'D{row}'] = '参考値(平均)'

        for col in ['A', 'B', 'C', 'D']:
            cell = ws[f'{col}{row}']
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        row += 1
        input_start_row = row

        for i, feature in enumerate(all_features):
            ws[f'A{row}'] = feature
            ws[f'A{row}'].border = border

            if feature in original_features.columns:
                mean_val = float(original_features[feature].mean())
                min_val = float(original_features[feature].min())
                max_val = float(original_features[feature].max())

                ws[f'B{row}'] = mean_val
                ws[f'C{row}'] = f'{min_val:.3f} - {max_val:.3f}'
                ws[f'D{row}'] = f'{mean_val:.3f}'
            else:
                ws[f'B{row}'] = 0
                ws[f'C{row}'] = '設定値'
                ws[f'D{row}'] = '0'

            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].border = border

            ws[f'B{row}'].fill = input_fill
            row += 1

        input_end_row = row - 1

        row += 2
        ws[f'A{row}'] = '【予測結果】'
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1

        ws[f'A{row}'] = '目的変数'
        ws[f'B{row}'] = '変換後予測値'
        ws[f'C{row}'] = '実測値スケール予測値'
        ws[f'D{row}'] = 'モデル'
        ws[f'E{row}'] = '変換方法'

        for col in ['A', 'B', 'C', 'D', 'E']:
            cell = ws[f'{col}{row}']
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        row += 1

        for target_name, info in prediction_info.items():
            ws[f'A{row}'] = target_name
            ws[f'A{row}'].border = border

            prediction_formula = self._create_excel_prediction_formula(
                info, all_features, input_start_row
            )
            ws[f'B{row}'] = prediction_formula
            ws[f'B{row}'].fill = transform_fill
            ws[f'B{row}'].border = border

            inverse_formula = InverseTransformer.create_inverse_formula(
                info['transformation_info'], f'B{row}'
            )
            ws[f'C{row}'] = inverse_formula
            ws[f'C{row}'].fill = output_fill
            ws[f'C{row}'].border = border

            ws[f'D{row}'] = info.get('model_name', 'Linear')
            ws[f'D{row}'].border = border

            transformation_method = info['transformation_info'].get('method', 'none')
            ws[f'E{row}'] = transformation_method
            ws[f'E{row}'].border = border

            row += 1

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12

    def _create_excel_prediction_formula(self, info, all_features, input_start_row):
        """Excel予測式作成"""
        feature_names = info['feature_names']
        coefficients = info['coefficients']
        intercept = info['intercept']
        scaling_params = info['scaling_params']

        formula_parts = [str(intercept)]

        for feature, coef in zip(feature_names, coefficients):
            feature_row = input_start_row + all_features.index(feature)
            input_cell = f'B{feature_row}'

            if scaling_params and 'method' in scaling_params:
                if scaling_params['method'] == 'robust':
                    center = scaling_params['centers'].get(feature, 0)
                    scale = scaling_params['scales'].get(feature, 1)
                    scaled_value = f'(({input_cell}-{center})/{scale})'
                elif scaling_params['method'] == 'standard':
                    mean = scaling_params['means'].get(feature, 0)
                    std = scaling_params['stds'].get(feature, 1)
                    scaled_value = f'(({input_cell}-{mean})/{std})'
                else:
                    scaled_value = input_cell
            else:
                scaled_value = input_cell

            if coef >= 0:
                formula_parts.append(f'+{coef}*{scaled_value}')
            else:
                formula_parts.append(f'{coef}*{scaled_value}')

        if info['model_type'] == 'classification':
            linear_formula = ''.join(formula_parts)
            probability_formula = f'1/(1+EXP(-({linear_formula})))'
            binary_formula = f'IF({probability_formula}>0.5,1,0)'
            return f'={binary_formula}'
        else:
            return f'={"".join(formula_parts)}'

    def _create_parameters_sheet_with_transformation(self, ws, prediction_info):
        """パラメーターシート作成"""
        ws['A1'] = 'モデルパラメーター詳細（変換情報含む）'
        ws['A1'].font = Font(size=16, bold=True)

        row = 3

        for target_name, info in prediction_info.items():
            ws[f'A{row}'] = f'【{target_name}】'
            ws[f'A{row}'].font = Font(size=14, bold=True)
            row += 1

            ws[f'A{row}'] = 'モデル種別:'
            ws[f'B{row}'] = info['model_type']
            row += 1

            # 精度指標の取得
            model_info = self.models.get(target_name, {})
            metrics = model_info.get('final_metrics', {})
            if info['model_type'] == 'classification':
                score = metrics.get('f1_score', 'N/A')
                ws[f'A{row}'] = 'F1スコア:'
                ws[f'B{row}'] = score
                row += 1
            elif info['model_type'] == 'regression':
                score = metrics.get('r2', 'N/A')
                ws[f'A{row}'] = 'R²値:'
                ws[f'B{row}'] = score
                row += 1

            ws[f'A{row}'] = '切片:'
            ws[f'B{row}'] = info['intercept']
            row += 1

            transformation_info = info['transformation_info']
            ws[f'A{row}'] = '変換適用:'
            ws[f'B{row}'] = 'はい' if transformation_info.get('applied', False) else 'いいえ'
            row += 1

            if transformation_info.get('applied', False):
                ws[f'A{row}'] = '変換方法:'
                ws[f'B{row}'] = transformation_info.get('method', 'unknown')
                row += 1

                params = transformation_info.get('parameters', {})
                if params:
                    for param_name, param_value in params.items():
                        ws[f'A{row}'] = f'  {param_name}:'
                        ws[f'B{row}'] = param_value
                        row += 1

            ws[f'A{row}'] = '特徴量'
            ws[f'B{row}'] = '係数'
            ws[f'C{row}'] = '標準化中心'
            ws[f'D{row}'] = '標準化スケール'
            row += 1

            for feature, coef in zip(info['feature_names'], info['coefficients']):
                ws[f'A{row}'] = feature
                ws[f'B{row}'] = coef

                scaling_params = info['scaling_params']
                if scaling_params:
                    if scaling_params['method'] == 'robust':
                        ws[f'C{row}'] = scaling_params['centers'].get(feature, 0)
                        ws[f'D{row}'] = scaling_params['scales'].get(feature, 1)
                    elif scaling_params['method'] == 'standard':
                        ws[f'C{row}'] = scaling_params['means'].get(feature, 0)
                        ws[f'D{row}'] = scaling_params['stds'].get(feature, 1)

                row += 1

            if info['model_type'] == 'classification':
                classification_info = info['classification_info']
                ws[f'A{row}'] = 'クラス:'
                ws[f'B{row}'] = ', '.join(map(str, classification_info.get('classes', [])))
                row += 1

            row += 2

    def _create_manual_sheet_with_transformation(self, ws, prediction_info):
        """使用方法説明シート作成"""
        manual_text = [
            'XEBEC加工条件予測システム 使用方法（逆変換対応版）',
            '',
            '【概要】',
            'このExcelファイルは、加工条件パラメーターを入力すると自動的に予測結果を計算します。',
            '目的変数に変換が適用されている場合、実測値スケールでの予測値も自動計算されます。',
            '',
            '【使用手順】',
            '1. 「予測計算機」シートを開く',
            '2. 「入力データ」セクションの「入力値」列（B列）に実際の値を入力',
            '3. 「予測結果」セクションで自動計算された予測値を確認',
            '   - 「変換後予測値」：モデルが直接出力する値（変換されたスケール）',
            '   - 「実測値スケール予測値」：元のスケールに戻した実用的な予測値',
            '',
            '【重要な注意事項】',
            '・実用的な予測値は「実測値スケール予測値」列（C列）を参照してください',
            '・変換後予測値は技術的な参考値です',
            '',
            '【入力パラメーター説明】',
        ]

        all_features = set()
        for info in prediction_info.values():
            all_features.update(info['feature_names'])

        for feature in sorted(all_features):
            if feature in self.data['processed']['features_original'].columns:
                feature_data = self.data['processed']['features_original'][feature]
                min_val = feature_data.min()
                max_val = feature_data.max()
                mean_val = feature_data.mean()
                manual_text.append(f'・{feature}: {min_val:.3f} ～ {max_val:.3f} (平均: {mean_val:.3f})')

        manual_text.extend([
            '',
            '【予測結果説明】',
        ])

        for target_name, info in prediction_info.items():
            transformation_applied = info['transformation_info'].get('applied', False)
            transformation_method = info['transformation_info'].get('method', 'none')

            if info['model_type'] == 'regression':
                if transformation_applied:
                    manual_text.append(f'・{target_name}: 連続値（{transformation_method}変換適用済み→逆変換で実測値スケール）')
                else:
                    manual_text.append(f'・{target_name}: 連続値（変換なし）')
            else:
                manual_text.append(f'・{target_name}: 分類（0または1、1の場合に該当条件を満たす）')

        manual_text.extend([
            '',
            '【変換について】',
            '・一部の目的変数には統計的最適化のため変換が適用されています',
            '・log変換: 対数変換（広い値域の圧縮）',
            '・boxcox変換: Box-Cox変換（正規分布化）',
            '・sqrt変換: 平方根変換（分散安定化）',
            '・逆変換により、実用的な単位での予測値を提供します',
            '',
            '【注意事項】',
            '・入力値は訓練データの範囲内で使用することを推奨',
            '・範囲外の値では予測精度が低下する可能性があります',
            '・A32, A21, A11の値は必ず設定してください（ツール能力影響）',
            '',
            '【モデル詳細】',
            '・詳細なパラメーターは「モデルパラメーター」シートを参照',
            '・線形モデルベースで解釈性を重視',
            '・Excel上で完結するためPythonは不要',
        ])

        for i, text in enumerate(manual_text, 1):
            ws[f'A{i}'] = text
            if '【' in text and '】' in text:
                ws[f'A{i}'].font = Font(bold=True, size=12)
            elif text.startswith('・'):
                ws[f'A{i}'].font = Font(size=10)

        ws.column_dimensions['A'].width = 100

# --- ここからmain部 ---
if __name__ == "__main__":
    config = PipelineConfig()
    config.TRANSFORMATION['enable'] = True
    #config.TRANSFORMATION['mode'] = 'simple'  # または 'advanced'
    config.TRANSFORMATION['mode'] = 'advanced'  # より詳細な変換分析
    config.FEATURE_SELECTION['method'] = 'importance'
    config.FEATURE_SELECTION['k_features'] = 10
    config.PREPROCESSING['noise_augmentation_ratio'] = 0.3  # データ拡張を増加
    config.TRANSFORMATION['improvement_threshold'] = 0.005  # より感度良く変換適用

    pipeline = IntegratedMLPipeline(base_dir="xebec_analysis_v2", config=config)
    data_file = "20250701_A32総実験データ.xlsx"

    try:
        results = pipeline.run_full_pipeline(
            file_path=data_file,
            index_col='Index'
        )

        transformation_info = results.get('transformation_info', {})
        for target, trans_info in transformation_info.items():
            if trans_info.get('applied', False):
                method = trans_info.get('method', 'unknown')
                print(f"{target}: {method}変換適用")
            else:
                reason = trans_info.get('reason', 'unknown')
                print(f"{target}: 変換なし ({reason})")

        feature_selection_results = results.get('feature_selection', {})
        if feature_selection_results:
            selected_features = feature_selection_results.get('selected_features', [])
            mandatory_features = feature_selection_results.get('mandatory_features', [])
            preserved_mandatory = [f for f in mandatory_features if f in selected_features]
            print(f"必須特徴量保持: {preserved_mandatory}")

        successful_models = 0
        excel_compatible_models = 0
        transformed_models = 0

        for target, model_info in results['models'].items():
            if model_info['model'] is not None:
                successful_models += 1
                model_name = model_info['model_name']

                is_linear = model_name in ['LinearRegression', 'Ridge', 'Lasso', 'ElasticNet',
                                          'LogisticRegression', 'RidgeClassifier', 'LinearDiscriminantAnalysis']
                if is_linear:
                    excel_compatible_models += 1

                trans_info = transformation_info.get(target, {})
                if trans_info.get('applied', False):
                    transformed_models += 1
                    transform_label = f"({trans_info['method']}変換)"
                else:
                    transform_label = ""

                if model_info['task_type'] == 'regression':
                    metrics = model_info['final_metrics']
                    print(f"{target} ({model_name}): R²={metrics['r2']:.3f} {transform_label}")
                else:
                    metrics = model_info['final_metrics']
                    print(f"{target} ({model_name}): F1={metrics['f1_score']:.3f}")
            else:
                error_type = model_info.get('error', 'unknown')
                print(f"{target}: 訓練失敗 ({error_type})")

        print(f"成功モデル: {successful_models}/{len(results['models'])}")
        print(f"Excel対応: {excel_compatible_models}/{successful_models}")
        print(f"変換適用: {transformed_models}個")

        excel_calculator = results.get('excel_calculator')
        if excel_calculator:
            print(f"Excel予測計算機: {excel_calculator}")
            print("逆変換対応: 実測値スケールで予測値表示")
            print("使用方法: Excelファイルを開いて「予測計算機」シートで条件入力")
            print("実用的な予測値は「実測値スケール予測値」列（C列）を参照")

        print("XEBEC加工条件予測システム構築完了")

    except FileNotFoundError:
        print(f"エラー: ファイルが見つかりません: {data_file}")
        print("ファイル名が正しいか確認してください")

    except Exception as e:
        print(f"実行エラー: {e}")
        import traceback
        traceback.print_exc()


# ======================================
# FUNCIONES DE CONEXIÓN CON LA APLICACIÓN
# ======================================



def run_advanced_linear_analysis_from_db(db_manager, filters, output_folder):
    """
    Ejecuta el análisis lineal avanzado usando datos filtrados de la base de datos
    
    Args:
        db_manager: Instancia del gestor de base de datos
        filters: Diccionario con los filtros aplicados
        output_folder: Carpeta donde guardar los resultados
    
    Returns:
        dict: Resultados del análisis
    """
    try:
        print("🔧 Iniciando análisis lineal desde base de datos...")
        print(f"🔧 Filtros aplicados: {filters}")
        print(f"🔧 Carpeta de salida: {output_folder}")
        
        # Obtener datos filtrados de la base de datos
        print("📊 Obteniendo datos filtrados de la base de datos...")
        
        # Construir consulta SQL con filtros
        query = "SELECT * FROM main_results WHERE 1=1"
        params = []
        
        # Aplicar filtros de cepillo (A13, A11, A21, A32 son columnas directas en la tabla)
        brush_selections = []
        if 'すべて' in filters and filters['すべて']:
            # Si "すべて" está seleccionado, filtrar por cualquier cepillo que tenga valor 1
            brush_condition = " OR ".join([f"{brush} = 1" for brush in ['A13', 'A11', 'A21', 'A32']])
            query += f" AND ({brush_condition})"
            print("🔧 Filtro 'すべて' seleccionado - aplicando filtro para cualquier cepillo con valor 1")
        else:
            # Filtrar por cepillos específicos seleccionados
            for brush_type in ['A13', 'A11', 'A21', 'A32']:
                if brush_type in filters and filters[brush_type]:
                    brush_selections.append(brush_type)
            
            if brush_selections:
                brush_condition = " OR ".join([f"{brush} = 1" for brush in brush_selections])
                query += f" AND ({brush_condition})"
                print(f"🔧 Filtros de cepillo específicos aplicados: {brush_selections}")
            else:
                print("🔧 No se aplicaron filtros de cepillo específicos")
        
        # Aplicar filtros de rango
        range_filters_applied = []
        for field_name, filter_value in filters.items():
            if field_name in ['すべて', 'A13', 'A11', 'A21', 'A32']:
                continue
                
            # Verificar si es un filtro de rango (tupla) o valor único
            if isinstance(filter_value, tuple) and len(filter_value) == 2:
                desde, hasta = filter_value
                if desde is not None and hasta is not None:
                    if field_name == "実験日":
                        # Filtro de fecha
                        desde_str = desde.toString("yyyyMMdd") if hasattr(desde, 'toString') else str(desde)
                        hasta_str = hasta.toString("yyyyMMdd") if hasattr(hasta, 'toString') else str(hasta)
                        query += f" AND {field_name} BETWEEN ? AND ?"
                        params.extend([desde_str, hasta_str])
                        range_filters_applied.append(f"{field_name}: {desde_str} - {hasta_str}")
                    else:
                        # Filtro numérico - convertir a números
                        try:
                            desde_num = float(desde) if isinstance(desde, str) else desde
                            hasta_num = float(hasta) if isinstance(hasta, str) else hasta
                            query += f" AND {field_name} BETWEEN ? AND ?"
                            params.extend([desde_num, hasta_num])
                            range_filters_applied.append(f"{field_name}: {desde_num} - {hasta_num}")
                        except (ValueError, TypeError) as e:
                            print(f"⚠️ Error convirtiendo valores de filtro para {field_name}: {e}")
                            continue
            elif isinstance(filter_value, (str, int, float)) and filter_value:
                # Filtro de valor único - convertir a número si es posible
                try:
                    if field_name in ['線材長', '回転速度', '送り速度', 'UPカット', '突出量', 'パス数', 'バリ除去']:
                        # Columnas enteras
                        value_num = int(filter_value) if isinstance(filter_value, str) else filter_value
                    else:
                        # Columnas decimales
                        value_num = float(filter_value) if isinstance(filter_value, str) else filter_value
                    
                    query += f" AND {field_name} = ?"
                    params.append(value_num)
                    range_filters_applied.append(f"{field_name}: {value_num}")
                except (ValueError, TypeError) as e:
                    print(f"⚠️ Error convirtiendo valor de filtro para {field_name}: {e}")
                    continue
        
        if range_filters_applied:
            print(f"🔧 Filtros de rango aplicados: {range_filters_applied}")
        else:
            print("🔧 No se aplicaron filtros de rango")
        
        print(f"🔧 Query SQL: {query}")
        print(f"🔧 Parámetros: {params}")
        
        # Ejecutar consulta usando el método correcto del DBManager
        try:
            # Verificar qué tablas existen
            cursor = db_manager.conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
            available_tables = [row[0] for row in cursor.fetchall()]
            print(f"🔧 Tablas disponibles: {available_tables}")
            
            # Verificar qué tabla tiene datos
            target_table = None
            for table_name in available_tables:
                cursor.execute(f"SELECT COUNT(*) FROM {table_name};")
                count = cursor.fetchone()[0]
                print(f"📊 Tabla {table_name}: {count} registros")
                if count > 0:
                    target_table = table_name
                    break
            
            if not target_table:
                print("❌ No se encontraron datos en ninguna tabla")
                return {'success': False, 'error': 'No se encontraron datos en la base de datos. Por favor, asegúrese de que haya datos disponibles.'}
            
            # Usar el método fetch_filtered para aplicar filtros
            if hasattr(db_manager, 'fetch_filtered'):
                filtered_data = db_manager.fetch_filtered(target_table, query.replace("Results", target_table), params)
            else:
                # Fallback: obtener todos los datos si no hay método de filtrado
                filtered_data = db_manager.fetch_all(target_table)
            
            
            
            print(f"✅ Datos obtenidos: {len(filtered_data)} registros")
            
        except Exception as e:
            print(f"❌ Error obteniendo datos: {e}")
            return {'success': False, 'error': f'Error obteniendo datos: {str(e)}'}
        
        if not filtered_data:
            return {'success': False, 'error': 'No se encontraron datos con los filtros aplicados'}
        
        # Convertir a DataFrame
        import pandas as pd
        
        # Crear DataFrame con nombres de columnas correctos
        if filtered_data and len(filtered_data) > 0:
            # Nombres de columnas basados en la estructura real de main_results
            column_names = [
                'id', '実験日', 'バリ除去', '上面ダレ量', '側面ダレ量', '摩耗量', 
                '面粗度前', '面粗度後', 'A13', 'A11', 'A21', 'A32', '直径', '材料', 
                '線材長', '回転速度', '送り速度', 'UPカット', '切込量', '突出量', 
                '載せ率', 'パス数', '加工時間'
            ]
            
            df = pd.DataFrame(filtered_data, columns=column_names)
            print(f"✅ DataFrame creado con {len(df)} registros y {len(df.columns)} columnas")
            

        else:
            df = pd.DataFrame()
        
        # Obtener nombres de columnas
        if len(df) > 0:
            column_names = list(df.columns)
            print(f"📊 Columnas disponibles: {column_names}")
        else:
            return {'success': False, 'error': 'DataFrame vacío después de aplicar filtros'}
        
        # Crear estructura de carpetas si no existe
        os.makedirs(output_folder, exist_ok=True)
        models_folder = os.path.join(output_folder, "01_学習モデル")
        os.makedirs(models_folder, exist_ok=True)
        
        # DEBUG: Verificar datos justo antes de generar el archivo filtered_data.xlsx

        
        # Guardar datos filtrados
        filtered_data_path = os.path.join(models_folder, "filtered_data.xlsx")
        df.to_excel(filtered_data_path, index=False)
        print(f"✅ Datos filtrados guardados en: {filtered_data_path}")
        
        # Configurar y ejecutar el pipeline de análisis lineal
        print("🔧 Configurando pipeline de análisis lineal...")
        
        # Crear configuración personalizada
        config = PipelineConfig()
        config.TRANSFORMATION['enable'] = True
        config.TRANSFORMATION['mode'] = 'advanced'
        config.FEATURE_SELECTION['method'] = 'importance'
        config.FEATURE_SELECTION['k_features'] = 10
        config.PREPROCESSING['noise_augmentation_ratio'] = 0.3
        config.TRANSFORMATION['improvement_threshold'] = 0.005
        
        # Crear pipeline con la carpeta de salida personalizada
        pipeline = IntegratedMLPipeline(base_dir=output_folder, config=config)
        
        # DEBUG: Verificar datos justo antes de empezar el análisis lineal

        
        # Ejecutar análisis completo
        print("🚀 Ejecutando análisis lineal completo...")
        results = pipeline.run_full_pipeline(
            file_path=filtered_data_path,
            index_col='Index'
        )
        
        
        
        # Preparar resultados para la aplicación
        analysis_results = {
            'success': True,
            'data_count': len(df),
            'models_trained': len(results.get('models', {})),
            'output_folder': output_folder,
            'filters_applied': range_filters_applied,
            'data_range': f"線材長: {df['線材長'].min()}-{df['線材長'].max()}, 送り速度: {df['送り速度'].min()}-{df['送り速度'].max()}" if len(df) > 0 else "N/A",
            'excel_calculator': results.get('excel_calculator'),
            'transformation_info': results.get('transformation_info', {}),
            'feature_selection': results.get('feature_selection', {}),
            'target_info': results.get('target_info', {}),
            'models': results.get('models', {})
        }
        
        # Crear resumen de resultados
        summary = []
        for target_name, model_info in results.get('models', {}).items():
            if model_info.get('model') is not None:
                if model_info['task_type'] == 'regression':
                    metrics = model_info.get('final_metrics', {})
                    summary.append({
                        'target': target_name,
                        'model': model_info.get('model_name', 'Unknown'),
                        'r2': metrics.get('r2', 'N/A'),
                        'mae': metrics.get('mae', 'N/A'),
                        'rmse': metrics.get('rmse', 'N/A'),
                        'transformation': results.get('transformation_info', {}).get(target_name, {}).get('method', 'none')
                    })
                else:
                    metrics = model_info.get('final_metrics', {})
                    summary.append({
                        'target': target_name,
                        'model': model_info.get('model_name', 'Unknown'),
                        'accuracy': metrics.get('accuracy', 'N/A'),
                        'f1_score': metrics.get('f1_score', 'N/A'),
                        'transformation': 'none'
                    })
        
        analysis_results['summary'] = summary
        
        print("✅ Análisis lineal completado exitosamente")
        print(f"📊 Modelos entrenados: {len(results.get('models', {}))}")
        print(f"📁 Resultados guardados en: {output_folder}")
        
        return analysis_results
        
    except Exception as e:
        print(f"❌ Error en análisis lineal: {e}")
        import traceback
        traceback.print_exc()
        return {'success': False, 'error': str(e)}

def create_analysis_summary_table(results):
    """
    Crea una tabla de resumen para mostrar en la aplicación
    
    Args:
        results: Resultados del análisis lineal
        
    Returns:
        list: Lista de listas para crear tabla
    """
    if not results.get('success', False):
        return []
    
    summary = results.get('summary', [])
    if not summary:
        return []
    
    # Crear encabezados
    headers = ['目的変数', 'モデル', 'R²/精度', 'MAE', 'RMSE', '変換']
    table_data = [headers]
    
    # Agregar datos
    for item in summary:
        target = item.get('target', 'Unknown')
        model = item.get('model', 'Unknown')
        
        # Determinar métrica principal
        if 'r2' in item:
            # Regresión
            metric = f"R² = {item.get('r2', 'N/A'):.3f}" if isinstance(item.get('r2'), (int, float)) else 'N/A'
            mae = f"{item.get('mae', 'N/A'):.3f}" if isinstance(item.get('mae'), (int, float)) else 'N/A'
            rmse = f"{item.get('rmse', 'N/A'):.3f}" if isinstance(item.get('rmse'), (int, float)) else 'N/A'
        else:
            # Clasificación
            metric = f"精度 = {item.get('accuracy', 'N/A'):.3f}" if isinstance(item.get('accuracy'), (int, float)) else 'N/A'
            mae = 'N/A'
            rmse = 'N/A'
        
        transformation = item.get('transformation', 'none')
        
        table_data.append([target, model, metric, mae, rmse, transformation])
    
    return table_data
