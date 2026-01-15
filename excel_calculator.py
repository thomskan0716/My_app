#!/usr/bin/env python
# coding: utf-8

"""
Excel Calculator Module for Advanced Linear Analysis
Replicates the Excel functionality from the reference file
"""

import os
import joblib
from pathlib import Path
from typing import Dict, Any, Optional, Callable

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Import the loading dialog
try:
    from excel_loading_dialog import ExcelProcessingDialog
    LOADING_DIALOG_AVAILABLE = True
except ImportError:
    LOADING_DIALOG_AVAILABLE = False

class ExcelCalculator:
    """Excel予測計算機クラス"""

    def __init__(self, output_dir: str = "xebec_analysis_v2", parent_widget=None):
        self.output_dir = Path(output_dir)
        self.predictions_dir = self.output_dir / '06_predictions'
        self.parent_widget = parent_widget
        self.loading_dialog = None

    def _show_loading_dialog(self, total_rows: int) -> bool:
        """Show loading dialog and return True if user wants to continue"""
        if LOADING_DIALOG_AVAILABLE and self.parent_widget:
            self.loading_dialog = ExcelProcessingDialog(self.parent_widget, total_rows)
            self.loading_dialog.set_total_rows(total_rows)
            return self.loading_dialog.exec() == ExcelProcessingDialog.Accepted
        return True

    def _update_progress(self, processed_rows: int, current_task: str = ""):
        """Update progress in loading dialog"""
        if self.loading_dialog:
            self.loading_dialog.update_progress(processed_rows, current_task)

    def _add_detail(self, message: str):
        """Add detail message to loading dialog"""
        if self.loading_dialog:
            self.loading_dialog.add_detail(message)

    def _complete_processing(self):
        """Mark processing as complete"""
        if self.loading_dialog:
            self.loading_dialog.processing_complete()

    def _error_processing(self, error_message: str):
        """Mark processing as failed"""
        if self.loading_dialog:
            self.loading_dialog.processing_error(error_message)

    def create_excel_prediction_calculator_with_inverse(self, models: Dict, transformation_info: Dict) -> str:
        """逆変換対応Excel予測計算機作成"""
        if not OPENPYXL_AVAILABLE:
            print("⚠️ openpyxl no disponible, no se puede crear la calculadora Excel")
            return None

        try:
            prediction_info = self.load_models_for_excel_prediction(models)

            if not prediction_info:
                print("⚠️ No hay modelos compatibles para Excel")
                return None

            # Calculate total operations for progress tracking
            total_operations = len(prediction_info) * 3  # 3 operations per model: main sheet, params sheet, manual sheet
            total_operations += sum(len(info['feature_names']) for info in prediction_info.values())  # Formula creation per feature
            
            # Show loading dialog
            if not self._show_loading_dialog(total_operations):
                return None

            self._add_detail("Iniciando creación del archivo Excel...")
            
            wb = Workbook()
            wb.remove(wb.active)
            processed_ops = 0

            # Create main prediction sheet
            self._add_detail("Creando hoja principal de predicción...")
            ws_main = wb.create_sheet("予測計算機")
            self._create_main_prediction_sheet_with_inverse(ws_main, prediction_info)
            processed_ops += 1
            self._update_progress(processed_ops, "Hoja principal creada")

            # Create parameters sheet
            self._add_detail("Creando hoja de parámetros...")
            ws_params = wb.create_sheet("モデルパラメーター")
            self._create_parameters_sheet_with_transformation(ws_params, prediction_info)
            processed_ops += 1
            self._update_progress(processed_ops, "Hoja de parámetros creada")

            # Create manual sheet
            self._add_detail("Creando hoja de instrucciones...")
            ws_manual = wb.create_sheet("使用方法")
            self._create_manual_sheet_with_transformation(ws_manual, prediction_info)
            processed_ops += 1
            self._update_progress(processed_ops, "Hoja de instrucciones creada")

            wb.active = ws_main

            # Save Excel file
            self._add_detail("Guardando archivo Excel...")
            excel_file_path = self.predictions_dir / 'XEBEC_予測計算機_逆変換対応.xlsx'
            wb.save(excel_file_path)
            
            self._complete_processing()
            self._add_detail(f"✅ Archivo guardado: {excel_file_path}")
            
            return str(excel_file_path)
            
        except Exception as e:
            error_msg = f"Error creando calculadora Excel: {str(e)}"
            print(f"❌ {error_msg}")
            self._error_processing(error_msg)
            return None

    def load_models_for_excel_prediction(self, models: Dict) -> Dict[str, Dict]:
        """Excel予測用モデル情報読み込み"""
        excel_prediction_info = {}

        for target_name, model_info in models.items():
            if model_info.get('model') is None:
                continue

            model = model_info['model']
            if not hasattr(model, 'coef_') or not hasattr(model, 'intercept_'):
                continue

            excel_info = self._extract_excel_prediction_info_with_transformation(
                model, model_info, target_name, model_info.get('task_type', 'regression')
            )
            excel_prediction_info[target_name] = excel_info

        return excel_prediction_info

    def _extract_excel_prediction_info_with_transformation(self, model, model_info, target_name, model_type):
        """Excel予測に必要な情報抽出"""
        feature_names = model_info.get('feature_names', [])
        coefficients = model.coef_
        intercept = model.intercept_

        scaling_params = {}
        if 'scaler' in model_info:
            scaler = model_info['scaler']
            if hasattr(scaler, 'center_') and hasattr(scaler, 'scale_'):
                scaling_params = {
                    'method': 'robust',
                    'centers': dict(zip(feature_names, scaler.center_)),
                    'scales': dict(zip(feature_names, scaler.scale_))
                }
            elif hasattr(scaler, 'mean_') and hasattr(scaler, 'scale_'):
                scaling_params = {
                    'method': 'standard',
                    'means': dict(zip(feature_names, scaler.mean_)),
                    'stds': dict(zip(feature_names, scaler.scale_))
                }

        transformation_info = model_info.get('transformation_info', {'applied': False, 'method': 'none', 'parameters': {}})

        classification_info = {}
        if model_type == 'classification':
            if 'label_encoder' in model_info:
                label_encoder = model_info['label_encoder']
                classification_info = {
                    'classes': label_encoder.classes_.tolist()
                }

            if coefficients.ndim > 1:
                coefficients = coefficients[0]

        return {
            'target_name': target_name,
            'model_type': model_type,
            'model_name': type(model).__name__,
            'feature_names': feature_names,
            'coefficients': coefficients.tolist() if hasattr(coefficients, 'tolist') else coefficients,
            'intercept': float(intercept) if hasattr(intercept, '__float__') else intercept,
            'scaling_params': scaling_params,
            'classification_info': classification_info,
            'transformation_info': transformation_info
        }

    def _create_main_prediction_sheet_with_inverse(self, ws, prediction_info):
        """メイン予測シート作成（逆変換対応）"""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        input_fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
        output_fill = PatternFill(start_color="F0F8E7", end_color="F0F8E7", fill_type="solid")
        transform_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))

        ws['A1'] = 'XEBEC加工条件予測システム（逆変換対応版）'
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:J1')

        all_features = set()
        for info in prediction_info.values():
            all_features.update(info['feature_names'])
        all_features = sorted(list(all_features))

        row = 3
        ws[f'A{row}'] = '【入力データ】'
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1

        ws[f'A{row}'] = '特徴量'
        ws[f'B{row}'] = '入力値'
        ws[f'C{row}'] = '単位/範囲'
        ws[f'D{row}'] = '参考値(平均)'

        for col in ['A', 'B', 'C', 'D']:
            cell = ws[f'{col}{row}']
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        row += 1
        input_start_row = row

        # Process input features with progress tracking
        for i, feature in enumerate(all_features):
            self._update_progress(i + 1, f"Procesando característica: {feature}")
            
            ws[f'A{row}'] = feature
            ws[f'A{row}'].border = border

            ws[f'B{row}'] = 0
            ws[f'C{row}'] = '設定値'
            ws[f'D{row}'] = '0'

            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].border = border

            ws[f'B{row}'].fill = input_fill
            row += 1

        input_end_row = row - 1

        row += 2
        ws[f'A{row}'] = '【予測結果】'
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1

        ws[f'A{row}'] = '目的変数'
        ws[f'B{row}'] = '変換後予測値'
        ws[f'C{row}'] = '実測値スケール予測値'
        ws[f'D{row}'] = 'モデル'
        ws[f'E{row}'] = '変換方法'

        for col in ['A', 'B', 'C', 'D', 'E']:
            cell = ws[f'{col}{row}']
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        row += 1

        # Process prediction formulas with progress tracking
        for i, (target_name, info) in enumerate(prediction_info.items()):
            self._update_progress(len(all_features) + i + 1, f"Creando fórmula para: {target_name}")
            
            ws[f'A{row}'] = target_name
            ws[f'A{row}'].border = border

            prediction_formula = self._create_excel_prediction_formula(
                info, all_features, input_start_row
            )
            ws[f'B{row}'] = prediction_formula
            ws[f'B{row}'].fill = transform_fill
            ws[f'B{row}'].border = border

            inverse_formula = self._create_inverse_formula(
                info['transformation_info'], f'B{row}'
            )
            ws[f'C{row}'] = inverse_formula
            ws[f'C{row}'].fill = output_fill
            ws[f'C{row}'].border = border

            ws[f'D{row}'] = info.get('model_name', 'Linear')
            ws[f'D{row}'].border = border

            transformation_method = info['transformation_info'].get('method', 'none')
            ws[f'E{row}'] = transformation_method
            ws[f'E{row}'].border = border

            row += 1

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12

    def _create_excel_prediction_formula(self, info, all_features, input_start_row):
        """Excel予測式作成"""
        feature_names = info['feature_names']
        coefficients = info['coefficients']
        intercept = info['intercept']
        scaling_params = info['scaling_params']

        formula_parts = [str(intercept)]

        for feature, coef in zip(feature_names, coefficients):
            feature_row = input_start_row + all_features.index(feature)
            input_cell = f'B{feature_row}'

            if scaling_params and 'method' in scaling_params:
                if scaling_params['method'] == 'robust':
                    center = scaling_params['centers'].get(feature, 0)
                    scale = scaling_params['scales'].get(feature, 1)
                    scaled_value = f'(({input_cell}-{center})/{scale})'
                elif scaling_params['method'] == 'standard':
                    mean = scaling_params['means'].get(feature, 0)
                    std = scaling_params['stds'].get(feature, 1)
                    scaled_value = f'(({input_cell}-{mean})/{std})'
                else:
                    scaled_value = input_cell
            else:
                scaled_value = input_cell

            if coef >= 0:
                formula_parts.append(f'+{coef}*{scaled_value}')
            else:
                formula_parts.append(f'{coef}*{scaled_value}')

        if info['model_type'] == 'classification':
            linear_formula = ''.join(formula_parts)
            probability_formula = f'1/(1+EXP(-({linear_formula})))'
            binary_formula = f'IF({probability_formula}>0.5,1,0)'
            return f'={binary_formula}'
        else:
            return f'={"".join(formula_parts)}'

    def _create_inverse_formula(self, transformation_info, prediction_cell):
        """逆変換式作成"""
        if not transformation_info.get('applied', False):
            return f"={prediction_cell}"

        method = transformation_info.get('method', 'none')
        params = transformation_info.get('parameters', {})

        if method == 'log':
            return f"=EXP({prediction_cell})"
        elif method == 'log10':
            return f"=POWER(10,{prediction_cell})"
        elif method == 'sqrt':
            return f"=POWER({prediction_cell},2)"
        elif method == 'boxcox':
            lambda_val = params.get('lambda', 0)
            if abs(lambda_val) < 1e-6:
                return f"=EXP({prediction_cell})"
            else:
                return f"=POWER({lambda_val}*{prediction_cell}+1,1/{lambda_val})"
        elif method == 'yeo_johnson':
            lambda_val = params.get('lambda', 0)
            if abs(lambda_val) < 1e-6:
                return f"=EXP({prediction_cell})-1"
            else:
                return f"=POWER({lambda_val}*{prediction_cell}+1,1/{lambda_val})-1"
        else:
            return f"={prediction_cell}"

    def _create_parameters_sheet_with_transformation(self, ws, prediction_info):
        """パラメーターシート作成"""
        ws['A1'] = 'モデルパラメーター詳細（変換情報含む）'
        ws['A1'].font = Font(size=16, bold=True)

        row = 3

        for target_name, info in prediction_info.items():
            ws[f'A{row}'] = f'【{target_name}】'
            ws[f'A{row}'].font = Font(size=14, bold=True)
            row += 1

            ws[f'A{row}'] = 'モデル種別:'
            ws[f'B{row}'] = info['model_type']
            row += 1

            ws[f'A{row}'] = '切片:'
            ws[f'B{row}'] = info['intercept']
            row += 1

            transformation_info = info['transformation_info']
            ws[f'A{row}'] = '変換適用:'
            ws[f'B{row}'] = 'はい' if transformation_info.get('applied', False) else 'いいえ'
            row += 1

            if transformation_info.get('applied', False):
                ws[f'A{row}'] = '変換方法:'
                ws[f'B{row}'] = transformation_info.get('method', 'unknown')
                row += 1

                params = transformation_info.get('parameters', {})
                if params:
                    for param_name, param_value in params.items():
                        ws[f'A{row}'] = f'  {param_name}:'
                        ws[f'B{row}'] = param_value
                        row += 1

            ws[f'A{row}'] = '特徴量'
            ws[f'B{row}'] = '係数'
            ws[f'C{row}'] = '標準化中心'
            ws[f'D{row}'] = '標準化スケール'
            row += 1

            for feature, coef in zip(info['feature_names'], info['coefficients']):
                ws[f'A{row}'] = feature
                ws[f'B{row}'] = coef

                scaling_params = info['scaling_params']
                if scaling_params:
                    if scaling_params['method'] == 'robust':
                        ws[f'C{row}'] = scaling_params['centers'].get(feature, 0)
                        ws[f'D{row}'] = scaling_params['scales'].get(feature, 1)
                    elif scaling_params['method'] == 'standard':
                        ws[f'C{row}'] = scaling_params['means'].get(feature, 0)
                        ws[f'D{row}'] = scaling_params['stds'].get(feature, 1)

                row += 1

            if info['model_type'] == 'classification':
                classification_info = info['classification_info']
                ws[f'A{row}'] = 'クラス:'
                ws[f'B{row}'] = ', '.join(map(str, classification_info.get('classes', [])))
                row += 1

            row += 2

    def _create_manual_sheet_with_transformation(self, ws, prediction_info):
        """使用方法説明シート作成"""
        manual_text = [
            'XEBEC加工条件予測システム 使用方法（逆変換対応版）',
            '',
            '【概要】',
            'このExcelファイルは、加工条件パラメーターを入力すると自動的に予測結果を計算します。',
            '目的変数に変換が適用されている場合、実測値スケールでの予測値も自動計算されます。',
            '',
            '【使用手順】',
            '1. 「予測計算機」シートを開く',
            '2. 「入力データ」セクションの「入力値」列（B列）に実際の値を入力',
            '3. 「予測結果」セクションで自動計算された予測値を確認',
            '   - 「変換後予測値」：モデルが直接出力する値（変換されたスケール）',
            '   - 「実測値スケール予測値」：元のスケールに戻した実用的な予測値',
            '',
            '【重要な注意事項】',
            '・実用的な予測値は「実測値スケール予測値」列（C列）を参照してください',
            '・変換後予測値は技術的な参考値です',
            '',
            '【入力パラメーター説明】',
        ]

        all_features = set()
        for info in prediction_info.values():
            all_features.update(info['feature_names'])

        for feature in sorted(all_features):
            manual_text.append(f'・{feature}: 設定値')

        manual_text.extend([
            '',
            '【予測結果説明】',
        ])

        for target_name, info in prediction_info.items():
            transformation_applied = info['transformation_info'].get('applied', False)
            transformation_method = info['transformation_info'].get('method', 'none')

            if info['model_type'] == 'regression':
                if transformation_applied:
                    manual_text.append(f'・{target_name}: 連続値（{transformation_method}変換適用済み→逆変換で実測値スケール）')
                else:
                    manual_text.append(f'・{target_name}: 連続値（変換なし）')
            else:
                manual_text.append(f'・{target_name}: 分類（0または1、1の場合に該当条件を満たす）')

        manual_text.extend([
            '',
            '【変換について】',
            '・一部の目的変数には統計的最適化のため変換が適用されています',
            '・log変換: 対数変換（広い値域の圧縮）',
            '・boxcox変換: Box-Cox変換（正規分布化）',
            '・sqrt変換: 平方根変換（分散安定化）',
            '・逆変換により、実用的な単位での予測値を提供します',
            '',
            '【注意事項】',
            '・入力値は訓練データの範囲内で使用することを推奨',
            '・範囲外の値では予測精度が低下する可能性があります',
            '・A32, A21, A11の値は必ず設定してください（ツール能力影響）',
            '',
            '【モデル詳細】',
            '・詳細なパラメーターは「モデルパラメーター」シートを参照',
            '・線形モデルベースで解釈性を重視',
            '・Excel上で完結するためPythonは不要',
        ])

        for i, text in enumerate(manual_text, 1):
            ws[f'A{i}'] = text
            if '【' in text and '】' in text:
                ws[f'A{i}'].font = Font(bold=True, size=12)
            elif text.startswith('・'):
                ws[f'A{i}'].font = Font(size=10)

        ws.column_dimensions['A'].width = 100

